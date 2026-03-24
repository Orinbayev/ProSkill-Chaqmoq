from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import PricingFeature, PricingPlan

logger = logging.getLogger(__name__)

HIGHLIGHT_FEATURE_LIMIT = 6

EXPORT_COLUMNS = [
    "id",
    "name",
    "student_range",
    "duration_months",
    "current_price",
    "old_price",
    "badge_text",
    "discount_label",
    "is_recommended",
    "is_active",
    "order",
    "name_uz",
    "name_ru",
    "name_en",
    "student_range_uz",
    "student_range_ru",
    "student_range_en",
    "badge_text_uz",
    "badge_text_ru",
    "badge_text_en",
    "discount_label_uz",
    "discount_label_ru",
    "discount_label_en",
    "highlight_features",
    "detail_features",
    "highlight_features_uz",
    "highlight_features_ru",
    "highlight_features_en",
    "detail_features_uz",
    "detail_features_ru",
    "detail_features_en",
]

IMPORT_REQUIRED_COLUMNS = {"name", "student_range", "duration_months", "current_price"}

IMPORT_MODE_CREATE_ONLY = "create_only"
IMPORT_MODE_UPDATE_EXISTING = "update_existing"
IMPORT_MODE_SKIP_DUPLICATES = "skip_duplicates"

IMPORT_MODES = {
    IMPORT_MODE_CREATE_ONLY,
    IMPORT_MODE_UPDATE_EXISTING,
    IMPORT_MODE_SKIP_DUPLICATES,
}

SUPPORTED_IMPORT_FORMATS = {"xlsx", "json"}
SUPPORTED_EXPORT_FORMATS = {"xlsx", "json", "csv"}

TRUE_VALUES = {
    "1",
    "true",
    "t",
    "yes",
    "y",
    "on",
    "ha",
    "faol",
}
FALSE_VALUES = {
    "0",
    "false",
    "f",
    "no",
    "n",
    "off",
    "yoq",
    "yo'q",
    "nofaol",
}

COLUMN_ALIASES = {
    "fallback_name": "name",
    "name_fallback": "name",
    "tarif_nomi_fallback": "name",
    "plan_name": "name",
    "fallback_student_limit": "student_range",
    "student_limit_fallback": "student_range",
    "oquvchi_limiti_fallback": "student_range",
    "student_limit": "student_range",
    "duration": "duration_months",
    "duration_in_months": "duration_months",
    "joriy_narx": "current_price",
    "eski_narx": "old_price",
    "badge": "badge_text",
    "badge_matni": "badge_text",
    "discount": "discount_label",
    "chegirma_yorligi": "discount_label",
    "sort_order": "order",
    "tartib": "order",
    "faol": "is_active",
    "tavsiya_etilgan": "is_recommended",
    "name_uzbek": "name_uz",
    "name_russian": "name_ru",
    "name_english": "name_en",
    "tarif_nomi_uz": "name_uz",
    "tarif_nomi_ru": "name_ru",
    "tarif_nomi_en": "name_en",
    "student_limit_uz": "student_range_uz",
    "student_limit_ru": "student_range_ru",
    "student_limit_en": "student_range_en",
    "oquvchi_limiti_uz": "student_range_uz",
    "oquvchi_limiti_ru": "student_range_ru",
    "oquvchi_limiti_en": "student_range_en",
    "badge_uz": "badge_text_uz",
    "badge_ru": "badge_text_ru",
    "badge_en": "badge_text_en",
    "discount_uz": "discount_label_uz",
    "discount_ru": "discount_label_ru",
    "discount_en": "discount_label_en",
    "chegirma_yorligi_uz": "discount_label_uz",
    "chegirma_yorligi_ru": "discount_label_ru",
    "chegirma_yorligi_en": "discount_label_en",
}

PLAN_STRING_FIELDS = [
    "name",
    "name_uz",
    "name_ru",
    "name_en",
    "student_range",
    "student_range_uz",
    "student_range_ru",
    "student_range_en",
    "discount_label",
    "discount_label_uz",
    "discount_label_ru",
    "discount_label_en",
    "badge_text",
    "badge_text_uz",
    "badge_text_ru",
    "badge_text_en",
]

PLAN_FIELD_MAX_LENGTH = {
    field_name: PricingPlan._meta.get_field(field_name).max_length
    for field_name in PLAN_STRING_FIELDS
}
FEATURE_TEXT_MAX_LENGTH = PricingFeature._meta.get_field("text").max_length


SAMPLE_PLAN_DATA = [
    {
        "id": "",
        "name": "Standart",
        "name_uz": "Standart",
        "name_ru": "Стандарт",
        "name_en": "Standard",
        "student_range": "0-200 ta",
        "student_range_uz": "0-200 ta",
        "student_range_ru": "0-200 учеников",
        "student_range_en": "0-200 students",
        "duration_months": 3,
        "current_price": Decimal("390000"),
        "old_price": Decimal("450000"),
        "badge_text": "Yangi",
        "badge_text_uz": "Yangi",
        "badge_text_ru": "Новый",
        "badge_text_en": "New",
        "discount_label": "-13%",
        "discount_label_uz": "-13%",
        "discount_label_ru": "-13%",
        "discount_label_en": "-13%",
        "is_recommended": False,
        "is_active": True,
        "order": 1,
        "highlight_features": [
            {"text": "Guruhlar va davomat", "text_uz": "Guruhlar va davomat", "text_ru": "Группы и посещаемость", "text_en": "Groups and attendance"},
            {"text": "To'lov nazorati", "text_uz": "To'lov nazorati", "text_ru": "Контроль платежей", "text_en": "Payment control"},
            {"text": "Dashboard", "text_uz": "Dashboard", "text_ru": "Дашборд", "text_en": "Dashboard"},
        ],
        "detail_features": [
            {"text": "Davomat: Sababli va sababsiz nazorat", "text_uz": "Davomat: Sababli va sababsiz nazorat", "text_ru": "Посещаемость: учёт причин", "text_en": "Attendance: reason tracking"},
            {"text": "To'lovlar: To'lov kiritish", "text_uz": "To'lovlar: To'lov kiritish", "text_ru": "Платежи: ввод оплаты", "text_en": "Payments: payment entry"},
            {"text": "Hisobotlar: Oylik statistika", "text_uz": "Hisobotlar: Oylik statistika", "text_ru": "Отчёты: месячная статистика", "text_en": "Reports: monthly statistics"},
        ],
    },
    {
        "id": "",
        "name": "Premium",
        "name_uz": "Premium",
        "name_ru": "Премиум",
        "name_en": "Premium",
        "student_range": "0-500 ta",
        "student_range_uz": "0-500 ta",
        "student_range_ru": "0-500 учеников",
        "student_range_en": "0-500 students",
        "duration_months": 6,
        "current_price": Decimal("690000"),
        "old_price": Decimal("790000"),
        "badge_text": "Tavsiya",
        "badge_text_uz": "Tavsiya",
        "badge_text_ru": "Рекомендуем",
        "badge_text_en": "Recommended",
        "discount_label": "-12%",
        "discount_label_uz": "-12%",
        "discount_label_ru": "-12%",
        "discount_label_en": "-12%",
        "is_recommended": True,
        "is_active": True,
        "order": 2,
        "highlight_features": [
            {"text": "To'liq CRM modullar", "text_uz": "To'liq CRM modullar", "text_ru": "Полный CRM", "text_en": "Full CRM modules"},
            {"text": "Telegram xabarnoma", "text_uz": "Telegram xabarnoma", "text_ru": "Telegram уведомления", "text_en": "Telegram notifications"},
            {"text": "O'qituvchi KPI", "text_uz": "O'qituvchi KPI", "text_ru": "KPI преподавателей", "text_en": "Teacher KPI"},
        ],
        "detail_features": [
            {"text": "Lidlar: Sotuv voronkasi", "text_uz": "Lidlar: Sotuv voronkasi", "text_ru": "Лиды: воронка продаж", "text_en": "Leads: sales funnel"},
            {"text": "Hisobotlar: Filial kesimi", "text_uz": "Hisobotlar: Filial kesimi", "text_ru": "Отчёты: по филиалам", "text_en": "Reports: by branch"},
            {"text": "Support: Prioritet yordam", "text_uz": "Support: Prioritet yordam", "text_ru": "Поддержка: приоритет", "text_en": "Support: priority help"},
        ],
    },
    {
        "id": "",
        "name": "Pro",
        "name_uz": "Pro",
        "name_ru": "Про",
        "name_en": "Pro",
        "student_range": "500+ ta",
        "student_range_uz": "500+ ta",
        "student_range_ru": "500+ учеников",
        "student_range_en": "500+ students",
        "duration_months": 12,
        "current_price": Decimal("1090000"),
        "old_price": Decimal("1290000"),
        "badge_text": "Enterprise",
        "badge_text_uz": "Enterprise",
        "badge_text_ru": "Enterprise",
        "badge_text_en": "Enterprise",
        "discount_label": "-16%",
        "discount_label_uz": "-16%",
        "discount_label_ru": "-16%",
        "discount_label_en": "-16%",
        "is_recommended": False,
        "is_active": True,
        "order": 3,
        "highlight_features": [
            {"text": "Cheksiz filial", "text_uz": "Cheksiz filial", "text_ru": "Неограниченные филиалы", "text_en": "Unlimited branches"},
            {"text": "Avtomatik billing", "text_uz": "Avtomatik billing", "text_ru": "Автобиллинг", "text_en": "Automated billing"},
            {"text": "KPI + AI tavsiya", "text_uz": "KPI + AI tavsiya", "text_ru": "KPI + AI рекомендации", "text_en": "KPI + AI insights"},
        ],
        "detail_features": [
            {"text": "Integratsiya: API va webhook", "text_uz": "Integratsiya: API va webhook", "text_ru": "Интеграция: API и webhook", "text_en": "Integration: API and webhooks"},
            {"text": "Monitoring: SLA panel", "text_uz": "Monitoring: SLA panel", "text_ru": "Мониторинг: SLA панель", "text_en": "Monitoring: SLA panel"},
            {"text": "Security: Audit log", "text_uz": "Security: Audit log", "text_ru": "Безопасность: аудит", "text_en": "Security: audit log"},
        ],
    },
]


@dataclass
class ImportErrorRow:
    row: str
    field: str
    message: str


@dataclass
class ParsedPlanRow:
    row_label: str
    import_id: int | None
    fields: dict[str, Any]
    features: list[dict[str, str]]


@dataclass
class PricingPlanImportResult:
    filename: str
    file_format: str
    mode: str
    total_rows: int = 0
    created_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    errors: list[ImportErrorRow] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    def add_error(self, *, row: str, field: str, message: str) -> None:
        self.errors.append(ImportErrorRow(row=row, field=field, message=message))


class PricingPlanExportService:
    def export_xlsx(self, queryset) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tariflar"

        sheet.append(EXPORT_COLUMNS)

        for plan in queryset:
            row = self._serialize_plan_for_table(plan)
            sheet.append([row.get(column, "") for column in EXPORT_COLUMNS])

        self._style_sheet(sheet)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def export_csv(self, queryset) -> bytes:
        stream = io.StringIO()
        writer = csv.DictWriter(stream, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()

        for plan in queryset:
            writer.writerow(self._serialize_plan_for_table(plan))

        return stream.getvalue().encode("utf-8-sig")

    def export_json(self, queryset) -> bytes:
        payload = {
            "meta": {
                "exported_at": timezone.now().isoformat(),
                "total_plans": queryset.count() if hasattr(queryset, "count") else len(queryset),
                "highlight_limit": HIGHLIGHT_FEATURE_LIMIT,
            },
            "plans": [self._serialize_plan_for_json(plan) for plan in queryset],
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def sample_template_xlsx(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tarif import template"

        sheet.append(EXPORT_COLUMNS)
        for sample in SAMPLE_PLAN_DATA:
            row = self._serialize_sample_for_table(sample)
            sheet.append([row.get(column, "") for column in EXPORT_COLUMNS])

        self._style_sheet(sheet)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def sample_template_json(self) -> bytes:
        payload = {
            "meta": {
                "description": "Tarif import uchun sample JSON template",
                "highlight_limit": HIGHLIGHT_FEATURE_LIMIT,
            },
            "plans": [self._serialize_sample_for_json(sample) for sample in SAMPLE_PLAN_DATA],
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def _style_sheet(self, sheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, column in enumerate(sheet.columns, start=1):
            max_length = 12
            for cell in column:
                value = cell.value
                if value is None:
                    continue
                value_len = len(str(value))
                if value_len > max_length:
                    max_length = value_len
            sheet.column_dimensions[chr(64 + index) if index <= 26 else sheet.cell(1, index).column_letter].width = min(max_length + 3, 60)

    def _serialize_plan_for_table(self, plan: PricingPlan) -> dict[str, Any]:
        features = list(plan.features.all().order_by("order", "id"))
        highlight = features[:HIGHLIGHT_FEATURE_LIMIT]
        detail = features[HIGHLIGHT_FEATURE_LIMIT:]

        return {
            "id": plan.pk,
            "name": plan.name,
            "student_range": plan.student_range,
            "duration_months": plan.duration_months,
            "current_price": self._decimal_or_blank(plan.current_price),
            "old_price": self._decimal_or_blank(plan.old_price),
            "badge_text": plan.badge_text,
            "discount_label": plan.discount_label,
            "is_recommended": plan.is_recommended,
            "is_active": plan.is_active,
            "order": plan.order,
            "name_uz": plan.name_uz,
            "name_ru": plan.name_ru,
            "name_en": plan.name_en,
            "student_range_uz": plan.student_range_uz,
            "student_range_ru": plan.student_range_ru,
            "student_range_en": plan.student_range_en,
            "badge_text_uz": plan.badge_text_uz,
            "badge_text_ru": plan.badge_text_ru,
            "badge_text_en": plan.badge_text_en,
            "discount_label_uz": plan.discount_label_uz,
            "discount_label_ru": plan.discount_label_ru,
            "discount_label_en": plan.discount_label_en,
            "highlight_features": self._join_feature_strings(highlight, "text"),
            "detail_features": self._join_feature_strings(detail, "text"),
            "highlight_features_uz": self._join_feature_strings(highlight, "text_uz", fallback_attr="text"),
            "highlight_features_ru": self._join_feature_strings(highlight, "text_ru", fallback_attr="text"),
            "highlight_features_en": self._join_feature_strings(highlight, "text_en", fallback_attr="text"),
            "detail_features_uz": self._join_feature_strings(detail, "text_uz", fallback_attr="text"),
            "detail_features_ru": self._join_feature_strings(detail, "text_ru", fallback_attr="text"),
            "detail_features_en": self._join_feature_strings(detail, "text_en", fallback_attr="text"),
        }

    def _serialize_plan_for_json(self, plan: PricingPlan) -> dict[str, Any]:
        features = list(plan.features.all().order_by("order", "id"))
        highlight = features[:HIGHLIGHT_FEATURE_LIMIT]
        detail = features[HIGHLIGHT_FEATURE_LIMIT:]

        return {
            "id": plan.pk,
            "name": {
                "fallback": plan.name,
                "uz": plan.name_uz,
                "ru": plan.name_ru,
                "en": plan.name_en,
            },
            "student_range": {
                "fallback": plan.student_range,
                "uz": plan.student_range_uz,
                "ru": plan.student_range_ru,
                "en": plan.student_range_en,
            },
            "duration_months": plan.duration_months,
            "current_price": self._decimal_or_blank(plan.current_price),
            "old_price": self._decimal_or_blank(plan.old_price),
            "badge_text": {
                "fallback": plan.badge_text,
                "uz": plan.badge_text_uz,
                "ru": plan.badge_text_ru,
                "en": plan.badge_text_en,
            },
            "discount_label": {
                "fallback": plan.discount_label,
                "uz": plan.discount_label_uz,
                "ru": plan.discount_label_ru,
                "en": plan.discount_label_en,
            },
            "is_recommended": plan.is_recommended,
            "is_active": plan.is_active,
            "order": plan.order,
            "features": {
                "highlight": [self._feature_to_json_item(feature) for feature in highlight],
                "detail": [self._feature_to_json_item(feature) for feature in detail],
            },
        }

    def _serialize_sample_for_table(self, sample: dict[str, Any]) -> dict[str, Any]:
        highlight = sample.get("highlight_features", [])
        detail = sample.get("detail_features", [])

        return {
            "id": sample.get("id", ""),
            "name": sample.get("name", ""),
            "student_range": sample.get("student_range", ""),
            "duration_months": sample.get("duration_months", ""),
            "current_price": self._decimal_or_blank(sample.get("current_price")),
            "old_price": self._decimal_or_blank(sample.get("old_price")),
            "badge_text": sample.get("badge_text", ""),
            "discount_label": sample.get("discount_label", ""),
            "is_recommended": sample.get("is_recommended", False),
            "is_active": sample.get("is_active", True),
            "order": sample.get("order", 1),
            "name_uz": sample.get("name_uz", ""),
            "name_ru": sample.get("name_ru", ""),
            "name_en": sample.get("name_en", ""),
            "student_range_uz": sample.get("student_range_uz", ""),
            "student_range_ru": sample.get("student_range_ru", ""),
            "student_range_en": sample.get("student_range_en", ""),
            "badge_text_uz": sample.get("badge_text_uz", ""),
            "badge_text_ru": sample.get("badge_text_ru", ""),
            "badge_text_en": sample.get("badge_text_en", ""),
            "discount_label_uz": sample.get("discount_label_uz", ""),
            "discount_label_ru": sample.get("discount_label_ru", ""),
            "discount_label_en": sample.get("discount_label_en", ""),
            "highlight_features": self._join_feature_dicts(highlight, "text"),
            "detail_features": self._join_feature_dicts(detail, "text"),
            "highlight_features_uz": self._join_feature_dicts(highlight, "text_uz", fallback_key="text"),
            "highlight_features_ru": self._join_feature_dicts(highlight, "text_ru", fallback_key="text"),
            "highlight_features_en": self._join_feature_dicts(highlight, "text_en", fallback_key="text"),
            "detail_features_uz": self._join_feature_dicts(detail, "text_uz", fallback_key="text"),
            "detail_features_ru": self._join_feature_dicts(detail, "text_ru", fallback_key="text"),
            "detail_features_en": self._join_feature_dicts(detail, "text_en", fallback_key="text"),
        }

    def _serialize_sample_for_json(self, sample: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": sample.get("id") or None,
            "name": {
                "fallback": sample.get("name", ""),
                "uz": sample.get("name_uz", ""),
                "ru": sample.get("name_ru", ""),
                "en": sample.get("name_en", ""),
            },
            "student_range": {
                "fallback": sample.get("student_range", ""),
                "uz": sample.get("student_range_uz", ""),
                "ru": sample.get("student_range_ru", ""),
                "en": sample.get("student_range_en", ""),
            },
            "duration_months": sample.get("duration_months", 1),
            "current_price": self._decimal_or_blank(sample.get("current_price")),
            "old_price": self._decimal_or_blank(sample.get("old_price")),
            "badge_text": {
                "fallback": sample.get("badge_text", ""),
                "uz": sample.get("badge_text_uz", ""),
                "ru": sample.get("badge_text_ru", ""),
                "en": sample.get("badge_text_en", ""),
            },
            "discount_label": {
                "fallback": sample.get("discount_label", ""),
                "uz": sample.get("discount_label_uz", ""),
                "ru": sample.get("discount_label_ru", ""),
                "en": sample.get("discount_label_en", ""),
            },
            "is_recommended": sample.get("is_recommended", False),
            "is_active": sample.get("is_active", True),
            "order": sample.get("order", 1),
            "features": {
                "highlight": sample.get("highlight_features", []),
                "detail": sample.get("detail_features", []),
            },
        }

    def _join_feature_strings(self, features: list[PricingFeature], attr: str, fallback_attr: str | None = None) -> str:
        values: list[str] = []
        for feature in features:
            value = (getattr(feature, attr) or "").strip()
            if not value and fallback_attr:
                value = (getattr(feature, fallback_attr) or "").strip()
            values.append(value)

        while values and not values[-1]:
            values.pop()
        return " | ".join(values)

    def _join_feature_dicts(self, features: list[dict[str, str]], key: str, fallback_key: str | None = None) -> str:
        values: list[str] = []
        for feature in features:
            value = (feature.get(key) or "").strip()
            if not value and fallback_key:
                value = (feature.get(fallback_key) or "").strip()
            values.append(value)

        while values and not values[-1]:
            values.pop()
        return " | ".join(values)

    def _feature_to_json_item(self, feature: PricingFeature) -> dict[str, str]:
        return {
            "text": feature.text,
            "text_uz": feature.text_uz,
            "text_ru": feature.text_ru,
            "text_en": feature.text_en,
        }

    def _decimal_or_blank(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, Decimal):
            return format(value.normalize(), "f") if value == value.to_integral() else format(value, "f")
        return str(value)


class PricingPlanImportService:
    def __init__(self, *, actor=None):
        self.actor = actor

    def import_file(self, *, uploaded_file, mode: str, file_format: str = "auto") -> PricingPlanImportResult:
        filename = getattr(uploaded_file, "name", "import-file")
        resolved_format = self._resolve_file_format(filename=filename, selected_format=file_format)

        result = PricingPlanImportResult(
            filename=filename,
            file_format=resolved_format or "unknown",
            mode=mode,
        )

        if mode not in IMPORT_MODES:
            result.add_error(row="-", field="import_mode", message="Noto'g'ri import strategiyasi tanlandi.")
            return result

        if not resolved_format:
            result.add_error(
                row="-",
                field="file",
                message="Faqat .xlsx yoki .json fayl qo'llab-quvvatlanadi.",
            )
            return result

        raw_bytes = uploaded_file.read()
        if not raw_bytes:
            result.add_error(row="-", field="file", message="Yuklangan fayl bo'sh.")
            return result

        if resolved_format == "xlsx":
            rows = self._parse_xlsx(raw_bytes, result)
        else:
            rows = self._parse_json(raw_bytes, result)

        result.total_rows = len(rows)
        if result.has_errors:
            return result

        actions = self._build_actions(rows=rows, mode=mode, result=result)
        if result.has_errors:
            return result

        self._apply_actions(actions=actions, result=result)
        if result.has_errors:
            result.created_count = 0
            result.updated_count = 0
            result.skipped_count = 0
            return result

        actor_name = "anonymous"
        if self.actor and hasattr(self.actor, "get_username"):
            actor_name = self.actor.get_username() or str(self.actor.pk)

        logger.info(
            "Pricing plan import completed by=%s filename=%s format=%s mode=%s total=%s created=%s updated=%s skipped=%s",
            actor_name,
            result.filename,
            result.file_format,
            result.mode,
            result.total_rows,
            result.created_count,
            result.updated_count,
            result.skipped_count,
        )
        return result

    def _resolve_file_format(self, *, filename: str, selected_format: str) -> str | None:
        if selected_format and selected_format != "auto":
            return selected_format if selected_format in SUPPORTED_IMPORT_FORMATS else None

        lower_name = filename.lower()
        if lower_name.endswith(".xlsx"):
            return "xlsx"
        if lower_name.endswith(".json"):
            return "json"
        return None

    def _parse_xlsx(self, raw_bytes: bytes, result: PricingPlanImportResult) -> list[ParsedPlanRow]:
        try:
            workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as exc:
            result.add_error(row="-", field="file", message=f"Excel faylini o'qib bo'lmadi: {exc}")
            return []

        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            result.add_error(row="-", field="header", message="Excel header qatori topilmadi.")
            return []

        header_map: dict[str, int] = {}
        for index, raw_header in enumerate(header_row):
            normalized = self._normalize_column_name(raw_header)
            if not normalized:
                continue
            canonical = COLUMN_ALIASES.get(normalized, normalized)
            if canonical in header_map:
                result.add_error(row="1", field="header", message=f"Takrorlangan column: {canonical}")
                continue
            header_map[canonical] = index

        missing_columns = sorted(column for column in IMPORT_REQUIRED_COLUMNS if column not in header_map)
        if missing_columns:
            result.add_error(
                row="1",
                field="header",
                message=f"Majburiy ustun(lar) yo'q: {', '.join(missing_columns)}",
            )
            return []

        parsed_rows: list[ParsedPlanRow] = []
        for excel_row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_payload = {
                field_name: values[column_index] if column_index < len(values) else None
                for field_name, column_index in header_map.items()
            }
            if self._is_empty_row(row_payload.values()):
                continue

            parsed = self._parse_plan_row(row_payload=row_payload, row_label=str(excel_row_index), result=result)
            if parsed:
                parsed_rows.append(parsed)

        return parsed_rows

    def _parse_json(self, raw_bytes: bytes, result: PricingPlanImportResult) -> list[ParsedPlanRow]:
        try:
            decoded = raw_bytes.decode("utf-8-sig")
            payload = json.loads(decoded)
        except Exception as exc:
            result.add_error(row="-", field="file", message=f"JSON fayl noto'g'ri: {exc}")
            return []

        if isinstance(payload, dict):
            plans = payload.get("plans")
        else:
            plans = payload

        if not isinstance(plans, list):
            result.add_error(row="-", field="plans", message="JSON formatda `plans` ro'yxati bo'lishi kerak.")
            return []

        parsed_rows: list[ParsedPlanRow] = []
        for index, item in enumerate(plans, start=1):
            if not isinstance(item, dict):
                result.add_error(row=str(index), field="row", message="Har bir element object bo'lishi kerak.")
                continue

            normalized_item = self._normalize_json_item(item)
            parsed = self._parse_plan_row(
                row_payload=normalized_item,
                row_label=str(index),
                result=result,
            )
            if parsed:
                parsed_rows.append(parsed)

        return parsed_rows

    def _normalize_json_item(self, item: dict[str, Any]) -> dict[str, Any]:
        normalized: dict[str, Any] = {}

        for raw_key, value in item.items():
            canonical_key = COLUMN_ALIASES.get(self._normalize_column_name(raw_key), self._normalize_column_name(raw_key))
            if canonical_key:
                normalized[canonical_key] = value

        for field_name in ("name", "student_range", "badge_text", "discount_label"):
            value = normalized.get(field_name)
            if isinstance(value, dict):
                normalized[field_name] = value.get("fallback") or value.get("base") or ""
                normalized[f"{field_name}_uz"] = value.get("uz") or ""
                normalized[f"{field_name}_ru"] = value.get("ru") or ""
                normalized[f"{field_name}_en"] = value.get("en") or ""

        features = normalized.get("features")
        if isinstance(features, dict):
            normalized["highlight_features"] = features.get("highlight") or features.get("highlight_features") or []
            normalized["detail_features"] = features.get("detail") or features.get("detail_features") or []
        elif isinstance(features, list):
            normalized["highlight_features"] = features

        return normalized

    def _parse_plan_row(
        self,
        *,
        row_payload: dict[str, Any],
        row_label: str,
        result: PricingPlanImportResult,
    ) -> ParsedPlanRow | None:
        row_errors: list[ImportErrorRow] = []

        import_id = self._parse_optional_positive_int(
            value=row_payload.get("id"),
            field="id",
            row_label=row_label,
            row_errors=row_errors,
        )

        fields: dict[str, Any] = {}

        for field_name in PLAN_STRING_FIELDS:
            fields[field_name] = self._parse_string_field(
                value=row_payload.get(field_name),
                field=field_name,
                row_label=row_label,
                row_errors=row_errors,
                required=field_name in {"name", "student_range"},
            )

        fields["duration_months"] = self._parse_positive_int(
            value=row_payload.get("duration_months"),
            field="duration_months",
            row_label=row_label,
            row_errors=row_errors,
            required=True,
            default=1,
        )

        fields["current_price"] = self._parse_decimal(
            value=row_payload.get("current_price"),
            field="current_price",
            row_label=row_label,
            row_errors=row_errors,
            required=True,
        )

        fields["old_price"] = self._parse_decimal(
            value=row_payload.get("old_price"),
            field="old_price",
            row_label=row_label,
            row_errors=row_errors,
            required=False,
        )

        fields["is_recommended"] = self._parse_bool(
            value=row_payload.get("is_recommended"),
            field="is_recommended",
            row_label=row_label,
            row_errors=row_errors,
            default=False,
        )
        fields["is_active"] = self._parse_bool(
            value=row_payload.get("is_active"),
            field="is_active",
            row_label=row_label,
            row_errors=row_errors,
            default=True,
        )

        fields["order"] = self._parse_positive_int(
            value=row_payload.get("order"),
            field="order",
            row_label=row_label,
            row_errors=row_errors,
            required=False,
            default=1,
        )

        current_price = fields.get("current_price")
        old_price = fields.get("old_price")
        if current_price is not None and current_price <= 0:
            row_errors.append(ImportErrorRow(row=row_label, field="current_price", message="0 dan katta bo'lishi kerak."))
        if old_price is not None and old_price < 0:
            row_errors.append(ImportErrorRow(row=row_label, field="old_price", message="Manfiy bo'lishi mumkin emas."))

        features = self._parse_features(
            row_payload=row_payload,
            row_label=row_label,
            row_errors=row_errors,
        )

        if row_errors:
            result.errors.extend(row_errors)
            return None

        return ParsedPlanRow(
            row_label=row_label,
            import_id=import_id,
            fields=fields,
            features=features,
        )

    def _parse_features(
        self,
        *,
        row_payload: dict[str, Any],
        row_label: str,
        row_errors: list[ImportErrorRow],
    ) -> list[dict[str, str]]:
        highlight_rows = self._build_feature_rows(
            base_value=row_payload.get("highlight_features"),
            uz_value=row_payload.get("highlight_features_uz"),
            ru_value=row_payload.get("highlight_features_ru"),
            en_value=row_payload.get("highlight_features_en"),
            row_label=row_label,
            field_prefix="highlight_features",
            row_errors=row_errors,
        )
        detail_rows = self._build_feature_rows(
            base_value=row_payload.get("detail_features"),
            uz_value=row_payload.get("detail_features_uz"),
            ru_value=row_payload.get("detail_features_ru"),
            en_value=row_payload.get("detail_features_en"),
            row_label=row_label,
            field_prefix="detail_features",
            row_errors=row_errors,
        )
        return highlight_rows + detail_rows

    def _build_feature_rows(
        self,
        *,
        base_value: Any,
        uz_value: Any,
        ru_value: Any,
        en_value: Any,
        row_label: str,
        field_prefix: str,
        row_errors: list[ImportErrorRow],
    ) -> list[dict[str, str]]:
        base_items = self._parse_feature_value(base_value)

        if base_items and any(isinstance(item, dict) for item in base_items):
            normalized_rows = []
            for item in base_items:
                if isinstance(item, dict):
                    text = self._stringify(item.get("text") or item.get("fallback"))
                    text_uz = self._stringify(item.get("text_uz") or item.get("uz"))
                    text_ru = self._stringify(item.get("text_ru") or item.get("ru"))
                    text_en = self._stringify(item.get("text_en") or item.get("en"))
                else:
                    text = self._stringify(item)
                    text_uz = ""
                    text_ru = ""
                    text_en = ""

                normalized = self._normalize_feature_row(
                    text=text,
                    text_uz=text_uz,
                    text_ru=text_ru,
                    text_en=text_en,
                    row_label=row_label,
                    field_prefix=field_prefix,
                    row_errors=row_errors,
                )
                if normalized:
                    normalized_rows.append(normalized)
            return normalized_rows

        base_list = self._parse_plain_feature_list(base_value)
        uz_list = self._parse_plain_feature_list(uz_value)
        ru_list = self._parse_plain_feature_list(ru_value)
        en_list = self._parse_plain_feature_list(en_value)

        max_len = max(len(base_list), len(uz_list), len(ru_list), len(en_list))
        rows: list[dict[str, str]] = []
        for index in range(max_len):
            text = base_list[index] if index < len(base_list) else ""
            text_uz = uz_list[index] if index < len(uz_list) else ""
            text_ru = ru_list[index] if index < len(ru_list) else ""
            text_en = en_list[index] if index < len(en_list) else ""

            normalized = self._normalize_feature_row(
                text=text,
                text_uz=text_uz,
                text_ru=text_ru,
                text_en=text_en,
                row_label=row_label,
                field_prefix=field_prefix,
                row_errors=row_errors,
            )
            if normalized:
                rows.append(normalized)

        return rows

    def _normalize_feature_row(
        self,
        *,
        text: str,
        text_uz: str,
        text_ru: str,
        text_en: str,
        row_label: str,
        field_prefix: str,
        row_errors: list[ImportErrorRow],
    ) -> dict[str, str] | None:
        text = (text or "").strip()
        text_uz = (text_uz or "").strip()
        text_ru = (text_ru or "").strip()
        text_en = (text_en or "").strip()

        if not any((text, text_uz, text_ru, text_en)):
            return None

        if not text:
            text = text_uz or text_ru or text_en

        for field_name, value in {
            field_prefix: text,
            f"{field_prefix}_uz": text_uz,
            f"{field_prefix}_ru": text_ru,
            f"{field_prefix}_en": text_en,
        }.items():
            if value and len(value) > FEATURE_TEXT_MAX_LENGTH:
                row_errors.append(
                    ImportErrorRow(
                        row=row_label,
                        field=field_name,
                        message=f"{FEATURE_TEXT_MAX_LENGTH} belgidan oshmasligi kerak.",
                    )
                )

        return {
            "text": text,
            "text_uz": text_uz,
            "text_ru": text_ru,
            "text_en": text_en,
        }

    def _build_actions(
        self,
        *,
        rows: list[ParsedPlanRow],
        mode: str,
        result: PricingPlanImportResult,
    ) -> list[dict[str, Any]]:
        actions: list[dict[str, Any]] = []
        seen_row_keys: dict[str, str] = {}

        for row in rows:
            row_key = self._row_key(row)
            if row_key in seen_row_keys:
                result.add_error(
                    row=row.row_label,
                    field="duplicate",
                    message=f"Fayl ichida dublikat yozuv. Avvalgi qator: {seen_row_keys[row_key]}",
                )
            else:
                seen_row_keys[row_key] = row.row_label

        if result.has_errors:
            return []

        for row in rows:
            existing, find_error = self._find_existing_plan(row)
            if find_error:
                result.add_error(row=row.row_label, field="lookup", message=find_error)
                continue

            if mode == IMPORT_MODE_CREATE_ONLY:
                if existing:
                    result.add_error(
                        row=row.row_label,
                        field="duplicate",
                        message="Mavjud tarif topildi. `Hammasini yangi qo'shish` rejimida bu qatordan import qilib bo'lmaydi.",
                    )
                    continue
                actions.append({"action": "create", "row": row, "existing": None})
                continue

            if mode == IMPORT_MODE_SKIP_DUPLICATES:
                if existing:
                    actions.append({"action": "skip", "row": row, "existing": existing})
                else:
                    actions.append({"action": "create", "row": row, "existing": None})
                continue

            if mode == IMPORT_MODE_UPDATE_EXISTING:
                if existing:
                    actions.append({"action": "update", "row": row, "existing": existing})
                else:
                    actions.append({"action": "create", "row": row, "existing": None})
                continue

        return actions

    def _apply_actions(self, *, actions: list[dict[str, Any]], result: PricingPlanImportResult) -> None:
        try:
            with transaction.atomic():
                for item in actions:
                    action = item["action"]
                    row: ParsedPlanRow = item["row"]
                    existing: PricingPlan | None = item.get("existing")

                    if action == "skip":
                        result.skipped_count += 1
                        continue

                    plan = existing or PricingPlan()
                    for field_name, value in row.fields.items():
                        setattr(plan, field_name, value)

                    try:
                        plan.full_clean()
                    except ValidationError as exc:
                        self._append_validation_errors(
                            result=result,
                            row_label=row.row_label,
                            validation_error=exc,
                        )
                        raise

                    plan.save()
                    self._replace_plan_features(plan=plan, features=row.features)

                    if action == "update":
                        result.updated_count += 1
                    else:
                        result.created_count += 1
        except ValidationError:
            return
        except Exception as exc:
            result.add_error(row="-", field="server", message=f"Import jarayonida xatolik: {exc}")

    def _replace_plan_features(self, *, plan: PricingPlan, features: list[dict[str, str]]) -> None:
        plan.features.all().delete()
        if not features:
            return

        feature_objects = [
            PricingFeature(
                pricing_plan=plan,
                text=feature["text"],
                text_uz=feature.get("text_uz", ""),
                text_ru=feature.get("text_ru", ""),
                text_en=feature.get("text_en", ""),
                order=index,
            )
            for index, feature in enumerate(features, start=1)
        ]
        PricingFeature.objects.bulk_create(feature_objects)

    def _find_existing_plan(self, row: ParsedPlanRow) -> tuple[PricingPlan | None, str | None]:
        if row.import_id:
            by_id = PricingPlan.objects.filter(pk=row.import_id).first()
            if by_id:
                return by_id, None

        name = row.fields.get("name", "")
        duration = row.fields.get("duration_months")
        if not name or not duration:
            return None, None

        queryset = PricingPlan.objects.filter(name__iexact=name, duration_months=duration).order_by("id")
        candidate_count = queryset.count()
        if candidate_count > 1:
            return None, "Bir xil name + duration bo'yicha bir nechta tarif topildi. id bilan import qiling."
        return queryset.first(), None

    def _append_validation_errors(
        self,
        *,
        result: PricingPlanImportResult,
        row_label: str,
        validation_error: ValidationError,
    ) -> None:
        if hasattr(validation_error, "message_dict"):
            for field_name, messages in validation_error.message_dict.items():
                for message in messages:
                    result.add_error(row=row_label, field=field_name, message=str(message))
            return

        for message in validation_error.messages:
            result.add_error(row=row_label, field="row", message=str(message))

    def _row_key(self, row: ParsedPlanRow) -> str:
        if row.import_id:
            return f"id:{row.import_id}"

        name = (row.fields.get("name") or "").strip().lower()
        duration = row.fields.get("duration_months")
        return f"name:{name}|duration:{duration}"

    def _parse_string_field(
        self,
        *,
        value: Any,
        field: str,
        row_label: str,
        row_errors: list[ImportErrorRow],
        required: bool,
    ) -> str:
        parsed = self._stringify(value)
        if required and not parsed:
            row_errors.append(ImportErrorRow(row=row_label, field=field, message="Majburiy maydon bo'sh."))
            return ""

        max_length = PLAN_FIELD_MAX_LENGTH.get(field)
        if max_length and parsed and len(parsed) > max_length:
            row_errors.append(
                ImportErrorRow(
                    row=row_label,
                    field=field,
                    message=f"{max_length} belgidan oshmasligi kerak.",
                )
            )
        return parsed

    def _parse_bool(
        self,
        *,
        value: Any,
        field: str,
        row_label: str,
        row_errors: list[ImportErrorRow],
        default: bool,
    ) -> bool:
        if value in (None, ""):
            return default

        if isinstance(value, bool):
            return value

        normalized = self._stringify(value).lower()
        if normalized in TRUE_VALUES:
            return True
        if normalized in FALSE_VALUES:
            return False

        row_errors.append(
            ImportErrorRow(
                row=row_label,
                field=field,
                message="Boolean qiymat bo'lishi kerak (true/false, 1/0, ha/yo'q).",
            )
        )
        return default

    def _parse_optional_positive_int(
        self,
        *,
        value: Any,
        field: str,
        row_label: str,
        row_errors: list[ImportErrorRow],
    ) -> int | None:
        if value in (None, ""):
            return None

        parsed = self._parse_positive_int(
            value=value,
            field=field,
            row_label=row_label,
            row_errors=row_errors,
            required=False,
            default=1,
        )
        return parsed

    def _parse_positive_int(
        self,
        *,
        value: Any,
        field: str,
        row_label: str,
        row_errors: list[ImportErrorRow],
        required: bool,
        default: int,
    ) -> int:
        if value in (None, ""):
            if required:
                row_errors.append(ImportErrorRow(row=row_label, field=field, message="Majburiy maydon bo'sh."))
            return default

        parsed: int | None = None
        if isinstance(value, int):
            parsed = value
        elif isinstance(value, float):
            if value.is_integer():
                parsed = int(value)
        else:
            normalized = self._stringify(value)
            if re.fullmatch(r"\d+", normalized):
                parsed = int(normalized)

        if parsed is None or parsed <= 0:
            row_errors.append(
                ImportErrorRow(row=row_label, field=field, message="Musbat butun son bo'lishi kerak.")
            )
            return default

        return parsed

    def _parse_decimal(
        self,
        *,
        value: Any,
        field: str,
        row_label: str,
        row_errors: list[ImportErrorRow],
        required: bool,
    ) -> Decimal | None:
        if value in (None, ""):
            if required:
                row_errors.append(ImportErrorRow(row=row_label, field=field, message="Majburiy maydon bo'sh."))
            return None

        if isinstance(value, Decimal):
            return value

        normalized = self._normalize_decimal_text(value)
        try:
            return Decimal(normalized)
        except (InvalidOperation, TypeError):
            row_errors.append(ImportErrorRow(row=row_label, field=field, message="Raqamli qiymat kiriting."))
            return None

    def _normalize_decimal_text(self, value: Any) -> str:
        text = self._stringify(value)
        text = text.replace(" ", "").replace("\u00a0", "")

        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            tail = text.rsplit(",", maxsplit=1)[-1]
            if tail.isdigit() and len(tail) <= 2:
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")

        return text

    def _parse_feature_value(self, value: Any) -> list[Any]:
        if value in (None, ""):
            return []
        if isinstance(value, (list, tuple)):
            return list(value)
        if isinstance(value, str):
            chunks = re.split(r"\||\n", value)
            return [chunk.strip() for chunk in chunks if chunk.strip()]
        return [value]

    def _parse_plain_feature_list(self, value: Any) -> list[str]:
        items = self._parse_feature_value(value)
        normalized: list[str] = []
        for item in items:
            if isinstance(item, dict):
                normalized.append(self._stringify(item.get("text") or item.get("fallback")))
            else:
                normalized.append(self._stringify(item))
        return normalized

    def _stringify(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, Decimal):
            return format(value, "f")
        return str(value).strip()

    def _normalize_column_name(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        text = str(value).strip().lower()
        text = text.replace("-", "_")
        text = re.sub(r"\s+", "_", text)
        return text

    def _is_empty_row(self, values) -> bool:
        for value in values:
            if self._stringify(value):
                return False
        return True


__all__ = [
    "EXPORT_COLUMNS",
    "HIGHLIGHT_FEATURE_LIMIT",
    "IMPORT_MODE_CREATE_ONLY",
    "IMPORT_MODE_SKIP_DUPLICATES",
    "IMPORT_MODE_UPDATE_EXISTING",
    "PricingPlanExportService",
    "PricingPlanImportService",
    "PricingPlanImportResult",
]
