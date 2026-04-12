from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from accounts.models import Center, User
from education.models import Enrollment, Group, Payment, PaymentAllocation, TuitionMonth


class PaymentExcelExportTests(TestCase):
    def setUp(self):
        self.center = Center.objects.create(name="Export Center", slug="export-center")
        self.manager = User.objects.create_user(
            email="manager@export.test",
            password="testpass123",
            role="manager",
            center=self.center,
            ism="Export",
            familya="Manager",
        )
        self.teacher = User.objects.create_user(
            email="teacher@export.test",
            password="testpass123",
            role="teacher",
            center=self.center,
            ism="Export",
            familya="Teacher",
        )
        self.group = Group.objects.create(
            center=self.center,
            nom="IELTS Export",
            oqituvchi=self.teacher,
            kurs_narxi=300_000,
            oqituvchi_foiz=40,
            oy_dars_soni=12,
        )

        self.alpha = User.objects.create_user(
            email="alpha@export.test",
            password="testpass123",
            role="student",
            center=self.center,
            ism="Alpha",
            familya="Export",
            telefon1="+998901234500",
        )
        self.beta = User.objects.create_user(
            email="beta@export.test",
            password="testpass123",
            role="student",
            center=self.center,
            ism="Beta",
            familya="Export",
            telefon1="+998901234501",
        )

        self.alpha_enrollment = Enrollment.objects.create(
            group=self.group,
            student=self.alpha,
            center=self.center,
            kurs_narhi=300_000,
            oqituvchi_foiz=40,
        )
        beta_enrollment = Enrollment.objects.create(
            group=self.group,
            student=self.beta,
            center=self.center,
            kurs_narhi=300_000,
            oqituvchi_foiz=40,
        )

        april_alpha = TuitionMonth.objects.create(
            enrollment=self.alpha_enrollment,
            center=self.center,
            month=date(2026, 4, 1),
            fee_amount=300_000,
        )
        april_beta = TuitionMonth.objects.create(
            enrollment=beta_enrollment,
            center=self.center,
            month=date(2026, 4, 1),
            fee_amount=300_000,
        )

        first_payment = Payment.objects.create(
            enrollment=self.alpha_enrollment,
            center=self.center,
            student=self.alpha,
            group=self.group,
            payment_type="cash",
            cash_amount=100_000,
            paid_date=date(2026, 4, 5),
            created_by=self.manager,
            note="Birinchi to'lov",
        )
        PaymentAllocation.objects.create(
            center=self.center,
            payment=first_payment,
            tuition_month=april_alpha,
            amount=100_000,
        )

        second_payment = Payment.objects.create(
            enrollment=self.alpha_enrollment,
            center=self.center,
            student=self.alpha,
            group=self.group,
            payment_type="card",
            card_amount=150_000,
            paid_date=date(2026, 4, 10),
            created_by=self.manager,
            note="Ikkinchi to'lov",
        )
        PaymentAllocation.objects.create(
            center=self.center,
            payment=second_payment,
            tuition_month=april_alpha,
            amount=150_000,
        )

        beta_payment = Payment.objects.create(
            enrollment=beta_enrollment,
            center=self.center,
            student=self.beta,
            group=self.group,
            payment_type="cash",
            cash_amount=300_000,
            paid_date=date(2026, 4, 12),
            created_by=self.manager,
        )
        PaymentAllocation.objects.create(
            center=self.center,
            payment=beta_payment,
            tuition_month=april_beta,
            amount=300_000,
        )

        self.client.force_login(self.manager)
        self.export_url = f"/{self.center.slug}{reverse('education:payment_export_xlsx')}"

    def test_export_respects_filters_and_contains_summary_and_details(self):
        response = self.client.get(
            self.export_url,
            {
                "q": "Alpha",
                "date_from": "2026-04-01",
                "date_to": "2026-04-30",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("tolovlar_export_2026-04-01_2026-04-30.xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Umumiy ro'yxat", "To'lov yozuvlari"])

        summary_sheet = workbook["Umumiy ro'yxat"]
        self.assertEqual(summary_sheet["A1"].value, "To'lovlar eksporti")
        self.assertEqual(summary_sheet["B6"].value, "1 ta")
        self.assertEqual(summary_sheet["B7"].value, "2 ta")
        self.assertEqual(summary_sheet["J17"].value, 250000)
        self.assertEqual(summary_sheet["B17"].value, self.alpha.get_full_name())
        self.assertEqual(summary_sheet["G17"].value, 2)

        detail_sheet = workbook["To'lov yozuvlari"]
        self.assertEqual(detail_sheet["A1"].value, "To'lov yozuvlari")
        self.assertEqual(detail_sheet.max_row, 6)
        self.assertEqual(detail_sheet["C5"].value, self.alpha.get_full_name())
        self.assertEqual(detail_sheet["J5"].value, 150000)
        self.assertEqual(detail_sheet["K5"].value, "Karta")
        self.assertEqual(detail_sheet["C6"].value, self.alpha.get_full_name())
