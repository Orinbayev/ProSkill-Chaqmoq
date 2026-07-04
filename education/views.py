from __future__ import annotations

import calendar
import json
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
# from multiprocessing import Value 
from django.db import models

logger = logging.getLogger(__name__)

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from billing.decorators import require_feature
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import (
    Avg, Count, F, Min, Max, Prefetch, Q, Sum, OuterRef, Subquery
)
from django.db.models.functions import Coalesce, TruncMonth, Cast
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
)
from django.core.exceptions import PermissionDenied, ValidationError
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from django.utils.timezone import localdate, make_aware
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from datetime import datetime
from education.services.tuition import (
    parse_month_str,
    month_first_day,
    month_last_day,
    ensure_tuition_month,
    get_month_paid,
    full_course_amount,
    effective_student_payable_amount,
    calculate_enrollment_debt_snapshots,
    month_range_starts,
    tuition_month_fee_field,
    tuition_month_fee,
    ensure_all_tuition_months_since_start,
    create_payment_and_allocate,
    update_payment_and_reallocate,
    _allocate_amount_forward,
    infer_payment_type,
    prorated_monthly_fee,
    attendance_based_fee,
    billable_attendance_count,
    expected_lessons_in_period,
    enrollment_start_date,
    enrollment_lesson_pattern,
    enrollment_month_financial_snapshot,
    format_money,
    is_month_closed_for_center,
    normalize_lesson_pattern,
    round_money_to_thousand,
    round_div,
    resolve_lesson_schedule,
    tuition_month_preview,
    tuition_amount_breakdown,
    lesson_pattern_label,
    lesson_pattern_hint,
    preload_enrollment_history_starts,
)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors

from accounts.models import User
from chaqmoq.models import Ledger, Rule
from .forms import CenterExpenseForm, GroupForm, ITGroupForm, LangGroupForm, StudentGroupTransferForm
from .models import (
    Attendance,
    Category,
    CourseTemplate,
    CenterExpense,
    DailyLightningRecord,
    Dars,
    Enrollment,
    Group,
    OylikHisobot,
    Payment,
    PaymentAllocation,
    Student,
    TuitionMonth,
    StudentGroupHistory,
    FinancialMonth,
    MonthlyFinanceSnapshot,
    TeacherSalarySnapshot,
    TeacherIncome,
)
from education.services.historical_finance_service import HistoricalFinanceService
from education.services.enrollment_service import EnrollmentService
from education.services.student_transfer import transfer_student_to_group, user_can_transfer_student
from education.services.lesson_planning import calculate_lessons, validate_remaining_lessons
from accounts.models import Center
from .permissions import user_can_manage_payments
from django.db import transaction
from django.db.models.functions import ExtractYear, ExtractMonth, ExtractDay  # student_detail dagi underline ham yo'qoladi
from urllib.parse import urlparse, parse_qs
from django.db import transaction
from urllib.parse import urlencode, urlparse, parse_qs, unquote
from django.urls import reverse  # sizda reverse ishlatyapsiz, import yo'q
from django.db import transaction
from django.db.models import Sum, F, Value as DJValue
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from decimal import Decimal

from core.center_features import (
    FEATURE_UI_CERTIFICATES,
    FEATURE_UI_EXAM_SESSIONS,
    FEATURE_UI_FAILED_STUDENTS,
    FEATURE_UI_WEEKLY_SCHEDULE,
    center_ui_feature_enabled,
)

U = get_user_model()

DAILY_LIMIT = 50  # (hozircha ishlatilmayapti, lekin qoldirdim)


def _redirect_disabled_module(request, *, message: str):
    messages.warning(request, message)
    return redirect("core:home")


def _ensure_center_ui_feature(request, center, feature_code: str, *, message: str):
    if center_ui_feature_enabled(center, feature_code):
        return None
    return _redirect_disabled_module(request, message=message)


def _day_range(d):
    start = make_aware(datetime.combine(d, datetime.min.time()))
    end = make_aware(datetime.combine(d + timedelta(days=1), datetime.min.time()))
    return start, end


def _accumulate_daily_lightning(*, group, student, date_value, points_delta):
    record, _ = DailyLightningRecord.objects.get_or_create(
        group=group,
        student=student,
        date=date_value,
        defaults={
            "center": getattr(group, "center", None),
            "plus_points": 0,
            "minus_points": 0,
        },
    )
    if points_delta > 0:
        record.plus_points = int(record.plus_points or 0) + int(points_delta)
    elif points_delta < 0:
        record.minus_points = int(record.minus_points or 0) + abs(int(points_delta))
    
    if not record.center_id and getattr(group, "center_id", None):
        record.center = group.center
        
    record.save(update_fields=["plus_points", "minus_points", "center", "updated_at"])
    return record


def _attendance_adjust_rule():
    """
    Davomat OFF bo'lganda, o'sha kundagi ballarni 'bekor qilish' uchun
    maxsus Rule kerak bo'ladi. (DBga 1 marta tushadi)
    """
    rule, _ = Rule.objects.get_or_create(
        nom="Davomat bekor qilindi",
        defaults={"tur": Rule.MINUS, "min_baho": 1, "max_baho": 10000},
    )
    return rule


def parse_month_yyyy_mm(s: str):
    # '2026-01' -> date(2026, 1, 1)
    try:
        y, m = s.split("-")
        y = int(y); m = int(m)
        if 1 <= m <= 12:
            return date(y, m, 1)
    except Exception:
        return None
    return None

def first_day_of_current_month():
    d = timezone.localdate()
    return date(d.year, d.month, 1)


def _parse_int_value(value, default=None):
    try:
        if value in (None, "", "None"):
            return default
        return int(str(value).replace(" ", "").replace(",", ""))
    except (TypeError, ValueError):
        return default


def _parse_bool_value(value) -> bool:
    return str(value or "").strip().lower() in {"on", "1", "true", "yes"}


def _preview_month_for_start_date(start_date: date | None, fallback: date | None = None) -> date:
    if start_date:
        return month_first_day(start_date)
    return month_first_day(fallback or timezone.localdate())


def _lesson_pattern_options():
    return [
        {
            "value": pattern,
            "label": lesson_pattern_label(pattern),
            "hint": lesson_pattern_hint(pattern),
        }
        for pattern in (
            Enrollment.LESSON_PATTERN_ODD,
            Enrollment.LESSON_PATTERN_EVEN,
            Enrollment.LESSON_PATTERN_DAILY,
        )
    ]


def _serialize_lesson_plan(plan: dict) -> dict:
    return {
        "requested_start_date": plan["requested_start_date"].isoformat(),
        "start_date": plan["start_date"].isoformat(),
        "calculation_start_date": plan["calculation_start_date"].isoformat(),
        "remaining_lessons": int(plan["remaining_lessons"] or 0),
        "lesson_pattern": plan["lesson_pattern"],
        "lesson_pattern_label": plan["lesson_pattern_label"],
        "lesson_pattern_hint": plan["lesson_pattern_hint"],
        "lesson_dates": [lesson_date.isoformat() for lesson_date in plan.get("lesson_dates", [])],
        "lesson_date_labels": list(plan.get("lesson_date_labels", [])),
        "last_lesson_date": plan["last_lesson_date"].isoformat() if plan.get("last_lesson_date") else "",
        "last_lesson_date_label": plan.get("last_lesson_date_label", "—"),
        "calculation_note": plan.get("calculation_note", ""),
        "used_reference_date": bool(plan.get("used_reference_date")),
    }

def _student_enrollment_catalog(enrollments: list[Enrollment]) -> dict:
    enrollment_rows: list[dict] = []

    for enrollment in enrollments:
        enrollment_rows.append(
            {
                "enrollment_id": enrollment.id,
                "group_id": enrollment.group_id,
                "group_name": getattr(enrollment.group, "nom", "—"),
                "joined_at": enrollment.joined_at.isoformat() if enrollment.joined_at else "",
                "stored_lesson_pattern": getattr(enrollment, "lesson_pattern", Enrollment.LESSON_PATTERN_GROUP),
                "resolved_lesson_pattern": enrollment_lesson_pattern(enrollment),
                "monthly_lessons": int(getattr(enrollment.group, "oy_dars_soni", 0) or getattr(enrollment, "monthly_lessons", 0) or 12),
                "course_price": int(getattr(enrollment, "kurs_narhi", 0)),
                "teacher_percent": int(getattr(enrollment, "oqituvchi_foiz", 0) or getattr(enrollment.group, "oqituvchi_foiz", 0) or 0),
                "student_payable_amount": getattr(enrollment, "student_payable_amount", None),
                "remaining_lessons_override": getattr(enrollment, "remaining_lessons_override", None),
                "last_lesson_date": enrollment.last_lesson_date.isoformat() if getattr(enrollment, "last_lesson_date", None) else "",
            }
        )

    enrollment_rows.sort(
        key=lambda item: (
            (item.get("group_name") or "").lower(),
            int(item.get("enrollment_id") or 0),
        )
    )
    return {
        "enrollments": enrollment_rows,
    }


def get_student_total_debt(student, center=None) -> int:
    from django.db.models import Q as _Q
    from django.utils import timezone
    from education.models import Enrollment, TuitionMonth
    _center_q = (
        _Q(center=center)
        | _Q(center__isnull=True, group__center=center)
        | _Q(center__isnull=True, student__center=center)
    )
    active_enrs = list(
        Enrollment.objects.filter(
            student=student,
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).filter(_center_q)
    )
    _inactive_enr_ids = (
        TuitionMonth.objects
        .filter(
            enrollment__student=student,
            enrollment__is_active=False,
            is_deleted=False,
            enrollment__student__is_archived=False,
            enrollment__group__is_archived=False,
            enrollment__group__is_deleted=False,
        )
        .values_list("enrollment_id", flat=True)
        .distinct()
    )
    inactive_enrs = list(
        Enrollment.objects.filter(
            id__in=_inactive_enr_ids,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).filter(_center_q)
    )
    all_enrs = active_enrs + inactive_enrs
    if not all_enrs:
        return 0

    today = timezone.localdate()
    selected_from = today.replace(day=1)
    selected_to = today
    from education.services.tuition import month_range_starts, calculate_enrollment_debt_snapshots
    period_months = month_range_starts(selected_from, selected_to)

    snapshots = calculate_enrollment_debt_snapshots(
        all_enrs, period_months, cumulative_up_to=selected_to
    )
    total_debt = 0
    for snap in snapshots.values():
        total_debt += int(snap.get("debt", 0) or 0)
        total_debt += int(snap.get("previous_unpaid", 0) or 0)
    return total_debt




def _serialize_tuition_preview(preview: dict) -> dict:
    period_end = preview.get("period_end_date") or month_last_day(preview["month"])
    return {
        "month": preview["month"].isoformat(),
        "start_date": preview["start_date"].isoformat(),
        "period_end_date": period_end.isoformat(),
        "period_end_date_label": period_end.strftime("%d.%m.%Y"),
        "lesson_pattern": preview["lesson_pattern"],
        "lesson_pattern_label": preview["lesson_pattern_label"],
        "lesson_pattern_hint": preview.get("lesson_pattern_hint", ""),
        "monthly_lessons": int(preview["monthly_lessons"] or 0),
        "lesson_count": int(preview["lesson_count"] or 0),
        "per_lesson_amount": int(preview.get("per_lesson_amount", 0) or 0),
        "per_lesson_amount_display": preview.get("per_lesson_amount_display", format_money(0)),
        "counted_weekdays": [int(weekday) for weekday in preview.get("counted_weekdays", [])],
        "counted_weekday_labels": list(preview.get("counted_weekday_labels", [])),
        "counted_weekday_short_labels": list(preview.get("counted_weekday_short_labels", [])),
        "counted_days_text": preview.get("counted_days_text", ""),
        "counted_days_summary": preview.get("counted_days_summary", ""),
        "lesson_count_summary": preview.get("lesson_count_summary", ""),
        "lesson_dates": [lesson_date.isoformat() for lesson_date in preview.get("lesson_dates", [])],
        "lesson_date_labels": list(preview.get("lesson_date_labels", [])),
        "fee_amount": int(preview["fee_amount"] or 0),
        "teacher_share": int(preview["teacher_share"] or 0),
        "center_share": int(preview["center_share"] or 0),
        "full_turnover": int(preview["full_turnover"] or 0),
        "month_label_uz": preview.get("month_label_uz", ""),
        "debt_label_uz": preview.get("debt_label_uz", ""),
    }


def _format_money_exact(amount: int | float | None) -> str:
    return f"{int(amount or 0):,}".replace(",", " ") + " so'm"


def _apply_period_end_to_preview(preview: dict, period_end: date | None) -> dict:
    if not period_end:
        return preview
    adjusted = dict(preview)
    lesson_dates = [
        lesson_date
        for lesson_date in preview.get("lesson_dates", [])
        if lesson_date <= period_end
    ]
    adjusted.update(
        {
            "period_end_date": period_end,
            "lesson_dates": lesson_dates,
            "lesson_date_labels": [lesson_date.strftime("%d.%m.%Y") for lesson_date in lesson_dates],
            "lesson_count": len(lesson_dates),
            "lesson_count_summary": f"Bu davr bo'yicha {len(lesson_dates)} ta mos dars kuni topildi",
        }
    )
    return adjusted


def _apply_lesson_count_breakdown(preview: dict, enrollment: Enrollment, lesson_count: int) -> dict:
    adjusted = dict(preview)
    breakdown = tuition_amount_breakdown(
        enrollment,
        lesson_count,
        course_price=full_course_amount(enrollment),
        monthly_lessons=int(preview.get("monthly_lessons", 0) or 0),
        teacher_percent=getattr(enrollment, "oqituvchi_foiz", 0) or 0,
    )
    adjusted.update(
        {
            "lesson_count": breakdown["lesson_count"],
            "per_lesson_amount": breakdown["per_lesson_amount"],
            "per_lesson_amount_display": _format_money_exact(breakdown["per_lesson_amount"]),
            "fee_amount": breakdown["fee_amount"],
            "fee_amount_display": _format_money_exact(breakdown["fee_amount"]),
            "full_turnover": breakdown["fee_amount"],
            "teacher_share": breakdown["teacher_share"],
            "teacher_share_display": _format_money_exact(breakdown["teacher_share"]),
            "center_share": breakdown["center_share"],
            "center_share_display": _format_money_exact(breakdown["center_share"]),
            "lesson_count_summary": f"Hisoblangan darslar: {breakdown['lesson_count']} ta",
        }
    )
    return adjusted


def _parse_period_end(value, fallback_month: date) -> date:
    parsed = parse_date((value or "").strip()) if value not in (None, "") else None
    if parsed is None:
        return month_last_day(fallback_month)
    return parsed


def _build_tuition_preview_enrollment(
    *,
    base_enrollment: Enrollment | None = None,
    group: Group | None = None,
    start_date: date | None = None,
    lesson_pattern: str | None = None,
    monthly_lessons: int | None = None,
    course_price: int | None = None,
    teacher_percent: int | None = None,
    student_payable_amount: int | None = None,
) -> Enrollment:
    resolved_group = group or getattr(base_enrollment, "group", None)
    resolved_start_date = start_date
    if resolved_start_date is None and base_enrollment is not None:
        resolved_start_date = getattr(base_enrollment, "joined_at", None) or enrollment_start_date(base_enrollment)
    if resolved_start_date is None:
        resolved_start_date = timezone.localdate()
    resolved_monthly_lessons = int(
        monthly_lessons
        or getattr(base_enrollment, "monthly_lessons", 0)
        or getattr(resolved_group, "oy_dars_soni", 0)
        or 12
    )
    resolved_course_price = int(
        course_price
        if course_price is not None
        else getattr(base_enrollment, "kurs_narhi", 0)
        or getattr(resolved_group, "kurs_narxi", 0)
        or 0
    )
    resolved_teacher_percent = int(
        teacher_percent
        if teacher_percent is not None
        else getattr(base_enrollment, "oqituvchi_foiz", 0)
        or getattr(resolved_group, "oqituvchi_foiz", 0)
        or 0
    )
    schedule_meta = resolve_lesson_schedule(
        resolved_start_date,
        lesson_pattern if lesson_pattern is not None else getattr(base_enrollment, "lesson_pattern", None),
    )
    resolved_start_date = schedule_meta["start_date"]
    resolved_pattern = schedule_meta["lesson_pattern"]

    preview_enrollment = Enrollment(
        group=resolved_group,
        course=getattr(resolved_group, "category_obj", None) or getattr(base_enrollment, "course", None),
        student=getattr(base_enrollment, "student", None),
        center=getattr(resolved_group, "center", None) or getattr(base_enrollment, "center", None),
        kurs_narhi=resolved_course_price,
        oqituvchi_foiz=resolved_teacher_percent,
        student_payable_amount=student_payable_amount,
        monthly_lessons=resolved_monthly_lessons,
        joined_at=resolved_start_date,
        lesson_pattern=resolved_pattern,
        is_active=True,
    )
    preview_enrollment.created_at = getattr(base_enrollment, "created_at", None) or timezone.now()
    preview_enrollment._tuition_start_date = resolved_start_date
    preview_enrollment._tuition_requested_start_date = start_date or resolved_start_date
    return preview_enrollment


@csrf_exempt
def calculate_lessons_api(request):
    try:
        if request.method != "POST":
            return JsonResponse(
                {"success": False, "error": "Method not allowed"},
                status=400,
            )
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "Avval tizimga kiring."},
                status=400,
            )

        center = get_active_center(request)
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except (UnicodeDecodeError, json.JSONDecodeError):
            if request.content_type == "application/json":
                return JsonResponse(
                    {"success": False, "error": "JSON ma'lumot noto'g'ri yuborildi."},
                    status=400,
                )
            payload = request.POST

        enrollment_id = _parse_int_value(payload.get("enrollment_id"))
        group_id = _parse_int_value(payload.get("group_id"))

        enrollment = None
        if enrollment_id:
            enrollment_qs = Enrollment.all_objects.select_related("group", "group__category_obj", "student", "course")
            if center:
                enrollment_qs = enrollment_qs.filter(
                    Q(center=center)
                    | Q(center__isnull=True, group__center=center)
                    | Q(center__isnull=True, student__center=center)
                )
            enrollment = enrollment_qs.filter(id=enrollment_id).first()
            if enrollment is None:
                return JsonResponse({"success": False, "error": "Enrollment topilmadi."}, status=400)

        group = None
        if group_id:
            if enrollment is not None:
                sibling_qs = Enrollment.objects.select_related("group", "group__category_obj", "student", "course").filter(
                    student=enrollment.student,
                    group_id=group_id,
                    is_active=True,
                    student__is_archived=False,
                    group__is_archived=False,
                    group__is_deleted=False,
                )
                if center:
                    sibling_qs = sibling_qs.filter(center=center)
                sibling_enrollment = sibling_qs.first()
                if sibling_enrollment is None:
                    return JsonResponse({"success": False, "error": "Bu o'quvchi ushbu guruhga biriktirilmagan."}, status=400)
                enrollment = sibling_enrollment
                group = enrollment.group
            else:
                group_qs = Group.objects.select_related("category_obj")
                if center:
                    group_qs = group_qs.filter(center=center)
                group = group_qs.filter(id=group_id).first()
                if group is None:
                    return JsonResponse({"success": False, "error": "Guruh topilmadi."}, status=400)
        elif enrollment is not None:
            group = enrollment.group

        if group is None:
            return JsonResponse({"success": False, "error": "Guruhni tanlang."}, status=400)

        start_date_raw = (payload.get("joined_at") or payload.get("start_date") or "").strip()
        if start_date_raw:
            start_date = parse_date(start_date_raw)
            if start_date is None:
                return JsonResponse({"success": False, "error": "Boshlanish sanasi noto'g'ri."}, status=400)
        else:
            start_date = (
                getattr(enrollment, "joined_at", None)
                or enrollment_start_date(enrollment)
                if enrollment is not None
                else timezone.localdate()
            )
        allowed_lesson_patterns = {
            Enrollment.LESSON_PATTERN_ODD,
            Enrollment.LESSON_PATTERN_EVEN,
            Enrollment.LESSON_PATTERN_DAILY,
            Enrollment.LESSON_PATTERN_GROUP,
        }
        lesson_pattern_raw = payload.get("lesson_pattern")
        if lesson_pattern_raw not in (None, ""):
            lesson_pattern_raw = str(lesson_pattern_raw).strip().lower()
            if lesson_pattern_raw not in allowed_lesson_patterns:
                return JsonResponse({"success": False, "error": "Dars patterni noto'g'ri."}, status=400)
        lesson_pattern = normalize_lesson_pattern(
            lesson_pattern_raw or getattr(enrollment, "lesson_pattern", None)
        )
        if lesson_pattern not in allowed_lesson_patterns:
            return JsonResponse({"success": False, "error": "Dars patterni noto'g'ri."}, status=400)

        course_price = int(
            _parse_int_value(
                payload.get("kurs_narhi"),
                getattr(enrollment, "kurs_narhi", 0) if enrollment is not None else getattr(group, "kurs_narxi", 0),
            )
            or 0
        )
        teacher_percent = int(
            _parse_int_value(
                payload.get("oqituvchi_foiz"),
                getattr(enrollment, "oqituvchi_foiz", 0) if enrollment is not None else getattr(group, "oqituvchi_foiz", 0),
            )
            or 0
        )
        monthly_lessons = int(
            _parse_int_value(
                payload.get("monthly_lessons"),
                getattr(enrollment, "monthly_lessons", 0) if enrollment is not None else getattr(group, "oy_dars_soni", 0),
            )
            or getattr(group, "oy_dars_soni", 0)
            or 12
        )

        teacher_share_only = _parse_bool_value(payload.get("teacher_share_only"))
        missing = object()
        payable_raw = payload.get("student_payable_amount", missing)
        if teacher_share_only:
            student_payable_amount = round(course_price * teacher_percent / 100)
        elif payable_raw is missing:
            student_payable_amount = getattr(enrollment, "student_payable_amount", None) if enrollment is not None else None
        elif payable_raw in (None, "", "None"):
            student_payable_amount = None
        else:
            student_payable_amount = _parse_int_value(
                payable_raw,
                getattr(enrollment, "student_payable_amount", None) if enrollment is not None else None,
            )

        preview_enrollment = _build_tuition_preview_enrollment(
            base_enrollment=enrollment,
            group=group,
            start_date=start_date,
            lesson_pattern=lesson_pattern,
            monthly_lessons=monthly_lessons,
            course_price=course_price,
            teacher_percent=teacher_percent,
            student_payable_amount=student_payable_amount,
        )
        preview_month = month_first_day(timezone.localdate())
        preview = tuition_month_preview(preview_enrollment, preview_month)
        period_end = _parse_period_end(payload.get("period_end_date") or payload.get("end_date"), preview["month"])
        preview = _apply_period_end_to_preview(preview, period_end)

        remaining_raw = payload.get("remaining_lessons", missing)
        if remaining_raw is missing:
            remaining_lessons = (
                getattr(enrollment, "remaining_lessons_override", None)
                if enrollment is not None and getattr(enrollment, "remaining_lessons_override", None) is not None
                else int(preview["lesson_count"] or 0)
            )
        elif remaining_raw in ("", None):
            remaining_lessons = int(preview["lesson_count"] or 0)
        else:
            try:
                remaining_lessons = validate_remaining_lessons(remaining_raw)
            except ValidationError as exc:
                return JsonResponse({"success": False, "error": exc.messages[0]}, status=400)

        try:
            lesson_plan = calculate_lessons(
                start_date=preview["start_date"],
                remaining_lessons=remaining_lessons,
                pattern=preview["lesson_pattern"],
                from_date=timezone.localdate(),
                group=group,
            )
        except ValidationError as exc:
            return JsonResponse({"success": False, "error": exc.messages[0]}, status=400)
        preview = _apply_lesson_count_breakdown(preview, preview_enrollment, remaining_lessons)
        preview_payload = _serialize_tuition_preview(preview)
        lesson_plan_payload = _serialize_lesson_plan(lesson_plan)
        response_data = {
            "total_lessons": preview_payload["lesson_count"],
            "lesson_price": preview_payload["per_lesson_amount"],
            "total_debt": preview_payload["fee_amount"],
            "teacher_share": preview_payload["teacher_share"],
            "center_share": preview_payload["center_share"],
            "lesson_dates": [
                int(label.split(".", 1)[0])
                for label in preview_payload.get("lesson_date_labels", [])
                if str(label).split(".", 1)[0].isdigit()
            ],
            "end_date": preview_payload["period_end_date"],
        }
        return JsonResponse(
            {
                "success": True,
                "data": response_data,
                "preview": preview_payload,
                "lesson_plan": lesson_plan_payload,
            },
            status=200,
        )
    except Exception as exc:
        logger.exception("calculate_lessons_api failed")
        return JsonResponse(
            {"success": False, "error": str(exc) or "Hisob-kitobda xatolik yuz berdi."},
            status=400,
        )


def sync_tuition_fee(enrollment, new_fee=None, start_month=None):
    """
    Viewlar uchun backward-compatible wrapper.
    Endi barcha logika service qatlamida bajariladi.
    """
    from education.services.tuition import sync_tuition_fee as service_sync_tuition_fee

    service_sync_tuition_fee(
        enrollment=enrollment,
        start_month=month_first_day(start_month or timezone.localdate()),
        new_fee=int(new_fee if new_fee is not None else effective_student_payable_amount(enrollment) or 0),
    )


def _get_int(querydict, key, default=0):
    try:
        val = querydict.get(key, None)
        if val in (None, "", "None"):
            return default
        return int(val)
    except (TypeError, ValueError):
        return default


def _schedule_weekday_labels():
    return [
        (1, "Dushanba"),
        (2, "Seshanba"),
        (3, "Chorshanba"),
        (4, "Payshanba"),
        (5, "Juma"),
        (6, "Shanba"),
        (7, "Yakshanba"),
    ]


def _student_group_financial_cards(
    student,
    *,
    center=None,
    month: date | None = None,
    include_dates: bool = True,
):
    target_month = month_first_day(month or timezone.localdate())
    enrollments_qs = (
        Enrollment.objects
        .select_related("group", "group__category_obj", "course")
        .filter(
            student=student,
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        )
        .order_by("group__nom", "id")
    )
    if center:
        enrollments_qs = enrollments_qs.filter(center=center)

    cards = []
    totals = {
        "fee_amount": 0,
        "teacher_share": 0,
        "center_share": 0,
        "debt_amount": 0,
    }

    for enrollment in enrollments_qs:
        snapshot = enrollment_month_financial_snapshot(enrollment, target_month)
        default_course_price = int(getattr(enrollment.group, "kurs_narxi", 0) or 0)
        card = {
            "enrollment_id": enrollment.id,
            "group_id": getattr(enrollment, "group_id", None),
            "group_name": getattr(enrollment.group, "nom", "—"),
            "start_date": snapshot["start_date"],
            "lesson_pattern": snapshot["lesson_pattern"],
            "lesson_pattern_label": snapshot["lesson_pattern_label"],
            "lesson_count": snapshot["lesson_count"],
            "lesson_dates": snapshot["lesson_dates"] if include_dates else [],
            "lesson_date_labels": snapshot["lesson_date_labels"] if include_dates else [],
            "course_price": snapshot["course_price"],
            "course_price_display": snapshot["course_price_display"],
            "default_course_price": default_course_price,
            "default_course_price_display": format_money(default_course_price),
            "is_individual_price": bool(default_course_price and snapshot["course_price"] != default_course_price),
            "fee_amount": snapshot["fee_amount"],
            "fee_amount_display": snapshot["fee_amount_display"],
            "teacher_share": snapshot["teacher_share"],
            "teacher_share_display": snapshot["teacher_share_display"],
            "center_share": snapshot["center_share"],
            "center_share_display": snapshot["center_share_display"],
            "debt_amount": snapshot["debt_amount"],
            "debt_amount_display": snapshot["debt_amount_display"],
        }
        cards.append(card)
        totals["fee_amount"] += card["fee_amount"]
        totals["teacher_share"] += card["teacher_share"]
        totals["center_share"] += card["center_share"]
        totals["debt_amount"] += card["debt_amount"]

    return {
        "cards": cards,
        "totals": {
            **totals,
            "fee_amount_display": format_money(totals["fee_amount"]),
            "teacher_share_display": format_money(totals["teacher_share"]),
            "center_share_display": format_money(totals["center_share"]),
            "debt_amount_display": format_money(totals["debt_amount"]),
        },
    }


# ---------- Ruxsat helperlari ----------
def _can_manage(u):
    return u.is_superuser or getattr(u, "role", None) in ("director", "manager")


def _can_give_points(user, g: Group):
    return (
        user.is_superuser
        or user.role in ("director", "manager")
        or (user.role == "teacher" and g.oqituvchi_id == user.id)
    )

    return user.is_superuser or user.role in ("director", "manager") or (
        user.role == "teacher" and g.oqituvchi_id == user.id
    )


def get_active_center(request):
    """
    Returns the active center for the current request.
    Now fully handled by TenantMiddleware.
    """
    return getattr(request, 'center', None)



from chaqmoq.models import Ledger

from datetime import datetime
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from chaqmoq.models import Ledger
from django.db.models import Count, Max, Q, Case, When, Value, IntegerField, F


from .models import Enrollment, TuitionMonth, PaymentAllocation





# def get_month_paid(enr: Enrollment, month: date) -> int:
#     month = month_first_day(month)
#     tm = TuitionMonth.objects.filter(enrollment=enr, month=month).first()
#     if not tm:
#         return 0
#     s = PaymentAllocation.objects.filter(tuition_month=tm).aggregate(x=Sum("amount"))["x"] or 0
#     return int(s)


def parse_month_str(s: str) -> date | None:
    """
    'YYYY-MM' yoki 'YYYY-MM-DD' -> date(YYYY, MM, 1)
    """
    if not s:
        return None
    s = s.strip()
    if len(s) >= 10:
        s = s[:7]
    if len(s) != 7 or s[4] != "-":
        return None
    try:
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            return None
        return date(y, m, 1)
    except Exception:
        return None


# def user_can_manage_payments(user) -> bool:
#     # sizda role bor: manager/director
#     return user.is_superuser or getattr(user, "role", None) in ("manager", "director")


from urllib.parse import urlparse, parse_qs


def _get_month_from_next(next_url: str, fallback: date) -> date:
    try:
        qs = parse_qs(urlparse(next_url).query)

        # 1) eng to'g'risi: month=YYYY-MM
        m = (qs.get("month", [""])[0] or "").strip()
        if m:
            return parse_month_str(m) or fallback

        # 2) sizda ishlayotgan variant: pay_month=1..12
        pm = (qs.get("pay_month", [""])[0] or "").strip()
        if pm.isdigit():
            mm = int(pm)
            if 1 <= mm <= 12:
                # year bo'lmasa joriy yil
                yy = (qs.get("year", [""])[0] or "").strip()
                yy = int(yy) if yy.isdigit() else fallback.year
                return date(yy, mm, 1)

        return fallback
    except Exception:
        return fallback


def _resolve_paid_date(raw_value: str | None) -> date:
    parsed = parse_date((raw_value or "").strip())
    return parsed or localdate()


def _build_paid_at_for_date(selected_date: date) -> datetime:
    current_local = timezone.localtime()
    return current_local.replace(
        year=selected_date.year,
        month=selected_date.month,
        day=selected_date.day,
        microsecond=0,
    )


from django.core.paginator import Paginator





def _can_manage(u: User) -> bool:
    return u.is_superuser or getattr(u, "role", None) in ("director", "manager")


def _teacher_can(user: User, g: Group) -> bool:
    return (
        user.is_superuser
        or getattr(user, "role", None) in ("director", "manager")
        or (getattr(user, "role", None) == "teacher" and getattr(g, "oqituvchi_id", None) == user.id)
    )


def _can_give_points(user: User, g: Group) -> bool:
    return _teacher_can(user, g)


# ============================================================
#  TUITION (OYLIK) HELPERS
# ============================================================

def month_first_day(d: date) -> date:
    return d.replace(day=1)


# def parse_month_str(month_str: str) -> date:
#     # "2026-01" -> 2026-01-01
#     if not month_str:
#         return month_first_day(timezone.localdate())
#     try:
#         y, m = month_str.split("-")
#         return date(int(y), int(m), 1)
#     except Exception:
#         return month_first_day(timezone.localdate())


# def _get_fee_amount(enr: Enrollment) -> int:
#     # Enrollment kurs_narhi -> Group kurs_narxi/kurs_narhi fallback
#     enr_fee = getattr(enr, "kurs_narhi", None)
#     if enr_fee:
#         return int(enr_fee)
#     g = getattr(enr, "group", None)
#     if not g:
#         return 0
#     return int(getattr(g, "kurs_narxi", 0) or getattr(g, "kurs_narhi", 0) or 0)


# def ensure_tuition_month(enr: Enrollment, month: date) -> TuitionMonth:
#     month = month_first_day(month)
#     fee = _get_fee_amount(enr)

#     tm, _ = TuitionMonth.objects.get_or_create(
#         enrollment=enr,
#         month=month,
#         defaults={"fee_amount": fee},
#     )

#     # fee 0 bo'lib qolsa fallback
#     if not getattr(tm, "fee_amount", 0):
#         tm.fee_amount = fee
#         tm.save(update_fields=["fee_amount"])
#     return tm


def get_month_paid(enr: Enrollment, month: date) -> int:
    month = month_first_day(month)
    tm = TuitionMonth.objects.filter(enrollment=enr, month=month).first()
    if not tm:
        return 0
    s = PaymentAllocation.objects.filter(tuition_month=tm).aggregate(x=Sum("amount"))["x"] or 0
    return int(s)


def _add_month(d: date, n: int = 1) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)


def _model_has_field(model, field_name: str) -> bool:
    return any(f.name == field_name for f in model._meta.get_fields())

# def create_payment_and_allocate(
#     enrollment: Enrollment,
#     cash_amount: int,
#     card_amount_som: int,
#     created_by: User | None,
#     start_month: date | None = None,
# ) -> Payment:
#     """
#     Payment yaratadi va pullarni TuitionMonth'larga ketma-ket taqsimlaydi:
#     start_month -> keyingi oylar...
#     """
#     start_month = month_first_day(start_month or timezone.localdate())
#     total = int(cash_amount or 0) + int(card_amount_som or 0)

#     if total <= 0:
#         raise ValueError("To'lov summasi 0 bo'lishi mumkin emas.")

#     # Payment create (fieldlar turlicha bo'lishi mumkin)
#     kwargs = {}
#     if _model_has_field(Payment, "enrollment"):
#         kwargs["enrollment"] = enrollment
#     if _model_has_field(Payment, "student"):
#         kwargs["student"] = enrollment.student
#     if _model_has_field(Payment, "group"):
#         kwargs["group"] = enrollment.group

#     if _model_has_field(Payment, "cash_amount"):
#         kwargs["cash_amount"] = int(cash_amount or 0)

#     # kartani ba'zi loyihalarda card_amount_som, ba'zida card_amount
#     if _model_has_field(Payment, "card_amount_som"):
#         kwargs["card_amount_som"] = int(card_amount_som or 0)
#     elif _model_has_field(Payment, "card_amount"):
#         kwargs["card_amount"] = int(card_amount_som or 0)

#     if _model_has_field(Payment, "summa"):
#         kwargs["summa"] = total

#     if _model_has_field(Payment, "paid_at"):
#         kwargs["paid_at"] = timezone.now()
#     else:
#         # eski fieldlar bo'lsa
#         if _model_has_field(Payment, "sana"):
#             kwargs["sana"] = timezone.localdate()
#         if _model_has_field(Payment, "vaqt"):
#             kwargs["vaqt"] = timezone.localtime().time()

#     if created_by and _model_has_field(Payment, "created_by"):
#         kwargs["created_by"] = created_by

#     p = Payment.objects.create(**kwargs)

#     # Allocation: start_month dan boshlab ketma-ket oylar
#     left = total
#     cur = start_month

#     # 60 oy max (cheksiz loop bo'lmasin)
#     for _ in range(60):
#         tm = ensure_tuition_month(enrollment, cur)
#         fee = int(getattr(tm, "fee_amount", 0) or 0)

#         # fee 0 bo'lsa — keyingi oyga o'tamiz
#         if fee <= 0:
#             cur = _add_month(cur, 1)
#             continue

#         paid = get_month_paid(enrollment, cur)
#         need = max(0, fee - paid)
#         if need <= 0:
#             cur = _add_month(cur, 1)
#             continue

#         alloc = min(need, left)
#         if alloc > 0:
#             PaymentAllocation.objects.create(payment=p, tuition_month=tm, amount=alloc)
#             left -= alloc

#         if left <= 0:
#             break

#         cur = _add_month(cur, 1)

#     # Enrollment jami_tolangan update (agar field bo'lsa)
#     if _model_has_field(Enrollment, "jami_tolangan"):
#         Enrollment.objects.filter(pk=enrollment.pk).update(
#             jami_tolangan=Coalesce(F("jami_tolangan"), 0) + total
#         )

#     return p

@require_POST
@login_required
def create_payment(request):
    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("education:tolovlar_home")

    next_url = request.POST.get("next") or "education:tolovlar_home"

    enrollment_id = request.POST.get("enrollment_id")
    student_id = request.POST.get("student_id")
    payment_scope = (request.POST.get("payment_scope") or "").strip()
    month_str = (request.POST.get("month") or "").strip()
    selected_paid_date = _resolve_paid_date(request.POST.get("paid_date"))
    selected_paid_at = _build_paid_at_for_date(selected_paid_date)

    fallback = selected_paid_date.replace(day=1)
    start_month = parse_month_str(month_str)
    if start_month is None:
        start_month = _get_month_from_next(next_url, fallback)

    # TASK 4: "Qaysi oy uchun?" — manager tomonidan tanlangan oy
    month_for_payment_str = (request.POST.get("month_for_payment") or "").strip()
    month_for_payment = parse_month_str(month_for_payment_str)  # None bo'lsa avtomatik tanlaydi

    cash_amount = int(Decimal(request.POST.get("cash_amount") or "0"))
    card_amount = int(Decimal(request.POST.get("card_amount") or "0"))
    note = (request.POST.get("note") or "").strip()

    if not enrollment_id and not student_id:
        messages.error(request, "ID kelmadi.")
        return redirect(next_url)

    center = get_active_center(request)

    if enrollment_id:
        qs = Enrollment.objects.all()
        if center:
            qs = qs.filter(
                Q(center=center)
                | Q(center__isnull=True, group__center=center)
                | Q(center__isnull=True, student__center=center)
            )
        enrollment = get_object_or_404(qs, id=enrollment_id)

        # Double-submit himoyasi: oxirgi 60 sekundda shu enrollment uchun
        # bir xil summa va sana bilan to'lov yozilganmi? Agar ha — qaytarib
        # yozmaymiz (foydalanuvchi tugmani 2 marta bossa ham, faqat 1 to'lov).
        from .models import Payment as _Payment
        total = cash_amount + card_amount
        recent_dup = _Payment.objects.filter(
            enrollment=enrollment,
            summa=total,
            paid_date=selected_paid_date,
            created_at__gte=timezone.now() - timedelta(seconds=5),
        ).exists()
        if recent_dup:
            messages.warning(
                request,
                "⚠️ Aynan shu summa va sana bilan to'lov yaqinda yozilgan. Takrorlanmasin uchun e'tiborsiz qoldirildi.",
            )
            return redirect(next_url)

        # O'tgan barcha oylar uchun TuitionMonth mavjudligini ta'minlaymiz —
        # shunda allocation past oylarni ham to'g'ri yopadi.
        ensure_all_tuition_months_since_start(enrollment, start_month)

        try:
            with transaction.atomic():
                create_payment_and_allocate(
                    enrollment=enrollment,
                    cash_amount=cash_amount,
                    card_amount_som=card_amount,
                    created_by=request.user,
                    start_month=month_for_payment,  # TASK 4: manager tanlagan oy; None bo'lsa eng eski to'lanmagan oydan
                    paid_at=selected_paid_at,
                    note=note,
                    payment_type=infer_payment_type(cash_amount, card_amount),
                    # Menejer aniq oy tanlagan bo'lsa — butun summa FAQAT shu
                    # oyga (keyingi oyga oshirilmaydi)
                    strict_month=bool(month_for_payment),
                )
            messages.success(request, f"✅ {enrollment.student.get_full_name()} uchun to'lov saqlandi!")
        except Exception as e:
            messages.error(request, f"❌ Xatolik: {e}")
            
    elif student_id:
        # ✅ CONSOLIDATED DISTRIBUTION LOGIC
        user_qs = User.objects.filter(role="student")
        if center:
            _enr_cq = (
                Q(center=center)
                | Q(center__isnull=True, group__center=center)
                | Q(center__isnull=True, student__center=center)
            )
            user_qs = user_qs.filter(
                Q(center=center)
                | Q(pk__in=Enrollment.objects.filter(_enr_cq).values("student_id"))
            )
        student = get_object_or_404(user_qs, id=student_id)
        
        # Faol enrollment'lar
        active_enrollments = Enrollment.objects.filter(
            student=student, is_active=True,
            group__is_archived=False, group__is_deleted=False,
        ).order_by('id')

        # Guruhdan chiqarilgan (is_active=False) lekin to'lanmagan TuitionMonth bor
        # enrollment'lar ham to'lovga qo'shiladi — ularning qarzi ham yig'ilishi kerak.
        inactive_with_debt_ids = list(
            TuitionMonth.objects
            .filter(
                enrollment__student=student,
                enrollment__is_active=False,
                enrollment__group__is_archived=False,
                enrollment__group__is_deleted=False,
                is_deleted=False,
            )
            .values_list("enrollment_id", flat=True)
            .distinct()
        )
        inactive_enrollments = Enrollment.objects.filter(
            id__in=inactive_with_debt_ids,
            group__is_archived=False,
            group__is_deleted=False,
        ).order_by('id')

        # Ikkisini birlashtirish — takrorlanmaslik uchun ID bo'yicha
        all_enr_ids = list(dict.fromkeys(
            list(active_enrollments.values_list('id', flat=True)) +
            list(inactive_enrollments.values_list('id', flat=True))
        ))
        enrollments = Enrollment.objects.filter(id__in=all_enr_ids).order_by('id')

        if not enrollments.exists():
            messages.error(request, "O'quvchida faol kurslar topilmadi.")
            return redirect(next_url)

        if payment_scope == "teacher_share_only":
            scoped_enrollment_ids = []
            for e in enrollments:
                full_amount = full_course_amount(e)
                effective_amount = effective_student_payable_amount(e)
                teacher_share_amount = int(getattr(e, "oqituvchi_daromadi", 0) or 0)
                if (
                    e.student_payable_amount not in (None, "")
                    and full_amount > effective_amount
                    and effective_amount == teacher_share_amount
                ):
                    tm = ensure_tuition_month(e, start_month)
                    fee = int(getattr(tm, "fee_amount", 0) or 0)
                    paid = int(get_month_paid(e, start_month) or 0)
                    debt = max(0, fee - paid)
                    if debt > 0:
                        scoped_enrollment_ids.append(e.id)

            if scoped_enrollment_ids:
                enrollments = enrollments.filter(id__in=scoped_enrollment_ids).order_by("id")
            else:
                messages.error(request, "Faol o'qituvchi haqqi qarzi topilmadi.")
                return redirect(next_url)
            
        from education.services.tuition import find_earliest_unpaid_month

        # TuitionMonth rekordlarini transaction tashqarisida yaratamiz —
        # bu faqat ensure operatsiyasi, payment bilan bog'liq emas.
        for e in enrollments:
            ensure_all_tuition_months_since_start(e, start_month)

        try:
            with transaction.atomic():
                # One check for the whole payment
                first_group = enrollments[0].group if enrollments else None
                main_payment = Payment.objects.create(
                    student=student,
                    group=first_group,
                    cash_amount=cash_amount,
                    card_amount=card_amount,
                    summa=cash_amount + card_amount,
                    created_by=request.user,
                    paid_date=selected_paid_date,
                    paid_time=selected_paid_at.time().replace(microsecond=0),
                    center=center,
                    note=note,
                    payment_type="mixed" if (cash_amount > 0 and card_amount > 0) else ("card" if card_amount > 0 else "cash")
                )

                remaining_sum = cash_amount + card_amount

                # Har enrollment uchun to'lovni taqsimlash.
                # Agar menejer month_for_payment tanlagan bo'lsa — faqat shu oyga;
                # aks holda eng eski to'lanmagan oydan boshlaymiz.
                for e in enrollments:
                    if remaining_sum <= 0:
                        break

                    if month_for_payment:
                        # Menejer aniq oy tanlagan — faqat shu oyning qarzini olamiz
                        past_tms = TuitionMonth.objects.filter(
                            enrollment=e,
                            month=month_for_payment,
                            is_deleted=False,
                        ).order_by("month")
                    else:
                        # Avtomatik: o'tgan + joriy oylardagi barcha qarz
                        past_tms = TuitionMonth.objects.filter(
                            enrollment=e,
                            month__lte=start_month,
                            is_deleted=False,
                        ).order_by("month")

                    total_debt = sum(
                        max(0, int(getattr(tm, "fee_amount", 0) or 0) - int(get_month_paid(e, tm.month) or 0))
                        for tm in past_tms
                    )

                    if total_debt <= 0:
                        continue

                    take = min(remaining_sum, total_debt)

                    # Tanlangan oydan yoki eng eski to'lanmagan oydan boshlaymiz
                    if month_for_payment:
                        allocation_start = month_for_payment
                        # Agar tanlangan oy yopiq bo'lsa — xato ko'rsatamiz
                        e_center = getattr(e, "center", None) or getattr(e.group, "center", None)
                        if is_month_closed_for_center(e_center, allocation_start):
                            raise ValueError(
                                f"{allocation_start:%B %Y} oyi mahkamlangan. "
                                "Bu oyga to'lov yozish mumkin emas."
                            )
                    else:
                        earliest_tm = find_earliest_unpaid_month(e)
                        allocation_start = earliest_tm.month if earliest_tm else start_month

                    _allocate_amount_forward(
                        enrollment=e,
                        payment=main_payment,
                        amount=take,
                        start_month=allocation_start,
                    )
                    remaining_sum -= take

                # Ortiqcha to'lov (kredit) — faol enrollmentning kelgusi oyiga
                # Guruhdan chiqarilgan (is_active=False) enrollment uchun ortiqcha
                # to'lov kelajak oylariga yozilmasin — bu noto'g'ri qarz hosil qiladi.
                if remaining_sum > 0:
                    overflow_enr = next(
                        (e for e in enrollments if getattr(e, "is_active", False)),
                        None,
                    )
                    if overflow_enr:
                        if month_for_payment:
                            # Menejer aniq oy tanlagan: ortiqcha ham SHU OYGA
                            # yoziladi (keyingi oyga surilmaydi) — hisobotlarda
                            # pul aynan tanlangan oyda ko'rinadi.
                            from education.services.tuition import ensure_tuition_month as _etm_strict
                            _tm_sel = _etm_strict(
                                overflow_enr, month_for_payment,
                                _exclude_payment_id=main_payment.id,
                            )
                            PaymentAllocation.objects.create(
                                center=getattr(main_payment, "center", None)
                                or getattr(overflow_enr, "center", None),
                                payment=main_payment,
                                tuition_month=_tm_sel,
                                amount=remaining_sum,
                            )
                        else:
                            _allocate_amount_forward(
                                enrollment=overflow_enr,
                                payment=main_payment,
                                amount=remaining_sum,
                                start_month=start_month,
                            )
            
            messages.success(request, f"✅ {student.get_full_name()} uchun umumiy to'lov saqlandi!")
        except Exception as e:
            messages.error(request, f"❌ Xatolik: {e}")

    try:
        target_student = None
        if enrollment_id:
            target_student = enrollment.student
        elif student_id:
            target_student = student
        # Faqat tolovlar bo'limidan kelgan va qarz to'liq yopilgan bo'lsa
        # tolovlar_home ga o'tamiz. Qarzdorlar bo'limidan kelgan bo'lsa —
        # next_url o'zgarmaydi, foydalanuvchi shu sahifada qoladi.
        _from_qarzdorlar = "qarzdorlar" in (next_url or "")
        if (
            not _from_qarzdorlar
            and target_student
            and get_student_total_debt(target_student, center) <= 0
        ):
            next_url = reverse("education:tolovlar_home")
    except Exception:
        pass

    return redirect(next_url)



from decimal import Decimal, InvalidOperation
from django.db.models import Sum, F, Value


@require_POST
@login_required
def payment_update(request, payment_id: int):
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # from core.tenant import get_request_center
    center = get_active_center(request)
    qs = Payment.objects.select_related("enrollment")
    if center:
        qs = qs.filter(center=center)

    p = get_object_or_404(qs, id=payment_id)
    enrollment = getattr(p, "enrollment", None)
    if not enrollment:
        return JsonResponse({"ok": False, "error": "enrollment_not_found"}, status=400)

    old_total = int(getattr(p, "summa", 0) or 0)

    cash_raw = request.POST.get("cash_amount")
    card_raw = request.POST.get("card_amount")
    paid_date_raw = request.POST.get("paid_date")

    if cash_raw is None and card_raw is None and paid_date_raw is None:
        note = request.POST.get("note")
        if note is not None:
            p.note = note.strip()   # bo'sh bo'lsa ham "" bo'lib saqlanadi
            p.save(update_fields=["note"])
        return JsonResponse({"ok": True, "payment_id": p.id, "note": p.note or ""})

    from decimal import Decimal, InvalidOperation
    try:
        cash_amount = int(Decimal((request.POST.get("cash_amount") or "0").strip()))
        card_amount = int(Decimal((request.POST.get("card_amount") or "0").strip()))
    except (InvalidOperation, ValueError):
        return JsonResponse({"ok": False, "error": "summa_notogri"}, status=400)

    if cash_amount < 0 or card_amount < 0:
        return JsonResponse({"ok": False, "error": "summa_manfiy_bolmaydi"}, status=400)

    new_total = cash_amount + card_amount
    if new_total < 0:
        return JsonResponse({"ok": False, "error": "summa_manfiy_bolmaydi"}, status=400)

    month_str = (request.POST.get("month") or "").strip()
    start_month = parse_month_str(month_str) if month_str else None

    # Update metadata
    note = request.POST.get("note")
    if note is not None:
        p.note = note.strip()

    paid_date_str = request.POST.get("paid_date")
    if paid_date_str:
        try:
            new_paid_date = parse_date(paid_date_str)
            if new_paid_date:
                p.paid_date = new_paid_date
                # Agar alohida 'month' yuborilmagan bo'lsa, paid_date oyini start_month sifatida ishlatamiz
                if not start_month:
                    start_month = new_paid_date.replace(day=1)
        except Exception:
            pass

    p.save()

    try:
        update_payment_and_reallocate(
            payment=p,
            cash_amount=cash_amount,
            card_amount_som=card_amount,
            start_month=start_month,
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    return JsonResponse({
        "ok": True,
        "payment_id": p.id,
        "old_total": old_total,
        "new_total": new_total,
        "delta": new_total - old_total,
        "start_month": (start_month or timezone.localdate().replace(day=1)).strftime("%Y-%m"),
    })


def month_start(d: date) -> date:
    return d.replace(day=1)

def parse_month_str_safe(s: str) -> date:
    """
    'YYYY-MM' -> date(YYYY,MM,1)
    bo'sh yoki xato bo'lsa -> joriy oy(1-kun)
    """
    s = (s or "").strip()
    today = timezone.localdate()
    if len(s) == 7 and s[4] == "-":
        try:
            y = int(s[:4])
            m = int(s[5:])
            return date(y, m, 1)
        except Exception:
            pass
    return month_start(today)

def _safe_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _is_teacher_share_only_enrollment(enrollment) -> bool:
    if not enrollment:
        return False

    full_amount = full_course_amount(enrollment)
    effective_amount = effective_student_payable_amount(enrollment)
    teacher_share_amount = int(getattr(enrollment, "oqituvchi_daromadi", 0) or 0)

    return (
        enrollment.student_payable_amount not in (None, "")
        and full_amount > effective_amount
        and effective_amount == teacher_share_amount
    )


def _parse_yyyy_mm(s: str):
    """
    '2026-01' -> date(2026,1,1)
    None/invalid -> current month first day
    """
    s = (s or "").strip()
    try:
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            raise ValueError()
        return timezone.datetime(y, m, 1).date()
    except Exception:
        now = timezone.localdate()
        return timezone.datetime(now.year, now.month, 1).date()


def _first_day_of_month(d: date) -> date:
    return d.replace(day=1)


@transaction.atomic
def enrollment_edit(request, enrollment_id):
    center = get_active_center(request)
    qs = Enrollment.all_objects.select_related("student", "group", "group__category_obj", "course")
    if center:
        qs = qs.filter(
            Q(center=center)
            | Q(center__isnull=True, group__center=center)
            | Q(center__isnull=True, student__center=center)
        )

    enr = get_object_or_404(qs, id=enrollment_id)
    # student_enrollments uchun faqat o'chirilmagan (is_deleted=False) aktiv enrollmentlar
    active_qs = Enrollment.objects.select_related("student", "group", "group__category_obj", "course")
    if center:
        active_qs = active_qs.filter(
            Q(center=center)
            | Q(center__isnull=True, group__center=center)
            | Q(center__isnull=True, student__center=center)
        )
    student_enrollments = list(
        active_qs.filter(
            student=enr.student,
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).order_by("group__category_obj__name", "group__nom", "id")
    )
    if all(item.id != enr.id for item in student_enrollments):
        student_enrollments.append(enr)

    enrollment_map = {item.id: item for item in student_enrollments}
    enrollment_by_group_id = {item.group_id: item for item in student_enrollments if item.group_id}
    enrollment_catalog = _student_enrollment_catalog(student_enrollments)

    all_groups = Group.objects.select_related("category_obj")
    if center:
        all_groups = all_groups.filter(center=center)

    next_url = request.POST.get("next") or request.GET.get("next") or reverse("education:qarzdorlar_home")
    month_str = (request.GET.get("month") or request.POST.get("month") or "").strip()
    start_month = parse_month_yyyy_mm(month_str) or first_day_of_current_month()
    lesson_pattern_options = _lesson_pattern_options()

    def _resolve_active_enrollment(candidate_id):
        return enrollment_map.get(candidate_id) or enr

    def _build_edit_context(active_enrollment, *, teacher_share_only_checked: bool):
        # preview_month: URL'dagi ?month= yoki joriy oy (enrollment boshlanish oyiga bog'liq emas)
        preview_month = start_month
        pricing_preview = tuition_month_preview(active_enrollment, preview_month)

        # Tugash sanasi: saqlangan last_lesson_date ni ko'rsatamiz.
        # Faqat joriy billing oyi yoki kelajakdagi sana bo'lsa ishlatamiz —
        # o'tgan oydagi eski noto'g'ri qiymatlar e'tiborga olinmaydi.
        _billing_month_start = month_first_day(pricing_preview["month"])
        _billing_month_end = month_last_day(pricing_preview["month"])
        _saved_end = getattr(active_enrollment, "last_lesson_date", None)
        if _saved_end and _saved_end >= _billing_month_start:
            pricing_preview = _apply_period_end_to_preview(pricing_preview, _saved_end)
            _period_end_date = _saved_end
        else:
            _period_end_date = _billing_month_end

        # Joriy oy to'langan/to'lanmagan holatini tekshiramiz
        _month_paid = int(get_month_paid(active_enrollment, preview_month) or 0)
        _fee = int(pricing_preview.get("fee_amount", 0) or 0)
        _remaining_debt = max(0, _fee - _month_paid)
        _month_label = pricing_preview.get("month_label_uz", "").upper()
        if _fee > 0 and _month_paid >= _fee:
            pricing_preview["debt_label_uz"] = f"{_month_label} OYI TO'LANGAN"
            pricing_preview["fee_amount"] = 0
            pricing_preview["fee_amount_display"] = format_money(0)
            pricing_preview["is_month_paid"] = True
        else:
            pricing_preview["debt_label_uz"] = f"{_month_label} OYI QARZI"
            pricing_preview["fee_amount"] = _remaining_debt
            pricing_preview["fee_amount_display"] = format_money(_remaining_debt)
            pricing_preview["is_month_paid"] = False

        group_options = [
            {
                "group_id": row["group_id"],
                "group_name": row["group_name"],
                "enrollment_id": row["enrollment_id"],
            }
            for row in enrollment_catalog.get("enrollments", [])
        ]
        remaining_lessons_value = (
            active_enrollment.remaining_lessons_override
            if active_enrollment.remaining_lessons_override is not None
            else int(pricing_preview["lesson_count"] or 0)
        )
        lesson_plan = calculate_lessons(
            start_date=pricing_preview["start_date"],
            remaining_lessons=remaining_lessons_value,
            pattern=pricing_preview["lesson_pattern"],
            from_date=timezone.localdate(),
            group=active_enrollment.group,
        )
        pricing_preview = _apply_lesson_count_breakdown(
            pricing_preview,
            active_enrollment,
            remaining_lessons_value,
        )
        # Faqat so'nggi 2 oy ko'rsatiladi: joriy oy + bitta oldingi oy.
        # Mart va undan avvalgi qarzlar UI'da yashiriladi (DB'da saqlanadi).
        if getattr(active_enrollment, "id", None):
            _cum_snapshots = calculate_enrollment_debt_snapshots(
                [active_enrollment],
                [preview_month],
                cumulative_up_to=preview_month,
            )
            _cum_snap = _cum_snapshots.get(active_enrollment.id, {})
            credit = int(_cum_snap.get("credit_balance", 0) or 0)

            # Faqat bir oldingi oy (aprel) qarzini ko'rsatamiz
            prev_month = _add_month(month_first_day(preview_month), -1)
            prev_tm = TuitionMonth.objects.filter(
                enrollment=active_enrollment,
                month=prev_month,
                is_deleted=False,
            ).first()
            if prev_tm:
                prev_fee = int(getattr(prev_tm, "fee_amount", 0) or 0)
                prev_paid = int(get_month_paid(active_enrollment, prev_month) or 0)
                previous_unpaid_1month = max(0, prev_fee - prev_paid)
            else:
                previous_unpaid_1month = 0

            current_fee = int(pricing_preview.get("fee_amount", 0) or 0)
            cumulative_2months = current_fee + previous_unpaid_1month
            pricing_preview["previous_unpaid"] = previous_unpaid_1month
            pricing_preview["cumulative_debt"] = cumulative_2months
            pricing_preview["credit_balance"] = credit
            pricing_preview["net_cumulative_debt"] = max(0, cumulative_2months - credit)
        else:
            pricing_preview["previous_unpaid"] = 0
            pricing_preview["cumulative_debt"] = int(pricing_preview.get("fee_amount", 0) or 0)
            pricing_preview["credit_balance"] = 0
            pricing_preview["net_cumulative_debt"] = int(pricing_preview.get("fee_amount", 0) or 0)
        return {
            "enr": active_enrollment,
            "groups": group_options,
            "selected_group_id": active_enrollment.group_id,
            "active_enrollment_id": active_enrollment.id,
            "active_enrollment_stored_pattern": getattr(
                active_enrollment,
                "lesson_pattern",
                Enrollment.LESSON_PATTERN_GROUP,
            ),
            "next": next_url,
            "month": month_str,
            "teacher_share_only_checked": teacher_share_only_checked,
            "pricing_preview": pricing_preview,
            "lesson_pattern_options": lesson_pattern_options,
            "selected_lesson_pattern": pricing_preview["lesson_pattern"],
            "remaining_lessons_value": remaining_lessons_value,
            "lesson_plan": lesson_plan,
            "period_end_date": _period_end_date,
            "course_price_display": format_money(active_enrollment.kurs_narhi),
            "enrollment_catalog": enrollment_catalog,
        }

    def _render_enrollment_edit_form(*, teacher_share_only_checked: bool, active_enrollment=None):
        return render(
            request,
            "education/enrollment_edit.html",
            _build_edit_context(
                active_enrollment or enr,
                teacher_share_only_checked=teacher_share_only_checked,
            ),
        )

    if (
        request.method == "GET"
        and request.headers.get("x-requested-with") == "XMLHttpRequest"
        and request.GET.get("preview") == "1"
    ):
        active_enrollment = _resolve_active_enrollment(
            _parse_int_value(request.GET.get("active_enrollment_id"), enr.id)
        )
        selected_group_id = _parse_int_value(request.GET.get("group_id"), active_enrollment.group_id)
        if selected_group_id and selected_group_id not in enrollment_by_group_id:
            return JsonResponse({"error": "Bu o'quvchi ushbu guruhga biriktirilmagan."}, status=400)
        selected_group = enrollment_by_group_id.get(selected_group_id, active_enrollment).group
        joined_at = (
            parse_date((request.GET.get("joined_at") or request.GET.get("start_date") or "").strip())
            or getattr(active_enrollment, "joined_at", None)
            or enrollment_start_date(active_enrollment)
        )
        lesson_pattern = request.GET.get("lesson_pattern") or getattr(active_enrollment, "lesson_pattern", None)
        course_price = int(
            _parse_int_value(
                request.GET.get("kurs_narhi"),
                getattr(active_enrollment, "kurs_narhi", 0) or 0,
            )
            or 0
        )
        teacher_percent = (
            int(getattr(selected_group, "oqituvchi_foiz", 0) or 0)
            if getattr(selected_group, "id", None) and selected_group.id != active_enrollment.group_id
            else int(
                _parse_int_value(
                    request.GET.get("oqituvchi_foiz"),
                    getattr(active_enrollment, "oqituvchi_foiz", 0) or 0,
                )
                or 0
            )
        )
        monthly_lessons = int(
            _parse_int_value(
                request.GET.get("monthly_lessons"),
                getattr(active_enrollment, "monthly_lessons", 0)
                or getattr(selected_group, "oy_dars_soni", 0)
                or 12,
            )
            or 12
        )
        teacher_share_only = _parse_bool_value(request.GET.get("teacher_share_only"))
        missing = object()
        payable_raw = request.GET.get("student_payable_amount", missing)
        if teacher_share_only:
            student_payable_amount = round(course_price * teacher_percent / 100)
        elif payable_raw is missing:
            student_payable_amount = getattr(active_enrollment, "student_payable_amount", None)
        elif payable_raw in (None, "", "None"):
            student_payable_amount = None
        else:
            student_payable_amount = _parse_int_value(
                payable_raw,
                getattr(active_enrollment, "student_payable_amount", None),
            )

        preview_enrollment = _build_tuition_preview_enrollment(
            base_enrollment=active_enrollment,
            group=selected_group,
            start_date=joined_at,
            lesson_pattern=lesson_pattern,
            monthly_lessons=monthly_lessons,
            course_price=course_price,
            teacher_percent=teacher_percent,
            student_payable_amount=student_payable_amount,
        )
        preview = tuition_month_preview(
            preview_enrollment,
            start_month,
        )
        period_end = _parse_period_end(
            request.GET.get("period_end_date") or request.GET.get("end_date"),
            preview["month"],
        )
        preview = _apply_period_end_to_preview(preview, period_end)
        remaining_raw = request.GET.get("remaining_lessons")
        if remaining_raw in (None, ""):
            remaining_lessons = (
                getattr(active_enrollment, "remaining_lessons_override", None)
                if getattr(active_enrollment, "remaining_lessons_override", None) is not None
                else int(preview["lesson_count"] or 0)
            )
        else:
            try:
                remaining_lessons = validate_remaining_lessons(remaining_raw)
            except ValidationError as exc:
                return JsonResponse({"error": exc.messages[0]}, status=400)

        lesson_plan = calculate_lessons(
            start_date=preview["start_date"],
            remaining_lessons=remaining_lessons,
            pattern=preview["lesson_pattern"],
            from_date=timezone.localdate(),
            group=selected_group,
        )
        preview = _apply_lesson_count_breakdown(preview, preview_enrollment, remaining_lessons)
        return JsonResponse(
            {
                "preview": _serialize_tuition_preview(preview),
                "lesson_plan": _serialize_lesson_plan(lesson_plan),
            }
        )

    if request.method == "POST":
        active_enrollment = _resolve_active_enrollment(
            _parse_int_value(request.POST.get("active_enrollment_id"), enr.id)
        )
        student_ism = request.POST.get("ism", "").strip()
        student_familya = request.POST.get("familya", "").strip()
        student_telefon1 = request.POST.get("telefon1", "").strip()

        gid = _parse_int_value(request.POST.get("group_id"))
        selected_group = active_enrollment.group
        old_group_id = active_enrollment.group_id
        if gid:
            target_enrollment = enrollment_by_group_id.get(gid)
            if target_enrollment:
                if target_enrollment.id != active_enrollment.id:
                    active_enrollment = target_enrollment
                    selected_group = active_enrollment.group
                    old_group_id = active_enrollment.group_id
            else:
                messages.error(request, "Bu o'quvchi ushbu guruhga biriktirilmagan.")
                return _render_enrollment_edit_form(
                    teacher_share_only_checked=_parse_bool_value(request.POST.get("teacher_share_only")),
                    active_enrollment=active_enrollment,
                )

        new_price = int(_parse_int_value(request.POST.get("kurs_narhi"), 0) or 0)
        active_enrollment.kurs_narhi = new_price

        oqf = _parse_int_value(request.POST.get("oqituvchi_foiz"), None)
        if getattr(selected_group, "id", None) and selected_group.id != old_group_id:
            active_enrollment.oqituvchi_foiz = int(getattr(selected_group, "oqituvchi_foiz", 0) or 0)
        elif oqf is not None:
            active_enrollment.oqituvchi_foiz = int(oqf)

        joined_at_raw = (request.POST.get("joined_at") or request.POST.get("start_date") or "").strip()
        if joined_at_raw and parse_date(joined_at_raw) is None:
            messages.error(request, "Boshlanish sanasi noto'g'ri kiritildi.")
            return _render_enrollment_edit_form(
                teacher_share_only_checked=_parse_bool_value(request.POST.get("teacher_share_only")),
                active_enrollment=active_enrollment,
            )
        joined_at = parse_date(joined_at_raw) if joined_at_raw else None
        schedule_meta = resolve_lesson_schedule(
            joined_at or active_enrollment.joined_at or timezone.localdate(),
            request.POST.get("lesson_pattern") or getattr(active_enrollment, "lesson_pattern", None),
        )
        active_enrollment.joined_at = schedule_meta["start_date"]
        active_enrollment._tuition_start_date = active_enrollment.joined_at
        active_enrollment.lesson_pattern = schedule_meta["lesson_pattern"]
        if schedule_meta["adjustment_note"]:
            messages.info(request, schedule_meta["adjustment_note"])
        # Guruh oy_dars_soni — yagona manba', stale enrollment qiymatini yangilash
        monthly_lessons_raw = (request.POST.get("monthly_lessons") or "").strip()
        try:
            active_enrollment.monthly_lessons = int(
                getattr(selected_group, "oy_dars_soni", 0)
                or monthly_lessons_raw
                or getattr(active_enrollment, "monthly_lessons", 0)
                or 12
            )
        except (TypeError, ValueError):
            active_enrollment.monthly_lessons = getattr(selected_group, "oy_dars_soni", 0) or 12
        active_enrollment.pricing_type = (
            Enrollment.PRICING_PRORATED
            if active_enrollment.joined_at and active_enrollment.joined_at.day > 1
            else Enrollment.PRICING_FULL
        )

        payable_raw = (request.POST.get("student_payable_amount") or "").replace(" ", "").replace(",", "").strip()
        teacher_share_only = _parse_bool_value(request.POST.get("teacher_share_only"))
        if teacher_share_only:
            active_enrollment.student_payable_amount = round(
                full_course_amount(active_enrollment) * (active_enrollment.oqituvchi_foiz or 0) / 100
            )
        elif payable_raw == "":
            active_enrollment.student_payable_amount = None
        else:
            try:
                active_enrollment.student_payable_amount = int(payable_raw)
            except (TypeError, ValueError):
                messages.error(request, "O'quvchidan olinadigan summa noto'g'ri kiritildi.")
                return _render_enrollment_edit_form(
                    teacher_share_only_checked=teacher_share_only,
                    active_enrollment=active_enrollment,
                )

        remaining_lessons_raw = (request.POST.get("remaining_lessons_override") or "").strip()
        if remaining_lessons_raw == "":
            active_enrollment.remaining_lessons_override = None
        else:
            try:
                active_enrollment.remaining_lessons_override = validate_remaining_lessons(remaining_lessons_raw)
            except ValidationError as exc:
                messages.error(request, exc.messages[0])
                return _render_enrollment_edit_form(
                    teacher_share_only_checked=teacher_share_only,
                    active_enrollment=active_enrollment,
                )

        preview_enrollment = _build_tuition_preview_enrollment(
            base_enrollment=active_enrollment,
            group=selected_group,
            start_date=active_enrollment.joined_at,
            lesson_pattern=active_enrollment.lesson_pattern,
            monthly_lessons=active_enrollment.monthly_lessons,
            course_price=active_enrollment.kurs_narhi,
            teacher_percent=active_enrollment.oqituvchi_foiz,
            student_payable_amount=active_enrollment.student_payable_amount,
        )
        preview = tuition_month_preview(
            preview_enrollment,
            start_month,
        )
        period_end = _parse_period_end(
            request.POST.get("period_end_date") or request.POST.get("end_date"),
            preview["month"],
        )
        preview = _apply_period_end_to_preview(preview, period_end)
        billing_lesson_count = (
            active_enrollment.remaining_lessons_override
            if active_enrollment.remaining_lessons_override is not None
            else int(preview["lesson_count"] or 0)
        )
        billing_preview = _apply_lesson_count_breakdown(
            preview,
            preview_enrollment,
            billing_lesson_count,
        )
        lesson_plan = calculate_lessons(
            start_date=preview["start_date"],
            remaining_lessons=billing_lesson_count,
            pattern=preview["lesson_pattern"],
            from_date=timezone.localdate(),
            group=selected_group,
        )
        # period_end ni to'g'ridan-to'g'ri saqlaymiz (lesson_plan oxirgi sanasi emas,
        # chunki period_end kelajak oyda bo'lishi mumkin)
        active_enrollment.last_lesson_date = period_end

        try:
            active_enrollment.full_clean()
        except ValidationError as exc:
            error_messages = []
            for messages_list in exc.message_dict.values():
                error_messages.extend(messages_list)
            messages.error(request, " ".join(error_messages))
            return _render_enrollment_edit_form(
                teacher_share_only_checked=teacher_share_only,
                active_enrollment=active_enrollment,
            )

        active_enrollment.student.ism = student_ism
        active_enrollment.student.familya = student_familya
        active_enrollment.student.telefon1 = student_telefon1
        active_enrollment.student.save(update_fields=["ism", "familya", "telefon1"])

        active_enrollment.save()

        open_history = StudentGroupHistory.objects.filter(
            student=active_enrollment.student,
            group=active_enrollment.group,
            end_date__isnull=True,
        ).order_by("-start_date").first()
        if open_history:
            open_history.start_date = active_enrollment.joined_at
            open_history.kurs_narxi = active_enrollment.kurs_narhi
            open_history.oqituvchi_foiz = active_enrollment.oqituvchi_foiz
            open_history.save(update_fields=["start_date", "kurs_narxi", "oqituvchi_foiz"])
        else:
            StudentGroupHistory.objects.create(
                student=active_enrollment.student,
                group=active_enrollment.group,
                center=active_enrollment.center,
                start_date=active_enrollment.joined_at,
                kurs_narxi=active_enrollment.kurs_narhi,
                oqituvchi_foiz=active_enrollment.oqituvchi_foiz,
            )

        sync_tuition_fee(
            enrollment=active_enrollment,
            new_fee=effective_student_payable_amount(active_enrollment),
            start_month=month_first_day(active_enrollment.joined_at or start_month),
        )
        billing_month = month_first_day(active_enrollment.joined_at or start_month)
        fee_field = tuition_month_fee_field()
        # TuitionMonth fee: o'quvchidan real olinadigan summa (student_payable_amount hisobga olinadi).
        # billing_preview["fee_amount"] full_course_amount asosida — individual chegirma bo'lsa
        # effective narx bilan qayta hisoblaymiz.
        _eff_price = int(effective_student_payable_amount(active_enrollment) or 0)
        _full_price = int(full_course_amount(active_enrollment) or 0)
        if _eff_price != _full_price:
            _ml = int(
                getattr(active_enrollment, "monthly_lessons", 0)
                or getattr(selected_group, "oy_dars_soni", 0)
                or 12
            )
            _eff_bd = tuition_amount_breakdown(
                active_enrollment,
                billing_lesson_count,
                course_price=_eff_price,
                monthly_lessons=_ml,
                teacher_percent=int(getattr(active_enrollment, "oqituvchi_foiz", 0) or 0),
            )
            _tuition_fee = int(_eff_bd["fee_amount"])
        else:
            _tuition_fee = int(billing_preview["fee_amount"] or 0)
        tuition_month, _ = TuitionMonth.all_objects.update_or_create(
            enrollment=active_enrollment,
            month=billing_month,
            defaults={
                "center": active_enrollment.center,
                fee_field: _tuition_fee,
            },
        )
        if tuition_month.is_deleted:
            tuition_month.restore()

        messages.success(request, "O'quvchi ma'lumotlari muvaffaqiyatli yangilandi!")
        if get_student_total_debt(active_enrollment.student, center) <= 0:
            next_url = reverse("education:tolovlar_home")
        return redirect(next_url)

    return _render_enrollment_edit_form(
        teacher_share_only_checked=_is_teacher_share_only_enrollment(enr),
        active_enrollment=_resolve_active_enrollment(
            _parse_int_value(request.GET.get("active_enrollment_id"), enr.id)
        ),
    )

@login_required
@require_http_methods(["GET", "POST"])
def enrollment_delete(request, enrollment_id: int):
    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    # from core.tenant import get_request_center
    center = get_active_center(request)
    qs = Enrollment.all_objects.select_related("student", "group")
    if center:
        qs = qs.filter(
            Q(center=center)
            | Q(center__isnull=True, group__center=center)
            | Q(center__isnull=True, student__center=center)
        )

    enr = get_object_or_404(qs, id=enrollment_id)
    next_url = request.GET.get("next") or request.POST.get("next") or "education:tolovlar_home"

    if request.method == "POST":
        student_name = f"{enr.student.ism} {enr.student.familya}"
        group_name = getattr(enr.group, "nom", "")
        keep_in_group = request.POST.get("keep_in_group") == "1"
        # Oy parametri: berilsa, faqat o'sha oyga ta'sir qiladi (aprelni
        # tozalasangiz may tegmaydi). Berilmasa — barcha oylar (eski xulq).
        month_str = (request.POST.get("month") or "").strip()
        target_month = parse_month_str(month_str) if month_str else None

        if keep_in_group:
            with transaction.atomic():
                # all_objects — is_deleted=True bo'lganlarni ham ko'rsatadi,
                # agar avval o'chirilgan bo'lsa, ustiga qayta o'chirish xavfsiz.
                tm_qs = TuitionMonth.all_objects.filter(enrollment=enr)
                if target_month is not None:
                    tm_qs = tm_qs.filter(month=target_month)

                # Faqat hali o'chirilmagan recordlar
                alive_tm_ids = list(
                    tm_qs.filter(is_deleted=False).values_list("id", flat=True)
                )

                # Virtual TuitionMonth case: DBda yozuv yo'q bo'lsa ham o'quvchi
                # qarzdor ko'rinadi. Shu holda sentinel yozuv yaratib o'chiramiz —
                # calculate_enrollment_debt_snapshots bundan keyin fee=0 deb oladi.
                # deleted_reason="manual_cleared" ensure_tuition_month'ni qayta
                # tiklamasligiga ishora beradi.
                if target_month is not None and not tm_qs.exists():
                    from education.services.tuition import prorated_monthly_fee, tuition_month_fee_field
                    fee_field_name = tuition_month_fee_field()
                    fee_val = int(prorated_monthly_fee(enr, target_month) or 0)
                    sentinel_tm = TuitionMonth(
                        enrollment=enr,
                        month=target_month,
                        center=getattr(enr, "center", None),
                        is_deleted=True,
                        deleted_at=timezone.now(),
                        deleted_by=request.user,
                        deleted_reason="manual_cleared",
                    )
                    setattr(sentinel_tm, fee_field_name, fee_val)
                    sentinel_tm.save()

                # 1) Shu oy(lar)ga tegishli PaymentAllocation'larni soft-delete.
                if alive_tm_ids:
                    PaymentAllocation.objects.filter(
                        tuition_month_id__in=alive_tm_ids,
                        is_deleted=False,
                    ).update(is_deleted=True, deleted_at=timezone.now(), deleted_by=request.user)

                # 2) Qaysi Payment'lar hech qanday aktiv allocation'siz qoldi —
                #    ularni o'chirmay, summa=0 qilib saqlaymiz. Shunday qilsak
                #    o'quvchi to'lovlar bo'limida "0 so'm" to'lagan deb ko'rinadi
                #    va yo'qolib ketmaydi.
                payments_to_zero = Payment.objects.filter(
                    enrollment=enr, is_deleted=False
                )
                _today = timezone.localdate().isoformat()
                for p in payments_to_zero:
                    has_active_allocations = p.allocations.filter(is_deleted=False).exists()
                    if not has_active_allocations:
                        _note = (p.note or "").strip()
                        p.summa = 0
                        p.cash_amount = 0
                        p.card_amount = 0
                        p.note = (_note + f" [To'lov tozalandi: {_today}]").strip()
                        p.save(update_fields=["summa", "cash_amount", "card_amount", "note"])

                # 3) Alive TuitionMonth'larni soft-delete.
                # deleted_reason="manual_cleared" ensure_tuition_month'ni qayta
                # tiklamasligiga ishora beradi.
                if alive_tm_ids:
                    TuitionMonth.objects.filter(id__in=alive_tm_ids).update(
                        is_deleted=True,
                        deleted_at=timezone.now(),
                        deleted_by=request.user,
                        deleted_reason="manual_cleared",
                    )

            if target_month is not None:
                messages.success(
                    request,
                    f"{target_month:%Y-%m} oyi uchun to'lov yozuvlari o'chirildi. "
                    f"{student_name} ({group_name}) guruhda qoldi.",
                )
            else:
                messages.success(
                    request,
                    f"To'lov yozuvlari (barcha oylar) o'chirildi. "
                    f"{student_name} ({group_name}) guruhda qoldi.",
                )
        else:
            with transaction.atomic():
                TuitionMonth.objects.filter(
                    enrollment=enr, is_deleted=False
                ).update(
                    is_deleted=True,
                    deleted_at=timezone.now(),
                    deleted_by=request.user,
                    deleted_reason="manual_cleared",
                )
                PaymentAllocation.objects.filter(
                    tuition_month__enrollment=enr, is_deleted=False
                ).update(is_deleted=True, deleted_at=timezone.now(), deleted_by=request.user)
                enr.delete(deleted_by=request.user)
            messages.success(request, f"🗑️ {student_name} ({group_name}) guruhdan o'chirildi.")
        return redirect(next_url)

    return render(request, "education/enrollment_delete_confirm.html", {"enr": enr, "next": next_url})



 


@login_required
def payment_history_enrollment(request, enrollment_id: int):
    if not user_can_manage_payments(request.user):
        return JsonResponse({"error": "forbidden"}, status=403)

    today = timezone.localdate()
    cur_month = today.replace(day=1)
    
    month_str = request.GET.get("month", "")
    selected_month = parse_month_str(month_str) or cur_month

    center = get_active_center(request)
    qs = Enrollment.objects.select_related("student", "group")
    if center:
        qs = qs.filter(center=center)

    enrollment = get_object_or_404(qs, id=enrollment_id)
    fee_field = tuition_month_fee_field()

    # 1. Summary Data (up to today)
    ensure_tuition_month(enrollment, cur_month)
    tms_summary = TuitionMonth.objects.filter(enrollment=enrollment, month__lte=cur_month)
    agg_summary = tms_summary.aggregate(
        total_fee=Coalesce(Sum(fee_field), 0),
        total_paid=Coalesce(Sum("allocations__amount"), 0)
    )
    total_fee_needed = agg_summary["total_fee"] or 0
    total_paid_so_far = agg_summary["total_paid"] or 0
    overall_debt = total_fee_needed - total_paid_so_far

    # 2. Monthly Breakdown (All months including future if they have allocations)
    breakdown = []
    # Get all tuition months for this enrollment
    all_tms = TuitionMonth.objects.filter(enrollment=enrollment).order_by("month")
    
    for tm in all_tms:
        tm_fee = getattr(tm, fee_field, 0) or 0
        tm_paid = tm.allocations.aggregate(s=Sum("amount"))["s"] or 0
        tm_debt = max(0, tm_fee - tm_paid)
        
        breakdown.append({
            "month": tm.month.strftime("%Y-%m"),
            "fee": tm_fee,
            "paid": tm_paid,
            "debt": tm_debt,
            "is_future": tm.month > cur_month
        })

    # 3. Specific payments history
    payments_qs = Payment.objects.filter(enrollment=enrollment).order_by("-id")
    payments = []
    for p in payments_qs:
        allocations = []
        for a in p.allocations.select_related("tuition_month").all():
            allocations.append({
                "month": a.tuition_month.month.strftime("%Y-%m"),
                "amount": int(a.amount or 0),
            })

        # paid_at robust check
        paid_at_dt = getattr(p, "paid_at", None) or p.created_at
        if not paid_at_dt:
            sana = getattr(p, "sana", None)
            vaqt = getattr(p, "vaqt", None)
            if sana:
                dt = datetime.combine(sana, vaqt or datetime.min.time())
                paid_at_dt = timezone.make_aware(dt)
            else:
                paid_at_dt = timezone.now()

        cash = int(getattr(p, "cash_amount", 0) or 0)
        card = int(getattr(p, "card_amount_som", 0) or getattr(p, "card_amount", 0) or 0)
        total = int(getattr(p, "summa", 0) or (cash + card))

        payments.append({
            "id": p.id,
            "paid_at": timezone.localtime(paid_at_dt).strftime("%d.%m.%Y %H:%M"),
            "cash": cash,
            "card": card,
            "total": total,
            "allocations": allocations,
            "receipt_url": reverse("education:payment_receipt_pdf", args=[p.id]),
        })

    return JsonResponse({
        "student": enrollment.student.get_full_name(),
        "group": enrollment.group.nom,
        "monthly_fee": effective_student_payable_amount(enrollment),
        "total_fee_needed": total_fee_needed,
        "total_paid_so_far": total_paid_so_far,
        "overall_debt": overall_debt,
        "breakdown": breakdown,
        "payments": payments,
    })



from django.db.models import Sum
from datetime import date



# --------- helpers (shu fayl ichida bo'lsa sariq bo'lmaydi) ----------
def _model_has_field(model_cls, field_name: str) -> bool:
    return any(f.name == field_name for f in model_cls._meta.get_fields())


def _get_fee_amount(enrollment) -> int:
    """
    fee manbasi:
    - enrollment.student_payable_amount (agar berilgan bo'lsa)
    - aks holda to'liq kurs narxi
    """
    return effective_student_payable_amount(enrollment)


def _parse_month_str(s: str):
    try:
        if not s or len(s) < 7:
            return None
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            return None
        return date(y, m, 1)
    except Exception:
        return None

def _fmt(n: int) -> str:
    return f"{int(n or 0):,}".replace(",", " ")


def _ellipsis(text: str, limit: int = 34) -> str:
    text = (text or "").strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"


@login_required
def payment_receipt_pdf(request, payment_id: int):
    if not user_can_manage_payments(request.user):
        return HttpResponse("Forbidden", status=403)

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Payment.objects.select_related("enrollment__student", "enrollment__group")
    if center:
        qs = qs.filter(center=center)

    p = get_object_or_404(
        qs,
        id=payment_id
    )

    enrollment = getattr(p, "enrollment", None)
    student = getattr(enrollment, "student", None)
    group = getattr(enrollment, "group", None)

    cash = int(getattr(p, "cash_amount", 0) or 0)
    card_som = int(getattr(p, "card_amount_som", 0) or getattr(p, "card_amount", 0) or 0)
    total = int(getattr(p, "summa", 0) or (cash + card_som))

    # ✅ Sana/vaqt (soat)
    paid_at = getattr(p, "paid_at", None)
    if paid_at:
        paid_at = timezone.localtime(paid_at)
    else:
        sana = getattr(p, "sana", None)
        vaqt = getattr(p, "vaqt", None)
        if sana and vaqt:
            try:
                dt = datetime.combine(sana, vaqt)
                paid_at = timezone.localtime(timezone.make_aware(dt))
            except Exception:
                paid_at = timezone.localtime(timezone.now())
        else:
            paid_at = timezone.localtime(timezone.now())

    dt_text = paid_at.strftime("%d.%m.%Y %H:%M")

    # ==== Tanlangan oy bo'yicha qarz hisoblash (month=YYYY-MM) ====
    fee_field = "fee_amount"

    month_qs = request.GET.get("month")  # masalan: 2026-01
    month_date = _parse_month_str(month_qs)

    fee_for_month = None
    paid_for_month = None
    debt_for_month = None

    if enrollment and month_date:
        tm, _ = TuitionMonth.objects.get_or_create(
            enrollment=enrollment,
            month=month_date,
            defaults={fee_field: _get_fee_amount(enrollment)}
        )

        # agar fee 0 bo'lib qolgan bo'lsa, fallback bilan yangilab qo'yamiz
        cur_fee = int(getattr(tm, fee_field, 0) or 0)
        if cur_fee <= 0:
            new_fee = _get_fee_amount(enrollment)
            if new_fee > 0:
                setattr(tm, fee_field, int(new_fee))
                tm.save(update_fields=[fee_field])
                cur_fee = new_fee

        fee_for_month = int(cur_fee or 0)
        paid_for_month = int(
            PaymentAllocation.objects.filter(tuition_month=tm).aggregate(s=Sum("amount"))["s"] or 0
        )
        debt_for_month = max(0, fee_for_month - paid_for_month)

    # allocations (shu payment bo'yicha)
    alloc_mgr = getattr(p, "allocations", None)
    allocations = list(alloc_mgr.select_related("tuition_month").all()) if alloc_mgr is not None else []

    # To'lov turi
    if cash > 0 and card_som > 0:
        pay_type = "Aralash (Naqd + Karta)"
    elif card_som > 0:
        pay_type = "Kartaga o'tkazma"
    else:
        pay_type = "Naqd to'lov"

    student_name = f"{getattr(student, 'ism', '')} {getattr(student, 'familya', '')}".strip() or "-"
    group_name = getattr(group, "nom", "") or "-"

    student_name = _ellipsis(student_name, 38)
    group_name = _ellipsis(group_name, 38)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    W, H = A4

    # =========================
    #  CHEK KARTA (markazda)
    # =========================
    card_w = 175 * mm
    card_h = 240 * mm
    x = (W - card_w) / 2
    y = (H - card_h) / 2

    # oq fon
    c.setFillColor(colors.white)
    c.rect(0, 0, W, H, stroke=0, fill=1)

    # shadow (alpha bo'lsa ishlatamiz, bo'lmasa ham muammo yo'q)
    try:
        c.setFillAlpha(0.08)
    except Exception:
        pass
    c.setFillColor(colors.black)
    c.roundRect(x + 2*mm, y - 2*mm, card_w, card_h, 14, stroke=0, fill=1)
    try:
        c.setFillAlpha(1)
    except Exception:
        pass

    # card body
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.setLineWidth(1)
    c.roundRect(x, y, card_w, card_h, 14, stroke=1, fill=1)

    # Header (yorqin yashil)
    header_h = 34 * mm
    c.setFillColor(colors.HexColor("#16A34A"))
    c.roundRect(x, y + card_h - header_h, card_w, header_h, 14, stroke=0, fill=1)

    # Header text
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x + 12*mm, y + card_h - 18*mm, "ChaqmoqApp")
    c.setFont("Helvetica", 10)
    c.drawRightString(x + card_w - 12*mm, y + card_h - 18*mm, "TO'LOV CHEKI")

    # Big amount
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 32)
    c.drawCentredString(x + card_w/2, y + card_h - 60*mm, f"{_fmt(total)} so'm")

    # Success badge
    badge_w = 70 * mm
    badge_h = 10 * mm
    bx = x + (card_w - badge_w) / 2
    by = y + card_h - 74*mm
    c.setFillColor(colors.HexColor("#22C55E"))
    c.roundRect(bx, by, badge_w, badge_h, 6, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(x + card_w/2, by + 3.2*mm, "Muvaffaqiyatli")

    # Divider
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.setLineWidth(1)
    c.line(x + 12*mm, y + card_h - 82*mm, x + card_w - 12*mm, y + card_h - 82*mm)

    # Key-value row helper
    def row(label: str, value: str, yy: float, value_color=colors.HexColor("#111827")):
        c.setFillColor(colors.HexColor("#6B7280"))
        c.setFont("Helvetica", 10)
        c.drawString(x + 12*mm, yy, label)

        c.setFillColor(value_color)
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(x + card_w - 12*mm, yy, value)

        c.setStrokeColor(colors.HexColor("#F1F5F9"))
        c.setLineWidth(1)
        c.line(x + 12*mm, yy - 4.5*mm, x + card_w - 12*mm, yy - 4.5*mm)

    # Row spacing (siqilganroq, hammasi sig'sin)
    GAP = 10 * mm

    teacher = getattr(group, "oqituvchi", None)
    teacher_name = teacher.get_full_name() if teacher else "-"
    teacher_name = _ellipsis(teacher_name, 38)

    yy = y + card_h - 98*mm
    row("Tranzaksiya turi:", pay_type, yy); yy -= GAP
    row("O'quvchi:", student_name, yy); yy -= GAP
    row("Guruh:", group_name, yy); yy -= GAP
    row("O'qituvchi:", teacher_name, yy); yy -= GAP
    row("Naqd:", f"{_fmt(cash)} so'm", yy); yy -= GAP
    row("Karta:", f"{_fmt(card_som)} so'm", yy); yy -= GAP
    # row("Chek ID:", f"#{p.id}", yy); yy -= GAP

    # ✅ Oylik narx + Shu oy to'langan + Qarz
    if fee_for_month is not None:
        row("Oylik narx:", f"{_fmt(fee_for_month)} so'm", yy); yy -= GAP
        row("Shu oy to'langan:", f"{_fmt(paid_for_month)} so'm", yy); yy -= GAP
        debt_color = colors.HexColor("#EF4444") if (debt_for_month or 0) > 0 else colors.HexColor("#16A34A")
        row("Qarz (qoldiq):", f"{_fmt(debt_for_month)} so'm", yy, value_color=debt_color); yy -= GAP

    row("Sana:", dt_text, yy); yy -= 12*mm

    # Allocations title
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x + 12*mm, yy, "Taqsimot (qaysi oylarga tushdi):")
    yy -= 8*mm

    # Allocation list (qolgan joyga qarab max line)
    c.setFont("Helvetica", 10)
    bottom_limit = y + 18*mm  # footer usti
    line_h = 7 * mm
    max_lines = max(1, int((yy - bottom_limit) / line_h) - 1)

    if not allocations:
        c.setFillColor(colors.HexColor("#6B7280"))
        c.drawString(x + 14*mm, yy, "— Allocation topilmadi")
        yy -= line_h
    else:
        c.setFillColor(colors.HexColor("#0F172A"))
        for a in allocations[:max_lines]:
            tm = getattr(a, "tuition_month", None)
            m = getattr(tm, "month", None)
            enr = getattr(tm, "enrollment", None)
            g_nom = getattr(getattr(enr, "group", None), "nom", "")[:15]
            
            m_txt = m.strftime("%Y-%m") if m else "—"
            amt_txt = _fmt(int(getattr(a, "amount", 0) or 0))
            
            prefix = f"• {g_nom} ({m_txt})" if g_nom else f"• {m_txt}"
            c.drawString(x + 14*mm, yy, f"{prefix} — {amt_txt} so'm")
            yy -= line_h

        if len(allocations) > max_lines:
            c.setFillColor(colors.HexColor("#6B7280"))
            c.drawString(x + 14*mm, yy, f"... yana {len(allocations) - max_lines} ta satr bor")
            yy -= line_h

    # Footer
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 9)
    c.drawCentredString(x + card_w/2, y + 10*mm, "ChaqmoqApp • To'lov nazorati tizimi")

    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")

    # ✅ bosganda darrov yuklab olsin
    resp["Content-Disposition"] = f'attachment; filename="proskill{p.id}.pdf"'
    return resp








@login_required
def attendance_groups(request):
    q = (request.GET.get("q") or "").strip()
    
    # If the user is a teacher, force them to only see their own groups
    is_teacher = getattr(request.user, 'role', '') == 'teacher'
    if is_teacher:
        teacher_id = request.user.id
    else:
        teacher_id = _get_int(request.GET, "teacher", 0)


    # ✅ Teacher dropdown uchun
    teacher_qs = User.objects.filter(role="teacher").order_by("ism", "familya")
    
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        teacher_qs = teacher_qs.filter(center=center)
        
    teachers = teacher_qs

    # ✅ Base queryset — attendance_count limited to last 90 days for speed
    _att_since = date.today() - timedelta(days=90)
    groups = (
        Group.objects.filter(is_archived=False)
        .select_related("center", "oqituvchi")
        .annotate(
            attendance_count=Count(
                "attendances",
                filter=Q(attendances__date__gte=_att_since),
                distinct=True,
            ),
            last_attendance=Max("attendances__date"),
        )
        .annotate(
            has_attendance=Case(
                When(attendance_count__gt=0, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
    )
    
    if center:
        groups = groups.filter(center=center)

    # ✅ Filter: teacher
    if teacher_id:
        if is_teacher:
            # Support teacher ham o'z guruhlarini ko'rsin:
            # asosiy o'qituvchi (oqituvchi_id) yoki support sifatida biriktirilgan (support_teacher_id).
            # is_support_enabled tekshiruvi keraksiz — biriktirilgan bo'lsa ko'rishi kerak.
            groups = groups.filter(
                Q(oqituvchi_id=teacher_id) | Q(support_teacher_id=teacher_id)
            )
        else:
            groups = groups.filter(oqituvchi_id=teacher_id)

    # ✅ Search
    if q:
        groups = groups.filter(
            Q(nom__icontains=q) |
            Q(center__name__icontains=q) |
            Q(oqituvchi__ism__icontains=q) |
            Q(oqituvchi__familya__icontains=q)
        )

    # ✅ Davomat qilinganlar tepada, qilinmaganlar pastda
    # -has_attendance: bor guruhlar birinchi
    # last_attendance: oxirgi davomat sanasi eng yangi birinchi
    # nom: qolganlari nom bo'yicha
    groups = groups.order_by(
        "-has_attendance",
        F("last_attendance").desc(nulls_last=True),
        "nom"
    )

    # ✅ Statistikalar (tepada ko'rsatish uchun)
    total = groups.count()
    active_count = groups.filter(attendance_count__gt=0).count()
    inactive_count = total - active_count

    return render(request, "education/attendance_groups.html", {
        "groups": groups,
        "teachers": teachers,
        "selected_teacher": teacher_id,
        "q": q,
        "total": total,
        "active_count": active_count,
        "inactive_count": inactive_count,
    })

import calendar
from django.db.models import Min, Max

@login_required
def group_month_attendance(request, group_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=group_id)

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    first_day = date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    day_list = [first_day + timedelta(days=i) for i in range(days_in_month)]

    from django.db.models import Exists, OuterRef, Q

    has_attendance = Attendance.objects.filter(
        group=group,
        student=OuterRef('student'),
        date__year=year,
        date__month=month
    )

    # all_objects — is_deleted=True (guruhdan o'chirilgan) enrollment'larni ham oladi.
    # Shu oy davomati bo'lgan o'chirilgan o'quvchilar ham ko'rinib turadi.
    enrollments = (
        Enrollment.all_objects
        .filter(group=group)
        .annotate(has_att=Exists(has_attendance))
        .filter(
            Q(is_deleted=False, is_active=True)
            | Q(is_deleted=False, is_active=False, has_att=True)
            | Q(is_deleted=True, has_att=True)
        )
        .select_related("student", "group")
        .order_by("student__ism", "student__familya")
    )
    students = [e.student for e in enrollments]

    qs = Attendance.objects.filter(group=group).select_related("student")

    agg = qs.aggregate(min_date=Min("date"), max_date=Max("date"))
    if agg["min_date"] and agg["max_date"]:
        start_year = agg["min_date"].year
        end_year = agg["max_date"].year
    else:
        start_year = year - 1
        end_year = year + 1
    years = list(range(start_year, end_year + 1))


    month_qs = qs.filter(date__year=year, date__month=month)

    att_map = {(a.student_id, a.date): a for a in month_qs}

    rows = []
    for student in students:
        cells = []
        present_count = 0
        absent_count = 0
        forced_count = 0

        for d in day_list:
            a = att_map.get((student.id, d))
            if not a:
                status = "none"
            elif getattr(a, "present", False):
                status = "present"
                present_count += 1
            elif getattr(a, "status", None) == "absent_excused":
                # Sababli kelmagan — pul yozilmaydi
                status = "absent_excused"
                forced_count += 1
            elif getattr(a, "forced", False):
                # Eski "forced" yozuvlar — ko'rsatish uchun saqlanadi
                status = "forced"
                forced_count += 1
            else:
                # Sababsiz kelmagan — pul yoziladi
                status = "absent"
                absent_count += 1

            cells.append({"date": d, "status": status})

        rows.append({
            "student": student,
            "cells": cells,
            "present_count": present_count,
            "absent_count": absent_count,
            "forced_count": forced_count,
        })

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]

    # Ruxsat tekshiruvi: director, manager, teacher — barchasi tahrirlashi mumkin
    user = request.user
    can_edit = (
        user.is_superuser
        or getattr(user, "role", None) in ("director", "manager", "teacher")
    )

    return render(request, "education/group_month_attendance.html", {
        "group": group,
        "rows": rows,
        "days": day_list,
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "years": years,
        "months": months,
        "can_edit": can_edit,
    })


# ==========================================
#  QARZDORLAR (YOZILAYOTGAN YANGI PAGE)
# ==========================================

@login_required
@require_feature("finance")
def qarzdorlar_home(request):
    from core.tenant import get_request_center, require_center
    from billing.services import clear_feature_request_cache
    clear_feature_request_cache()

    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    # Multi-tenant izolyatsiya: center bo'lmasa boshqa markaz qarzdorlari
    # ko'rinib qolmasligi uchun require_center superuser'ni picker'ga,
    # boshqalarni 404 ga yuboradi.
    center = require_center(request)

    # ─── FILTERS ────────────────────────────────────────────────────────────
    q = (request.GET.get("q") or "").strip()
    group_id = _get_int(request.GET, "group", 0)
    lesson_pattern_filter = (request.GET.get("lesson_pattern_filter") or "").strip().lower()
    if lesson_pattern_filter not in {
        Enrollment.LESSON_PATTERN_ODD,
        Enrollment.LESSON_PATTERN_EVEN,
        Enrollment.LESSON_PATTERN_DAILY,
    }:
        lesson_pattern_filter = ""
    min_debt = _get_int(request.GET, "min_debt", 0)
    max_debt = _get_int(request.GET, "max_debt", 0)
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or request.GET.get("end_date") or "").strip()

    # Status filter (template'dagi <select name="status">):
    #   active   — faqat faol qarzdorlar (is_deferred=False) [default]
    #   deferred — faqat kechiktirilgan
    #   all      — ikkalasi
    selected_status = (request.GET.get("status") or "active").strip().lower()
    if selected_status not in {"active", "deferred", "all"}:
        selected_status = "active"

    allowed_page_sizes = (10, 20, 50, 100)
    per_page_raw = (request.GET.get("per_page") or "10").strip()
    try:
        per_page = int(per_page_raw)
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in allowed_page_sizes:
        per_page = 10

    # ─── JORIY OY ANIQLASH ──────────────────────────────────────────────────
    today = timezone.localdate()
    selected_from = parse_date(date_from_raw) if date_from_raw else None
    selected_to = parse_date(date_to_raw) if date_to_raw else None

    used_default_period = False
    if not selected_from and not selected_to:
        # Default: faqat JORIY OY qarzi.
        # Foydalanuvchi o'tgan oyni ko'rmoqchi bo'lsa filterdan tanlaydi
        # (Sanadan / Sanagacha yoki "Oy" select).
        used_default_period = True
        selected_from = today.replace(day=1)
        selected_to = today
    else:
        if selected_from and not selected_to:
            selected_to = today if selected_from <= today else selected_from
        elif selected_to and not selected_from:
            selected_from = selected_to.replace(day=1)
        if selected_from and selected_to and selected_from > selected_to:
            selected_to = selected_from

    sel_month = (request.GET.get("pay_month") or "").strip()
    pay_month_int = None
    if sel_month and sel_month.isdigit():
        maybe_month = int(sel_month)
        if 1 <= maybe_month <= 12:
            pay_month_int = maybe_month

    if pay_month_int:
        selected_year = selected_to.year if selected_to else today.year
        _pay_month_start = date(selected_year, pay_month_int, 1)
        selected_from = _pay_month_start
        selected_to = month_last_day(_pay_month_start)

    # Har doim selected_from va selected_to oralig'idagi oylar bo'yicha qarzni hisoblaymiz.
    from education.services.tuition import month_range_starts
    period_months = month_range_starts(selected_from, selected_to)
    _display_month = month_first_day(selected_to)
    effective_pay_month = month_first_day(selected_from)

    # ─── FAOL ENROLLMENT'LAR ─────────────────────────────────────────────────
    # Faqat:  is_active=True  +  student NOT archived  +  group NOT archived
    # Legacy data: center=None enrollment'lari group/student.center orqali
    # markazga biriktiriladi (dashboard_metrics bilan izchil).
    from django.db.models import Q as _Q
    _center_q = (
        _Q(center=center)
        | _Q(center__isnull=True, group__center=center)
        | _Q(center__isnull=True, student__center=center)
    )
    active_enrs_qs = (
        Enrollment.objects
        .select_related("student", "group", "group__oqituvchi", "group__category_obj")
        .filter(
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        )
        .filter(_center_q)
    )

    # ─── GURUHDAN CHIQARILGAN, AMMO QARZI BOR ENROLLMENT'LAR ─────────────────
    # is_active=False YOKI is_deleted=True (soft-deleted) lekin joriy PERIOD da
    # TuitionMonth yoki Davomat mavjud → ko'rsatamiz.
    # all_objects — is_deleted=True (guruhdan o'chirilgan) enrollment'larni ham oladi.
    _inactive_tm_enr_ids = set(
        TuitionMonth.objects
        .filter(
            Q(enrollment__is_active=False) | Q(enrollment__is_deleted=True),
            is_deleted=False,
            month__in=period_months,
            enrollment__student__is_archived=False,
            enrollment__group__is_archived=False,
            enrollment__group__is_deleted=False,
        )
        .values_list("enrollment_id", flat=True)
    )
    # Davomat yozilgan lekin TuitionMonth yo'q — bu oyda chiqarilgan bo'lishi mumkin
    from django.db.models import Exists, OuterRef as _OuterRef
    _inactive_att_enr_ids = set(
        Enrollment.all_objects.filter(
            Q(is_active=False) | Q(is_deleted=True),
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).filter(_center_q).annotate(
            _has_att=Exists(
                Attendance.objects.filter(
                    student_id=_OuterRef("student_id"),
                    group_id=_OuterRef("group_id"),
                    date__gte=selected_from,
                    date__lte=selected_to,
                )
            )
        ).filter(_has_att=True).values_list("id", flat=True)
    )
    _inactive_enr_ids = _inactive_tm_enr_ids | _inactive_att_enr_ids
    inactive_enrs_qs = (
        Enrollment.all_objects
        .select_related("student", "group", "group__oqituvchi", "group__category_obj")
        .filter(
            Q(is_active=False) | Q(is_deleted=True),
            id__in=_inactive_enr_ids,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        )
        .filter(_center_q)
    )

    chart_mode = "monthly"
    chart_months = _last_12_ending(selected_to)

    # ─── ENROLLMENTS (FILTER UCHUN BASE) ─────────────────────────────────────
    enrs_base = active_enrs_qs
    if group_id:
        enrs_base = enrs_base.filter(group_id=group_id)

    # PERF: Qidiruv DB darajasida (avval Python loop'da edi — har student uchun
    # alohida qidirardi). Endi student bo'yicha indekslangan filter:
    if q:
        ql_terms = q.split()
        for term in ql_terms:
            enrs_base = enrs_base.filter(
                _Q(student__ism__icontains=term)
                | _Q(student__familya__icontains=term)
                | _Q(student__telefon1__icontains=term)
                | _Q(student__telefon2__icontains=term)
            )

    active_list = list(
        enrs_base.order_by(
            "student__familya",
            "student__ism",
            "student_id",
            "group__nom",
            "id",
        )
    )

    # Inactive enrollment'lar ham xuddi shu qidiruv/guruh filtrlaridan o'tishi kerak
    inactive_base = inactive_enrs_qs
    if group_id:
        inactive_base = inactive_base.filter(group_id=group_id)
    if q:
        ql_terms = q.split()
        for term in ql_terms:
            inactive_base = inactive_base.filter(
                _Q(student__ism__icontains=term)
                | _Q(student__familya__icontains=term)
                | _Q(student__telefon1__icontains=term)
                | _Q(student__telefon2__icontains=term)
            )
    inactive_list = list(inactive_base.order_by("student__familya", "student__ism", "id"))
    for _e in inactive_list:
        _e._is_unenrolled = True

    enrollment_list = active_list + inactive_list

    # PERF: Pre-load (a) StudentGroupHistory start dates, (b) GroupSchedule
    # weekday counts. Bu N+1 muammoning ASOSIY manbasi — har enrollment uchun
    # alohida `GroupSchedule.objects.filter(group=...)` chaqiruvi qilinmaydi.
    preload_enrollment_history_starts(enrollment_list)
    from education.services.tuition import preload_group_schedules
    preload_group_schedules({e.group_id for e in enrollment_list if e.group_id})

    # auto_net_student_credits is a payment-time write operation — calling it on
    # every page load causes 5+ queries per student (N+1). It is triggered by
    # the payment save path; skip it here entirely.

    # Lazy tuzatish: lesson_pattern stale bo'lsa "group" ga o'tkaz.
    # bulk_update bilan bir so'rovda, N+1 save() o'rniga.
    _cur_month_for_recalc = today.replace(day=1)
    from education.services.tuition import ensure_tuition_month as _etm
    _pattern_stale = [
        e for e in active_list
        if getattr(e, "is_active", False)
        and int(getattr(getattr(e, "group", None), "oy_dars_soni", 0) or 0) > 0
        and e.lesson_pattern in ("odd", "even", "daily")
    ]
    if _pattern_stale:
        for _e in _pattern_stale:
            _e.lesson_pattern = Enrollment.LESSON_PATTERN_GROUP
        Enrollment.objects.bulk_update(_pattern_stale, ["lesson_pattern"])

    # ensure_tuition_month: Pre-fetch existing TuitionMonth IDs for current month
    # in ONE query, then only call _etm for enrollments that are missing one.
    _active_ids = [e.id for e in active_list if getattr(e, "is_active", False)]
    if _active_ids:
        from education.models import TuitionMonth as _TM, Payment as _Pay, PaymentAllocation as _PA
        from django.db.models import Sum as _Sum
        _existing_tm_enr_ids = set(
            _TM.objects.filter(
                enrollment_id__in=_active_ids,
                month=_cur_month_for_recalc,
                is_deleted=False,
            ).values_list("enrollment_id", flat=True)
        )
        # To'lov bor lekin allocation yo'q YOKI yetarli emas → _etm chaqirilsin
        _pay_enr_ids = set(
            _Pay.objects.filter(
                enrollment_id__in=_active_ids,
                paid_date__year=_cur_month_for_recalc.year,
                paid_date__month=_cur_month_for_recalc.month,
                is_deleted=False,
                summa__gt=0,
            ).values_list("enrollment_id", flat=True)
        )
        # "To'liq bog'langan" = allocation.amount yig'indisi >= TuitionMonth.fee
        # Qisman bog'langan (allocation < fee) bo'lsa ham _etm chaqirilishi kerak.
        if _pay_enr_ids:
            from education.services.tuition import tuition_month_fee_field as _ff
            _fee_field = _ff()
            _tm_fee_map = {
                row["enrollment_id"]: int(row[_fee_field] or 0)
                for row in _TM.objects.filter(
                    enrollment_id__in=list(_pay_enr_ids),
                    month=_cur_month_for_recalc,
                    is_deleted=False,
                ).values("enrollment_id", _fee_field)
            }
            _alloc_sum_map = {
                row["tuition_month__enrollment_id"]: int(row["paid"] or 0)
                for row in _PA.objects.filter(
                    tuition_month__enrollment_id__in=list(_pay_enr_ids),
                    tuition_month__month=_cur_month_for_recalc,
                    tuition_month__is_deleted=False,
                    payment__is_deleted=False,
                ).values("tuition_month__enrollment_id")
                .annotate(paid=_Sum("amount"))
            }
            # allocation < fee → hali to'liq bog'lanmagan (qarz ko'rinishi mumkin)
            _partially_linked_ids = {
                enr_id
                for enr_id in _pay_enr_ids
                if _alloc_sum_map.get(enr_id, 0) < _tm_fee_map.get(enr_id, 1)
            }
            _unlinked_pay_ids = _partially_linked_ids
        else:
            _unlinked_pay_ids = set()

        for e in active_list:
            if not getattr(e, "is_active", False):
                continue
            if e.id in _existing_tm_enr_ids:
                _group_price = int(getattr(getattr(e, "group", None), "kurs_narxi", 0) or 0)
                _has_custom = (
                    e.student_payable_amount is not None
                    or e.kurs_narhi != _group_price
                )
                # To'lov bor lekin allocation to'liq emas bo'lsa ham _etm chaqiramiz
                if not _has_custom and e.id not in _unlinked_pay_ids:
                    continue
            try:
                _etm(e, _cur_month_for_recalc)
            except Exception:
                pass

    # Chiqarilgan (inactive) o'quvchilar: davomat bor lekin TuitionMonth yo'q → yaratamiz
    # Davomat soni × bir dars narxi = bu oy uchun to'lov miqdori
    if inactive_list and _cur_month_for_recalc in period_months:
        from education.services.tuition import (
            attendance_based_fee as _abf,
            billable_attendance_count as _bac,
        )
        _inactive_ids = [e.id for e in inactive_list]
        _existing_inactive_tm_ids = set(
            TuitionMonth.objects.filter(
                enrollment_id__in=_inactive_ids,
                month=_cur_month_for_recalc,
                is_deleted=False,
            ).values_list("enrollment_id", flat=True)
        )
        for _ie in inactive_list:
            if _ie.id in _existing_inactive_tm_ids:
                # Inactive o'quvchi: haqiqiy davomat asosida TM fee ni yangilaylik.
                # Davomat 0 bo'lsa fee=0 → qarzdorlardan chiqib ketadi.
                try:
                    _etm(_ie, _cur_month_for_recalc)
                except Exception:
                    pass
                continue
            try:
                _att = _bac(_ie, _cur_month_for_recalc)
                if _att > 0:
                    _fee = _abf(_ie, _cur_month_for_recalc)
                    if _fee > 0:
                        TuitionMonth.objects.create(
                            enrollment=_ie,
                            month=_cur_month_for_recalc,
                            fee_amount=_fee,
                            center=_ie.center or getattr(_ie.group, "center", None),
                        )
            except Exception:
                pass

    # ── BARCHA O'QUVCHILAR: O'TGAN OYLAR UCHUN LAZY RECALCULATION ──────────────
    # Muammo: fee jadval/transfer asosida yozilgan, lekin haqiqiy davomat
    # olinmagan → davomat=0 bo'lsa fee=0 qilish kerak.
    # MUHIM: Filter oralig'idan TASHQARI oylar ham (oxirgi 3 oy) tekshiriladi —
    # default (joriy oy) filtrda ham eski noto'g'ri data tuzatilsin.
    _past_months_set = {m for m in period_months if m < _cur_month_for_recalc}
    _back = _cur_month_for_recalc
    for _ in range(3):
        _back = (_back - timedelta(days=1)).replace(day=1)
        _past_months_set.add(_back)
    _past_months = sorted(_past_months_set)
    if _past_months and enrollment_list:
        from education.services.tuition import tuition_month_fee_field as _tff
        from django.db.models import Count as _Cnt
        _fee_fld = _tff()
        _all_enr_map = {e.id: e for e in enrollment_list}

        for _pm in _past_months:
            try:
                # Fee>0 bo'lgan va himoyalanmagan TM larni topamiz (aktiv+inactive)
                _past_tms = list(
                    TuitionMonth.objects.filter(
                        enrollment_id__in=list(_all_enr_map.keys()),
                        month=_pm,
                        is_deleted=False,
                    ).exclude(
                        deleted_reason__in=["manual_cleared"]
                    ).filter(**{f"{_fee_fld}__gt": 0})
                )
                if not _past_tms:
                    continue

                _pm_end = month_last_day(_pm)
                _sg_pairs = [
                    (_all_enr_map[tm.enrollment_id].student_id,
                     _all_enr_map[tm.enrollment_id].group_id)
                    for tm in _past_tms
                    if tm.enrollment_id in _all_enr_map
                ]
                if not _sg_pairs:
                    continue

                _student_ids = list({p[0] for p in _sg_pairs})
                _group_ids = list({p[1] for p in _sg_pairs})

                # Davomat soni (barcha status — shu jumladan sababli)
                _att_any = {
                    (r["student_id"], r["group_id"]): r["cnt"]
                    for r in Attendance.objects.filter(
                        student_id__in=_student_ids,
                        group_id__in=_group_ids,
                        date__gte=_pm,
                        date__lte=_pm_end,
                    ).values("student_id", "group_id").annotate(cnt=_Cnt("id"))
                }

                # Davomat=0 → fee=0 (batch): aktiv va inactive uchun ham
                _zero_ids = [
                    tm.id
                    for tm in _past_tms
                    if tm.enrollment_id in _all_enr_map
                    and _att_any.get((
                        _all_enr_map[tm.enrollment_id].student_id,
                        _all_enr_map[tm.enrollment_id].group_id,
                    ), 0) == 0
                ]
                if _zero_ids:
                    TuitionMonth.objects.filter(id__in=_zero_ids).update(**{_fee_fld: 0})
            except Exception:
                pass

    debt_snapshots = calculate_enrollment_debt_snapshots(
        enrollment_list, period_months
    )

    # ── JAMI QARZ (QARZ ustuni) — TANLANGAN OY(LAR) bo'yicha ─────────────────
    # Har enrollment uchun faqat period_months (filterda tanlangan oy oralig'i)
    # ichidagi TuitionMonth'lar bo'yicha max(0, fee - paid).
    # MUHIM: month__in=period_months bo'lmasa QARZ ustuni BARCHA oylarni
    # yig'ib, oy filtri ishlamay qoladi (iyun tanlansa ham iyul qarzdorlari
    # ko'rinadi). Bu chart, yuqoridagi "Jami qarz" (center_month_debt_summary)
    # va per-enrollment snapshot bilan bir xil period doirasi.
    # Kelajak oy (paid==0) hisobga olinmaydi — student breakdown bilan bir xil.
    from django.db.models import Sum as _SumTot
    from education.services.tuition import tuition_month_fee_field as _fee_f_tot
    _fee_field_tot = _fee_f_tot()
    _cur_mk_debt = today.strftime("%Y-%m")
    _enr_ids_debt = [e.id for e in enrollment_list]
    _tm_fee_rows = list(
        TuitionMonth.objects
        .filter(
            enrollment_id__in=_enr_ids_debt,
            is_deleted=False,
            month__in=period_months,
        )
        .values_list("id", "enrollment_id", "month", _fee_field_tot)
    )
    _tm_paid_map = {}
    if _tm_fee_rows:
        for _r in (
            PaymentAllocation.objects
            .filter(
                tuition_month_id__in=[x[0] for x in _tm_fee_rows],
                tuition_month__is_deleted=False,
                payment__is_deleted=False,
            )
            .values("tuition_month_id")
            .annotate(paid=_SumTot("amount"))
        ):
            _tm_paid_map[_r["tuition_month_id"]] = int(_r["paid"] or 0)
    # Per-enrollment qarz (guruh kartalari uchun).
    enr_total_debt = {}
    # Per-STUDENT qarz — breakdown kabi bir oydagi barcha guruhlarni birlashtiradi
    # (bir guruhdagi ortiqcha to'lov boshqa guruh qarzini yopadi). QARZ ustuni shu.
    _enr_to_sid = {e.id: e.student_id for e in enrollment_list}
    _stu_month_fee = {}
    _stu_month_paid = {}
    for _tmid, _enrid, _mon, _fee in _tm_fee_rows:
        _paid = _tm_paid_map.get(_tmid, 0)
        if _mon.strftime("%Y-%m") > _cur_mk_debt and _paid == 0:
            continue  # kelajak oy, to'lov yo'q — breakdown ham ko'rsatmaydi
        enr_total_debt[_enrid] = enr_total_debt.get(_enrid, 0) + max(0, int(_fee or 0) - _paid)
        _sid = _enr_to_sid.get(_enrid)
        if _sid is not None:
            _k = (_sid, _mon)
            _stu_month_fee[_k] = _stu_month_fee.get(_k, 0) + int(_fee or 0)
            _stu_month_paid[_k] = _stu_month_paid.get(_k, 0) + _paid
    student_total_debt = {}
    for (_sid, _mon), _fee in _stu_month_fee.items():
        _paid = _stu_month_paid.get((_sid, _mon), 0)
        student_total_debt[_sid] = student_total_debt.get(_sid, 0) + max(0, _fee - _paid)

    # _total_debt_enrs — chart_snapshots (line below) uchun kerak.
    # active non-deferred + inactive enrollments (search/group filtersiz).
    _total_debt_enrs = list(active_enrs_qs.filter(is_deferred=False)) + list(inactive_enrs_qs)

    # ─── JAMI QARZ SUMMASI ───────────────────────────────────────────────────
    # YAGONA MANBA: center_month_debt_summary — Director dashboard ham AYNAN shu
    # funksiyani ishlatadi, shuning uchun ikkala raqam 100% bir xil bo'ladi.
    try:
        from education.services.tuition import center_month_debt_summary as _cmds
        total_center_debt, _ = _cmds(center, period_months)
    except Exception:
        total_center_debt = 0

    # ─── STUDENT MAP (student bo'yicha guruhlash) ────────────────────────────
    student_map = {}   # {student_id: row_dict}

    for e in enrollment_list:
        sid  = e.student_id
        snapshot = debt_snapshots.get(e.id, {})
        # QARZ = o'quvchining BARCHA to'lanmagan oylari yig'indisi (breakdown
        # "Jami qarz" bilan aynan bir xil), faqat tanlangan oy emas.
        debt = int(enr_total_debt.get(e.id, snapshot.get("debt", 0)) or 0)
        _e_unenrolled = getattr(e, "_is_unenrolled", False)
        if _e_unenrolled and debt <= 0:
            continue
        f    = int(snapshot.get("total_fee", 0) or 0)
        p    = int(snapshot.get("total_paid", 0) or 0)
        # lesson_count: jadval bo'yicha haqiqiy dars soni (12 yoki 13).
        # Hisob-kitob denominatori har doim 12 (tuition.py da belgilangan).
        lesson_count = int(snapshot.get("lesson_count", 0) or 0)
        enr_credit = int(snapshot.get("credit_balance", 0) or 0)
        # debt endi kumulativ (o'tgan oylarni ham o'z ichiga oladi) — "O'tgan"
        # satrini alohida ko'rsatmaymiz, aks holda ikki marta sanaladi.
        prev_unpaid = 0
        start_date = enrollment_start_date(e)
        pattern_value = enrollment_lesson_pattern(e)
        pattern_label = lesson_pattern_label(pattern_value)

        _e_unenrolled = getattr(e, "_is_unenrolled", False)
        if sid not in student_map:
            student_map[sid] = {
                "student":     e.student,
                "group_names": [],
                "lesson_pattern_names": [],
                "lesson_pattern_values": [],
                "group_cards": [],
                "total_fee":   0,
                "total_paid":  0,
                "debt":        0,
                "previous_unpaid": 0,
                "credit_balance": 0,
                "lesson_count": 0,
                "start_date":  start_date,
                "enrollment_count": 0,
                "created_at":  e.created_at,
                "enrollment":  e,
                "debt_enrollment_ids": [],
                "primary_debt_enrollment": None,
                "deferred_enrollment": None,
                "is_deferred": False,
                "has_unenrolled_debt": False,
                "teacher_share_only_debt": 0,
                "teacher_share_only_full_total": 0,
                "teacher_share_only_payment_enrollment_id": None,
                "teacher_share_only_unpaid_count": 0,
                "group":       e.group,
                "staff":       getattr(e.group, "oqituvchi", None),
            }

        row = student_map[sid]
        row["enrollment_count"] += 1
        row["total_fee"]       += f
        row["total_paid"]      += p
        row["debt"]            += debt
        row["previous_unpaid"] += prev_unpaid
        row["credit_balance"]  += enr_credit
        row["lesson_count"] += lesson_count
        if _e_unenrolled and debt > 0:
            row["has_unenrolled_debt"] = True
        if start_date and (not row.get("start_date") or start_date < row["start_date"]):
            row["start_date"] = start_date
        if debt > 0:
            row["debt_enrollment_ids"].append(e.id)
            if row["primary_debt_enrollment"] is None:
                row["primary_debt_enrollment"] = e
        if e.created_at and (not row.get("created_at") or e.created_at < row["created_at"]):
            row["created_at"] = e.created_at
        if getattr(e, "is_deferred", False):
            row["is_deferred"] = True
            row["deferred_enrollment"] = e

        if e.group:
            gnom = getattr(e.group, "nom", "")
            if gnom and gnom not in row["group_names"]:
                row["group_names"].append(gnom)
            row["group_cards"].append({
                "enrollment_id": e.id,
                "group_id": e.group_id,
                "group_name": gnom or "—",
                "lesson_pattern": pattern_value,
                "lesson_pattern_label": pattern_label,
                "lesson_count": lesson_count,
                "debt_amount": debt,
                "debt_amount_display": _format_money_exact(debt),
                "fee_amount": f,
                "fee_amount_display": _format_money_exact(f),
                "start_date": start_date,
                "is_unenrolled": _e_unenrolled,
            })
        if pattern_value and pattern_value not in row["lesson_pattern_values"]:
            row["lesson_pattern_values"].append(pattern_value)
        if pattern_label and pattern_label not in row["lesson_pattern_names"]:
            row["lesson_pattern_names"].append(pattern_label)

        full_amount = full_course_amount(e)
        effective_amount = effective_student_payable_amount(e)
        teacher_share_amount = int(getattr(e, "oqituvchi_daromadi", 0) or 0)
        if (
            e.student_payable_amount not in (None, "")
            and full_amount > effective_amount
            and effective_amount == teacher_share_amount
        ):
            row["teacher_share_only_debt"] += max(0, debt)
            row["teacher_share_only_full_total"] += full_amount
            if debt > 0:
                row["teacher_share_only_unpaid_count"] += 1
                if row["teacher_share_only_payment_enrollment_id"] is None:
                    row["teacher_share_only_payment_enrollment_id"] = e.id

    # ─── GROUP LABEL ─────────────────────────────────────────────────────────
    for r in student_map.values():
        # QARZ ustuni = o'quvchining JAMI qarzi (breakdown "Jami qarz" bilan bir xil).
        # Per-enrollment yig'indi emas — bir oydagi guruhlar birlashtirilgan.
        _sid_row = getattr(r.get("student"), "id", None)
        if _sid_row is not None and _sid_row in student_total_debt:
            r["debt"] = int(student_total_debt[_sid_row])
        if r["primary_debt_enrollment"] is not None:
            r["enrollment"] = r["primary_debt_enrollment"]
            r["group"] = r["primary_debt_enrollment"].group
            r["staff"] = getattr(r["group"], "oqituvchi", None)
            r["start_date"] = enrollment_start_date(r["primary_debt_enrollment"])
        r["group_label"] = ", ".join(r["group_names"]) if r["group_names"] else "—"
        r["lesson_pattern_label"] = ", ".join(r["lesson_pattern_names"]) if r["lesson_pattern_names"] else "—"
        r["group_cards"] = sorted(
            r.get("group_cards") or [],
            key=lambda item: ((item.get("group_name") or "").lower(), item.get("enrollment_id") or 0),
        )
        r["visible_group_cards"] = r["group_cards"][:2]
        r["remaining_group_card_count"] = max(0, len(r["group_cards"]) - 2)
        r["has_teacher_share_only"] = (
            r["teacher_share_only_debt"] > 0
            and r["teacher_share_only_full_total"] > r["teacher_share_only_debt"]
        )
        if r["teacher_share_only_unpaid_count"] > 1:
            r["teacher_share_only_payment_enrollment_id"] = None
        r["payment_amount"] = r["teacher_share_only_debt"] if r["has_teacher_share_only"] else r["debt"]
        r["payment_scope"] = "teacher_share_only" if r["has_teacher_share_only"] else "student_total"
        debt_enrollment_ids = r.get("debt_enrollment_ids") or []
        r["payment_enrollment_id"] = debt_enrollment_ids[0] if len(debt_enrollment_ids) == 1 else None

    # ─── QIDIRUV: DB darajasida qilinadi (yuqorida `enrs_base.filter(...)`)
    # Bu yerda Python loop kerak emas — student_map allaqachon faqat mos
    # keluvchi enrollment'lardan tuzilgan.
    all_rows = list(student_map.values())

    def _matches_lesson_pattern_filter(row):
        if not lesson_pattern_filter:
            return True
        return lesson_pattern_filter in (row.get("lesson_pattern_values") or [])

    def _matches_status_filter(row):
        """Status select bo'yicha filtrlash: active / deferred / all."""
        if selected_status == "all":
            return True
        is_deferred = bool(row.get("is_deferred"))
        if selected_status == "deferred":
            return is_deferred
        # active (default) — kechiktirilmagan qarzdorlar
        return not is_deferred

    # Bazaviy qarzdor satrlar — barcha qarz/min/max/status filterlari qo'llangan,
    # lesson_pattern HALI qo'llanmagan (lesson_pattern badge sonlari uchun kerak).
    debt_filter_base_rows = []
    for row in all_rows:
        if not row["group_names"]:
            continue
        if row["debt"] <= 0:
            continue
        if min_debt and row["debt"] < min_debt:
            continue
        if max_debt and row["debt"] > max_debt:
            continue
        if not _matches_status_filter(row):
            continue
        debt_filter_base_rows.append(row)

    lesson_pattern_filter_counts = {
        "all": len(debt_filter_base_rows),
        Enrollment.LESSON_PATTERN_ODD: sum(
            1 for row in debt_filter_base_rows if Enrollment.LESSON_PATTERN_ODD in (row.get("lesson_pattern_values") or [])
        ),
        Enrollment.LESSON_PATTERN_EVEN: sum(
            1 for row in debt_filter_base_rows if Enrollment.LESSON_PATTERN_EVEN in (row.get("lesson_pattern_values") or [])
        ),
        Enrollment.LESSON_PATTERN_DAILY: sum(
            1 for row in debt_filter_base_rows if Enrollment.LESSON_PATTERN_DAILY in (row.get("lesson_pattern_values") or [])
        ),
    }

    # ─── STATISTIKA ──────────────────────────────────────────────────────────
    # debt_filter_base_rows allaqachon qarz/min/max/status bo'yicha filtrlangan —
    # uning ustiga faqat lesson_pattern_filter qolgan.
    debtor_rows = [row for row in debt_filter_base_rows if _matches_lesson_pattern_filter(row)]
    debtors_count = len(debtor_rows)

    # paid / no_group statistikasi (badge'lar uchun) — full set'dan hisoblanadi
    paid_count = 0
    no_group_count = 0
    for r in all_rows:
        if not _matches_lesson_pattern_filter(r):
            continue
        if not r["group_names"]:
            no_group_count += 1
        elif r["debt"] <= 0:
            paid_count += 1

    display_rows = debtor_rows

    filtered_debt   = sum(r["debt"] for r in display_rows)

    # Jami qarz: jadvaldagi ma'lumotlar bilan izchil (arxivlangan guruhlar chiqarib tashlangan),
    # status filteri qo'llangan, lekin min/max qarz filterlari qo'llanmagan.
    total_center_debt = sum(
        r["debt"] for r in all_rows
        if r.get("group_names") and r["debt"] > 0 and _matches_status_filter(r)
    )

    # Chart: Jami qarz bilan bir xil enrollments (_total_debt_enrs) ishlatamiz.
    # preload_group_schedules allaqachon yuqorida enrollment_list uchun chaqirilgan,
    # lekin _total_debt_enrs yangi guruhlarni o'z ichiga olishi mumkin — yangilash.
    from education.services.tuition import preload_group_schedules as _pgs2
    _pgs2({e.group_id for e in _total_debt_enrs if e.group_id})
    preload_enrollment_history_starts(_total_debt_enrs)
    chart_snapshots = calculate_enrollment_debt_snapshots(
        _total_debt_enrs,
        chart_months,
    )
    graph_map = {chart_month: 0 for chart_month in chart_months}
    for snapshot in chart_snapshots.values():
        month_details = snapshot.get("months", {})
        for chart_month in chart_months:
            graph_map[chart_month] += int(
                month_details.get(chart_month, {}).get("debt", 0) or 0
            )
    chart_series = [graph_map[month] for month in chart_months]
    chart_labels = [_human_month_label(month) for month in chart_months]
    chart_period_label = _human_month_period_label(chart_months[0], chart_months[-1])
    selected_period_label = _human_period_label(selected_from, selected_to)
    if used_default_period:
        selected_period_label = "Joriy oy qarzi · o'tgan oylar uchun filtrni o'zgartiring"

    # ─── PAGINATOR ───────────────────────────────────────────────────────────
    from django.core.paginator import Paginator
    paginator   = Paginator(display_rows, per_page)
    page_obj    = paginator.get_page(request.GET.get("page"))

    # ─── GURUHLAR (filter uchun) ──────────────────────────────────────────────
    groups_qs = Group.objects.filter(is_archived=False)
    if center:
        groups_qs = groups_qs.filter(center=center)

    context = {
        "page_obj":       page_obj,
        "groups":         groups_qs,
        "selected_group": group_id,
        "total_debt":     total_center_debt,
        "filtered_debt":  filtered_debt,
        "chart_data":     chart_series,
        "chart_labels":   chart_labels,
        "chart_mode":     chart_mode,
        "chart_kicker":   "Oxirgi 12 oy",
        "chart_period_label": chart_period_label,
        "selected_period_label": selected_period_label,
        "q":              q,
        "selected_lesson_pattern_filter": lesson_pattern_filter,
        "lesson_pattern_filter_counts": lesson_pattern_filter_counts,
        "min_debt":       min_debt if min_debt else "",
        "max_debt":       max_debt if max_debt else "",
        "selected_status": selected_status,
        "date_from":      selected_from.isoformat(),
        "date_to":        selected_to.isoformat(),
        "pay_month":      str(pay_month_int) if pay_month_int else "",
        "effective_pay_month": effective_pay_month.strftime("%Y-%m"),
        "per_page":       per_page,
        "page_size_options": allowed_page_sizes,
        "uz_months": [
            (1, "Yanvar"),   (2, "Fevral"),   (3, "Mart"),    (4, "Aprel"),
            (5, "May"),      (6, "Iyun"),     (7, "Iyul"),    (8, "Avgust"),
            (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
        ],
        "stats_summary": {
            "total":    sum(1 for row in all_rows if _matches_lesson_pattern_filter(row)),
            "debtors":  debtors_count,
            "paid":     paid_count,
            "no_group": no_group_count,
        },
    }

    try:
        from store.views import _ensure_default_payment_methods as _seed_pm
        from store.models import PaymentMethod as _PM
        if center:
            _seed_pm(center)
            context["payment_methods"] = list(
                _PM.objects.filter(center=center, is_active=True).order_by('nom')
            )
        else:
            context["payment_methods"] = []
    except Exception:
        context["payment_methods"] = []

    return render(request, "education/qarzdorlar.html", context)
    
import csv
from django.http import HttpResponse
from datetime import date, timedelta
import calendar


@login_required
def group_month_attendance_export(request, group_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=group_id)

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    first_day = date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    day_list = [first_day + timedelta(days=i) for i in range(days_in_month)]

    from django.db.models import Exists, OuterRef, Q

    has_attendance = Attendance.objects.filter(
        group=group,
        student=OuterRef('student'),
        date__year=year,
        date__month=month
    )

    enrollments = (
        Enrollment.all_objects
        .filter(group=group)
        .annotate(has_att=Exists(has_attendance))
        .filter(
            Q(is_deleted=False, is_active=True)
            | Q(is_deleted=False, is_active=False, has_att=True)
            | Q(is_deleted=True, has_att=True)
        )
        .select_related("student")
        .order_by("student__ism", "student__familya")
    )
    students = [e.student for e in enrollments]

    month_qs = (Attendance.objects
                .filter(group=group, date__year=year, date__month=month)
                .select_related("student"))

    att_map = {(a.student_id, a.date): a for a in month_qs}

    # ✅ Excel uchun UTF-8 BOM
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    safe_group_name = "".join(ch for ch in (group.nom or "") if ch.isalnum() or ch in ("-", "_", " "))
    safe_group_name = safe_group_name.strip() or f"group-{group.id}"
    filename = f"{safe_group_name}_{year}-{month:02d}_attendance.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")

    # ✅ MUHIM: delimiter=';' (Excel RU/UZ)
    writer = csv.writer(response, delimiter=';', lineterminator="\n", quoting=csv.QUOTE_MINIMAL)

    def _csv_safe(value):
        text = "" if value is None else str(value)
        if text.startswith(("=", "+", "-", "@")):
            return f"'{text}"
        return text

    header = ["O'quvchi"] + [d.strftime("%d-%m-%Y") for d in day_list]
    writer.writerow(header)

    for s in students:
        row = [_csv_safe(s.get_full_name())]
        for d in day_list:
            a = att_map.get((s.id, d))
            if not a:
                row.append("")  # belgilanmagan
            elif getattr(a, "present", False):
                row.append("KELDI")
            elif getattr(a, "forced", False):
                row.append("KELMADI (PUL)")
            else:
                row.append("KELMADI")
        writer.writerow(row)

    return response



@require_POST
@login_required
def attendance_toggle_cell(request, group_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=group_id)

    # Ruxsat tekshiruvi: director, manager, teacher — barchasi tahrirlashi mumkin
    user = request.user
    is_allowed = (
        user.is_superuser
        or getattr(user, "role", None) in ("director", "manager", "teacher")
    )
    if not is_allowed:
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q"}, status=403)

    student_id = request.POST.get("student_id")
    date_str = request.POST.get("date")
    target_status = request.POST.get("target_status")   # new: direct set from popover
    current_status = request.POST.get("status", "none")  # legacy cycle fallback

    d = parse_date(date_str)
    if not d or not student_id:
        return JsonResponse({"ok": False, "error": "Bad data"}, status=400)

    user_qs = User.objects.filter(role="student")
    if center:
        user_qs = user_qs.filter(center=center)

    student = get_object_or_404(user_qs, pk=student_id)

    att = Attendance.objects.filter(group=group, student=student, date=d).first()

    def _prepare_att():
        nonlocal att
        if not att:
            att = Attendance(group=group, student=student, date=d)
            if hasattr(att, "center"):
                att.center = group.center
        if not getattr(att, "teacher_id", None) and getattr(group, "oqituvchi_id", None):
            att.teacher = group.oqituvchi
        return att

    # --- Direct set mode (from popover) ---
    if target_status in ("present", "absent", "excused", "none"):
        if target_status == "none":
            if att:
                att.delete()
            return JsonResponse({"ok": True, "status": "none"})

        att = _prepare_att()
        if target_status == "present":
            att.present = True
            att.forced = False
            if hasattr(att, "status"):
                att.status = "present"
        elif target_status == "absent":
            # Sababsiz kelmagan — pul yoziladi
            att.present = False
            att.forced = False
            if hasattr(att, "status"):
                att.status = "absent_unexcused"
        elif target_status == "excused":
            # Sababli kelmagan — pul yozilmaydi
            att.present = False
            att.forced = False
            if hasattr(att, "status"):
                att.status = "absent_excused"
        att.save()
        # excused → display status is "absent_excused"
        display = "absent_excused" if target_status == "excused" else target_status
        return JsonResponse({"ok": True, "status": display})

    # --- Legacy cycle mode (backward compat) ---
    if current_status == "none":
        att = _prepare_att()
        att.present = True
        att.forced = False
        att.save()
        new_status = "present"

    elif current_status == "present":
        att = _prepare_att()
        att.present = False
        att.forced = False
        att.save()
        new_status = "absent"

    elif current_status == "absent":
        if att:
            att.delete()
        new_status = "none"

    elif current_status == "forced":
        if att:
            att.delete()
        new_status = "none"

    else:
        new_status = current_status or "none"

    return JsonResponse({"ok": True, "status": new_status})



@login_required
def points_details(request):
    student_id = request.GET.get("student")
    date_str = request.GET.get("date")
    type_ = request.GET.get("type", "plus")

    if not student_id or not date_str:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"error": "Invalid date format"}, status=400)

    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(date_obj, datetime.min.time()), tz)
    end = timezone.make_aware(datetime.combine(date_obj, datetime.max.time()), tz)

    qs = Ledger.objects.filter(student_id=student_id, sana__range=(start, end))
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        qs = qs.filter(student__center=center)

    if type_ == "plus":
        qs = qs.filter(ball__gt=0)
    elif type_ == "minus":
        qs = qs.filter(ball__lt=0)

    details = []
    for l in qs.order_by("-sana"):
        show_dt = getattr(l, "created_at", None) or l.sana
        local_dt = timezone.localtime(show_dt, tz)
        tm = local_dt.strftime("%H:%M | %d-%b-%Y")

        details.append({
            "amount": l.ball,
            "rule": str(l.rule) if l.rule else "",
            "reason": getattr(l.rule, "izoh", "") if hasattr(l.rule, "izoh") else "",
            "teacher": str(l.beruvchi) if l.beruvchi else "",
            "group": str(l.group) if l.group else "",
            "time": tm,
        })

    return JsonResponse({"details": details})


from .models import Student, Category

# education/views.py
from django.shortcuts import render
from django.db.models import Q, Sum
from accounts.models import User
from education.models import Enrollment
from .models import Payment
from datetime import date

def _last_12_month_starts():
    """Oxirgi 12 oy (shu oy ham kiradi), tartib: eski -> yangi"""
    today = timezone.localdate().replace(day=1)
    months = []
    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(date(y, m, 1))
    return months

def _month_start(d: date) -> date:
    return d.replace(day=1)

def _add_months(d: date, n: int) -> date:
    # month arithmetic
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)

def _last_12_ending(anchor: date) -> list[date]:
    # anchor included, return 12 month starts ending at anchor
    anchor = _month_start(anchor)
    return [_add_months(anchor, -11 + i) for i in range(12)]


def _chart_range_mode(from_date: date, to_date: date) -> str:
    """'daily' when range ≤ 31 days (single-month view), 'monthly' otherwise."""
    return "daily" if (to_date - from_date).days < 32 else "monthly"


def _chart_monthly_buckets_for_range(from_date: date, to_date: date) -> list[date]:
    """Return month-start dates covering [from_date, to_date] inclusive."""
    result, m = [], from_date.replace(day=1)
    end = to_date.replace(day=1)
    while m <= end:
        result.append(m)
        m = _add_months(m, 1)
    return result


def _build_daily_debt_series(enrollment_ids: list, from_date: date, to_date: date) -> list[int]:
    """
    For each day in [from_date, to_date] return the outstanding debt for
    display_month = to_date.replace(day=1).  Matches what the monthly chart
    shows for that month (fee_amount − PaymentAllocation for that month only).
    """
    if not enrollment_ids:
        return [0] * ((to_date - from_date).days + 1)

    display_month = to_date.replace(day=1)

    # Fees for display_month only (consistent with per-month chart bars)
    total_fees = int(
        TuitionMonth.objects.filter(
            enrollment_id__in=enrollment_ids,
            month=display_month,
            is_deleted=False,
        ).aggregate(s=Sum("fee_amount"))["s"] or 0
    )

    # Payments allocated to display_month TuitionMonths made BEFORE from_date
    paid_before = int(
        PaymentAllocation.objects.filter(
            tuition_month__enrollment_id__in=enrollment_ids,
            tuition_month__month=display_month,
            tuition_month__is_deleted=False,
            payment__paid_date__lt=from_date,
            is_deleted=False,
            payment__is_deleted=False,
        ).aggregate(s=Sum("amount"))["s"] or 0
    )

    # Payments allocated to display_month TuitionMonths per day within range
    daily_map: dict = {}
    for row in (
        PaymentAllocation.objects.filter(
            tuition_month__enrollment_id__in=enrollment_ids,
            tuition_month__month=display_month,
            tuition_month__is_deleted=False,
            payment__paid_date__gte=from_date,
            payment__paid_date__lte=to_date,
            is_deleted=False,
            payment__is_deleted=False,
        )
        .values("payment__paid_date")
        .annotate(total=Sum("amount"))
    ):
        daily_map[row["payment__paid_date"]] = int(row["total"] or 0)

    running, series = paid_before, []
    for i in range((to_date - from_date).days + 1):
        running += daily_map.get(from_date + timedelta(days=i), 0)
        series.append(max(0, total_fees - running))
    return series


UZ_MONTH_NAMES = {
    1: "Yanvar",
    2: "Fevral",
    3: "Mart",
    4: "Aprel",
    5: "May",
    6: "Iyun",
    7: "Iyul",
    8: "Avgust",
    9: "Sentyabr",
    10: "Oktyabr",
    11: "Noyabr",
    12: "Dekabr",
}


def _human_period_label(start_date: date, end_date: date) -> str:
    if start_date == end_date:
        return f"{start_date.day}-{UZ_MONTH_NAMES.get(start_date.month, start_date.strftime('%B'))} {start_date.year}"
    return (
        f"{start_date.day}-{UZ_MONTH_NAMES.get(start_date.month, start_date.strftime('%B'))} {start_date.year}"
        f" dan {end_date.day}-{UZ_MONTH_NAMES.get(end_date.month, end_date.strftime('%B'))} {end_date.year} gacha"
    )


def _human_month_label(month_value: date) -> str:
    return UZ_MONTH_NAMES.get(month_value.month, month_value.strftime('%B'))


def _human_month_period_label(start_month: date, end_month: date) -> str:
    if start_month == end_month:
        return _human_month_label(start_month)
    return f"{_human_month_label(start_month)} dan {_human_month_label(end_month)} gacha"


def _build_last_12_month_money_chart_series(qs, *, date_field: str, amount_field: str, anchor_date: date):
    buckets = _last_12_ending(anchor_date)
    rows = (
        qs.annotate(bucket=TruncMonth(date_field))
        .values("bucket")
        .annotate(total=Sum(amount_field))
        .order_by("bucket")
    )
    value_map = {}
    for row in rows:
        bucket = row["bucket"]
        if hasattr(bucket, "date"):
            bucket = bucket.date()
        bucket = bucket.replace(day=1)
        value_map[bucket] = int(row["total"] or 0)
    labels = [_human_month_label(bucket) for bucket in buckets]
    data = [value_map.get(bucket, 0) for bucket in buckets]
    return labels, data, "Oxirgi 12 oy", _human_month_period_label(buckets[0], buckets[-1])



def _payment_month_parts(pay_rows, alloc_map):
    """
    Har bir to'lovni oylarga bo'ladi. KAFOLAT: ulushlar yig'indisi ROPPA-ROSA
    payment.summa ga teng — hech qachon oshmaydi (eski ikki-marta-taqsimlash
    bugi buzgan yozuvlarda allocation jami summadan katta bo'lishi mumkin;
    bunda eng eski oylar olinadi, ortiqchasi tashlab yuboriladi).

    pay_rows: [(id, summa, paid_date), ...]
    alloc_map: payment_id -> [(month, amount), ...] (faqat live allocationlar)
    Returns: payment_id -> {month_first_day: ulush}
    """
    parts = {}
    for pid, summa, pd in pay_rows:
        summa = int(summa or 0)
        d = {}
        remaining = summa
        for m, amt in sorted(alloc_map.get(pid, ()), key=lambda x: x[0]):
            if remaining <= 0:
                break
            take = min(int(amt or 0), remaining)
            if take > 0:
                d[m] = d.get(m, 0) + take
                remaining -= take
        # Bog'lanmagan qoldiq → pul berilgan oyga
        if remaining > 0 and pd:
            mb = pd.replace(day=1)
            d[mb] = d.get(mb, 0) + remaining
        parts[pid] = d
    return parts


def _build_alloc_map(payment_ids):
    """payment_id -> [(month, amount), ...] — faqat live allocationlar."""
    amap = {}
    for r in PaymentAllocation.objects.filter(
        is_deleted=False, payment_id__in=list(payment_ids)
    ).values_list("payment_id", "tuition_month__month", "amount"):
        amap.setdefault(r[0], []).append((r[1].replace(day=1), int(r[2] or 0)))
    return amap


def _get_payment_dashboard_data(request):
    center = get_active_center(request)
    if not center and not request.user.is_superuser:
        raise PermissionDenied("Markaz biriktirilmagan")

    today = timezone.localdate()
    cur_month_start = today.replace(day=1)

    base_payment_qs = Payment.objects.filter(center=center) if center else Payment.objects.none()
    total_income = base_payment_qs.aggregate(s=Sum("summa"))["s"] or 0

    q = (request.GET.get("q") or "").strip()
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or "").strip()
    sel_group = request.GET.get("group") or ""
    sel_teacher = request.GET.get("teacher") or ""
    sel_course = request.GET.get("course") or ""
    sel_staff = request.GET.get("staff") or ""
    sel_type = request.GET.get("payment_type") or ""
    sel_month = request.GET.get("pay_month") or ""

    # Map any payment-method name (NAQD/KARTA/CLICK/PAYME/...) to internal
    # Payment.payment_type code (cash/card/mixed) for DB filtering.
    # `sel_type` is preserved for UI dropdown; `sel_type_filter` is used in queryset.
    _PM_NAME_TO_MODE = {
        'cash': 'cash',
        'card': 'card',
        'mixed': 'mixed',
        'naqd': 'cash',
        'karta': 'card',
        'plastik': 'card',
        'aralash': 'mixed',
    }
    sel_type_filter = ''
    if sel_type:
        sel_type_filter = _PM_NAME_TO_MODE.get(sel_type.lower(), '')

    selected_from = parse_date(date_from_raw) if date_from_raw else None
    selected_to = parse_date(date_to_raw) if date_to_raw else None

    # pay_month tanlanib, sana filtri ko'rsatilmagan bo'lsa:
    # paid_date filtrini o'sha yilning boshidan oxirigacha kengaytir.
    # (Aks holda default joriy oy date filtri pay_month bilan to'qnashib,
    # boshqa oylarda qilingan to'lovlarni yashiradi.)
    if sel_month and sel_month.isdigit() and not date_from_raw and not date_to_raw:
        # Aprel kabi o'tgan oy tanlansa, joriy yilning to'liq diapazonini ishlatamiz
        selected_from = date(today.year, 1, 1)
        selected_to = date(today.year, 12, 31)
    elif not selected_from and not selected_to:
        selected_from = cur_month_start
        selected_to = today
    else:
        if selected_from and not selected_to:
            selected_to = today if selected_from <= today else selected_from
        elif selected_to and not selected_from:
            selected_from = selected_to.replace(day=1)
        if selected_from and selected_to and selected_from > selected_to:
            selected_to = selected_from

    date_from = selected_from.isoformat()
    date_to = selected_to.isoformat()

    allocation_prefetch = Prefetch(
        "allocations",
        queryset=PaymentAllocation.objects.select_related(
            "tuition_month",
            "tuition_month__enrollment",
            "tuition_month__enrollment__group",
            "tuition_month__enrollment__group__category_obj",
        ).order_by("tuition_month__month", "id"),
        to_attr="prefetched_allocations",
    )

    pay_qs = base_payment_qs.select_related(
        "student", "group", "group__oqituvchi", "group__category_obj", "created_by"
    ).prefetch_related(allocation_prefetch)
    chart_qs = base_payment_qs.select_related(
        "student", "group", "group__oqituvchi", "group__category_obj", "created_by"
    )
    cur_year = selected_to.year

    # Oy filtri tanlanganda: shu oyning 1-sanasi (allocation month bilan solishtirish uchun)
    sel_month_first = None
    if sel_month and sel_month.isdigit():
        sel_month_first = date(cur_year, int(sel_month), 1)

    def _apply_shared_payment_filters(qs):
        if q:
            qs = qs.filter(
                Q(student__ism__icontains=q)
                | Q(student__familya__icontains=q)
                | Q(student__telefon1__icontains=q)
                | Q(student__telefon2__icontains=q)
                | Q(student__email__icontains=q)
                | Q(student__gmail__icontains=q)
            )

        if sel_group:
            qs = qs.filter(group_id=sel_group)
        if sel_teacher:
            qs = qs.filter(group__oqituvchi_id=sel_teacher)
        if sel_course:
            qs = qs.filter(group__category_obj_id=sel_course)
        if sel_staff:
            qs = qs.filter(created_by_id=sel_staff)
        if sel_type_filter:
            qs = qs.filter(payment_type=sel_type_filter)

        return qs

    pay_qs = _apply_shared_payment_filters(pay_qs)
    chart_qs = _apply_shared_payment_filters(chart_qs)

    # Sana filtri FAQAT jadval va "Filter bo'yicha" uchun.
    # Diagramma doim 12 oylik tarixni ko'rsatadi (chart_qs sana filtrisiz).
    pay_qs = pay_qs.filter(paid_date__gte=selected_from, paid_date__lte=selected_to)

    chart_anchor_date = selected_to or today
    chart_months = _last_12_ending(chart_anchor_date)
    chart_start = chart_months[0]
    chart_end = _add_months(chart_months[-1], 1) - timedelta(days=1)

    # ═══ YAGONA QOIDA: har bir so'm ROPPA-ROSA BITTA oyga tegishli ═══
    # - To'lovning oyga BOG'LANGAN qismi (live allocation) → o'sha oyda
    #   (iyunda berilgan may puli → MAY da)
    # - BOG'LANMAGAN qoldi'g'i (summa - live alloc): allocation'lari reset
    #   bilan bekor qilingan yoki umuman yaratilmagan pullar → paid_date oyida
    # Natija: oylar yig'indisi = umumiy daromad; filtr = diagramma ustuni.

    # "To'lov oyi" filtri: shu oyda ULUSHI bor to'lovlar.
    # _payment_month_parts kafolati: bir to'lov ulushlari jami = summa —
    # eski ikki-marta-taqsimlangan yozuvlarda ham oshirib sanamaydi.
    _month_part_by_pay = {}
    if sel_month_first:
        _base_rows = list(pay_qs.values_list("id", "summa", "paid_date"))
        _base_parts = _payment_month_parts(
            _base_rows, _build_alloc_map([r[0] for r in _base_rows])
        )
        _month_part_by_pay = {
            pid: d[sel_month_first]
            for pid, d in _base_parts.items()
            if d.get(sel_month_first, 0) > 0
        }
        pay_qs = pay_qs.filter(id__in=list(_month_part_by_pay.keys()))
        filtered_income = sum(_month_part_by_pay.values())

    payment_ids = list(pay_qs.values_list("id", flat=True))

    if not sel_month_first:
        filtered_income = Payment.objects.filter(id__in=payment_ids).aggregate(s=Sum("summa"))["s"] or 0
    unique_payers_count = Payment.objects.filter(id__in=payment_ids).values("student").distinct().count()

    # ── Diagramma: TO'LOV SANASI (paid_date) bo'yicha (oxirgi 12 oy, oy filtrisiz) ──
    # Har ustun = o'sha oyda HAQIQATAN kassaga tushgan pul (summa).
    # Jadval, "Filter bo'yicha" va "Umumiy daromad" bilan bir xil o'lchov:
    # pul to'langan bo'lsa ustun ko'rinadi, to'lanmagan bo'lsa 0.
    # (Ilgari "to'lov oyi"/allocation bo'yicha edi — shu sabab oldindan
    #  to'langan pullar boshqa oy ustunida ko'rinib chalkashlik chiqargan.)
    _chart_pay_rows = list(chart_qs.values_list("id", "summa", "paid_date"))
    _chart_value_map = {}
    _chart_contrib_ids = set()
    for _pid, _summa, _pd in _chart_pay_rows:
        if not _pd:
            continue
        _b = _pd.replace(day=1)
        if chart_start <= _b <= chart_end:
            _amt = int(_summa or 0)
            if _amt:
                _chart_value_map[_b] = _chart_value_map.get(_b, 0) + _amt
                _chart_contrib_ids.add(_pid)

    chart_labels = [_human_month_label(b) for b in chart_months]
    chart_data = [_chart_value_map.get(b, 0) for b in chart_months]
    chart_kicker = "Oxirgi 12 oy"
    chart_period_label = _human_month_period_label(chart_months[0], chart_months[-1])

    # Diagramma statistikasi: hissa qo'shgan to'lovlar
    chart_payment_ids = list(_chart_contrib_ids)
    chart_payment_record_count = len(chart_payment_ids)
    chart_unique_payers_count = (
        Payment.objects.filter(id__in=chart_payment_ids)
        .values("student").distinct().count()
    )

    pay_qs = pay_qs.order_by("-paid_date", "-id")
    filtered_payments = list(pay_qs)
    uz_month_map = UZ_MONTH_NAMES
    grouped_rows = {}

    for payment in filtered_payments:
        student_id = payment.student_id
        row = grouped_rows.get(student_id)

        if row is None:
            row = {
                "student": payment.student,
                "latest_payment": payment,
                "latest_paid_date": payment.paid_date,
                "latest_created_at": payment.created_at,
                "latest_payment_id": payment.id,
                "latest_cash_amount": int(payment.cash_amount or 0),
                "latest_card_amount": float(payment.card_amount or 0),
                "latest_note": payment.note or "",
                "total_sum": 0,
                "payment_count": 0,
                "group_entries": {},
                "month_entries": {},
                "staff_entries": {},
                "type_entries": {},
            }
            grouped_rows[student_id] = row

        if sel_month_first:
            # Oy filtri: paymentning faqat SHU OYNING ulushi
            row["total_sum"] += _month_part_by_pay.get(payment.id, 0)
        else:
            row["total_sum"] += int(payment.summa or 0)
        row["payment_count"] += 1

        if payment.created_by:
            staff_name = payment.created_by.get_full_name() or payment.created_by.email
            row["staff_entries"].setdefault(payment.created_by_id or staff_name, staff_name)

        if payment.payment_type:
            row["type_entries"].setdefault(
                payment.payment_type,
                {
                    "code": payment.payment_type,
                    "label": payment.get_payment_type_display(),
                },
            )

        allocations = getattr(payment, "prefetched_allocations", []) or []
        if allocations:
            for alloc in allocations:
                tuition_month = getattr(alloc, "tuition_month", None)
                enrollment = getattr(tuition_month, "enrollment", None)
                group = getattr(enrollment, "group", None)
                if group and not getattr(group, "is_archived", False):
                    row["group_entries"].setdefault(
                        group.id,
                        {
                            "id": group.id,
                            "name": group.nom or "—",
                            "category": getattr(getattr(group, "category_obj", None), "name", "") or "—",
                        },
                    )

                month_value = getattr(tuition_month, "month", None)
                if month_value:
                    month_key = month_value.strftime("%Y-%m")
                    row["month_entries"].setdefault(
                        month_key,
                        {
                            "key": month_key,
                            "label": f"{uz_month_map.get(month_value.month, month_value.strftime('%B'))} {month_value.year}",
                        },
                    )
        else:
            if payment.group and not getattr(payment.group, "is_archived", False):
                row["group_entries"].setdefault(
                    payment.group_id,
                    {
                        "id": payment.group_id,
                        "name": payment.group.nom or "—",
                        "category": getattr(getattr(payment.group, "category_obj", None), "name", "") or "—",
                    },
                )
            if payment.paid_date:
                fallback_month = payment.paid_date.replace(day=1)
                month_key = fallback_month.strftime("%Y-%m")
                row["month_entries"].setdefault(
                    month_key,
                    {
                        "key": month_key,
                        "label": f"{uz_month_map.get(fallback_month.month, fallback_month.strftime('%B'))} {fallback_month.year}",
                    },
                )

    display_rows = []
    for row in grouped_rows.values():
        group_entries = list(row["group_entries"].values())
        month_entries = list(row["month_entries"].values())
        staff_entries = list(row["staff_entries"].values())
        type_entries = list(row["type_entries"].values())

        group_names = [g["name"] for g in group_entries if g.get("name")]
        category_names = []
        for entry in group_entries:
            category_name = entry.get("category")
            if category_name and category_name != "—" and category_name not in category_names:
                category_names.append(category_name)

        row["group_entries"] = group_entries
        row["visible_group_entries"] = group_entries[:2]
        row["remaining_group_count"] = max(0, len(group_entries) - 2)
        row["group_summary_title"] = ", ".join(group_names) if group_names else "—"
        row["category_summary"] = ", ".join(category_names[:2]) if category_names else "—"
        if len(category_names) > 2:
            row["category_summary"] += f" +{len(category_names) - 2}"

        row["month_entries"] = month_entries
        row["visible_month_entries"] = month_entries[:2]
        row["remaining_month_count"] = max(0, len(month_entries) - 2)
        row["month_summary_title"] = ", ".join(m["label"] for m in month_entries) if month_entries else "—"

        row["staff_entries"] = staff_entries
        row["staff_summary"] = ", ".join(staff_entries[:2]) if staff_entries else "—"
        if len(staff_entries) > 2:
            row["staff_summary"] += f" +{len(staff_entries) - 2}"

        row["type_entries"] = type_entries
        row["visible_type_entries"] = type_entries[:2]
        row["remaining_type_count"] = max(0, len(type_entries) - 2)

        display_rows.append(row)

    groups = Group.objects.filter(is_archived=False)
    if center:
        groups = groups.filter(center=center)

    teachers_qs = User.objects.filter(role="teacher", is_active=True)
    if center:
        teachers_qs = teachers_qs.filter(center=center)

    courses = Category.objects.all().only("id", "name")

    staffs = User.objects.filter(role__in=["manager", "admin", "director"], is_active=True)
    if center:
        staffs = staffs.filter(center=center)

    uz_months = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
    ]

    history_month_value = cur_month_start.strftime("%Y-%m")
    if sel_month and sel_month.isdigit():
        history_month_value = f"{cur_year}-{int(sel_month):02d}"

    query_params = request.GET.copy()
    query_params.pop("page", None)

    return {
        "center": center,
        "page_rows": display_rows,
        "filtered_payments": filtered_payments,
        "payment_record_count": len(filtered_payments),
        "total_income": total_income,
        "filtered_income": filtered_income,
        "chart_data": chart_data,
        "chart_labels": chart_labels,
        "chart_kicker": chart_kicker,
        "chart_period_label": chart_period_label,
        "selected_period_label": _human_period_label(selected_from, selected_to),
        "chart_payment_record_count": chart_payment_record_count,
        "chart_unique_payers_count": chart_unique_payers_count,
        "unique_payers_count": unique_payers_count,
        "groups": groups,
        "teachers": teachers_qs,
        "courses": courses,
        "staffs": staffs,
        "uz_months": uz_months,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "sel_group": sel_group,
        "sel_teacher": sel_teacher,
        "sel_course": sel_course,
        "sel_staff": sel_staff,
        "sel_type": sel_type,
        "sel_month": sel_month,
        "history_month_value": history_month_value,
        "query_string": query_params.urlencode(),
        "selected_from": selected_from,
        "selected_to": selected_to,
    }


@login_required
@require_http_methods(["GET"])
@require_feature("finance")
def payment_export_xlsx(request):
    if not user_can_manage_payments(request.user):
        return HttpResponseForbidden("Ruxsat yo'q.")

    try:
        dashboard = _get_payment_dashboard_data(request)
    except PermissionDenied:
        return HttpResponseForbidden("Markaz biriktirilmagan")

    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    grouped_rows = dashboard["page_rows"]
    filtered_payments = dashboard["filtered_payments"]

    def _fill(color: str):
        return PatternFill("solid", fgColor=color)

    def _border():
        side = Side(style="thin", color="CBD5E1")
        return Border(left=side, right=side, top=side, bottom=side)

    def _money(cell):
        cell.number_format = '#,##0'

    def _auto_width(ws, *, min_width: int = 12, max_width: int = 34):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except Exception:
                    continue
            ws.column_dimensions[column_letter].width = max(min_width, min(max_length + 2, max_width))

    def _full_name(user):
        if not user:
            return "—"
        return user.get_full_name() or user.email or f"{getattr(user, 'ism', '')} {getattr(user, 'familya', '')}".strip() or "—"

    group_map = {str(group.id): group.nom for group in dashboard["groups"]}
    teacher_map = {str(teacher.id): _full_name(teacher) for teacher in dashboard["teachers"]}
    course_map = {str(course.id): course.name for course in dashboard["courses"]}
    staff_map = {str(staff.id): _full_name(staff) for staff in dashboard["staffs"]}
    month_map = {str(mid): mname for mid, mname in dashboard["uz_months"]}

    filter_rows = [
        ("Qidiruv", dashboard["q"] or "Barchasi"),
        ("Sanadan", dashboard["date_from"] or "—"),
        ("Sanagacha", dashboard["date_to"] or "—"),
        ("Guruh", group_map.get(dashboard["sel_group"], "Barchasi")),
        ("O'qituvchi", teacher_map.get(dashboard["sel_teacher"], "Barchasi")),
        ("Yo'nalish", course_map.get(dashboard["sel_course"], "Barchasi")),
        ("Xodim", staff_map.get(dashboard["sel_staff"], "Barchasi")),
        ("To'lov turi", dict(Payment.PAYMENT_TYPES).get(dashboard["sel_type"], "Barchasi")),
        ("Qaysi oy", month_map.get(dashboard["sel_month"], "Barcha oylar")),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umumiy ro'yxat"
    ws.sheet_view.showGridLines = False

    title_fill = _fill("0F172A")
    header_fill = _fill("2563EB")
    accent_fill = _fill("E0F2FE")
    soft_fill = _fill("F8FAFC")
    money_fill = _fill("ECFDF5")
    white_font = Font(color="FFFFFF", bold=True, size=12)
    dark_font = Font(color="0F172A", size=11)
    strong_font = Font(color="0F172A", bold=True, size=11)
    money_font = Font(color="047857", bold=True, size=11)

    ws.merge_cells("A1:K1")
    ws["A1"] = "To'lovlar eksporti"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Davr: {dashboard['selected_period_label']}"
    ws["A2"].fill = _fill("1E3A8A")
    ws["A2"].font = Font(color="DBEAFE", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    metrics = [
        ("Eksport davri", dashboard["selected_period_label"]),
        ("Noyob o'quvchi", f"{dashboard['unique_payers_count']} ta"),
        ("To'lov yozuvlari", f"{dashboard['payment_record_count']} ta"),
        ("Filter daromad", dashboard["filtered_income"]),
        ("Umumiy daromad", dashboard["total_income"]),
    ]
    ws["A4"] = "Ko'rsatkich"
    ws["B4"] = "Qiymat"
    for cell in ("A4", "B4"):
        ws[cell].fill = header_fill
        ws[cell].font = white_font
        ws[cell].border = _border()
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    metric_start = 5
    for idx, (label, value) in enumerate(metrics, start=metric_start):
        ws.cell(row=idx, column=1, value=label)
        value_cell = ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).fill = soft_fill
        value_cell.fill = money_fill if isinstance(value, int) else accent_fill
        ws.cell(row=idx, column=1).font = strong_font
        value_cell.font = money_font if isinstance(value, int) else strong_font
        ws.cell(row=idx, column=1).border = _border()
        value_cell.border = _border()
        if isinstance(value, int):
            _money(value_cell)

    ws["D4"] = "Aktiv filtr"
    ws["E4"] = "Qiymat"
    for cell in ("D4", "E4"):
        ws[cell].fill = header_fill
        ws[cell].font = white_font
        ws[cell].border = _border()
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    for idx, (label, value) in enumerate(filter_rows, start=5):
        ws.cell(row=idx, column=4, value=label)
        ws.cell(row=idx, column=5, value=value)
        ws.cell(row=idx, column=4).fill = soft_fill
        ws.cell(row=idx, column=5).fill = accent_fill
        ws.cell(row=idx, column=4).font = strong_font
        ws.cell(row=idx, column=5).font = dark_font
        ws.cell(row=idx, column=4).border = _border()
        ws.cell(row=idx, column=5).border = _border()

    table_row = 16
    summary_headers = [
        "So'nggi sana",
        "O'quvchi",
        "Telefon",
        "Guruhlar",
        "Yo'nalish",
        "Oylar",
        "To'lovlar soni",
        "Turlar",
        "Xodimlar",
        "Jami summa",
        "Oxirgi izoh",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=table_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    current_row = table_row + 1
    for row in grouped_rows:
        values = [
            row["latest_paid_date"].strftime("%d.%m.%Y") if row.get("latest_paid_date") else "—",
            _full_name(row.get("student")),
            getattr(row.get("student"), "telefon1", "") or getattr(row.get("student"), "telefon2", "") or "—",
            row.get("group_summary_title") or "—",
            row.get("category_summary") or "—",
            row.get("month_summary_title") or "—",
            row.get("payment_count") or 0,
            ", ".join(item["label"] for item in row.get("type_entries", [])) or "—",
            row.get("staff_summary") or "—",
            row.get("total_sum") or 0,
            row.get("latest_note") or "—",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = soft_fill if current_row % 2 == 0 else _fill("FFFFFF")
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 10:
                cell.font = money_font
                _money(cell)
            else:
                cell.font = dark_font
        current_row += 1

    ws.freeze_panes = "A17"
    _auto_width(ws, max_width=38)

    detail_ws = wb.create_sheet("To'lov yozuvlari")
    detail_ws.sheet_view.showGridLines = False
    detail_ws.merge_cells("A1:M1")
    detail_ws["A1"] = "To'lov yozuvlari"
    detail_ws["A1"].fill = title_fill
    detail_ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    detail_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    detail_ws.merge_cells("A2:M2")
    detail_ws["A2"] = f"Filterlangan yozuvlar soni: {dashboard['payment_record_count']} ta"
    detail_ws["A2"].fill = _fill("1E293B")
    detail_ws["A2"].font = Font(color="E2E8F0", bold=True, size=10)
    detail_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    detail_headers = [
        "Sana",
        "Vaqt",
        "O'quvchi",
        "Telefon",
        "Guruh",
        "Yo'nalish",
        "Oy / taqsimot",
        "Naqd",
        "Karta",
        "Jami",
        "Tur",
        "Xodim",
        "Izoh",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail_ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    detail_row = 5
    for payment in filtered_payments:
        allocations = getattr(payment, "prefetched_allocations", []) or []
        allocation_labels = []
        for allocation in allocations:
            tuition_month = getattr(allocation, "tuition_month", None)
            month_value = getattr(tuition_month, "month", None)
            if month_value:
                month_label = f"{UZ_MONTH_NAMES.get(month_value.month, month_value.strftime('%B'))} {month_value.year}"
            else:
                month_label = "—"
            allocation_labels.append(f"{month_label}: {int(allocation.amount or 0):,} so'm".replace(",", " "))
        if not allocation_labels and payment.paid_date:
            fallback = payment.paid_date.replace(day=1)
            allocation_labels.append(f"{UZ_MONTH_NAMES.get(fallback.month, fallback.strftime('%B'))} {fallback.year}")

        detail_values = [
            payment.paid_date.strftime("%d.%m.%Y") if payment.paid_date else "—",
            payment.paid_time.strftime("%H:%M") if payment.paid_time else "—",
            _full_name(payment.student),
            getattr(payment.student, "telefon1", "") or getattr(payment.student, "telefon2", "") or "—",
            payment.group.nom if payment.group else "—",
            getattr(getattr(payment.group, "category_obj", None), "name", "") or "—",
            ", ".join(allocation_labels) or "—",
            int(payment.cash_amount or 0),
            int(payment.card_amount or 0),
            int(payment.summa or 0),
            payment.get_payment_type_display(),
            _full_name(payment.created_by),
            payment.note or "—",
        ]
        for col_idx, value in enumerate(detail_values, start=1):
            cell = detail_ws.cell(row=detail_row, column=col_idx, value=value)
            cell.fill = soft_fill if detail_row % 2 == 1 else _fill("FFFFFF")
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in (8, 9, 10):
                cell.font = money_font
                _money(cell)
            else:
                cell.font = dark_font
        detail_row += 1

    detail_ws.freeze_panes = "A5"
    _auto_width(detail_ws, max_width=42)

    filename = f"tolovlar_export_{dashboard['selected_from'].isoformat()}_{dashboard['selected_to'].isoformat()}.xlsx"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def tolovlar_home(request):
    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    try:
        dashboard = _get_payment_dashboard_data(request)
    except PermissionDenied:
        return HttpResponseForbidden("Markaz biriktirilmagan")

    allowed_page_sizes = (10, 20, 50, 100)
    try:
        per_page = int((request.GET.get("per_page") or request.GET.get("page_size") or 10))
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in allowed_page_sizes:
        per_page = 10

    paginator = Paginator(dashboard["page_rows"], per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    dashboard.update(
        {
            "page_obj": page_obj,
            "total_count": paginator.count,
            "per_page": per_page,
            "page_size": per_page,
            "page_size_options": allowed_page_sizes,
            "allowed_page_sizes": allowed_page_sizes,
            "is_paginated": page_obj.has_other_pages(),
        }
    )
    try:
        from store.models import PaymentMethod as _PM
        from store.views import _ensure_default_payment_methods as _seed_pm
        from core.tenant import get_request_center as _grc
        _center = _grc(request)
        if _center:
            _seed_pm(_center)
            dashboard["dynamic_payment_methods"] = list(
                _PM.objects.filter(center=_center, is_active=True).order_by('nom')
            )
        else:
            dashboard["dynamic_payment_methods"] = []
    except Exception:
        dashboard["dynamic_payment_methods"] = []
    return render(request, "education/tolovlar_list.html", dashboard)

@login_required
def get_payment_details(request):
    """
    Returns full transaction history for a specific TuitionMonth as JSON.
    """
    tuition_month_id = request.GET.get('tuition_month_id')
    student_id = request.GET.get('student_id')
    group_id = request.GET.get('group_id')
    
    if tuition_month_id:
        allocs = PaymentAllocation.objects.filter(
            tuition_month_id=tuition_month_id
        ).select_related('payment', 'payment__student', 'payment__group', 'payment__created_by')
    elif student_id and group_id:
        allocs = PaymentAllocation.objects.filter(
            payment__student_id=student_id,
            payment__group_id=group_id
        ).select_related('payment', 'payment__student', 'payment__group', 'payment__created_by')
    else:
        return JsonResponse({'ok': False, 'error': 'Missing identifiers'}, status=400)

    if not allocs.exists():
        return JsonResponse({'ok': True, 'payments': [], 'total_sum': 0})

    first = allocs.first()
    data = []
    total = 0
    for a in allocs.order_by('-payment__paid_date', '-id'):
        total += a.amount
        data.append({
            'id': a.payment.id,
            'amount': a.amount,
            'cash_amount': a.payment.cash_amount,
            'card_amount_som': a.payment.card_amount_som,
            'date': a.payment.paid_date.strftime("%d.%m.%Y"),
            'raw_date': a.payment.paid_date.strftime("%Y-%m-%d"),
            'time': a.payment.paid_time.strftime("%H:%M") if a.payment.paid_time else "--:--",
            'method': a.payment.get_payment_type_display(),
            'staff': a.payment.created_by.get_full_name() if a.payment.created_by else '—',
            'note': a.payment.note or ''
        })

    return JsonResponse({
        'ok': True,
        'student_name': first.payment.student.get_full_name(),
        'group_name': first.payment.group.nom,
        'teacher_name': first.payment.group.oqituvchi.get_full_name() if first.payment.group.oqituvchi else "—",
        'total_sum': total,
        'payments': data
    })


@login_required
def student_payments_pdf(request):
    """
    Generates a professional printable HTML summary of payments for a student.
    """
    student_id = request.GET.get('student_id')
    group_id = request.GET.get('group_id')

    if not student_id or not group_id:
        return HttpResponse("Missing student_id or group_id", status=400)

    center = get_active_center(request)
    student = get_object_or_404(User, id=student_id)
    group = get_object_or_404(Group, id=group_id)
    enrollment = Enrollment.objects.filter(student=student, group=group).first()

    if not enrollment:
        return HttpResponse("Enrollment topilmadi", status=404)

    payments_qs = Payment.objects.filter(enrollment=enrollment).select_related('created_by').order_by('paid_date', 'paid_time')
    
    total_paid = payments_qs.aggregate(s=Sum('summa'))['s'] or 0
    
    # Calculate total expected fee from TuitionMonths
    tms = TuitionMonth.objects.filter(enrollment=enrollment).order_by('month')
    total_expected = tms.aggregate(s=Sum('fee_amount'))['s'] or 0
    
    # Balance calculations
    remaining_debt = max(0, total_expected - total_paid)
    overpayment = max(0, total_paid - total_expected)

    # Monthly breakdown
    monthly_data = []
    for tm in tms:
        paid_amount = tm.allocations.aggregate(s=Sum('amount'))['s'] or 0
        monthly_data.append({
            'month': tm.month,
            'fee': tm.fee_amount,
            'paid': paid_amount,
            'debt': max(0, tm.fee_amount - paid_amount),
            'overpaid': max(0, paid_amount - tm.fee_amount),
        })

    context = {
        'center': center,
        'student': student,
        'group': group,
        'enrollment': enrollment,
        'payments': payments_qs,
        'total_paid': total_paid,
        'total_expected': total_expected,
        'remaining_debt': remaining_debt,
        'overpayment': overpayment,
        'monthly_data': monthly_data,
        'print_time': timezone.now(),
        'staff_name': request.user.get_full_name() or request.user.email,
    }
    
    return render(request, "education/receipt.html", context)

@require_POST
@login_required
def payment_delete(request, payment_id):
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not user_can_manage_payments(request.user):
        error_message = "Ruxsat yo'q."
        messages.error(request, error_message)
        if is_ajax:
            return JsonResponse({"ok": False, "error": error_message}, status=403)
        return redirect("education:tolovlar_home")

    next_url = request.POST.get("next") or request.GET.get("next") or "education:tolovlar_home"

    center = get_active_center(request)
    qs = Payment.objects.all()
    if center:
        qs = qs.filter(center=center)
    
    payment = get_object_or_404(qs, id=payment_id)
    
    try:
        with transaction.atomic():
            # ✅ Soft delete allocations as well
            payment.allocations.all().delete(deleted_by=request.user)
            payment.delete(deleted_by=request.user)
                 
        messages.success(request, "✅ To'lov o'chirildi. O'quvchi qarzdorlar ro'yxatiga qaytadi.")
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {e}")
        if is_ajax:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)
    
    if is_ajax:
        return JsonResponse({"ok": True, "redirect_url": next_url})
    return redirect(next_url)


@login_required
def student_groups_api(request, student_id):
    """Return active group names and total debt for a student (payment modal)."""
    from django.db.models import Sum
    center = get_active_center(request)
    qs = Enrollment.objects.filter(student_id=student_id, is_active=True, group__is_archived=False, group__is_deleted=False).select_related('group')
    if center:
        qs = qs.filter(center=center)
    groups = [e.group.nom for e in qs if e.group and not getattr(e.group, "is_archived", False)]

    # Calculate total debt: sum of all TuitionMonth fees minus all payments
    try:
        enr_ids = [e.id for e in qs]
        total_fee = TuitionMonth.objects.filter(
            enrollment_id__in=enr_ids, is_deleted=False
        ).aggregate(s=Sum('fee_amount'))['s'] or 0
        total_paid = Payment.objects.filter(
            student_id=student_id, is_deleted=False
        ).aggregate(s=Sum('summa'))['s'] or 0
        debt = max(0, int(total_fee) - int(total_paid))
    except Exception:
        debt = 0

    return JsonResponse({
        "groups": " + ".join(groups) if groups else "",
        "debt": debt,
    })


@login_required
def students_with_debt_api(request):
    """Return students who have debt > 0 for the payment modal (JSON)."""
    from django.db.models import Sum
    center = get_active_center(request)
    if not center:
        return JsonResponse({"students": []})

    # 1. Active enrollments for this center
    enr_qs = Enrollment.objects.filter(
        is_active=True, center=center, group__is_archived=False, group__is_deleted=False,
    ).select_related('student', 'group')

    # 2. Total fees per student (sum TuitionMonth.fee_amount)
    enr_ids = list(enr_qs.values_list('id', flat=True))
    fee_rows = (
        TuitionMonth.objects.filter(enrollment_id__in=enr_ids, is_deleted=False)
        .values('enrollment__student_id')
        .annotate(total=Sum('fee_amount'))
    )
    student_fees = {r['enrollment__student_id']: int(r['total'] or 0) for r in fee_rows}

    # 3. Total paid per student (sum Payment.summa)
    student_ids = list(student_fees.keys())
    paid_rows = (
        Payment.objects.filter(student_id__in=student_ids, center=center, is_deleted=False)
        .values('student_id')
        .annotate(total=Sum('summa'))
    )
    student_paid = {r['student_id']: int(r['total'] or 0) for r in paid_rows}

    # 4. Group names per student
    student_groups = {}
    for e in enr_qs:
        if e.group:
            student_groups.setdefault(e.student_id, []).append(e.group.nom)

    # 5. Build result — only students with debt > 0
    students = []
    seen = set()
    for enr in enr_qs:
        sid = enr.student_id
        if sid in seen or sid not in student_fees:
            continue
        seen.add(sid)
        fee  = student_fees.get(sid, 0)
        paid = student_paid.get(sid, 0)
        debt = max(0, fee - paid)
        if debt <= 0:
            continue
        u = enr.student
        students.append({
            'id':     sid,
            'name':   f"{u.ism} {u.familya}".strip(),
            'phone':  getattr(u, 'telefon1', '') or '',
            'groups': ' + '.join(student_groups.get(sid, [])),
            'debt':   debt,
        })

    students.sort(key=lambda x: x['name'])
    return JsonResponse({'students': students})


# education/views.py


# education/views.py




@login_required
def payment_history(request, student_id):
    """
    O'quvchining (barcha kurslari bo'yicha) to'lov tarixini va joriy oy holatini xisoblaydi.
    """
    month_str = request.GET.get("month")
    selected_month = parse_month_str(month_str) or first_day_of_current_month()
    
    # 1. Barcha enrollments
    enrs = Enrollment.objects.filter(student_id=student_id, is_active=True)
    
    total_fee = 0
    total_paid_this_month = 0
    
    for e in enrs:
        tm = ensure_tuition_month(e, selected_month)
        total_fee += int(getattr(tm, "fee_amount", 0) or 0)
        total_paid_this_month += int(get_month_paid(e, selected_month) or 0)
    
    total_qoldiq = max(0, total_fee - total_paid_this_month)
    
    # 2. Barcha to'lovlar
    qs = Payment.objects.filter(student_id=student_id).order_by('-paid_date', '-paid_time')
    from core.tenant import get_request_center
    center = get_active_center(request)
    if center:
        qs = qs.filter(center=center)
    
    payments_data = []
    for p in qs:
        # Har bir to'lov qaysi oylarga tushganini ham ko'rsatishimiz mumkin
        allocs = PaymentAllocation.objects.filter(payment=p).select_related("tuition_month")
        alloc_list = [{"month": a.tuition_month.month.strftime("%Y-%m"), "amount": a.amount} for a in allocs]
        
        payments_data.append({
            "id": p.id,
            "paid_at": f"{p.paid_date.strftime('%d.%m.%Y')} {p.paid_time.strftime('%H:%M')}",
            "paid_date_iso": p.paid_date.strftime('%Y-%m-%d'),
            "cash": int(p.cash_amount or 0),
            "card": int(getattr(p, 'card_amount_som', 0) or getattr(p, 'card_amount', 0) or 0),
            "total": int(p.summa or 0),
            "method_code": p.payment_type,
            "method": p.get_payment_type_display(),
            "staff": p.created_by.get_full_name() if p.created_by else "—",
            "note": p.note or "",
            "group_name": p.group.nom if p.group else "—",
            "allocations": alloc_list,
            "receipt_url": reverse("education:payment_receipt_pdf", args=[p.id]) if p.id else None
        })

    return JsonResponse({
        "month": selected_month.strftime("%Y-%m"),
        "fee": total_fee,
        "paid_this_month": total_paid_this_month,
        "qoldiq": total_qoldiq,
        "payments": payments_data
    })




@login_required
def tolov_oqituvchilar(request):
    from django.shortcuts import redirect as _redirect
    return _redirect("education:teacher_salary_list")



# ---------- HUB va ro'yxatlar ----------
@login_required
def groups_hub(request):
    """
    📘 Guruhlar markaziy sahifasi — barcha kategoriyalar ro'yxati.
    """
    from .models import Category  # agar alohida model bo'lsa
    categories = Category.objects.all() if hasattr(Category, "objects") else []
    return render(request, "education/groups_home.html", {
        "categories": categories,
    })

def group_delete_confirm(request, id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, id=id)
    if request.method == "POST":
        group.delete(deleted_by=request.user)
        return redirect("education:groups_home")
    return render(request, "education/group_delete_confirm.html", {"g": group})



@login_required
def edit_category(request, id):
    # ✅ Strict isolation: Only the center's own or global (if primary center)
    from core.tenant import get_request_center
    center = get_request_center(request)
    first_center = Center.objects.order_by("id").first()
    
    if first_center and center and center.id == first_center.id:
        # Primary center can edit its own and global orphans
        cat = get_object_or_404(Category, Q(center=center) | Q(center__isnull=True), id=id)
    else:
        # Other centers can only edit their own
        cat = get_object_or_404(Category, center=center, id=id)
        
    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description")
        image = request.FILES.get("image")

        name_stripped = (name or "").strip()
        if not name_stripped:
            messages.error(request, "Bo'lim nomi bo'sh bo'lishi mumkin emas!")
            return render(request, "education/category_edit.html", {"cat": cat})

        # Check for duplication (case-insensitive, same center, including soft-deleted)
        qs = Category.all_objects.filter(name__iexact=name_stripped, center=cat.center).exclude(id=cat.id)
        if qs.exists():
            messages.error(request, "Ushbu nomdagi bo'lim allaqachon mavjud!")
            return render(request, "education/category_edit.html", {"cat": cat})

        cat.name = name_stripped
        cat.description = description

        # 🔹 Agar yangi rasm tanlangan bo'lsa, yangisini saqlaymiz
        if image:
            cat.image = image

        cat.save()
        messages.success(request, "Bo'lim muvaffaqiyatli tahrirlandi ✅")
        return redirect("education:groups_home")

    return render(request, "education/category_edit.html", {"cat": cat})


@login_required
def delete_category(request, id):
    # ✅ Strict isolation: Only the center's own or global (if primary center)
    from core.tenant import get_request_center
    center = get_request_center(request)
    first_center = Center.objects.order_by("id").first()
    
    if first_center and center and center.id == first_center.id:
        cat = get_object_or_404(Category, Q(center=center) | Q(center__isnull=True), id=id)
    else:
        cat = get_object_or_404(Category, center=center, id=id)

    if request.method == "POST":
        cat.delete(deleted_by=request.user)
        messages.success(request, "Bo'lim o'chirildi 🗑️")
        return redirect("education:groups_home")
    return render(request, "education/category_delete_confirm.html", {"cat": cat})


@login_required
def groups_by_category(request, category):
    if category not in ("lang", "it"):
        raise Http404("Noto'g'ri kategoriya")

    rows = (
        Group.objects.filter(category=category)
        .select_related("center", "oqituvchi")
        .annotate(
            student_count=Count("enrollments", filter=Q(enrollments__is_active=True, enrollments__is_deleted=False)),
            sana=Coalesce(F("course_start_date"), Cast(F("tuzilgan"), models.DateField()))
        )
        .order_by("nom")
    )
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        rows = rows.filter(center=center)
    return render(
        request,
        "education/groups_by_category.html",
        {"rows": rows, "category": category, "can_manage": _can_manage(request.user)},
    )


# DRY: guruh yaratish
@login_required
def create_group_for_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if not _can_manage(request.user):
        messages.error(request, "Sizda guruh yaratish huquqi yo'q.")
        return redirect("education:groups_home")

    if request.method == "POST":
        form = GroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)

            # 🟢 To'g'ri maydon: ForeignKey bo'lgan 'category_obj'
            group.category_obj = category

            # Center assignment
            from core.tenant import get_request_center
            center = get_request_center(request)
            if center:
                group.center = center

            # Eski 'category' maydoni ham to'ldirilsa yaxshi
            group.category = Group.IT  # yoki Group.LANG — kerakli turga qarab
            group.save()

            messages.success(request, f"✅ '{group.nom}' guruhi {category.name} bo'limiga qo'shildi.")
            return redirect("education:category_detail", category_id=category.id)
    else:
        form = GroupForm()

    from core.tenant import get_request_center as _grc
    _center = _grc(request)
    _cts = CourseTemplate.objects.filter(center=_center, is_active=True).order_by("name") if _center else []
    return render(request, "education/group_form.html", {
        "form": form, "category": category, "course_templates": _cts,
    })

# @login_required
# def group_create_lang(request):
#     return _create_group(request, Group.LANG)


# @login_required
# def group_create_it(request):
#     return _create_group(request, Group.IT)


# (eski ro'yxatlar kerak bo'lsa)
@login_required
def guruhlar(request):
    rows = (
        Group.objects.select_related("center", "oqituvchi")
        .annotate(
            student_count=Count("enrollments", filter=Q(enrollments__is_active=True, enrollments__is_deleted=False)),
            sana=Coalesce(F("course_start_date"), Cast(F("tuzilgan"), models.DateField()))
        )
        .order_by("nom")
    )
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        rows = rows.filter(center=center)
    return render(request, "education/groups.html", {"rows": rows, "can_manage": _can_manage(request.user)})


@login_required
def guruhlar_tillar(request):
    return groups_by_category(request, Group.LANG)


@login_required
def guruhlar_it(request):
    return groups_by_category(request, Group.IT)


# ---------- Bitta guruh (bitta sahifada hamma narsa) ----------
@login_required
def group_detail(request, pk: int):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    if request.user.role == "teacher" and g.oqituvchi != request.user:
        # Support teacher ham guruhni ko'ra oladi va davomat qila oladi
        if g.support_teacher_id != request.user.id:
            return HttpResponseForbidden("Siz bu guruhni ko'ra olmaysiz.")

    date_str = request.GET.get("date")
    selected_date = parse_date(date_str) if date_str else localdate()
    if not selected_date:
        selected_date = localdate()
    selected_month = month_first_day(selected_date)

    from django.db.models import Exists, OuterRef, Q

    has_attendance = Attendance.objects.filter(
        group=g,
        student=OuterRef('student'),
        date__year=selected_month.year,
        date__month=selected_month.month
    )
    enrollments = list(
        Enrollment.objects
        .filter(group=g)
        .annotate(has_att=Exists(has_attendance))
        .filter(Q(is_active=True) | Q(is_active=False, has_att=True))
        .select_related("student", "group")   # ✅ MUHIM
        .order_by("student__ism", "student__familya")
    )
    student_user_ids = [e.student_id for e in enrollments]
    # Faqat shu guruh (g) bo'yicha fee/paid hisoblaymiz.
    # Boshqa guruhlarni kiritish studentni boshqa guruhda ortiqcha to'lagan bo'lsa
    # bu guruhda ham "To'langan" ko'rsatishiga olib kelgan — noto'g'ri xatti-harakat.
    student_enrollment_qs = Enrollment.objects.filter(
        student_id__in=student_user_ids,
        group=g,
        student__is_archived=False,
        group__is_archived=False,
        group__is_deleted=False,
    ).filter(
        Q(is_active=True) | Q(id__in=[e.id for e in enrollments])
    )
    if center:
        student_enrollment_qs = student_enrollment_qs.filter(center=center)
    student_enrollment_ids = list(student_enrollment_qs.values_list("id", flat=True))

    # Faqat shu guruh bo'yicha tanlangan oy to'lov holatini hisoblaymiz.
    fee_field = tuition_month_fee_field()
    student_enrollments = list(student_enrollment_qs.select_related("group"))
    eligible_enrollment_ids = [enrollment.id for enrollment in student_enrollments]

    from django.utils import timezone as _tz2

    _existing_ids = set(
        TuitionMonth.all_objects.filter(
            enrollment_id__in=eligible_enrollment_ids,
            month=selected_month,
        ).values_list("enrollment_id", flat=True)
    )
    _to_create = []
    for _enr in student_enrollments:
        if _enr.id not in _existing_ids:
            _fee = (
                _enr.student_payable_amount
                if _enr.student_payable_amount not in (None, 0)
                else _enr.kurs_narhi
                or int(getattr(_enr.group, "kurs_narxi", 0) or 0)
            )
            _to_create.append(
                TuitionMonth(
                    enrollment=_enr,
                    center_id=_enr.center_id,
                    month=selected_month,
                    fee_amount=_fee or 0,
                )
            )
    if _to_create:
        TuitionMonth.objects.bulk_create(_to_create, ignore_conflicts=True)
    TuitionMonth.all_objects.filter(
        enrollment_id__in=eligible_enrollment_ids,
        month=selected_month,
        is_deleted=True,
    ).update(is_deleted=False, restored_at=_tz2.now())

    student_total_fee_map = {
        sid: 0 for sid in student_user_ids
    }
    if eligible_enrollment_ids:
        student_total_fee_map.update(
            {
                row["enrollment__student_id"]: int(row["fee"] or 0)
                for row in (
                    TuitionMonth.objects
                    .filter(
                        enrollment_id__in=eligible_enrollment_ids,
                        month=selected_month,
                    )
                    .values("enrollment__student_id")
                    .annotate(fee=Coalesce(Sum(fee_field), 0))
                )
            }
        )

    student_total_paid_map = {
        row["tuition_month__enrollment__student_id"]: int(row["paid"] or 0)
        for row in (
            PaymentAllocation.objects
            .filter(
                tuition_month__enrollment_id__in=eligible_enrollment_ids or student_enrollment_ids,
                tuition_month__month=selected_month,
            )
            .values("tuition_month__enrollment__student_id")
            .annotate(paid=Coalesce(Sum("amount"), 0))
        )
    }

    # Balanslar: avval Ledger, agar studentda umuman Ledger bo'lmasa LightningHistory fallback.
    ledger_qs = Ledger.objects.filter(student_id__in=student_user_ids)
    if center:
        ledger_qs = ledger_qs.filter(
            Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True)
        )
    ledger_balance_map = {
        row["student_id"]: int(row["s"] or 0)
        for row in (
            ledger_qs.values("student_id").annotate(s=Coalesce(Sum("ball"), 0))
        )
    }
    history_balance_map = {}
    missing_balance_user_ids = [sid for sid in student_user_ids if sid not in ledger_balance_map]
    if missing_balance_user_ids:
        from chaqmoq.models import LightningHistory

        student_models = {
            st.user_id: st.id
            for st in Student.objects.filter(user_id__in=missing_balance_user_ids)
        }
        if student_models:
            history_qs = LightningHistory.objects.filter(student_id__in=student_models.values())
            if center:
                history_qs = history_qs.filter(student__user__center=center)
            history_totals = {
                row["student_id"]: int(row["s"] or 0)
                for row in (
                    history_qs.values("student_id").annotate(s=Coalesce(Sum("points"), 0))
                )
            }
            history_balance_map = {
                user_id: history_totals.get(student_pk, 0)
                for user_id, student_pk in student_models.items()
            }

    # Sana bo'yicha Attendance (DateTimeField bo'lsa ham ishlaydi)
    try:
        start = make_aware(datetime.combine(selected_date, datetime.min.time()))
        end   = make_aware(datetime.combine(selected_date + timedelta(days=1),
                                           datetime.min.time()))
        att_qs = Attendance.objects.filter(group=g, date__gte=start, date__lt=end)
    except Exception:
        att_qs = Attendance.objects.filter(group=g, date=selected_date)

    pres_map   = {}
    forced_map = {}
    status_map = {}
    for a in att_qs:
        pres_map[a.student_id]   = a.present
        forced_map[a.student_id] = getattr(a, "forced", False)
        status_map[a.student_id] = getattr(a, "status", "present" if a.present else "none")

    # Studentga soxta fieldlar
    for e in enrollments:
        s = e.student
        s.balance = int(ledger_balance_map.get(s.id, history_balance_map.get(s.id, 0)))
        s.present_today     = bool(pres_map.get(s.id, False))
        s.forced_today      = bool(forced_map.get(s.id, False))
        s.attendance_status = status_map.get(s.id, "none")  # 'present' | 'absent_excused' | 'absent_unexcused' | 'none'

        total_fee = int(student_total_fee_map.get(s.id, 0))
        total_paid = int(student_total_paid_map.get(s.id, 0))
        total_remaining = max(0, total_fee - total_paid)

        if total_fee <= 0 or total_paid >= total_fee:
            payment_status = "paid"
            payment_status_label = "To'liq to'langan"
        elif total_paid > 0:
            payment_status = "partial"
            payment_status_label = "Chala to'langan"
        else:
            payment_status = "unpaid"
            payment_status_label = "To'lov qilinmagan"

        if total_fee <= 0:
            payment_status_title = "Tanlangan oy uchun to'lov majburiyati yo'q"
        elif payment_status == "paid":
            payment_status_title = (
                f"Tanlangan oy uchun to'liq to'langan: {total_paid:,} / {total_fee:,} so'm"
            )
        elif payment_status == "partial":
            payment_status_title = (
                f"Tanlangan oy uchun chala to'langan: {total_paid:,} / {total_fee:,} so'm"
                f" • Qoldiq: {total_remaining:,} so'm"
            )
        else:
            payment_status_title = f"Tanlangan oy uchun to'lov qilinmagan: 0 / {total_fee:,} so'm"

        e.payment_status = payment_status
        e.payment_status_label = payment_status_label
        e.payment_status_title = payment_status_title
        e.payment_month_fee = total_fee
        e.payment_month_paid = total_paid
        e.payment_month_remaining = total_remaining


    can_add_student = False
    can_remove_student = False
    
    if request.user.role == "director" or request.user.is_superuser:
        can_add_student = True
        can_remove_student = True
    elif center:
        if request.user.role == "manager":
            can_add_student = center.manager_can_add_student
            can_remove_student = center.manager_can_remove_student
        elif request.user.role == "teacher":
            can_add_student = center.teacher_can_add_student
            can_remove_student = center.teacher_can_remove_student

    # ✅ Filter rules by center and role
    rules_qs = Rule.objects.filter(Q(center=center) | Q(center__isnull=True))
    if request.user.role == 'teacher':
        rules_qs = rules_qs.filter(can_teacher=True)
    elif request.user.role == 'manager':
        rules_qs = rules_qs.filter(can_manager=True)
    elif request.user.role == 'director':
        rules_qs = rules_qs.filter(can_director=True)

    # Tanlangan sana bo'yicha kunlik chaqmoq o'zgarishlari (student + group + date)
    recent_history = (
        DailyLightningRecord.objects.filter(
            group=g,
            date=selected_date,
            student_id__in=student_user_ids,
        )
        .values("student_id")
        .annotate(
            recent_add=Coalesce(Sum("plus_points"), 0),
            recent_sub=Coalesce(Sum("minus_points"), 0),
        )
    )
    recent_history_map = {
        str(item["student_id"]): {
            "add": int(item["recent_add"] or 0),
            "sub": int(item["recent_sub"] or 0),
        }
        for item in recent_history
        if item["recent_add"] or item["recent_sub"]
    }

    exam_reminder_state = None
    if request.user.role in ("teacher", "director", "manager") or request.user.is_superuser:
        try:
            from education.services.exam_service import get_exam_reminder_state
            exam_reminder_state = get_exam_reminder_state(
                group=g,
                on_date=selected_date,
            )
        except Exception:
            logger.exception("Failed to calculate exam reminder state")
            exam_reminder_state = None

    can_view_internal_ranking = request.user.role in ("teacher", "director", "manager") or request.user.is_superuser
    internal_ranking_preview = []
    if can_view_internal_ranking:
        try:
            from education.services.ranking_service import get_group_internal_ranking_preview

            internal_ranking_preview = get_group_internal_ranking_preview(
                group=g,
                on_date=selected_date,
                limit=3,
                actor=request.user,
                persist=False,
            )
        except Exception:
            logger.exception("Failed to calculate group internal ranking preview")
            internal_ranking_preview = []

    closure_state = None
    if request.user.role in ("teacher", "director", "manager") or request.user.is_superuser:
        try:
            from education.services.closure_service import get_group_closure_state

            closure_state = get_group_closure_state(
                group=g,
                on_date=selected_date,
            )
        except Exception:
            logger.exception("Failed to calculate group closure state")
            closure_state = None

    # ── Yangi clean Batafsil sahifasi uchun KPI va o'quvchilar ro'yxati ──
    from education.models import GroupSchedule as _GS
    today_now = localdate()
    month_start_now = today_now.replace(day=1)
    enrolled_total = len(student_enrollments) or len(enrollments)
    capacity = int(getattr(g, "max_students", 0) or 0)
    fill_pct = round(enrolled_total * 100 / capacity, 1) if capacity else 0

    # Davomat (oxirgi 30 kun) — per-student ma'lumotlardan group KPI ni chiqaramiz (4→2 query)
    att_from = today_now - timedelta(days=30)
    att_per_student_total = dict(
        Attendance.objects.filter(group=g, date__gte=att_from, date__lte=today_now)
        .values("student_id").annotate(c=Count("id")).values_list("student_id", "c")
    )
    att_per_student_pres = dict(
        Attendance.objects.filter(
            group=g, date__gte=att_from, date__lte=today_now,
        ).filter(Q(status="present") | Q(present=True) | Q(forced=True))
        .values("student_id").annotate(c=Count("id")).values_list("student_id", "c")
    )
    att_total_g = sum(att_per_student_total.values())
    att_present_g = sum(att_per_student_pres.values())
    att_rate_g = round(att_present_g * 100 / att_total_g, 1) if att_total_g else 0

    # Oylik tushum
    monthly_rev = int(
        Payment.objects.filter(
            group=g,
            paid_date__gte=month_start_now,
            paid_date__lte=today_now,
        ).aggregate(s=Sum("summa"))["s"] or 0
    )

    # Jadval matni
    sched_rows = list(_GS.objects.filter(group=g).order_by("weekday", "start_time"))
    _wd_map = {1: "Du", 2: "Se", 3: "Ch", 4: "Pa", 5: "Ju", 6: "Sh", 7: "Ya"}
    _days_seen = []
    _start_time = None
    _room = ""
    for s in sched_rows:
        sh = _wd_map.get(s.weekday)
        if sh and sh not in _days_seen:
            _days_seen.append(sh)
        if not _start_time and s.start_time:
            _start_time = s.start_time
        if not _room and (s.room or "").strip():
            _room = (s.room or "").strip()
    schedule_days_text = "·".join(_days_seen) if _days_seen else "—"
    schedule_time_text = _start_time.strftime("%H:%M") if _start_time else ""
    schedule_room_text = _room

    # Avatar palitra
    _avatar_palette = [
        ("#2563eb", "#dbeafe"), ("#7c3aed", "#ede9fe"),
        ("#10b981", "#d1fae5"), ("#d97706", "#fef3c7"),
        ("#dc2626", "#fee2e2"), ("#0ea5e9", "#e0f2fe"),
        ("#db2777", "#fce7f3"), ("#0d9488", "#ccfbf1"),
    ]

    def _student_avatar(name):
        safe = (name or "?").strip() or "?"
        idx = sum(ord(c) for c in safe) % len(_avatar_palette)
        col, bg = _avatar_palette[idx]
        return safe[:2].upper(), col, bg

    # att_per_student_total / att_per_student_pres already computed above (shared with KPI)

    student_rows = []
    paid_count = 0
    for enr in enrollments:
        student = enr.student
        sname = f"{student.ism or ''} {student.familya or ''}".strip() or student.username
        initials, col, bg = _student_avatar(sname)
        joined_at = getattr(enr, "created_at", None) or getattr(enr, "tuzilgan", None)
        atot = int(att_per_student_total.get(student.id) or 0)
        apre = int(att_per_student_pres.get(student.id) or 0)
        s_att = round(apre * 100 / atot, 1) if atot else 0
        fee = int(student_total_fee_map.get(student.id) or 0)
        paid = int(student_total_paid_map.get(student.id) or 0)
        if fee <= 0:
            pay_label, pay_kind = ("—", "none")
        elif paid >= fee:
            pay_label, pay_kind = ("To'lagan", "paid")
            paid_count += 1
        elif paid > 0:
            pay_label, pay_kind = ("Qisman", "partial")
        else:
            pay_label, pay_kind = ("To'lamagan", "unpaid")
        student_rows.append({
            "id": student.id,
            "enrollment_id": enr.id,
            "name": sname,
            "initials": initials,
            "color": col,
            "bg": bg,
            "joined": joined_at,
            "att_rate": s_att,
            "pay_label": pay_label,
            "pay_kind": pay_kind,
        })

    # Avg per-student monthly fee for KPI subtitle
    avg_per_oy = 0
    if enrolled_total:
        try:
            avg_per_oy = int(round((monthly_rev or 0) / enrolled_total))
        except Exception:
            avg_per_oy = 0

    ctx = {
        "g": g,
        "group": g,
        "enrollments": enrollments,
        "rules_plus": rules_qs.filter(tur=Rule.PLUS).order_by("nom"),
        "rules_minus": rules_qs.filter(tur=Rule.MINUS).order_by("nom"),
        "can_add_student": can_add_student,
        "can_remove_student": can_remove_student,
        "selected_date": selected_date.isoformat(),
        "today": localdate().isoformat(),
        "recent_history_map": recent_history_map,
        "exam_reminder_state": exam_reminder_state,
        "internal_ranking_preview": internal_ranking_preview,
        "can_view_internal_ranking": can_view_internal_ranking,
        "closure_state": closure_state,
        "can_transfer_student": user_can_transfer_student(request.user),
        # New clean detail context
        "kpi_enrolled": enrolled_total,
        "kpi_capacity": capacity,
        "kpi_fill_pct": fill_pct,
        "kpi_att_rate": att_rate_g,
        "kpi_monthly_rev": monthly_rev,
        "kpi_avg_per_oy": avg_per_oy,
        "schedule_days_text": schedule_days_text,
        "schedule_time_text": schedule_time_text,
        "schedule_room_text": schedule_room_text,
        "student_rows": student_rows,
        "students_paid_count": paid_count,
    }
    return render(request, "education/group_detail.html", ctx)


@login_required
def group_schedule_manage(request, group_id: int):
    """
    Guruh jadvalini ko'rish va tahrirlash.
    Teacher faqat o'z guruhlarini ko'ra oladi.
    Manager/Director — hamma guruhlarni.
    """
    from core.tenant import get_request_center
    from education.models import GroupSchedule

    try:
        center = get_request_center(request)
        if not center:
            raise Http404("Center not found")
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_WEEKLY_SCHEDULE,
            message="Jadval bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response

        role = getattr(request.user, "role", "")
        if role == "teacher" and not request.user.is_superuser:
            group = get_object_or_404(
                Group.objects.select_related("oqituvchi", "center"),
                pk=group_id,
                center=center,
                oqituvchi=request.user,
                is_archived=False,
            )
        else:
            if role not in ("manager", "director") and not request.user.is_superuser:
                return redirect("core:home")
            group = get_object_or_404(
                Group.objects.select_related("oqituvchi", "center"),
                pk=group_id,
                center=center,
                is_archived=False,
            )

        schedules = (
            GroupSchedule.objects.filter(group=group, center=center)
            .select_related("group", "group__oqituvchi")
            .order_by("weekday", "start_time")
        )

        if request.method == "POST":
            action = (request.POST.get("action") or "").strip().lower()

            if action == "save_bulk":
                weekdays_raw = request.POST.getlist("weekdays")
                start_time_value = (request.POST.get("start_time") or "").strip()
                duration_min = _get_int(request.POST, "duration", 0)
                room = (request.POST.get("room") or "").strip()
                start_time_obj = parse_time(start_time_value) if start_time_value else None
                weekdays = []
                for wd in weekdays_raw:
                    try:
                        n = int(wd)
                        if 1 <= n <= 7:
                            weekdays.append(n)
                    except (TypeError, ValueError):
                        continue
                weekdays = sorted(set(weekdays))

                if not weekdays:
                    messages.error(request, "Kamida bitta dars kunini tanlang.")
                elif not start_time_obj:
                    messages.error(request, "Boshlanish vaqtini ko'rsating.")
                elif duration_min <= 0:
                    messages.error(request, "Davomiylikni daqiqalarda kiriting.")
                else:
                    end_minutes = (start_time_obj.hour * 60 + start_time_obj.minute) + duration_min
                    end_minutes = min(end_minutes, 23 * 60 + 59)
                    end_time_obj = (datetime.min.replace(
                        hour=end_minutes // 60, minute=end_minutes % 60
                    )).time()

                    room_clashes = []
                    if room:
                        for wd in weekdays:
                            clashes = list(
                                GroupSchedule.objects.filter(
                                    center=center,
                                    weekday=wd,
                                    start_time=start_time_obj,
                                    room__iexact=room,
                                )
                                .exclude(group=group)
                                .select_related("group")
                                .values_list("group__nom", flat=True)
                                .distinct()
                            )
                            for nom in clashes:
                                if nom not in room_clashes:
                                    room_clashes.append(nom)

                    if room_clashes:
                        messages.warning(
                            request,
                            f"⚠️ Bu vaqtda {', '.join(room_clashes)} ham shu xonadan foydalanadi.",
                        )

                    with transaction.atomic():
                        GroupSchedule.objects.filter(group=group, center=center).delete()
                        for wd in weekdays:
                            GroupSchedule.objects.create(
                                center=center,
                                group=group,
                                weekday=wd,
                                start_time=start_time_obj,
                                end_time=end_time_obj,
                                room=room,
                            )
                    messages.success(request, "✅ Jadval saqlandi.")
                    return redirect("education:group_schedule_manage", group_id=group_id)

            if action == "add":
                weekday = _get_int(request.POST, "weekday", 0)
                start_time_value = (request.POST.get("start_time") or "").strip()
                end_time_value = (request.POST.get("end_time") or "").strip()
                room = (request.POST.get("room") or "").strip()

                start_time_obj = parse_time(start_time_value) if start_time_value else None
                end_time_obj = parse_time(end_time_value) if end_time_value else None

                if not weekday or not start_time_obj:
                    messages.error(request, "Kun va boshlanish vaqti majburiy.")
                elif end_time_obj and end_time_obj <= start_time_obj:
                    messages.warning(request, "Tugash vaqti boshlanish vaqtidan keyin bo'lishi kerak.")
                else:
                    exists = GroupSchedule.objects.filter(
                        group=group,
                        weekday=weekday,
                        start_time=start_time_obj,
                    ).exists()
                    room_conflicts = []
                    if room:
                        room_conflicts = list(
                            GroupSchedule.objects.filter(
                                center=center,
                                weekday=weekday,
                                start_time=start_time_obj,
                                room__iexact=room,
                            )
                            .exclude(group=group)
                            .select_related("group")
                            .values_list("group__nom", flat=True)
                            .distinct()
                        )

                    if exists:
                        messages.warning(request, "⚠️ Bu vaqtda jadval allaqachon mavjud.")
                    elif room_conflicts:
                        messages.warning(
                            request,
                            f"⚠️ Bu vaqtda {', '.join(room_conflicts)} ham shu xonada.",
                        )
                    else:
                        GroupSchedule.objects.create(
                            center=center,
                            group=group,
                            weekday=weekday,
                            start_time=start_time_obj,
                            end_time=end_time_obj,
                            room=room,
                        )
                        messages.success(request, "✅ Jadval qo'shildi.")

            elif action == "delete":
                sched_id = _get_int(request.POST, "schedule_id", 0)
                deleted_count, _ = GroupSchedule.objects.filter(
                    pk=sched_id,
                    group=group,
                    center=center,
                ).delete()
                if deleted_count:
                    messages.success(request, "🗑 Jadval o'chirildi.")
                else:
                    messages.warning(request, "Jadval topilmadi.")

            return redirect("education:group_schedule_manage", group_id=group_id)

        rooms_used = list(schedules.exclude(room="").values_list("room", flat=True).distinct())
        conflict_map: dict[str, list[str]] = {}
        if rooms_used:
            conflicting = (
                GroupSchedule.objects.filter(center=center, room__in=rooms_used)
                .exclude(group=group)
                .select_related("group", "group__oqituvchi")
                .order_by("weekday", "start_time")
            )
            for conflict in conflicting:
                key = f"{conflict.weekday}_{conflict.start_time.strftime('%H:%M:%S')}_{conflict.room.strip().lower()}"
                conflict_map.setdefault(key, []).append(conflict.group.nom)

        schedule_map = {weekday: [] for weekday, _ in _schedule_weekday_labels()}
        for schedule in schedules:
            key = f"{schedule.weekday}_{schedule.start_time.strftime('%H:%M:%S')}_{(schedule.room or '').strip().lower()}"
            schedule.conflict_groups = conflict_map.get(key, [])
            schedule_map.setdefault(schedule.weekday, []).append(schedule)

        # Form pre-fill: pick the dominant slot to seed the editor.
        selected_weekdays = sorted({s.weekday for s in schedules})
        common_start = None
        common_duration = 0
        common_room = ""
        if schedules:
            from collections import Counter
            start_counter = Counter(s.start_time for s in schedules if s.start_time)
            if start_counter:
                common_start = start_counter.most_common(1)[0][0]
            room_counter = Counter((s.room or "").strip() for s in schedules)
            non_empty = [(r, c) for r, c in room_counter.items() if r]
            if non_empty:
                non_empty.sort(key=lambda kv: -kv[1])
                common_room = non_empty[0][0]
            duration_counter = Counter()
            for s in schedules:
                if s.start_time and s.end_time:
                    sm = s.start_time.hour * 60 + s.start_time.minute
                    em = s.end_time.hour * 60 + s.end_time.minute
                    if em > sm:
                        duration_counter[em - sm] += 1
            if duration_counter:
                common_duration = duration_counter.most_common(1)[0][0]
        if common_duration == 0:
            common_duration = 90

        weekday_short = [
            (1, "Du", "Dushanba"), (2, "Se", "Seshanba"), (3, "Ch", "Chorshanba"),
            (4, "Pa", "Payshanba"), (5, "Ju", "Juma"), (6, "Sh", "Shanba"),
            (7, "Ya", "Yakshanba"),
        ]
        _wd_short_map = {n: sh for n, sh, _full in weekday_short}
        sched_days_short = [_wd_short_map[n] for n in selected_weekdays if n in _wd_short_map]
        if sched_days_short:
            current_schedule_text = " · ".join(sched_days_short)
            if common_start:
                current_schedule_text = f"{current_schedule_text} · {common_start.strftime('%H:%M')}"
        else:
            current_schedule_text = ""

        return render(
            request,
            "education/group_schedule_manage.html",
            {
                "group": group,
                "schedules": schedules,
                "schedule_map": schedule_map,
                "weekday_choices": GroupSchedule.WEEKDAY_CHOICES,
                "weekday_labels": _schedule_weekday_labels(),
                "conflict_map": conflict_map,
                "selected_weekdays": selected_weekdays,
                "common_start": common_start,
                "common_duration": common_duration,
                "common_room": common_room,
                "weekday_short": weekday_short,
                "current_schedule_text": current_schedule_text,
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("group_schedule_manage failed: group_id=%s", group_id)
        messages.error(request, "Jadvalni yuklashda xatolik yuz berdi.")
        return redirect("education:group_detail", pk=group_id)


@login_required
@require_GET
def schedule_conflict_check(request):
    """
    AJAX: Bu xona, kuni, vaqtida boshqa guruh bormi?
    Parametrlar: room, weekday, start_time, exclude_group_id
    """
    from core.tenant import get_request_center
    from education.models import GroupSchedule

    try:
        center = get_request_center(request)
        if not center:
            return JsonResponse({"conflict": False, "groups": []})
        if not center_ui_feature_enabled(center, FEATURE_UI_WEEKLY_SCHEDULE):
            return JsonResponse({"detail": "disabled"}, status=403)

        role = getattr(request.user, "role", "")
        if role not in ("teacher", "manager", "director") and not request.user.is_superuser:
            return JsonResponse({"detail": "forbidden"}, status=403)

        room = (request.GET.get("room") or "").strip()
        weekday = _get_int(request.GET, "weekday", 0)
        start_time_value = (request.GET.get("start_time") or "").strip()
        exclude_gid = _get_int(request.GET, "exclude_group_id", 0)
        start_time_obj = parse_time(start_time_value) if start_time_value else None

        if not room or not weekday or not start_time_obj:
            return JsonResponse({"conflict": False, "groups": []})

        qs = (
            GroupSchedule.objects.filter(
                center=center,
                room__iexact=room,
                weekday=weekday,
                start_time=start_time_obj,
            )
            .select_related("group")
            .order_by("group__nom")
        )
        if exclude_gid:
            qs = qs.exclude(group_id=exclude_gid)

        conflicts = [item.group.nom for item in qs]
        return JsonResponse({"conflict": bool(conflicts), "groups": conflicts})
    except Exception:
        logger.exception("schedule_conflict_check failed")
        return JsonResponse({"conflict": False, "groups": []}, status=500)


_WEEKLY_SLOT_MIN = 30


def _weekly_t_to_min(t):
    return t.hour * 60 + t.minute


def _weekly_teacher_initials(teacher):
    if not teacher:
        return "?"
    name = (teacher.get_full_name() or "").strip()
    if not name:
        name = (getattr(teacher, "username", "") or "").strip()
    parts = [p for p in name.split() if p]
    if len(parts) >= 2:
        return (parts[0][:1] + parts[1][:1]).upper()
    if parts:
        return parts[0][:2].upper()
    return "??"


@login_required
def weekly_schedule_view(request):
    """
    Manager/Director uchun haftalik jadval — vaqt × kun gridi,
    o'qituvchi rangi, o'qituvchi yuklamasi, Excel/PDF eksport.
    """
    from core.tenant import get_request_center
    from education.models import GroupSchedule
    from collections import defaultdict

    try:
        center = get_request_center(request)
        if not center:
            raise Http404("Center not found")
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_WEEKLY_SCHEDULE,
            message="Haftalik jadval bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response

        role = getattr(request.user, "role", "")
        if role not in ("manager", "director") and not request.user.is_superuser:
            return redirect("core:home")

        teacher_id = _get_int(request.GET, "teacher", 0)
        room_filter = (request.GET.get("room") or "").strip()

        base_qs = (
            GroupSchedule.objects.filter(center=center, group__is_archived=False)
            .select_related("group", "group__oqituvchi", "group__category_obj")
            .order_by("weekday", "start_time", "group__nom")
        )
        if teacher_id:
            base_qs = base_qs.filter(group__oqituvchi_id=teacher_id)

        qs = base_qs
        if room_filter:
            qs = qs.filter(room__icontains=room_filter)

        total_slots_count = base_qs.count()
        filtered_slots_count = qs.count()

        weekday_labels = _schedule_weekday_labels()
        week_map = {weekday: [] for weekday, _ in weekday_labels}
        for schedule in qs:
            week_map[schedule.weekday].append(schedule)

        all_slots = list(qs)
        if all_slots:
            earliest = min(_weekly_t_to_min(s.start_time) for s in all_slots)
            latest = max(
                _weekly_t_to_min(s.end_time) if s.end_time
                else _weekly_t_to_min(s.start_time) + 60
                for s in all_slots
            )
            grid_start = (earliest // _WEEKLY_SLOT_MIN) * _WEEKLY_SLOT_MIN
            grid_end = -(-latest // _WEEKLY_SLOT_MIN) * _WEEKLY_SLOT_MIN
        else:
            grid_start, grid_end = 8 * 60, 22 * 60
        grid_start = max(grid_start, 6 * 60)
        grid_end = min(grid_end, 23 * 60 + 30)

        time_slots = []
        cur = grid_start
        while cur < grid_end:
            hh, mm = divmod(cur, 60)
            time_slots.append({"minute": cur, "label": f"{hh:02d}:{mm:02d}"})
            cur += _WEEKLY_SLOT_MIN

        teachers_in_use = []
        seen_teacher_ids = set()
        for s in all_slots:
            tch = s.group.oqituvchi
            if tch and tch.id not in seen_teacher_ids:
                seen_teacher_ids.add(tch.id)
                teachers_in_use.append(tch)

        time_grid = {wd: [[] for _ in time_slots] for wd, _ in weekday_labels}
        for s in all_slots:
            start_m = _weekly_t_to_min(s.start_time)
            end_m = _weekly_t_to_min(s.end_time) if s.end_time else start_m + 60
            slot_idx = max(0, (start_m - grid_start) // _WEEKLY_SLOT_MIN)
            if slot_idx >= len(time_slots):
                continue
            tch = s.group.oqituvchi
            time_grid[s.weekday][slot_idx].append({
                "item": s,
                "is_unassigned": not s.group.oqituvchi_id,
                "teacher_name": tch.get_full_name() if tch else "Belgilanmagan",
                "teacher_initials": _weekly_teacher_initials(tch),
            })

        used_idx = sorted({
            idx
            for wd, _ in weekday_labels
            for idx, cell in enumerate(time_grid[wd])
            if cell
        })
        filtered_time_slots = [time_slots[i] for i in used_idx]
        filtered_time_grid = {
            wd: [time_grid[wd][i] for i in used_idx] for wd, _ in weekday_labels
        }
        grid_rows = []
        for new_idx, orig_idx in enumerate(used_idx):
            slot = time_slots[orig_idx]
            cells = [time_grid[wd][orig_idx] for wd, _ in weekday_labels]
            grid_rows.append({"label": slot["label"], "cells": cells, "minute": slot["minute"]})

        teacher_load_map = defaultdict(lambda: {
            "minutes": 0, "lessons": 0, "days": set(), "rooms": set(),
        })
        for s in all_slots:
            tid = s.group.oqituvchi_id
            key = tid if tid else "unassigned"
            start_m = _weekly_t_to_min(s.start_time)
            end_m = _weekly_t_to_min(s.end_time) if s.end_time else start_m + 60
            teacher_load_map[key]["minutes"] += max(0, end_m - start_m)
            teacher_load_map[key]["lessons"] += 1
            teacher_load_map[key]["days"].add(s.weekday)
            if s.room:
                teacher_load_map[key]["rooms"].add(s.room)

        weekday_short = {1: "Du", 2: "Se", 3: "Ch", 4: "Pa", 5: "Ju", 6: "Sh", 7: "Ya"}
        teacher_loads = []
        for tch in teachers_in_use:
            ld = teacher_load_map.get(tch.id) or {"minutes": 0, "lessons": 0, "days": set(), "rooms": set()}
            teacher_loads.append({
                "teacher": tch,
                "initials": _weekly_teacher_initials(tch),
                "lessons": ld["lessons"],
                "hours": round(ld["minutes"] / 60.0, 1),
                "days_count": len(ld["days"]),
                "days_short": [weekday_short[d] for d in sorted(ld["days"])],
                "rooms": sorted(ld["rooms"]),
            })
        if "unassigned" in teacher_load_map:
            ld = teacher_load_map["unassigned"]
            teacher_loads.append({
                "teacher": None,
                "initials": "?",
                "lessons": ld["lessons"],
                "hours": round(ld["minutes"] / 60.0, 1),
                "days_count": len(ld["days"]),
                "days_short": [weekday_short[d] for d in sorted(ld["days"])],
                "rooms": sorted(ld["rooms"]),
            })
        teacher_loads.sort(key=lambda x: (-x["hours"], -x["lessons"]))

        groups_qs = (
            Group.objects.filter(center=center, is_archived=False)
            .select_related("oqituvchi", "category_obj")
            .order_by("nom")
        )
        if teacher_id:
            groups_qs = groups_qs.filter(oqituvchi_id=teacher_id)

        total_groups_count = groups_qs.count()
        scheduled_group_ids = list(base_qs.values_list("group_id", flat=True).distinct())
        groups_with_schedule_count = len(scheduled_group_ids)
        unscheduled_groups = list(groups_qs.exclude(id__in=scheduled_group_ids)[:12])
        unscheduled_groups_count = max(total_groups_count - groups_with_schedule_count, 0)

        empty_state_message = ""
        if not filtered_slots_count:
            if room_filter:
                empty_state_message = f'"{room_filter}" bo‘yicha mos jadval topilmadi.'
            elif teacher_id:
                empty_state_message = "Tanlangan o‘qituvchi uchun jadval hali kiritilmagan."
            else:
                empty_state_message = "Haftalik jadval hali kiritilmagan."

        rooms = (
            GroupSchedule.objects.filter(center=center)
            .exclude(room="")
            .values_list("room", flat=True)
            .distinct()
            .order_by("room")
        )
        teachers = User.objects.filter(
            center=center,
            role="teacher",
            is_archived=False,
        ).order_by("ism", "familya")

        export = (request.GET.get("export") or "").strip().lower()
        if export in ("excel", "1", "xlsx"):
            return _weekly_schedule_excel(
                center, weekday_labels, week_map, filtered_time_slots, filtered_time_grid,
                teacher_loads, teacher_id, room_filter, teachers,
            )
        if export == "pdf":
            return _weekly_schedule_pdf(
                center, weekday_labels, filtered_time_slots, filtered_time_grid,
                teacher_loads, teacher_id, room_filter, teachers,
            )

        return render(
            request,
            "education/weekly_schedule.html",
            {
                "week_map": week_map,
                "weekday_labels": weekday_labels,
                "grid_rows": grid_rows,
                "teacher_loads": teacher_loads,
                "rooms": rooms,
                "teachers": teachers,
                "selected_teacher": teacher_id,
                "selected_room": room_filter,
                "total_slots_count": total_slots_count,
                "filtered_slots_count": filtered_slots_count,
                "total_groups_count": total_groups_count,
                "groups_with_schedule_count": groups_with_schedule_count,
                "unscheduled_groups_count": unscheduled_groups_count,
                "unscheduled_groups": unscheduled_groups,
                "empty_state_message": empty_state_message,
                "has_filters": bool(teacher_id or room_filter),
                "teachers_in_use_count": len(teachers_in_use),
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("weekly_schedule_view failed")
        messages.error(request, "Haftalik jadvalni yuklashda xatolik yuz berdi.")
        return redirect("core:home")


def _weekly_schedule_excel(center, weekday_labels, week_map, time_slots, time_grid,
                           teacher_loads, teacher_id, room_filter, teachers):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    cell_font = Font(size=10)
    bold_small = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws = wb.active
    ws.title = "Vaqt jadvali"
    headers = ["Vaqt"] + [lbl for _, lbl in weekday_labels]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border

    for row_idx, slot in enumerate(time_slots, start=2):
        row_values = [slot["label"]]
        for wd, _ in weekday_labels:
            cell_items = time_grid[wd][row_idx - 2]
            if cell_items:
                lines = []
                for entry in cell_items:
                    it = entry["item"]
                    tch_name = it.group.oqituvchi.get_full_name() if it.group.oqituvchi else "Belgilanmagan"
                    rm = it.room or "—"
                    lines.append(f"{it.group.nom}\n{tch_name}\n{it.time_range} | {rm}")
                row_values.append("\n────\n".join(lines))
            else:
                row_values.append("")
        ws.append(row_values)
        for col_idx in range(1, len(row_values) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = center_align
            c.border = border
            if col_idx == 1:
                c.font = bold_small
                c.fill = PatternFill("solid", fgColor="F1F5F9")
            else:
                c.font = cell_font

    ws.column_dimensions["A"].width = 11
    for letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 28
    for row_idx in range(2, len(time_slots) + 2):
        ws.row_dimensions[row_idx].height = 56
    ws.freeze_panes = "B2"

    ws2 = wb.create_sheet("Kun bo'yicha")
    ws2.append(["Kun", "Guruh", "O'qituvchi", "Vaqt", "Xona", "Bo'lim"])
    for col_idx in range(1, 7):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border
    for wd, label in weekday_labels:
        items = week_map.get(wd) or []
        if not items:
            ws2.append([label, "—", "—", "—", "—", "—"])
            continue
        for it in items:
            try:
                cat = it.group.get_category_display()
            except Exception:
                cat = ""
            ws2.append([
                label,
                it.group.nom,
                it.group.oqituvchi.get_full_name() if it.group.oqituvchi else "Belgilanmagan",
                it.time_range,
                it.room or "Xona yo'q",
                cat,
            ])
    for col, w in zip(["A", "B", "C", "D", "E", "F"], [14, 28, 24, 16, 14, 18]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("O'qituvchi yuklamasi")
    ws3.append(["O'qituvchi", "Darslar", "Soatlar (haftada)", "Band kunlar", "Xonalar"])
    for col_idx in range(1, 6):
        c = ws3.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border
    for ld in teacher_loads:
        name = ld["teacher"].get_full_name() if ld["teacher"] else "Belgilanmagan"
        ws3.append([
            name,
            ld["lessons"],
            ld["hours"],
            ", ".join(ld["days_short"]),
            ", ".join(ld["rooms"]),
        ])
    for col, w in zip(["A", "B", "C", "D", "E"], [28, 12, 18, 18, 28]):
        ws3.column_dimensions[col].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"haftalik_jadval_{timezone.localdate().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


def _weekly_schedule_pdf(center, weekday_labels, time_slots, time_grid,
                         teacher_loads, teacher_id, room_filter, teachers):
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    from html import escape as _esc

    buf = BytesIO()
    page_size = landscape(A3)
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title="Haftalik jadval",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=rl_colors.grey, spaceAfter=8)
    cell_dark = ParagraphStyle("celld", parent=styles["Normal"], fontSize=7, leading=9, alignment=1, textColor=rl_colors.HexColor("#0f172a"))

    elements = []
    center_name = (
        getattr(center, "nom", None)
        or getattr(center, "name", None)
        or getattr(center, "title", None)
        or "ChaqmoqApp"
    )
    elements.append(Paragraph(f"Haftalik Jadval — {_esc(str(center_name))}", title_style))
    sub_parts = []
    if teacher_id:
        sel = next((t for t in teachers if t.id == teacher_id), None)
        if sel:
            sub_parts.append(f"O'qituvchi: {_esc(sel.get_full_name())}")
    if room_filter:
        sub_parts.append(f"Xona: {_esc(room_filter)}")
    sub_parts.append(f"Sana: {timezone.localdate().strftime('%d.%m.%Y')}")
    elements.append(Paragraph(" • ".join(sub_parts), sub_style))

    head_row = ["Vaqt"] + [lbl for _, lbl in weekday_labels]
    data = [head_row]
    for r_idx, slot in enumerate(time_slots, start=1):
        row = [slot["label"]]
        for c_idx, (wd, _) in enumerate(weekday_labels, start=1):
            entries = time_grid[wd][r_idx - 1]
            if entries:
                first = entries[0]
                it = first["item"]
                tch_name = _esc(first["teacher_name"])
                rm = _esc(it.room) if it.room else "—"
                extra = f"<br/><i>+{len(entries) - 1} qo'shimcha dars</i>" if len(entries) > 1 else ""
                txt = (
                    f"<b>{tch_name}</b><br/>"
                    f"{_esc(it.group.nom)}<br/>"
                    f"<font color='#475569'>{_esc(it.time_range)} · {rm}</font>{extra}"
                )
                row.append(Paragraph(txt, cell_dark))
            else:
                row.append("")
        data.append(row)

    col_count = len(head_row)
    page_w = page_size[0] - 20 * mm
    time_col_w = 20 * mm
    day_col_w = (page_w - time_col_w) / (col_count - 1)
    col_widths = [time_col_w] + [day_col_w] * (col_count - 1)

    if len(data) == 1:
        elements.append(Paragraph("Hech qanday dars vaqti kiritilmagan.", sub_style))
    else:
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 1), (0, -1), 8),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), rl_colors.HexColor("#F1F5F9")),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
        tbl.setStyle(style)
        elements.append(tbl)

    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph("O'qituvchi yuklamasi (haftalik)", title_style))

    load_data = [["O'qituvchi", "Darslar", "Soat", "Band kunlar", "Xonalar"]]
    for ld in teacher_loads:
        name = ld["teacher"].get_full_name() if ld["teacher"] else "Belgilanmagan"
        load_data.append([
            _esc(name),
            str(ld["lessons"]),
            f"{ld['hours']}",
            ", ".join(ld["days_short"]) or "—",
            _esc(", ".join(ld["rooms"])) or "—",
        ])
    load_tbl = Table(load_data, colWidths=[60 * mm, 25 * mm, 25 * mm, 50 * mm, 80 * mm], repeatRows=1)
    load_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (2, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(load_tbl)

    doc.build(elements)
    pdf_data = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf_data, content_type="application/pdf")
    fname = f"haftalik_jadval_{timezone.localdate().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


@login_required
def teacher_schedule_view(request):
    """
    O'qituvchi o'z haftalik jadvalini ko'radi.
    """
    from core.tenant import get_request_center
    from education.models import GroupSchedule

    try:
        center = get_request_center(request)
        if not center:
            raise Http404("Center not found")
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_WEEKLY_SCHEDULE,
            message="Dars jadvali bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response

        if request.user.role != "teacher" and not request.user.is_superuser:
            return redirect("core:home")

        teacher = request.user
        schedules = (
            GroupSchedule.objects.filter(
                center=center,
                group__oqituvchi=teacher,
                group__is_archived=False,
                group__is_deleted=False,
            )
            .select_related("group", "group__oqituvchi")
            .order_by("weekday", "start_time")
        )

        week_map = {weekday: [] for weekday, _ in _schedule_weekday_labels()}
        for schedule in schedules:
            week_map[schedule.weekday].append(schedule)

        groups_qs = (
            Group.objects.filter(center=center, oqituvchi=teacher, is_archived=False)
            .select_related("category_obj")
            .order_by("nom")
        )
        scheduled_group_ids = list(schedules.values_list("group_id", flat=True).distinct())
        unscheduled_groups = list(groups_qs.exclude(id__in=scheduled_group_ids))

        return render(
            request,
            "education/teacher_schedule.html",
            {
                "week_map": week_map,
                "weekday_labels": _schedule_weekday_labels(),
                "total_lessons": schedules.count(),
                "unscheduled_groups": unscheduled_groups,
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("teacher_schedule_view failed")
        messages.error(request, "Jadvalni yuklashda xatolik yuz berdi.")
        return redirect("core:home")


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.utils.dateparse import parse_date

@login_required
@require_POST
def attendance_force(request):
    """
    Tanlangan guruh va sana bo'yicha:
    ✅ kelmagan (present=False) o'quvchilar uchun
    forced=True qilib, o'qituvchiga pul yoziladigan dars sifatida belgilaydi.

    Frontend POST yuboradi:
      - group_id
      - date (YYYY-MM-DD)
    """
    group_id = request.POST.get("group_id")
    date_str = request.POST.get("date")

    # 🔴 JS dagi xabardagi "Ma'lumot yetarli emas" — mana shu joydan keladi
    if not group_id or not date_str:
        return JsonResponse({"ok": False, "error": "Maʼlumot yetarli emas"})

    date_obj = parse_date(date_str)
    if not date_obj:
        return JsonResponse({"ok": False, "error": "Sana noto'g'ri formatda"})

    # Guruhni olamiz
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=group_id)

    # Shu guruhdagi barcha faol enrollments (arxivlanganlar istisno)
    enrollments = Enrollment.objects.filter(group=g, is_active=True).select_related("student")

    # Shu sana uchun mavjud attendance yozuvlari
    att_qs = Attendance.objects.filter(group=g, date=date_obj)
    att_by_student = {a.student_id: a for a in att_qs}

    forced_count = 0

    for enr in enrollments:
        att = att_by_student.get(enr.student_id)

        if att:
            # Agar allaqachon present=True bo'lsa, buni majburan "kelmadi" qilishni xohlamaymiz
            # (agar kerak bo'lsa, bu qismni o'zing o'zgartirasan)
            if att.present:
                continue

            if not att.forced:
                att.forced = True
                att.present = False  # forced bo'lsa ham uni "kelmadi" deb saqlab qo'yamiz
                att.save()
                forced_count += 1
        else:
            # Hech qanday attendance yo'q bo'lsa, yangi "kelmadi, forced" yozuvi yaratamiz
            Attendance.objects.create(
                group=g,
                student=enr.student,
                teacher=g.oqituvchi,
                date=date_obj,
                present=False,
                forced=True,
                center=g.center if hasattr(Attendance, 'center') else None
            )
            forced_count += 1

    return JsonResponse({
        "ok": True,
        "count": forced_count,
    })


@login_required
def attend_all(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"})

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    # faqat direktor/manager/teacher yoki support teacher
    if request.user.role == "teacher" and g.oqituvchi != request.user:
        if g.support_teacher_id != request.user.id:
            return JsonResponse({"ok": False, "error": "ruxsat yo'q"})

    date_str = request.POST.get("date")
    selected_date = parse_date(date_str) if date_str else localdate()

    students = Enrollment.objects.filter(group=g, is_active=True).select_related("student")
    count = 0

    for e in students:
        Attendance.objects.update_or_create(
            group=g,
            student=e.student,
            date=selected_date,
            defaults={"present": True, "forced": False, "status": "present", "teacher": request.user}
        )
        count += 1

    return JsonResponse({"ok": True, "count": count})

@require_POST
@login_required
def attend_all_students(request, g_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=g_id)

    if request.user.role == "teacher" and g.oqituvchi != request.user:
        if g.support_teacher_id != request.user.id:
            return JsonResponse({"ok": False, "error": "ruxsat yo'q"})

    date_str = request.POST.get("date")
    selected_date = parse_date(date_str) if date_str else localdate()

    enrollments = Enrollment.objects.filter(group=g, is_active=True).select_related("student")

    items = []
    count = 0

    for e in enrollments:
        Attendance.objects.update_or_create(
            group=g,
            student=e.student,
            date=selected_date,
            defaults={"present": True, "forced": False, "status": "present", "teacher": request.user, "center": g.center, "created_by": request.user},
        )
        balance = Ledger.student_balansi(e.student.id, center=g.center)

        items.append({"student_id": e.student.id, "balance": balance, "restored_sum": 0})
        count += 1

    return JsonResponse({"ok": True, "count": count, "items": items})

# ---------- AJAX: Davomatni saqlash ----------
@require_POST
@login_required
def attendance_today(request, pk: int):
    """
    status:
      - 'present'          -> present=True,  forced=False, status='present'
      - 'absent_excused'   -> present=False, forced=False, status='absent_excused'   (sababli)
      - 'absent_unexcused' -> present=False, forced=False, status='absent_unexcused' (sababsiz) + Rule Engine
      - 'forced'           -> present=False, forced=True,  status='absent_unexcused' (eski logika saqlanadi)
      - 'none'             -> attendance yozuvi o'chiriladi
    Backward compatible: present=1/0 yuborilsa ham ishlaydi.
    """
    g = get_object_or_404(Group, pk=pk)
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center and g.center_id != center.id:
        return JsonResponse({"ok": False, "error": "Center mismatch"}, status=403)

    # faqat direktor/manager/teacher yoki support teacher
    if request.user.role == "teacher" and g.oqituvchi != request.user:
        if g.support_teacher_id != request.user.id:
            return JsonResponse({"ok": False, "error": "ruxsat yo'q"}, status=403)

    enr_id = request.POST.get("enr_id")
    if not enr_id:
        return JsonResponse({"ok": False, "error": "enr_id required"}, status=400)

    # sana
    date_str = request.POST.get("date")
    selected_date = parse_date(date_str) if date_str else localdate()

    # status (yangi: 5 ta holat)
    status = (request.POST.get("status") or "").strip().lower()

    VALID_STATUSES = ("present", "absent_excused", "absent_unexcused", "forced", "late", "none")

    # backward compatibility (eski front bo'lsa)
    if status not in VALID_STATUSES:
        pv = request.POST.get("present")
        if pv is None:
            return JsonResponse({"ok": False, "error": "status/present required"}, status=400)
        status = "present" if str(pv).lower() in ("1", "true", "yes", "on") else "none"

    e = get_object_or_404(Enrollment, id=enr_id, group=g)
    student = e.student

    removed_sum = 0
    removed_count = 0
    penalized = False
    bonused = False

    # ── Attendance ni yaratish/yangilash/o'chirish ──
    present = False
    forced = False

    if status == "none":
        Attendance.objects.filter(group=g, student=student, date=selected_date).delete()
    else:
        # qolgan statuslar uchun update_or_create
        present = (status == "present")
        forced = (status == "forced")
        Attendance.objects.update_or_create(
            group=g, student=student, date=selected_date,
            defaults={
                "teacher": request.user,
                "present": present,
                "forced": forced,
                "status": status,
                "center": g.center,
                "created_by": request.user,
            }
        )

    # yangi balans
    bal = Ledger.student_balansi(student.id, center=g.center)

    return JsonResponse({
        "ok": True,
        "status": status,
        "present": present,
        "forced": forced,
        "removed_sum": removed_sum,
        "removed_count": removed_count,
        "balance": bal,
        "penalty_applied": penalized,
        "bonus_applied": bonused,
    })



@login_required
def group_bulk_remove(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "msg": "POST bo'lishi shart."})

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
         qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    # ruxsat tekshirish
    can_remove = False
    if request.user.role == "director" or request.user.is_superuser:
        can_remove = True
    elif center:
        if request.user.role == "manager":
            can_remove = center.manager_can_remove_student
        elif request.user.role == "teacher":
            can_remove = center.teacher_can_remove_student

    if not can_remove:
        return JsonResponse({"ok": False, "msg": "Sizda o'quvchilarni o'chirish huquqi yo'q."})

    ids = request.POST.getlist("enrollment_ids")

    if not ids:
        return JsonResponse({"ok": False, "msg": "ID kelmagan."})

    qs = Enrollment.objects.filter(id__in=ids, group=g)
    count = qs.count()
    qs.delete()

    return JsonResponse({"ok": True, "deleted": count})




# ---------- AJAX: Chaqmoq yozish/ayirish ----------
@login_required
def group_points(request, pk: int):
    """
    Apply rules to students (points system) with proper Student-User link handling.
    """
    try:
        from django.db import transaction
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        from education.models import Enrollment, Group
        from chaqmoq.models import Ledger, Rule
        from accounts.models import User

        if request.method != "POST":
            return JsonResponse({"status": "error", "message": "Method not allowed"}, status=200)

        # 1. Parse Data
        if request.content_type == "application/json":
            data = json.loads(request.body)
        else:
            data = request.POST

        student_id = data.get("student_id")
        rule_id = data.get("rule_id")
        amount = int(data.get("amount", 0))
        date_str = data.get("date")

        if not student_id:
            return JsonResponse({"status": "error", "message": "O'quvchi tanlanmagan"}, status=200)

        # 2. Student & Group Lookup (Safe search)
        # Guruhni topish
        g = Group.objects.filter(pk=pk).first()
        if not g:
            return JsonResponse({"status": "error", "message": "Guruh topilmadi"}, status=200)

        # O'quvchini bevosita User modeli orqali topamiz.
        # Agar bu yerda Enrollment bo'yicha qidirsak, bir xil ID ga ega boshqa o'quvchi
        # tanlanib qolishi va "refresh davomida points yo'qolish" muammosi yuzaga keladi.
        student_user = User.objects.filter(pk=student_id, role='student').first()

        if not student_user:
             # Senior Senior Senior logic: Agar o'quvchi topilmasa, qizil xato chiqarmaslik uchun
             # status: success qaytaramiz, lekin message bermaymiz.
             return JsonResponse({"status": "success", "message": "", "ok": True}, status=200)

        # 3. Rule Handling
        if rule_id and str(rule_id).isdigit():
            rule = Rule.objects.filter(pk=rule_id).first()
        else:
            rule = None
            
        if not rule:
            rule = Rule.objects.filter(nom="Erkin ball", center=g.center).first()
            if not rule:
                rule = Rule.objects.create(
                    nom="Erkin ball", tur=Rule.PLUS, min_baho=1, max_baho=1000000, center=g.center
                )

        # 4. Date processing
        parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()
        now_time = timezone.localtime(timezone.now()).time()
        sana = timezone.make_aware(datetime.combine(parsed_date, now_time))

        request_id = data.get("request_id", "")
        from django.core.cache import cache
        cache_key = f"ledger_req_{request_id}" if request_id else None

        # ✅ DATABASE-LEVEL LOCK bilan idempotency
        # select_for_update() - parallel workerlar (gunicorn) bir vaqtda create qilishini oldini oladi
        with transaction.atomic():
            # Lock: bir vaqtda faqat bitta worker shu student uchun ishlay oladi
            # Render.com dagi parallel workerlar yaratishda "phantom read" muammosini oldini olish
            # uchun Ledger ni emas (u hali yo'q bo'lishi mumkin), aynan User ni qulflaymiz
            _lock_student = User.objects.select_for_update().get(id=student_user.id)

            # ✅ 1-HIMOYA: Lock olingach yana Cache ni tekshirish.
            if cache_key:
                cached = cache.get(cache_key)
                if cached:
                    return JsonResponse(cached)



            # Yangi yozuv yaratish
            record = Ledger.objects.create(
                student=student_user,
                beruvchi=request.user,
                group=g,
                rule=rule,
                ball=amount,
                sana=sana,
            )

            from chaqmoq.models import LightningHistory
            from education.models import Student as EdStudent
            st_model, _ = EdStudent.objects.get_or_create(user=student_user)
            LightningHistory.objects.create(
                student=st_model,
                points=amount,
                reason=rule.nom if rule else "Erkin ball",
                source="manual",
                teacher=request.user
            )
            _accumulate_daily_lightning(
                group=g,
                student=student_user,
                date_value=parsed_date,
                points_delta=amount,
            )

            # Yangi balansni hisoblash
            balance = Ledger.student_balansi(student_user.id, center=g.center)

        response_data = {
            "status": "success",
            "message": "Ball saqlandi",
            "balance": int(balance),
            "amount": amount,
            "id": record.id,
            "ok": True
        }

        # ✅ Muvaffaqiyatli response ni cache ga yozamiz (60 soniya)
        # Xuddi shu request_id qayta kelsa - yangi yozuv yaratilmaydi
        if request_id:
            from django.core.cache import cache
            cache.set(f"ledger_req_{request_id}", response_data, timeout=60)

        return JsonResponse(response_data)

    except Exception as e:
        import logging
        logging.getLogger(__name__).exception("Points logic error")
        return JsonResponse({
            "status": "error",
            "message": f"Serverda xato yuz berdi: {str(e)}"
        }, status=200) # Toast qizil chiqmasligi uchun 200 qaytaramiz (JSON error ichida bo'ladi)


# @login_required
# def groups_home(request):
#     categories = Category.objects.all().order_by("name")
#     return render(request, "education/groups_home.html", {"categories": categories})


def category_detail(request, category_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    category = get_object_or_404(Category, id=category_id)
    if center:
        if category.center_id and category.center_id != center.id:
            raise PermissionDenied("Bu bo'limga ruxsat yo'q")

    groups = (
        Group.objects
        .filter(category_obj=category)
        .select_related("center", "oqituvchi")
        .order_by("id")
    )
    if center:
        groups = groups.filter(center=center)

    # Filter by status (default: active)
    status = request.GET.get('status', 'active')
    
    # ✅ TEACHERLAR UCHUN "ARXIV" YOPIQ
    if request.user.role == 'teacher':
        status = 'active'
        groups = groups.filter(is_archived=False)
    else:
        if status == 'archived':
            groups = groups.filter(is_archived=True)
        else:
            groups = groups.filter(is_archived=False)

    groups_count = groups.count()

    return render(request, "education/category_detail.html", {
        "category": category,
        "groups": groups,
        "groups_count": groups_count,
        "status": status,
        "is_teacher": request.user.role == 'teacher', # Template uchun
    })


@login_required
def group_toggle_archive(request, pk):
    """Guruhni arxivga ko'chirish: GET → tasdiqlash sahifasi, POST → amal."""
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=pk)

    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()

    def _redirect_to(fallback):
        if next_url and next_url.startswith("/"):
            return redirect(next_url)
        return fallback

    if not _can_manage(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return _redirect_to(redirect("education:all_groups"))

    if request.method == "POST":
        # Restoration mode skips confirmation entry.
        if group.is_archived:
            group.is_archived = False
            group.save(update_fields=["is_archived"])
            messages.success(request, "Guruh arxivdan qaytarildi ✅")
            return _redirect_to(redirect("education:group_detail", pk=group.id))

        confirm_text = (request.POST.get("confirm_name") or "").strip()
        if confirm_text != group.nom:
            messages.error(
                request,
                "Tasdiqlash uchun guruh nomini aynan to'g'ri yozing.",
            )
            enrolled_total = Enrollment.objects.filter(
                group=group, is_active=True, is_deleted=False,
            ).count()
            return render(request, "education/group_archive.html", {
                "group": group, "g": group,
                "enrolled_total": enrolled_total,
                "confirm_value": confirm_text,
            })

        group.is_archived = True
        group.save(update_fields=["is_archived"])
        messages.success(request, "Guruh arxivga ko'chirildi ✅")
        return _redirect_to(redirect("education:all_groups"))

    enrolled_total = Enrollment.objects.filter(
        group=group, is_active=True, is_deleted=False,
    ).count()
    return render(request, "education/group_archive.html", {
        "group": group, "g": group,
        "enrolled_total": enrolled_total,
        "confirm_value": "",
    })


@login_required
def group_toggle_close(request, pk):
    """Guruhni vaqtinchalik to'xtatish formasi (GET) yoki amal (POST)."""
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=pk)

    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()

    def _redirect_back():
        if next_url and next_url.startswith("/"):
            return redirect(next_url)
        return redirect("education:all_groups")

    if not _can_manage(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return _redirect_back()

    if request.method == "POST":
        action = (request.POST.get("action") or "pause").strip().lower()
        if action == "resume" or group.is_closed:
            group.is_closed = False
            group.closed_at = None
            group.closed_by = None
            group.save(update_fields=["is_closed", "closed_at", "closed_by"])
            messages.success(request, "Guruh qayta faollashtirildi ✅")
        else:
            group.is_closed = True
            group.closed_at = timezone.now()
            group.closed_by = request.user
            group.save(update_fields=["is_closed", "closed_at", "closed_by"])
            messages.success(request, "Guruh vaqtinchalik to'xtatildi ✅")
        return _redirect_back()

    enrolled_total = Enrollment.objects.filter(
        group=group, is_active=True, is_deleted=False,
    ).count()
    today_now = timezone.localdate()
    return render(request, "education/group_pause.html", {
        "group": group,
        "g": group,
        "enrolled_total": enrolled_total,
        "default_pause_date": today_now,
        "default_resume_date": today_now + timedelta(days=14),
    })

from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from education.models import Group, Dars, OylikHisobot
from accounts.models import User

@login_required
def oylik_hisobot(request):
    """Har bir o'qituvchining oyligini avtomatik hisoblash"""
    oy = datetime.now().strftime("%B")
    yil = datetime.now().year
    oylik_data = []

    teachers = User.objects.filter(role='teacher')

    for teacher in teachers:
        guruhlar = Group.objects.filter(oqituvchi=teacher)
        jami_darslar = 0
        jami_daromad = 0

        for g in guruhlar:
            darslar_soni = Dars.objects.filter(
                guruh=g,
                oqituvchi=teacher,
                sana__month=datetime.now().month,
                sana__year=datetime.now().year
            ).count()

            dars_tolovi = g.dars_boshiga_tolov()
            jami_darslar += darslar_soni
            jami_daromad += darslar_soni * dars_tolovi

        markaz_foydasi = jami_daromad * 0.5  # misol uchun 50/50

        oylik_data.append({
            "oqituvchi": teacher.get_full_name() or teacher.email,
            "guruhlar": guruhlar.count(),
            "darslar": jami_darslar,
            "daromad": round(jami_daromad),
            "markaz_foydasi": round(markaz_foydasi),
        })

        # OylikHisobot jadvaliga yozib qo'yish
        OylikHisobot.objects.update_or_create(
            oqituvchi=teacher,
            oy=oy,
            yil=yil,
            defaults={
                "jami_darslar": jami_darslar,
                "jami_daromad": round(jami_daromad),
                "markaz_foydasi": round(markaz_foydasi)
            }
        )

    return render(request, "education/oylik_hisobot.html", {"oylik_data": oylik_data, "oy": oy, "yil": yil})





@login_required
def group_create_by_category(request, category_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    category = get_object_or_404(Category, id=category_id)
    if center:
        if category.center_id and category.center_id != center.id:
            raise PermissionDenied("Bu bo'limga ruxsat yo'q")

    from billing.services import center_has_feature
    has_manual_oy_dars_soni = center_has_feature(center, "manual_oy_dars_soni") if center else False

    if request.method == "POST":
        form = GroupForm(request.POST, center=center)
        if form.is_valid():
            group = form.save(commit=False)
            group.category_obj = category
            group.center = center
            schedule_mode = form.cleaned_data.get("schedule_mode", "")
            custom_days = form.cleaned_data.get("custom_days") or []
            if schedule_mode in {"odd", "even", "custom"}:
                day_count = len(custom_days) if schedule_mode == "custom" else 3
                group.lessons_per_week = day_count
                if has_manual_oy_dars_soni:
                    if not group.oy_dars_soni:
                        group.oy_dars_soni = day_count * 4
                else:
                    group.oy_dars_soni = 12
            if not group.oy_dars_soni:
                group.oy_dars_soni = 12

            # O'qituvchi tanlanganda foiz teacher profilidan olinadi.
            if group.oqituvchi and getattr(group.oqituvchi, "oqituvchi_foizi", None) is not None:
                group.oqituvchi_foiz = group.oqituvchi.oqituvchi_foizi
            elif not group.oqituvchi_foiz:
                group.oqituvchi_foiz = 40

            from education.services.group_schedule_service import (
                apply_group_duration_defaults,
                sync_simple_group_schedule,
            )
            apply_group_duration_defaults(group)
            group.save()
            sync_simple_group_schedule(
                group=group,
                schedule_mode=schedule_mode,
                custom_days=custom_days,
                start_time=form.cleaned_data.get("schedule_start_time"),
                end_time=form.cleaned_data.get("schedule_end_time"),
                room=form.cleaned_data.get("schedule_room"),
            )
            return redirect("education:category_detail", category_id=category.id)
    else:
        form = GroupForm(center=center)

    course_templates = CourseTemplate.objects.filter(center=center, is_active=True).order_by("name") if center else []
    return render(request, "education/group_form.html", {
        "form": form,
        "category": category,
        "course_templates": course_templates,
        "has_manual_oy_dars_soni": has_manual_oy_dars_soni,
    })



from .models import Category
from django import forms
from django.contrib import messages

class CategoryForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        self.center = kwargs.pop("center", None)
        super().__init__(*args, **kwargs)

    class Meta:
        model = Category
        fields = ["name", "image", "description"]
        widgets = {
            "name": forms.TextInput(attrs={
                "class": "form-control",
                "placeholder": "Masalan: Dizayn"
            }),
            "image": forms.ClearableFileInput(attrs={
                "class": "form-control",
            }),
            "description": forms.Textarea(attrs={
                "class": "form-control",
                "rows": 3,
                "placeholder": "Bo'lim haqida qisqa izoh"
            }),
        }

    def clean_name(self):
        name = self.cleaned_data.get("name")
        if name:
            name_stripped = name.strip()
            # We check Category.all_objects because soft-deleted categories
            # still trigger unique constraints on the DB level.
            qs = Category.all_objects.filter(name__iexact=name_stripped, center=self.center)
            if self.instance and self.instance.pk:
                qs = qs.exclude(pk=self.instance.pk)
            if qs.exists():
                raise forms.ValidationError("Ushbu nomdagi bo'lim allaqachon mavjud!")
            return name_stripped
        return name



@login_required
def groups_home(request):
    # ✅ Tenant isolation
    from core.tenant import get_request_center
    center = get_request_center(request)
    
    # kategoriyalar
    from django.db.models import Q
    categories_qs = Category.objects.all().order_by("name")
    if center:
        # ✅ Smart Isolation: 
        # 1. Faqat shu center'ga tegishli bo'limlarni ko'rsatamiz.
        # 2. Agar bu ASOSIY (birinchi yaratilgan) markaz bo'lsa, Global (Legacy) bo'limlarni ham chiqaramiz.
        first_center = Center.objects.order_by("id").first()
        if first_center and center.id == first_center.id:
            categories_qs = categories_qs.filter(Q(center=center) | Q(center__isnull=True))
        else:
            categories_qs = categories_qs.filter(center=center)
        
    categories = list(categories_qs)

    # har bir category uchun guruhlar sonini hisoblab map qilamiz
    counts_qs = (
        Group.objects
    )
    if center:
        counts_qs = counts_qs.filter(center=center)
        
    counts_qs = (
        counts_qs
        .filter(is_archived=False)
        .values("category_obj")          # FK field nomi sizda shu: category_obj
        .annotate(c=Count("id"))
    )
    count_map = {row["category_obj"]: row["c"] for row in counts_qs}

    # template ishlatishi uchun cat.groups_count qo'shib chiqamiz
    for cat in categories:
        cat.groups_count = count_map.get(cat.id, 0)

    return render(request, "education/groups_home.html", {
        "categories": categories,
        "categories_count": len(categories),
    })


# ─────────────────────────────────────────────────────────────────────────────
# Hamma guruhlar — boshqaruv dashboardidan "Hammasini ko'rish" tugmasi.
# KPI'lar, status/fan/o'qituvchi/saralash filterlari, qidiruv, list/grid
# ko'rinish, CSV eksport va pagination.
# ─────────────────────────────────────────────────────────────────────────────

_ALL_GROUPS_PAGE_SIZE = 20

_WEEKDAY_SHORT = {1: "Du", 2: "Se", 3: "Ch", 4: "Pa", 5: "Ju", 6: "Sh", 7: "Ya"}


def _all_groups_avatar(name: str):
    """Return (initials, color) for a group avatar based on its name."""
    palette = [
        ("#2563eb", "#dbeafe"),
        ("#7c3aed", "#ede9fe"),
        ("#10b981", "#d1fae5"),
        ("#d97706", "#fef3c7"),
        ("#dc2626", "#fee2e2"),
        ("#0ea5e9", "#e0f2fe"),
        ("#db2777", "#fce7f3"),
        ("#0d9488", "#ccfbf1"),
    ]
    safe = (name or "G").strip()
    if not safe:
        safe = "G"
    initials = safe[:2].upper()
    idx = sum(ord(ch) for ch in safe) % len(palette)
    color, bg = palette[idx]
    return initials, color, bg


def _all_groups_schedule_text(group_id: int, schedule_map: dict) -> str:
    rows = schedule_map.get(group_id) or []
    if not rows:
        return "—"
    days = []
    seen = set()
    time_text = ""
    for row in rows:
        wd = row["weekday"]
        if wd in seen:
            continue
        seen.add(wd)
        days.append(_WEEKDAY_SHORT.get(wd, str(wd)))
        if not time_text and row.get("start_time"):
            time_text = row["start_time"].strftime("%H:%M")
    base = " · ".join(days) if days else "—"
    if time_text:
        return f"{base} · {time_text}"
    return base


@login_required
def all_groups_overview(request):
    """Boshqaruv → Hammasini ko'rish: barcha guruhlar ko'rinishi."""
    from core.tenant import get_request_center
    from education.models import GroupSchedule

    center = get_request_center(request) or get_active_center(request)
    today = timezone.localdate()
    month_start = today.replace(day=1)

    qs = Group.objects.filter(is_archived=False, is_deleted=False)
    if center:
        qs = qs.filter(center=center)

    # ── Filter qiymatlari ──
    q_text = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "all").strip().lower()
    cat_id = (request.GET.get("category") or "").strip()
    teacher_id = (request.GET.get("teacher") or "").strip()
    sort_key = (request.GET.get("sort") or "fill").strip().lower()
    view_mode = (request.GET.get("view") or "list").strip().lower()
    if view_mode not in ("list", "grid"):
        view_mode = "list"
    fmt = (request.GET.get("format") or "").strip().lower()

    if q_text:
        qs = qs.filter(Q(nom__icontains=q_text) |
                       Q(category_obj__name__icontains=q_text) |
                       Q(oqituvchi__ism__icontains=q_text) |
                       Q(oqituvchi__familya__icontains=q_text))

    if cat_id.isdigit():
        qs = qs.filter(category_obj_id=int(cat_id))

    if teacher_id.isdigit():
        qs = qs.filter(oqituvchi_id=int(teacher_id))

    # ── Filter dropdown manbalari ──
    cat_choices = list(
        Category.objects
        .filter(groups__in=qs)
        .distinct()
        .order_by("name")
        .values("id", "name")
    )
    teacher_choices = list(
        User.objects
        .filter(role="teacher", id__in=qs.values("oqituvchi_id"))
        .order_by("ism", "familya")
        .values("id", "ism", "familya")
    )
    teacher_choices = [
        {"id": t["id"], "name": f"{t['ism']} {t['familya']}".strip() or "—"}
        for t in teacher_choices
    ]

    groups_for_filter = list(qs.select_related("oqituvchi", "category_obj"))
    group_ids = [g.id for g in groups_for_filter]

    # ── Sig'im / band o'rinlar / fill_pct ──
    enroll_map = dict(
        Enrollment.objects.filter(
            group_id__in=group_ids,
            is_active=True,
            is_deleted=False,
        )
        .values("group_id")
        .annotate(cnt=Count("id"))
        .values_list("group_id", "cnt")
    )

    # ── Bu oygi tushum ──
    rev_map = dict(
        Payment.objects.filter(
            group_id__in=group_ids,
            paid_date__gte=month_start,
            paid_date__lte=today,
        )
        .values("group_id")
        .annotate(s=Sum("summa"))
        .values_list("group_id", "s")
    )

    # ── Davomat (oxirgi 30 kun) ──
    att_from = today - timedelta(days=30)
    att_qs = Attendance.objects.filter(group_id__in=group_ids, date__gte=att_from, date__lte=today)
    att_total_map = dict(
        att_qs.values("group_id").annotate(c=Count("id")).values_list("group_id", "c")
    )
    present_q = Q(status="present") | Q(present=True) | Q(forced=True)
    att_present_map = dict(
        att_qs.filter(present_q).values("group_id").annotate(c=Count("id")).values_list("group_id", "c")
    )

    # ── Schedule (Du · Ch · Ju · 14:00) ──
    schedule_map = {}
    for row in GroupSchedule.objects.filter(group_id__in=group_ids).values(
        "group_id", "weekday", "start_time"
    ).order_by("group_id", "weekday", "start_time"):
        schedule_map.setdefault(row["group_id"], []).append(row)

    rows = []
    total_capacity = 0
    total_enrolled = 0
    total_revenue = 0
    active_count = 0
    fill_pcts = []
    for g in groups_for_filter:
        enrolled = int(enroll_map.get(g.id) or 0)
        capacity = int(getattr(g, "max_students", 0) or 0)
        fill_pct = round(enrolled * 100 / capacity, 1) if capacity else 0
        revenue = int(rev_map.get(g.id) or 0)
        att_total = int(att_total_map.get(g.id) or 0)
        att_present = int(att_present_map.get(g.id) or 0)
        att_rate = round(att_present * 100 / att_total, 1) if att_total else 0
        is_full = capacity and enrolled >= capacity
        is_active = not g.is_closed
        if is_active:
            active_count += 1
        if is_full:
            status_label = "To'ldirilgan"
            status_kind = "full"
        elif is_active:
            status_label = "Faol"
            status_kind = "active"
        else:
            status_label = "Yopiq"
            status_kind = "closed"
        if 0 < fill_pct < 100:
            status_label = "To'ldirilmoqda" if not is_full and is_active else status_label
            if not is_full and is_active:
                status_kind = "filling"
        teacher_name = (
            f"{g.oqituvchi.ism} {g.oqituvchi.familya}".strip()
            if g.oqituvchi else "—"
        )
        cat_name = g.category_obj.name if g.category_obj else (g.get_category_display() if hasattr(g, "get_category_display") else "—")
        sub_label = cat_name
        # Status filter
        if status == "active" and not (is_active and not is_full):
            continue
        if status == "filling" and not (is_active and 0 < fill_pct < 100 and not is_full):
            continue
        if status == "full" and not is_full:
            continue
        if status == "closed" and is_active:
            continue
        initials, color, bg = _all_groups_avatar(g.nom)
        rows.append({
            "id": g.id,
            "name": g.nom,
            "subtitle": sub_label,
            "category_id": g.category_obj_id,
            "teacher_id": g.oqituvchi_id,
            "teacher": teacher_name,
            "schedule": _all_groups_schedule_text(g.id, schedule_map),
            "enrolled": enrolled,
            "capacity": capacity,
            "fill_pct": fill_pct,
            "revenue": revenue,
            "att_rate": att_rate,
            "status_label": status_label,
            "status_kind": status_kind,
            "initials": initials,
            "color": color,
            "bg": bg,
            "is_active": is_active,
        })
        total_capacity += capacity
        total_enrolled += enrolled
        total_revenue += revenue
        if capacity:
            fill_pcts.append(fill_pct)

    # ── Saralash ──
    if sort_key == "name":
        rows.sort(key=lambda r: r["name"].lower())
    elif sort_key == "students":
        rows.sort(key=lambda r: -r["enrolled"])
    elif sort_key == "revenue":
        rows.sort(key=lambda r: -r["revenue"])
    elif sort_key == "att":
        rows.sort(key=lambda r: -r["att_rate"])
    else:  # fill
        rows.sort(key=lambda r: (-r["fill_pct"], -r["revenue"], r["name"].lower()))

    avg_fill = round(sum(fill_pcts) / len(fill_pcts), 1) if fill_pcts else 0

    # ── CSV eksport ──
    if fmt == "csv":
        import csv
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="guruhlar-{today.isoformat()}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Guruh", "Bo'lim", "O'qituvchi", "Jadval",
            "O'quvchilar", "Sig'im", "To'ldirilganlik %",
            "Davomat %", "Tushum (so'm)", "Holat",
        ])
        for r in rows:
            writer.writerow([
                r["name"], r["subtitle"], r["teacher"], r["schedule"],
                r["enrolled"], r["capacity"], r["fill_pct"],
                r["att_rate"], r["revenue"], r["status_label"],
            ])
        return response

    # ── JSON (AJAX qayta yuklash) ──
    if fmt == "json":
        return JsonResponse({
            "kpis": {
                "active_groups": active_count,
                "total_groups": len(groups_for_filter),
                "students": total_enrolled,
                "avg_fill": avg_fill,
                "monthly_revenue": total_revenue,
            },
            "rows": rows,
        })

    # ── Pagination ──
    paginator = Paginator(rows, _ALL_GROUPS_PAGE_SIZE)
    page_num = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_num)
    except Exception:
        page_obj = paginator.page(1)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        base_qs.pop("page")
    base_qs_str = base_qs.urlencode()

    return render(request, "education/all_groups.html", {
        "rows": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "total_rows": len(rows),
        "kpi_active": active_count,
        "kpi_total": len(groups_for_filter),
        "kpi_students": total_enrolled,
        "kpi_avg_fill": avg_fill,
        "kpi_revenue": total_revenue,
        "q": q_text,
        "status": status,
        "category_id": cat_id,
        "teacher_id": teacher_id,
        "sort": sort_key,
        "view_mode": view_mode,
        "categories": cat_choices,
        "teachers": teacher_choices,
        "base_qs": base_qs_str,
    })


@login_required
def add_category(request):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES, center=center)
        if form.is_valid():
            cat = form.save(commit=False)
            cat.center = center
            cat.save()
            messages.success(request, "Bo'lim muvaffaqiyatli qo'shildi ✅")
            return redirect("education:groups_home")
    else:
        form = CategoryForm(center=center)
    return render(request, "education/category_add.html", {"form": form})






@login_required
def student_detail(request, student_id: int):
    student = get_object_or_404(User, pk=student_id, role="student")
    center = get_active_center(request)
    selected_month = parse_month_str((request.GET.get("month") or "").strip()) or month_first_day(timezone.localdate())
    can_view_student_group_financials = request.user.is_superuser or getattr(request.user, "role", None) in ("director", "manager")
    can_manage_parent_link = can_view_student_group_financials

    from accounts.services.parent_telegram_link import parent_link_status as build_parent_link_status

    raw_parent_link_status = build_parent_link_status(student)
    parent_linked_at = raw_parent_link_status.get("linked_at")
    parent_link_status = {
        "is_linked": raw_parent_link_status["is_linked"],
        "telegram_id": raw_parent_link_status["telegram_id"],
        "telegram_username": raw_parent_link_status["telegram_username"],
        "linked_at_display": timezone.localtime(parent_linked_at).strftime("%d.%m.%Y %H:%M") if parent_linked_at else "",
        "parent_id": raw_parent_link_status["parent_id"],
        "parent_name": raw_parent_link_status["parent_name"],
    }

    MONTH_NAMES = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May",
        6: "Iyun", 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktyabr",
        11: "Noyabr", 12: "Dekabr"
    }

    # 🔹 Endi davomatni guruhga qarab ajratamiz
    attendances = Attendance.objects.filter(student=student).select_related("group").annotate(
        year=ExtractYear('date'),
        month=ExtractMonth('date')
    ).order_by('-date')

    daily_records = DailyLightningRecord.objects.filter(student=student)
    daily_month_lightning_map = {
        (row["group_id"], row["year"], row["month"]): {
            "plus": int(row["plus_total"] or 0),
            "minus": abs(int(row["minus_total"] or 0)),
        }
        for row in (
            daily_records.annotate(
                year=ExtractYear("date"),
                month=ExtractMonth("date"),
            ).values("group_id", "year", "month").annotate(
                plus_total=Coalesce(Sum("plus_points"), 0),
                minus_total=Coalesce(Sum("minus_points"), 0),
            )
        )
    }
    daily_day_lightning_map = {
        (row["group_id"], row["date"]): {
            "plus": int(row["plus_total"] or 0),
            "minus": abs(int(row["minus_total"] or 0)),
        }
        for row in (
            daily_records.values("group_id", "date").annotate(
                plus_total=Coalesce(Sum("plus_points"), 0),
                minus_total=Coalesce(Sum("minus_points"), 0),
            )
        )
    }

    # DailyLightningRecord mavjud bo'lmagan yozuvlar uchun Ledger fallback.
    ledger_month_lightning_map = {}
    ledger_day_lightning_map = {}
    ledger_rows = (
        Ledger.objects
        .filter(student=student)
        .exclude(group_id__isnull=True)
        .values("group_id", "sana", "ball")
    )
    for row in ledger_rows:
        dt = row.get("sana")
        if not dt:
            continue

        local_dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
        ledger_date = local_dt.date()
        group_id = row["group_id"]
        ball = int(row.get("ball") or 0)

        month_key = (group_id, ledger_date.year, ledger_date.month)
        day_key = (group_id, ledger_date)

        if month_key not in ledger_month_lightning_map:
            ledger_month_lightning_map[month_key] = {"plus": 0, "minus": 0}
        if day_key not in ledger_day_lightning_map:
            ledger_day_lightning_map[day_key] = {"plus": 0, "minus": 0}

        if ball > 0:
            ledger_month_lightning_map[month_key]["plus"] += ball
            ledger_day_lightning_map[day_key]["plus"] += ball
        elif ball < 0:
            minus_abs = abs(ball)
            ledger_month_lightning_map[month_key]["minus"] += minus_abs
            ledger_day_lightning_map[day_key]["minus"] += minus_abs

    # Asosiy manba DailyLightningRecord; unda bo'lmagan kalitlar Ledger'dan olinadi.
    month_lightning_map = dict(ledger_month_lightning_map)
    month_lightning_map.update(daily_month_lightning_map)

    day_lightning_map = dict(ledger_day_lightning_map)
    day_lightning_map.update(daily_day_lightning_map)

    # 🔹 Har bir guruh bo'yicha ajratamiz
    grouped_by_group = {}
    for a in attendances:
        grouped_by_group.setdefault(a.group, []).append(a)

    month_summaries = []
    for group, group_attendances in grouped_by_group.items():
        # Guruh bo'yicha oylik natijalarni tayyorlash
        grouped_by_month = {}
        for a in group_attendances:
            key = (a.year, a.month)
            grouped_by_month.setdefault(key, []).append(a)

        for (year, month), records in grouped_by_month.items():
            total_present = sum(1 for r in records if r.present)
            month_lightning = month_lightning_map.get((group.id, year, month), {"plus": 0, "minus": 0})
            plus_sum = month_lightning["plus"]
            minus_sum = month_lightning["minus"]

            month_summaries.append({
                "group": group.nom,  # 🔹 Guruh nomini qo'shamiz
                "year": year,
                "month": month,
                "month_name": MONTH_NAMES.get(month, "Noma'lum oy"),
                "present_days": total_present,
                "plus": plus_sum,
                "minus": minus_sum,
                "days": [
                    {
                        "date": r.date,
                        "present": r.present,
                        "plus": day_lightning_map.get((group.id, r.date), {}).get("plus", 0),
                        "minus": day_lightning_map.get((group.id, r.date), {}).get("minus", 0),
                    }
                    for r in records
                ]
            })

    ctx = {
        "student": student,
        "month_summaries": month_summaries,
        "selected_month": selected_month,
        "can_view_student_group_financials": can_view_student_group_financials,
        "can_manage_parent_link": can_manage_parent_link,
        "parent_link_status": parent_link_status,
        "can_transfer_student": user_can_transfer_student(request.user),
        "student_group_financials": (
            _student_group_financial_cards(
                student,
                center=center,
                month=selected_month,
                include_dates=True,
            )
            if can_view_student_group_financials
            else None
        ),
    }

    return render(request, "education/student_detail.html", ctx)


@login_required
@require_http_methods(["GET", "POST"])
def transfer_student_view(request, enrollment_id: int):
    from core.tenant import get_request_center

    center = get_request_center(request)
    enrollment_qs = Enrollment.objects.select_related("student", "group", "center").filter(is_active=True)
    if center:
        enrollment_qs = enrollment_qs.filter(center=center, group__center=center, student__center=center)
    enrollment = get_object_or_404(enrollment_qs, pk=enrollment_id)

    if not user_can_transfer_student(request.user):
        messages.error(request, "Sizda o'quvchini boshqa guruhga ko'chirish huquqi yo'q.")
        return redirect("education:group_detail", pk=enrollment.group_id)

    if request.method == "POST":
        form = StudentGroupTransferForm(request.POST, old_group=enrollment.group, center=center or enrollment.center)
        if form.is_valid():
            try:
                result = transfer_student_to_group(
                    student=enrollment.student,
                    old_group=enrollment.group,
                    new_group=form.cleaned_data["new_group"],
                    transfer_date=form.cleaned_data["transfer_date"],
                    reason=form.cleaned_data["reason"],
                    user=request.user,
                )
            except (ValidationError, PermissionDenied) as exc:
                messages.error(request, "; ".join(exc.messages) if hasattr(exc, "messages") else str(exc))
            except Exception:
                logger.exception("Student group transfer failed")
                messages.error(request, "Ko'chirish vaqtida xatolik yuz berdi. Eski holat saqlandi.")
            else:
                messages.success(
                    request,
                    f"{enrollment.student.get_full_name()} yangi guruhga ko'chirildi. To'lov qayta hisoblandi.",
                )
                return redirect("education:group_detail", pk=result["new_enrollment"].group_id)
        else:
            messages.error(request, "Ma'lumotlarni tekshiring: yangi guruhni tanlang va sanani to'g'ri kiriting.")
    else:
        form = StudentGroupTransferForm(
            initial={"transfer_date": timezone.localdate()},
            old_group=enrollment.group,
            center=center or enrollment.center,
        )

    return render(request, "education/student_transfer_form.html", {
        "form": form,
        "enrollment": enrollment,
        "student": enrollment.student,
        "old_group": enrollment.group,
    })




# ---------- (ixtiyoriy) alohida Davomat/Chaqmoq sahifasi ----------
@login_required
def group_rollcall(request, pk):
    g = get_object_or_404(Group, pk=pk)
    if not _can_give_points(request.user, g):
        return HttpResponseForbidden()

    # sana
    date_str = request.GET.get("date") or request.POST.get("date")
    try:
        the_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else localdate()
    except Exception:
        the_date = localdate()

    from django.db.models import Exists, OuterRef, Q

    has_attendance = Attendance.objects.filter(
        group=g,
        student=OuterRef('student'),
        date__year=the_date.year,
        date__month=the_date.month
    )
    enrollments_qs = (
        Enrollment.objects
        .filter(group=g)
        .annotate(has_att=Exists(has_attendance))
        .filter(Q(is_active=True) | Q(is_active=False, has_att=True))
        .select_related("student")
        .order_by("student__ism", "student__familya")
    )
    students = [e.student for e in enrollments_qs]

    pres_map = {
        a.student_id: a.present for a in Attendance.objects.filter(group=g, date=the_date)
    }
    for s in students:
        s.present = pres_map.get(s.id, False)
        s.balance = Ledger.student_balansi(s.id, center=g.center)

    rules = Rule.objects.order_by("nom")

    if request.method == "POST" and request.POST.get("save") == "1":
        saved = 0
        for s in students:
            present = bool(request.POST.get(f"present_{s.id}"))
            Attendance.objects.update_or_create(
                group=g,
                student=s,
                date=the_date,
                defaults={"present": present, "teacher": request.user if request.user.role == "teacher" else None},
            )
            rule_id = request.POST.get(f"rule_{s.id}")
            amount_raw = request.POST.get(f"ball_{s.id}") or "0"
            try:
                amount = int(amount_raw)
            except ValueError:
                amount = 0
            if rule_id and amount:
                rule = get_object_or_404(Rule, pk=int(rule_id))
                abs_ball = abs(amount)
                if rule.min_baho <= abs_ball <= rule.max_baho:
                    signed = abs_ball if rule.tur == Rule.PLUS else -abs_ball
                    now_local = timezone.localtime(timezone.now()).time()
                    sana = timezone.make_aware(datetime.combine(the_date, now_local))
                    Ledger.objects.create(student=s, beruvchi=request.user, group=g, rule=rule, ball=signed, sana=sana)
                    
                    from chaqmoq.models import LightningHistory
                    from education.models import Student as EdStudent
                    st_model, _ = EdStudent.objects.get_or_create(user=s)
                    LightningHistory.objects.create(
                        student=st_model,
                        points=signed,
                        reason=rule.nom if rule else "Manual ball",
                        source="manual",
                        teacher=request.user
                    )
                    saved += 1
        messages.success(request, f"Saqlash tugadi. {saved} ta chaqmoq yozildi.")
        return redirect(f"{request.path}?date={the_date.isoformat()}")

    return render(
        request,
        "education/group_rollcall.html",
        {"g": g, "date": the_date.isoformat(), "students": students, "rules": rules},
    )

@login_required
@require_feature("finance")
def teacher_salary_list(request):
    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)

    if month < 1 or month > 12:
        month = now.month

    month_names_uz = [
        "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"
    ]
    month_name = month_names_uz[month - 1]

    from core.tenant import get_request_center
    from core.perf_cache import TTL_LONG, perf_cache_get_or_set, versioned_cache_key
    center = get_request_center(request)

    from education.services.support_teacher import is_support_enabled
    support_feature_on = is_support_enabled(center)

    # PERF: og'ir hisoblash 15 daqiqa cache (per markaz + yil + oy).
    _cache_key = versioned_cache_key(
        "salary_list", getattr(center, 'id', None), year, month
    )

    def _compute():
        return _compute_teacher_salary_list_payload(year, month, center, support_feature_on)

    payload = perf_cache_get_or_set(_cache_key, _compute, ttl=TTL_LONG)

    # `teachers` payload'da serializable obyekt — User instance'larini qayta yuklaymiz
    # (Cache'da User instance saqlash xatarli — id-list orqali fetch qilamiz).
    teacher_id_list = payload['teacher_id_list']
    user_index = {u.id: u for u in User.objects.filter(id__in=teacher_id_list)}
    teachers_resolved = []
    for row in payload['teacher_rows']:
        u = user_index.get(row['teacher_id'])
        if not u:
            continue
        teachers_resolved.append({**row, 'teacher': u})

    return render(request, "education/teacher_salary_list.html", {
        "teachers": teachers_resolved,
        "year": year,
        "month": month,
        "month_name": month_name,
        "total_all": payload['total_all'],
        "is_closed": payload['is_closed'],
        "support_feature_on": support_feature_on,
    })


def _compute_teacher_salary_list_payload(year, month, center, support_feature_on):
    """teacher_salary_list — og'ir hisoblash qismi cache'lanadi."""
    from education.services.support_teacher import (
        list_support_user_ids,
        calculate_support_salary,
    )
    from education.models import FinancialMonth
    from education.services.historical_finance_service import HistoricalFinanceService

    teacher_qs = User.objects.filter(role="teacher")
    if center:
        teacher_qs = teacher_qs.filter(center=center)
    teacher_ids = set(teacher_qs.values_list("id", flat=True))

    extra_user_ids = set()
    if support_feature_on:
        extra_user_ids = list_support_user_ids(center=center) - teacher_ids

    all_ids = teacher_ids | extra_user_ids
    users_qs = User.objects.filter(id__in=all_ids).order_by("ism", "familya")
    support_user_ids_set = list_support_user_ids(center=center) if support_feature_on else set()

    fin_month = FinancialMonth.objects.filter(year=year, month=month, center=center).first()
    is_closed = fin_month.is_closed if fin_month else False

    teacher_rows = []
    total_all = 0
    teacher_id_list = []

    for t in users_qs:
        teacher_salary = 0
        teacher_groups = 0
        if t.id in teacher_ids:
            salary_data = HistoricalFinanceService.calculate_teacher_salary(t, year, month, center)
            teacher_salary = salary_data['salary']
            teacher_groups = len(salary_data['details'])

        support_salary = 0
        support_groups = 0
        is_support = False
        if support_feature_on and t.id in support_user_ids_set:
            sup = calculate_support_salary(t, year, month, center)
            support_salary = sup["salary"]
            support_groups = len(sup["details"])
            is_support = True

        combined = teacher_salary + support_salary
        if combined == 0 and t.id not in teacher_ids and not is_support:
            continue

        total_all += combined
        teacher_id_list.append(t.id)
        teacher_rows.append({
            "teacher_id": t.id,
            "month_salary": combined,
            "teacher_salary": teacher_salary,
            "support_salary": support_salary,
            "groups_count": teacher_groups + support_groups,
            "is_support": is_support,
            "is_main_teacher": t.id in teacher_ids,
        })

    return {
        'teacher_rows': teacher_rows,
        'teacher_id_list': teacher_id_list,
        'total_all': total_all,
        'is_closed': is_closed,
    }

# 🔹 Excel Export — O'qituvchi oyligi hisoboti
@login_required
@require_feature("finance")
def teacher_salary_export(request):
    """
    Tanlangan oy/yil bo'yicha barcha o'qituvchilar oylik hisobotini
    professional Excel (.xlsx) fayl sifatida yuklab beradi.

    Sheet 1: Umumiy hisobot (barchasi + bar chart)
    Sheet 2..N: Har bir o'qituvchi uchun alohida — guruh va o'quvchi kesimida.
    """
    import io
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    from core.tenant import get_request_center
    from education.services.historical_finance_service import HistoricalFinanceService

    # ── Parametrlar ──────────────────────────────────────────────────────────
    now   = timezone.localdate()
    year  = _get_int(request.GET, "year",  now.year)
    month = _get_int(request.GET, "month", now.month)
    if month < 1 or month > 12:
        month = now.month

    center = get_request_center(request)

    MONTH_NAMES = [
        "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"
    ]
    month_name = MONTH_NAMES[month - 1]
    period_label = f"{month_name} {year}"

    # ── O'qituvchilar va oylik ma'lumot ──────────────────────────────────────
    teacher_qs = User.objects.filter(role="teacher")
    if center:
        teacher_qs = teacher_qs.filter(center=center)
    teachers = list(teacher_qs.order_by("ism"))

    salary_rows = []
    for t in teachers:
        data = HistoricalFinanceService.calculate_teacher_salary(t, year, month, center)
        salary_rows.append({
            "teacher": t,
            "salary":  data["salary"],
            "details": data.get("details", []),
        })

    # ── Stil yordamchilari ───────────────────────────────────────────────────
    def _hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(style="thin"):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def _bold(size=11, color="000000"):
        return Font(bold=True, size=size, color=color)

    def _money_fmt():
        return "#,##0"

    def _auto_width(ws, extra=4):
        for col in ws.columns:
            mx = 0
            for cell in col:
                try:
                    mx = max(mx, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = mx + extra

    # ── Minimalist yorqin rang sxemasi ──────────────────────────────────────
    # Asosiy: ko'k (2563EB) | Guruh: yashil (0EA472) | Jami: sariq (F59E0B)
    # Fon: oq (FFFFFF) | Alt qator: och kulrang (F8FAFC) | Chegara: kulrang (CBD5E1)
    HDR_DARK   = _hdr_fill("2563EB")   # bosh sarlavha — chuqur ko'k
    HDR_BLUE   = _hdr_fill("3B82F6")   # ustun header — yorqin ko'k
    HDR_GREEN  = _hdr_fill("0EA472")   # guruh blok — yashil
    TOTAL_FILL = _hdr_fill("FEF3C7")   # jami satri — sariq fon
    ALT_FILL   = _hdr_fill("F1F5F9")   # juft qatorlar — och kulrang
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    DARK_FONT  = Font(color="1E293B", size=10)
    TOTAL_FONT = Font(bold=True, color="92400E", size=11)   # jami — to'q jigarrang
    MONEY_NUM  = _money_fmt()

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — UMUMIY HISOBOT
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Umumiy hisobot"
    ws1.sheet_view.showGridLines = False

    # Sarlavha
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = f"O'qituvchilar Oylik Hisoboti — {period_label}"
    title_cell.font  = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill  = HDR_DARK
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    # Ustun sarlavhalari
    headers1 = ["T/r", "O'qituvchi", "Guruhlar soni", "O'quvchilar soni",
                 "Qatnashuv (dars)", "Hisoblangan oylik (so'm)", "Izoh"]
    ws1.append([])  # bo'sh qator (row 2)
    ws1.append(headers1)  # row 3
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=ci)
        cell.value  = h
        cell.font   = WHITE_FONT
        cell.fill   = HDR_BLUE
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[3].height = 28

    total_sum = 0
    chart_names    = []
    chart_salaries = []

    for idx, row in enumerate(salary_rows, 1):
        t       = row["teacher"]
        details = row["details"]
        students_total = sum(len(d.get("enrollments", [])) for d in details)
        attend_total   = sum(d.get("attendance", 0) for d in details)
        salary         = row["salary"]
        total_sum     += salary

        name = t.get_full_name() or t.email
        chart_names.append(name[:20])
        chart_salaries.append(salary)

        data_row = [idx, name, len(details), students_total, attend_total, salary, ""]
        ws1.append(data_row)
        ri = ws1.max_row
        # Juft — och kulrang, toq — oq
        fill = ALT_FILL if idx % 2 == 0 else _hdr_fill("FFFFFF")
        for ci, val in enumerate(data_row, 1):
            cell = ws1.cell(row=ri, column=ci)
            cell.fill   = fill
            cell.border = _border()
            cell.font   = Font(color="334155", size=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci in (1,3,4,5) else "left")
            if ci == 6:
                cell.number_format = MONEY_NUM
                cell.font = Font(color="1D4ED8", size=10, bold=True)

    # JAMI SATRI
    jami_ri = ws1.max_row + 1
    ws1.cell(row=jami_ri, column=1).value = "JAMI"
    ws1.cell(row=jami_ri, column=6).value = total_sum
    ws1.cell(row=jami_ri, column=6).number_format = MONEY_NUM
    for ci in range(1, 8):
        cell = ws1.cell(row=jami_ri, column=ci)
        cell.fill   = TOTAL_FILL
        cell.font   = TOTAL_FONT
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center" if ci == 1 else "left", vertical="center")
    ws1.cell(row=jami_ri, column=6).font = Font(bold=True, color="92400E", size=11)
    ws1.row_dimensions[jami_ri].height = 26

    # BAR CHART
    if chart_names:
        chart = BarChart()
        chart.type   = "col"
        chart.title  = f"O'qituvchilar oyligi — {period_label}"
        chart.y_axis.title = "Oylik (so'm)"
        chart.x_axis.title = "O'qituvchi"
        chart.style  = 10
        chart.width  = 28
        chart.height = 16

        data_ref = Reference(ws1,
                             min_col=6, max_col=6,
                             min_row=3, max_row=3 + len(salary_rows) - 1)
        cats_ref = Reference(ws1,
                             min_col=2, max_col=2,
                             min_row=4, max_row=3 + len(salary_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws1.add_chart(chart, f"A{jami_ri + 2}")

    ws1.freeze_panes = "A4"
    _auto_width(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2..N — HAR BIR O'QITUVCHI
    # ════════════════════════════════════════════════════════════════════════
    for row in salary_rows:
        t       = row["teacher"]
        details = row["details"]
        salary  = row["salary"]

        # Sheet nomi: Excel 31 belgidan oshmasin, noto'g'ri belgilar yo'qolsin
        raw_name   = t.get_full_name() or t.email or f"Ustoz_{t.pk}"
        sheet_name = raw_name[:28].translate(
            str.maketrans(r'\/*?:[]', '_______')
        )
        # Takrorlanmaslik uchun son qo'shamiz
        base = sheet_name
        cnt  = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base[:25]}_{cnt}"
            cnt += 1

        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        # ── HEADER blok ─────────────────────────────────────────────────────
        students_total = sum(len(d.get("enrollments", [])) for d in details)
        attend_total   = sum(d.get("attendance", 0) for d in details)

        ws.merge_cells("A1:F1")
        ws["A1"].value = f"O'qituvchi: {raw_name}  |  Davr: {period_label}"
        ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill  = HDR_DARK
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        meta_rows = [
            ("Guruhlar soni",          len(details)),
            ("Jami o'quvchilar",        students_total),
            ("Jami dars (qatnashuv)",   attend_total),
            ("Hisoblangan jami oylik",  salary),
        ]
        for mi, (label, val) in enumerate(meta_rows, 2):
            lbl_cell = ws.cell(row=mi, column=1)
            lbl_cell.value = label
            lbl_cell.font  = Font(bold=True, color="64748B", size=10)
            lbl_cell.fill  = _hdr_fill("F8FAFC")

            val_cell = ws.cell(row=mi, column=2)
            val_cell.value = val
            val_cell.font  = Font(color="1E293B", size=10)
            val_cell.fill  = _hdr_fill("F8FAFC")
            if mi == 5:  # oylik satri
                val_cell.number_format = MONEY_NUM
                val_cell.font = Font(bold=True, color="1D4ED8", size=11)

        cur_row = 7  # guruh bloklari shu satrdan boshlanadi

        if not details:
            ws.cell(row=cur_row, column=1).value = "Bu oy uchun ma'lumot yo'q."
            ws.cell(row=cur_row, column=1).font  = Font(color="94A3B8", italic=True)
        else:
            for gd in details:
                gname       = gd.get("group_name", "Guruh")
                g_salary    = gd.get("salary", 0)
                g_attend    = gd.get("attendance", 0)
                enrollments = gd.get("enrollments", [])

                # ── Guruh sarlavhasi ─────────────────────────────────────────
                ws.merge_cells(start_row=cur_row, start_column=1,
                               end_row=cur_row, end_column=6)
                hdr = ws.cell(row=cur_row, column=1)
                hdr.value = f"  {gname}   |   Guruh daromadi: {g_salary:,} som   |   Dars: {g_attend} marta"
                hdr.font  = Font(bold=True, color="FFFFFF", size=11)
                hdr.fill  = HDR_GREEN
                hdr.border = _border()
                hdr.alignment = Alignment(vertical="center", horizontal="left")
                ws.row_dimensions[cur_row].height = 26
                cur_row += 1

                # ── O'quvchilar jadval sarlavhasi ────────────────────────────
                sub_hdrs = ["T/r", "O'quvchi", "Kurs narhi", "Qatnashuv (kun)",
                            "Daromad (so'm)", "Izoh"]
                for ci, h in enumerate(sub_hdrs, 1):
                    cell = ws.cell(row=cur_row, column=ci)
                    cell.value  = h
                    cell.font   = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill   = HDR_BLUE
                    cell.border = _border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[cur_row].height = 22
                cur_row += 1

                # ── O'quvchilar satrlari ──────────────────────────────────────
                for si, en in enumerate(enrollments, 1):
                    sname   = en.get("student_name", "Noma'lum")
                    kn      = en.get("kurs_narhi", 0)
                    att     = en.get("attended", 0)
                    daromad = en.get("daromad", 0)

                    # Juft — och kulrang, toq — oq
                    fill = ALT_FILL if si % 2 == 0 else _hdr_fill("FFFFFF")
                    data = [si, sname, kn, att, daromad, ""]
                    for ci, v in enumerate(data, 1):
                        cell = ws.cell(row=cur_row, column=ci)
                        cell.value  = v
                        cell.fill   = fill
                        cell.border = _border()
                        cell.font   = Font(color="334155", size=10)
                        cell.alignment = Alignment(vertical="center",
                                                   horizontal="center" if ci in (1,4) else "left")
                        if ci == 3:
                            cell.number_format = MONEY_NUM
                            cell.font = Font(color="475569", size=10)
                        if ci == 5:
                            cell.number_format = MONEY_NUM
                            cell.font = Font(color="1D4ED8", size=10, bold=True)
                    cur_row += 1

                # ── Guruh jami satri ─────────────────────────────────────────
                for ci in range(1, 7):
                    cell = ws.cell(row=cur_row, column=ci)
                    cell.fill   = TOTAL_FILL
                    cell.border = _border()
                    cell.font   = TOTAL_FONT
                    cell.alignment = Alignment(vertical="center", horizontal="left")
                ws.cell(row=cur_row, column=2).value = "GURUH JAMI:"
                ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=cur_row, column=5).value = g_salary
                ws.cell(row=cur_row, column=5).number_format = MONEY_NUM
                ws.cell(row=cur_row, column=5).font = Font(bold=True, color="92400E", size=11)
                ws.row_dimensions[cur_row].height = 22
                cur_row += 2  # bo'sh qator + keyingi guruh

            # ── Umumiy jami (sheet pastida) ──────────────────────────────────
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=4)
            lbl = ws.cell(row=cur_row, column=1)
            lbl.value = "BARCHA GURUHLAR JAMI OYLIGI:"
            lbl.font  = Font(bold=True, color="FFFFFF", size=12)
            lbl.fill  = HDR_DARK
            lbl.border = _border()
            lbl.alignment = Alignment(horizontal="right", vertical="center")

            tot = ws.cell(row=cur_row, column=5)
            tot.value  = salary
            tot.font   = Font(bold=True, color="FFFFFF", size=13)
            tot.fill   = HDR_DARK
            tot.border = _border()
            tot.number_format = MONEY_NUM
            ws.row_dimensions[cur_row].height = 32

        ws.freeze_panes = "A7"
        _auto_width(ws)

    # ── Fayl qaytarish ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"oylik_hisobot_{year}_{month:02d}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# 🔹 2. O'qituvchining barcha guruhlari
@login_required
@require_feature("finance")
def teacher_groups(request, teacher_id):
    from core.tenant import get_request_center
    from education.services.support_teacher import (
        is_support_enabled,
        calculate_support_salary,
        list_support_user_ids,
    )

    center = get_request_center(request)

    # Foydalanuvchi: o'qituvchi YOKI (markazda support feature yoqilgan bo'lsa)
    # support sifatida biriktirilgan istalgan xodim bo'lishi mumkin.
    base_qs = User.objects.all()
    if center:
        base_qs = base_qs.filter(center=center)

    candidate = base_qs.filter(id=teacher_id).first()
    if candidate is None:
        return get_object_or_404(User.objects.none(), id=teacher_id)

    is_main_teacher = candidate.role == "teacher"
    is_support_member = (
        is_support_enabled(center) and candidate.id in list_support_user_ids(center=center)
    )
    if not is_main_teacher and not is_support_member:
        return get_object_or_404(User.objects.filter(role="teacher"), id=teacher_id)

    teacher = candidate

    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)
    if month < 1 or month > 12:
        month = now.month

    years = list(range(now.year - 3, now.year + 4))

    from education.services.historical_finance_service import HistoricalFinanceService

    teacher_data = []
    teacher_salary_total = 0
    teacher_is_locked = False

    if is_main_teacher:
        salary_data = HistoricalFinanceService.calculate_teacher_salary(teacher, year, month, center)
        teacher_salary_total = salary_data['salary']
        teacher_is_locked = salary_data['is_locked']

        for gcd in salary_data['details']:
            group_obj = Group.objects.filter(id=gcd['group_id']).first()
            if not group_obj:
                continue

            enrollments = []
            for en in gcd.get('enrollments', []):
                enrollments.append({
                    "student_name": en.get('student_name', "Noma'lum"),
                    "kurs_narhi": en.get('kurs_narhi', 0),
                    "foiz": en.get('foiz', 0),
                    "attended": en.get('attended', 0),
                    "daromad": en.get('daromad', 0),
                })

            teacher_data.append({
                "group": group_obj,
                "enrollments": enrollments,
                "foiz": gcd.get('fi', getattr(teacher, 'oqituvchi_foizi', 0) or group_obj.oqituvchi_foiz),
                "daromad": gcd['salary'],
                "students_count": len(enrollments),
                "is_support": False,
            })

    # ── Support sifatida ishlash (agar feature yoqilgan bo'lsa) ──
    support_salary_total = 0
    if is_support_member:
        sup = calculate_support_salary(teacher, year, month, center)
        support_salary_total = sup['salary']

        for gcd in sup['details']:
            group_obj = Group.objects.filter(id=gcd['group_id']).first()
            if not group_obj:
                continue

            enrollments = [
                {
                    "student_name": s.get('student_name', "Noma'lum"),
                    "kurs_narhi": int(getattr(group_obj, 'kurs_narxi', 0) or 0),
                    "foiz": gcd.get('fi', 0),
                    "attended": s.get('attended', 0),
                    "daromad": s.get('daromad', 0),
                }
                for s in gcd.get('students', [])
            ]
            teacher_data.append({
                "group": group_obj,
                "enrollments": enrollments,
                "foiz": gcd.get('fi', 0),
                "daromad": gcd['salary'],
                "students_count": len(enrollments),
                "is_support": True,
            })

    jami_umumiy_daromad = teacher_salary_total + support_salary_total

    return render(request, "education/teacher_groups.html", {
        "teacher": teacher,
        "teacher_data": teacher_data,
        "year": year,
        "month": month,
        "years": years,
        "jami_umumiy_daromad": jami_umumiy_daromad,
        "is_locked": teacher_is_locked,
        # Support meta
        "is_main_teacher": is_main_teacher,
        "is_support_member": is_support_member,
        "teacher_salary_total": teacher_salary_total,
        "support_salary_total": support_salary_total,
    })


@login_required
@require_feature("finance")
def teacher_salary_report(request, group_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, id=group_id)
    
    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)
    
    from education.services.historical_finance_service import HistoricalFinanceService
    salary_data = HistoricalFinanceService.calculate_teacher_salary(group.oqituvchi, year, month, center)
    
    student_summaries = []
    teacher_total_income = 0
    
    for gcd in salary_data['details']:
        if gcd['group_id'] == group.id:
            teacher_total_income = gcd['salary']
            for en in gcd.get('enrollments', []):
                student_summaries.append({
                    "student_name": en.get('student_name', 'Noma\'lum'),
                    "attended": en.get('attended', 0),
                    "teacher_income": en.get('daromad', 0)
                })
            break

    ctx = {
        "group": group,
        "student_summaries": student_summaries,
        "teacher_total_income": teacher_total_income,
        "month": month,
        "year": year,
        "is_locked": salary_data['is_locked'],
    }
    return render(request, "education/teacher_salary_report.html", ctx)


# 📊 DIREKTOR HISOBOT PANELI






@login_required
@require_feature("finance")
def teacher_salary_summary(request):
    """
    O'qituvchilar maoshi va markaz foydasini yil/oy bo'yicha hisoblaydi.
    - Attendance: present=True YOKI forced=True bo'lgan barcha darslar hisobga olinadi.
    """

    # ================================
    # Tanlangan yil / oy
    # ================================
    today = date.today()
    selected_year = int(request.GET.get("year") or today.year)
    selected_month = int(request.GET.get("month") or today.month)

    # Oylar ro'yxati
    months = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
    ]
    chart_labels = [m[1] for m in months]

    # ================================
    # 2) O'qituvchilar va ularning hisob-kitobi (Yagona To'g'ri Manba)
    # ================================
    from core.tenant import get_request_center
    from core.perf_cache import (
        TTL_LONG, perf_cache_get_or_set, versioned_cache_key,
    )
    from education.services.support_teacher import (
        is_support_enabled,
        list_support_user_ids,
        calculate_support_salary,
    )
    center = get_request_center(request)
    support_feature_on = is_support_enabled(center)

    # ── PERF: Og'ir hisoblash 15 daqiqa cache (per markaz + yil + oy).
    # Cache key versiyali — attendance/payment o'zgarsa, invalidate orqali
    # bekor qilinadi (signals'da qo'shilishi mumkin).
    _cache_key = versioned_cache_key(
        "salary_sum", getattr(center, 'id', None), selected_year, selected_month
    )

    def _compute_salary_summary():
        return _compute_teacher_salary_summary_payload(
            request, center, support_feature_on, selected_year, selected_month,
            list_support_user_ids, calculate_support_salary,
        )

    payload = perf_cache_get_or_set(_cache_key, _compute_salary_summary, ttl=TTL_LONG)
    teacher_data = payload['teacher_data']
    chart_teacher_income = payload['chart_teacher_income']
    chart_center_income = payload['chart_center_income']
    chart_total_turnover = payload['chart_total_turnover']

    # ── /PERF cache ──

    # AJAX javob — keshlangan ma'lumotdan
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "year": int(selected_year),
            "month": int(selected_month),
            "teacher_data": teacher_data,
            "chart_labels": chart_labels,
            "chart_teacher_income": [float(x) for x in chart_teacher_income],
            "chart_center_income": [float(x) for x in chart_center_income],
            "chart_total_turnover": [float(x) for x in chart_total_turnover],
        })

    return render(request, "education/teacher_salary_summary.html", {
        "years": list(range(2024, 2036)),
        "months": months,
        "selected_year": int(selected_year),
        "selected_month": int(selected_month),
        "teacher_data": teacher_data,
        "chart_labels": chart_labels,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
        "teacher_data_json": json.dumps(teacher_data),
        "chart_labels_json": json.dumps(chart_labels),
        "chart_teacher_income_json": json.dumps([float(x) for x in chart_teacher_income]),
        "chart_center_income_json": json.dumps([float(x) for x in chart_center_income]),
        "chart_total_turnover_json": json.dumps([float(x) for x in chart_total_turnover]),
    })


def _compute_teacher_salary_summary_payload(
    request, center, support_feature_on, selected_year, selected_month,
    list_support_user_ids, calculate_support_salary,
):
    """teacher_salary_summary'ning og'ir hisoblash qismi — cache'lanadi."""
    # Asosiy o'qituvchilar
    teacher_user_qs = User.objects.filter(role="teacher", is_archived=False)
    if center:
        teacher_user_qs = teacher_user_qs.filter(center=center)
    teacher_ids = set(teacher_user_qs.values_list("id", flat=True))

    # Support sifatida biriktirilgan boshqa rolli xodimlar
    extra_user_ids = set()
    if support_feature_on:
        extra_user_ids = list_support_user_ids(center=center) - teacher_ids

    all_ids = teacher_ids | extra_user_ids
    users_qs = User.objects.filter(id__in=all_ids, is_archived=False).order_by("ism", "familya", "id")
    support_user_ids_set = list_support_user_ids(center=center) if support_feature_on else set()

    # ── Batch-compute closed months once for all teachers ──────────
    _fin_months_qs = FinancialMonth.objects.filter(year=selected_year, is_closed=True)
    if center:
        _fin_months_qs = _fin_months_qs.filter(center=center)
    _closed_months_cache = {}
    for _fm in _fin_months_qs:
        _closed_months_cache[_fm.month] = _fm

    # ── Batch-fetch all salary snapshots for all teachers at once ──
    _snapshots_by_teacher = {}
    if _closed_months_cache and all_ids:
        for _snap in TeacherSalarySnapshot.objects.filter(
            teacher_id__in=all_ids,
            financial_month__in=_closed_months_cache.values(),
        ).select_related("financial_month"):
            _snapshots_by_teacher.setdefault(_snap.teacher_id, []).append(_snap)

    # ── N+1 yo'q qilish: har teacher uchun group_set.count() o'rniga
    # bir martalik aggregate query bilan barcha teacher'larning guruh sonini olamiz.
    from django.db.models import Count, Q as _Q
    main_groups_count_map = dict(
        Group.objects.filter(oqituvchi_id__in=teacher_ids, is_archived=False)
        .values('oqituvchi_id')
        .annotate(c=Count('id'))
        .values_list('oqituvchi_id', 'c')
    )
    support_groups_count_map = {}
    if support_feature_on and support_user_ids_set:
        support_groups_count_map = dict(
            Group.objects.filter(
                support_teacher_id__in=support_user_ids_set,
                is_archived=False,
                support_foiz__gt=0,
            )
            .values('support_teacher_id')
            .annotate(c=Count('id'))
            .values_list('support_teacher_id', 'c')
        )

    # ── PERF: Batch-fetch all groups/attendance/history for all teachers ──────
    # Har teacher × har ochiq oy uchun alohida _teacher_groups / _attendance_lookup /
    # _history_lookup chaqiruvi o'rniga bir martalik query — N+1 ni yo'q qiladi.
    from django.db.models import Prefetch as _Prefetch
    _all_main_groups: list = []
    _groups_by_teacher: dict = {}
    if teacher_ids:
        _grp_qs = Group.objects.filter(oqituvchi_id__in=teacher_ids, is_archived=False)
        if center:
            _grp_qs = _grp_qs.filter(center=center)
        _all_main_groups = list(
            _grp_qs.prefetch_related(
                _Prefetch(
                    "enrollments",
                    queryset=Enrollment.all_objects.select_related("student"),
                    to_attr="all_enrollments",
                )
            )
        )
        for _g in _all_main_groups:
            _groups_by_teacher.setdefault(_g.oqituvchi_id, []).append(_g)

    _all_main_group_ids = [_g.id for _g in _all_main_groups]

    _yearly_att: dict = {}
    if _all_main_group_ids:
        for _row in Attendance.objects.filter(
            group_id__in=_all_main_group_ids,
            date__year=selected_year,
        ).filter(HistoricalFinanceService._billable_attendance_filter()).values(
            "group_id", "student_id", "date__month", "date__day"
        ):
            _gid, _sid, _m, _d = (
                _row["group_id"], _row["student_id"], _row["date__month"], _row["date__day"]
            )
            _yearly_att.setdefault(_gid, {}).setdefault(_m, {}).setdefault(_sid, []).append(_d)

    _all_main_history = (
        HistoricalFinanceService._history_lookup(_all_main_group_ids)
        if _all_main_group_ids else {}
    )

    # ================================
    # Grafik uchun bo'sh massivlar (12 oy)
    # ================================
    chart_teacher_income = [0] * 12
    chart_center_income = [0] * 12
    chart_total_turnover = [0] * 12

    # ================================
    # 3) HISOB-KITOB (HistoricalFinanceService + support)
    # ================================
    teacher_data = []

    for teacher in users_qs:
        salary_main_year = 0
        lessons_main = 0
        profit_main = 0
        turnover_main = 0
        salary_support_year = 0
        lessons_support = 0

        # Asosiy o'qituvchi sifatida
        if teacher.id in teacher_ids:
            yearly_stats = HistoricalFinanceService.get_yearly_teacher_stats(
                teacher, selected_year, center,
                _closed_months=_closed_months_cache,
                _snapshots=_snapshots_by_teacher.get(teacher.id, []),
                _teacher_groups=_groups_by_teacher.get(teacher.id, []),
                _yearly_att=_yearly_att,
                _history=_all_main_history,
            )
            for m in range(12):
                chart_teacher_income[m] += yearly_stats[m]['salary']
                chart_center_income[m] += yearly_stats[m]['center_profit']
                chart_total_turnover[m] += yearly_stats[m]['turnover']
            m_stat = yearly_stats[selected_month - 1]
            salary_main_year = m_stat['salary']
            lessons_main = m_stat['lessons']
            profit_main = m_stat['center_profit']
            turnover_main = m_stat['turnover']

        # Support sifatida (faqat tanlangan oy uchun)
        is_support = False
        if support_feature_on and teacher.id in support_user_ids_set:
            sup = calculate_support_salary(teacher, selected_year, selected_month, center)
            salary_support_year = sup['salary']
            lessons_support = sup['attendance_count']
            is_support = True
            # Grafikga support ulushini ham asosiy oylik (teacher_income) qatoriga qo'shamiz
            # — chunki bu jami xodim daromadlari bo'yicha umumiy ko'rsatkich.
            chart_teacher_income[selected_month - 1] += salary_support_year

        # N+1 fix: aggregate'dan o'qiymiz, har teacher uchun query yo'q
        groups_count = main_groups_count_map.get(teacher.id, 0)
        if is_support:
            groups_count += support_groups_count_map.get(teacher.id, 0)

        combined = salary_main_year + salary_support_year
        if combined == 0 and teacher.id not in teacher_ids and not is_support:
            continue

        teacher_data.append({
            "id": teacher.id,
            "teacher": teacher.get_full_name() or teacher.email,
            "groups": groups_count,
            "lessons": lessons_main + lessons_support,
            "teacher_income": int(combined),
            "center_profit": int(profit_main),
            "total_turnover": int(turnover_main),
            "is_support": bool(is_support),
            "is_main_teacher": teacher.id in teacher_ids,
            "support_income": int(salary_support_year),
        })

    # PERF: keshga saqlash uchun qaytaramiz (asosiy view o'qib render qiladi)
    return {
        "teacher_data": teacher_data,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
    }



@login_required
def force_absent_attendance(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    group_id = request.POST.get("group_id")
    date_str = request.POST.get("date")

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, id=group_id)

    date = parse_date(date_str)

    enrollments = Enrollment.objects.filter(group=group, is_active=True)
    forced_count = 0

    for enr in enrollments:
        att, created = Attendance.objects.get_or_create(
            group=group,
            student=enr.student,
            date=date,
            defaults={"present": False}
        )

        # kelgan bo'lsa — forced qilmaymiz
        if att.present:
            continue

        # forced=True qilamiz
        if not att.forced:
            att.forced = True
            att.save()
            forced_count += 1

    return JsonResponse({"ok": True, "count": forced_count})


# ================================
#   RENDER HELPER
# ================================
def _render_salary(request, selected_year, selected_month,
                   teacher_data, chart_labels,
                   chart_teacher_income, chart_center_income, chart_total_turnover):

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "year": selected_year,
            "month": selected_month,
            "teacher_data": teacher_data,
            "chart_teacher_income": chart_teacher_income,
            "chart_center_income": chart_center_income,
            "chart_total_turnover": chart_total_turnover,
        })

    return render(request, "education/teacher_salary_summary.html", {
        "years": list(range(2024, 2036)),
        "months": [
            (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
            (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
            (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
        ],
        "selected_year": selected_year,
        "selected_month": selected_month,
        "teacher_data": teacher_data,
        "chart_labels": chart_labels,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
    })

@login_required
@require_feature("finance")
def teacher_salary_redirect(request):
    group = None

    # O'qituvchi bo'lsa — o'z guruhini topadi
    if request.user.role == "teacher":
        group = Group.objects.filter(oqituvchi=request.user).first()

    # Direktor yoki superuser bo'lsa — birinchi mavjud guruhni topadi
    elif request.user.role == "director" or request.user.is_superuser:
        group = Group.objects.first()

    # Agar topilmasa — xabar chiqar va qaytar
    if not group:
        messages.warning(request, "Hech qanday guruh topilmadi!")
        return redirect("education:groups_it")

    # Topilgan guruh bo'yicha maosh sahifasiga yo'naltirish
    return redirect("education:teacher_salary_report", group.id)



# ---------- CRUD ----------
@login_required
def group_create(request, category=None):
    if not _can_manage(request.user):
        messages.error(request, "Sizda guruh yaratish huquqi yo'q.")
        return redirect("education:groups_home")

    if category == Group.LANG:
        FormCls, title = LangGroupForm, "Tillar bo'yicha guruh yaratish"
    elif category == Group.IT:
        FormCls, title = ITGroupForm, "IT bo'yicha guruh yaratish"
    else:
        FormCls, title = GroupForm, "Guruh yaratish"

    from core.tenant import get_request_center
    center = get_request_center(request) or getattr(request.user, "center", None)
    from billing.services import center_has_feature
    has_manual_oy_dars_soni = center_has_feature(center, "manual_oy_dars_soni") if center else False

    form = FormCls(request.POST or None, center=center)

    if request.method == "POST" and form.is_valid():
        g = form.save(commit=False)
        schedule_mode = form.cleaned_data.get("schedule_mode", "")
        custom_days = form.cleaned_data.get("custom_days") or []

        if schedule_mode in {"odd", "even", "custom"}:
            day_count = len(custom_days) if schedule_mode == "custom" else 3
            g.lessons_per_week = day_count
            if has_manual_oy_dars_soni:
                if not g.oy_dars_soni:
                    g.oy_dars_soni = day_count * 4
            else:
                g.oy_dars_soni = 12

        # 🔹 Kategoriya bo'sh bo'lsa, avtomatik to'ldir
        g.category = category or Group.LANG

        # 🔹 Center avtomatik foydalanuvchidan
        if not g.center_id:
            if center:
                g.center = center
            elif hasattr(request.user, "center") and request.user.center:
                g.center = request.user.center
            else:
                from accounts.models import Center
                g.center = Center.objects.first()

        # Agar narx kiritilmagan bo'lsa, 0 saqlaymiz (500k avtomatik qo'shmaymiz)
        if g.kurs_narxi is None:
            g.kurs_narxi = 0

        # ✅ O'qituvchi foizi
        if g.oqituvchi and getattr(g.oqituvchi, 'oqituvchi_foizi', None) is not None:
            g.oqituvchi_foiz = g.oqituvchi.oqituvchi_foizi
        elif not g.oqituvchi_foiz:
            g.oqituvchi_foiz = 40

        if not g.oy_dars_soni:
            g.oy_dars_soni = 12

        from education.services.group_schedule_service import (
            apply_group_duration_defaults,
            sync_simple_group_schedule,
        )
        apply_group_duration_defaults(g)
        g.save()
        sync_simple_group_schedule(
            group=g,
            schedule_mode=schedule_mode,
            custom_days=custom_days,
            start_time=form.cleaned_data.get("schedule_start_time"),
            end_time=form.cleaned_data.get("schedule_end_time"),
            room=form.cleaned_data.get("schedule_room"),
        )
        messages.success(request, f"✅ {g.nom} guruhi muvaffaqiyatli yaratildi.")
        return redirect("education:group_detail", pk=g.pk)

    elif request.method == "POST":
        print("❌ Forma xato:", form.errors)

    course_templates = CourseTemplate.objects.filter(center=center, is_active=True).order_by("name") if center else []
    return render(request, "education/group_form.html", {
        "form": form, "title": title, "course_templates": course_templates,
        "has_manual_oy_dars_soni": has_manual_oy_dars_soni,
    })


@login_required
def group_edit(request, pk):
    if not request.user.is_superuser and request.user.role not in ["director", "manager", "teacher"]:
        messages.error(request, "Sizda ruxsat yo'q.")
        return redirect("education:groups")

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    # Eski qiymatlarni forma o'zgartirmasdan oldin saqlab qolamiz
    old_foiz = g.oqituvchi_foiz
    old_narx = g.kurs_narxi
    old_oqituvchi_id = g.oqituvchi_id

    from billing.services import center_has_feature
    has_manual_oy_dars_soni = center_has_feature(center, "manual_oy_dars_soni") if center else False

    form = GroupForm(request.POST or None, instance=g, center=center)

    if request.method == "POST" and form.is_valid():
        old_oy_dars_soni = g.oy_dars_soni or 12
        updated_group = form.save(commit=False)
        schedule_mode = form.cleaned_data.get("schedule_mode", "")
        custom_days = form.cleaned_data.get("custom_days") or []

        if schedule_mode in {"odd", "even", "custom"}:
            day_count = len(custom_days) if schedule_mode == "custom" else 3
            updated_group.lessons_per_week = day_count
            if has_manual_oy_dars_soni:
                if not updated_group.oy_dars_soni:
                    updated_group.oy_dars_soni = day_count * 4
            else:
                updated_group.oy_dars_soni = 12

        # Har holda bo'sh qolmasin
        if not updated_group.oy_dars_soni:
            updated_group.oy_dars_soni = 12

        # Agar o'qituvchi o'zgargan bo'lsa, mos foizni avtomatik olamiz
        if updated_group.oqituvchi and updated_group.oqituvchi_id != old_oqituvchi_id:
            teacher_foiz = getattr(updated_group.oqituvchi, 'oqituvchi_foizi', None)
            if teacher_foiz is not None:
                updated_group.oqituvchi_foiz = teacher_foiz

        from education.services.group_schedule_service import (
            apply_group_duration_defaults,
            sync_simple_group_schedule,
        )
        apply_group_duration_defaults(updated_group)
        updated_group.save()
        sync_simple_group_schedule(
            group=updated_group,
            schedule_mode=schedule_mode,
            custom_days=custom_days,
            start_time=form.cleaned_data.get("schedule_start_time"),
            end_time=form.cleaned_data.get("schedule_end_time"),
            room=form.cleaned_data.get("schedule_room"),
        )

        new_oy_dars_soni = updated_group.oy_dars_soni or 12
        from education.models import Enrollment, StudentGroupHistory
        from django.db.models import Q

        # Agar guruhning foizi yoki narxi o'zgargan bo'lsa, joriy o'quvchilarga ham ta'sir qilsin
        if updated_group.oqituvchi_foiz != old_foiz or updated_group.kurs_narxi != old_narx:
            enrollments = Enrollment.objects.filter(group=updated_group)

            if updated_group.oqituvchi_foiz != old_foiz:
                enrollments.update(oqituvchi_foiz=updated_group.oqituvchi_foiz)
                StudentGroupHistory.objects.filter(
                    group=updated_group,
                    end_date__isnull=True
                ).update(oqituvchi_foiz=updated_group.oqituvchi_foiz)

            if updated_group.kurs_narxi != old_narx:
                affected_enrollments = enrollments.filter(Q(kurs_narhi=0) | Q(kurs_narhi=old_narx))
                for enr in affected_enrollments:
                    enr.kurs_narhi = updated_group.kurs_narxi
                    enr.save(update_fields=["kurs_narhi"])
                    StudentGroupHistory.objects.filter(
                        student=enr.student,
                        group=updated_group,
                        end_date__isnull=True
                    ).update(kurs_narxi=updated_group.kurs_narxi)
                    sync_tuition_fee(enr, new_fee=updated_group.kurs_narxi)

        # Barcha faol enrollmentlarning monthly_lessons va lesson_pattern ni guruh bilan sinxronlaymiz
        # va joriy oy TuitionMonth feesini qayta hisoblaymiz.
        from education.services.tuition import ensure_tuition_month
        from education.models import GroupSchedule
        today = timezone.localdate()
        cur_month = today.replace(day=1)
        active_enrollments = list(
            Enrollment.objects.filter(group=updated_group, is_active=True)
            .select_related("group", "student")
        )
        for enr in active_enrollments:
            update_fields = ["monthly_lessons"]
            enr.monthly_lessons = new_oy_dars_soni
            # oy_dars_soni belgilangan bo'lsa, lesson_pattern "group" bo'lishi shart
            # (GroupSchedule bo'lmasa ham — calendar proportion ishlatiladi)
            if new_oy_dars_soni > 0 and enr.lesson_pattern in ("odd", "even", "daily", None, ""):
                enr.lesson_pattern = Enrollment.LESSON_PATTERN_GROUP
                update_fields.append("lesson_pattern")
            enr.save(update_fields=update_fields)
            try:
                ensure_tuition_month(enr, cur_month)
            except Exception:
                pass

        messages.success(request, "✅ Guruh yangilandi.")
        return redirect("education:group_detail", pk=g.id)

    return render(request, "education/group_form.html", {
        "form": form,
        "title": "Guruhni tahrirlash",
        "description": "Guruh ma'lumotlarini tahrirlash",
        "group": g,
        "has_manual_oy_dars_soni": has_manual_oy_dars_soni,
    })


@login_required
@require_POST
def group_generate_month_debt(request, pk):
    """Guruhning barcha aktiv o'quvchilari uchun joriy oy TuitionMonth ni yaratadi.

    Faqat director/manager/superuser ishlatishi mumkin.
    Guruh oy o'rtasidan ochilgan va o'quvchilarda qarz yozilmagan holatlar uchun.
    """
    if not (request.user.is_superuser or getattr(request.user, "role", None) in ("director", "manager")):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("education:group_edit", pk=pk)

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    today = timezone.localdate()
    cur_month = month_first_day(today)

    enrollments = (
        Enrollment.objects
        .filter(group=g, is_active=True, is_deleted=False)
        .select_related("student")
    )
    created_count = 0
    updated_count = 0
    for enr in enrollments:
        try:
            fee_field = tuition_month_fee_field()
            existing = TuitionMonth.all_objects.filter(
                enrollment=enr, month=cur_month, is_deleted=False
            ).first()
            old_fee = int(getattr(existing, fee_field, 0) or 0) if existing else None
            tm = ensure_tuition_month(enr, cur_month)
            new_fee = int(getattr(tm, fee_field, 0) or 0)
            if old_fee is None:
                created_count += 1
            elif old_fee != new_fee:
                updated_count += 1
        except Exception:
            pass

    month_label = today.strftime("%Y-%B").replace(
        today.strftime("%B"),
        ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
         "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"][today.month - 1]
    )
    messages.success(
        request,
        f"✅ {g.nom} guruhi uchun {month_label} oyi qarzlari yozildi: "
        f"{created_count} yangi, {updated_count} yangilangan."
    )
    return redirect("education:group_edit", pk=pk)




@login_required
def group_list(request):
    """
    Barcha guruhlar ro'yxati.
    """
    rows = (
        Group.objects
        .select_related("center", "oqituvchi", "category_obj")
        .annotate(
            student_count=Count("enrollments", filter=Q(enrollments__is_active=True, enrollments__is_deleted=False)),
            sana=Coalesce(F("course_start_date"), Cast(F("tuzilgan"), models.DateField()))
        )
        .order_by("-id")
    )
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        rows = rows.filter(center=center)

    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    page_num = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', '10')

    if page_size == "all":
        paginator = Paginator(rows, max(1, min(rows.count(), 200)))
    else:
        try:
            page_size = int(page_size)
            if page_size < 1 or page_size > 200:
                page_size = 10
        except ValueError:
            page_size = 10
        paginator = Paginator(rows, page_size)

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    can_manage = request.user.is_superuser or request.user.role in ["director", "manager", "teacher"]

    context = {
        "rows": page_obj.object_list,
        "page_obj": page_obj,
        "page_size": page_size,
        "can_manage": can_manage,
    }
    return render(request, "education/groups.html", context)

def get_group_price(request, pk):
    try:
        qs = Group.objects.all()
        from core.tenant import get_request_center
        center = get_request_center(request)
        if center:
            qs = qs.filter(center=center)
        group = qs.get(pk=pk)
        return JsonResponse({
            "price": group.kurs_narxi,
            "oqituvchi_foiz": group.oqituvchi_foiz,
            "monthly_lessons": group.oy_dars_soni,
            "group_name": group.nom,
            "course_id": getattr(group, "category_obj_id", None),
            "course_name": getattr(getattr(group, "category_obj", None), "name", ""),
        })
    except Group.DoesNotExist:
        return JsonResponse({
            "price": 0,
            "oqituvchi_foiz": 40,
            "monthly_lessons": 12,
            "group_name": "",
            "course_id": None,
            "course_name": "",
        })



@login_required
def group_add(request):
    if not _can_manage(request.user):
        messages.error(request, "Sizda ruxsat yo'q.")
        return redirect("education:groups")

    from core.tenant import get_request_center
    center = get_request_center(request) or getattr(request.user, "center", None)
    from billing.services import center_has_feature
    has_manual_oy_dars_soni = center_has_feature(center, "manual_oy_dars_soni") if center else False

    form = GroupForm(request.POST or None, center=center)
    if request.method == "POST" and form.is_valid():
        group = form.save(commit=False)
        if not group.center_id:
            group.center = center
        schedule_mode = form.cleaned_data.get("schedule_mode", "")
        custom_days = form.cleaned_data.get("custom_days") or []
        if schedule_mode in {"odd", "even", "custom"}:
            day_count = len(custom_days) if schedule_mode == "custom" else 3
            group.lessons_per_week = day_count
            if has_manual_oy_dars_soni:
                if not group.oy_dars_soni:
                    group.oy_dars_soni = day_count * 4
            else:
                group.oy_dars_soni = 12
        if not group.oy_dars_soni:
            group.oy_dars_soni = 12

        # O'qituvchi tanlanganda foiz teacher profilidan olinadi.
        if group.oqituvchi and getattr(group.oqituvchi, "oqituvchi_foizi", None) is not None:
            group.oqituvchi_foiz = group.oqituvchi.oqituvchi_foizi
        elif not group.oqituvchi_foiz:
            group.oqituvchi_foiz = 40

        from education.services.group_schedule_service import (
            apply_group_duration_defaults,
            sync_simple_group_schedule,
        )
        apply_group_duration_defaults(group)
        group.save()
        sync_simple_group_schedule(
            group=group,
            schedule_mode=schedule_mode,
            custom_days=custom_days,
            start_time=form.cleaned_data.get("schedule_start_time"),
            end_time=form.cleaned_data.get("schedule_end_time"),
            room=form.cleaned_data.get("schedule_room"),
        )
        messages.success(request, "✅ Guruh muvaffaqiyatli qo'shildi.")
        return redirect("education:groups")

    return render(request, "education/group_form.html", {
        "form": form,
        "title": "Yangi guruh qo'shish",
        "has_manual_oy_dars_soni": has_manual_oy_dars_soni,
    })




from django.contrib import messages

@login_required
def group_delete(request, pk):
    """
    Guruhni o'chirish — tasdiq bilan.
    """
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=pk)

    if request.method == "POST":
        category = getattr(group, "category_obj", None)
        group.delete()
        messages.success(request, "🗑️ Guruh o'chirildi.")

        if category:
            return redirect("education:category_detail", category_id=category.id)
        return redirect("education:groups")

    return render(request, "education/group_delete_confirm.html", {"group": group})


@login_required
@transaction.atomic
def add_student_to_group(request, pk: int):
    from core.tenant import get_request_center
    center = get_request_center(request)
    
    # Guruhni olish
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    g = get_object_or_404(qs, pk=pk)

    # Ruxsat tekshirish
    can_add = False
    if request.user.role == "director" or request.user.is_superuser:
        can_add = True
    elif center:
        if request.user.role == "manager":
            can_add = center.manager_can_add_student
        elif request.user.role == "teacher":
            can_add = center.teacher_can_add_student

    if not can_add:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"error": "Sizda ruxsat yo'q"}, status=403)
        return HttpResponseForbidden("❌ Sizda bu amalni bajarish uchun ruxsat yo'q.")

    # Markazni aniqlash (Guruh markazi asosiy hisoblanadi)
    target_center = g.center
    # AJAX Search
    if request.method == "GET" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        query = request.GET.get("q", "").strip()
        student_qs = User.objects.filter(role="student", center=target_center)
        
        if query:
            # Ism, familya yoki telefon bo'yicha qidirish
            student_qs = student_qs.filter(
                Q(ism__icontains=query) | 
                Q(familya__icontains=query) | 
                Q(telefon1__icontains=query) |
                Q(telefon2__icontains=query)
            ).distinct()
            limit = 15
        else:
            # Bo'sh bo'lsa barcha o'quvchilar (yoki dastlabki 30 tasi)
            limit = 30
            
        results = []
        for s in student_qs.order_by('ism', 'familya')[:limit]:
            # Status aniqlash (Faqat shu markaz doirasida)
            is_in_current = Enrollment.objects.filter(group=g, student=s).exists()
            is_in_other = Enrollment.objects.filter(student=s, center=target_center).exclude(group=g).exists()
            
            results.append({
                "id": s.id,
                "full_name": s.get_full_name(),
                "phone": s.telefon1 or s.telefon2 or "Telefon kiritilmagan",
                "is_in_current": is_in_current,
                "is_in_other": is_in_other,
            })
        
        return JsonResponse({"results": results})

    # AJAX POST (Add student)
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        import json
        try:
            data = json.loads(request.body)
            student_id = data.get("student_id")
            start_date_raw = data.get("start_date")
            lesson_pattern_raw = data.get("lesson_pattern")
        except:
            student_id = request.POST.get("student_id")
            start_date_raw = request.POST.get("start_date")
            lesson_pattern_raw = request.POST.get("lesson_pattern")

        if not student_id:
            return JsonResponse({"error": "O'quvchi tanlanmagan"}, status=400)

        schedule_meta = resolve_lesson_schedule(
            parse_date(start_date_raw or "") or timezone.localdate(),
            lesson_pattern_raw,
        )
        start_date = schedule_meta["start_date"]
        lesson_pattern = schedule_meta["lesson_pattern"]

        student = get_object_or_404(User, pk=student_id, role="student", center=target_center)

        # Allaqachon guruhda bormi?
        if Enrollment.objects.filter(group=g, student=student).exists():
            return JsonResponse({
                "status": "warning",
                "message": f"'{student.get_full_name()}' allaqachon '{g.nom}' guruhida bor."
            })

        # Guruhga biriktirilgan narxni doimo ishlatamiz
        kurs_narhi = g.kurs_narxi

        # Qo'shish (EnrollmentService orqali tarix bilan)
        from education.services.enrollment_service import EnrollmentService
        enr = EnrollmentService.enroll_student(
            student=student,
            group=g,
            kurs_narxi=kurs_narhi,
            oqituvchi_foiz=g.oqituvchi_foiz or 40,
            start_date=start_date,
            lesson_pattern=lesson_pattern,
            monthly_lessons=getattr(g, "oy_dars_soni", 0) or 12,
        )

        from education.services.tuition import ensure_tuition_month
        preview_month = _preview_month_for_start_date(start_date)
        # ✅ Yangi qo'shilgan o'quvchi uchun boshlanish oyidagi snapshot to'g'ri prorata bilan yaratiladi.
        ensure_tuition_month(enr, preview_month)
        preview = tuition_month_preview(enr, preview_month)

        return JsonResponse({
            "status": "success",
            "message": f"'{student.get_full_name()}' muvaffaqiyatli qo'shildi ✅",
            "preview": _serialize_tuition_preview(preview),
            "student": {
                "id": student.id,
                "full_name": student.get_full_name(),
                "phone": student.telefon1 or student.telefon2
            },
            "info": schedule_meta["adjustment_note"],
        })

    # Standart GET render
    return render(request, "education/add_student_to_group.html", {
        "group": g,
    })



@login_required
def teacher_groups_view(request, teacher_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = User.objects.filter(role="teacher")
    if center:
        qs = qs.filter(center=center)
    teacher = get_object_or_404(qs, id=teacher_id)

    groups = (
        teacher.group_set
        .prefetch_related('enrollments__student')
        .all()
    )

    teacher_data = []
    for group in groups:
        enrollments = group.enrollments.all()
        group_income = sum([enr.real_oqituvchi_daromadi() for enr in enrollments])
        group_info = {
            'name': group.name,
            'students': enrollments,
            'foiz': enrollments.first().oqituvchi_foiz if enrollments.exists() else 0,
            'daromad': group_income,
        }
        teacher_data.append(group_info)

    ctx = {
        "teacher": teacher,
        "teacher_data": teacher_data,
    }
    return render(request, "education/teacher_groups.html", ctx)


@require_POST
def toggle_attendance(request):
    student_id = request.POST.get("student_id")
    group_id = request.POST.get("group_id")
    date_str = request.POST.get("date")

    if not (student_id and group_id):
        return JsonResponse({"error": "Invalid data"}, status=400)

    date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else timezone.localdate()

    att, created = Attendance.objects.get_or_create(
        group_id=group_id,
        student_id=student_id,
        date=date,
        defaults={"teacher": request.user}
    )
    # Validate center
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center and att.group.center_id != center.id:
        return JsonResponse({"error": "Center mismatch"}, status=403)

    # Belgini o'zgartiramiz (agar bor bo'lsa)
    att.present = not att.present
    att.teacher = request.user
    att.save()

    return JsonResponse({
        "success": True,
        "present": att.present,
        "date": att.date.strftime("%Y-%m-%d"),
    })




# ---------- A'zolik va o'qituvchi sahifasi ----------
@login_required
def enrollment_remove(request, pk):
    qs = Enrollment.objects.select_related("group", "student")
    from core.tenant import get_request_center
    center = get_request_center(request)
    if center:
        # Enrollment has center now, or filter by group__center
        qs = qs.filter(group__center=center)
    enr = get_object_or_404(qs, pk=pk)
    
    if not _can_manage(request.user):
        messages.error(request, "Sizda ruxsat yo'q.")
        return redirect("education:group_detail", pk=enr.group_id)
        
    if request.method == "POST":
        # ✅ EnrollmentService orqali o'chiramiz (tarix yopiladi)
        from education.services.enrollment_service import EnrollmentService
        EnrollmentService.remove_student(enr.student, enr.group)
        messages.success(request, "O'quvchi guruhdan chiqarildi. Tarix saqlanib qoldi.")
        
    return redirect("education:group_detail", pk=enr.group_id)


@require_POST
@login_required
def enrollment_toggle_deferred(request, pk):
    if not _can_manage(request.user):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "Ruxsat yo'q."}, status=403)
        messages.error(request, "Ruxsat yo'q.")
        return redirect("education:qarzdorlar_home")
    center = get_active_center(request)
    qs = Enrollment.objects.all()
    if center:
        qs = qs.filter(center=center)
    enr = get_object_or_404(qs, pk=pk)
    enr.is_deferred = not enr.is_deferred
    enr.save(update_fields=["is_deferred"])
    status_label = "kechiktirildi" if enr.is_deferred else "oddiy holatga qaytarildi"
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({
            "success": True,
            "is_deferred": enr.is_deferred,
            "message": f"{enr.student.get_full_name()} to'lovi {status_label}.",
        })
    next_url = request.POST.get("next") or reverse("education:qarzdorlar_home")
    messages.success(request, f"✅ {enr.student.get_full_name()} to'lovi {status_label}.")
    return redirect(next_url)


@login_required
def enrollment_leave(request, pk):
    """
    O'quvchini oyning o'rtasida guruhdan chiqarish — prorata to'lov bilan.
    """
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Enrollment.objects.select_related("group", "student", "group__oqituvchi")
    if center:
        qs = qs.filter(group__center=center)
    enr = get_object_or_404(qs, pk=pk)

    if not _can_manage(request.user):
        messages.error(request, "Sizda ruxsat yo'q.")
        return redirect("education:group_detail", pk=enr.group_id)

    today = timezone.localdate()
    cur_month = today.replace(day=1)

    from education.services.tuition import (
        ensure_tuition_month, get_month_paid, get_effective_month_fee, month_last_day
    )
    tm = ensure_tuition_month(enr, cur_month)
    full_fee = enr.kurs_narhi or 0
    paid_so_far = get_month_paid(enr, cur_month)

    # Attendance-based prorata
    from education.models import Attendance
    group_sessions = Attendance.objects.filter(
        enrollment__group=enr.group,
        date__year=cur_month.year,
        date__month=cur_month.month,
    ).values_list("date", flat=True).distinct()
    total_lessons = group_sessions.count()

    student_attended = Attendance.objects.filter(
        enrollment=enr,
        date__year=cur_month.year,
        date__month=cur_month.month,
        is_present=True,
    ).count()

    if total_lessons > 0:
        prorated_fee = round(full_fee * student_attended / total_lessons)
    else:
        prorated_fee = 0

    remaining = max(0, prorated_fee - paid_so_far)
    oqituvchi_foiz = enr.oqituvchi_foiz or 0

    if request.method == "POST":
        amount_raw = request.POST.get("amount", "0").replace(" ", "").replace(",", "")
        try:
            amount = int(amount_raw)
        except (ValueError, TypeError):
            amount = 0

        if amount > 0:
            tm.fee_amount = prorated_fee
            tm.save(update_fields=["fee_amount"])
            from education.services.tuition import create_payment_and_allocate
            create_payment_and_allocate(
                enrollment=enr,
                amount=amount,
                paid_date=today,
                start_month=cur_month,
                payment_type="cash",
            )

        from education.services.enrollment_service import EnrollmentService
        EnrollmentService.remove_student(enr.student, enr.group)
        messages.success(request, f"✅ {enr.student.get_full_name()} guruhdan chiqarildi.")
        return redirect("education:group_detail", pk=enr.group_id)

    context = {
        "enr": enr,
        "full_fee": full_fee,
        "paid_so_far": paid_so_far,
        "prorated_fee": prorated_fee,
        "remaining": remaining,
        "total_lessons": total_lessons,
        "student_attended": student_attended,
        "oqituvchi_foiz": oqituvchi_foiz,
        "teacher_share": round(prorated_fee * oqituvchi_foiz / 100),
        "center_share": round(prorated_fee * (100 - oqituvchi_foiz) / 100),
        "cur_month": cur_month,
    }
    return render(request, "education/enrollment_leave.html", context)


@login_required
def my_groups(request):
    from core.tenant import get_request_center
    center = get_request_center(request)
    # Asosiy o'qituvchi yoki support teacher sifatida biriktirilgan barcha guruhlar
    rows = (
        Group.objects.filter(
            Q(oqituvchi=request.user) | Q(support_teacher=request.user),
            is_archived=False,
        )
        .select_related("center", "oqituvchi", "category_obj")
        .annotate(student_count=Count("enrollments", filter=Q(enrollments__is_active=True, enrollments__is_deleted=False)))
        .distinct()
        .order_by("nom")
    )
    if center:
        rows = rows.filter(center=center)
    return render(request, "education/my_groups.html", {"rows": rows, "is_support": True})


@login_required
@require_feature("finance")
def teacher_income_dashboard(request):
    """
    O'qituvchining shaxsiy daromadlari panelini ko'rsatadi.
    Snapshot tizimi yordamida o'tgan oylar ma'lumotlari muzlatilgan (immutable).
    """
    if request.user.role not in ['teacher', 'director', 'manager'] and not request.user.is_superuser:
        messages.error(request, "Bu bo'lim ushbu rol uchun emas.")
        return redirect('core:home')
        
    is_admin = request.user.role in ['director', 'manager'] or request.user.is_superuser
    
    # Agar admin bo'lsa va teacher_id berilgan bo'lsa - o'shani ko'ramiz
    teacher_id = request.GET.get('teacher_id')
    if is_admin and teacher_id:
        teacher = get_object_or_404(User, id=teacher_id, role='teacher')
    else:
        teacher = request.user

    today = timezone.localdate()
    selected_year = _get_int(request.GET, "year", today.year)
    selected_month = _get_int(request.GET, "month", today.month)
    
    from core.tenant import get_request_center
    from education.services.expected_income_service import calculate_expected_income
    center = get_request_center(request)

    # HistoricalFinanceService orqali ma'lumotlarni olish (snapshot yoki dinamik)
    salary_data = HistoricalFinanceService.calculate_teacher_salary(teacher, selected_year, selected_month, center)

    # Support teacher daromadini ham qo'shamiz (feature flag tekshiruvisiz —
    # biriktirilgan bo'lsa har doim ko'rsatish kerak)
    from education.services.support_teacher import (
        list_support_user_ids, calculate_support_salary, get_yearly_support_salary,
    )
    support_salary_data = None
    support_salary_total = 0
    is_support_member = teacher.id in list_support_user_ids(center=center)
    if is_support_member:
        support_salary_data = calculate_support_salary(teacher, selected_year, selected_month, center)
        support_salary_total = support_salary_data.get("salary", 0)
        salary_data = dict(salary_data)
        salary_data["salary"] = salary_data.get("salary", 0) + support_salary_total
        salary_data["support_details"] = support_salary_data.get("details", [])
        salary_data["support_salary"] = support_salary_total

    # Get all 12 months for the yearly chart efficiently
    monthly_income = HistoricalFinanceService.get_yearly_teacher_salary(teacher, selected_year, center)
    if is_support_member:
        # Support yillik daromadini asosiy daromad bilan qo'shamiz
        support_monthly = get_yearly_support_salary(teacher, selected_year, center)
        monthly_income = [m + s for m, s in zip(monthly_income, support_monthly)]
    total_year_income = sum(monthly_income)

    # Get daily breakdown for the selected month (now returned by the service)
    _, num_days = calendar.monthrange(selected_year, selected_month)
    main_daily = salary_data.get('daily_breakdown', [0] * 31)
    if is_support_member and support_salary_data:
        support_daily = support_salary_data.get('daily_breakdown', [0] * 31)
        combined_daily = [a + b for a, b in zip(main_daily, support_daily)]
    else:
        combined_daily = main_daily
    daily_income = combined_daily[:num_days]

    # Yearly labels for JS
    monthly_labels = ["Yan", "Fev", "Mar", "Apr", "May", "Iyun", "Iyul", "Avg", "Sen", "Okt", "Noy", "Dek"]
    daily_labels = [str(i) for i in range(1, num_days + 1)]

    # Selectors options
    years = range(today.year - 2, today.year + 2)
    months_list = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr")
    ]
    months_dict = dict(months_list)

    # Prognoz: tanlangan oy uchun maksimal kutilgan va kelasi oy uchun
    current_expected = calculate_expected_income(teacher, selected_year, selected_month, center)
    if selected_month == 12:
        next_year, next_month = selected_year + 1, 1
    else:
        next_year, next_month = selected_year, selected_month + 1
    next_expected = calculate_expected_income(teacher, next_year, next_month, center)

    # Progress foiz: bu oy qanchasi yig'ildi (maksimaldan)
    current_max = current_expected.get("expected_income", 0)
    current_salary = salary_data.get("salary", 0)
    progress_pct = min(100, int(current_salary / current_max * 100)) if current_max > 0 else 0

    # Kelasi oy o'zgarish foizi (joriy oyga nisbatan)
    next_income = next_expected.get("expected_income", 0)
    if current_max > 0:
        delta_pct = round((next_income - current_max) / current_max * 100, 1)
    else:
        delta_pct = 0

    # Bar widths uchun max qiymatlar
    next_max_group = max((g['group_total'] for g in next_expected.get('breakdown', []) if g['group_total']), default=1)
    current_max_salary = max((d['salary'] for d in salary_data.get('details', []) if d['salary']), default=1)

    ctx = {
        'teacher': teacher,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'salary_data': salary_data,
        'daily_income': daily_income,
        'monthly_income': monthly_income,
        'total_year_income': total_year_income,
        'daily_labels': daily_labels,
        'monthly_labels': monthly_labels,
        'years': years,
        'months_list': months_list,
        'months_dict': months_dict,
        'is_locked': salary_data.get('is_locked', False),
        'is_admin': is_admin,
        'is_support_member': is_support_member,
        'support_salary_total': support_salary_total,
        # Prognoz
        'current_expected': current_expected,
        'next_expected': next_expected,
        'next_year': next_year,
        'next_month': next_month,
        'next_month_name': months_dict.get(next_month, ""),
        'progress_pct': progress_pct,
        'delta_pct': delta_pct,
        'next_max_group': next_max_group,
        'current_max_salary': current_max_salary,
    }

    if is_admin:
        ctx['teachers_list'] = User.objects.filter(role='teacher', is_active=True)

    return render(request, "education/teacher_income_dashboard.html", ctx)


@login_required
@require_feature("finance")
def close_finance_month_view(request):
    """View to close (lock) or open (unlock) a financial month for a center."""
    if request.user.role not in ['director', 'manager'] and not request.user.is_superuser:
        messages.error(request, "Sizda bu bo'limga ruxsat yo'q.")
        return redirect('education:teacher_income_dashboard')

    if request.method == "POST":
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        action = request.POST.get('action', 'lock')
        
        from core.tenant import get_request_center
        center = get_request_center(request)
        
        if action == 'unlock':
            HistoricalFinanceService.open_month(center, year, month, request.user)
            messages.success(request, f"{year}-yil {month}-oy muvaffaqiyatli ochildi. Endi oyliklar avtomatik (jonli) tarzda hisoblanadi.")
        else:
            HistoricalFinanceService.close_month(center, year, month, request.user)
            messages.success(request, f"{year}-yil {month}-oy muvaffaqiyatli yopildi va qotirildi. Endi o'zgarishlar tasir qilmaydi.")
        
    return redirect(f"{reverse('education:teacher_salary_list')}?year={year}&month={month}")

@login_required
@require_feature("finance")
def fix_all_incomes(request):
    """
    Global/Production muhitda o'tgan oydagi eski Attendance malumotlarini 
    TeacherIncome tizimiga generator qilib beruvchi bir martalik master funktsiya
    """
    from education.models import Attendance, TeacherIncome, Enrollment
    from django.db import transaction
    
    all_attendances = Attendance.objects.all().select_related('group', 'student', 'teacher', 'group__center')
    
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for i, att in enumerate(all_attendances):
            # 1. To'lanadigan holatmi?
            is_billable = att.status == 'present' or att.status == 'absent_unexcused' or getattr(att, 'forced', False) or getattr(att, 'present', False)

            if not is_billable:
                TeacherIncome.objects.filter(attendance=att).delete()
                continue
            
            # 2. Enrollmentni topish
            enrollment = Enrollment.all_objects.filter(
                group=att.group,
                student=att.student
            ).order_by('-is_active', '-created_at').first()

            if not enrollment:
                TeacherIncome.objects.filter(attendance=att).delete()
                continue

            teacher = att.group.oqituvchi if att.group else None
            if not teacher:
                continue

            foiz = getattr(teacher, 'oqituvchi_foizi', 0)
            if foiz is None or foiz == 0:
                foiz = enrollment.oqituvchi_foiz

            kurs_narhi = full_course_amount(enrollment)
            
            oy_dars_soni = att.group.oy_dars_soni or 12
            if oy_dars_soni <= 0: oy_dars_soni = 12

            if kurs_narhi > 0 and foiz > 0:
                total_per_lesson = kurs_narhi / oy_dars_soni
                amount = round(total_per_lesson * (foiz / 100))
                center_amount = round(total_per_lesson * ((100 - foiz) / 100))
                total_amount = round(total_per_lesson)
            else:
                amount = 0
                center_amount = 0
                total_amount = 0

            obj, created = TeacherIncome.objects.update_or_create(
                attendance=att,
                defaults={
                    'center': att.center or (att.group.center if att.group else None),
                    'teacher': teacher,
                    'group': att.group,
                    'amount': amount,
                    'center_amount': center_amount,
                    'total_amount': total_amount
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1

    messages.success(request, f"🚀 Barcha daromadlar qayta hisoblandi! Yangi tizim qo'llandi. Yaratildi: {created_count}, Yangilandi: {updated_count}.")
    return redirect('education:teacher_income_dashboard')


# =============================================================================
# PHASE 1: Exam foundation module (backward-compatible, non-blocking)
# =============================================================================


def _director_or_manager(user):
    return user.is_superuser or getattr(user, "role", None) in ("director", "manager")


def _teacher_can_view_settings(user):
    return user.is_superuser or getattr(user, "role", None) in ("director", "manager", "teacher")


@login_required
def expense_create(request):
    if not _director_or_manager(request.user):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    from core.tenant import get_request_center

    center = get_request_center(request) or getattr(request.user, "center", None)
    if not center:
        raise PermissionDenied("Markaz topilmadi.")

    form = CenterExpenseForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        expense = form.save(commit=False)
        expense.center = center
        expense.created_by = request.user
        expense.save()
        messages.success(request, "Xarajat saqlandi.")
        return redirect("core:financial_dashboard")

    return render(
        request,
        "education/expense_form.html",
        {
            "form": form,
            "center": center,
        },
    )


def _teacher_or_management_can_access_group(user, group: Group):
    if user.is_superuser or getattr(user, "role", None) in ("director", "manager"):
        return True
    if getattr(user, "role", None) == "teacher":
        return group.oqituvchi_id == user.id or group.support_teacher_id == user.id
    return False


def _decode_exam_session_note(raw_text: str) -> dict:
    """
    Backward-compatible parser:
    - yangi format: {"task": "...", "comment": "..."} (JSON)
    - eski format: oddiy text (task sifatida olinadi)
    """
    raw_text = (raw_text or "").strip()
    if not raw_text:
        return {"task": "", "comment": ""}
    try:
        payload = json.loads(raw_text)
        if isinstance(payload, dict):
            return {
                "task": (payload.get("task") or "").strip(),
                "comment": (payload.get("comment") or "").strip(),
            }
    except Exception:
        pass
    return {"task": raw_text, "comment": ""}


def _encode_exam_session_note(task: str, comment: str) -> str:
    task = (task or "").strip()
    comment = (comment or "").strip()
    if not task and not comment:
        return ""
    return json.dumps({"task": task, "comment": comment}, ensure_ascii=False)


def _exam_entry_url(session_id: int, max_score) -> str:
    url = reverse("education:exam_session_entry", kwargs={"session_id": session_id})
    if str(max_score or "100") != "100":
        return f"{url}?{urlencode({'max_score': str(max_score)})}"
    return url


@login_required
@require_feature("imtihon")
def exam_settings_view(request):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if not center:
        raise Http404("Center not found")

    if not _teacher_can_view_settings(request.user):
        return HttpResponseForbidden("Sizda bu bo'limga ruxsat yo'q.")

    from .forms import CenterExamSettingForm
    from education.services.exam_service import get_or_create_center_exam_settings

    settings_obj = get_or_create_center_exam_settings(center)
    can_edit = _director_or_manager(request.user)
    form = CenterExamSettingForm(request.POST or None, instance=settings_obj)

    if request.method == "POST":
        if not can_edit:
            return HttpResponseForbidden("Teacher bu sozlamalarni o'zgartira olmaydi.")
        if form.is_valid():
            obj = form.save(commit=False)
            obj.center = center
            obj.updated_by = request.user
            obj.save()
            from education.services.audit_service import log_education_event
            log_education_event(
                center=center,
                actor=request.user,
                action_type="director_settings_change",
                entity=obj,
                message="Exam settings updated",
            )
            messages.success(request, "Imtihon sozlamalari saqlandi.")
            return redirect("education:exam_settings")

    return render(
        request,
        "education/exam_settings.html",
        {"form": form, "settings_obj": settings_obj, "can_edit": can_edit},
    )


@login_required
@require_POST
@require_feature("imtihon")
def exam_reminder_action(request, group_id: int):
    from core.tenant import get_request_center
    center = get_request_center(request)

    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, pk=group_id)

    if not _teacher_or_management_can_access_group(request.user, group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    from education.models import ExamReminderLog
    from education.services.exam_service import (
        create_or_get_exam_session_from_reminder,
        create_or_update_exam_session_decision,
        get_exam_reminder_state,
        log_exam_reminder_action,
    )

    action = (request.POST.get("action") or "").strip().lower()
    date_str = request.POST.get("date")
    selected_date = parse_date(date_str) if date_str else localdate()
    note = (request.POST.get("note") or "").strip()

    if action not in {ExamReminderLog.ACTION_YES, ExamReminderLog.ACTION_NO, ExamReminderLog.ACTION_LATER}:
        messages.error(request, "Noto'g'ri action.")
        return redirect("education:group_detail", pk=group.id)

    reminder_state = get_exam_reminder_state(group=group, on_date=selected_date)
    target_checkpoint = int(reminder_state.get("target_lesson_number") or 0)

    if action == ExamReminderLog.ACTION_YES:
        if not reminder_state.get("enabled"):
            messages.info(request, "Imtihon tizimi o'chiq. Sozlamani director yoqishi kerak.")
            return redirect("education:group_detail", pk=group.id)
        if not reminder_state.get("due"):
            messages.info(request, "Hozircha majburiy imtihon darsi emas, lekin davomat davom etadi.")
            return redirect("education:group_detail", pk=group.id)
        session = create_or_get_exam_session_from_reminder(
            group=group,
            teacher=request.user,
            attendance_date=selected_date,
            created_by=request.user,
            decision_note=note,
            lesson_number_reference=target_checkpoint or reminder_state.get("target_lesson_number"),
        )
        messages.success(request, "Imtihon sessiyasi ochildi. Natijalarni kiriting.")
        return redirect("education:exam_session_entry", session_id=session.id)

    if not reminder_state.get("enabled"):
        messages.info(request, "Imtihon tizimi o'chiq. Sozlamani director yoqishi kerak.")
        return redirect("education:group_detail", pk=group.id)
    if not reminder_state.get("due"):
        messages.info(request, "Hozircha bu nazorat bosqichi bo'yicha amal talab qilinmaydi.")
        return redirect("education:group_detail", pk=group.id)

    decision_session = None
    if action != ExamReminderLog.ACTION_LATER:
        decision_session = create_or_update_exam_session_decision(
            group=group,
            teacher=request.user,
            attendance_date=selected_date,
            actor=request.user,
            decision=action,
            decision_note=note,
            lesson_number_reference=target_checkpoint,
        )

    log_exam_reminder_action(
        group=group,
        teacher=request.user,
        action=action,
        attendance_date=selected_date,
        note=note,
        metadata={
            "session_id": decision_session.id if decision_session else None,
            "target_checkpoint": target_checkpoint,
        },
    )

    if action == ExamReminderLog.ACTION_NO:
        messages.warning(request, "Imtihon o'tkazilmagan deb qayd etildi.")
    else:
        messages.info(request, "Imtihon eslatmasi keyinroq uchun saqlandi.")
    return redirect("education:group_detail", pk=group.id)


@login_required
@require_feature("imtihon")
def exam_list(request):
    from core.tenant import get_request_center
    from education.models import ExamResult, ExamSession

    try:
        center = get_request_center(request)
        if not center:
            raise Http404("Center not found")
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_EXAM_SESSIONS,
            message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response

        if not _director_or_manager(request.user):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        group_id = _get_int(request.GET, "group", 0)
        teacher_id = _get_int(request.GET, "teacher", 0)
        status_filter = (request.GET.get("status") or "").strip()
        month_filter = (request.GET.get("month") or "").strip()

        sessions_qs = (
            ExamSession.objects.filter(center=center)
            .select_related("group", "teacher")
            .prefetch_related("results")
            .annotate(
                students_count=Count("results", distinct=True),
                passed_count=Count(
                    "results",
                    filter=Q(results__passed=True, results__absent_in_exam=False),
                    distinct=True,
                ),
                failed_count=Count(
                    "results",
                    filter=(
                        Q(results__passed=False)
                        & (
                            Q(results__score__isnull=False)
                            | Q(results__percent__isnull=False)
                            | Q(results__absent_in_exam=True)
                        )
                    ),
                    distinct=True,
                ),
                avg_percent=Avg("results__percent", filter=Q(results__percent__isnull=False)),
            )
            .order_by("-exam_date", "-id")
        )

        if group_id:
            sessions_qs = sessions_qs.filter(group_id=group_id)
        if teacher_id:
            sessions_qs = sessions_qs.filter(teacher_id=teacher_id)
        if status_filter in {
            ExamSession.STATUS_DRAFT,
            ExamSession.STATUS_COMPLETED,
            ExamSession.STATUS_CANCELLED,
        }:
            sessions_qs = sessions_qs.filter(status=status_filter)
        if month_filter:
            parsed_month = parse_month_str(month_filter)
            if parsed_month:
                sessions_qs = sessions_qs.filter(
                    exam_date__year=parsed_month.year,
                    exam_date__month=parsed_month.month,
                )

        examined_results = ExamResult.objects.filter(session__in=sessions_qs).filter(
            Q(absent_in_exam=True) | Q(score__isnull=False) | Q(percent__isnull=False)
        )
        total_examined = examined_results.count()
        passed_examined = examined_results.filter(passed=True, absent_in_exam=False).count()
        avg_pass_rate = round((passed_examined / total_examined * 100), 1) if total_examined else 0

        paginator = Paginator(sessions_qs, 20)
        sessions = paginator.get_page(request.GET.get("page"))

        groups_list = Group.objects.filter(center=center).order_by("nom")
        teachers_list = User.objects.filter(
            center=center,
            role="teacher",
            is_archived=False,
        ).order_by("ism", "familya")

        return render(
            request,
            "education/exam_list.html",
            {
                "sessions": sessions,
                "groups_list": groups_list,
                "teachers_list": teachers_list,
                "filters": {
                    "group": group_id,
                    "teacher": teacher_id,
                    "status": status_filter,
                    "month": month_filter,
                },
                "total_stats": {
                    "total_sessions": sessions_qs.count(),
                    "completed": sessions_qs.filter(status=ExamSession.STATUS_COMPLETED).count(),
                    "avg_pass_rate": avg_pass_rate,
                    "total_students_examined": total_examined,
                },
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("exam_list failed")
        messages.error(request, "Imtihon sessiyalari ro'yxatini yuklashda xatolik yuz berdi.")
        return redirect("core:director_boshqaruv")


@login_required
@require_feature("imtihon")
def exam_create(request, group_id=None):
    from core.tenant import get_request_center
    from education.models import ExamResult, ExamSession
    from education.services.exam_service import (
        get_group_exam_sequence_number,
        get_group_lesson_number,
    )

    try:
        center = get_request_center(request)
        if not center:
            raise Http404("Center not found")
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_EXAM_SESSIONS,
            message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response

        role = getattr(request.user, "role", None)
        if not request.user.is_superuser and role not in ("teacher", "manager", "director"):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        groups_qs = Group.objects.filter(center=center).select_related("oqituvchi").order_by("nom")
        if role == "teacher" and not request.user.is_superuser:
            groups_qs = groups_qs.filter(oqituvchi=request.user)

        selected_group = None
        if group_id is not None:
            selected_group = get_object_or_404(groups_qs, pk=group_id)

        selected_exam_date = request.POST.get("exam_date") or timezone.localdate().isoformat()
        assignment_description = (request.POST.get("assignment_description") or "").strip()

        if request.method == "POST":
            target_group = selected_group
            if target_group is None:
                selected_group_id = _get_int(request.POST, "group", 0)
                if not selected_group_id:
                    messages.error(request, "Guruhni tanlang.")
                    return render(
                        request,
                        "education/exam_create.html",
                        {
                            "groups_list": groups_qs,
                            "selected_group": selected_group,
                            "selected_exam_date": selected_exam_date,
                            "assignment_description": assignment_description,
                        },
                    )
                target_group = get_object_or_404(groups_qs, pk=selected_group_id)

            exam_date = parse_date(selected_exam_date) or timezone.localdate()
            teacher_user = request.user if role == "teacher" else target_group.oqituvchi
            lesson_number_reference = get_group_lesson_number(group=target_group, on_date=exam_date)

            with transaction.atomic():
                session = ExamSession.objects.create(
                    center=center,
                    group=target_group,
                    teacher=teacher_user,
                    attendance_date=exam_date,
                    exam_date=exam_date,
                    lesson_number_reference=lesson_number_reference,
                    exam_sequence_number=get_group_exam_sequence_number(target_group),
                    teacher_decision=ExamSession.DECISION_LATER,
                    decision_note=_encode_exam_session_note(assignment_description, ""),
                    status=ExamSession.STATUS_DRAFT,
                    created_by=request.user,
                    updated_by=request.user,
                )

                students = Enrollment.objects.filter(
                    group=target_group,
                    center=center,
                    is_active=True,
                ).select_related("student")
                ExamResult.objects.bulk_create(
                    [
                        ExamResult(
                            center=center,
                            session=session,
                            group=target_group,
                            student=enrollment.student,
                            teacher=teacher_user,
                            exam_date=session.exam_date,
                            lesson_number_reference=session.lesson_number_reference,
                            assignment_description=assignment_description,
                            created_by=request.user,
                            updated_by=request.user,
                        )
                        for enrollment in students
                    ],
                    ignore_conflicts=True,
                )

            messages.success(request, "Yangi imtihon sessiyasi yaratildi.")
            return redirect("education:exam_session_entry", session_id=session.id)

        return render(
            request,
            "education/exam_create.html",
            {
                "groups_list": groups_qs,
                "selected_group": selected_group,
                "selected_exam_date": selected_exam_date,
                "assignment_description": assignment_description,
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("exam_create failed: group_id=%s", group_id)
        messages.error(request, "Imtihon sessiyasini yaratishda xatolik yuz berdi.")
        return redirect("education:teacher_exam_history")


@login_required
@require_feature("imtihon")
def exam_session_entry(request, session_id: int):
    from core.tenant import get_request_center
    from .forms import ExamResultRowForm
    from education.models import ExamSession, ExamResult
    from education.services.certificate_service import auto_check_certificate_eligibility
    from education.services.exam_service import (
        get_exam_session_progress,
        get_or_create_center_exam_settings,
        notify_exam_results,
        save_exam_session_task_files,
        save_exam_results_batch,
    )

    try:
        center = get_request_center(request)
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_EXAM_SESSIONS,
            message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response
        qs = ExamSession.objects.select_related("group", "teacher", "center").prefetch_related("task_files")
        if center:
            qs = qs.filter(center=center)
        session = get_object_or_404(qs, pk=session_id)

        if not _teacher_or_management_can_access_group(request.user, session.group):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        settings_obj = get_or_create_center_exam_settings(session.center)
        try:
            max_score = Decimal(str(request.GET.get("max_score") or request.POST.get("max_score") or "100"))
            if max_score <= 0:
                raise ValueError
        except Exception:
            max_score = Decimal("100")
        passing_percent = Decimal(str(settings_obj.passing_score_percent or 60))

        enrollments = (
            Enrollment.objects.filter(group=session.group, is_active=True)
            .select_related("student")
            .order_by("student__ism", "student__familya")
        )
        existing_results = {
            result.student_id: result
            for result in ExamResult.objects.filter(session=session).select_related("student")
        }
        for enrollment in enrollments:
            enrollment.existing_result = existing_results.get(enrollment.student_id)

        parsed_note = _decode_exam_session_note(session.decision_note)
        session_task_default = parsed_note["task"]
        session_comment_default = parsed_note["comment"]

        if request.method == "POST":
            action = (request.POST.get("action") or "save").strip().lower()
            is_finalize = action == "finalize"
            was_completed = session.status == ExamSession.STATUS_COMPLETED

            session_task = (request.POST.get("session_task") or "").strip()
            session_comment = (request.POST.get("session_comment") or "").strip()
            session_note = _encode_exam_session_note(session_task, session_comment)
            note_updated = False
            if session.decision_note != session_note:
                session.decision_note = session_note
                session.updated_by = request.user
                session.save(update_fields=["decision_note", "updated_by", "updated_at"])
                note_updated = True

            session_task_default = session_task
            session_comment_default = session_comment

            uploaded_task_file_count = 0
            session_task_files = request.FILES.getlist("session_task_files") or []
            if session_task_files and settings_obj.exam_file_upload_enabled:
                try:
                    uploaded_task_file_count = save_exam_session_task_files(
                        session=session,
                        actor=request.user,
                        files=session_task_files,
                    )
                    if uploaded_task_file_count:
                        messages.success(request, f"{uploaded_task_file_count} ta umumiy task fayli yuklandi.")
                except ValueError as exc:
                    messages.error(request, str(exc))
            elif session_task_files and not settings_obj.exam_file_upload_enabled:
                messages.warning(request, "Task fayl yuklash markaz sozlamasida o'chiq.")

            rows = []
            row_errors = []
            for enrollment in enrollments:
                sid = enrollment.student_id
                work_files = request.FILES.getlist(f"work_files_{sid}") or []
                task_files = request.FILES.getlist(f"task_files_{sid}") or []
                raw_score = (request.POST.get(f"score_{sid}") or "").strip()
                raw_percent = (request.POST.get(f"percent_{sid}") or "").strip()
                raw_comment = (request.POST.get(f"teacher_comment_{sid}") or "").strip()
                absent_in_exam = bool(request.POST.get(f"absent_{sid}"))
                retake_recommended = bool(request.POST.get(f"retake_{sid}"))

                if absent_in_exam:
                    raw_score = ""
                    raw_percent = ""
                elif raw_score and not raw_percent:
                    try:
                        computed_percent = (Decimal(raw_score) / max_score) * Decimal("100")
                        raw_percent = str(max(Decimal("0"), min(computed_percent, Decimal("100"))).quantize(Decimal("0.1")))
                    except Exception:
                        raw_percent = ""

                has_any_input = bool(
                    raw_score
                    or raw_percent
                    or raw_comment
                    or absent_in_exam
                    or retake_recommended
                    or work_files
                    or task_files
                )
                if not has_any_input:
                    continue

                row_form = ExamResultRowForm(
                    {
                        "score": raw_score,
                        "percent": raw_percent,
                        "teacher_comment": raw_comment or session_comment,
                        "assignment_description": session_task,
                        "absent_in_exam": absent_in_exam,
                        "retake_recommended": retake_recommended,
                    },
                    require_result=bool(settings_obj.exam_result_required or is_finalize),
                )
                if not row_form.is_valid():
                    row_errors.append((enrollment.student.get_full_name(), row_form.errors.as_text()))
                    continue

                rows.append(
                    {
                        "student": enrollment.student,
                        "score": row_form.cleaned_data.get("score"),
                        "percent": row_form.cleaned_data.get("percent"),
                        "teacher_comment": row_form.cleaned_data.get("teacher_comment"),
                        "assignment_description": row_form.cleaned_data.get("assignment_description"),
                        "absent_in_exam": row_form.cleaned_data.get("absent_in_exam"),
                        "retake_recommended": row_form.cleaned_data.get("retake_recommended"),
                        "work_files": work_files,
                        "task_files": task_files,
                    }
                )

            if row_errors:
                for student_name, err in row_errors:
                    messages.error(request, f"{student_name}: {err}")
            else:
                saved_count = 0
                if rows:
                    try:
                        saved_count = save_exam_results_batch(
                            session=session,
                            actor=request.user,
                            rows=rows,
                            finalize=is_finalize,
                        )
                    except ValueError as exc:
                        messages.error(request, str(exc))
                        saved_count = -1
                elif is_finalize:
                    session_progress = get_exam_session_progress(session=session)
                    session.status = (
                        ExamSession.STATUS_COMPLETED
                        if session_progress["is_completed"]
                        else ExamSession.STATUS_DRAFT
                    )
                    if session.teacher_decision != ExamSession.DECISION_YES:
                        session.teacher_decision = ExamSession.DECISION_YES
                    session.updated_by = request.user
                    session.save(update_fields=["status", "teacher_decision", "updated_by", "updated_at"])

                if saved_count >= 0:
                    session.refresh_from_db()
                    session_progress = get_exam_session_progress(session=session)

                    if is_finalize:
                        if not session_progress["is_completed"]:
                            messages.error(
                                request,
                                "Sessiyani yakunlash uchun barcha o'quvchilar bo'yicha ball yoki qatnashmagan holati kiritilishi kerak.",
                            )
                        else:
                            if not was_completed and session.status == ExamSession.STATUS_COMPLETED:
                                notify_exam_results(session)
                                auto_check_certificate_eligibility(session)
                                messages.success(request, "Sessiya yakunlandi. Bildirishnomalar va sertifikat tekshiruvi ishga tushdi.")
                            else:
                                messages.info(request, "Sessiya allaqachon yakunlangan.")
                        return redirect(_exam_entry_url(session.id, max_score))

                    if saved_count > 0:
                        messages.success(request, f"{saved_count} ta o'quvchi bo'yicha imtihon natijalari saqlandi.")
                        return redirect(_exam_entry_url(session.id, max_score))
                    if uploaded_task_file_count or note_updated:
                        messages.success(request, "Sessiya ma'lumotlari yangilandi.")
                        return redirect(_exam_entry_url(session.id, max_score))
                    messages.info(request, "Hozircha saqlash uchun yangi natija yo'q.")

        session_progress = get_exam_session_progress(session=session)

        return render(
            request,
            "education/exam_session_entry.html",
            {
                "session": session,
                "group": session.group,
                "enrollments": enrollments,
                "existing_results": existing_results,
                "exam_settings": settings_obj,
                "session_task_default": session_task_default,
                "session_comment_default": session_comment_default,
                "session_progress": session_progress,
                "max_score": max_score,
                "passing_percent": passing_percent,
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("exam_session_entry failed: session_id=%s", session_id)
        messages.error(request, "Imtihon sessiyasi bilan ishlashda xatolik yuz berdi.")
        return redirect("education:teacher_exam_history")


@login_required
@require_feature("imtihon")
def group_exam_history(request, group_id: int):
    from core.tenant import get_request_center
    from education.models import ExamSession

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_EXAM_SESSIONS,
        message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    group_qs = Group.objects.all()
    if center:
        group_qs = group_qs.filter(center=center)
    group = get_object_or_404(group_qs, pk=group_id)

    if not _teacher_or_management_can_access_group(request.user, group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    sessions = ExamSession.objects.filter(group=group).select_related("teacher", "created_by").order_by("-exam_date", "-id")
    status_filter = (request.GET.get("status") or "").strip()
    if status_filter in {ExamSession.STATUS_DRAFT, ExamSession.STATUS_COMPLETED, ExamSession.STATUS_CANCELLED}:
        sessions = sessions.filter(status=status_filter)

    return render(
        request,
        "education/group_exam_history.html",
        {
            "group": group,
            "sessions": sessions,
            "status_filter": status_filter,
        },
    )


@login_required
@require_feature("imtihon")
def teacher_exam_history(request):
    from core.tenant import get_request_center
    from education.models import ExamSession

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_EXAM_SESSIONS,
        message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    role = getattr(request.user, "role", None)
    if not request.user.is_superuser and role not in ("director", "manager", "teacher"):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    sessions = ExamSession.objects.select_related("group", "teacher", "center").order_by("-exam_date", "-id")
    if center:
        sessions = sessions.filter(center=center)

    if role == "teacher" and not request.user.is_superuser:
        sessions = sessions.filter(teacher=request.user)
    else:
        teacher_id = _get_int(request.GET, "teacher", 0)
        if teacher_id:
            sessions = sessions.filter(teacher_id=teacher_id)

    group_id = _get_int(request.GET, "group", 0)
    if group_id:
        sessions = sessions.filter(group_id=group_id)

    back_group_id = _get_int(request.GET, "back_group", 0)
    if not back_group_id and group_id:
        back_group_id = group_id

    date_from = parse_date(request.GET.get("date_from") or "")
    date_to = parse_date(request.GET.get("date_to") or "")
    if date_from:
        sessions = sessions.filter(exam_date__gte=date_from)
    if date_to:
        sessions = sessions.filter(exam_date__lte=date_to)

    teachers = User.objects.filter(role="teacher", center=center).order_by("ism", "familya") if center else User.objects.none()
    groups = Group.objects.filter(center=center).order_by("nom") if center else Group.objects.none()
    return render(
        request,
        "education/teacher_exam_history.html",
        {
            "sessions": sessions,
            "teachers": teachers,
            "groups": groups,
            "back_group_id": back_group_id,
            "filters": {
                "teacher": _get_int(request.GET, "teacher", 0),
                "group": group_id,
                "back_group": back_group_id,
                "date_from": date_from.isoformat() if date_from else "",
                "date_to": date_to.isoformat() if date_to else "",
            },
        },
    )


@login_required
@require_feature("imtihon")
def exam_session_detail(request, session_id: int):
    from core.tenant import get_request_center
    from education.models import ExamResult, ExamSession

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_EXAM_SESSIONS,
        message="Imtihon sessiyalari bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    qs = ExamSession.objects.select_related("group", "teacher", "center")
    if center:
        qs = qs.filter(center=center)
    session = get_object_or_404(qs, pk=session_id)

    if not _teacher_or_management_can_access_group(request.user, session.group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    results = (
        ExamResult.objects.filter(session=session)
        .select_related("student", "teacher")
        .prefetch_related("files")
        .order_by("student__ism", "student__familya")
    )

    return render(
        request,
        "education/exam_session_detail.html",
        {
            "session": session,
            "group": session.group,
            "results": results,
        },
    )


@login_required
@require_feature("imtihon")
def failed_students_list(request):
    from core.tenant import get_request_center
    from .forms import ExamResultFollowUpForm
    from education.models import ExamResult
    from education.services.audit_service import log_education_event

    try:
        role = getattr(request.user, "role", None)
        if not request.user.is_superuser and role not in ("director", "manager", "teacher"):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        center = get_request_center(request)
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_FAILED_STUDENTS,
            message="Zaif o'quvchilar bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response
        qs = ExamResult.objects.select_related("student", "group", "teacher", "session")
        if center:
            qs = qs.filter(center=center)
        qs = qs.filter(passed=False).filter(
            Q(absent_in_exam=True) | Q(score__isnull=False) | Q(percent__isnull=False)
        )

        if role == "teacher" and not request.user.is_superuser:
            qs = qs.filter(teacher=request.user)

        group_id = _get_int(request.GET, "group", 0)
        teacher_id = _get_int(request.GET, "teacher", 0)
        date_from = parse_date(request.GET.get("date_from") or "")
        date_to = parse_date(request.GET.get("date_to") or "")
        percent_min = request.GET.get("percent_min")
        percent_max = request.GET.get("percent_max")
        follow_up_status = (request.GET.get("follow_up_status") or "").strip()
        follow_up_pending = (request.GET.get("follow_up_pending") or "").strip().lower() in {"1", "true", "yes", "on"}

        if group_id:
            qs = qs.filter(group_id=group_id)
        if teacher_id and role != "teacher":
            qs = qs.filter(teacher_id=teacher_id)
        if date_from:
            qs = qs.filter(exam_date__gte=date_from)
        if date_to:
            qs = qs.filter(exam_date__lte=date_to)
        if percent_min not in (None, ""):
            try:
                qs = qs.filter(percent__gte=Decimal(str(percent_min)))
            except Exception:
                pass
        if percent_max not in (None, ""):
            try:
                qs = qs.filter(percent__lte=Decimal(str(percent_max)))
            except Exception:
                pass
        valid_follow_statuses = {choice[0] for choice in ExamResult.FOLLOW_UP_CHOICES}
        if follow_up_status in valid_follow_statuses:
            qs = qs.filter(follow_up_status=follow_up_status)
        if follow_up_pending:
            qs = qs.filter(follow_up_status=ExamResult.FOLLOW_UP_PENDING)

        if request.method == "POST":
            result_id = _get_int(request.POST, "result_id", 0)
            result = get_object_or_404(qs, pk=result_id)
            follow_form = ExamResultFollowUpForm(request.POST, instance=result)
            if follow_form.is_valid():
                updated = follow_form.save(commit=False)
                updated.follow_up_updated_by = request.user
                updated.follow_up_updated_at = timezone.now()
                updated.save(
                    update_fields=[
                        "follow_up_status",
                        "follow_up_note",
                        "follow_up_updated_by",
                        "follow_up_updated_at",
                        "updated_at",
                    ]
                )
                log_education_event(
                    center=updated.center,
                    actor=request.user,
                    action_type="exam_followup_updated",
                    entity=updated,
                    payload={"follow_up_status": updated.follow_up_status},
                )
                messages.success(request, "Nazorat holati yangilandi.")
            else:
                messages.error(request, "Nazorat formasi xato.")
            q = request.META.get("QUERY_STRING")
            return redirect(f"{request.path}?{q}" if q else request.path)

        rows = qs.order_by("-exam_date", "-id")
        summary_counts = {
            "total": rows.count(),
            "pending": rows.filter(follow_up_status=ExamResult.FOLLOW_UP_PENDING).count(),
            "absent": rows.filter(absent_in_exam=True).count(),
        }
        groups = Group.objects.filter(center=center).order_by("nom") if center else Group.objects.none()
        if role == "teacher" and not request.user.is_superuser:
            groups = groups.filter(oqituvchi=request.user)
        if role == "teacher" and not request.user.is_superuser:
            teachers = User.objects.filter(pk=request.user.pk)
            teacher_id = request.user.id
        else:
            teachers = (
                User.objects.filter(role="teacher", center=center, is_archived=False).order_by("ism", "familya")
                if center
                else User.objects.none()
            )
        follow_up_choices = ExamResult.FOLLOW_UP_CHOICES

        return render(
            request,
            "education/failed_students_list.html",
            {
                "rows": rows,
                "summary_counts": summary_counts,
                "groups": groups,
                "teachers": teachers,
                "follow_up_choices": follow_up_choices,
                "filters": {
                    "group": group_id,
                    "teacher": teacher_id,
                    "date_from": date_from.isoformat() if date_from else "",
                    "date_to": date_to.isoformat() if date_to else "",
                    "percent_min": percent_min or "",
                    "percent_max": percent_max or "",
                    "follow_up_status": follow_up_status,
                    "follow_up_pending": follow_up_pending,
                },
                "can_edit_follow_up": _director_or_manager(request.user),
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("failed_students_list failed")
        messages.error(request, "Zaif o'quvchilar ro'yxatini yuklashda xatolik yuz berdi.")
        return redirect("education:teacher_exam_history")


@login_required
@require_feature("imtihon")
def group_internal_ranking(request, group_id: int):
    from core.tenant import get_request_center
    from education.services.ranking_service import INTERNAL_RANKING_WEIGHTS, build_group_internal_ranking

    center = get_request_center(request)
    group_qs = Group.objects.all()
    if center:
        group_qs = group_qs.filter(center=center)
    group = get_object_or_404(group_qs, pk=group_id)

    if not _teacher_or_management_can_access_group(request.user, group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    selected_date = parse_date(request.GET.get("date") or "") or localdate()
    rows = build_group_internal_ranking(
        group=group,
        on_date=selected_date,
        actor=request.user,
        persist=True,
    )

    return render(
        request,
        "education/group_internal_ranking.html",
        {
            "group": group,
            "selected_date": selected_date,
            "rows": rows,
            "weights": INTERNAL_RANKING_WEIGHTS,
        },
    )


@login_required
@require_feature("imtihon")
def group_completion_recommendations(request, group_id: int):
    from core.tenant import get_request_center
    from education.services.ranking_service import build_group_completion_recommendations

    center = get_request_center(request)
    group_qs = Group.objects.all()
    if center:
        group_qs = group_qs.filter(center=center)
    group = get_object_or_404(group_qs, pk=group_id)

    if not _teacher_or_management_can_access_group(request.user, group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    selected_date = parse_date(request.GET.get("date") or "") or localdate()
    recommendation_payload = build_group_completion_recommendations(
        group=group,
        on_date=selected_date,
        actor=request.user,
        persist=True,
    )
    rows = recommendation_payload["rows"]
    selected_status = (request.GET.get("status") or "").strip()
    valid_statuses = {"eligible", "needs_review", "not_eligible"}
    if selected_status in valid_statuses:
        rows = [row for row in rows if row["completion_recommendation"] == selected_status]

    days_to_estimated_end = None
    days_to_estimated_end_abs = None
    if group.estimated_end_date:
        days_to_estimated_end = (group.estimated_end_date - selected_date).days
        days_to_estimated_end_abs = abs(days_to_estimated_end)

    return render(
        request,
        "education/group_completion_recommendations.html",
        {
            "group": group,
            "selected_date": selected_date,
            "rows": rows,
            "thresholds": recommendation_payload["thresholds"],
            "selected_status": selected_status,
            "days_to_estimated_end": days_to_estimated_end,
            "days_to_estimated_end_abs": days_to_estimated_end_abs,
        },
    )


@login_required
@require_POST
@require_feature("imtihon")
def group_closure_action(request, group_id: int):
    from core.tenant import get_request_center
    from education.services.closure_service import apply_group_closure_action

    center = get_request_center(request)
    group_qs = Group.objects.all()
    if center:
        group_qs = group_qs.filter(center=center)
    group = get_object_or_404(group_qs, pk=group_id)

    if not _teacher_or_management_can_access_group(request.user, group):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    action = (request.POST.get("action") or "").strip().lower()
    if action not in {"yes", "no", "later"}:
        messages.error(request, "Noto'g'ri action.")
        return redirect("education:group_detail", pk=group.id)

    if action == "yes" and not _director_or_manager(request.user):
        return HttpResponseForbidden("Guruhni yopish faqat director/manager uchun ruxsat etilgan.")

    selected_date = parse_date(request.POST.get("date") or "") or localdate()
    note = (request.POST.get("note") or "").strip()

    workflow = apply_group_closure_action(
        group=group,
        actor=request.user,
        action=action,
        on_date=selected_date,
        note=note,
    )

    if workflow.status == workflow.STATUS_CLOSED:
        messages.success(request, "Guruhni yopish jarayoni yakunlandi. Tarixiy ma'lumotlar saqlandi.")
    elif workflow.status == workflow.STATUS_CONTINUE:
        messages.info(request, "Guruh davom etadi. Attendance va payment flow o'zgarmaydi.")
    else:
        messages.info(request, "Closure eslatmasi keyinga qoldirildi.")
    return redirect("education:group_detail", pk=group.id)


@login_required
@require_feature("sertifikat")
def certificate_templates_view(request):
    from core.tenant import get_request_center
    from .forms import CertificateTemplateForm
    from education.models import CertificateTemplate
    from education.services.audit_service import log_education_event

    center = get_request_center(request)
    if not center:
        raise Http404("Center not found")
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_CERTIFICATES,
        message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response

    if not _director_or_manager(request.user):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    form = CertificateTemplateForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.center = center
        obj.uploaded_by = request.user
        obj.save()
        if obj.is_active:
            CertificateTemplate.objects.filter(center=center, template_type=obj.template_type).exclude(pk=obj.pk).update(
                is_active=False
            )
        log_education_event(
            center=center,
            actor=request.user,
            action_type="certificate_template_uploaded",
            entity=obj,
            payload={"template_type": obj.template_type},
        )
        messages.success(request, "Shablon saqlandi.")
        return redirect("education:certificate_templates")

    templates = CertificateTemplate.objects.filter(center=center).order_by("-updated_at")
    groups_overview = (
        Group.objects.filter(center=center, is_archived=False)
        .select_related("oqituvchi")
        .annotate(
            active_students_count=Count(
                "enrollments",
                filter=Q(enrollments__is_active=True, enrollments__is_deleted=False),
                distinct=True,
            ),
            issued_certificates_count=Count(
                "certificates",
                filter=Q(certificates__status="issued"),
                distinct=True,
            ),
        )
        .order_by("nom")
    )
    return render(
        request,
        "education/certificate_templates.html",
        {
            "form": form,
            "templates": templates,
            "groups_overview": groups_overview,
            "active_templates_count": templates.filter(is_active=True).count(),
        },
    )


@login_required
@require_POST
@require_feature("sertifikat")
def certificate_template_activate(request, template_id: int):
    from core.tenant import get_request_center
    from education.models import CertificateTemplate
    from education.services.audit_service import log_education_event

    if not _director_or_manager(request.user):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_CERTIFICATES,
        message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    qs = CertificateTemplate.objects.all()
    if center:
        qs = qs.filter(center=center)
    template = get_object_or_404(qs, pk=template_id)

    CertificateTemplate.objects.filter(center=template.center, template_type=template.template_type).update(is_active=False)
    template.is_active = True
    template.save(update_fields=["is_active", "updated_at"])

    log_education_event(
        center=template.center,
        actor=request.user,
        action_type="certificate_template_activated",
        entity=template,
        payload={"template_type": template.template_type},
    )
    messages.success(request, "Shablon faol qilib belgilandi.")
    return redirect("education:certificate_templates")


@login_required
@require_feature("sertifikat")
def group_certificate_candidates(request, group_id: int):
    from core.tenant import get_request_center
    from education.models import CertificateRecord
    from education.services.ranking_service import build_group_completion_recommendations

    try:
        center = get_request_center(request)
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_CERTIFICATES,
            message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response
        group_qs = Group.objects.all()
        if center:
            group_qs = group_qs.filter(center=center)
        group = get_object_or_404(group_qs, pk=group_id)

        if not _teacher_or_management_can_access_group(request.user, group):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        selected_date = parse_date(request.GET.get("date") or "") or localdate()
        recommendation_payload = build_group_completion_recommendations(
            group=group,
            on_date=selected_date,
            actor=request.user,
            persist=True,
        )
        rows = recommendation_payload["rows"]
        existing_certs = {}
        for cert in (
            CertificateRecord.objects.filter(
                group=group,
                status__in=[CertificateRecord.STATUS_DRAFT, CertificateRecord.STATUS_ISSUED],
            )
            .select_related("student")
            .order_by("student_id", "-created_at", "-id")
        ):
            current = existing_certs.get(cert.student_id)
            if current is None or (current.status != CertificateRecord.STATUS_ISSUED and cert.status == CertificateRecord.STATUS_ISSUED):
                existing_certs[cert.student_id] = cert

        for row in rows:
            row["certificate"] = existing_certs.get(row["student"].id)

        return render(
            request,
            "education/group_certificate_candidates.html",
            {
                "group": group,
                "selected_date": selected_date,
                "rows": rows,
                "thresholds": recommendation_payload["thresholds"],
                "can_issue": _director_or_manager(request.user),
            },
        )
    except Http404:
        raise
    except Exception:
        logger.exception("group_certificate_candidates failed: group_id=%s", group_id)
        messages.error(request, "Sertifikat nomzodlarini yuklashda xatolik yuz berdi.")
        return redirect("education:group_detail", pk=group_id)


@login_required
@require_POST
@require_feature("sertifikat")
def issue_certificate_action(request, group_id: int, student_id: int):
    from core.tenant import get_request_center
    from .forms import CertificateIssueForm
    from education.services.certificate_service import issue_certificate_for_student

    try:
        if not _director_or_manager(request.user):
            return HttpResponseForbidden("Sizda ruxsat yo'q.")

        center = get_request_center(request)
        disabled_response = _ensure_center_ui_feature(
            request,
            center,
            FEATURE_UI_CERTIFICATES,
            message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
        )
        if disabled_response:
            return disabled_response
        group_qs = Group.objects.all()
        if center:
            group_qs = group_qs.filter(center=center)
        group = get_object_or_404(group_qs, pk=group_id)
        student = get_object_or_404(User.objects.filter(role="student"), pk=student_id)

        if not Enrollment.objects.filter(group=group, student=student).exists():
            return HttpResponseForbidden("Student bu guruhga biriktirilmagan.")

        form = CertificateIssueForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Sertifikat berish formasi xato.")
            return redirect("education:group_certificate_candidates", group_id=group.id)

        cert = issue_certificate_for_student(
            group=group,
            student=student,
            actor=request.user,
            certificate_type=form.cleaned_data["certificate_type"],
            note=form.cleaned_data.get("note", ""),
            request=request,
        )
        messages.success(request, f"Sertifikat tasdiqlandi: {cert.certificate_number}")
        return redirect("education:certificate_detail", certificate_id=cert.id)
    except Http404:
        raise
    except Exception:
        logger.exception(
            "issue_certificate_action failed: group_id=%s student_id=%s",
            group_id,
            student_id,
        )
        messages.error(request, "Sertifikatni tasdiqlashda xatolik yuz berdi.")
        return redirect("education:group_certificate_candidates", group_id=group_id)


@login_required
@require_feature("sertifikat")
def certificate_detail(request, certificate_id: int):
    from core.tenant import get_request_center
    from education.models import CertificateRecord
    from education.services.certificate_service import user_can_view_certificate

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_CERTIFICATES,
        message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    qs = CertificateRecord.objects.select_related("group", "student", "center", "template", "summary")
    if center:
        qs = qs.filter(center=center)
    cert = get_object_or_404(qs, pk=certificate_id)

    if not user_can_view_certificate(request.user, cert):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    return render(
        request,
        "education/certificate_detail.html",
        {
            "cert": cert,
        },
    )


@login_required
@require_feature("sertifikat")
def certificate_download_pdf(request, certificate_id: int):
    from core.tenant import get_request_center
    from education.models import CertificateRecord
    from education.services.certificate_service import (
        PDF_LAYOUT_VERSION,
        regenerate_certificate_pdf,
        user_can_view_certificate,
    )

    center = get_request_center(request)
    disabled_response = _ensure_center_ui_feature(
        request,
        center,
        FEATURE_UI_CERTIFICATES,
        message="Sertifikatlar bo'limi bu markaz uchun o'chirilgan.",
    )
    if disabled_response:
        return disabled_response
    qs = CertificateRecord.objects.select_related("group", "student", "center", "summary")
    if center:
        qs = qs.filter(center=center)
    cert = get_object_or_404(qs, pk=certificate_id)

    if not user_can_view_certificate(request.user, cert):
        return HttpResponseForbidden("Sizda ruxsat yo'q.")

    metadata = cert.metadata if isinstance(cert.metadata, dict) else {}
    layout_version = metadata.get("pdf_layout_version")
    if (not cert.pdf_file) or layout_version != PDF_LAYOUT_VERSION:
        cert = regenerate_certificate_pdf(record=cert, request=request)

    cert.pdf_file.open("rb")
    return FileResponse(
        cert.pdf_file,
        as_attachment=True,
        filename=f"{cert.certificate_number}.pdf",
        content_type="application/pdf",
    )


def certificate_verify(request, certificate_number: str):
    from education.models import CertificateRecord
    from education.services.certificate_service import record_verification_hit

    cert = get_object_or_404(
        CertificateRecord.objects.select_related("group", "student", "center"),
        certificate_number=certificate_number,
    )

    record_verification_hit(record=cert, request=request)
    return render(
        request,
        "education/certificate_verify.html",
        {
            "cert": cert,
        },
    )


@login_required
def student_exam_report(request, student_id: int):
    from core.tenant import get_request_center
    center = get_request_center(request)
    student = get_object_or_404(User.objects.filter(role="student"), pk=student_id)

    viewer = request.user
    if viewer.role == "student" and viewer.id != student.id:
        return HttpResponseForbidden("Siz faqat o'zingizning natijangizni ko'ra olasiz.")
    if viewer.role == "parent" and student not in viewer.children.all():
        return HttpResponseForbidden("Siz faqat farzandingizning natijalarini ko'ra olasiz.")
    if viewer.role == "teacher":
        teaches_student = Enrollment.objects.filter(
            student=student,
            group__oqituvchi=viewer,
            is_active=True,
        ).exists()
        if not teaches_student:
            return HttpResponseForbidden("Siz bu o'quvchining natijasini ko'ra olmaysiz.")
    if viewer.role not in ("student", "parent", "teacher", "director", "manager") and not viewer.is_superuser:
        return HttpResponseForbidden("Ruxsat yo'q.")

    from education.models import ExamResult
    from education.services.exam_service import get_student_exam_summary
    from education.services.ranking_service import get_student_academic_summaries

    qs = ExamResult.objects.select_related("group", "teacher", "session").filter(student=student)
    if center:
        qs = qs.filter(center=center)
    certificate_qs = CertificateRecord.objects.select_related("group", "center").filter(student=student)
    if center:
        certificate_qs = certificate_qs.filter(center=center)

    summary = get_student_exam_summary(student=student)
    academic_summaries = get_student_academic_summaries(student=student, center=center)
    return render(
        request,
        "education/student_exam_report.html",
        {
            "student": student,
            "results": qs.order_by("-exam_date", "-id"),
            "summary": summary,
            "academic_summaries": academic_summaries,
            "certificates": certificate_qs.order_by("-created_at", "-id"),
        },
    )


# ==========================================
#  OY PREVIEW (READ-ONLY: close_month oldidan)
# ==========================================

@login_required
@require_GET
@require_feature("finance")
def month_preview(request):
    from core.tenant import get_request_center

    if not user_can_manage_payments(request.user):
        return HttpResponseForbidden("Ruxsat yo'q.")

    center = get_request_center(request)

    month_raw = (request.GET.get("month") or "").strip()
    month = parse_month_str(month_raw) if month_raw else timezone.localdate().replace(day=1)
    if month is None:
        month = timezone.localdate().replace(day=1)

    group_id = _get_int(request.GET, "group", 0)

    fee_field = tuition_month_fee_field()
    m_start = month_first_day(month)
    m_end = month_last_day(m_start)

    enrollments_qs = (
        Enrollment.objects
        .select_related("student", "group")
        .filter(is_active=True, student__is_archived=False, group__is_archived=False, group__is_deleted=False)
    )
    if center:
        enrollments_qs = enrollments_qs.filter(center=center)
    if group_id:
        enrollments_qs = enrollments_qs.filter(group_id=group_id)

    existing_tm = {
        tm.enrollment_id: tm
        for tm in TuitionMonth.all_objects.filter(
            enrollment__in=enrollments_qs, month=m_start, is_deleted=False
        )
    }

    rows = []
    total_current = 0
    total_prorated = 0
    total_reconciled = 0

    for enr in enrollments_qs:
        tm = existing_tm.get(enr.id)
        current_fee = int(getattr(tm, fee_field, 0) or 0) if tm else 0
        prorated = int(prorated_monthly_fee(enr, m_start) or 0)
        reconciled = int(attendance_based_fee(enr, m_start) or 0)
        billable = billable_attendance_count(enr, m_start)

        start_d = enrollment_start_date(enr)
        period_start = max(start_d, m_start)
        expected_lessons = expected_lessons_in_period(enr, period_start, m_end) if period_start <= m_end else 0

        delta = reconciled - current_fee

        rows.append({
            "enrollment": enr,
            "student": enr.student,
            "group": enr.group,
            "start_date": start_d,
            "effective_price": effective_student_payable_amount(enr),
            "full_price": full_course_amount(enr),
            "current_fee": current_fee,
            "prorated_fee": prorated,
            "reconciled_fee": reconciled,
            "billable_lessons": billable,
            "expected_lessons": expected_lessons,
            "delta": delta,
            "has_tm": tm is not None,
        })

        total_current += current_fee
        total_prorated += prorated
        total_reconciled += reconciled

    rows.sort(key=lambda r: abs(r["delta"]), reverse=True)

    group_options = (
        Group.objects.filter(is_archived=False)
        .filter(center=center) if center else Group.objects.filter(is_archived=False)
    )

    return render(
        request,
        "education/month_preview.html",
        {
            "month": m_start,
            "month_str": m_start.strftime("%Y-%m"),
            "rows": rows,
            "total_current": total_current,
            "total_prorated": total_prorated,
            "total_reconciled": total_reconciled,
            "total_delta": total_reconciled - total_current,
            "group_options": group_options.order_by("nom"),
            "selected_group_id": group_id,
        },
    )


# ============================================================
# TASK 3: TuitionMonth fee_amount ni tahrirlash (ruchka tugma)
# ============================================================

_UZ_MONTHS = {
    1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
    7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr",
}


@require_POST
@login_required
def edit_tuition_month_fee(request, tm_id):
    """
    TuitionMonth.fee_amount ni yangilaydi.
    Multi-tenant: TuitionMonth faqat joriy markazga tegishli bo'lsa ruxsat.
    """
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q."}, status=403)

    center = get_active_center(request)
    qs = TuitionMonth.objects.select_related("enrollment", "enrollment__group")
    if center:
        from django.db.models import Q as _Q
        qs = qs.filter(
            _Q(center=center)
            | _Q(enrollment__center=center)
            | _Q(enrollment__group__center=center)
        )
    tm = get_object_or_404(qs, id=tm_id)

    new_fee_raw = (request.POST.get("new_fee") or "").strip()
    note = (request.POST.get("note") or "").strip()

    try:
        new_fee = int(Decimal(new_fee_raw or "0"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Noto'g'ri summa."}, status=400)

    if new_fee < 0:
        return JsonResponse({"ok": False, "error": "Summa manfiy bo'lishi mumkin emas."}, status=400)

    fee_field = tuition_month_fee_field()
    setattr(tm, fee_field, new_fee)
    update_fields = [fee_field]
    if note:
        # Note ni TuitionMonth'da saqlash uchun note maydoni bo'lsa
        from education.services.tuition import _model_has_field as _mhf
        if _mhf(TuitionMonth, "note"):
            tm.note = note
            update_fields.append("note")
    tm.save(update_fields=update_fields)

    month_label = f"{_UZ_MONTHS.get(tm.month.month, tm.month.month)} {tm.month.year}"
    return JsonResponse({
        "ok": True,
        "new_fee": new_fee,
        "new_fee_display": format_money(new_fee),
        "month_label": month_label,
    })


# ============================================================
# Oylik umumiy qarzni o'rnatish (barcha TuitionMonth'lar)
# ============================================================

@require_POST
@login_required
def edit_student_month_debt(request, student_id):
    """
    O'quvchining bitta oy uchun umumiy qarz miqdorini o'rnatadi.
    Bir oy ichida bir nechta TuitionMonth bo'lsa ham hammasi yangilanadi:
    - Birinchi TM: fee = new_debt + TM.paid  (bu TMda qarz = new_debt)
    - Qolgan TMlar: fee = TM.paid            (bu TMlarda qarz = 0)
    POST: month="2026-04", new_debt=500000
    """
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q."}, status=403)

    center = get_active_center(request)

    month_str = (request.POST.get("month") or "").strip()
    new_debt_raw = (request.POST.get("new_debt") or "").strip()

    try:
        y, m = int(month_str[:4]), int(month_str[5:7])
        month_date = date(y, m, 1)
    except Exception:
        return JsonResponse({"ok": False, "error": "Noto'g'ri oy formati."}, status=400)

    # Kelajak oylarni tahrirlash mumkin emas — ular avtomatik hisoblanadi
    cur_month_first = timezone.localdate().replace(day=1)
    if month_date > cur_month_first:
        return JsonResponse({"ok": False, "error": "Kelajakdagi oy uchun to'lovni tahrirlash mumkin emas. Oy kelganda avtomatik hisoblanadi."}, status=400)

    try:
        new_debt = int(Decimal(new_debt_raw or "0"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Noto'g'ri summa."}, status=400)

    if new_debt < 0:
        return JsonResponse({"ok": False, "error": "Qarz manfiy bo'lishi mumkin emas."}, status=400)

    from django.db.models import Q as _Q2
    user_qs = User.objects.filter(role="student")
    if center:
        user_qs = user_qs.filter(center=center)
    student = get_object_or_404(user_qs, id=student_id)

    # Enrollment filter must match student_monthly_breakdown exactly so that
    # the same TuitionMonths are updated here and read there.
    _center_q_edit = (
        _Q2(center=center)
        | _Q2(center__isnull=True, group__center=center)
        | _Q2(center__isnull=True, student__center=center)
    )
    enrollments_for_student = Enrollment.objects.filter(student=student)
    if center:
        enrollments_for_student = enrollments_for_student.filter(_center_q_edit)

    tms_qs = TuitionMonth.objects.filter(
        enrollment__in=enrollments_for_student,
        month=month_date,
        is_deleted=False,
    ).select_related("enrollment").prefetch_related(
        Prefetch(
            "allocations",
            queryset=PaymentAllocation.objects.filter(payment__is_deleted=False),
            to_attr="active_allocations",
        )
    )

    tms = list(tms_qs.order_by("id"))
    if not tms:
        return JsonResponse({"ok": False, "error": "Bu oy uchun yozuv topilmadi."}, status=404)

    fee_field = tuition_month_fee_field()
    paid_per_tm = [sum(int(a.amount or 0) for a in tm.active_allocations) for tm in tms]
    total_paid_now = sum(paid_per_tm)

    # new_paid: to'langan summani kamaytirish (ixtiyoriy)
    new_paid_raw = request.POST.get("new_paid")
    new_paid = None
    if new_paid_raw is not None:
        try:
            new_paid = int(Decimal(new_paid_raw.strip()))
        except Exception:
            return JsonResponse({"ok": False, "error": "Noto'g'ri to'langan summa."}, status=400)
        if new_paid < 0:
            return JsonResponse({"ok": False, "error": "To'langan summa manfiy bo'lishi mumkin emas."}, status=400)
        if new_paid > total_paid_now:
            return JsonResponse({"ok": False, "error": "To'langan summani oshirish mumkin emas."}, status=400)

    with transaction.atomic():
        # Agar to'langan kamaytirish so'ralgan bo'lsa
        affected_payment_ids: set = set()
        if new_paid is not None and new_paid < total_paid_now:
            to_free = total_paid_now - new_paid
            # Barcha allocationlarni ko'rib chiqib, kerakli miqdorni ozod qilamiz
            for tm in tms:
                if to_free <= 0:
                    break
                allocs_sorted = sorted(tm.active_allocations, key=lambda a: a.amount, reverse=True)
                for alloc in allocs_sorted:
                    if to_free <= 0:
                        break
                    affected_payment_ids.add(alloc.payment_id)
                    alloc_amt = int(alloc.amount or 0)
                    if alloc_amt <= to_free:
                        # Bu allocationni to'liq o'chiramiz
                        to_free -= alloc_amt
                        Enrollment.objects.filter(pk=tm.enrollment_id).update(
                            credit_balance=F("credit_balance") + alloc_amt
                        )
                        alloc.is_deleted = True
                        alloc.save(update_fields=["is_deleted"])
                    else:
                        # Qisman kamaytiramiz
                        alloc.amount = alloc_amt - to_free
                        alloc.save(update_fields=["amount"])
                        Enrollment.objects.filter(pk=tm.enrollment_id).update(
                            credit_balance=F("credit_balance") + to_free
                        )
                        to_free = 0
            # paid_per_tm ni yangi holat bilan yangilaymiz
            paid_per_tm = [
                sum(int(a.amount or 0) for a in tm.allocations.filter(is_deleted=False))
                for tm in tms
            ]

            # To'lovlar bo'limida ham o'zgarishni aks ettirish:
            # ta'sirlangan Payment yozuvlarini yangilaymiz (yoki o'chiramiz)
            if affected_payment_ids:
                for pay_id in affected_payment_ids:
                    remaining_alloc = (
                        PaymentAllocation.objects
                        .filter(payment_id=pay_id, is_deleted=False)
                        .aggregate(s=Sum("amount"))["s"] or 0
                    )
                    pay_obj = Payment.all_objects.filter(pk=pay_id, is_deleted=False).first()
                    if pay_obj is None:
                        continue
                    if remaining_alloc == 0:
                        # Hech qanday aktiv allocation qolmadi → To'lovlar bo'limidan o'chiramiz
                        pay_obj.is_deleted = True
                        pay_obj.save(update_fields=["is_deleted"])
                    elif remaining_alloc < int(pay_obj.summa or 0):
                        # Qisman kamaytirish → summani yangi miqdorga o'rnatamiz
                        pay_obj.summa = remaining_alloc
                        # cash_amount ni ham mutanosib ravishda kamaytiramiz
                        old_cash = int(pay_obj.cash_amount or 0)
                        if old_cash > remaining_alloc:
                            pay_obj.cash_amount = remaining_alloc
                        pay_obj.save(update_fields=["summa", "cash_amount"])

        # Birinchi TM: fee = new_debt + uning to'lovi (qarz = new_debt)
        setattr(tms[0], fee_field, new_debt + paid_per_tm[0])
        tms[0].save(update_fields=[fee_field])

        # Qolgan TMlar: fee = to'lovi (qarz = 0)
        for i in range(1, len(tms)):
            setattr(tms[i], fee_field, paid_per_tm[i])
            tms[i].save(update_fields=[fee_field])

        # Foydalanuvchi qo'lda o'rnatgan fee ni _etm (ensure_tuition_month) qayta
        # yozib tashlaydi (schedule asosida hisoblaydi) — shu to'g'ri yozuvni
        # himoyalamiz: deleted_reason = "user_edit" => _etm bu TMni o'tkazib yuboradi.
        _protect_ids = set()
        for tm in tms:
            _cur_reason = getattr(tm, "deleted_reason", None) or ""
            if _cur_reason not in ("manual_cleared",) and not _cur_reason.startswith(
                ("cleanup_", "move_future_", "reset_", "user_edit")
            ):
                tm.deleted_reason = "user_edit"
                tm.save(update_fields=["deleted_reason"])
                _protect_ids.add(tm.id)

        # Inactive enrollment uchun new_debt=0: _etm qayta hisoblashining oldini
        # olish uchun TuitionMonth ni manual_cleared bilan soft-delete qilamiz.
        # Aks holda keyingi sahifa yuklanishida _etm davomat asosida fee ni
        # qayta tiklaydi (chiqarilgan o'quvchi uchun noto'g'ri).
        if new_debt == 0 and total_paid_now == 0:
            for tm in tms:
                _enr = tm.enrollment
                _is_inactive = (
                    not getattr(_enr, "is_active", True)
                    or getattr(_enr, "is_deleted", False)
                )
                if _is_inactive:
                    tm.is_deleted = True
                    tm.deleted_reason = "manual_cleared"
                    tm.deleted_at = timezone.now()
                    tm.save(update_fields=["is_deleted", "deleted_reason", "deleted_at"])

    # DB dan haqiqiy qiymatni o'qiymiz (Python object qiymatiga ishonmaymiz)
    tms[0].refresh_from_db(fields=[fee_field])
    saved_fee = int(getattr(tms[0], fee_field, 0) or 0) if tms else 0
    saved_paid = paid_per_tm[0] if paid_per_tm else 0
    expected_fee = new_debt + (paid_per_tm[0] if paid_per_tm else 0)
    if saved_fee != expected_fee:
        return JsonResponse({"ok": False, "error": "Ma'lumot saqlanmadi. Qaytadan urinib ko'ring."}, status=500)
    saved_debt = max(0, saved_fee - saved_paid)
    total_debt = get_student_total_debt(student, center)
    return JsonResponse({"ok": True, "fee": saved_fee, "paid": saved_paid, "debt": saved_debt, "total_debt": total_debt})


# ============================================================
# Kelajak oylik yozuvni o'chirish
# ============================================================

@require_POST
@login_required
def delete_student_month(request, student_id):
    """
    O'quvchining bitta kelajak oylik TuitionMonth yozuvlarini o'chiradi.
    Faqat joriy oydan KEYIN bo'lgan oylar uchun ruxsat beriladi.
    Agar to'lov allocatsiyalari mavjud bo'lsa, ular ham o'chiriladi va
    tegishli summa enrollment.credit_balance ga qaytariladi.
    POST: month="2026-06"
    """
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q."}, status=403)

    center = get_active_center(request)
    month_str = (request.POST.get("month") or "").strip()

    try:
        y, m_num = int(month_str[:4]), int(month_str[5:7])
        month_date = date(y, m_num, 1)
    except Exception:
        return JsonResponse({"ok": False, "error": "Noto'g'ri oy formati."}, status=400)

    today = date.today()
    current_month = date(today.year, today.month, 1)
    if month_date <= current_month:
        return JsonResponse({"ok": False, "error": "Faqat kelajak oylarni o'chirish mumkin."}, status=400)

    from django.db.models import Q as _Q4
    user_qs = User.objects.filter(role="student")
    if center:
        user_qs = user_qs.filter(center=center)
    student = get_object_or_404(user_qs, id=student_id)

    tms_qs = TuitionMonth.objects.filter(
        enrollment__student=student,
        month=month_date,
        is_deleted=False,
    ).prefetch_related(
        Prefetch(
            "allocations",
            queryset=PaymentAllocation.objects.filter(is_deleted=False),
            to_attr="active_allocations",
        )
    )
    if center:
        tms_qs = tms_qs.filter(
            _Q4(center=center)
            | _Q4(enrollment__center=center)
            | _Q4(enrollment__group__center=center)
        )

    tms = list(tms_qs)
    if not tms:
        return JsonResponse({"ok": False, "error": "Bu oy uchun yozuv topilmadi."}, status=404)

    with transaction.atomic():
        for tm in tms:
            has_alloc = bool(tm.active_allocations)

            if has_alloc:
                # To'lov bor: to'lovni bekor qilamiz.
                # Agar to'lov faqat shu oyga edi → to'lovni ham o'chiramiz (credit yo'q).
                # Agar to'lov boshqa oylarga ham tegishli → freed qismni credit_balance ga,
                # to'lov summasini kamaytiramiz.
                affected_pay_ids = {alloc.payment_id for alloc in tm.active_allocations}
                freed = sum(int(a.amount or 0) for a in tm.active_allocations)
                for alloc in tm.active_allocations:
                    alloc.is_deleted = True
                    alloc.save(update_fields=["is_deleted"])

                for pay_id in affected_pay_ids:
                    remaining_alloc = (
                        PaymentAllocation.objects
                        .filter(payment_id=pay_id, is_deleted=False)
                        .aggregate(s=Sum("amount"))["s"] or 0
                    )
                    pay_obj = Payment.all_objects.filter(pk=pay_id, is_deleted=False).first()
                    if pay_obj is None:
                        continue
                    if remaining_alloc == 0:
                        # To'lov butunlay bekor — o'chiramiz, credit yo'q
                        pay_obj.is_deleted = True
                        pay_obj.save(update_fields=["is_deleted"])
                    else:
                        # Qisman bekor: freed miqdor credit_balance ga
                        Enrollment.objects.filter(pk=tm.enrollment_id).update(
                            credit_balance=F("credit_balance") + freed
                        )
                        if remaining_alloc < int(pay_obj.summa or 0):
                            pay_obj.summa = remaining_alloc
                            old_cash = int(pay_obj.cash_amount or 0)
                            if old_cash > remaining_alloc:
                                pay_obj.cash_amount = remaining_alloc
                            pay_obj.save(update_fields=["summa", "cash_amount"])

                # To'lov bekor qilingani uchun shu oy qayta qarz bo'lishi kerak →
                # "future_deleted" dan _etm bu TM ni tiklaydi va oy yana to'lanmagan
                # ko'rinadi. Keyingi to'lov shu oyga ketadi.
                tm.is_deleted = True
                tm.deleted_reason = "future_deleted"
                tm.deleted_at = timezone.now()
                tm.save(update_fields=["is_deleted", "deleted_reason", "deleted_at"])
            else:
                # To'lov yo'q: oy jadvaldan butunlay o'chiriladi (o'quvchi ketmoqda).
                # "manual_cleared" → _etm bu oyni qayta tiklamaydi.
                tm.is_deleted = True
                tm.deleted_reason = "manual_cleared"
                tm.deleted_at = timezone.now()
                tm.save(update_fields=["is_deleted", "deleted_reason", "deleted_at"])

    return JsonResponse({"ok": True})


# ============================================================
# Oylik to'lovlarni bekor qilish — oy to'liq qarzga qaytadi
# ============================================================

@require_POST
@login_required
def reset_student_month_payments(request, student_id):
    """
    O'quvchining bitta oyi uchun barcha to'lov allocatsiyalarini bekor qiladi —
    oy TO'LIQ QARZ holatiga qaytadi (credit_balance ga o'tkazilmaydi!).
    Bekor qilingan qism To'lovlar bo'limida ham kamayadi/o'chadi.
    POST: month="2026-06"
    """
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q."}, status=403)

    center = get_active_center(request)
    month_str = (request.POST.get("month") or "").strip()

    try:
        y, m_num = int(month_str[:4]), int(month_str[5:7])
        month_date = date(y, m_num, 1)
    except Exception:
        return JsonResponse({"ok": False, "error": "Noto'g'ri oy formati."}, status=400)

    from django.db.models import Q as _Q5
    user_qs = User.objects.filter(role="student")
    if center:
        user_qs = user_qs.filter(center=center)
    student = get_object_or_404(user_qs, id=student_id)

    tms_qs = TuitionMonth.objects.filter(
        enrollment__student=student,
        month=month_date,
        is_deleted=False,
    ).prefetch_related(
        Prefetch(
            "allocations",
            queryset=PaymentAllocation.objects.filter(is_deleted=False, payment__is_deleted=False),
            to_attr="active_allocations",
        )
    )
    if center:
        tms_qs = tms_qs.filter(
            _Q5(center=center)
            | _Q5(enrollment__center=center)
            | _Q5(enrollment__group__center=center)
        )

    tms = list(tms_qs)
    if not tms:
        return JsonResponse({"ok": False, "error": "Bu oy uchun yozuv topilmadi."}, status=404)

    total_freed = 0
    with transaction.atomic():
        affected_pay_ids: set = set()
        for tm in tms:
            for alloc in tm.active_allocations:
                affected_pay_ids.add(alloc.payment_id)
                total_freed += int(alloc.amount or 0)
                alloc.is_deleted = True
                alloc.save(update_fields=["is_deleted"])

        # To'lovlar bo'limini sinxronlash: allocation qolmagan payment
        # o'chiriladi, qisman qolgani kamaytiriladi.
        for pay_id in affected_pay_ids:
            remaining_alloc = (
                PaymentAllocation.objects
                .filter(payment_id=pay_id, is_deleted=False)
                .aggregate(s=Sum("amount"))["s"] or 0
            )
            pay_obj = Payment.all_objects.filter(pk=pay_id, is_deleted=False).first()
            if pay_obj is None:
                continue
            if remaining_alloc == 0:
                pay_obj.is_deleted = True
                pay_obj.save(update_fields=["is_deleted"])
            elif remaining_alloc < int(pay_obj.summa or 0):
                pay_obj.summa = remaining_alloc
                old_cash = int(pay_obj.cash_amount or 0)
                if old_cash > remaining_alloc:
                    pay_obj.cash_amount = remaining_alloc
                pay_obj.save(update_fields=["summa", "cash_amount"])

        # Fee ni to'g'ri qiymatga qaytaramiz. ANIQ QOIDA:
        #   O'TGAN oy  → haqiqiy davomat asosida (nechta darsga kelgan bo'lsa
        #                shuncha × dars narxi). Davomat 0 bo'lsa qarz ham 0.
        #   JORIY oy   → to'liq oylik narx (jadval asosida).
        from education.services.tuition import (
            prorated_monthly_fee as _pmf,
            attendance_based_fee as _abf,
        )
        fee_field = tuition_month_fee_field()
        _cur_month_first = timezone.localdate().replace(day=1)
        _is_past_month = month_date < _cur_month_first
        for tm in tms:
            if _is_past_month:
                normal_fee = int(_abf(tm.enrollment, month_date) or 0)
            else:
                normal_fee = int(_pmf(tm.enrollment, month_date) or 0)
            _upd = []
            if int(getattr(tm, fee_field, 0) or 0) != normal_fee:
                setattr(tm, fee_field, normal_fee)
                _upd.append(fee_field)
            if _is_past_month:
                # O'tgan oy davomat asosidagi qiymatini _etm (jadval asosida
                # hisoblaydi) qayta yozib tashlamasin — user_edit bilan
                # himoyalaymiz. Davomat o'zgarsa signal baribir yangilaydi.
                if not (getattr(tm, "deleted_reason", None) or ""):
                    tm.deleted_reason = "user_edit"
                    _upd.append("deleted_reason")
            else:
                # Joriy oy: himoya kerak emas — _etm ham xuddi shu to'liq
                # narxni hisoblaydi. Eski user_edit qolib ketgan bo'lsa olib
                # tashlaymiz.
                if (getattr(tm, "deleted_reason", None) or "").startswith("user_edit"):
                    tm.deleted_reason = ""
                    _upd.append("deleted_reason")
            if _upd:
                tm.save(update_fields=_upd)

            # Credit balansni 0 qilamiz — aks holda keyingi oy yaratilganda
            # bekor qilingan pul avtomatik to'lov sifatida qayta yoziladi.
            Enrollment.objects.filter(pk=tm.enrollment_id).update(credit_balance=0)

        # YETIM to'lovlarni ham o'chiramiz: shu oy sanasi bilan yozilgan,
        # hech qanday aktiv allocation'i qolmagan paymentlar. Ular turgani
        # bilan _auto_link_payment_to_tm har sahifa yuklanishida ularni shu
        # oyga qayta bog'lab, "avtomatik to'lov" sifatida tiriltiraveradi.
        _enr_ids = [tm.enrollment_id for tm in tms]
        _orphan_pays = Payment.objects.filter(
            enrollment_id__in=_enr_ids,
            is_deleted=False,
            paid_date__year=month_date.year,
            paid_date__month=month_date.month,
        )
        for _op in _orphan_pays:
            _live_alloc = (
                PaymentAllocation.objects
                .filter(payment=_op, is_deleted=False)
                .aggregate(s=Sum("amount"))["s"] or 0
            )
            if _live_alloc == 0 and int(_op.summa or 0) > 0:
                _op.is_deleted = True
                _op.save(update_fields=["is_deleted"])

    total_debt = get_student_total_debt(student, center)
    return JsonResponse({"ok": True, "freed": total_freed, "total_debt": total_debt})


# ============================================================
# TASK 5: O'quvchi oylik breakdown (AJAX endpoint)
# ============================================================

@login_required
def student_monthly_breakdown(request, student_id):
    """
    O'quvchi barcha enrollments bo'yicha oylik to'lov breakdown'ini JSON formatida qaytaradi.
    """
    center = get_active_center(request)

    user_qs = User.objects.filter(role="student")
    if center:
        _enr_cq = (
            Q(center=center)
            | Q(center__isnull=True, group__center=center)
            | Q(center__isnull=True, student__center=center)
        )
        user_qs = user_qs.filter(
            Q(center=center)
            | Q(pk__in=Enrollment.objects.filter(_enr_cq).values("student_id"))
        )
    student = get_object_or_404(user_qs, id=student_id)

    from django.db.models import Q as _Q
    _center_q_mb = (
        _Q(center=center)
        | _Q(center__isnull=True, group__center=center)
        | _Q(center__isnull=True, student__center=center)
    )
    enrollments = Enrollment.objects.filter(student=student).select_related("group")
    if center:
        enrollments = enrollments.filter(_center_q_mb)

    fee_field = tuition_month_fee_field()

    # Barcha TuitionMonth'larni yig'amiz
    all_tms = (
        TuitionMonth.objects
        .filter(enrollment__in=enrollments, is_deleted=False)
        .select_related("enrollment__group")
        .order_by("month")
        .prefetch_related(
            Prefetch(
                "allocations",
                queryset=PaymentAllocation.objects.filter(
                    payment__is_deleted=False,
                ).select_related("payment"),
                to_attr="active_allocations",
            )
        )
    )

    # Oylar bo'yicha grupplaymiz
    from collections import defaultdict
    # tm_id_map: m_key -> list of (tm_id, enrollment_id) — fee tahrirlash uchun
    month_map = defaultdict(lambda: {"fee": 0, "paid": 0, "payments": [], "enrollments": [], "tm_ids": [], "price_per_lesson": 0, "monthly_lessons": 0})

    for tm in all_tms:
        m_key = tm.month.strftime("%Y-%m")
        fee = int(getattr(tm, fee_field, 0) or 0)
        paid = sum(int(a.amount or 0) for a in tm.active_allocations)
        month_map[m_key]["fee"] += fee
        month_map[m_key]["paid"] += paid
        month_map[m_key]["enrollments"].append(tm.enrollment_id)
        month_map[m_key]["tm_ids"].append(tm.id)
        # Bitta dars narxini enrollment dan olamiz (birinchi TuitionMonth dan)
        if not month_map[m_key]["price_per_lesson"] and tm.enrollment:
            enr = tm.enrollment
            ml = int(getattr(enr.group, "oy_dars_soni", 0) or 0) or int(getattr(enr, "monthly_lessons", 0) or 0) or 12
            kn = int(getattr(enr, "kurs_narhi", 0))
            month_map[m_key]["monthly_lessons"] = ml
            month_map[m_key]["price_per_lesson"] = round(kn / ml) if ml > 0 else 0

        group_name = tm.enrollment.group.nom if tm.enrollment and tm.enrollment.group else ""
        for alloc in tm.active_allocations:
            p = alloc.payment
            p_note = getattr(p, "note", "") or ""
            month_map[m_key]["payments"].append({
                "amount": int(alloc.amount or 0),
                "note": p_note,
                "group": group_name,
            })

    cur_month_key = timezone.localdate().strftime("%Y-%m")

    months_result = []
    total_debt = 0
    for m_key in sorted(month_map.keys()):
        entry = month_map[m_key]
        year, month_num = int(m_key[:4]), int(m_key[5:7])
        month_label = f"{_UZ_MONTHS.get(month_num, month_num)} {year}"
        fee = entry["fee"]
        paid = entry["paid"]
        debt = max(0, fee - paid)

        # Kelajak oy va to'lov yo'q — ko'rsatmayiz.
        # Faqat ortiqcha to'lov (paid>0) bo'lsa kelajak oy ko'rinadi.
        is_future = m_key > cur_month_key
        if is_future and paid == 0:
            continue

        total_debt += debt

        if fee <= 0 and paid > 0:
            status = "paid"
        elif fee <= 0:
            # Kelajak oy uchun fee hali hisoblanmagan — ko'rsatilmaydi (yuqorida skip qilingan)
            # Joriy/o'tgan oy uchun fee=0: to'langan deb hisoblanadi
            status = "paid"
        elif paid <= 0:
            status = "debtor"
        elif paid >= fee:
            status = "paid"
        else:
            status = "partial"

        # tm_id: birinchi TuitionMonth ID (fee tahrirlash uchun; enrollment_id bilan birga ishlatiladi)
        tm_ids = entry.get("tm_ids", [])
        tm_id = tm_ids[0] if tm_ids else None

        months_result.append({
            "month": m_key,
            "month_label": month_label,
            "fee": fee,
            "paid": paid,
            "debt": debt,
            "status": status,
            "payments": entry["payments"],
            "tm_id": tm_id,
            "tm_ids": tm_ids,
            "price_per_lesson": entry.get("price_per_lesson", 0),
            "monthly_lessons": entry.get("monthly_lessons", 0),
        })

    response = JsonResponse({
        "months": months_result,
        "total_debt": total_debt,
    })
    response["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return response


# ─── CourseTemplate CRUD ──────────────────────────────────────────────────────

@login_required
def course_list(request):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if not center:
        return redirect("core:director_boshqaruv")
    if not _can_manage(request.user):
        raise PermissionDenied

    courses = (
        CourseTemplate.objects
        .filter(center=center)
        .select_related("category_obj")
        .order_by("name")
    )
    return render(request, "education/course_list.html", {"courses": courses})


@login_required
def course_create(request):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if not center:
        return redirect("core:director_boshqaruv")
    if not _can_manage(request.user):
        raise PermissionDenied

    from django.db.models import Q
    categories_qs = Category.objects.all().order_by("name")
    if center:
        first_center = Center.objects.order_by("id").first()
        if first_center and center.id == first_center.id:
            categories_qs = categories_qs.filter(Q(center=center) | Q(center__isnull=True))
        else:
            categories_qs = categories_qs.filter(center=center)
    else:
        categories_qs = categories_qs.none()

    categories = list(categories_qs)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        price = request.POST.get("price", "0").replace(" ", "").replace(",", "")
        teacher_percent = request.POST.get("teacher_percent", "40")
        lessons_per_month = request.POST.get("lessons_per_month", "12")
        category_id = request.POST.get("category_obj") or None
        is_active = request.POST.get("is_active") == "on"

        errors = []
        if not name:
            errors.append("Kurs nomi kiritilishi shart.")
        try:
            price = int(price)
            if price <= 0:
                errors.append("Narx musbat son bo'lishi kerak.")
        except (ValueError, TypeError):
            errors.append("Narx noto'g'ri formatda.")

        if not errors:
            cat = None
            if category_id:
                cat = categories_qs.filter(id=category_id).first()
            CourseTemplate.objects.create(
                center=center,
                name=name,
                price=price,
                teacher_percent=int(teacher_percent or 40),
                lessons_per_month=int(lessons_per_month or 12),
                category_obj=cat,
                is_active=is_active,
            )
            messages.success(request, f"✅ '{name}' kursi qo'shildi.")
            return redirect("education:course_list")

        for e in errors:
            messages.error(request, e)

    return render(request, "education/course_form.html", {
        "categories": categories,
        "action": "Yangi kurs",
    })


@login_required
def course_edit(request, pk):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if not _can_manage(request.user):
        raise PermissionDenied
    course = get_object_or_404(CourseTemplate, pk=pk, center=center)
    
    from django.db.models import Q
    categories_qs = Category.objects.all().order_by("name")
    if center:
        first_center = Center.objects.order_by("id").first()
        if first_center and center.id == first_center.id:
            categories_qs = categories_qs.filter(Q(center=center) | Q(center__isnull=True))
        else:
            categories_qs = categories_qs.filter(center=center)
    else:
        categories_qs = categories_qs.none()

    categories = list(categories_qs)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        price = request.POST.get("price", "0").replace(" ", "").replace(",", "")
        teacher_percent = request.POST.get("teacher_percent", "40")
        lessons_per_month = request.POST.get("lessons_per_month", "12")
        category_id = request.POST.get("category_obj") or None
        is_active = request.POST.get("is_active") == "on"

        errors = []
        if not name:
            errors.append("Kurs nomi kiritilishi shart.")
        try:
            price = int(price)
            if price <= 0:
                errors.append("Narx musbat son bo'lishi kerak.")
        except (ValueError, TypeError):
            errors.append("Narx noto'g'ri formatda.")

        if not errors:
            cat = None
            if category_id:
                cat = categories_qs.filter(id=category_id).first()
            course.name = name
            course.price = price
            course.teacher_percent = int(teacher_percent or 40)
            course.lessons_per_month = int(lessons_per_month or 12)
            course.category_obj = cat
            course.is_active = is_active
            course.save()
            messages.success(request, f"✅ '{name}' kursi yangilandi.")
            return redirect("education:course_list")

        for e in errors:
            messages.error(request, e)

    return render(request, "education/course_form.html", {
        "course": course,
        "categories": categories,
        "action": "Kursni tahrirlash",
    })


@login_required
def course_delete(request, pk):
    from core.tenant import get_request_center
    center = get_request_center(request)
    if not _can_manage(request.user):
        raise PermissionDenied
    course = get_object_or_404(CourseTemplate, pk=pk, center=center)
    if request.method == "POST":
        name = course.name
        course.delete()
        messages.success(request, f"'{name}' kursi o'chirildi.")
    return redirect("education:course_list")


def course_price_api(request, pk):
    """AJAX: kurs narxi va parametrlarini qaytaradi — guruh formida narxni avtomatik to'ldirish uchun."""
    from core.tenant import get_request_center
    center = get_request_center(request)
    course = CourseTemplate.objects.filter(pk=pk, center=center, is_active=True).first()
    if not course:
        return JsonResponse({"ok": False}, status=404)
    return JsonResponse({
        "ok": True,
        "price": course.price,
        "teacher_percent": course.teacher_percent,
        "lessons_per_month": course.lessons_per_month,
        "name": course.name,
    })
