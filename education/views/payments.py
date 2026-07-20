"""
Auto-split from education/views.py (phase 7 god-file reduction).
Public API re-exported via education.views package.
"""
from __future__ import annotations

from .common import *  # noqa: F403


@require_POST
@login_required
def create_payment(request):
    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("education:tolovlar_home")

    next_url = request.POST.get("next") or "education:tolovlar_home"

    enrollment_id = request.POST.get("enrollment_id")
    student_id = request.POST.get("student_id")
    payment_scope = (request.POST.get("payment_scope") or "").strip()
    month_str = (request.POST.get("month") or "").strip()
    selected_paid_date = _resolve_paid_date(request.POST.get("paid_date"))
    selected_paid_at = _build_paid_at_for_date(selected_paid_date)

    fallback = selected_paid_date.replace(day=1)
    start_month = parse_month_str(month_str)
    if start_month is None:
        start_month = _get_month_from_next(next_url, fallback)

    # TASK 4: "Qaysi oy uchun?" — manager tomonidan tanlangan oy
    month_for_payment_str = (request.POST.get("month_for_payment") or "").strip()
    month_for_payment = parse_month_str(month_for_payment_str)  # None bo'lsa avtomatik tanlaydi

    cash_amount = int(Decimal(request.POST.get("cash_amount") or "0"))
    card_amount = int(Decimal(request.POST.get("card_amount") or "0"))
    note = (request.POST.get("note") or "").strip()

    if not enrollment_id and not student_id:
        messages.error(request, "ID kelmadi.")
        return redirect(next_url)

    center = get_active_center(request)

    if enrollment_id:
        # all_objects — guruhdan CHIQARILGAN (is_deleted=True) enrollment uchun ham
        # to'lov qabul qilinsin. Aks holda chiqarilgan o'quvchining o'qigan oyi qarzi
        # (mas. iyun) ko'rinadi-yu, to'lash uchun bosilganda 404 chiqardi.
        qs = Enrollment.all_objects.all()
        if center:
            qs = qs.filter(
                Q(center=center)
                | Q(center__isnull=True, group__center=center)
                | Q(center__isnull=True, student__center=center)
            )
        enrollment = get_object_or_404(qs, id=enrollment_id)

        # Double-submit himoyasi: oxirgi 60 sekundda shu enrollment uchun
        # bir xil summa va sana bilan to'lov yozilganmi? Agar ha — qaytarib
        # yozmaymiz (foydalanuvchi tugmani 2 marta bossa ham, faqat 1 to'lov).
        from ..models import Payment as _Payment
        total = cash_amount + card_amount
        recent_dup = _Payment.objects.filter(
            enrollment=enrollment,
            summa=total,
            paid_date=selected_paid_date,
            created_at__gte=timezone.now() - timedelta(seconds=5),
        ).exists()
        if recent_dup:
            messages.warning(
                request,
                "⚠️ Aynan shu summa va sana bilan to'lov yaqinda yozilgan. Takrorlanmasin uchun e'tiborsiz qoldirildi.",
            )
            return redirect(next_url)

        # O'tgan barcha oylar uchun TuitionMonth mavjudligini ta'minlaymiz —
        # shunda allocation past oylarni ham to'g'ri yopadi.
        ensure_all_tuition_months_since_start(enrollment, start_month)

        try:
            with transaction.atomic():
                create_payment_and_allocate(
                    enrollment=enrollment,
                    cash_amount=cash_amount,
                    card_amount_som=card_amount,
                    created_by=request.user,
                    start_month=month_for_payment,  # TASK 4: manager tanlagan oy; None bo'lsa eng eski to'lanmagan oydan
                    paid_at=selected_paid_at,
                    note=note,
                    payment_type=infer_payment_type(cash_amount, card_amount),
                    # Menejer aniq oy tanlagan bo'lsa — butun summa FAQAT shu
                    # oyga (keyingi oyga oshirilmaydi)
                    strict_month=bool(month_for_payment),
                )
            messages.success(request, f"✅ {enrollment.student.get_full_name()} uchun to'lov saqlandi!")
        except Exception as e:
            messages.error(request, f"❌ Xatolik: {e}")
            
    elif student_id:
        # ✅ CONSOLIDATED DISTRIBUTION LOGIC
        user_qs = User.objects.filter(role="student")
        if center:
            _enr_cq = (
                Q(center=center)
                | Q(center__isnull=True, group__center=center)
                | Q(center__isnull=True, student__center=center)
            )
            user_qs = user_qs.filter(
                Q(center=center)
                | Q(pk__in=Enrollment.objects.filter(_enr_cq).values("student_id"))
            )
        student = get_object_or_404(user_qs, id=student_id)
        
        # Faol enrollment'lar
        active_enrollments = Enrollment.objects.filter(
            student=student, is_active=True,
            group__is_archived=False, group__is_deleted=False,
        ).order_by('id')

        # Guruhdan CHIQARILGAN lekin to'lanmagan TuitionMonth bor enrollment'lar
        # ham to'lovga qo'shiladi. Chiqarish ikki xil: is_active=False YOKI
        # is_deleted=True (soft-delete). Ikkalasini ham qamraymiz, aks holda
        # soft-delete qilingan o'quvchining o'qigan oyi qarzini to'lab bo'lmaydi.
        inactive_with_debt_ids = list(
            TuitionMonth.objects
            .filter(
                Q(enrollment__is_active=False) | Q(enrollment__is_deleted=True),
                enrollment__student=student,
                enrollment__group__is_archived=False,
                enrollment__group__is_deleted=False,
                is_deleted=False,
            )
            .values_list("enrollment_id", flat=True)
            .distinct()
        )
        inactive_enrollments = Enrollment.all_objects.filter(
            id__in=inactive_with_debt_ids,
            group__is_archived=False,
            group__is_deleted=False,
        ).order_by('id')

        # Ikkisini birlashtirish — takrorlanmaslik uchun ID bo'yicha
        all_enr_ids = list(dict.fromkeys(
            list(active_enrollments.values_list('id', flat=True)) +
            list(inactive_enrollments.values_list('id', flat=True))
        ))
        enrollments = Enrollment.all_objects.filter(id__in=all_enr_ids).order_by('id')

        if not enrollments.exists():
            messages.error(request, "O'quvchida faol kurslar topilmadi.")
            return redirect(next_url)

        if payment_scope == "teacher_share_only":
            scoped_enrollment_ids = []
            for e in enrollments:
                full_amount = full_course_amount(e)
                effective_amount = effective_student_payable_amount(e)
                teacher_share_amount = int(getattr(e, "oqituvchi_daromadi", 0) or 0)
                if (
                    e.student_payable_amount not in (None, "")
                    and full_amount > effective_amount
                    and effective_amount == teacher_share_amount
                ):
                    tm = ensure_tuition_month(e, start_month)
                    fee = int(getattr(tm, "fee_amount", 0) or 0)
                    paid = int(get_month_paid(e, start_month) or 0)
                    debt = max(0, fee - paid)
                    if debt > 0:
                        scoped_enrollment_ids.append(e.id)

            if scoped_enrollment_ids:
                enrollments = enrollments.filter(id__in=scoped_enrollment_ids).order_by("id")
            else:
                messages.error(request, "Faol o'qituvchi haqqi qarzi topilmadi.")
                return redirect(next_url)
            
        from education.services.tuition import find_earliest_unpaid_month

        # TuitionMonth rekordlarini transaction tashqarisida yaratamiz —
        # bu faqat ensure operatsiyasi, payment bilan bog'liq emas.
        for e in enrollments:
            ensure_all_tuition_months_since_start(e, start_month)

        try:
            with transaction.atomic():
                # One check for the whole payment
                first_group = enrollments[0].group if enrollments else None
                main_payment = Payment.objects.create(
                    student=student,
                    group=first_group,
                    cash_amount=cash_amount,
                    card_amount=card_amount,
                    summa=cash_amount + card_amount,
                    created_by=request.user,
                    paid_date=selected_paid_date,
                    paid_time=selected_paid_at.time().replace(microsecond=0),
                    center=center,
                    note=note,
                    payment_type="mixed" if (cash_amount > 0 and card_amount > 0) else ("card" if card_amount > 0 else "cash")
                )

                remaining_sum = cash_amount + card_amount

                # Har enrollment uchun to'lovni taqsimlash.
                # Agar menejer month_for_payment tanlagan bo'lsa — faqat shu oyga;
                # aks holda eng eski to'lanmagan oydan boshlaymiz.
                for e in enrollments:
                    if remaining_sum <= 0:
                        break

                    if month_for_payment:
                        # Menejer aniq oy tanlagan — faqat shu oyning qarzini olamiz
                        past_tms = TuitionMonth.objects.filter(
                            enrollment=e,
                            month=month_for_payment,
                            is_deleted=False,
                        ).order_by("month")
                    else:
                        # Avtomatik: o'tgan + joriy oylardagi barcha qarz
                        past_tms = TuitionMonth.objects.filter(
                            enrollment=e,
                            month__lte=start_month,
                            is_deleted=False,
                        ).order_by("month")

                    total_debt = sum(
                        max(0, int(getattr(tm, "fee_amount", 0) or 0) - int(get_month_paid(e, tm.month) or 0))
                        for tm in past_tms
                    )

                    if total_debt <= 0:
                        continue

                    take = min(remaining_sum, total_debt)

                    # Tanlangan oydan yoki eng eski to'lanmagan oydan boshlaymiz
                    if month_for_payment:
                        allocation_start = month_for_payment
                        # Agar tanlangan oy yopiq bo'lsa — xato ko'rsatamiz
                        e_center = getattr(e, "center", None) or getattr(e.group, "center", None)
                        if is_month_closed_for_center(e_center, allocation_start):
                            raise ValueError(
                                f"{allocation_start:%B %Y} oyi mahkamlangan. "
                                "Bu oyga to'lov yozish mumkin emas."
                            )
                    else:
                        earliest_tm = find_earliest_unpaid_month(e)
                        allocation_start = earliest_tm.month if earliest_tm else start_month

                    _allocate_amount_forward(
                        enrollment=e,
                        payment=main_payment,
                        amount=take,
                        start_month=allocation_start,
                    )
                    remaining_sum -= take

                # Ortiqcha to'lov (kredit) — faol enrollmentning kelgusi oyiga
                # Guruhdan chiqarilgan (is_active=False) enrollment uchun ortiqcha
                # to'lov kelajak oylariga yozilmasin — bu noto'g'ri qarz hosil qiladi.
                if remaining_sum > 0:
                    overflow_enr = next(
                        (e for e in enrollments if getattr(e, "is_active", False)),
                        None,
                    )
                    if overflow_enr:
                        if month_for_payment:
                            # Menejer aniq oy tanlagan: ortiqcha ham SHU OYGA
                            # yoziladi (keyingi oyga surilmaydi) — hisobotlarda
                            # pul aynan tanlangan oyda ko'rinadi.
                            from education.services.tuition import ensure_tuition_month as _etm_strict
                            _tm_sel = _etm_strict(
                                overflow_enr, month_for_payment,
                                _exclude_payment_id=main_payment.id,
                            )
                            PaymentAllocation.objects.create(
                                center=getattr(main_payment, "center", None)
                                or getattr(overflow_enr, "center", None),
                                payment=main_payment,
                                tuition_month=_tm_sel,
                                amount=remaining_sum,
                            )
                        else:
                            _allocate_amount_forward(
                                enrollment=overflow_enr,
                                payment=main_payment,
                                amount=remaining_sum,
                                start_month=start_month,
                            )
            
            messages.success(request, f"✅ {student.get_full_name()} uchun umumiy to'lov saqlandi!")
        except Exception as e:
            messages.error(request, f"❌ Xatolik: {e}")

    try:
        target_student = None
        if enrollment_id:
            target_student = enrollment.student
        elif student_id:
            target_student = student
        # Faqat tolovlar bo'limidan kelgan va qarz to'liq yopilgan bo'lsa
        # tolovlar_home ga o'tamiz. Qarzdorlar bo'limidan kelgan bo'lsa —
        # next_url o'zgarmaydi, foydalanuvchi shu sahifada qoladi.
        _from_qarzdorlar = "qarzdorlar" in (next_url or "")
        if (
            not _from_qarzdorlar
            and target_student
            and get_student_total_debt(target_student, center) <= 0
        ):
            next_url = reverse("education:tolovlar_home")
    except Exception:
        pass

    return redirect(next_url)



@require_POST
@login_required
def payment_update(request, payment_id: int):
    if not user_can_manage_payments(request.user):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # from core.tenant import get_request_center
    center = get_active_center(request)
    qs = Payment.objects.select_related("enrollment")
    if center:
        qs = qs.filter(center=center)

    p = get_object_or_404(qs, id=payment_id)
    enrollment = getattr(p, "enrollment", None)
    if not enrollment:
        return JsonResponse({"ok": False, "error": "enrollment_not_found"}, status=400)

    old_total = int(getattr(p, "summa", 0) or 0)

    cash_raw = request.POST.get("cash_amount")
    card_raw = request.POST.get("card_amount")
    paid_date_raw = request.POST.get("paid_date")

    if cash_raw is None and card_raw is None and paid_date_raw is None:
        note = request.POST.get("note")
        if note is not None:
            p.note = note.strip()   # bo'sh bo'lsa ham "" bo'lib saqlanadi
            p.save(update_fields=["note"])
        return JsonResponse({"ok": True, "payment_id": p.id, "note": p.note or ""})

    from decimal import Decimal, InvalidOperation
    try:
        cash_amount = int(Decimal((request.POST.get("cash_amount") or "0").strip()))
        card_amount = int(Decimal((request.POST.get("card_amount") or "0").strip()))
    except (InvalidOperation, ValueError):
        return JsonResponse({"ok": False, "error": "summa_notogri"}, status=400)

    if cash_amount < 0 or card_amount < 0:
        return JsonResponse({"ok": False, "error": "summa_manfiy_bolmaydi"}, status=400)

    new_total = cash_amount + card_amount
    if new_total < 0:
        return JsonResponse({"ok": False, "error": "summa_manfiy_bolmaydi"}, status=400)

    month_str = (request.POST.get("month") or "").strip()
    start_month = parse_month_str(month_str) if month_str else None

    # Update metadata
    note = request.POST.get("note")
    if note is not None:
        p.note = note.strip()

    paid_date_str = request.POST.get("paid_date")
    if paid_date_str:
        try:
            new_paid_date = parse_date(paid_date_str)
            if new_paid_date:
                p.paid_date = new_paid_date
                # Agar alohida 'month' yuborilmagan bo'lsa, paid_date oyini start_month sifatida ishlatamiz
                if not start_month:
                    start_month = new_paid_date.replace(day=1)
        except Exception:
            pass

    p.save()

    try:
        update_payment_and_reallocate(
            payment=p,
            cash_amount=cash_amount,
            card_amount_som=card_amount,
            start_month=start_month,
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    return JsonResponse({
        "ok": True,
        "payment_id": p.id,
        "old_total": old_total,
        "new_total": new_total,
        "delta": new_total - old_total,
        "start_month": (start_month or timezone.localdate().replace(day=1)).strftime("%Y-%m"),
    })



@login_required
def payment_history_enrollment(request, enrollment_id: int):
    if not user_can_manage_payments(request.user):
        return JsonResponse({"error": "forbidden"}, status=403)

    today = timezone.localdate()
    cur_month = today.replace(day=1)
    
    month_str = request.GET.get("month", "")
    selected_month = parse_month_str(month_str) or cur_month

    center = get_active_center(request)
    qs = Enrollment.objects.select_related("student", "group")
    if center:
        qs = qs.filter(center=center)

    enrollment = get_object_or_404(qs, id=enrollment_id)
    fee_field = tuition_month_fee_field()

    # 1. Summary Data (up to today)
    ensure_tuition_month(enrollment, cur_month)
    tms_summary = TuitionMonth.objects.filter(enrollment=enrollment, month__lte=cur_month)
    agg_summary = tms_summary.aggregate(
        total_fee=Coalesce(Sum(fee_field), 0),
        total_paid=Coalesce(Sum("allocations__amount"), 0)
    )
    total_fee_needed = agg_summary["total_fee"] or 0
    total_paid_so_far = agg_summary["total_paid"] or 0
    overall_debt = total_fee_needed - total_paid_so_far

    # 2. Monthly Breakdown (All months including future if they have allocations)
    breakdown = []
    # Get all tuition months for this enrollment
    all_tms = TuitionMonth.objects.filter(enrollment=enrollment).order_by("month")
    
    for tm in all_tms:
        tm_fee = getattr(tm, fee_field, 0) or 0
        tm_paid = tm.allocations.aggregate(s=Sum("amount"))["s"] or 0
        tm_debt = max(0, tm_fee - tm_paid)
        
        breakdown.append({
            "month": tm.month.strftime("%Y-%m"),
            "fee": tm_fee,
            "paid": tm_paid,
            "debt": tm_debt,
            "is_future": tm.month > cur_month
        })

    # 3. Specific payments history
    payments_qs = Payment.objects.filter(enrollment=enrollment).order_by("-id")
    payments = []
    for p in payments_qs:
        allocations = []
        for a in p.allocations.select_related("tuition_month").all():
            allocations.append({
                "month": a.tuition_month.month.strftime("%Y-%m"),
                "amount": int(a.amount or 0),
            })

        # paid_at robust check
        paid_at_dt = getattr(p, "paid_at", None) or p.created_at
        if not paid_at_dt:
            sana = getattr(p, "sana", None)
            vaqt = getattr(p, "vaqt", None)
            if sana:
                dt = datetime.combine(sana, vaqt or datetime.min.time())
                paid_at_dt = timezone.make_aware(dt)
            else:
                paid_at_dt = timezone.now()

        cash = int(getattr(p, "cash_amount", 0) or 0)
        card = int(getattr(p, "card_amount_som", 0) or getattr(p, "card_amount", 0) or 0)
        total = int(getattr(p, "summa", 0) or (cash + card))

        payments.append({
            "id": p.id,
            "paid_at": timezone.localtime(paid_at_dt).strftime("%d.%m.%Y %H:%M"),
            "cash": cash,
            "card": card,
            "total": total,
            "allocations": allocations,
            "receipt_url": reverse("education:payment_receipt_pdf", args=[p.id]),
        })

    return JsonResponse({
        "student": enrollment.student.get_full_name(),
        "group": enrollment.group.nom,
        "monthly_fee": effective_student_payable_amount(enrollment),
        "total_fee_needed": total_fee_needed,
        "total_paid_so_far": total_paid_so_far,
        "overall_debt": overall_debt,
        "breakdown": breakdown,
        "payments": payments,
    })



@login_required
def payment_receipt_pdf(request, payment_id: int):
    if not user_can_manage_payments(request.user):
        return HttpResponse("Forbidden", status=403)

    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Payment.objects.select_related("enrollment__student", "enrollment__group")
    if center:
        qs = qs.filter(center=center)

    p = get_object_or_404(
        qs,
        id=payment_id
    )

    enrollment = getattr(p, "enrollment", None)
    student = getattr(enrollment, "student", None)
    group = getattr(enrollment, "group", None)

    cash = int(getattr(p, "cash_amount", 0) or 0)
    card_som = int(getattr(p, "card_amount_som", 0) or getattr(p, "card_amount", 0) or 0)
    total = int(getattr(p, "summa", 0) or (cash + card_som))

    # ✅ Sana/vaqt (soat)
    paid_at = getattr(p, "paid_at", None)
    if paid_at:
        paid_at = timezone.localtime(paid_at)
    else:
        sana = getattr(p, "sana", None)
        vaqt = getattr(p, "vaqt", None)
        if sana and vaqt:
            try:
                dt = datetime.combine(sana, vaqt)
                paid_at = timezone.localtime(timezone.make_aware(dt))
            except Exception:
                paid_at = timezone.localtime(timezone.now())
        else:
            paid_at = timezone.localtime(timezone.now())

    dt_text = paid_at.strftime("%d.%m.%Y %H:%M")

    # ==== Tanlangan oy bo'yicha qarz hisoblash (month=YYYY-MM) ====
    fee_field = "fee_amount"

    month_qs = request.GET.get("month")  # masalan: 2026-01
    month_date = _parse_month_str(month_qs)

    fee_for_month = None
    paid_for_month = None
    debt_for_month = None

    if enrollment and month_date:
        tm, _ = TuitionMonth.objects.get_or_create(
            enrollment=enrollment,
            month=month_date,
            defaults={fee_field: _get_fee_amount(enrollment)}
        )

        # agar fee 0 bo'lib qolgan bo'lsa, fallback bilan yangilab qo'yamiz
        cur_fee = int(getattr(tm, fee_field, 0) or 0)
        if cur_fee <= 0:
            new_fee = _get_fee_amount(enrollment)
            if new_fee > 0:
                setattr(tm, fee_field, int(new_fee))
                tm.save(update_fields=[fee_field])
                cur_fee = new_fee

        fee_for_month = int(cur_fee or 0)
        paid_for_month = int(
            PaymentAllocation.objects.filter(tuition_month=tm).aggregate(s=Sum("amount"))["s"] or 0
        )
        debt_for_month = max(0, fee_for_month - paid_for_month)

    # allocations (shu payment bo'yicha)
    alloc_mgr = getattr(p, "allocations", None)
    allocations = list(alloc_mgr.select_related("tuition_month").all()) if alloc_mgr is not None else []

    # To'lov turi
    if cash > 0 and card_som > 0:
        pay_type = "Aralash (Naqd + Karta)"
    elif card_som > 0:
        pay_type = "Kartaga o'tkazma"
    else:
        pay_type = "Naqd to'lov"

    student_name = f"{getattr(student, 'ism', '')} {getattr(student, 'familya', '')}".strip() or "-"
    group_name = getattr(group, "nom", "") or "-"

    student_name = _ellipsis(student_name, 38)
    group_name = _ellipsis(group_name, 38)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    W, H = A4

    # =========================
    #  CHEK KARTA (markazda)
    # =========================
    card_w = 175 * mm
    card_h = 240 * mm
    x = (W - card_w) / 2
    y = (H - card_h) / 2

    # oq fon
    c.setFillColor(colors.white)
    c.rect(0, 0, W, H, stroke=0, fill=1)

    # shadow (alpha bo'lsa ishlatamiz, bo'lmasa ham muammo yo'q)
    try:
        c.setFillAlpha(0.08)
    except Exception:
        pass
    c.setFillColor(colors.black)
    c.roundRect(x + 2*mm, y - 2*mm, card_w, card_h, 14, stroke=0, fill=1)
    try:
        c.setFillAlpha(1)
    except Exception:
        pass

    # card body
    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.setLineWidth(1)
    c.roundRect(x, y, card_w, card_h, 14, stroke=1, fill=1)

    # Header (yorqin yashil)
    header_h = 34 * mm
    c.setFillColor(colors.HexColor("#16A34A"))
    c.roundRect(x, y + card_h - header_h, card_w, header_h, 14, stroke=0, fill=1)

    # Header text
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x + 12*mm, y + card_h - 18*mm, "ChaqmoqApp")
    c.setFont("Helvetica", 10)
    c.drawRightString(x + card_w - 12*mm, y + card_h - 18*mm, "TO'LOV CHEKI")

    # Big amount
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 32)
    c.drawCentredString(x + card_w/2, y + card_h - 60*mm, f"{_fmt(total)} so'm")

    # Success badge
    badge_w = 70 * mm
    badge_h = 10 * mm
    bx = x + (card_w - badge_w) / 2
    by = y + card_h - 74*mm
    c.setFillColor(colors.HexColor("#22C55E"))
    c.roundRect(bx, by, badge_w, badge_h, 6, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(x + card_w/2, by + 3.2*mm, "Muvaffaqiyatli")

    # Divider
    c.setStrokeColor(colors.HexColor("#E5E7EB"))
    c.setLineWidth(1)
    c.line(x + 12*mm, y + card_h - 82*mm, x + card_w - 12*mm, y + card_h - 82*mm)

    # Key-value row helper
    def row(label: str, value: str, yy: float, value_color=colors.HexColor("#111827")):
        c.setFillColor(colors.HexColor("#6B7280"))
        c.setFont("Helvetica", 10)
        c.drawString(x + 12*mm, yy, label)

        c.setFillColor(value_color)
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(x + card_w - 12*mm, yy, value)

        c.setStrokeColor(colors.HexColor("#F1F5F9"))
        c.setLineWidth(1)
        c.line(x + 12*mm, yy - 4.5*mm, x + card_w - 12*mm, yy - 4.5*mm)

    # Row spacing (siqilganroq, hammasi sig'sin)
    GAP = 10 * mm

    teacher = getattr(group, "oqituvchi", None)
    teacher_name = teacher.get_full_name() if teacher else "-"
    teacher_name = _ellipsis(teacher_name, 38)

    yy = y + card_h - 98*mm
    row("Tranzaksiya turi:", pay_type, yy); yy -= GAP
    row("O'quvchi:", student_name, yy); yy -= GAP
    row("Guruh:", group_name, yy); yy -= GAP
    row("O'qituvchi:", teacher_name, yy); yy -= GAP
    row("Naqd:", f"{_fmt(cash)} so'm", yy); yy -= GAP
    row("Karta:", f"{_fmt(card_som)} so'm", yy); yy -= GAP
    # row("Chek ID:", f"#{p.id}", yy); yy -= GAP

    # ✅ Oylik narx + Shu oy to'langan + Qarz
    if fee_for_month is not None:
        row("Oylik narx:", f"{_fmt(fee_for_month)} so'm", yy); yy -= GAP
        row("Shu oy to'langan:", f"{_fmt(paid_for_month)} so'm", yy); yy -= GAP
        debt_color = colors.HexColor("#EF4444") if (debt_for_month or 0) > 0 else colors.HexColor("#16A34A")
        row("Qarz (qoldiq):", f"{_fmt(debt_for_month)} so'm", yy, value_color=debt_color); yy -= GAP

    row("Sana:", dt_text, yy); yy -= 12*mm

    # Allocations title
    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x + 12*mm, yy, "Taqsimot (qaysi oylarga tushdi):")
    yy -= 8*mm

    # Allocation list (qolgan joyga qarab max line)
    c.setFont("Helvetica", 10)
    bottom_limit = y + 18*mm  # footer usti
    line_h = 7 * mm
    max_lines = max(1, int((yy - bottom_limit) / line_h) - 1)

    if not allocations:
        c.setFillColor(colors.HexColor("#6B7280"))
        c.drawString(x + 14*mm, yy, "— Allocation topilmadi")
        yy -= line_h
    else:
        c.setFillColor(colors.HexColor("#0F172A"))
        for a in allocations[:max_lines]:
            tm = getattr(a, "tuition_month", None)
            m = getattr(tm, "month", None)
            enr = getattr(tm, "enrollment", None)
            g_nom = getattr(getattr(enr, "group", None), "nom", "")[:15]
            
            m_txt = m.strftime("%Y-%m") if m else "—"
            amt_txt = _fmt(int(getattr(a, "amount", 0) or 0))
            
            prefix = f"• {g_nom} ({m_txt})" if g_nom else f"• {m_txt}"
            c.drawString(x + 14*mm, yy, f"{prefix} — {amt_txt} so'm")
            yy -= line_h

        if len(allocations) > max_lines:
            c.setFillColor(colors.HexColor("#6B7280"))
            c.drawString(x + 14*mm, yy, f"... yana {len(allocations) - max_lines} ta satr bor")
            yy -= line_h

    # Footer
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.setFont("Helvetica", 9)
    c.drawCentredString(x + card_w/2, y + 10*mm, "ChaqmoqApp • To'lov nazorati tizimi")

    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")

    # ✅ bosganda darrov yuklab olsin
    resp["Content-Disposition"] = f'attachment; filename="proskill{p.id}.pdf"'
    return resp



@login_required
@require_http_methods(["GET"])
@require_feature("finance")
def payment_export_xlsx(request):
    if not user_can_manage_payments(request.user):
        return HttpResponseForbidden("Ruxsat yo'q.")

    try:
        dashboard = _get_payment_dashboard_data(request)
    except PermissionDenied:
        return HttpResponseForbidden("Markaz biriktirilmagan")

    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    grouped_rows = dashboard["page_rows"]
    filtered_payments = dashboard["filtered_payments"]

    def _fill(color: str):
        return PatternFill("solid", fgColor=color)

    def _border():
        side = Side(style="thin", color="CBD5E1")
        return Border(left=side, right=side, top=side, bottom=side)

    def _money(cell):
        cell.number_format = '#,##0'

    def _auto_width(ws, *, min_width: int = 12, max_width: int = 34):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except Exception:
                    continue
            ws.column_dimensions[column_letter].width = max(min_width, min(max_length + 2, max_width))

    def _full_name(user):
        if not user:
            return "—"
        return user.get_full_name() or user.email or f"{getattr(user, 'ism', '')} {getattr(user, 'familya', '')}".strip() or "—"

    group_map = {str(group.id): group.nom for group in dashboard["groups"]}
    teacher_map = {str(teacher.id): _full_name(teacher) for teacher in dashboard["teachers"]}
    course_map = {str(course.id): course.name for course in dashboard["courses"]}
    staff_map = {str(staff.id): _full_name(staff) for staff in dashboard["staffs"]}
    month_map = {str(mid): mname for mid, mname in dashboard["uz_months"]}

    filter_rows = [
        ("Qidiruv", dashboard["q"] or "Barchasi"),
        ("Sanadan", dashboard["date_from"] or "—"),
        ("Sanagacha", dashboard["date_to"] or "—"),
        ("Guruh", group_map.get(dashboard["sel_group"], "Barchasi")),
        ("O'qituvchi", teacher_map.get(dashboard["sel_teacher"], "Barchasi")),
        ("Yo'nalish", course_map.get(dashboard["sel_course"], "Barchasi")),
        ("Xodim", staff_map.get(dashboard["sel_staff"], "Barchasi")),
        ("To'lov turi", dict(Payment.PAYMENT_TYPES).get(dashboard["sel_type"], "Barchasi")),
        ("Qaysi oy", month_map.get(dashboard["sel_month"], "Barcha oylar")),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umumiy ro'yxat"
    ws.sheet_view.showGridLines = False

    title_fill = _fill("0F172A")
    header_fill = _fill("2563EB")
    accent_fill = _fill("E0F2FE")
    soft_fill = _fill("F8FAFC")
    money_fill = _fill("ECFDF5")
    white_font = Font(color="FFFFFF", bold=True, size=12)
    dark_font = Font(color="0F172A", size=11)
    strong_font = Font(color="0F172A", bold=True, size=11)
    money_font = Font(color="047857", bold=True, size=11)

    ws.merge_cells("A1:K1")
    ws["A1"] = "To'lovlar eksporti"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Davr: {dashboard['selected_period_label']}"
    ws["A2"].fill = _fill("1E3A8A")
    ws["A2"].font = Font(color="DBEAFE", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    metrics = [
        ("Eksport davri", dashboard["selected_period_label"]),
        ("Noyob o'quvchi", f"{dashboard['unique_payers_count']} ta"),
        ("To'lov yozuvlari", f"{dashboard['payment_record_count']} ta"),
        ("Filter daromad", dashboard["filtered_income"]),
        ("Umumiy daromad", dashboard["total_income"]),
    ]
    ws["A4"] = "Ko'rsatkich"
    ws["B4"] = "Qiymat"
    for cell in ("A4", "B4"):
        ws[cell].fill = header_fill
        ws[cell].font = white_font
        ws[cell].border = _border()
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    metric_start = 5
    for idx, (label, value) in enumerate(metrics, start=metric_start):
        ws.cell(row=idx, column=1, value=label)
        value_cell = ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).fill = soft_fill
        value_cell.fill = money_fill if isinstance(value, int) else accent_fill
        ws.cell(row=idx, column=1).font = strong_font
        value_cell.font = money_font if isinstance(value, int) else strong_font
        ws.cell(row=idx, column=1).border = _border()
        value_cell.border = _border()
        if isinstance(value, int):
            _money(value_cell)

    ws["D4"] = "Aktiv filtr"
    ws["E4"] = "Qiymat"
    for cell in ("D4", "E4"):
        ws[cell].fill = header_fill
        ws[cell].font = white_font
        ws[cell].border = _border()
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    for idx, (label, value) in enumerate(filter_rows, start=5):
        ws.cell(row=idx, column=4, value=label)
        ws.cell(row=idx, column=5, value=value)
        ws.cell(row=idx, column=4).fill = soft_fill
        ws.cell(row=idx, column=5).fill = accent_fill
        ws.cell(row=idx, column=4).font = strong_font
        ws.cell(row=idx, column=5).font = dark_font
        ws.cell(row=idx, column=4).border = _border()
        ws.cell(row=idx, column=5).border = _border()

    table_row = 16
    summary_headers = [
        "So'nggi sana",
        "O'quvchi",
        "Telefon",
        "Guruhlar",
        "Yo'nalish",
        "Oylar",
        "To'lovlar soni",
        "Turlar",
        "Xodimlar",
        "Jami summa",
        "Oxirgi izoh",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=table_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    current_row = table_row + 1
    for row in grouped_rows:
        values = [
            row["latest_paid_date"].strftime("%d.%m.%Y") if row.get("latest_paid_date") else "—",
            _full_name(row.get("student")),
            getattr(row.get("student"), "telefon1", "") or getattr(row.get("student"), "telefon2", "") or "—",
            row.get("group_summary_title") or "—",
            row.get("category_summary") or "—",
            row.get("month_summary_title") or "—",
            row.get("payment_count") or 0,
            ", ".join(item["label"] for item in row.get("type_entries", [])) or "—",
            row.get("staff_summary") or "—",
            row.get("total_sum") or 0,
            row.get("latest_note") or "—",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = soft_fill if current_row % 2 == 0 else _fill("FFFFFF")
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 10:
                cell.font = money_font
                _money(cell)
            else:
                cell.font = dark_font
        current_row += 1

    ws.freeze_panes = "A17"
    _auto_width(ws, max_width=38)

    detail_ws = wb.create_sheet("To'lov yozuvlari")
    detail_ws.sheet_view.showGridLines = False
    detail_ws.merge_cells("A1:M1")
    detail_ws["A1"] = "To'lov yozuvlari"
    detail_ws["A1"].fill = title_fill
    detail_ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    detail_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    detail_ws.merge_cells("A2:M2")
    detail_ws["A2"] = f"Filterlangan yozuvlar soni: {dashboard['payment_record_count']} ta"
    detail_ws["A2"].fill = _fill("1E293B")
    detail_ws["A2"].font = Font(color="E2E8F0", bold=True, size=10)
    detail_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    detail_headers = [
        "Sana",
        "Vaqt",
        "O'quvchi",
        "Telefon",
        "Guruh",
        "Yo'nalish",
        "Oy / taqsimot",
        "Naqd",
        "Karta",
        "Jami",
        "Tur",
        "Xodim",
        "Izoh",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail_ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    detail_row = 5
    for payment in filtered_payments:
        allocations = getattr(payment, "prefetched_allocations", []) or []
        allocation_labels = []
        for allocation in allocations:
            tuition_month = getattr(allocation, "tuition_month", None)
            month_value = getattr(tuition_month, "month", None)
            if month_value:
                month_label = f"{UZ_MONTH_NAMES.get(month_value.month, month_value.strftime('%B'))} {month_value.year}"
            else:
                month_label = "—"
            allocation_labels.append(f"{month_label}: {int(allocation.amount or 0):,} so'm".replace(",", " "))
        if not allocation_labels and payment.paid_date:
            fallback = payment.paid_date.replace(day=1)
            allocation_labels.append(f"{UZ_MONTH_NAMES.get(fallback.month, fallback.strftime('%B'))} {fallback.year}")

        detail_values = [
            payment.paid_date.strftime("%d.%m.%Y") if payment.paid_date else "—",
            payment.paid_time.strftime("%H:%M") if payment.paid_time else "—",
            _full_name(payment.student),
            getattr(payment.student, "telefon1", "") or getattr(payment.student, "telefon2", "") or "—",
            payment.group.nom if payment.group else "—",
            getattr(getattr(payment.group, "category_obj", None), "name", "") or "—",
            ", ".join(allocation_labels) or "—",
            int(payment.cash_amount or 0),
            int(payment.card_amount or 0),
            int(payment.summa or 0),
            payment.get_payment_type_display(),
            _full_name(payment.created_by),
            payment.note or "—",
        ]
        for col_idx, value in enumerate(detail_values, start=1):
            cell = detail_ws.cell(row=detail_row, column=col_idx, value=value)
            cell.fill = soft_fill if detail_row % 2 == 1 else _fill("FFFFFF")
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx in (8, 9, 10):
                cell.font = money_font
                _money(cell)
            else:
                cell.font = dark_font
        detail_row += 1

    detail_ws.freeze_panes = "A5"
    _auto_width(detail_ws, max_width=42)

    filename = f"tolovlar_export_{dashboard['selected_from'].isoformat()}_{dashboard['selected_to'].isoformat()}.xlsx"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



@login_required
def tolovlar_home(request):
    if not user_can_manage_payments(request.user):
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    try:
        dashboard = _get_payment_dashboard_data(request)
    except PermissionDenied:
        return HttpResponseForbidden("Markaz biriktirilmagan")

    allowed_page_sizes = (10, 20, 50, 100)
    try:
        per_page = int((request.GET.get("per_page") or request.GET.get("page_size") or 10))
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in allowed_page_sizes:
        per_page = 10

    paginator = Paginator(dashboard["page_rows"], per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    dashboard.update(
        {
            "page_obj": page_obj,
            "total_count": paginator.count,
            "per_page": per_page,
            "page_size": per_page,
            "page_size_options": allowed_page_sizes,
            "allowed_page_sizes": allowed_page_sizes,
            "is_paginated": page_obj.has_other_pages(),
        }
    )
    try:
        from store.models import PaymentMethod as _PM
        from store.views import _ensure_default_payment_methods as _seed_pm
        from core.tenant import get_request_center as _grc
        _center = _grc(request)
        if _center:
            _seed_pm(_center)
            dashboard["dynamic_payment_methods"] = list(
                _PM.objects.filter(center=_center, is_active=True).order_by('nom')
            )
        else:
            dashboard["dynamic_payment_methods"] = []
    except Exception:
        dashboard["dynamic_payment_methods"] = []
    return render(request, "education/tolovlar_list.html", dashboard)



@login_required
def get_payment_details(request):
    """
    Returns full transaction history for a specific TuitionMonth as JSON.
    """
    tuition_month_id = request.GET.get('tuition_month_id')
    student_id = request.GET.get('student_id')
    group_id = request.GET.get('group_id')
    
    if tuition_month_id:
        allocs = PaymentAllocation.objects.filter(
            tuition_month_id=tuition_month_id
        ).select_related('payment', 'payment__student', 'payment__group', 'payment__created_by')
    elif student_id and group_id:
        allocs = PaymentAllocation.objects.filter(
            payment__student_id=student_id,
            payment__group_id=group_id
        ).select_related('payment', 'payment__student', 'payment__group', 'payment__created_by')
    else:
        return JsonResponse({'ok': False, 'error': 'Missing identifiers'}, status=400)

    if not allocs.exists():
        return JsonResponse({'ok': True, 'payments': [], 'total_sum': 0})

    first = allocs.first()
    data = []
    total = 0
    for a in allocs.order_by('-payment__paid_date', '-id'):
        total += a.amount
        data.append({
            'id': a.payment.id,
            'amount': a.amount,
            'cash_amount': a.payment.cash_amount,
            'card_amount_som': a.payment.card_amount_som,
            'date': a.payment.paid_date.strftime("%d.%m.%Y"),
            'raw_date': a.payment.paid_date.strftime("%Y-%m-%d"),
            'time': a.payment.paid_time.strftime("%H:%M") if a.payment.paid_time else "--:--",
            'method': a.payment.get_payment_type_display(),
            'staff': a.payment.created_by.get_full_name() if a.payment.created_by else '—',
            'note': a.payment.note or ''
        })

    return JsonResponse({
        'ok': True,
        'student_name': first.payment.student.get_full_name(),
        'group_name': first.payment.group.nom,
        'teacher_name': first.payment.group.oqituvchi.get_full_name() if first.payment.group.oqituvchi else "—",
        'total_sum': total,
        'payments': data
    })



@login_required
def student_payments_pdf(request):
    """
    Generates a professional printable HTML summary of payments for a student.
    """
    student_id = request.GET.get('student_id')
    group_id = request.GET.get('group_id')

    if not student_id or not group_id:
        return HttpResponse("Missing student_id or group_id", status=400)

    from core.tenant import get_tenant_object_or_404

    center = get_active_center(request)
    # Tenant-scoped lookup: boshqa markaz o'quvchi/guruhini ko'rsatmaydi (IDOR).
    student = get_tenant_object_or_404(User, request, id=student_id, role="student")
    group = get_tenant_object_or_404(Group, request, id=group_id)
    enrollment = Enrollment.objects.filter(student=student, group=group).first()
    if enrollment is not None and center and getattr(enrollment, "center_id", None) not in (None, center.id):
        raise Http404("Enrollment topilmadi")

    if not enrollment:
        return HttpResponse("Enrollment topilmadi", status=404)

    payments_qs = Payment.objects.filter(enrollment=enrollment).select_related('created_by').order_by('paid_date', 'paid_time')
    
    total_paid = payments_qs.aggregate(s=Sum('summa'))['s'] or 0
    
    # Calculate total expected fee from TuitionMonths
    tms = TuitionMonth.objects.filter(enrollment=enrollment).order_by('month')
    total_expected = tms.aggregate(s=Sum('fee_amount'))['s'] or 0
    
    # Balance calculations
    remaining_debt = max(0, total_expected - total_paid)
    overpayment = max(0, total_paid - total_expected)

    # Monthly breakdown
    monthly_data = []
    for tm in tms:
        paid_amount = tm.allocations.aggregate(s=Sum('amount'))['s'] or 0
        monthly_data.append({
            'month': tm.month,
            'fee': tm.fee_amount,
            'paid': paid_amount,
            'debt': max(0, tm.fee_amount - paid_amount),
            'overpaid': max(0, paid_amount - tm.fee_amount),
        })

    context = {
        'center': center,
        'student': student,
        'group': group,
        'enrollment': enrollment,
        'payments': payments_qs,
        'total_paid': total_paid,
        'total_expected': total_expected,
        'remaining_debt': remaining_debt,
        'overpayment': overpayment,
        'monthly_data': monthly_data,
        'print_time': timezone.now(),
        'staff_name': request.user.get_full_name() or request.user.email,
    }
    
    return render(request, "education/receipt.html", context)



@require_POST
@login_required
def payment_delete(request, payment_id):
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not user_can_manage_payments(request.user):
        error_message = "Ruxsat yo'q."
        messages.error(request, error_message)
        if is_ajax:
            return JsonResponse({"ok": False, "error": error_message}, status=403)
        return redirect("education:tolovlar_home")

    next_url = request.POST.get("next") or request.GET.get("next") or "education:tolovlar_home"

    center = get_active_center(request)
    qs = Payment.objects.all()
    if center:
        qs = qs.filter(center=center)
    
    payment = get_object_or_404(qs, id=payment_id)
    
    try:
        with transaction.atomic():
            # ✅ Soft delete allocations as well
            payment.allocations.all().delete(deleted_by=request.user)
            payment.delete(deleted_by=request.user)
                 
        messages.success(request, "✅ To'lov o'chirildi. O'quvchi qarzdorlar ro'yxatiga qaytadi.")
    except Exception as e:
        messages.error(request, f"❌ Xatolik: {e}")
        if is_ajax:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)
    
    if is_ajax:
        return JsonResponse({"ok": True, "redirect_url": next_url})
    return redirect(next_url)



@login_required
def student_groups_api(request, student_id):
    """Return active group names and total debt for a student (payment modal)."""
    from django.db.models import Sum
    center = get_active_center(request)
    qs = Enrollment.objects.filter(student_id=student_id, is_active=True, group__is_archived=False, group__is_deleted=False).select_related('group')
    if center:
        qs = qs.filter(center=center)
    groups = [e.group.nom for e in qs if e.group and not getattr(e.group, "is_archived", False)]

    # Calculate total debt: sum of all TuitionMonth fees minus all payments
    try:
        enr_ids = [e.id for e in qs]
        total_fee = TuitionMonth.objects.filter(
            enrollment_id__in=enr_ids, is_deleted=False
        ).aggregate(s=Sum('fee_amount'))['s'] or 0
        total_paid = Payment.objects.filter(
            student_id=student_id, is_deleted=False
        ).aggregate(s=Sum('summa'))['s'] or 0
        debt = max(0, int(total_fee) - int(total_paid))
    except Exception:
        debt = 0

    return JsonResponse({
        "groups": " + ".join(groups) if groups else "",
        "debt": debt,
    })



@login_required
def students_with_debt_api(request):
    """Return students who have debt > 0 for the payment modal (JSON)."""
    from django.db.models import Sum
    center = get_active_center(request)
    if not center:
        return JsonResponse({"students": []})

    # 1. Active enrollments for this center
    enr_qs = Enrollment.objects.filter(
        is_active=True, center=center, group__is_archived=False, group__is_deleted=False,
    ).select_related('student', 'group')

    # 2. Total fees per student (sum TuitionMonth.fee_amount)
    enr_ids = list(enr_qs.values_list('id', flat=True))
    fee_rows = (
        TuitionMonth.objects.filter(enrollment_id__in=enr_ids, is_deleted=False)
        .values('enrollment__student_id')
        .annotate(total=Sum('fee_amount'))
    )
    student_fees = {r['enrollment__student_id']: int(r['total'] or 0) for r in fee_rows}

    # 3. Total paid per student (sum Payment.summa)
    student_ids = list(student_fees.keys())
    paid_rows = (
        Payment.objects.filter(student_id__in=student_ids, center=center, is_deleted=False)
        .values('student_id')
        .annotate(total=Sum('summa'))
    )
    student_paid = {r['student_id']: int(r['total'] or 0) for r in paid_rows}

    # 4. Group names per student
    student_groups = {}
    for e in enr_qs:
        if e.group:
            student_groups.setdefault(e.student_id, []).append(e.group.nom)

    # 5. Build result — only students with debt > 0
    students = []
    seen = set()
    for enr in enr_qs:
        sid = enr.student_id
        if sid in seen or sid not in student_fees:
            continue
        seen.add(sid)
        fee  = student_fees.get(sid, 0)
        paid = student_paid.get(sid, 0)
        debt = max(0, fee - paid)
        if debt <= 0:
            continue
        u = enr.student
        students.append({
            'id':     sid,
            'name':   f"{u.ism} {u.familya}".strip(),
            'phone':  getattr(u, 'telefon1', '') or '',
            'groups': ' + '.join(student_groups.get(sid, [])),
            'debt':   debt,
        })

    students.sort(key=lambda x: x['name'])
    return JsonResponse({'students': students})



@login_required
def payment_history(request, student_id):
    """
    O'quvchining (barcha kurslari bo'yicha) to'lov tarixini va joriy oy holatini xisoblaydi.
    """
    month_str = request.GET.get("month")
    selected_month = parse_month_str(month_str) or first_day_of_current_month()
    
    # 1. Barcha enrollments
    enrs = Enrollment.objects.filter(student_id=student_id, is_active=True)
    
    total_fee = 0
    total_paid_this_month = 0
    
    for e in enrs:
        tm = ensure_tuition_month(e, selected_month)
        total_fee += int(getattr(tm, "fee_amount", 0) or 0)
        total_paid_this_month += int(get_month_paid(e, selected_month) or 0)
    
    total_qoldiq = max(0, total_fee - total_paid_this_month)
    
    # 2. Barcha to'lovlar
    qs = Payment.objects.filter(student_id=student_id).order_by('-paid_date', '-paid_time')
    from core.tenant import get_request_center
    center = get_active_center(request)
    if center:
        qs = qs.filter(center=center)
    
    payments_data = []
    for p in qs:
        # Har bir to'lov qaysi oylarga tushganini ham ko'rsatishimiz mumkin
        allocs = PaymentAllocation.objects.filter(payment=p).select_related("tuition_month")
        alloc_list = [{"month": a.tuition_month.month.strftime("%Y-%m"), "amount": a.amount} for a in allocs]
        
        payments_data.append({
            "id": p.id,
            "paid_at": f"{p.paid_date.strftime('%d.%m.%Y')} {p.paid_time.strftime('%H:%M')}",
            "paid_date_iso": p.paid_date.strftime('%Y-%m-%d'),
            "cash": int(p.cash_amount or 0),
            "card": int(getattr(p, 'card_amount_som', 0) or getattr(p, 'card_amount', 0) or 0),
            "total": int(p.summa or 0),
            "method_code": p.payment_type,
            "method": p.get_payment_type_display(),
            "staff": p.created_by.get_full_name() if p.created_by else "—",
            "note": p.note or "",
            "group_name": p.group.nom if p.group else "—",
            "allocations": alloc_list,
            "receipt_url": reverse("education:payment_receipt_pdf", args=[p.id]) if p.id else None
        })

    return JsonResponse({
        "month": selected_month.strftime("%Y-%m"),
        "fee": total_fee,
        "paid_this_month": total_paid_this_month,
        "qoldiq": total_qoldiq,
        "payments": payments_data
    })



@login_required
def tolov_oqituvchilar(request):
    from django.shortcuts import redirect as _redirect
    return _redirect("education:teacher_salary_list")

