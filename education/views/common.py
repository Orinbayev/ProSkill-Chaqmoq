"""
Auto-split from education/views.py (phase 7 god-file reduction).
Public API re-exported via education.views package.
"""
from __future__ import annotations

from __future__ import annotations

import calendar
import json
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
# from multiprocessing import Value 
from django.db import models

logger = logging.getLogger(__name__)

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from billing.decorators import require_feature
from django.core.paginator import Paginator
from django.db.models import (
    Avg, Count, F, Min, Max, Prefetch, Q, Sum, OuterRef, Subquery
)
from django.db.models.functions import Coalesce, TruncMonth, Cast
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    JsonResponse,
)
from django.core.exceptions import PermissionDenied, ValidationError
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from django.utils.timezone import localdate, make_aware
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from datetime import datetime
from education.services.tuition import (
    parse_month_str,
    month_first_day,
    month_last_day,
    ensure_tuition_month,
    get_month_paid,
    full_course_amount,
    effective_student_payable_amount,
    calculate_enrollment_debt_snapshots,
    month_range_starts,
    tuition_month_fee_field,
    tuition_month_fee,
    ensure_all_tuition_months_since_start,
    create_payment_and_allocate,
    update_payment_and_reallocate,
    _allocate_amount_forward,
    infer_payment_type,
    prorated_monthly_fee,
    attendance_based_fee,
    billable_attendance_count,
    expected_lessons_in_period,
    enrollment_start_date,
    enrollment_lesson_pattern,
    enrollment_month_financial_snapshot,
    format_money,
    is_month_closed_for_center,
    normalize_lesson_pattern,
    round_money_to_thousand,
    round_div,
    resolve_lesson_schedule,
    tuition_month_preview,
    tuition_amount_breakdown,
    lesson_pattern_label,
    lesson_pattern_hint,
    preload_enrollment_history_starts,
)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors

from django import forms

from accounts.models import User
from chaqmoq.models import Ledger, Rule
from ..forms import CenterExpenseForm, GroupForm, ITGroupForm, LangGroupForm, StudentGroupTransferForm
from ..models import (
    Attendance,
    Category,
    CourseTemplate,
    CenterExpense,
    DailyLightningRecord,
    Dars,
    Enrollment,
    Group,
    OylikHisobot,
    Payment,
    PaymentAllocation,
    Student,
    TuitionMonth,
    StudentGroupHistory,
    FinancialMonth,
    MonthlyFinanceSnapshot,
    TeacherSalarySnapshot,
    TeacherIncome,
)
from education.services.historical_finance_service import HistoricalFinanceService
from education.services.enrollment_service import EnrollmentService
from education.services.student_transfer import transfer_student_to_group, user_can_transfer_student
from education.services.lesson_planning import calculate_lessons, validate_remaining_lessons
from accounts.models import Center
from ..permissions import user_can_manage_payments
from django.db import transaction
from django.db.models.functions import ExtractYear, ExtractMonth, ExtractDay  # student_detail dagi underline ham yo'qoladi
from urllib.parse import urlparse, parse_qs
from django.db import transaction
from urllib.parse import urlencode, urlparse, parse_qs, unquote
from django.urls import reverse  # sizda reverse ishlatyapsiz, import yo'q
from django.db import transaction
from django.db.models import Sum, F, Value as DJValue
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from decimal import Decimal

from core.center_features import (
    FEATURE_UI_CERTIFICATES,
    FEATURE_UI_EXAM_SESSIONS,
    FEATURE_UI_FAILED_STUDENTS,
    FEATURE_UI_WEEKLY_SCHEDULE,
    center_ui_feature_enabled,
)

U = get_user_model()

DAILY_LIMIT = 50  # (hozircha ishlatilmayapti, lekin qoldirdim)



def _redirect_disabled_module(request, *, message: str):
    messages.warning(request, message)
    return redirect("core:home")



def _ensure_center_ui_feature(request, center, feature_code: str, *, message: str):
    if center_ui_feature_enabled(center, feature_code):
        return None
    return _redirect_disabled_module(request, message=message)



def _day_range(d):
    start = make_aware(datetime.combine(d, datetime.min.time()))
    end = make_aware(datetime.combine(d + timedelta(days=1), datetime.min.time()))
    return start, end



def _accumulate_daily_lightning(*, group, student, date_value, points_delta):
    record, _ = DailyLightningRecord.objects.get_or_create(
        group=group,
        student=student,
        date=date_value,
        defaults={
            "center": getattr(group, "center", None),
            "plus_points": 0,
            "minus_points": 0,
        },
    )
    if points_delta > 0:
        record.plus_points = int(record.plus_points or 0) + int(points_delta)
    elif points_delta < 0:
        record.minus_points = int(record.minus_points or 0) + abs(int(points_delta))
    
    if not record.center_id and getattr(group, "center_id", None):
        record.center = group.center
        
    record.save(update_fields=["plus_points", "minus_points", "center", "updated_at"])
    return record



def _attendance_adjust_rule():
    """
    Davomat OFF bo'lganda, o'sha kundagi ballarni 'bekor qilish' uchun
    maxsus Rule kerak bo'ladi. (DBga 1 marta tushadi)
    """
    rule, _ = Rule.objects.get_or_create(
        nom="Davomat bekor qilindi",
        defaults={"tur": Rule.MINUS, "min_baho": 1, "max_baho": 10000},
    )
    return rule



def parse_month_yyyy_mm(s: str):
    # '2026-01' -> date(2026, 1, 1)
    try:
        y, m = s.split("-")
        y = int(y); m = int(m)
        if 1 <= m <= 12:
            return date(y, m, 1)
    except Exception:
        return None
    return None



def first_day_of_current_month():
    d = timezone.localdate()
    return date(d.year, d.month, 1)



def _parse_int_value(value, default=None):
    try:
        if value in (None, "", "None"):
            return default
        return int(str(value).replace(" ", "").replace(",", ""))
    except (TypeError, ValueError):
        return default



def _parse_bool_value(value) -> bool:
    return str(value or "").strip().lower() in {"on", "1", "true", "yes"}



def _preview_month_for_start_date(start_date: date | None, fallback: date | None = None) -> date:
    if start_date:
        return month_first_day(start_date)
    return month_first_day(fallback or timezone.localdate())



def _lesson_pattern_options():
    return [
        {
            "value": pattern,
            "label": lesson_pattern_label(pattern),
            "hint": lesson_pattern_hint(pattern),
        }
        for pattern in (
            Enrollment.LESSON_PATTERN_ODD,
            Enrollment.LESSON_PATTERN_EVEN,
            Enrollment.LESSON_PATTERN_DAILY,
        )
    ]



def _serialize_lesson_plan(plan: dict) -> dict:
    return {
        "requested_start_date": plan["requested_start_date"].isoformat(),
        "start_date": plan["start_date"].isoformat(),
        "calculation_start_date": plan["calculation_start_date"].isoformat(),
        "remaining_lessons": int(plan["remaining_lessons"] or 0),
        "lesson_pattern": plan["lesson_pattern"],
        "lesson_pattern_label": plan["lesson_pattern_label"],
        "lesson_pattern_hint": plan["lesson_pattern_hint"],
        "lesson_dates": [lesson_date.isoformat() for lesson_date in plan.get("lesson_dates", [])],
        "lesson_date_labels": list(plan.get("lesson_date_labels", [])),
        "last_lesson_date": plan["last_lesson_date"].isoformat() if plan.get("last_lesson_date") else "",
        "last_lesson_date_label": plan.get("last_lesson_date_label", "—"),
        "calculation_note": plan.get("calculation_note", ""),
        "used_reference_date": bool(plan.get("used_reference_date")),
    }



def _student_enrollment_catalog(enrollments: list[Enrollment]) -> dict:
    enrollment_rows: list[dict] = []

    for enrollment in enrollments:
        enrollment_rows.append(
            {
                "enrollment_id": enrollment.id,
                "group_id": enrollment.group_id,
                "group_name": getattr(enrollment.group, "nom", "—"),
                "joined_at": enrollment.joined_at.isoformat() if enrollment.joined_at else "",
                "stored_lesson_pattern": getattr(enrollment, "lesson_pattern", Enrollment.LESSON_PATTERN_GROUP),
                "resolved_lesson_pattern": enrollment_lesson_pattern(enrollment),
                "monthly_lessons": int(getattr(enrollment.group, "oy_dars_soni", 0) or getattr(enrollment, "monthly_lessons", 0) or 12),
                "course_price": int(getattr(enrollment, "kurs_narhi", 0)),
                "teacher_percent": int(getattr(enrollment, "oqituvchi_foiz", 0) or getattr(enrollment.group, "oqituvchi_foiz", 0) or 0),
                "student_payable_amount": getattr(enrollment, "student_payable_amount", None),
                "remaining_lessons_override": getattr(enrollment, "remaining_lessons_override", None),
                "last_lesson_date": enrollment.last_lesson_date.isoformat() if getattr(enrollment, "last_lesson_date", None) else "",
            }
        )

    enrollment_rows.sort(
        key=lambda item: (
            (item.get("group_name") or "").lower(),
            int(item.get("enrollment_id") or 0),
        )
    )
    return {
        "enrollments": enrollment_rows,
    }



def get_student_total_debt(student, center=None) -> int:
    from django.db.models import Q as _Q
    from django.utils import timezone
    from education.models import Enrollment, TuitionMonth
    _center_q = (
        _Q(center=center)
        | _Q(center__isnull=True, group__center=center)
        | _Q(center__isnull=True, student__center=center)
    )
    active_enrs = list(
        Enrollment.objects.filter(
            student=student,
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).filter(_center_q)
    )
    _inactive_enr_ids = (
        TuitionMonth.objects
        .filter(
            enrollment__student=student,
            enrollment__is_active=False,
            is_deleted=False,
            enrollment__student__is_archived=False,
            enrollment__group__is_archived=False,
            enrollment__group__is_deleted=False,
        )
        .values_list("enrollment_id", flat=True)
        .distinct()
    )
    inactive_enrs = list(
        Enrollment.objects.filter(
            id__in=_inactive_enr_ids,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        ).filter(_center_q)
    )
    all_enrs = active_enrs + inactive_enrs
    if not all_enrs:
        return 0

    today = timezone.localdate()
    selected_from = today.replace(day=1)
    selected_to = today
    from education.services.tuition import month_range_starts, calculate_enrollment_debt_snapshots
    period_months = month_range_starts(selected_from, selected_to)

    snapshots = calculate_enrollment_debt_snapshots(
        all_enrs, period_months, cumulative_up_to=selected_to,
        # O'tgan oyga TuitionMonth yozuvi bo'lmasa avtomatik (virtual) qarz
        # yozilmaydi — faqat haqiqiy yozuvlar hisoblanadi.
        synthesize_past_virtual=False,
    )
    total_debt = 0
    for snap in snapshots.values():
        total_debt += int(snap.get("debt", 0) or 0)
        total_debt += int(snap.get("previous_unpaid", 0) or 0)
    return total_debt



def _serialize_tuition_preview(preview: dict) -> dict:
    period_end = preview.get("period_end_date") or month_last_day(preview["month"])
    return {
        "month": preview["month"].isoformat(),
        "start_date": preview["start_date"].isoformat(),
        "period_end_date": period_end.isoformat(),
        "period_end_date_label": period_end.strftime("%d.%m.%Y"),
        "lesson_pattern": preview["lesson_pattern"],
        "lesson_pattern_label": preview["lesson_pattern_label"],
        "lesson_pattern_hint": preview.get("lesson_pattern_hint", ""),
        "monthly_lessons": int(preview["monthly_lessons"] or 0),
        "lesson_count": int(preview["lesson_count"] or 0),
        "per_lesson_amount": int(preview.get("per_lesson_amount", 0) or 0),
        "per_lesson_amount_display": preview.get("per_lesson_amount_display", format_money(0)),
        "counted_weekdays": [int(weekday) for weekday in preview.get("counted_weekdays", [])],
        "counted_weekday_labels": list(preview.get("counted_weekday_labels", [])),
        "counted_weekday_short_labels": list(preview.get("counted_weekday_short_labels", [])),
        "counted_days_text": preview.get("counted_days_text", ""),
        "counted_days_summary": preview.get("counted_days_summary", ""),
        "lesson_count_summary": preview.get("lesson_count_summary", ""),
        "lesson_dates": [lesson_date.isoformat() for lesson_date in preview.get("lesson_dates", [])],
        "lesson_date_labels": list(preview.get("lesson_date_labels", [])),
        "fee_amount": int(preview["fee_amount"] or 0),
        "teacher_share": int(preview["teacher_share"] or 0),
        "center_share": int(preview["center_share"] or 0),
        "full_turnover": int(preview["full_turnover"] or 0),
        "month_label_uz": preview.get("month_label_uz", ""),
        "debt_label_uz": preview.get("debt_label_uz", ""),
    }



def _format_money_exact(amount: int | float | None) -> str:
    return f"{int(amount or 0):,}".replace(",", " ") + " so'm"



def _apply_period_end_to_preview(preview: dict, period_end: date | None) -> dict:
    if not period_end:
        return preview
    adjusted = dict(preview)
    lesson_dates = [
        lesson_date
        for lesson_date in preview.get("lesson_dates", [])
        if lesson_date <= period_end
    ]
    adjusted.update(
        {
            "period_end_date": period_end,
            "lesson_dates": lesson_dates,
            "lesson_date_labels": [lesson_date.strftime("%d.%m.%Y") for lesson_date in lesson_dates],
            "lesson_count": len(lesson_dates),
            "lesson_count_summary": f"Bu davr bo'yicha {len(lesson_dates)} ta mos dars kuni topildi",
        }
    )
    return adjusted



def _apply_lesson_count_breakdown(preview: dict, enrollment: Enrollment, lesson_count: int) -> dict:
    adjusted = dict(preview)
    breakdown = tuition_amount_breakdown(
        enrollment,
        lesson_count,
        course_price=full_course_amount(enrollment),
        monthly_lessons=int(preview.get("monthly_lessons", 0) or 0),
        teacher_percent=getattr(enrollment, "oqituvchi_foiz", 0) or 0,
    )
    adjusted.update(
        {
            "lesson_count": breakdown["lesson_count"],
            "per_lesson_amount": breakdown["per_lesson_amount"],
            "per_lesson_amount_display": _format_money_exact(breakdown["per_lesson_amount"]),
            "fee_amount": breakdown["fee_amount"],
            "fee_amount_display": _format_money_exact(breakdown["fee_amount"]),
            "full_turnover": breakdown["fee_amount"],
            "teacher_share": breakdown["teacher_share"],
            "teacher_share_display": _format_money_exact(breakdown["teacher_share"]),
            "center_share": breakdown["center_share"],
            "center_share_display": _format_money_exact(breakdown["center_share"]),
            "lesson_count_summary": f"Hisoblangan darslar: {breakdown['lesson_count']} ta",
        }
    )
    return adjusted



def _parse_period_end(value, fallback_month: date) -> date:
    parsed = parse_date((value or "").strip()) if value not in (None, "") else None
    if parsed is None:
        return month_last_day(fallback_month)
    return parsed



def _build_tuition_preview_enrollment(
    *,
    base_enrollment: Enrollment | None = None,
    group: Group | None = None,
    start_date: date | None = None,
    lesson_pattern: str | None = None,
    monthly_lessons: int | None = None,
    course_price: int | None = None,
    teacher_percent: int | None = None,
    student_payable_amount: int | None = None,
) -> Enrollment:
    resolved_group = group or getattr(base_enrollment, "group", None)
    resolved_start_date = start_date
    if resolved_start_date is None and base_enrollment is not None:
        resolved_start_date = getattr(base_enrollment, "joined_at", None) or enrollment_start_date(base_enrollment)
    if resolved_start_date is None:
        resolved_start_date = timezone.localdate()
    resolved_monthly_lessons = int(
        monthly_lessons
        or getattr(base_enrollment, "monthly_lessons", 0)
        or getattr(resolved_group, "oy_dars_soni", 0)
        or 12
    )
    resolved_course_price = int(
        course_price
        if course_price is not None
        else getattr(base_enrollment, "kurs_narhi", 0)
        or getattr(resolved_group, "kurs_narxi", 0)
        or 0
    )
    resolved_teacher_percent = int(
        teacher_percent
        if teacher_percent is not None
        else getattr(base_enrollment, "oqituvchi_foiz", 0)
        or getattr(resolved_group, "oqituvchi_foiz", 0)
        or 0
    )
    schedule_meta = resolve_lesson_schedule(
        resolved_start_date,
        lesson_pattern if lesson_pattern is not None else getattr(base_enrollment, "lesson_pattern", None),
    )
    resolved_start_date = schedule_meta["start_date"]
    resolved_pattern = schedule_meta["lesson_pattern"]

    preview_enrollment = Enrollment(
        group=resolved_group,
        course=getattr(resolved_group, "category_obj", None) or getattr(base_enrollment, "course", None),
        student=getattr(base_enrollment, "student", None),
        center=getattr(resolved_group, "center", None) or getattr(base_enrollment, "center", None),
        kurs_narhi=resolved_course_price,
        oqituvchi_foiz=resolved_teacher_percent,
        student_payable_amount=student_payable_amount,
        monthly_lessons=resolved_monthly_lessons,
        joined_at=resolved_start_date,
        lesson_pattern=resolved_pattern,
        is_active=True,
    )
    preview_enrollment.created_at = getattr(base_enrollment, "created_at", None) or timezone.now()
    preview_enrollment._tuition_start_date = resolved_start_date
    preview_enrollment._tuition_requested_start_date = start_date or resolved_start_date
    return preview_enrollment



def _get_int(querydict, key, default=0):
    try:
        val = querydict.get(key, None)
        if val in (None, "", "None"):
            return default
        return int(val)
    except (TypeError, ValueError):
        return default



def _schedule_weekday_labels():
    return [
        (1, "Dushanba"),
        (2, "Seshanba"),
        (3, "Chorshanba"),
        (4, "Payshanba"),
        (5, "Juma"),
        (6, "Shanba"),
        (7, "Yakshanba"),
    ]



def _student_group_financial_cards(
    student,
    *,
    center=None,
    month: date | None = None,
    include_dates: bool = True,
):
    target_month = month_first_day(month or timezone.localdate())
    enrollments_qs = (
        Enrollment.objects
        .select_related("group", "group__category_obj", "course")
        .filter(
            student=student,
            is_active=True,
            student__is_archived=False,
            group__is_archived=False,
            group__is_deleted=False,
        )
        .order_by("group__nom", "id")
    )
    if center:
        enrollments_qs = enrollments_qs.filter(center=center)

    cards = []
    totals = {
        "fee_amount": 0,
        "teacher_share": 0,
        "center_share": 0,
        "debt_amount": 0,
    }

    for enrollment in enrollments_qs:
        snapshot = enrollment_month_financial_snapshot(enrollment, target_month)
        default_course_price = int(getattr(enrollment.group, "kurs_narxi", 0) or 0)
        card = {
            "enrollment_id": enrollment.id,
            "group_id": getattr(enrollment, "group_id", None),
            "group_name": getattr(enrollment.group, "nom", "—"),
            "start_date": snapshot["start_date"],
            "lesson_pattern": snapshot["lesson_pattern"],
            "lesson_pattern_label": snapshot["lesson_pattern_label"],
            "lesson_count": snapshot["lesson_count"],
            "lesson_dates": snapshot["lesson_dates"] if include_dates else [],
            "lesson_date_labels": snapshot["lesson_date_labels"] if include_dates else [],
            "course_price": snapshot["course_price"],
            "course_price_display": snapshot["course_price_display"],
            "default_course_price": default_course_price,
            "default_course_price_display": format_money(default_course_price),
            "is_individual_price": bool(default_course_price and snapshot["course_price"] != default_course_price),
            "fee_amount": snapshot["fee_amount"],
            "fee_amount_display": snapshot["fee_amount_display"],
            "teacher_share": snapshot["teacher_share"],
            "teacher_share_display": snapshot["teacher_share_display"],
            "center_share": snapshot["center_share"],
            "center_share_display": snapshot["center_share_display"],
            "debt_amount": snapshot["debt_amount"],
            "debt_amount_display": snapshot["debt_amount_display"],
        }
        cards.append(card)
        totals["fee_amount"] += card["fee_amount"]
        totals["teacher_share"] += card["teacher_share"]
        totals["center_share"] += card["center_share"]
        totals["debt_amount"] += card["debt_amount"]

    return {
        "cards": cards,
        "totals": {
            **totals,
            "fee_amount_display": format_money(totals["fee_amount"]),
            "teacher_share_display": format_money(totals["teacher_share"]),
            "center_share_display": format_money(totals["center_share"]),
            "debt_amount_display": format_money(totals["debt_amount"]),
        },
    }



# ---------- Ruxsat helperlari ----------
def _can_manage(u):
    return u.is_superuser or getattr(u, "role", None) in ("director", "manager")



def _can_give_points(user, g: Group):
    return (
        user.is_superuser
        or user.role in ("director", "manager")
        or (user.role == "teacher" and g.oqituvchi_id == user.id)
    )

    return user.is_superuser or user.role in ("director", "manager") or (
        user.role == "teacher" and g.oqituvchi_id == user.id
    )



def get_active_center(request):
    """
    Returns the active center for the current request.
    Now fully handled by TenantMiddleware.
    """
    return getattr(request, 'center', None)



def parse_month_str(s: str) -> date | None:
    """
    'YYYY-MM' yoki 'YYYY-MM-DD' -> date(YYYY, MM, 1)
    """
    if not s:
        return None
    s = s.strip()
    if len(s) >= 10:
        s = s[:7]
    if len(s) != 7 or s[4] != "-":
        return None
    try:
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            return None
        return date(y, m, 1)
    except Exception:
        return None



def _get_month_from_next(next_url: str, fallback: date) -> date:
    try:
        qs = parse_qs(urlparse(next_url).query)

        # 1) eng to'g'risi: month=YYYY-MM
        m = (qs.get("month", [""])[0] or "").strip()
        if m:
            return parse_month_str(m) or fallback

        # 2) sizda ishlayotgan variant: pay_month=1..12
        pm = (qs.get("pay_month", [""])[0] or "").strip()
        if pm.isdigit():
            mm = int(pm)
            if 1 <= mm <= 12:
                # year bo'lmasa joriy yil
                yy = (qs.get("year", [""])[0] or "").strip()
                yy = int(yy) if yy.isdigit() else fallback.year
                return date(yy, mm, 1)

        return fallback
    except Exception:
        return fallback



def _resolve_paid_date(raw_value: str | None) -> date:
    parsed = parse_date((raw_value or "").strip())
    return parsed or localdate()



def _build_paid_at_for_date(selected_date: date) -> datetime:
    current_local = timezone.localtime()
    return current_local.replace(
        year=selected_date.year,
        month=selected_date.month,
        day=selected_date.day,
        microsecond=0,
    )



def _can_manage(u: User) -> bool:
    return u.is_superuser or getattr(u, "role", None) in ("director", "manager")



def _teacher_can(user: User, g: Group) -> bool:
    return (
        user.is_superuser
        or getattr(user, "role", None) in ("director", "manager")
        or (getattr(user, "role", None) == "teacher" and getattr(g, "oqituvchi_id", None) == user.id)
    )



def _can_give_points(user: User, g: Group) -> bool:
    return _teacher_can(user, g)



def month_first_day(d: date) -> date:
    return d.replace(day=1)



def get_month_paid(enr: Enrollment, month: date) -> int:
    month = month_first_day(month)
    tm = TuitionMonth.objects.filter(enrollment=enr, month=month).first()
    if not tm:
        return 0
    s = PaymentAllocation.objects.filter(tuition_month=tm).aggregate(x=Sum("amount"))["x"] or 0
    return int(s)



def _add_month(d: date, n: int = 1) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)



def _model_has_field(model, field_name: str) -> bool:
    return any(f.name == field_name for f in model._meta.get_fields())



def _safe_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default



def _is_teacher_share_only_enrollment(enrollment) -> bool:
    if not enrollment:
        return False

    full_amount = full_course_amount(enrollment)
    effective_amount = effective_student_payable_amount(enrollment)
    teacher_share_amount = int(getattr(enrollment, "oqituvchi_daromadi", 0) or 0)

    return (
        enrollment.student_payable_amount not in (None, "")
        and full_amount > effective_amount
        and effective_amount == teacher_share_amount
    )



def _parse_yyyy_mm(s: str):
    """
    '2026-01' -> date(2026,1,1)
    None/invalid -> current month first day
    """
    s = (s or "").strip()
    try:
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            raise ValueError()
        return timezone.datetime(y, m, 1).date()
    except Exception:
        now = timezone.localdate()
        return timezone.datetime(now.year, now.month, 1).date()



def _first_day_of_month(d: date) -> date:
    return d.replace(day=1)



# --------- helpers (shu fayl ichida bo'lsa sariq bo'lmaydi) ----------
def _model_has_field(model_cls, field_name: str) -> bool:
    return any(f.name == field_name for f in model_cls._meta.get_fields())



def _get_fee_amount(enrollment) -> int:
    """
    fee manbasi:
    - enrollment.student_payable_amount (agar berilgan bo'lsa)
    - aks holda to'liq kurs narxi
    """
    return effective_student_payable_amount(enrollment)



def _parse_month_str(s: str):
    try:
        if not s or len(s) < 7:
            return None
        y = int(s[:4])
        m = int(s[5:7])
        if m < 1 or m > 12:
            return None
        return date(y, m, 1)
    except Exception:
        return None



def _fmt(n: int) -> str:
    return f"{int(n or 0):,}".replace(",", " ")



def _ellipsis(text: str, limit: int = 34) -> str:
    text = (text or "").strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"



def _last_12_month_starts():
    """Oxirgi 12 oy (shu oy ham kiradi), tartib: eski -> yangi"""
    today = timezone.localdate().replace(day=1)
    months = []
    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(date(y, m, 1))
    return months



def _month_start(d: date) -> date:
    return d.replace(day=1)



def _add_months(d: date, n: int) -> date:
    # month arithmetic
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)



def _last_12_ending(anchor: date) -> list[date]:
    # anchor included, return 12 month starts ending at anchor
    anchor = _month_start(anchor)
    return [_add_months(anchor, -11 + i) for i in range(12)]



def _chart_range_mode(from_date: date, to_date: date) -> str:
    """'daily' when range ≤ 31 days (single-month view), 'monthly' otherwise."""
    return "daily" if (to_date - from_date).days < 32 else "monthly"



def _chart_monthly_buckets_for_range(from_date: date, to_date: date) -> list[date]:
    """Return month-start dates covering [from_date, to_date] inclusive."""
    result, m = [], from_date.replace(day=1)
    end = to_date.replace(day=1)
    while m <= end:
        result.append(m)
        m = _add_months(m, 1)
    return result



def _build_daily_debt_series(enrollment_ids: list, from_date: date, to_date: date) -> list[int]:
    """
    For each day in [from_date, to_date] return the outstanding debt for
    display_month = to_date.replace(day=1).  Matches what the monthly chart
    shows for that month (fee_amount − PaymentAllocation for that month only).
    """
    if not enrollment_ids:
        return [0] * ((to_date - from_date).days + 1)

    display_month = to_date.replace(day=1)

    # Fees for display_month only (consistent with per-month chart bars)
    total_fees = int(
        TuitionMonth.objects.filter(
            enrollment_id__in=enrollment_ids,
            month=display_month,
            is_deleted=False,
        ).aggregate(s=Sum("fee_amount"))["s"] or 0
    )

    # Payments allocated to display_month TuitionMonths made BEFORE from_date
    paid_before = int(
        PaymentAllocation.objects.filter(
            tuition_month__enrollment_id__in=enrollment_ids,
            tuition_month__month=display_month,
            tuition_month__is_deleted=False,
            payment__paid_date__lt=from_date,
            is_deleted=False,
            payment__is_deleted=False,
        ).aggregate(s=Sum("amount"))["s"] or 0
    )

    # Payments allocated to display_month TuitionMonths per day within range
    daily_map: dict = {}
    for row in (
        PaymentAllocation.objects.filter(
            tuition_month__enrollment_id__in=enrollment_ids,
            tuition_month__month=display_month,
            tuition_month__is_deleted=False,
            payment__paid_date__gte=from_date,
            payment__paid_date__lte=to_date,
            is_deleted=False,
            payment__is_deleted=False,
        )
        .values("payment__paid_date")
        .annotate(total=Sum("amount"))
    ):
        daily_map[row["payment__paid_date"]] = int(row["total"] or 0)

    running, series = paid_before, []
    for i in range((to_date - from_date).days + 1):
        running += daily_map.get(from_date + timedelta(days=i), 0)
        series.append(max(0, total_fees - running))
    return series



def _human_period_label(start_date: date, end_date: date) -> str:
    if start_date == end_date:
        return f"{start_date.day}-{UZ_MONTH_NAMES.get(start_date.month, start_date.strftime('%B'))} {start_date.year}"
    return (
        f"{start_date.day}-{UZ_MONTH_NAMES.get(start_date.month, start_date.strftime('%B'))} {start_date.year}"
        f" dan {end_date.day}-{UZ_MONTH_NAMES.get(end_date.month, end_date.strftime('%B'))} {end_date.year} gacha"
    )



def _human_month_label(month_value: date) -> str:
    return UZ_MONTH_NAMES.get(month_value.month, month_value.strftime('%B'))



def _human_month_period_label(start_month: date, end_month: date) -> str:
    if start_month == end_month:
        return _human_month_label(start_month)
    return f"{_human_month_label(start_month)} dan {_human_month_label(end_month)} gacha"



def _build_last_12_month_money_chart_series(qs, *, date_field: str, amount_field: str, anchor_date: date):
    buckets = _last_12_ending(anchor_date)
    rows = (
        qs.annotate(bucket=TruncMonth(date_field))
        .values("bucket")
        .annotate(total=Sum(amount_field))
        .order_by("bucket")
    )
    value_map = {}
    for row in rows:
        bucket = row["bucket"]
        if hasattr(bucket, "date"):
            bucket = bucket.date()
        bucket = bucket.replace(day=1)
        value_map[bucket] = int(row["total"] or 0)
    labels = [_human_month_label(bucket) for bucket in buckets]
    data = [value_map.get(bucket, 0) for bucket in buckets]
    return labels, data, "Oxirgi 12 oy", _human_month_period_label(buckets[0], buckets[-1])



def _payment_month_parts(pay_rows, alloc_map):
    """
    Har bir to'lovni oylarga bo'ladi. KAFOLAT: ulushlar yig'indisi ROPPA-ROSA
    payment.summa ga teng — hech qachon oshmaydi (eski ikki-marta-taqsimlash
    bugi buzgan yozuvlarda allocation jami summadan katta bo'lishi mumkin;
    bunda eng eski oylar olinadi, ortiqchasi tashlab yuboriladi).

    pay_rows: [(id, summa, paid_date), ...]
    alloc_map: payment_id -> [(month, amount), ...] (faqat live allocationlar)
    Returns: payment_id -> {month_first_day: ulush}
    """
    parts = {}
    for pid, summa, pd in pay_rows:
        summa = int(summa or 0)
        d = {}
        remaining = summa
        for m, amt in sorted(alloc_map.get(pid, ()), key=lambda x: x[0]):
            if remaining <= 0:
                break
            take = min(int(amt or 0), remaining)
            if take > 0:
                d[m] = d.get(m, 0) + take
                remaining -= take
        # Bog'lanmagan qoldiq → pul berilgan oyga
        if remaining > 0 and pd:
            mb = pd.replace(day=1)
            d[mb] = d.get(mb, 0) + remaining
        parts[pid] = d
    return parts



def _build_alloc_map(payment_ids):
    """payment_id -> [(month, amount), ...] — faqat live allocationlar."""
    amap = {}
    for r in PaymentAllocation.objects.filter(
        is_deleted=False, payment_id__in=list(payment_ids)
    ).values_list("payment_id", "tuition_month__month", "amount"):
        amap.setdefault(r[0], []).append((r[1].replace(day=1), int(r[2] or 0)))
    return amap



def _get_payment_dashboard_data(request):
    center = get_active_center(request)
    if not center and not request.user.is_superuser:
        raise PermissionDenied("Markaz biriktirilmagan")

    today = timezone.localdate()
    cur_month_start = today.replace(day=1)

    base_payment_qs = Payment.objects.filter(center=center) if center else Payment.objects.none()
    total_income = base_payment_qs.aggregate(s=Sum("summa"))["s"] or 0

    q = (request.GET.get("q") or "").strip()
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or "").strip()
    sel_group = request.GET.get("group") or ""
    sel_teacher = request.GET.get("teacher") or ""
    sel_course = request.GET.get("course") or ""
    sel_staff = request.GET.get("staff") or ""
    sel_type = request.GET.get("payment_type") or ""
    sel_month = request.GET.get("pay_month") or ""

    # Map any payment-method name (NAQD/KARTA/CLICK/PAYME/...) to internal
    # Payment.payment_type code (cash/card/mixed) for DB filtering.
    # `sel_type` is preserved for UI dropdown; `sel_type_filter` is used in queryset.
    _PM_NAME_TO_MODE = {
        'cash': 'cash',
        'card': 'card',
        'mixed': 'mixed',
        'naqd': 'cash',
        'karta': 'card',
        'plastik': 'card',
        'aralash': 'mixed',
    }
    sel_type_filter = ''
    if sel_type:
        sel_type_filter = _PM_NAME_TO_MODE.get(sel_type.lower(), '')

    selected_from = parse_date(date_from_raw) if date_from_raw else None
    selected_to = parse_date(date_to_raw) if date_to_raw else None

    # pay_month tanlanib, sana filtri ko'rsatilmagan bo'lsa:
    # paid_date filtrini o'sha yilning boshidan oxirigacha kengaytir.
    # (Aks holda default joriy oy date filtri pay_month bilan to'qnashib,
    # boshqa oylarda qilingan to'lovlarni yashiradi.)
    if sel_month and sel_month.isdigit() and not date_from_raw and not date_to_raw:
        # Aprel kabi o'tgan oy tanlansa, joriy yilning to'liq diapazonini ishlatamiz
        selected_from = date(today.year, 1, 1)
        selected_to = date(today.year, 12, 31)
    elif not selected_from and not selected_to:
        selected_from = cur_month_start
        selected_to = today
    else:
        if selected_from and not selected_to:
            selected_to = today if selected_from <= today else selected_from
        elif selected_to and not selected_from:
            selected_from = selected_to.replace(day=1)
        if selected_from and selected_to and selected_from > selected_to:
            selected_to = selected_from

    date_from = selected_from.isoformat()
    date_to = selected_to.isoformat()

    allocation_prefetch = Prefetch(
        "allocations",
        queryset=PaymentAllocation.objects.select_related(
            "tuition_month",
            "tuition_month__enrollment",
            "tuition_month__enrollment__group",
            "tuition_month__enrollment__group__category_obj",
        ).order_by("tuition_month__month", "id"),
        to_attr="prefetched_allocations",
    )

    pay_qs = base_payment_qs.select_related(
        "student", "group", "group__oqituvchi", "group__category_obj", "created_by"
    ).prefetch_related(allocation_prefetch)
    chart_qs = base_payment_qs.select_related(
        "student", "group", "group__oqituvchi", "group__category_obj", "created_by"
    )
    cur_year = selected_to.year

    # Oy filtri tanlanganda: shu oyning 1-sanasi (allocation month bilan solishtirish uchun)
    sel_month_first = None
    if sel_month and sel_month.isdigit():
        sel_month_first = date(cur_year, int(sel_month), 1)

    def _apply_shared_payment_filters(qs):
        if q:
            qs = qs.filter(
                Q(student__ism__icontains=q)
                | Q(student__familya__icontains=q)
                | Q(student__telefon1__icontains=q)
                | Q(student__telefon2__icontains=q)
                | Q(student__email__icontains=q)
                | Q(student__gmail__icontains=q)
            )

        if sel_group:
            qs = qs.filter(group_id=sel_group)
        if sel_teacher:
            qs = qs.filter(group__oqituvchi_id=sel_teacher)
        if sel_course:
            qs = qs.filter(group__category_obj_id=sel_course)
        if sel_staff:
            qs = qs.filter(created_by_id=sel_staff)
        if sel_type_filter:
            qs = qs.filter(payment_type=sel_type_filter)

        return qs

    pay_qs = _apply_shared_payment_filters(pay_qs)
    chart_qs = _apply_shared_payment_filters(chart_qs)

    # Sana filtri FAQAT jadval va "Filter bo'yicha" uchun.
    # Diagramma doim 12 oylik tarixni ko'rsatadi (chart_qs sana filtrisiz).
    pay_qs = pay_qs.filter(paid_date__gte=selected_from, paid_date__lte=selected_to)

    chart_anchor_date = selected_to or today
    chart_months = _last_12_ending(chart_anchor_date)
    chart_start = chart_months[0]
    chart_end = _add_months(chart_months[-1], 1) - timedelta(days=1)

    # ═══ YAGONA QOIDA: har bir so'm ROPPA-ROSA BITTA oyga tegishli ═══
    # - To'lovning oyga BOG'LANGAN qismi (live allocation) → o'sha oyda
    #   (iyunda berilgan may puli → MAY da)
    # - BOG'LANMAGAN qoldi'g'i (summa - live alloc): allocation'lari reset
    #   bilan bekor qilingan yoki umuman yaratilmagan pullar → paid_date oyida
    # Natija: oylar yig'indisi = umumiy daromad; filtr = diagramma ustuni.

    # "To'lov oyi" filtri: shu oyda ULUSHI bor to'lovlar.
    # _payment_month_parts kafolati: bir to'lov ulushlari jami = summa —
    # eski ikki-marta-taqsimlangan yozuvlarda ham oshirib sanamaydi.
    _month_part_by_pay = {}
    if sel_month_first:
        _base_rows = list(pay_qs.values_list("id", "summa", "paid_date"))
        _base_parts = _payment_month_parts(
            _base_rows, _build_alloc_map([r[0] for r in _base_rows])
        )
        _month_part_by_pay = {
            pid: d[sel_month_first]
            for pid, d in _base_parts.items()
            if d.get(sel_month_first, 0) > 0
        }
        pay_qs = pay_qs.filter(id__in=list(_month_part_by_pay.keys()))
        filtered_income = sum(_month_part_by_pay.values())

    payment_ids = list(pay_qs.values_list("id", flat=True))

    if not sel_month_first:
        filtered_income = Payment.objects.filter(id__in=payment_ids).aggregate(s=Sum("summa"))["s"] or 0
    unique_payers_count = Payment.objects.filter(id__in=payment_ids).values("student").distinct().count()

    # ── Diagramma: TO'LOV SANASI (paid_date) bo'yicha (oxirgi 12 oy, oy filtrisiz) ──
    # Har ustun = o'sha oyda HAQIQATAN kassaga tushgan pul (summa).
    # Jadval, "Filter bo'yicha" va "Umumiy daromad" bilan bir xil o'lchov:
    # pul to'langan bo'lsa ustun ko'rinadi, to'lanmagan bo'lsa 0.
    # (Ilgari "to'lov oyi"/allocation bo'yicha edi — shu sabab oldindan
    #  to'langan pullar boshqa oy ustunida ko'rinib chalkashlik chiqargan.)
    _chart_pay_rows = list(chart_qs.values_list("id", "summa", "paid_date"))
    _chart_value_map = {}
    _chart_contrib_ids = set()
    for _pid, _summa, _pd in _chart_pay_rows:
        if not _pd:
            continue
        _b = _pd.replace(day=1)
        if chart_start <= _b <= chart_end:
            _amt = int(_summa or 0)
            if _amt:
                _chart_value_map[_b] = _chart_value_map.get(_b, 0) + _amt
                _chart_contrib_ids.add(_pid)

    chart_labels = [_human_month_label(b) for b in chart_months]
    chart_data = [_chart_value_map.get(b, 0) for b in chart_months]
    chart_kicker = "Oxirgi 12 oy"
    chart_period_label = _human_month_period_label(chart_months[0], chart_months[-1])

    # Diagramma statistikasi: hissa qo'shgan to'lovlar
    chart_payment_ids = list(_chart_contrib_ids)
    chart_payment_record_count = len(chart_payment_ids)
    chart_unique_payers_count = (
        Payment.objects.filter(id__in=chart_payment_ids)
        .values("student").distinct().count()
    )

    pay_qs = pay_qs.order_by("-paid_date", "-id")
    filtered_payments = list(pay_qs)
    uz_month_map = UZ_MONTH_NAMES
    grouped_rows = {}

    for payment in filtered_payments:
        student_id = payment.student_id
        row = grouped_rows.get(student_id)

        if row is None:
            row = {
                "student": payment.student,
                "latest_payment": payment,
                "latest_paid_date": payment.paid_date,
                "latest_created_at": payment.created_at,
                "latest_payment_id": payment.id,
                "latest_cash_amount": int(payment.cash_amount or 0),
                "latest_card_amount": float(payment.card_amount or 0),
                "latest_note": payment.note or "",
                "total_sum": 0,
                "payment_count": 0,
                "group_entries": {},
                "month_entries": {},
                "staff_entries": {},
                "type_entries": {},
            }
            grouped_rows[student_id] = row

        if sel_month_first:
            # Oy filtri: paymentning faqat SHU OYNING ulushi
            row["total_sum"] += _month_part_by_pay.get(payment.id, 0)
        else:
            row["total_sum"] += int(payment.summa or 0)
        row["payment_count"] += 1

        if payment.created_by:
            staff_name = payment.created_by.get_full_name() or payment.created_by.email
            row["staff_entries"].setdefault(payment.created_by_id or staff_name, staff_name)

        if payment.payment_type:
            row["type_entries"].setdefault(
                payment.payment_type,
                {
                    "code": payment.payment_type,
                    "label": payment.get_payment_type_display(),
                },
            )

        allocations = getattr(payment, "prefetched_allocations", []) or []
        if allocations:
            for alloc in allocations:
                tuition_month = getattr(alloc, "tuition_month", None)
                enrollment = getattr(tuition_month, "enrollment", None)
                group = getattr(enrollment, "group", None)
                if group and not getattr(group, "is_archived", False):
                    row["group_entries"].setdefault(
                        group.id,
                        {
                            "id": group.id,
                            "name": group.nom or "—",
                            "category": getattr(getattr(group, "category_obj", None), "name", "") or "—",
                        },
                    )

                month_value = getattr(tuition_month, "month", None)
                if month_value:
                    month_key = month_value.strftime("%Y-%m")
                    row["month_entries"].setdefault(
                        month_key,
                        {
                            "key": month_key,
                            "label": f"{uz_month_map.get(month_value.month, month_value.strftime('%B'))} {month_value.year}",
                        },
                    )
        else:
            if payment.group and not getattr(payment.group, "is_archived", False):
                row["group_entries"].setdefault(
                    payment.group_id,
                    {
                        "id": payment.group_id,
                        "name": payment.group.nom or "—",
                        "category": getattr(getattr(payment.group, "category_obj", None), "name", "") or "—",
                    },
                )
            if payment.paid_date:
                fallback_month = payment.paid_date.replace(day=1)
                month_key = fallback_month.strftime("%Y-%m")
                row["month_entries"].setdefault(
                    month_key,
                    {
                        "key": month_key,
                        "label": f"{uz_month_map.get(fallback_month.month, fallback_month.strftime('%B'))} {fallback_month.year}",
                    },
                )

    display_rows = []
    for row in grouped_rows.values():
        group_entries = list(row["group_entries"].values())
        month_entries = list(row["month_entries"].values())
        staff_entries = list(row["staff_entries"].values())
        type_entries = list(row["type_entries"].values())

        group_names = [g["name"] for g in group_entries if g.get("name")]
        category_names = []
        for entry in group_entries:
            category_name = entry.get("category")
            if category_name and category_name != "—" and category_name not in category_names:
                category_names.append(category_name)

        row["group_entries"] = group_entries
        row["visible_group_entries"] = group_entries[:2]
        row["remaining_group_count"] = max(0, len(group_entries) - 2)
        row["group_summary_title"] = ", ".join(group_names) if group_names else "—"
        row["category_summary"] = ", ".join(category_names[:2]) if category_names else "—"
        if len(category_names) > 2:
            row["category_summary"] += f" +{len(category_names) - 2}"

        row["month_entries"] = month_entries
        row["visible_month_entries"] = month_entries[:2]
        row["remaining_month_count"] = max(0, len(month_entries) - 2)
        row["month_summary_title"] = ", ".join(m["label"] for m in month_entries) if month_entries else "—"

        row["staff_entries"] = staff_entries
        row["staff_summary"] = ", ".join(staff_entries[:2]) if staff_entries else "—"
        if len(staff_entries) > 2:
            row["staff_summary"] += f" +{len(staff_entries) - 2}"

        row["type_entries"] = type_entries
        row["visible_type_entries"] = type_entries[:2]
        row["remaining_type_count"] = max(0, len(type_entries) - 2)

        display_rows.append(row)

    groups = Group.objects.filter(is_archived=False)
    if center:
        groups = groups.filter(center=center)

    teachers_qs = User.objects.filter(role="teacher", is_active=True)
    if center:
        teachers_qs = teachers_qs.filter(center=center)

    courses = Category.objects.all().only("id", "name")

    staffs = User.objects.filter(role__in=["manager", "admin", "director"], is_active=True)
    if center:
        staffs = staffs.filter(center=center)

    uz_months = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
    ]

    history_month_value = cur_month_start.strftime("%Y-%m")
    if sel_month and sel_month.isdigit():
        history_month_value = f"{cur_year}-{int(sel_month):02d}"

    query_params = request.GET.copy()
    query_params.pop("page", None)

    return {
        "center": center,
        "page_rows": display_rows,
        "filtered_payments": filtered_payments,
        "payment_record_count": len(filtered_payments),
        "total_income": total_income,
        "filtered_income": filtered_income,
        "chart_data": chart_data,
        "chart_labels": chart_labels,
        "chart_kicker": chart_kicker,
        "chart_period_label": chart_period_label,
        "selected_period_label": _human_period_label(selected_from, selected_to),
        "chart_payment_record_count": chart_payment_record_count,
        "chart_unique_payers_count": chart_unique_payers_count,
        "unique_payers_count": unique_payers_count,
        "groups": groups,
        "teachers": teachers_qs,
        "courses": courses,
        "staffs": staffs,
        "uz_months": uz_months,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "sel_group": sel_group,
        "sel_teacher": sel_teacher,
        "sel_course": sel_course,
        "sel_staff": sel_staff,
        "sel_type": sel_type,
        "sel_month": sel_month,
        "history_month_value": history_month_value,
        "query_string": query_params.urlencode(),
        "selected_from": selected_from,
        "selected_to": selected_to,
    }



def _weekly_t_to_min(t):
    return t.hour * 60 + t.minute



def _weekly_teacher_initials(teacher):
    if not teacher:
        return "?"
    name = (teacher.get_full_name() or "").strip()
    if not name:
        name = (getattr(teacher, "username", "") or "").strip()
    parts = [p for p in name.split() if p]
    if len(parts) >= 2:
        return (parts[0][:1] + parts[1][:1]).upper()
    if parts:
        return parts[0][:2].upper()
    return "??"



def _weekly_schedule_excel(center, weekday_labels, week_map, time_slots, time_grid,
                           teacher_loads, teacher_id, room_filter, teachers):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    cell_font = Font(size=10)
    bold_small = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws = wb.active
    ws.title = "Vaqt jadvali"
    headers = ["Vaqt"] + [lbl for _, lbl in weekday_labels]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border

    for row_idx, slot in enumerate(time_slots, start=2):
        row_values = [slot["label"]]
        for wd, _ in weekday_labels:
            cell_items = time_grid[wd][row_idx - 2]
            if cell_items:
                lines = []
                for entry in cell_items:
                    it = entry["item"]
                    tch_name = it.group.oqituvchi.get_full_name() if it.group.oqituvchi else "Belgilanmagan"
                    rm = it.room or "—"
                    lines.append(f"{it.group.nom}\n{tch_name}\n{it.time_range} | {rm}")
                row_values.append("\n────\n".join(lines))
            else:
                row_values.append("")
        ws.append(row_values)
        for col_idx in range(1, len(row_values) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = center_align
            c.border = border
            if col_idx == 1:
                c.font = bold_small
                c.fill = PatternFill("solid", fgColor="F1F5F9")
            else:
                c.font = cell_font

    ws.column_dimensions["A"].width = 11
    for letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 28
    for row_idx in range(2, len(time_slots) + 2):
        ws.row_dimensions[row_idx].height = 56
    ws.freeze_panes = "B2"

    ws2 = wb.create_sheet("Kun bo'yicha")
    ws2.append(["Kun", "Guruh", "O'qituvchi", "Vaqt", "Xona", "Bo'lim"])
    for col_idx in range(1, 7):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border
    for wd, label in weekday_labels:
        items = week_map.get(wd) or []
        if not items:
            ws2.append([label, "—", "—", "—", "—", "—"])
            continue
        for it in items:
            try:
                cat = it.group.get_category_display()
            except Exception:
                cat = ""
            ws2.append([
                label,
                it.group.nom,
                it.group.oqituvchi.get_full_name() if it.group.oqituvchi else "Belgilanmagan",
                it.time_range,
                it.room or "Xona yo'q",
                cat,
            ])
    for col, w in zip(["A", "B", "C", "D", "E", "F"], [14, 28, 24, 16, 14, 18]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("O'qituvchi yuklamasi")
    ws3.append(["O'qituvchi", "Darslar", "Soatlar (haftada)", "Band kunlar", "Xonalar"])
    for col_idx in range(1, 6):
        c = ws3.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = border
    for ld in teacher_loads:
        name = ld["teacher"].get_full_name() if ld["teacher"] else "Belgilanmagan"
        ws3.append([
            name,
            ld["lessons"],
            ld["hours"],
            ", ".join(ld["days_short"]),
            ", ".join(ld["rooms"]),
        ])
    for col, w in zip(["A", "B", "C", "D", "E"], [28, 12, 18, 18, 28]):
        ws3.column_dimensions[col].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"haftalik_jadval_{timezone.localdate().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response



def _weekly_schedule_pdf(center, weekday_labels, time_slots, time_grid,
                         teacher_loads, teacher_id, room_filter, teachers):
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    from html import escape as _esc

    buf = BytesIO()
    page_size = landscape(A3)
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title="Haftalik jadval",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, textColor=rl_colors.grey, spaceAfter=8)
    cell_dark = ParagraphStyle("celld", parent=styles["Normal"], fontSize=7, leading=9, alignment=1, textColor=rl_colors.HexColor("#0f172a"))

    elements = []
    center_name = (
        getattr(center, "nom", None)
        or getattr(center, "name", None)
        or getattr(center, "title", None)
        or "ChaqmoqApp"
    )
    elements.append(Paragraph(f"Haftalik Jadval — {_esc(str(center_name))}", title_style))
    sub_parts = []
    if teacher_id:
        sel = next((t for t in teachers if t.id == teacher_id), None)
        if sel:
            sub_parts.append(f"O'qituvchi: {_esc(sel.get_full_name())}")
    if room_filter:
        sub_parts.append(f"Xona: {_esc(room_filter)}")
    sub_parts.append(f"Sana: {timezone.localdate().strftime('%d.%m.%Y')}")
    elements.append(Paragraph(" • ".join(sub_parts), sub_style))

    head_row = ["Vaqt"] + [lbl for _, lbl in weekday_labels]
    data = [head_row]
    for r_idx, slot in enumerate(time_slots, start=1):
        row = [slot["label"]]
        for c_idx, (wd, _) in enumerate(weekday_labels, start=1):
            entries = time_grid[wd][r_idx - 1]
            if entries:
                first = entries[0]
                it = first["item"]
                tch_name = _esc(first["teacher_name"])
                rm = _esc(it.room) if it.room else "—"
                extra = f"<br/><i>+{len(entries) - 1} qo'shimcha dars</i>" if len(entries) > 1 else ""
                txt = (
                    f"<b>{tch_name}</b><br/>"
                    f"{_esc(it.group.nom)}<br/>"
                    f"<font color='#475569'>{_esc(it.time_range)} · {rm}</font>{extra}"
                )
                row.append(Paragraph(txt, cell_dark))
            else:
                row.append("")
        data.append(row)

    col_count = len(head_row)
    page_w = page_size[0] - 20 * mm
    time_col_w = 20 * mm
    day_col_w = (page_w - time_col_w) / (col_count - 1)
    col_widths = [time_col_w] + [day_col_w] * (col_count - 1)

    if len(data) == 1:
        elements.append(Paragraph("Hech qanday dars vaqti kiritilmagan.", sub_style))
    else:
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CBD5E1")),
            ("FONTSIZE", (0, 1), (0, -1), 8),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), rl_colors.HexColor("#F1F5F9")),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
        tbl.setStyle(style)
        elements.append(tbl)

    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph("O'qituvchi yuklamasi (haftalik)", title_style))

    load_data = [["O'qituvchi", "Darslar", "Soat", "Band kunlar", "Xonalar"]]
    for ld in teacher_loads:
        name = ld["teacher"].get_full_name() if ld["teacher"] else "Belgilanmagan"
        load_data.append([
            _esc(name),
            str(ld["lessons"]),
            f"{ld['hours']}",
            ", ".join(ld["days_short"]) or "—",
            _esc(", ".join(ld["rooms"])) or "—",
        ])
    load_tbl = Table(load_data, colWidths=[60 * mm, 25 * mm, 25 * mm, 50 * mm, 80 * mm], repeatRows=1)
    load_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (2, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(load_tbl)

    doc.build(elements)
    pdf_data = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf_data, content_type="application/pdf")
    fname = f"haftalik_jadval_{timezone.localdate().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response



class CategoryForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        self.center = kwargs.pop("center", None)
        super().__init__(*args, **kwargs)

    class Meta:
        model = Category
        fields = ["name", "image", "description"]
        widgets = {
            "name": forms.TextInput(attrs={
                "class": "form-control",
                "placeholder": "Masalan: Dizayn"
            }),
            "image": forms.ClearableFileInput(attrs={
                "class": "form-control",
            }),
            "description": forms.Textarea(attrs={
                "class": "form-control",
                "rows": 3,
                "placeholder": "Bo'lim haqida qisqa izoh"
            }),
        }

    def clean_name(self):
        name = self.cleaned_data.get("name")
        if name:
            name_stripped = name.strip()
            # We check Category.all_objects because soft-deleted categories
            # still trigger unique constraints on the DB level.
            qs = Category.all_objects.filter(name__iexact=name_stripped, center=self.center)
            if self.instance and self.instance.pk:
                qs = qs.exclude(pk=self.instance.pk)
            if qs.exists():
                raise forms.ValidationError("Ushbu nomdagi bo'lim allaqachon mavjud!")
            return name_stripped
        return name



def _all_groups_avatar(name: str):
    """Return (initials, color) for a group avatar based on its name."""
    palette = [
        ("#2563eb", "#dbeafe"),
        ("#7c3aed", "#ede9fe"),
        ("#10b981", "#d1fae5"),
        ("#d97706", "#fef3c7"),
        ("#dc2626", "#fee2e2"),
        ("#0ea5e9", "#e0f2fe"),
        ("#db2777", "#fce7f3"),
        ("#0d9488", "#ccfbf1"),
    ]
    safe = (name or "G").strip()
    if not safe:
        safe = "G"
    initials = safe[:2].upper()
    idx = sum(ord(ch) for ch in safe) % len(palette)
    color, bg = palette[idx]
    return initials, color, bg



def _all_groups_schedule_text(group_id: int, schedule_map: dict) -> str:
    rows = schedule_map.get(group_id) or []
    if not rows:
        return "—"
    days = []
    seen = set()
    time_text = ""
    for row in rows:
        wd = row["weekday"]
        if wd in seen:
            continue
        seen.add(wd)
        days.append(_WEEKDAY_SHORT.get(wd, str(wd)))
        if not time_text and row.get("start_time"):
            time_text = row["start_time"].strftime("%H:%M")
    base = " · ".join(days) if days else "—"
    if time_text:
        return f"{base} · {time_text}"
    return base



def _compute_teacher_salary_list_payload(year, month, center, support_feature_on):
    """teacher_salary_list — og'ir hisoblash qismi cache'lanadi."""
    from education.services.support_teacher import (
        list_support_user_ids,
        calculate_support_salary,
    )
    from education.models import FinancialMonth
    from education.services.historical_finance_service import HistoricalFinanceService

    teacher_qs = User.objects.filter(role="teacher")
    if center:
        teacher_qs = teacher_qs.filter(center=center)
    teacher_ids = set(teacher_qs.values_list("id", flat=True))

    extra_user_ids = set()
    if support_feature_on:
        extra_user_ids = list_support_user_ids(center=center) - teacher_ids

    all_ids = teacher_ids | extra_user_ids
    users_qs = User.objects.filter(id__in=all_ids).order_by("ism", "familya")
    support_user_ids_set = list_support_user_ids(center=center) if support_feature_on else set()

    fin_month = FinancialMonth.objects.filter(year=year, month=month, center=center).first()
    is_closed = fin_month.is_closed if fin_month else False

    teacher_rows = []
    total_all = 0
    teacher_id_list = []

    for t in users_qs:
        teacher_salary = 0
        teacher_groups = 0
        if t.id in teacher_ids:
            salary_data = HistoricalFinanceService.calculate_teacher_salary(t, year, month, center)
            teacher_salary = salary_data['salary']
            teacher_groups = len(salary_data['details'])

        support_salary = 0
        support_groups = 0
        is_support = False
        if support_feature_on and t.id in support_user_ids_set:
            sup = calculate_support_salary(t, year, month, center)
            support_salary = sup["salary"]
            support_groups = len(sup["details"])
            is_support = True

        combined = teacher_salary + support_salary
        if combined == 0 and t.id not in teacher_ids and not is_support:
            continue

        total_all += combined
        teacher_id_list.append(t.id)
        teacher_rows.append({
            "teacher_id": t.id,
            "month_salary": combined,
            "teacher_salary": teacher_salary,
            "support_salary": support_salary,
            "groups_count": teacher_groups + support_groups,
            "is_support": is_support,
            "is_main_teacher": t.id in teacher_ids,
        })

    return {
        'teacher_rows': teacher_rows,
        'teacher_id_list': teacher_id_list,
        'total_all': total_all,
        'is_closed': is_closed,
    }



def _compute_teacher_salary_summary_payload(
    request, center, support_feature_on, selected_year, selected_month,
    list_support_user_ids, calculate_support_salary,
):
    """teacher_salary_summary'ning og'ir hisoblash qismi — cache'lanadi."""
    # Asosiy o'qituvchilar
    teacher_user_qs = User.objects.filter(role="teacher", is_archived=False)
    if center:
        teacher_user_qs = teacher_user_qs.filter(center=center)
    teacher_ids = set(teacher_user_qs.values_list("id", flat=True))

    # Support sifatida biriktirilgan boshqa rolli xodimlar
    extra_user_ids = set()
    if support_feature_on:
        extra_user_ids = list_support_user_ids(center=center) - teacher_ids

    all_ids = teacher_ids | extra_user_ids
    users_qs = User.objects.filter(id__in=all_ids, is_archived=False).order_by("ism", "familya", "id")
    support_user_ids_set = list_support_user_ids(center=center) if support_feature_on else set()

    # ── Batch-compute closed months once for all teachers ──────────
    _fin_months_qs = FinancialMonth.objects.filter(year=selected_year, is_closed=True)
    if center:
        _fin_months_qs = _fin_months_qs.filter(center=center)
    _closed_months_cache = {}
    for _fm in _fin_months_qs:
        _closed_months_cache[_fm.month] = _fm

    # ── Batch-fetch all salary snapshots for all teachers at once ──
    _snapshots_by_teacher = {}
    if _closed_months_cache and all_ids:
        for _snap in TeacherSalarySnapshot.objects.filter(
            teacher_id__in=all_ids,
            financial_month__in=_closed_months_cache.values(),
        ).select_related("financial_month"):
            _snapshots_by_teacher.setdefault(_snap.teacher_id, []).append(_snap)

    # ── N+1 yo'q qilish: har teacher uchun group_set.count() o'rniga
    # bir martalik aggregate query bilan barcha teacher'larning guruh sonini olamiz.
    from django.db.models import Count, Q as _Q
    main_groups_count_map = dict(
        Group.objects.filter(oqituvchi_id__in=teacher_ids, is_archived=False)
        .values('oqituvchi_id')
        .annotate(c=Count('id'))
        .values_list('oqituvchi_id', 'c')
    )
    support_groups_count_map = {}
    if support_feature_on and support_user_ids_set:
        support_groups_count_map = dict(
            Group.objects.filter(
                support_teacher_id__in=support_user_ids_set,
                is_archived=False,
                support_foiz__gt=0,
            )
            .values('support_teacher_id')
            .annotate(c=Count('id'))
            .values_list('support_teacher_id', 'c')
        )

    # ── PERF: Batch-fetch all groups/attendance/history for all teachers ──────
    # Har teacher × har ochiq oy uchun alohida _teacher_groups / _attendance_lookup /
    # _history_lookup chaqiruvi o'rniga bir martalik query — N+1 ni yo'q qiladi.
    from django.db.models import Prefetch as _Prefetch
    _all_main_groups: list = []
    _groups_by_teacher: dict = {}
    if teacher_ids:
        _grp_qs = Group.objects.filter(oqituvchi_id__in=teacher_ids, is_archived=False)
        if center:
            _grp_qs = _grp_qs.filter(center=center)
        _all_main_groups = list(
            _grp_qs.prefetch_related(
                _Prefetch(
                    "enrollments",
                    queryset=Enrollment.all_objects.select_related("student"),
                    to_attr="all_enrollments",
                )
            )
        )
        for _g in _all_main_groups:
            _groups_by_teacher.setdefault(_g.oqituvchi_id, []).append(_g)

    _all_main_group_ids = [_g.id for _g in _all_main_groups]

    _yearly_att: dict = {}
    if _all_main_group_ids:
        for _row in Attendance.objects.filter(
            group_id__in=_all_main_group_ids,
            date__year=selected_year,
        ).filter(HistoricalFinanceService._billable_attendance_filter()).values(
            "group_id", "student_id", "date__month", "date__day"
        ):
            _gid, _sid, _m, _d = (
                _row["group_id"], _row["student_id"], _row["date__month"], _row["date__day"]
            )
            _yearly_att.setdefault(_gid, {}).setdefault(_m, {}).setdefault(_sid, []).append(_d)

    _all_main_history = (
        HistoricalFinanceService._history_lookup(_all_main_group_ids)
        if _all_main_group_ids else {}
    )

    # ================================
    # Grafik uchun bo'sh massivlar (12 oy)
    # ================================
    chart_teacher_income = [0] * 12
    chart_center_income = [0] * 12
    chart_total_turnover = [0] * 12

    # ================================
    # 3) HISOB-KITOB (HistoricalFinanceService + support)
    # ================================
    teacher_data = []

    for teacher in users_qs:
        salary_main_year = 0
        lessons_main = 0
        profit_main = 0
        turnover_main = 0
        salary_support_year = 0
        lessons_support = 0

        # Asosiy o'qituvchi sifatida
        if teacher.id in teacher_ids:
            yearly_stats = HistoricalFinanceService.get_yearly_teacher_stats(
                teacher, selected_year, center,
                _closed_months=_closed_months_cache,
                _snapshots=_snapshots_by_teacher.get(teacher.id, []),
                _teacher_groups=_groups_by_teacher.get(teacher.id, []),
                _yearly_att=_yearly_att,
                _history=_all_main_history,
            )
            for m in range(12):
                chart_teacher_income[m] += yearly_stats[m]['salary']
                chart_center_income[m] += yearly_stats[m]['center_profit']
                chart_total_turnover[m] += yearly_stats[m]['turnover']
            m_stat = yearly_stats[selected_month - 1]
            salary_main_year = m_stat['salary']
            lessons_main = m_stat['lessons']
            profit_main = m_stat['center_profit']
            turnover_main = m_stat['turnover']

        # Support sifatida (faqat tanlangan oy uchun)
        is_support = False
        if support_feature_on and teacher.id in support_user_ids_set:
            sup = calculate_support_salary(teacher, selected_year, selected_month, center)
            salary_support_year = sup['salary']
            lessons_support = sup['attendance_count']
            is_support = True
            # Grafikga support ulushini ham asosiy oylik (teacher_income) qatoriga qo'shamiz
            # — chunki bu jami xodim daromadlari bo'yicha umumiy ko'rsatkich.
            chart_teacher_income[selected_month - 1] += salary_support_year

        # N+1 fix: aggregate'dan o'qiymiz, har teacher uchun query yo'q
        groups_count = main_groups_count_map.get(teacher.id, 0)
        if is_support:
            groups_count += support_groups_count_map.get(teacher.id, 0)

        combined = salary_main_year + salary_support_year
        if combined == 0 and teacher.id not in teacher_ids and not is_support:
            continue

        teacher_data.append({
            "id": teacher.id,
            "teacher": teacher.get_full_name() or teacher.email,
            "groups": groups_count,
            "lessons": lessons_main + lessons_support,
            "teacher_income": int(combined),
            "center_profit": int(profit_main),
            "total_turnover": int(turnover_main),
            "is_support": bool(is_support),
            "is_main_teacher": teacher.id in teacher_ids,
            "support_income": int(salary_support_year),
        })

    # PERF: keshga saqlash uchun qaytaramiz (asosiy view o'qib render qiladi)
    return {
        "teacher_data": teacher_data,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
    }



# ================================
#   RENDER HELPER
# ================================
def _render_salary(request, selected_year, selected_month,
                   teacher_data, chart_labels,
                   chart_teacher_income, chart_center_income, chart_total_turnover):

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "year": selected_year,
            "month": selected_month,
            "teacher_data": teacher_data,
            "chart_teacher_income": chart_teacher_income,
            "chart_center_income": chart_center_income,
            "chart_total_turnover": chart_total_turnover,
        })

    return render(request, "education/teacher_salary_summary.html", {
        "years": list(range(2024, 2036)),
        "months": [
            (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
            (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
            (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
        ],
        "selected_year": selected_year,
        "selected_month": selected_month,
        "teacher_data": teacher_data,
        "chart_labels": chart_labels,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
    })



def _director_or_manager(user):
    return user.is_superuser or getattr(user, "role", None) in ("director", "manager")



def _teacher_can_view_settings(user):
    return user.is_superuser or getattr(user, "role", None) in ("director", "manager", "teacher")



def _teacher_or_management_can_access_group(user, group: Group):
    if user.is_superuser or getattr(user, "role", None) in ("director", "manager"):
        return True
    if getattr(user, "role", None) == "teacher":
        return group.oqituvchi_id == user.id or group.support_teacher_id == user.id
    return False



def _decode_exam_session_note(raw_text: str) -> dict:
    """
    Backward-compatible parser:
    - yangi format: {"task": "...", "comment": "..."} (JSON)
    - eski format: oddiy text (task sifatida olinadi)
    """
    raw_text = (raw_text or "").strip()
    if not raw_text:
        return {"task": "", "comment": ""}
    try:
        payload = json.loads(raw_text)
        if isinstance(payload, dict):
            return {
                "task": (payload.get("task") or "").strip(),
                "comment": (payload.get("comment") or "").strip(),
            }
    except Exception:
        pass
    return {"task": raw_text, "comment": ""}



def _encode_exam_session_note(task: str, comment: str) -> str:
    task = (task or "").strip()
    comment = (comment or "").strip()
    if not task and not comment:
        return ""
    return json.dumps({"task": task, "comment": comment}, ensure_ascii=False)



def _exam_entry_url(session_id: int, max_score) -> str:
    url = reverse("education:exam_session_entry", kwargs={"session_id": session_id})
    if str(max_score or "100") != "100":
        return f"{url}?{urlencode({'max_score': str(max_score)})}"
    return url

# Export helpers (incl. _private) for domain modules using `import *`.
__all__ = [k for k in globals() if not k.startswith('__')]
