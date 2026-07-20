"""
Auto-split from education/views.py (phase 7 god-file reduction).
Public API re-exported via education.views package.
"""
from __future__ import annotations

from .common import *  # noqa: F403


@login_required
@require_feature("finance")
def teacher_salary_list(request):
    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)

    if month < 1 or month > 12:
        month = now.month

    month_names_uz = [
        "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"
    ]
    month_name = month_names_uz[month - 1]

    from core.tenant import get_request_center
    from core.perf_cache import TTL_LONG, perf_cache_get_or_set, versioned_cache_key
    center = get_request_center(request)

    from education.services.support_teacher import is_support_enabled
    support_feature_on = is_support_enabled(center)

    # PERF: og'ir hisoblash 15 daqiqa cache (per markaz + yil + oy).
    _cache_key = versioned_cache_key(
        "salary_list", getattr(center, 'id', None), year, month
    )

    def _compute():
        return _compute_teacher_salary_list_payload(year, month, center, support_feature_on)

    payload = perf_cache_get_or_set(_cache_key, _compute, ttl=TTL_LONG)

    # `teachers` payload'da serializable obyekt — User instance'larini qayta yuklaymiz
    # (Cache'da User instance saqlash xatarli — id-list orqali fetch qilamiz).
    teacher_id_list = payload['teacher_id_list']
    user_index = {u.id: u for u in User.objects.filter(id__in=teacher_id_list)}
    teachers_resolved = []
    for row in payload['teacher_rows']:
        u = user_index.get(row['teacher_id'])
        if not u:
            continue
        teachers_resolved.append({**row, 'teacher': u})

    return render(request, "education/teacher_salary_list.html", {
        "teachers": teachers_resolved,
        "year": year,
        "month": month,
        "month_name": month_name,
        "total_all": payload['total_all'],
        "is_closed": payload['is_closed'],
        "support_feature_on": support_feature_on,
    })



# 🔹 Excel Export — O'qituvchi oyligi hisoboti
@login_required
@require_feature("finance")
def teacher_salary_export(request):
    """
    Tanlangan oy/yil bo'yicha barcha o'qituvchilar oylik hisobotini
    professional Excel (.xlsx) fayl sifatida yuklab beradi.

    Sheet 1: Umumiy hisobot (barchasi + bar chart)
    Sheet 2..N: Har bir o'qituvchi uchun alohida — guruh va o'quvchi kesimida.
    """
    import io
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    from core.tenant import get_request_center
    from education.services.historical_finance_service import HistoricalFinanceService

    # ── Parametrlar ──────────────────────────────────────────────────────────
    now   = timezone.localdate()
    year  = _get_int(request.GET, "year",  now.year)
    month = _get_int(request.GET, "month", now.month)
    if month < 1 or month > 12:
        month = now.month

    center = get_request_center(request)

    MONTH_NAMES = [
        "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"
    ]
    month_name = MONTH_NAMES[month - 1]
    period_label = f"{month_name} {year}"

    # ── O'qituvchilar va oylik ma'lumot ──────────────────────────────────────
    teacher_qs = User.objects.filter(role="teacher")
    if center:
        teacher_qs = teacher_qs.filter(center=center)
    teachers = list(teacher_qs.order_by("ism"))

    salary_rows = []
    for t in teachers:
        data = HistoricalFinanceService.calculate_teacher_salary(t, year, month, center)
        salary_rows.append({
            "teacher": t,
            "salary":  data["salary"],
            "details": data.get("details", []),
        })

    # ── Stil yordamchilari ───────────────────────────────────────────────────
    def _hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(style="thin"):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def _bold(size=11, color="000000"):
        return Font(bold=True, size=size, color=color)

    def _money_fmt():
        return "#,##0"

    def _auto_width(ws, extra=4):
        for col in ws.columns:
            mx = 0
            for cell in col:
                try:
                    mx = max(mx, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = mx + extra

    # ── Minimalist yorqin rang sxemasi ──────────────────────────────────────
    # Asosiy: ko'k (2563EB) | Guruh: yashil (0EA472) | Jami: sariq (F59E0B)
    # Fon: oq (FFFFFF) | Alt qator: och kulrang (F8FAFC) | Chegara: kulrang (CBD5E1)
    HDR_DARK   = _hdr_fill("2563EB")   # bosh sarlavha — chuqur ko'k
    HDR_BLUE   = _hdr_fill("3B82F6")   # ustun header — yorqin ko'k
    HDR_GREEN  = _hdr_fill("0EA472")   # guruh blok — yashil
    TOTAL_FILL = _hdr_fill("FEF3C7")   # jami satri — sariq fon
    ALT_FILL   = _hdr_fill("F1F5F9")   # juft qatorlar — och kulrang
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    DARK_FONT  = Font(color="1E293B", size=10)
    TOTAL_FONT = Font(bold=True, color="92400E", size=11)   # jami — to'q jigarrang
    MONEY_NUM  = _money_fmt()

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — UMUMIY HISOBOT
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Umumiy hisobot"
    ws1.sheet_view.showGridLines = False

    # Sarlavha
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = f"O'qituvchilar Oylik Hisoboti — {period_label}"
    title_cell.font  = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill  = HDR_DARK
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    # Ustun sarlavhalari
    headers1 = ["T/r", "O'qituvchi", "Guruhlar soni", "O'quvchilar soni",
                 "Qatnashuv (dars)", "Hisoblangan oylik (so'm)", "Izoh"]
    ws1.append([])  # bo'sh qator (row 2)
    ws1.append(headers1)  # row 3
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=ci)
        cell.value  = h
        cell.font   = WHITE_FONT
        cell.fill   = HDR_BLUE
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[3].height = 28

    total_sum = 0
    chart_names    = []
    chart_salaries = []

    for idx, row in enumerate(salary_rows, 1):
        t       = row["teacher"]
        details = row["details"]
        students_total = sum(len(d.get("enrollments", [])) for d in details)
        attend_total   = sum(d.get("attendance", 0) for d in details)
        salary         = row["salary"]
        total_sum     += salary

        name = t.get_full_name() or t.email
        chart_names.append(name[:20])
        chart_salaries.append(salary)

        data_row = [idx, name, len(details), students_total, attend_total, salary, ""]
        ws1.append(data_row)
        ri = ws1.max_row
        # Juft — och kulrang, toq — oq
        fill = ALT_FILL if idx % 2 == 0 else _hdr_fill("FFFFFF")
        for ci, val in enumerate(data_row, 1):
            cell = ws1.cell(row=ri, column=ci)
            cell.fill   = fill
            cell.border = _border()
            cell.font   = Font(color="334155", size=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci in (1,3,4,5) else "left")
            if ci == 6:
                cell.number_format = MONEY_NUM
                cell.font = Font(color="1D4ED8", size=10, bold=True)

    # JAMI SATRI
    jami_ri = ws1.max_row + 1
    ws1.cell(row=jami_ri, column=1).value = "JAMI"
    ws1.cell(row=jami_ri, column=6).value = total_sum
    ws1.cell(row=jami_ri, column=6).number_format = MONEY_NUM
    for ci in range(1, 8):
        cell = ws1.cell(row=jami_ri, column=ci)
        cell.fill   = TOTAL_FILL
        cell.font   = TOTAL_FONT
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center" if ci == 1 else "left", vertical="center")
    ws1.cell(row=jami_ri, column=6).font = Font(bold=True, color="92400E", size=11)
    ws1.row_dimensions[jami_ri].height = 26

    # BAR CHART
    if chart_names:
        chart = BarChart()
        chart.type   = "col"
        chart.title  = f"O'qituvchilar oyligi — {period_label}"
        chart.y_axis.title = "Oylik (so'm)"
        chart.x_axis.title = "O'qituvchi"
        chart.style  = 10
        chart.width  = 28
        chart.height = 16

        data_ref = Reference(ws1,
                             min_col=6, max_col=6,
                             min_row=3, max_row=3 + len(salary_rows) - 1)
        cats_ref = Reference(ws1,
                             min_col=2, max_col=2,
                             min_row=4, max_row=3 + len(salary_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws1.add_chart(chart, f"A{jami_ri + 2}")

    ws1.freeze_panes = "A4"
    _auto_width(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2..N — HAR BIR O'QITUVCHI
    # ════════════════════════════════════════════════════════════════════════
    for row in salary_rows:
        t       = row["teacher"]
        details = row["details"]
        salary  = row["salary"]

        # Sheet nomi: Excel 31 belgidan oshmasin, noto'g'ri belgilar yo'qolsin
        raw_name   = t.get_full_name() or t.email or f"Ustoz_{t.pk}"
        sheet_name = raw_name[:28].translate(
            str.maketrans(r'\/*?:[]', '_______')
        )
        # Takrorlanmaslik uchun son qo'shamiz
        base = sheet_name
        cnt  = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base[:25]}_{cnt}"
            cnt += 1

        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        # ── HEADER blok ─────────────────────────────────────────────────────
        students_total = sum(len(d.get("enrollments", [])) for d in details)
        attend_total   = sum(d.get("attendance", 0) for d in details)

        ws.merge_cells("A1:F1")
        ws["A1"].value = f"O'qituvchi: {raw_name}  |  Davr: {period_label}"
        ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill  = HDR_DARK
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        meta_rows = [
            ("Guruhlar soni",          len(details)),
            ("Jami o'quvchilar",        students_total),
            ("Jami dars (qatnashuv)",   attend_total),
            ("Hisoblangan jami oylik",  salary),
        ]
        for mi, (label, val) in enumerate(meta_rows, 2):
            lbl_cell = ws.cell(row=mi, column=1)
            lbl_cell.value = label
            lbl_cell.font  = Font(bold=True, color="64748B", size=10)
            lbl_cell.fill  = _hdr_fill("F8FAFC")

            val_cell = ws.cell(row=mi, column=2)
            val_cell.value = val
            val_cell.font  = Font(color="1E293B", size=10)
            val_cell.fill  = _hdr_fill("F8FAFC")
            if mi == 5:  # oylik satri
                val_cell.number_format = MONEY_NUM
                val_cell.font = Font(bold=True, color="1D4ED8", size=11)

        cur_row = 7  # guruh bloklari shu satrdan boshlanadi

        if not details:
            ws.cell(row=cur_row, column=1).value = "Bu oy uchun ma'lumot yo'q."
            ws.cell(row=cur_row, column=1).font  = Font(color="94A3B8", italic=True)
        else:
            for gd in details:
                gname       = gd.get("group_name", "Guruh")
                g_salary    = gd.get("salary", 0)
                g_attend    = gd.get("attendance", 0)
                enrollments = gd.get("enrollments", [])

                # ── Guruh sarlavhasi ─────────────────────────────────────────
                ws.merge_cells(start_row=cur_row, start_column=1,
                               end_row=cur_row, end_column=6)
                hdr = ws.cell(row=cur_row, column=1)
                hdr.value = f"  {gname}   |   Guruh daromadi: {g_salary:,} som   |   Dars: {g_attend} marta"
                hdr.font  = Font(bold=True, color="FFFFFF", size=11)
                hdr.fill  = HDR_GREEN
                hdr.border = _border()
                hdr.alignment = Alignment(vertical="center", horizontal="left")
                ws.row_dimensions[cur_row].height = 26
                cur_row += 1

                # ── O'quvchilar jadval sarlavhasi ────────────────────────────
                sub_hdrs = ["T/r", "O'quvchi", "Kurs narhi", "Qatnashuv (kun)",
                            "Daromad (so'm)", "Izoh"]
                for ci, h in enumerate(sub_hdrs, 1):
                    cell = ws.cell(row=cur_row, column=ci)
                    cell.value  = h
                    cell.font   = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill   = HDR_BLUE
                    cell.border = _border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[cur_row].height = 22
                cur_row += 1

                # ── O'quvchilar satrlari ──────────────────────────────────────
                for si, en in enumerate(enrollments, 1):
                    sname   = en.get("student_name", "Noma'lum")
                    kn      = en.get("kurs_narhi", 0)
                    att     = en.get("attended", 0)
                    daromad = en.get("daromad", 0)

                    # Juft — och kulrang, toq — oq
                    fill = ALT_FILL if si % 2 == 0 else _hdr_fill("FFFFFF")
                    data = [si, sname, kn, att, daromad, ""]
                    for ci, v in enumerate(data, 1):
                        cell = ws.cell(row=cur_row, column=ci)
                        cell.value  = v
                        cell.fill   = fill
                        cell.border = _border()
                        cell.font   = Font(color="334155", size=10)
                        cell.alignment = Alignment(vertical="center",
                                                   horizontal="center" if ci in (1,4) else "left")
                        if ci == 3:
                            cell.number_format = MONEY_NUM
                            cell.font = Font(color="475569", size=10)
                        if ci == 5:
                            cell.number_format = MONEY_NUM
                            cell.font = Font(color="1D4ED8", size=10, bold=True)
                    cur_row += 1

                # ── Guruh jami satri ─────────────────────────────────────────
                for ci in range(1, 7):
                    cell = ws.cell(row=cur_row, column=ci)
                    cell.fill   = TOTAL_FILL
                    cell.border = _border()
                    cell.font   = TOTAL_FONT
                    cell.alignment = Alignment(vertical="center", horizontal="left")
                ws.cell(row=cur_row, column=2).value = "GURUH JAMI:"
                ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=cur_row, column=5).value = g_salary
                ws.cell(row=cur_row, column=5).number_format = MONEY_NUM
                ws.cell(row=cur_row, column=5).font = Font(bold=True, color="92400E", size=11)
                ws.row_dimensions[cur_row].height = 22
                cur_row += 2  # bo'sh qator + keyingi guruh

            # ── Umumiy jami (sheet pastida) ──────────────────────────────────
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=4)
            lbl = ws.cell(row=cur_row, column=1)
            lbl.value = "BARCHA GURUHLAR JAMI OYLIGI:"
            lbl.font  = Font(bold=True, color="FFFFFF", size=12)
            lbl.fill  = HDR_DARK
            lbl.border = _border()
            lbl.alignment = Alignment(horizontal="right", vertical="center")

            tot = ws.cell(row=cur_row, column=5)
            tot.value  = salary
            tot.font   = Font(bold=True, color="FFFFFF", size=13)
            tot.fill   = HDR_DARK
            tot.border = _border()
            tot.number_format = MONEY_NUM
            ws.row_dimensions[cur_row].height = 32

        ws.freeze_panes = "A7"
        _auto_width(ws)

    # ── Fayl qaytarish ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"oylik_hisobot_{year}_{month:02d}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



# 🔹 2. O'qituvchining barcha guruhlari
@login_required
@require_feature("finance")
def teacher_groups(request, teacher_id):
    from core.tenant import get_request_center
    from education.services.support_teacher import (
        is_support_enabled,
        calculate_support_salary,
        list_support_user_ids,
    )

    center = get_request_center(request)

    # Foydalanuvchi: o'qituvchi YOKI (markazda support feature yoqilgan bo'lsa)
    # support sifatida biriktirilgan istalgan xodim bo'lishi mumkin.
    base_qs = User.objects.all()
    if center:
        base_qs = base_qs.filter(center=center)

    candidate = base_qs.filter(id=teacher_id).first()
    if candidate is None:
        return get_object_or_404(User.objects.none(), id=teacher_id)

    is_main_teacher = candidate.role == "teacher"
    is_support_member = (
        is_support_enabled(center) and candidate.id in list_support_user_ids(center=center)
    )
    if not is_main_teacher and not is_support_member:
        return get_object_or_404(User.objects.filter(role="teacher"), id=teacher_id)

    teacher = candidate

    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)
    if month < 1 or month > 12:
        month = now.month

    years = list(range(now.year - 3, now.year + 4))

    from education.services.historical_finance_service import HistoricalFinanceService

    teacher_data = []
    teacher_salary_total = 0
    teacher_is_locked = False

    if is_main_teacher:
        salary_data = HistoricalFinanceService.calculate_teacher_salary(teacher, year, month, center)
        teacher_salary_total = salary_data['salary']
        teacher_is_locked = salary_data['is_locked']

        for gcd in salary_data['details']:
            group_obj = Group.objects.filter(id=gcd['group_id']).first()
            if not group_obj:
                continue

            enrollments = []
            for en in gcd.get('enrollments', []):
                enrollments.append({
                    "student_name": en.get('student_name', "Noma'lum"),
                    "kurs_narhi": en.get('kurs_narhi', 0),
                    "foiz": en.get('foiz', 0),
                    "attended": en.get('attended', 0),
                    "daromad": en.get('daromad', 0),
                })

            teacher_data.append({
                "group": group_obj,
                "enrollments": enrollments,
                "foiz": gcd.get('fi', getattr(teacher, 'oqituvchi_foizi', 0) or group_obj.oqituvchi_foiz),
                "daromad": gcd['salary'],
                "students_count": len(enrollments),
                "is_support": False,
            })

    # ── Support sifatida ishlash (agar feature yoqilgan bo'lsa) ──
    support_salary_total = 0
    if is_support_member:
        sup = calculate_support_salary(teacher, year, month, center)
        support_salary_total = sup['salary']

        for gcd in sup['details']:
            group_obj = Group.objects.filter(id=gcd['group_id']).first()
            if not group_obj:
                continue

            enrollments = [
                {
                    "student_name": s.get('student_name', "Noma'lum"),
                    "kurs_narhi": int(getattr(group_obj, 'kurs_narxi', 0) or 0),
                    "foiz": gcd.get('fi', 0),
                    "attended": s.get('attended', 0),
                    "daromad": s.get('daromad', 0),
                }
                for s in gcd.get('students', [])
            ]
            teacher_data.append({
                "group": group_obj,
                "enrollments": enrollments,
                "foiz": gcd.get('fi', 0),
                "daromad": gcd['salary'],
                "students_count": len(enrollments),
                "is_support": True,
            })

    jami_umumiy_daromad = teacher_salary_total + support_salary_total

    return render(request, "education/teacher_groups.html", {
        "teacher": teacher,
        "teacher_data": teacher_data,
        "year": year,
        "month": month,
        "years": years,
        "jami_umumiy_daromad": jami_umumiy_daromad,
        "is_locked": teacher_is_locked,
        # Support meta
        "is_main_teacher": is_main_teacher,
        "is_support_member": is_support_member,
        "teacher_salary_total": teacher_salary_total,
        "support_salary_total": support_salary_total,
    })



@login_required
@require_feature("finance")
def teacher_salary_report(request, group_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = Group.objects.all()
    if center:
        qs = qs.filter(center=center)
    group = get_object_or_404(qs, id=group_id)
    
    now = timezone.localdate()
    year = _get_int(request.GET, "year", now.year)
    month = _get_int(request.GET, "month", now.month)
    
    from education.services.historical_finance_service import HistoricalFinanceService
    salary_data = HistoricalFinanceService.calculate_teacher_salary(group.oqituvchi, year, month, center)
    
    student_summaries = []
    teacher_total_income = 0
    
    for gcd in salary_data['details']:
        if gcd['group_id'] == group.id:
            teacher_total_income = gcd['salary']
            for en in gcd.get('enrollments', []):
                student_summaries.append({
                    "student_name": en.get('student_name', 'Noma\'lum'),
                    "attended": en.get('attended', 0),
                    "teacher_income": en.get('daromad', 0)
                })
            break

    ctx = {
        "group": group,
        "student_summaries": student_summaries,
        "teacher_total_income": teacher_total_income,
        "month": month,
        "year": year,
        "is_locked": salary_data['is_locked'],
    }
    return render(request, "education/teacher_salary_report.html", ctx)



@login_required
@require_feature("finance")
def teacher_salary_summary(request):
    """
    O'qituvchilar maoshi va markaz foydasini yil/oy bo'yicha hisoblaydi.
    - Attendance: present=True YOKI forced=True bo'lgan barcha darslar hisobga olinadi.
    """

    # ================================
    # Tanlangan yil / oy
    # ================================
    today = date.today()
    selected_year = int(request.GET.get("year") or today.year)
    selected_month = int(request.GET.get("month") or today.month)

    # Oylar ro'yxati
    months = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr"),
    ]
    chart_labels = [m[1] for m in months]

    # ================================
    # 2) O'qituvchilar va ularning hisob-kitobi (Yagona To'g'ri Manba)
    # ================================
    from core.tenant import get_request_center
    from core.perf_cache import (
        TTL_LONG, perf_cache_get_or_set, versioned_cache_key,
    )
    from education.services.support_teacher import (
        is_support_enabled,
        list_support_user_ids,
        calculate_support_salary,
    )
    center = get_request_center(request)
    support_feature_on = is_support_enabled(center)

    # ── PERF: Og'ir hisoblash 15 daqiqa cache (per markaz + yil + oy).
    # Cache key versiyali — attendance/payment o'zgarsa, invalidate orqali
    # bekor qilinadi (signals'da qo'shilishi mumkin).
    _cache_key = versioned_cache_key(
        "salary_sum", getattr(center, 'id', None), selected_year, selected_month
    )

    def _compute_salary_summary():
        return _compute_teacher_salary_summary_payload(
            request, center, support_feature_on, selected_year, selected_month,
            list_support_user_ids, calculate_support_salary,
        )

    payload = perf_cache_get_or_set(_cache_key, _compute_salary_summary, ttl=TTL_LONG)
    teacher_data = payload['teacher_data']
    chart_teacher_income = payload['chart_teacher_income']
    chart_center_income = payload['chart_center_income']
    chart_total_turnover = payload['chart_total_turnover']

    # ── /PERF cache ──

    # AJAX javob — keshlangan ma'lumotdan
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "year": int(selected_year),
            "month": int(selected_month),
            "teacher_data": teacher_data,
            "chart_labels": chart_labels,
            "chart_teacher_income": [float(x) for x in chart_teacher_income],
            "chart_center_income": [float(x) for x in chart_center_income],
            "chart_total_turnover": [float(x) for x in chart_total_turnover],
        })

    return render(request, "education/teacher_salary_summary.html", {
        "years": list(range(2024, 2036)),
        "months": months,
        "selected_year": int(selected_year),
        "selected_month": int(selected_month),
        "teacher_data": teacher_data,
        "chart_labels": chart_labels,
        "chart_teacher_income": chart_teacher_income,
        "chart_center_income": chart_center_income,
        "chart_total_turnover": chart_total_turnover,
        "teacher_data_json": json.dumps(teacher_data),
        "chart_labels_json": json.dumps(chart_labels),
        "chart_teacher_income_json": json.dumps([float(x) for x in chart_teacher_income]),
        "chart_center_income_json": json.dumps([float(x) for x in chart_center_income]),
        "chart_total_turnover_json": json.dumps([float(x) for x in chart_total_turnover]),
    })



@login_required
@require_feature("finance")
def teacher_salary_redirect(request):
    group = None

    # O'qituvchi bo'lsa — o'z guruhini topadi
    if request.user.role == "teacher":
        group = Group.objects.filter(oqituvchi=request.user).first()

    # Direktor yoki superuser bo'lsa — birinchi mavjud guruhni topadi
    elif request.user.role == "director" or request.user.is_superuser:
        group = Group.objects.first()

    # Agar topilmasa — xabar chiqar va qaytar
    if not group:
        messages.warning(request, "Hech qanday guruh topilmadi!")
        return redirect("education:groups_it")

    # Topilgan guruh bo'yicha maosh sahifasiga yo'naltirish
    return redirect("education:teacher_salary_report", group.id)



@login_required
def teacher_groups_view(request, teacher_id):
    from core.tenant import get_request_center
    center = get_request_center(request)
    qs = User.objects.filter(role="teacher")
    if center:
        qs = qs.filter(center=center)
    teacher = get_object_or_404(qs, id=teacher_id)

    groups = (
        teacher.group_set
        .prefetch_related('enrollments__student')
        .all()
    )

    teacher_data = []
    for group in groups:
        enrollments = group.enrollments.all()
        group_income = sum([enr.real_oqituvchi_daromadi() for enr in enrollments])
        group_info = {
            'name': group.name,
            'students': enrollments,
            'foiz': enrollments.first().oqituvchi_foiz if enrollments.exists() else 0,
            'daromad': group_income,
        }
        teacher_data.append(group_info)

    ctx = {
        "teacher": teacher,
        "teacher_data": teacher_data,
    }
    return render(request, "education/teacher_groups.html", ctx)



@login_required
@require_feature("finance")
def teacher_income_dashboard(request):
    """
    O'qituvchining shaxsiy daromadlari panelini ko'rsatadi.
    Snapshot tizimi yordamida o'tgan oylar ma'lumotlari muzlatilgan (immutable).
    """
    if request.user.role not in ['teacher', 'director', 'manager'] and not request.user.is_superuser:
        messages.error(request, "Bu bo'lim ushbu rol uchun emas.")
        return redirect('core:home')
        
    is_admin = request.user.role in ['director', 'manager'] or request.user.is_superuser
    
    # Agar admin bo'lsa va teacher_id berilgan bo'lsa - o'shani ko'ramiz
    from core.tenant import get_request_center, get_tenant_object_or_404
    from education.services.expected_income_service import calculate_expected_income

    center = get_request_center(request)
    teacher_id = request.GET.get('teacher_id')
    if is_admin and teacher_id:
        # Boshqa markaz o'qituvchisi IDOR dan himoyalangan.
        teacher = get_tenant_object_or_404(User, request, id=teacher_id, role="teacher")
    else:
        teacher = request.user

    today = timezone.localdate()
    selected_year = _get_int(request.GET, "year", today.year)
    selected_month = _get_int(request.GET, "month", today.month)

    # HistoricalFinanceService orqali ma'lumotlarni olish (snapshot yoki dinamik)
    salary_data = HistoricalFinanceService.calculate_teacher_salary(teacher, selected_year, selected_month, center)

    # Support teacher daromadini ham qo'shamiz (feature flag tekshiruvisiz —
    # biriktirilgan bo'lsa har doim ko'rsatish kerak)
    from education.services.support_teacher import (
        list_support_user_ids, calculate_support_salary, get_yearly_support_salary,
    )
    support_salary_data = None
    support_salary_total = 0
    is_support_member = teacher.id in list_support_user_ids(center=center)
    if is_support_member:
        support_salary_data = calculate_support_salary(teacher, selected_year, selected_month, center)
        support_salary_total = support_salary_data.get("salary", 0)
        salary_data = dict(salary_data)
        salary_data["salary"] = salary_data.get("salary", 0) + support_salary_total
        salary_data["support_details"] = support_salary_data.get("details", [])
        salary_data["support_salary"] = support_salary_total

    # Get all 12 months for the yearly chart efficiently
    monthly_income = HistoricalFinanceService.get_yearly_teacher_salary(teacher, selected_year, center)
    if is_support_member:
        # Support yillik daromadini asosiy daromad bilan qo'shamiz
        support_monthly = get_yearly_support_salary(teacher, selected_year, center)
        monthly_income = [m + s for m, s in zip(monthly_income, support_monthly)]
    total_year_income = sum(monthly_income)

    # Get daily breakdown for the selected month (now returned by the service)
    _, num_days = calendar.monthrange(selected_year, selected_month)
    main_daily = salary_data.get('daily_breakdown', [0] * 31)
    if is_support_member and support_salary_data:
        support_daily = support_salary_data.get('daily_breakdown', [0] * 31)
        combined_daily = [a + b for a, b in zip(main_daily, support_daily)]
    else:
        combined_daily = main_daily
    daily_income = combined_daily[:num_days]

    # Yearly labels for JS
    monthly_labels = ["Yan", "Fev", "Mar", "Apr", "May", "Iyun", "Iyul", "Avg", "Sen", "Okt", "Noy", "Dek"]
    daily_labels = [str(i) for i in range(1, num_days + 1)]

    # Selectors options
    years = range(today.year - 2, today.year + 2)
    months_list = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"),
        (5, "May"), (6, "Iyun"), (7, "Iyul"), (8, "Avgust"),
        (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr")
    ]
    months_dict = dict(months_list)

    # Prognoz: tanlangan oy uchun maksimal kutilgan va kelasi oy uchun
    current_expected = calculate_expected_income(teacher, selected_year, selected_month, center)
    if selected_month == 12:
        next_year, next_month = selected_year + 1, 1
    else:
        next_year, next_month = selected_year, selected_month + 1
    next_expected = calculate_expected_income(teacher, next_year, next_month, center)

    # Progress foiz: bu oy qanchasi yig'ildi (maksimaldan)
    current_max = current_expected.get("expected_income", 0)
    current_salary = salary_data.get("salary", 0)
    progress_pct = min(100, int(current_salary / current_max * 100)) if current_max > 0 else 0

    # Kelasi oy o'zgarish foizi (joriy oyga nisbatan)
    next_income = next_expected.get("expected_income", 0)
    if current_max > 0:
        delta_pct = round((next_income - current_max) / current_max * 100, 1)
    else:
        delta_pct = 0

    # Bar widths uchun max qiymatlar
    next_max_group = max((g['group_total'] for g in next_expected.get('breakdown', []) if g['group_total']), default=1)
    current_max_salary = max((d['salary'] for d in salary_data.get('details', []) if d['salary']), default=1)

    ctx = {
        'teacher': teacher,
        'selected_year': selected_year,
        'selected_month': selected_month,
        'salary_data': salary_data,
        'daily_income': daily_income,
        'monthly_income': monthly_income,
        'total_year_income': total_year_income,
        'daily_labels': daily_labels,
        'monthly_labels': monthly_labels,
        'years': years,
        'months_list': months_list,
        'months_dict': months_dict,
        'is_locked': salary_data.get('is_locked', False),
        'is_admin': is_admin,
        'is_support_member': is_support_member,
        'support_salary_total': support_salary_total,
        # Prognoz
        'current_expected': current_expected,
        'next_expected': next_expected,
        'next_year': next_year,
        'next_month': next_month,
        'next_month_name': months_dict.get(next_month, ""),
        'progress_pct': progress_pct,
        'delta_pct': delta_pct,
        'next_max_group': next_max_group,
        'current_max_salary': current_max_salary,
    }

    if is_admin:
        ctx['teachers_list'] = User.objects.filter(role='teacher', is_active=True)

    return render(request, "education/teacher_income_dashboard.html", ctx)



@login_required
@require_feature("finance")
def close_finance_month_view(request):
    """View to close (lock) or open (unlock) a financial month for a center."""
    if request.user.role not in ['director', 'manager'] and not request.user.is_superuser:
        messages.error(request, "Sizda bu bo'limga ruxsat yo'q.")
        return redirect('education:teacher_income_dashboard')

    if request.method == "POST":
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        action = request.POST.get('action', 'lock')
        
        from core.tenant import get_request_center
        center = get_request_center(request)
        
        if action == 'unlock':
            HistoricalFinanceService.open_month(center, year, month, request.user)
            messages.success(request, f"{year}-yil {month}-oy muvaffaqiyatli ochildi. Endi oyliklar avtomatik (jonli) tarzda hisoblanadi.")
        else:
            HistoricalFinanceService.close_month(center, year, month, request.user)
            messages.success(request, f"{year}-yil {month}-oy muvaffaqiyatli yopildi va qotirildi. Endi o'zgarishlar tasir qilmaydi.")
        
    return redirect(f"{reverse('education:teacher_salary_list')}?year={year}&month={month}")



@login_required
@require_feature("finance")
def fix_all_incomes(request):
    """
    Global/Production muhitda o'tgan oydagi eski Attendance malumotlarini 
    TeacherIncome tizimiga generator qilib beruvchi bir martalik master funktsiya
    """
    from education.models import Attendance, TeacherIncome, Enrollment
    from django.db import transaction
    
    all_attendances = Attendance.objects.all().select_related('group', 'student', 'teacher', 'group__center')
    
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for i, att in enumerate(all_attendances):
            # 1. To'lanadigan holatmi?
            is_billable = att.status == 'present' or att.status == 'absent_unexcused' or getattr(att, 'forced', False) or getattr(att, 'present', False)

            if not is_billable:
                TeacherIncome.objects.filter(attendance=att).delete()
                continue
            
            # 2. Enrollmentni topish
            enrollment = Enrollment.all_objects.filter(
                group=att.group,
                student=att.student
            ).order_by('-is_active', '-created_at').first()

            if not enrollment:
                TeacherIncome.objects.filter(attendance=att).delete()
                continue

            teacher = att.group.oqituvchi if att.group else None
            if not teacher:
                continue

            foiz = getattr(teacher, 'oqituvchi_foizi', 0)
            if foiz is None or foiz == 0:
                foiz = enrollment.oqituvchi_foiz

            kurs_narhi = full_course_amount(enrollment)
            
            oy_dars_soni = att.group.oy_dars_soni or 12
            if oy_dars_soni <= 0: oy_dars_soni = 12

            if kurs_narhi > 0 and foiz > 0:
                total_per_lesson = kurs_narhi / oy_dars_soni
                amount = round(total_per_lesson * (foiz / 100))
                center_amount = round(total_per_lesson * ((100 - foiz) / 100))
                total_amount = round(total_per_lesson)
            else:
                amount = 0
                center_amount = 0
                total_amount = 0

            obj, created = TeacherIncome.objects.update_or_create(
                attendance=att,
                defaults={
                    'center': att.center or (att.group.center if att.group else None),
                    'teacher': teacher,
                    'group': att.group,
                    'amount': amount,
                    'center_amount': center_amount,
                    'total_amount': total_amount
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1

    messages.success(request, f"🚀 Barcha daromadlar qayta hisoblandi! Yangi tizim qo'llandi. Yaratildi: {created_count}, Yangilandi: {updated_count}.")
    return redirect('education:teacher_income_dashboard')

