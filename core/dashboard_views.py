"""
ChaqmoqApp — 8 ta statistik dashboard uchun view va API endpointlar.
Har bir dashboard:
  - Page view (HTML render)
  - JSON API endpoint (Chart.js uchun ma'lumot)
  - 5 daqiqa cache
  - Director/Manager role check
"""

import calendar
from collections import defaultdict
from datetime import date, datetime, time, timedelta

from django.contrib.auth.decorators import login_required
from billing.decorators import require_feature
from django.core.cache import cache
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from accounts.models import Branch, Center, User
from core.alerts import get_director_alerts
from education.models import (
    Attendance,
    Category,
    CenterExpense,
    Enrollment,
    ExamResult,
    Group,
    MonthlyFinanceSnapshot,
    Payment,
    PaymentAllocation,
    StudentGroupHistory,
)
from store.models import Expense, Lead, LeadStatus, Product, PurchaseRequest, Sale


# ─────────────────────────── Helpers ────────────────────────────

def _get_center(request):
    """Return center for the request user (director / manager / superuser)."""
    user = request.user
    if not user.is_authenticated:
        return None
    if user.is_superuser:
        return getattr(request, "center", None) or getattr(user, "center", None)
    role = getattr(user, "role", None)
    if role in ("director", "manager"):
        return getattr(request, "center", None) or getattr(user, "center", None)
    return None


def _parse_dates(request):
    """Parse date_from / date_to query params; default = current month → today."""
    today = timezone.localdate()
    try:
        d_from = date.fromisoformat(request.GET.get("date_from", ""))
    except (ValueError, TypeError):
        d_from = today.replace(day=1)
    try:
        d_to = date.fromisoformat(request.GET.get("date_to", ""))
    except (ValueError, TypeError):
        d_to = today
    if d_from > d_to:
        d_from, d_to = d_to, d_from
    return d_from, d_to


UZ_MONTHS = {
    1: "Yan", 2: "Fev", 3: "Mar", 4: "Apr", 5: "May", 6: "Iyn",
    7: "Iyl", 8: "Avg", 9: "Sen", 10: "Okt", 11: "Noy", 12: "Dek",
}


def _six_month_range(anchor: date):
    """Yield (m_start, m_end, label) for 6 months ending with anchor's month."""
    for i in range(5, -1, -1):
        raw = anchor.replace(day=1) - timedelta(days=30 * i)
        m_start = raw.replace(day=1)
        m_end = m_start.replace(day=calendar.monthrange(m_start.year, m_start.month)[1])
        yield m_start, m_end, UZ_MONTHS[m_start.month]


def _403():
    return JsonResponse({"error": "Ruxsat yo'q"}, status=403)


def _period_stats(d_from: date, d_to: date):
    days = (d_to - d_from).days + 1
    prev_to = d_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=days - 1)
    return {
        "days": days,
        "prev_from": prev_from,
        "prev_to": prev_to,
    }


def _pct_change(current, previous):
    current = float(current or 0)
    previous = float(previous or 0)
    if previous == 0:
        return None if current else 0.0
    return round((current - previous) / previous * 100, 1)


def _attendance_present_filter():
    return Q(status="present") | Q(present=True) | Q(forced=True)


def _payments_for_center(center):
    """
    Productiondagi eski to'lovlarda center bo'sh qolgan bo'lishi mumkin.
    Shunda group/student/enrollment orqali markazni aniqlab olamiz.
    """
    return Payment.objects.filter(
        Q(center=center)
        | Q(center__isnull=True, group__center=center)
        | Q(center__isnull=True, student__center=center)
        | Q(center__isnull=True, enrollment__center=center)
    )


def _payment_allocations_for_center(center):
    """
    Productionda tarixiy daromad ko'pincha PaymentAllocation orqali saqlanadi.
    center null bo'lgan legacy allocation/paymentlar uchun ham fallback qilamiz.
    """
    return PaymentAllocation.objects.filter(
        payment__is_deleted=False,
    ).filter(
        Q(center=center)
        | Q(center__isnull=True, payment__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__group__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__student__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__enrollment__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, tuition_month__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, tuition_month__enrollment__center=center)
    )


def _deleted_payment_allocations_for_center(center):
    return PaymentAllocation.all_objects.filter(
        is_deleted=True,
    ).filter(
        Q(center=center)
        | Q(center__isnull=True, payment__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__group__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__student__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, payment__enrollment__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, tuition_month__center=center)
        | Q(center__isnull=True, payment__center__isnull=True, tuition_month__enrollment__center=center)
    )


def _deleted_payments_for_center(center):
    return Payment.all_objects.filter(
        is_deleted=True,
    ).filter(
        Q(center=center)
        | Q(center__isnull=True, group__center=center)
        | Q(center__isnull=True, student__center=center)
        | Q(center__isnull=True, enrollment__center=center)
    )


def _monthly_snapshot_for_center(center, m_start):
    return MonthlyFinanceSnapshot.objects.filter(
        financial_month__center=center,
        financial_month__year=m_start.year,
        financial_month__month=m_start.month,
    ).first()


# ──────────────────────────────────────────────────────────────────
#  Bulk monthly aggregators — bitta GROUP BY query bilan ko'p oyni
#  qaytaradi. Avvalgi per-oy chaqiruv loop'lari (12-24 ta) o'rniga
#  ishlatiladi. Empty oylar uchun per-month fallback saqlanadi.
# ──────────────────────────────────────────────────────────────────

def _bulk_monthly_turnover(center, months):
    """
    months: list of (m_start, m_end) tuples.
    Returns dict {(year, month): int_total} for ALL months in span,
    using a fixed number of queries (5 — independent of months count).

    Priority: active_total > snapshot > deleted_total. Identical to the
    single-month helper but vectorised.
    """
    if not months:
        return {}
    span_start = min(m[0] for m in months)
    span_end = max(m[1] for m in months)
    keys = [(m[0].year, m[0].month) for m in months]
    result = {key: 0 for key in keys}

    def _bucket(rows, value_field):
        out = {key: 0 for key in keys}
        for row in rows:
            m = row["_m"]
            if m is None:
                continue
            key = (m.year, m.month)
            if key in out:
                out[key] += int(row[value_field] or 0)
        return out

    # 1) Active allocations
    active_alloc = _bucket(
        _payment_allocations_for_center(center)
        .filter(tuition_month__month__range=(span_start, span_end))
        .annotate(_m=TruncMonth("tuition_month__month"))
        .values("_m").annotate(total=Sum("amount")),
        "total",
    )

    # 2) Active unallocated payments
    active_unalloc = _bucket(
        _payments_for_center(center)
        .filter(paid_date__range=(span_start, span_end), allocations__isnull=True)
        .annotate(_m=TruncMonth("paid_date"))
        .values("_m").annotate(total=Sum("summa")),
        "total",
    )

    # 3) Snapshots — single query, indexed lookup
    snapshot_map = {}
    snap_qs = MonthlyFinanceSnapshot.objects.filter(
        financial_month__center=center,
        financial_month__year__in={k[0] for k in keys},
        financial_month__month__in={k[1] for k in keys},
    ).values("financial_month__year", "financial_month__month", "total_income")
    for row in snap_qs:
        snapshot_map[(row["financial_month__year"], row["financial_month__month"])] = int(row["total_income"] or 0)

    # 4) Deleted allocations
    deleted_alloc = _bucket(
        _deleted_payment_allocations_for_center(center)
        .filter(tuition_month__month__range=(span_start, span_end))
        .annotate(_m=TruncMonth("tuition_month__month"))
        .values("_m").annotate(total=Sum("amount")),
        "total",
    )

    # 5) Deleted unallocated payments
    deleted_unalloc = _bucket(
        _deleted_payments_for_center(center)
        .filter(paid_date__range=(span_start, span_end), allocations__isnull=True)
        .annotate(_m=TruncMonth("paid_date"))
        .values("_m").annotate(total=Sum("summa")),
        "total",
    )

    for key in keys:
        active_total = active_alloc[key] + active_unalloc[key]
        if active_total:
            result[key] = active_total
            continue
        snap = snapshot_map.get(key, 0)
        if snap:
            result[key] = snap
            continue
        result[key] = deleted_alloc[key] + deleted_unalloc[key]

    return result


def _bulk_monthly_expenses(center, months):
    """
    months: list of (m_start, m_end) tuples.
    Returns dict {(year, month): int_total} via 3 queries total.
    Priority: active_total > snapshot.total_expense.
    """
    if not months:
        return {}
    span_start = min(m[0] for m in months)
    span_end = max(m[1] for m in months)
    keys = [(m[0].year, m[0].month) for m in months]
    result = {key: 0 for key in keys}

    def _bucket(rows, value_field):
        out = {key: 0 for key in keys}
        for row in rows:
            m = row["_m"]
            if m is None:
                continue
            key = (m.year, m.month)
            if key in out:
                out[key] += int(row[value_field] or 0)
        return out

    legacy = _bucket(
        _expenses_for_center(center)
        .filter(sana__date__range=(span_start, span_end))
        .annotate(_m=TruncMonth("sana"))
        .values("_m").annotate(total=Sum("summa")),
        "total",
    )
    center_exp = _bucket(
        _center_expenses_for_center(center)
        .filter(date__range=(span_start, span_end))
        .annotate(_m=TruncMonth("date"))
        .values("_m").annotate(total=Sum("amount")),
        "total",
    )

    snapshot_map = {}
    snap_qs = MonthlyFinanceSnapshot.objects.filter(
        financial_month__center=center,
        financial_month__year__in={k[0] for k in keys},
        financial_month__month__in={k[1] for k in keys},
    ).values("financial_month__year", "financial_month__month", "total_expense")
    for row in snap_qs:
        snapshot_map[(row["financial_month__year"], row["financial_month__month"])] = int(row["total_expense"] or 0)

    for key in keys:
        active_total = legacy[key] + center_exp[key]
        result[key] = active_total if active_total else snapshot_map.get(key, 0)

    return result


def _monthly_turnover_for_center(center, m_start, m_end):
    """
    Oy kesimidagi aylanma:
    1. Allocation mavjud bo'lsa, to'lovni aynan tegishli oyga yozamiz.
    2. Allocation yo'q legacy paymentlar bo'lsa, paid_date bo'yicha fallback qilamiz.
    """
    allocated_total = int(
        _payment_allocations_for_center(center)
        .filter(tuition_month__month__range=(m_start, m_end))
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    unallocated_total = int(
        _payments_for_center(center)
        .filter(paid_date__range=(m_start, m_end), allocations__isnull=True)
        .aggregate(s=Sum("summa"))["s"] or 0
    )
    active_total = allocated_total + unallocated_total
    if active_total:
        return active_total

    snapshot = _monthly_snapshot_for_center(center, m_start)
    if snapshot and snapshot.total_income:
        return int(snapshot.total_income or 0)

    deleted_allocated_total = int(
        _deleted_payment_allocations_for_center(center)
        .filter(tuition_month__month__range=(m_start, m_end))
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    deleted_unallocated_total = int(
        _deleted_payments_for_center(center)
        .filter(paid_date__range=(m_start, m_end), allocations__isnull=True)
        .aggregate(s=Sum("summa"))["s"] or 0
    )
    return deleted_allocated_total + deleted_unallocated_total


def _monthly_expenses_for_center(center, m_start, m_end):
    legacy_total = int(
        _expenses_for_center(center)
        .filter(sana__date__range=(m_start, m_end))
        .aggregate(s=Sum("summa"))["s"] or 0
    )
    center_total = int(
        _center_expenses_for_center(center)
        .filter(date__range=(m_start, m_end))
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    active_total = legacy_total + center_total
    if active_total:
        return active_total

    snapshot = _monthly_snapshot_for_center(center, m_start)
    if snapshot and snapshot.total_expense:
        return int(snapshot.total_expense or 0)
    return 0


def _expenses_for_center(center):
    """
    Legacy xarajatlardagi center=null holati uchun worker/product orqali fallback.
    """
    return Expense.objects.filter(
        Q(center=center)
        | Q(center__isnull=True, worker__center=center)
        | Q(center__isnull=True, product__center=center)
    )


def _center_expenses_for_center(center):
    return CenterExpense.objects.filter(center=center)


def _teacher_compensation(center, d_from, d_to):
    att_map = {
        (row["group_id"], row["student_id"]): row["cnt"]
        for row in Attendance.objects.filter(
            group__center=center,
            date__range=(d_from, d_to),
        ).filter(_attendance_present_filter())
        .values("group_id", "student_id")
        .annotate(cnt=Count("id"))
    }

    total = 0
    enrollments = Enrollment.objects.filter(
        group__center=center,
        is_active=True,
        student__is_archived=False,
    ).select_related("group")
    for enr in enrollments:
        present_count = att_map.get((enr.group_id, enr.student_id), 0)
        if not present_count:
            continue
        lessons = enr.group.oy_dars_soni or 12
        fee = enr.kurs_narhi or enr.group.kurs_narxi or 0
        teacher_pct = enr.oqituvchi_foiz or enr.group.oqituvchi_foiz or 0
        if lessons <= 0 or fee <= 0 or teacher_pct <= 0:
            continue
        per_lesson_share = (fee * teacher_pct / 100) / lessons
        total += per_lesson_share * present_count
    return int(round(total))


def _payment_type_breakdown(pay_qs):
    type_map = {"cash": "Naqd", "card": "Karta", "mixed": "Aralash"}
    return [
        {
            "label": type_map.get(row["payment_type"], row["payment_type"]),
            "value": int(row["total"] or 0),
            "count": row["cnt"],
        }
        for row in pay_qs.values("payment_type").annotate(total=Sum("summa"), cnt=Count("id"))
    ]


def _financial_payload(center, d_from, d_to):
    cache_key = f"financial_payload_{center.id}_{d_from}_{d_to}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    period = _period_stats(d_from, d_to)
    pay_qs = _payments_for_center(center).filter(paid_date__range=(d_from, d_to))
    revenue = int(pay_qs.aggregate(s=Sum("summa"))["s"] or 0)
    pay_count = pay_qs.count()
    avg_pay = int(revenue / pay_count) if pay_count else 0

    legacy_exp_qs = _expenses_for_center(center).filter(sana__date__range=(d_from, d_to))
    center_exp_qs = _center_expenses_for_center(center).filter(date__range=(d_from, d_to))
    legacy_expenses = int(legacy_exp_qs.aggregate(s=Sum("summa"))["s"] or 0)
    center_expenses = int(center_exp_qs.aggregate(s=Sum("amount"))["s"] or 0)
    expenses = legacy_expenses + center_expenses
    teacher_comp = _teacher_compensation(center, d_from, d_to)
    total_cost = expenses + teacher_comp
    net_profit = revenue - total_cost
    profit_margin = round(net_profit / revenue * 100, 1) if revenue else 0

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_pay_qs = _payments_for_center(center).filter(paid_date__range=(prev_from, prev_to))
    prev_revenue = int(prev_pay_qs.aggregate(s=Sum("summa"))["s"] or 0)
    prev_expenses = _monthly_expenses_for_center(center, prev_from, prev_to)
    prev_teacher_comp = _teacher_compensation(center, prev_from, prev_to)
    prev_total_cost = prev_expenses + prev_teacher_comp
    prev_net_profit = prev_revenue - prev_total_cost

    six_months_meta = list(_six_month_range(d_to))
    six_months_pairs = [(ms, me) for ms, me, _ in six_months_meta]
    bulk_six_turnover = _bulk_monthly_turnover(center, six_months_pairs)
    bulk_six_expenses = _bulk_monthly_expenses(center, six_months_pairs)

    m_labels, m_inc, m_exp, m_teacher, m_cost = [], [], [], [], []
    for ms, me, lbl in six_months_meta:
        m_labels.append(lbl)
        key = (ms.year, ms.month)
        month_income = int(bulk_six_turnover.get(key, 0) or 0)
        month_expense = int(bulk_six_expenses.get(key, 0) or 0)
        month_teacher = _teacher_compensation(center, ms, me)
        m_inc.append(month_income)
        m_exp.append(month_expense)
        m_teacher.append(month_teacher)
        m_cost.append(month_expense + month_teacher)

    cat_breakdown = [
        {
            "label": row.get("group__category_obj__name") or "Bo'limsiz",
            "value": int(row["total"] or 0),
        }
        for row in pay_qs.values("group__category_obj__name").annotate(total=Sum("summa")).order_by("-total")[:6]
    ]

    expense_category_map = dict(CenterExpense.CATEGORY_CHOICES)
    expense_breakdown = [
        {
            "category": row["category"],
            "label": expense_category_map.get(row["category"], row["category"]),
            "value": int(row["total"] or 0),
        }
        for row in center_exp_qs.values("category").annotate(total=Sum("amount")).order_by("-total")
    ]
    if legacy_expenses:
        expense_breakdown.append(
            {
                "category": "legacy",
                "label": "Legacy / boshqa",
                "value": legacy_expenses,
            }
        )

    today = timezone.localdate()
    year_months = []
    for month in range(1, 13):
        ts = date(today.year, month, 1)
        te = ts.replace(day=calendar.monthrange(ts.year, ts.month)[1])
        ls = date(today.year - 1, month, 1)
        le = ls.replace(day=calendar.monthrange(ls.year, ls.month)[1])
        year_months.append((ts, te))
        year_months.append((ls, le))
    bulk_year = _bulk_monthly_turnover(center, year_months)
    this_year = [
        int(bulk_year.get((today.year, m), 0) or 0)
        for m in range(1, 13)
    ]
    last_year = [
        int(bulk_year.get((today.year - 1, m), 0) or 0)
        for m in range(1, 13)
    ]
    this_year_ytd = sum(this_year[: today.month])
    last_year_ytd = sum(last_year[: today.month])
    growth_pct = round((this_year_ytd - last_year_ytd) / max(last_year_ytd, 1) * 100, 1)

    result = {
        "kpis": {
            "revenue": revenue,
            "net_profit": net_profit,
            "expenses": expenses,
            "teacher_comp": teacher_comp,
            "total_cost": total_cost,
            "profit_margin": profit_margin,
            "avg_pay": avg_pay,
            "pay_count": pay_count,
            "changes": {
                "revenue": _pct_change(revenue, prev_revenue),
                "expenses": _pct_change(expenses, prev_expenses),
                "teacher_comp": _pct_change(teacher_comp, prev_teacher_comp),
                "total_cost": _pct_change(total_cost, prev_total_cost),
                "net_profit": _pct_change(net_profit, prev_net_profit),
                "pay_count": _pct_change(pay_count, prev_pay_qs.count()),
            },
        },
        "charts": {
            "m_labels": m_labels,
            "m_inc": m_inc,
            "m_exp": m_exp,
            "m_teacher": m_teacher,
            "m_cost": m_cost,
            "pay_types": _payment_type_breakdown(pay_qs),
            "cat_breakdown": cat_breakdown,
        },
        "breakdowns": {
            "expenses": expense_breakdown,
        },
        "year_compare": {
            "labels": [UZ_MONTHS[month] for month in range(1, 13)],
            "this_year": this_year,
            "last_year": last_year,
            "growth_pct": growth_pct,
        },
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
        "revenue": revenue,
        "expense": expenses,
        "teacher_comp": teacher_comp,
        "profit": net_profit,
        "profit_margin": profit_margin,
    }
    cache.set(cache_key, result, timeout=900)
    return result


def _student_payload(center, d_from, d_to):
    period = _period_stats(d_from, d_to)
    students_qs = User.objects.filter(center=center, role="student", is_archived=False)
    total_students = students_qs.count()
    active_enroll = Enrollment.objects.filter(
        group__center=center,
        is_active=True,
        student__is_archived=False,
    )
    active_count = active_enroll.values("student").distinct().count()
    new_students = students_qs.filter(date_joined__date__range=(d_from, d_to)).count()

    att_qs = Attendance.objects.filter(group__center=center, date__range=(d_from, d_to))
    present_filter = _attendance_present_filter()
    total_att = att_qs.count()
    present_att = att_qs.filter(present_filter).count()
    avg_att_rate = round(present_att / total_att * 100, 1) if total_att else 0

    per_student = list(
        att_qs.values("student_id")
        .annotate(tot=Count("id"), pres=Count("id", filter=present_filter))
    )
    low_activity = sum(1 for row in per_student if (row["pres"] / row["tot"] * 100 if row["tot"] else 0) < 70)
    churn_risk = sum(1 for row in per_student if (row["pres"] / row["tot"] * 100 if row["tot"] else 0) < 30)

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_students = students_qs.filter(date_joined__date__range=(prev_from, prev_to)).count()
    prev_att_qs = Attendance.objects.filter(group__center=center, date__range=(prev_from, prev_to))
    prev_att_total = prev_att_qs.count()
    prev_att_present = prev_att_qs.filter(present_filter).count()
    prev_att_rate = round(prev_att_present / prev_att_total * 100, 1) if prev_att_total else 0

    group_dist = list(
        active_enroll.values("group__nom")
        .annotate(cnt=Count("student", distinct=True))
        .order_by("-cnt")[:8]
    )

    m_labels, m_new, m_att = [], [], []
    for ms, me, lbl in _six_month_range(d_to):
        m_labels.append(lbl)
        m_new.append(students_qs.filter(date_joined__date__range=(ms, me)).count())
        monthly_att = Attendance.objects.filter(group__center=center, date__range=(ms, me))
        tot = monthly_att.count()
        pres = monthly_att.filter(present_filter).count()
        m_att.append(round(pres / tot * 100, 1) if tot else 0)

    top_students = []
    raw_students = list(
        att_qs.values("student__ism", "student__familya")
        .annotate(tot=Count("id"), pres=Count("id", filter=present_filter))
    )
    for row in raw_students:
        rate = round(row["pres"] / row["tot"] * 100, 1) if row["tot"] else 0
        top_students.append({
            "name": f"{row['student__ism']} {row['student__familya']}",
            "rate": rate,
            "sessions": row["pres"],
        })
    top_students.sort(key=lambda row: (-row["rate"], -row["sessions"], row["name"]))

    return {
        "kpis": {
            "total_students": total_students,
            "active_count": active_count,
            "new_students": new_students,
            "avg_att_rate": avg_att_rate,
            "low_activity": low_activity,
            "churn_risk": churn_risk,
            "changes": {
                "new_students": _pct_change(new_students, prev_students),
                "avg_att_rate": round(avg_att_rate - prev_att_rate, 1),
            },
        },
        "charts": {
            "m_labels": m_labels,
            "m_new": m_new,
            "m_att": m_att,
            "group_labels": [row["group__nom"] for row in group_dist],
            "group_counts": [row["cnt"] for row in group_dist],
        },
        "top_students": top_students[:8],
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
    }


def _teacher_payload(center, d_from, d_to):
    cache_key = f"teacher_payload_{center.id}_{d_from}_{d_to}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    period = _period_stats(d_from, d_to)
    teachers = list(User.objects.filter(center=center, role="teacher"))
    total_teachers = len(teachers)
    teacher_ids = [teacher.id for teacher in teachers]
    active_enrollments = Enrollment.objects.filter(group__center=center, is_active=True, student__is_archived=False)
    distinct_students = active_enrollments.values("student").distinct().count()

    rev_map = dict(
        _payments_for_center(center)
        .filter(
            group__oqituvchi_id__in=teacher_ids,
            paid_date__range=(d_from, d_to),
        )
        .values("group__oqituvchi_id")
        .annotate(s=Sum("summa"))
        .values_list("group__oqituvchi_id", "s")
    )
    stu_map = dict(
        active_enrollments
        .filter(group__oqituvchi_id__in=teacher_ids)
        .values("group__oqituvchi_id")
        .annotate(cnt=Count("student", distinct=True))
        .values_list("group__oqituvchi_id", "cnt")
    )
    grp_map = dict(
        Group.objects.filter(
            center=center,
            is_archived=False,
            is_deleted=False,
            oqituvchi_id__in=teacher_ids,
        )
        .values("oqituvchi_id")
        .annotate(cnt=Count("id"))
        .values_list("oqituvchi_id", "cnt")
    )
    att_base = Attendance.objects.filter(
        group__center=center,
        group__oqituvchi_id__in=teacher_ids,
        date__range=(d_from, d_to),
    )
    att_total_map = dict(
        att_base.values("group__oqituvchi_id")
        .annotate(cnt=Count("id"))
        .values_list("group__oqituvchi_id", "cnt")
    )
    att_present_map = dict(
        att_base.filter(_attendance_present_filter())
        .values("group__oqituvchi_id")
        .annotate(cnt=Count("id"))
        .values_list("group__oqituvchi_id", "cnt")
    )
    dropout_map = dict(
        StudentGroupHistory.objects.filter(
            center=center,
            group__oqituvchi_id__in=teacher_ids,
            end_date__range=(d_from, d_to),
        )
        .values("group__oqituvchi_id")
        .annotate(cnt=Count("student", distinct=True))
        .values_list("group__oqituvchi_id", "cnt")
    )
    exam_base = ExamResult.objects.filter(
        center=center,
        teacher_id__in=teacher_ids,
        exam_date__range=(d_from, d_to),
    )
    pct_map = dict(
        exam_base.exclude(percent__isnull=True)
        .values("teacher_id")
        .annotate(a=Avg("percent"))
        .values_list("teacher_id", "a")
    )
    score_map = dict(
        exam_base.exclude(score__isnull=True)
        .values("teacher_id")
        .annotate(a=Avg("score"))
        .values_list("teacher_id", "a")
    )

    stats = []
    for teacher in teachers:
        teacher_id = teacher.id
        revenue = int(rev_map.get(teacher_id) or 0)
        students = int(stu_map.get(teacher_id) or 0)
        groups = int(grp_map.get(teacher_id) or 0)
        total_att = int(att_total_map.get(teacher_id) or 0)
        present_att = int(att_present_map.get(teacher_id) or 0)
        avg_attendance_pct = round(present_att / total_att * 100, 1) if total_att else 0
        hours = present_att
        dropout_count = int(dropout_map.get(teacher_id) or 0)
        percent_avg = pct_map.get(teacher_id)
        score_avg = score_map.get(teacher_id)
        exam_avg_score = round(float(percent_avg if percent_avg is not None else score_avg or 0), 1)
        dropout_rate = dropout_count / max(students + dropout_count, 1)
        score = round(
            (avg_attendance_pct * 0.4)
            + ((1 - dropout_rate) * 100 * 0.4)
            + (exam_avg_score * 0.2),
            1,
        )
        stats.append({
            "name": teacher.get_full_name(),
            "revenue": revenue,
            "students": students,
            "students_count": students,
            "groups": groups,
            "groups_count": groups,
            "hours": hours,
            "avg_attendance_pct": avg_attendance_pct,
            "dropout_count": dropout_count,
            "exam_avg_score": exam_avg_score,
            "dropout_rate": round(dropout_rate * 100, 1),
            "score": max(0, min(score, 100)),
        })
    stats.sort(key=lambda row: (-row["score"], -row["avg_attendance_pct"], -row["students"], row["name"]))

    total_rev = sum(row["revenue"] for row in stats)
    avg_rev = int(total_rev / total_teachers) if total_teachers else 0
    ratio = round(distinct_students / total_teachers, 1) if total_teachers else 0

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_total_rev = int(
        _payments_for_center(center).filter(
            group__oqituvchi__role="teacher",
            paid_date__range=(prev_from, prev_to),
        ).aggregate(s=Sum("summa"))["s"] or 0
    )

    top10 = stats[:10]
    result = {
        "kpis": {
            "total_teachers": total_teachers,
            "avg_rev": avg_rev,
            "total_rev": total_rev,
            "ratio": ratio,
            "total_hours": sum(row["hours"] for row in stats),
            "changes": {
                "total_rev": _pct_change(total_rev, prev_total_rev),
            },
        },
        "charts": {
            "names": [row["name"] for row in top10],
            "revenues": [row["revenue"] for row in top10],
            "students": [row["students"] for row in top10],
            "hours": [row["hours"] for row in top10],
            "scores": [row["score"] for row in top10],
            "attendance": [row["avg_attendance_pct"] for row in top10],
        },
        "teacher_list": stats[:15],
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
        "total_teachers": total_teachers,
        "avg_revenue_per_teacher": avg_rev,
        "total_revenue": total_rev,
        "student_teacher_ratio": ratio,
        "teachers": stats[:15],
    }
    cache.set(cache_key, result, timeout=900)
    return result


def _groups_payload(center, d_from, d_to):
    cache_key = f"groups_payload_{center.id}_{d_from}_{d_to}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    period = _period_stats(d_from, d_to)
    groups_qs = Group.objects.filter(center=center, is_archived=False, is_deleted=False)
    total_groups = groups_qs.count()
    active_groups = groups_qs.filter(is_closed=False).count()
    active_enroll_total = Enrollment.objects.filter(group__center=center, is_active=True).count()
    avg_size = round(active_enroll_total / active_groups, 1) if active_groups else 0

    top_groups = list(groups_qs.select_related("oqituvchi", "category_obj")[:40])
    top_group_ids = [group.id for group in top_groups]

    enroll_map = dict(
        Enrollment.objects.filter(
            group_id__in=top_group_ids,
            is_active=True,
            is_deleted=False,
        )
        .values("group_id")
        .annotate(cnt=Count("id"))
        .values_list("group_id", "cnt")
    )
    rev_map = dict(
        Payment.objects.filter(
            group_id__in=top_group_ids,
            paid_date__range=(d_from, d_to),
        )
        .values("group_id")
        .annotate(s=Sum("summa"))
        .values_list("group_id", "s")
    )
    att_base = Attendance.objects.filter(
        group_id__in=top_group_ids,
        date__range=(d_from, d_to),
    )
    att_total_map = dict(
        att_base.values("group_id")
        .annotate(cnt=Count("id"))
        .values_list("group_id", "cnt")
    )
    att_present_map = dict(
        att_base.filter(_attendance_present_filter())
        .values("group_id")
        .annotate(cnt=Count("id"))
        .values_list("group_id", "cnt")
    )

    group_stats = []
    total_capacity = 0
    total_active_students = 0
    for group in top_groups:
        group_id = group.id
        enrolled = int(enroll_map.get(group_id) or 0)
        capacity = int(getattr(group, "max_students", 0) or 0)
        fill_pct = round(enrolled * 100 / capacity, 1) if capacity else 0
        revenue = int(rev_map.get(group_id) or 0)
        att_tot = int(att_total_map.get(group_id) or 0)
        att_pre = int(att_present_map.get(group_id) or 0)
        total_capacity += capacity
        total_active_students += enrolled
        group_stats.append({
            "name": group.nom,
            "teacher": f"{group.oqituvchi.ism} {group.oqituvchi.familya}" if group.oqituvchi else "—",
            "category": group.category_obj.name if group.category_obj else group.category,
            "student_count": enrolled,
            "enrolled": enrolled,
            "capacity": capacity,
            "fill_pct": fill_pct,
            "revenue": revenue,
            "att_rate": round(att_pre / att_tot * 100, 1) if att_tot else 0,
        })
    group_stats.sort(key=lambda row: (-row["fill_pct"], -row["revenue"], row["name"]))

    avg_fill_rate = round(total_active_students * 100 / total_capacity, 1) if total_capacity else 0

    cat_dist = list(
        Enrollment.objects.filter(group__center=center, is_active=True)
        .values("group__category_obj__name")
        .annotate(cnt=Count("student", distinct=True))
        .order_by("-cnt")[:6]
    )

    m_labels, m_enroll, m_rev = [], [], []
    for ms, me, lbl in _six_month_range(d_to):
        m_labels.append(lbl)
        m_enroll.append(Enrollment.objects.filter(group__center=center, created_at__date__range=(ms, me)).count())
        m_rev.append(int(Payment.objects.filter(center=center, paid_date__range=(ms, me)).aggregate(s=Sum("summa"))["s"] or 0))

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_rev = int(Payment.objects.filter(center=center, paid_date__range=(prev_from, prev_to)).aggregate(s=Sum("summa"))["s"] or 0)

    top8 = group_stats[:8]
    result = {
        "kpis": {
            "total_groups": total_groups,
            "active_groups": active_groups,
            "active_enroll_total": active_enroll_total,
            "avg_size": avg_size,
            "total_capacity": total_capacity,
            "avg_fill_rate": avg_fill_rate,
            "total_rev": sum(group["revenue"] for group in group_stats),
            "changes": {
                "total_rev": _pct_change(sum(group["revenue"] for group in group_stats), prev_rev),
            },
        },
        "charts": {
            "m_labels": m_labels,
            "m_enroll": m_enroll,
            "m_rev": m_rev,
            "cat_labels": [row["group__category_obj__name"] or "Boshqa" for row in cat_dist],
            "cat_counts": [row["cnt"] for row in cat_dist],
            "grp_names": [group["name"] for group in top8],
            "grp_revs": [group["revenue"] for group in top8],
            "grp_enroll": [group["enrolled"] for group in top8],
        },
        "group_list": group_stats[:12],
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
        "total_groups": total_groups,
        "active_groups": active_groups,
        "total_enrolled": active_enroll_total,
        "avg_fill_rate": avg_fill_rate,
        "total_revenue": sum(group["revenue"] for group in group_stats),
        "groups": group_stats[:12],
        "enrollment_trend": {
            "labels": m_labels,
            "enrolled": m_enroll,
        },
        "category_distribution": [
            {"name": row["group__category_obj__name"] or "Boshqa", "count": row["cnt"]}
            for row in cat_dist
        ],
    }
    cache.set(cache_key, result, timeout=900)
    return result


def _billing_payload(center, d_from, d_to):
    period = _period_stats(d_from, d_to)
    pay_qs = Payment.objects.filter(center=center, paid_date__range=(d_from, d_to))
    total_pay = int(pay_qs.aggregate(s=Sum("summa"))["s"] or 0)
    pay_count = pay_qs.count()
    avg_pay = int(total_pay / pay_count) if pay_count else 0

    daily = list(pay_qs.values("paid_date").annotate(tot=Sum("summa")).order_by("paid_date"))
    daily_labels = [str(row["paid_date"]) for row in daily]
    daily_amounts = [int(row["tot"] or 0) for row in daily]

    sub_info = {}
    try:
        sub = center.subscriptions.filter(status="ACTIVE").first()
        if sub:
            sub_info = {
                "plan": sub.plan.title,
                "status": sub.status,
                "days_left": sub.days_left(),
                "expires_at": str(sub.expires_at.date()) if sub.expires_at else None,
                "monthly_price": sub.plan.monthly_price,
            }
    except Exception:
        pass

    def debt_snapshot(target_date):
        try:
            from django.db.models import OuterRef, Subquery
            from django.db.models.functions import Coalesce
            from education.models import PaymentAllocation, TuitionMonth
            from education.services.tuition import tuition_month_fee_field

            debt_month = target_date.replace(day=1)
            fee_field = tuition_month_fee_field()
            fee_sub = (
                TuitionMonth.objects
                .filter(enrollment=OuterRef("pk"), month=debt_month, is_deleted=False)
                .values("enrollment")
                .annotate(s=Sum(fee_field))
                .values("s")
            )
            paid_sub = (
                PaymentAllocation.objects
                .filter(
                    tuition_month__enrollment=OuterRef("pk"),
                    tuition_month__month=debt_month,
                    tuition_month__is_deleted=False,
                    payment__is_deleted=False,
                )
                .values("tuition_month__enrollment")
                .annotate(s=Sum("amount"))
                .values("s")
            )
            debt_qs = (
                Enrollment.objects.filter(
                    group__center=center,
                    is_active=True,
                    student__is_archived=False,
                    group__is_archived=False,
                    group__is_deleted=False,
                    is_deferred=False,
                )
                .annotate(f=Coalesce(Subquery(fee_sub), 0), p=Coalesce(Subquery(paid_sub), 0))
                .annotate(d=F("f") - F("p")).filter(d__gt=0)
            )
            return int(debt_qs.aggregate(s=Sum("d"))["s"] or 0), debt_qs.values("student").distinct().count()
        except Exception:
            return 0, 0

    total_debt, total_debtors = debt_snapshot(d_to)

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_total_pay = int(Payment.objects.filter(center=center, paid_date__range=(prev_from, prev_to)).aggregate(s=Sum("summa"))["s"] or 0)
    prev_debt, _ = debt_snapshot(prev_to)

    return {
        "kpis": {
            "total_pay": total_pay,
            "pay_count": pay_count,
            "avg_pay": avg_pay,
            "total_debt": total_debt,
            "total_debtors": total_debtors,
            "daily_avg_count": round(pay_count / period["days"], 1) if period["days"] else 0,
            "changes": {
                "total_pay": _pct_change(total_pay, prev_total_pay),
                "total_debt": _pct_change(total_debt, prev_debt),
            },
        },
        "subscription": sub_info,
        "charts": {
            "daily_labels": daily_labels,
            "daily_amounts": daily_amounts,
            "pay_types": _payment_type_breakdown(pay_qs),
        },
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
    }


def _marketing_payload(center, d_from, d_to):
    period = _period_stats(d_from, d_to)
    converted_filter = Q(converted_to_student=True) | Q(converted_user__isnull=False)
    leads_qs = Lead.objects.filter(center=center, qoshilgan_sana__date__range=(d_from, d_to))
    total_leads = leads_qs.count()
    converted = leads_qs.filter(converted_filter).count()
    conv_rate = round(converted / total_leads * 100, 1) if total_leads else 0

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_leads_qs = Lead.objects.filter(center=center, qoshilgan_sana__date__range=(prev_from, prev_to))
    prev_leads = prev_leads_qs.count()

    by_source = list(
        leads_qs.values("manba__nom")
        .annotate(total=Count("id"), conv=Count("id", filter=converted_filter))
        .order_by("-total")[:8]
    )

    m_labels, m_leads, m_conv = [], [], []
    for ms, me, lbl in _six_month_range(d_to):
        m_labels.append(lbl)
        qs = Lead.objects.filter(center=center, qoshilgan_sana__date__range=(ms, me))
        m_leads.append(qs.count())
        m_conv.append(qs.filter(converted_filter).count())

    return {
        "kpis": {
            "total_leads": total_leads,
            "converted": converted,
            "conv_rate": conv_rate,
            "leads_growth": _pct_change(total_leads, prev_leads),
        },
        "charts": {
            "m_labels": m_labels,
            "m_leads": m_leads,
            "m_conv": m_conv,
            "src_labels": [row["manba__nom"] or "Noma'lum" for row in by_source],
            "src_counts": [row["total"] for row in by_source],
            "src_conv": [row["conv"] for row in by_source],
        },
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
    }


def _inventory_payload(center, d_from, d_to):
    period = _period_stats(d_from, d_to)
    products_qs = Product.objects.filter(center=center)
    total_products = products_qs.count()

    approved_requests = PurchaseRequest.objects.filter(
        center=center,
        status=PurchaseRequest.APPROVED,
        sana__date__range=(d_from, d_to),
    ).select_related("product")

    total_sale_count = approved_requests.count()
    total_qty = 0
    total_sale_rev = 0
    product_map = defaultdict(lambda: {"name": "", "price_som": 0, "qty_sold": 0, "revenue": 0})
    for req in approved_requests:
        if not req.product_id:
            continue
        revenue = (req.product.narx_som or 0) * (req.qty or 0)
        total_qty += req.qty or 0
        total_sale_rev += revenue
        bucket = product_map[req.product_id]
        bucket["name"] = req.product.nom
        bucket["price_som"] = req.product.narx_som or 0
        bucket["qty_sold"] += req.qty or 0
        bucket["revenue"] += revenue

    if not total_sale_count:
        sale_qs = Sale.objects.filter(
            Q(center=center) | Q(product__center=center),
            sana__date__range=(d_from, d_to),
        ).select_related("product")
        total_sale_count = sale_qs.count()
        for sale in sale_qs:
            if not sale.product_id:
                continue
            price_som = sale.narx_som or sale.product.narx_som or 0
            revenue = price_som * (sale.qty or 0)
            total_qty += sale.qty or 0
            total_sale_rev += revenue
            bucket = product_map[sale.product_id]
            bucket["name"] = sale.product.nom
            bucket["price_som"] = price_som
            bucket["qty_sold"] += sale.qty or 0
            bucket["revenue"] += revenue

    product_list = list(product_map.values())
    product_list.sort(key=lambda row: (-row["revenue"], -row["qty_sold"], row["name"]))
    top8 = product_list[:8]

    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    prev_requests = PurchaseRequest.objects.filter(
        center=center,
        status=PurchaseRequest.APPROVED,
        sana__date__range=(prev_from, prev_to),
    ).select_related("product")
    prev_rev = sum((req.product.narx_som or 0) * (req.qty or 0) for req in prev_requests if req.product_id)

    return {
        "kpis": {
            "total_products": total_products,
            "total_sale_count": total_sale_count,
            "total_sale_rev": int(total_sale_rev),
            "total_qty": int(total_qty),
            "changes": {
                "total_sale_rev": _pct_change(total_sale_rev, prev_rev),
            },
        },
        "charts": {
            "p_names": [row["name"] for row in top8],
            "p_revs": [row["revenue"] for row in top8],
            "p_qty": [row["qty_sold"] for row in top8],
        },
        "product_list": product_list[:12],
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
    }


def _analytics_payload(center, d_from, d_to):
    period = _period_stats(d_from, d_to)
    today = timezone.localdate()
    all_users = User.objects.filter(center=center)
    total_users = all_users.count()
    active_today = all_users.filter(last_login__date=today).count()
    active_week = all_users.filter(last_login__date__gte=today - timedelta(days=6)).count()
    active_month = all_users.filter(last_login__date__gte=today - timedelta(days=29)).count()

    m_labels, m_new, m_act = [], [], []
    for ms, me, lbl in _six_month_range(d_to):
        m_labels.append(lbl)
        m_new.append(all_users.filter(date_joined__date__range=(ms, me)).count())
        m_act.append(all_users.filter(last_login__date__range=(ms, me)).count())

    role_map = {
        "director": "Direktor",
        "manager": "Manager",
        "teacher": "O'qituvchi",
        "student": "O'quvchi",
        "parent": "Ota-ona",
    }
    role_dist = list(all_users.values("role").annotate(cnt=Count("id")).order_by("-cnt"))
    prev_from = period["prev_from"]
    prev_to = period["prev_to"]
    new_users = all_users.filter(date_joined__date__range=(d_from, d_to)).count()
    prev_new_users = all_users.filter(date_joined__date__range=(prev_from, prev_to)).count()

    return {
        "kpis": {
            "total_users": total_users,
            "active_today": active_today,
            "active_week": active_week,
            "active_month": active_month,
            "students": all_users.filter(role="student", is_archived=False).count(),
            "teachers": all_users.filter(role="teacher").count(),
            "managers": all_users.filter(role="manager").count(),
            "changes": {
                "new_users": _pct_change(new_users, prev_new_users),
            },
        },
        "charts": {
            "m_labels": m_labels,
            "m_new": m_new,
            "m_act": m_act,
            "role_labels": [role_map.get(row["role"], row["role"]) for row in role_dist],
            "role_counts": [row["cnt"] for row in role_dist],
        },
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
            "prev_from": str(prev_from),
            "prev_to": str(prev_to),
        },
    }


def _overview_headline(financial, students, billing, marketing):
    revenue_change = financial["kpis"]["changes"]["revenue"]
    attendance = students["kpis"]["avg_att_rate"]
    debt = billing["kpis"]["total_debt"]
    conv_rate = marketing["kpis"]["conv_rate"]
    if revenue_change is not None and revenue_change > 0 and attendance >= 85:
        return "Daromad o'smoqda, davomat barqaror. Hozir o'sishni ushlab qolish uchun qarzdorlikni bosib turish kerak."
    if debt > financial["kpis"]["revenue"] * 0.35:
        return "Qarzdorlik darajasi yuqori. Birinchi navbatda to'lov intizomi va undirish oqimini kuchaytirish kerak."
    if conv_rate < 35:
        return "Lead oqimi bor, lekin konversiya past. Trial va follow-up sifati direktor uchun asosiy nazorat nuqtasi bo'lib turibdi."
    return "Asosiy bloklar muvozanatda. Daromad, davomat va lead sifati bo'yicha joriy davrni davomiy nazorat qilish kifoya."


def _overview_payload(center, d_from, d_to):
    financial = _financial_payload(center, d_from, d_to)
    students = _student_payload(center, d_from, d_to)
    teachers = _teacher_payload(center, d_from, d_to)
    groups = _groups_payload(center, d_from, d_to)
    billing = _billing_payload(center, d_from, d_to)
    marketing = _marketing_payload(center, d_from, d_to)
    inventory = _inventory_payload(center, d_from, d_to)
    analytics = _analytics_payload(center, d_from, d_to)

    return {
        "summary": {
            "headline": _overview_headline(financial, students, billing, marketing),
            "cards": {
                "revenue": {
                    "value": financial["kpis"]["revenue"],
                    "change": financial["kpis"]["changes"]["revenue"],
                    "meta": f"{financial['kpis']['pay_count']} ta to'lov",
                },
                "profit": {
                    "value": financial["kpis"]["net_profit"],
                    "change": financial["kpis"]["changes"]["net_profit"],
                    "meta": f"Marja {financial['kpis']['profit_margin']}%",
                },
                "active_students": {
                    "value": students["kpis"]["active_count"],
                    "change": students["kpis"]["changes"]["new_students"],
                    "meta": f"Yangi {students['kpis']['new_students']} ta",
                },
                "attendance": {
                    "value": students["kpis"]["avg_att_rate"],
                    "change": students["kpis"]["changes"]["avg_att_rate"],
                    "meta": f"Risk {students['kpis']['churn_risk']} ta",
                },
                "debt": {
                    "value": billing["kpis"]["total_debt"],
                    "change": billing["kpis"]["changes"].get("total_debt"),
                    "meta": f"Qarzdorlar {billing['kpis']['total_debtors']} ta",
                },
                "leads": {
                    "value": marketing["kpis"]["total_leads"],
                    "change": marketing["kpis"]["leads_growth"],
                    "meta": f"Konversiya {marketing['kpis']['conv_rate']}%",
                },
            },
        },
        "period": financial["period"],
        "fin": financial,
        "st": students,
        "te": teachers,
        "gr": groups,
        "bi": billing,
        "mk": marketing,
        "om": inventory,
        "an": analytics,
    }


# ══════════════════════════════════════════════════════════════════
#  1. MOLIYAVIY DASHBOARD
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def financial_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/financial.html", {
        "center": center,
        "page_title": "Moliyaviy Panel",
        "active_dash": "financial",
    })


@login_required
@require_feature("finance")
def financial_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_fin:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _financial_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  2. O'QUVCHI SAMARADORLIGI
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def student_performance_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/student_performance.html", {
        "center": center,
        "page_title": "O'quvchi Samaradorligi",
        "active_dash": "student_performance",
    })


@login_required
@require_feature("finance")
def student_performance_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_stud:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _student_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  3. USTOZ SAMARADORLIGI
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def teacher_performance_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/teacher_performance.html", {
        "center": center,
        "page_title": "Ustoz Samaradorligi",
        "active_dash": "teacher_performance",
    })


@login_required
@require_feature("finance")
def teacher_performance_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_teach:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _teacher_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  4. GURUH VA KURSLAR
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def groups_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/groups.html", {
        "center": center,
        "page_title": "Guruh va Kurslar",
        "active_dash": "groups",
    })


@login_required
def groups_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_grp:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _groups_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  5. TO'LOV VA OBUNALAR
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def billing_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/billing.html", {
        "center": center,
        "page_title": "To'lov va Obunalar",
        "active_dash": "billing",
    })


@login_required
def billing_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_bill:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _billing_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  6. MARKETING VA LIDLAR
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def marketing_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/marketing.html", {
        "center": center,
        "page_title": "Marketing va Lidlar",
        "active_dash": "marketing",
    })


@login_required
@require_feature("finance")
def marketing_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_mkt:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _marketing_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  7. INVENTAR VA MAHSULOTLAR
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("finance")
def inventory_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/inventory.html", {
        "center": center,
        "page_title": "Inventar va Mahsulotlar",
        "active_dash": "inventory",
    })


@login_required
@require_feature("finance")
def inventory_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_inv:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _inventory_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  8. SISTEMA ANALITIKASI
# ══════════════════════════════════════════════════════════════════

@login_required
@require_feature("analytics")
def analytics_dashboard(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/analytics.html", {
        "center": center,
        "page_title": "Sistema Analitikasi",
        "active_dash": "analytics",
    })


@login_required
@require_feature("analytics")
def analytics_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_ana:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _analytics_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  DASHBOARD HUB
# ══════════════════════════════════════════════════════════════════

@login_required
def dashboard_hub(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    return render(request, "core/dashboards/hub.html", {
        "center": center,
        "page_title": "Dashboard Markazi",
        "active_dash": "hub",
        "director_alerts": get_director_alerts(center),
    })


# ══════════════════════════════════════════════════════════════════
#  DIRECTOR OVERVIEW — yagona sahifali umumiy dashboard
# ══════════════════════════════════════════════════════════════════

@login_required
def director_overview(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    d_from, d_to = _parse_dates(request)
    return render(request, "core/dashboards/overview.html", {
        "center": center,
        "page_title": "Direktor Paneli",
        "active_dash": "overview",
        "date_from": d_from,
        "date_to": d_to,
    })


@login_required
def overview_api(request):
    center = _get_center(request)
    if not center:
        return _403()

    d_from, d_to = _parse_dates(request)
    ck = f"dash_overview:{center.id}:{d_from}:{d_to}"
    if (hit := cache.get(ck)):
        return JsonResponse(hit)

    data = _overview_payload(center, d_from, d_to)
    cache.set(ck, data, 900)
    return JsonResponse(data)


# ══════════════════════════════════════════════════════════════════
#  DIRECTOR BOSHQARUV — Yangi asosiy dashboard (noldan)
# ══════════════════════════════════════════════════════════════════

def _month_range_list(anchor, n=12):
    """anchor sanasidan orqaga n ta oy qaytaradi: [(m_start, m_end, label), ...]"""
    result = []
    for i in range(n - 1, -1, -1):
        raw = anchor.replace(day=1)
        for _ in range(i):
            raw = (raw.replace(day=1) - timedelta(days=1)).replace(day=1)
        m_start = raw.replace(day=1)
        m_end = m_start.replace(day=calendar.monthrange(m_start.year, m_start.month)[1])
        result.append((m_start, m_end, UZ_MONTHS[m_start.month]))
    return result


def _payments_for_scope(center, branch=None):
    qs = _payments_for_center(center)
    if branch:
        qs = qs.filter(group__branch=branch)
    return qs


def _payment_allocations_for_scope(center, branch=None):
    qs = _payment_allocations_for_center(center)
    if branch:
        qs = qs.filter(payment__group__branch=branch)
    return qs


def _deleted_payments_for_scope(center, branch=None):
    qs = _deleted_payments_for_center(center)
    if branch:
        qs = qs.filter(group__branch=branch)
    return qs


def _deleted_payment_allocations_for_scope(center, branch=None):
    qs = _deleted_payment_allocations_for_center(center)
    if branch:
        qs = qs.filter(payment__group__branch=branch)
    return qs


def _monthly_turnover_for_scope(center, m_start, m_end, branch=None):
    allocated_total = int(
        _payment_allocations_for_scope(center, branch)
        .filter(tuition_month__month__range=(m_start, m_end))
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    unallocated_total = int(
        _payments_for_scope(center, branch)
        .filter(paid_date__range=(m_start, m_end), allocations__isnull=True)
        .aggregate(s=Sum("summa"))["s"] or 0
    )
    active_total = allocated_total + unallocated_total
    if active_total:
        return active_total

    if branch is None:
        snapshot = _monthly_snapshot_for_center(center, m_start)
        if snapshot and snapshot.total_income:
            return int(snapshot.total_income or 0)

    deleted_allocated_total = int(
        _deleted_payment_allocations_for_scope(center, branch)
        .filter(tuition_month__month__range=(m_start, m_end))
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    deleted_unallocated_total = int(
        _deleted_payments_for_scope(center, branch)
        .filter(paid_date__range=(m_start, m_end), allocations__isnull=True)
        .aggregate(s=Sum("summa"))["s"] or 0
    )
    return deleted_allocated_total + deleted_unallocated_total


def _boshqaruv_payload(center, d_from, d_to, branch=None):
    """Boshqaruv dashboard uchun barcha ma'lumotlar."""
    from billing.services import clear_feature_request_cache
    clear_feature_request_cache()
    today = timezone.localdate()
    present_filter = _attendance_present_filter()

    # ── O'quvchilar ────────────────────────────────────────────
    students_qs = User.objects.filter(center=center, role="student", is_archived=False)
    students_history_qs = User.all_objects.filter(center=center, role="student")
    active_enroll = Enrollment.objects.filter(
        group__center=center,
        is_active=True,
        student__is_archived=False,
        group__is_archived=False,
        group__is_deleted=False,
    )
    if branch:
        students_qs = students_qs.filter(enrollments__group__branch=branch).distinct()
        students_history_qs = students_history_qs.filter(enrollments__group__branch=branch).distinct()
        active_enroll = active_enroll.filter(group__branch=branch)

    total_students = students_qs.count()
    active_students = active_enroll.values("student").distinct().count()
    new_this_month = students_qs.filter(date_joined__date__range=(d_from, d_to)).count()

    # ── Daromad ────────────────────────────────────────────────
    pay_qs = _payments_for_scope(center, branch).filter(paid_date__range=(d_from, d_to))
    revenue = int(pay_qs.aggregate(s=Sum("summa"))["s"] or 0)
    pay_count = pay_qs.count()

    exp_qs = _expenses_for_center(center).filter(sana__date__range=(d_from, d_to))
    expenses = int(exp_qs.aggregate(s=Sum("summa"))["s"] or 0)
    net_profit = revenue - expenses  # teacher_salary_total keyin ayiriladi

    # Oldingi davr
    period = _period_stats(d_from, d_to)
    prev_rev = int(
        _payments_for_scope(center, branch).filter(
            paid_date__range=(period["prev_from"], period["prev_to"])
        ).aggregate(s=Sum("summa"))["s"] or 0
    )
    prev_students = students_qs.filter(
        date_joined__date__range=(period["prev_from"], period["prev_to"])
    ).count()

    # ── Guruhlar ───────────────────────────────────────────────
    groups_qs = Group.objects.filter(center=center, is_archived=False)
    if branch:
        groups_qs = groups_qs.filter(branch=branch)
    total_groups = groups_qs.count()
    active_groups = groups_qs.filter(is_closed=False).count()

    # ── Davomat ────────────────────────────────────────────────
    att_qs = Attendance.objects.filter(group__center=center, date__range=(d_from, d_to))
    if branch:
        att_qs = att_qs.filter(group__branch=branch)
    att_total = att_qs.count()
    att_present = att_qs.filter(present_filter).count()
    avg_attendance = round(att_present / att_total * 100, 1) if att_total else 0
    att_excused = att_qs.filter(status="absent_excused").exclude(present_filter).count()
    att_absent = max(att_total - att_present - att_excused, 0)

    # ── O'qituvchilar va Managerlar ────────────────────────────
    teachers_qs = User.objects.filter(center=center, role="teacher", is_archived=False)
    if branch:
        teachers_qs = teachers_qs.filter(group__branch=branch).distinct()
    teachers_count = teachers_qs.count()
    managers_count = User.objects.filter(center=center, role="manager", is_archived=False).count()

    # ── Lidlar ─────────────────────────────────────────────────
    converted_filter = Q(converted_to_student=True) | Q(converted_user__isnull=False)
    leads_qs = Lead.objects.filter(center=center, qoshilgan_sana__date__range=(d_from, d_to))
    total_leads = leads_qs.count()
    converted_leads = leads_qs.filter(converted_filter).count()
    conv_rate = round(converted_leads / total_leads * 100, 1) if total_leads else 0

    # ── Chart 1 & 2: 12 oylik moliyaviy + o'quvchilar ──────────
    # Doim today dan orqaga 12 oy — filter d_from/d_to faqat KPI ga ta'sir qiladi
    monthly_labels   = []
    monthly_turnover = []
    monthly_expenses = []
    monthly_profit   = []
    monthly_students = []
    monthly_left     = []

    months_list = _month_range_list(today, 12)
    chart_start  = months_list[0][0]
    chart_end    = months_list[-1][1]

    # ── Pre-aggregate student joins & leaves (24 queries → 3) ──
    _join_map = {
        (r["y"], r["m"]): r["cnt"]
        for r in students_history_qs
        .filter(date_joined__date__range=(chart_start, chart_end))
        .values(y=F("date_joined__year"), m=F("date_joined__month"))
        .annotate(cnt=Count("id"))
    }
    _arch_map = {
        (r["y"], r["m"]): r["cnt"]
        for r in students_history_qs
        .filter(deleted_at__date__range=(chart_start, chart_end))
        .values(y=F("deleted_at__year"), m=F("deleted_at__month"))
        .annotate(cnt=Count("id"))
    }

    if branch is None:
        month_pairs = [(ms, me) for ms, me, _ in months_list]
        _bulk_turn = _bulk_monthly_turnover(center, month_pairs)
        _bulk_exp = _bulk_monthly_expenses(center, month_pairs)
        for ms, me, lbl in months_list:
            monthly_labels.append(lbl)
            m_turn = _bulk_turn.get((ms.year, ms.month), 0)
            m_exp = _bulk_exp.get((ms.year, ms.month), 0)
            monthly_turnover.append(m_turn)
            monthly_expenses.append(m_exp)
            monthly_profit.append(m_turn - m_exp)
            monthly_students.append(_join_map.get((ms.year, ms.month), 0))
            monthly_left.append(_arch_map.get((ms.year, ms.month), 0))
    else:
        for ms, me, lbl in months_list:
            monthly_labels.append(lbl)
            m_turn = _monthly_turnover_for_scope(center, ms, me, branch)
            monthly_turnover.append(m_turn)
            m_exp = _monthly_expenses_for_center(center, ms, me)
            monthly_expenses.append(m_exp)
            monthly_profit.append(m_turn - m_exp)
            monthly_students.append(_join_map.get((ms.year, ms.month), 0))
            monthly_left.append(_arch_map.get((ms.year, ms.month), 0))

    # ── Chart 3: Guruh to'ldirilganlik (N+1 → 2 queries) ──────
    _open_groups_qs = groups_qs.filter(is_closed=False).only("id", "nom")
    _fill_enroll_map = dict(
        Enrollment.objects.filter(
            group__in=_open_groups_qs,
            is_active=True,
        ).values("group_id").annotate(cnt=Count("id")).values_list("group_id", "cnt")
    )
    group_fill = [
        {"name": g.nom, "enrolled": _fill_enroll_map.get(g.id, 0)}
        for g in _open_groups_qs
    ]
    group_fill.sort(key=lambda x: -x["enrolled"])

    # ── Chart 4: Guruhlar davomat ranking ────────────────────────
    grp_att_rows = list(
        Attendance.objects.filter(
            group__center=center,
            date__range=(d_from, d_to),
            group__is_archived=False,
            group__is_deleted=False,
            **({"group__branch": branch} if branch else {}),
        )
        .values("group__nom")
        .annotate(
            tot=Count("id"),
            pres=Count("id", filter=present_filter),
        )
        .order_by("-tot")
    )
    grp_att_ranking = []
    for row in grp_att_rows:
        if not row["tot"]:
            continue
        rate = round(row["pres"] / row["tot"] * 100, 1)
        grp_att_ranking.append({
            "name": row["group__nom"] or "Guruh",
            "rate": rate,
            "present": row["pres"],
            "total": row["tot"],
        })
    # davomat foizi bo'yicha saralash (pasayish tartibida)
    grp_att_ranking.sort(key=lambda x: -x["rate"])
    grp_att_ranking = grp_att_ranking[:12]

    # ── Chart 5: Lid bosqichlari ───────────────────────────────
    # Bu blok marketing "hozirgi holat"ni ko'rsatadi: eski o'quvchiga aylangan
    # leadlar ham manba va conversion bosqichida ko'rinishi kerak.
    all_leads_scope = Lead.objects.filter(center=center)
    converted_leads_qs = all_leads_scope.filter(converted_filter)
    if branch:
        converted_leads_qs = converted_leads_qs.filter(
            converted_user__enrollments__group__branch=branch
        ).distinct()

    open_leads_scope = all_leads_scope.exclude(converted_filter)
    lost_q = (
        Q(status__code=LeadStatus.Code.LOST)
        | Q(status__nom__icontains="bekor")
        | Q(status__nom__icontains="yo'qot")
        | Q(status__nom__icontains="yo‘qot")
    )
    trial_q = (
        Q(status__code__in=[LeadStatus.Code.TRIAL_SCHEDULED, LeadStatus.Code.TRIAL_ATTENDED])
        | Q(status__nom__icontains="trial")
        | Q(status__nom__icontains="sinov")
    )
    contacted_q = (
        Q(status__code__in=[LeadStatus.Code.CONTACTED, LeadStatus.Code.NO_ANSWER])
        | Q(status__nom__icontains="bog'")
        | Q(status__nom__icontains="bog‘")
        | Q(status__nom__icontains="aloqa")
        | Q(status__nom__icontains="javob")
    )
    confirmed_q = (
        Q(is_confirmed=True)
        | Q(status__code=LeadStatus.Code.REGISTERED)
        | Q(status__nom__icontains="tasdiq")
        | Q(status__nom__icontains="register")
        | Q(status__nom__icontains="ro'yxat")
        | Q(status__nom__icontains="ro‘yxat")
    )
    funnel_trial = open_leads_scope.filter(trial_q).exclude(lost_q).count()
    funnel_contacted = open_leads_scope.filter(contacted_q).exclude(lost_q | trial_q).count()
    funnel_confirmed = open_leads_scope.filter(confirmed_q).exclude(lost_q | trial_q | contacted_q).count()
    funnel_new = open_leads_scope.exclude(lost_q | trial_q | contacted_q | confirmed_q).count()
    funnel_converted = converted_leads_qs.count()
    funnel_registered = funnel_confirmed + funnel_converted

    source_rows = list(
        converted_leads_qs
        .values("manba__nom")
        .annotate(cnt=Count("id", distinct=True))
        .order_by("-cnt", "manba__nom")[:5]
    )
    source_labels = [row["manba__nom"] or "Manba ko'rsatilmagan" for row in source_rows]
    source_counts = [int(row["cnt"] or 0) for row in source_rows]

    # ── Chart 6: Kategoriya taqsimoti (1-donut uchun) ─────────
    cat_dist = list(
        active_enroll.values("group__category_obj__name")
        .annotate(cnt=Count("student", distinct=True))
        .order_by("-cnt")[:6]
    )
    if not cat_dist:
        cat_dist = list(
            active_enroll.values("group__category")
            .annotate(cnt=Count("student", distinct=True))
            .order_by("-cnt")[:6]
        )
        cat_labels = [r.get("group__category") or "Boshqa" for r in cat_dist]
        cat_counts = [r["cnt"] for r in cat_dist]
    else:
        cat_labels = [r["group__category_obj__name"] or "Boshqa" for r in cat_dist]
        cat_counts = [r["cnt"] for r in cat_dist]

    if not cat_labels:
        cat_labels = ["IT", "Til kurslari"]
        cat_counts = [active_students // 2, active_students - active_students // 2]

    # ── Chart 7: To'lov holati (donut + kategoriya breakdown) ─────────
    # To'liq, to'lanmagan va qisman to'lovlar alohida ko'rsatiladi.
    pay_toliq = pay_tolamagan = pay_qisman = 0
    pay_category_map = {}
    enrollments_all = list(
        active_enroll.select_related("group", "group__category_obj", "student")
    )
    student_group_pairs = [
        (e.student_id, e.group_id, e.kurs_narhi or e.group.kurs_narxi or 0)
        for e in enrollments_all
    ]

    # Joriy oy uchun to'lovlarni batch olamiz
    from django.db.models import Sum as _Sum
    pay_map = {}
    if student_group_pairs:
        pay_rows = list(
            _payments_for_scope(center, branch).filter(
                paid_date__range=(d_from, d_to),
            )
            .values("student_id", "group_id")
            .annotate(total=_Sum("summa"))
        )
        for row in pay_rows:
            pay_map[(row["student_id"], row["group_id"])] = int(row["total"] or 0)

    for enr in enrollments_all:
        student_id = enr.student_id
        group_id = enr.group_id
        fee = enr.kurs_narhi or enr.group.kurs_narxi or 0
        if fee <= 0:
            continue
        paid = pay_map.get((student_id, group_id), 0)
        category_name = (
            getattr(getattr(enr.group, "category_obj", None), "name", None)
            or enr.group.get_category_display()
            or "Boshqa"
        )
        bucket = pay_category_map.setdefault(category_name, {
            "name": category_name,
            "total": 0,
            "paid": 0,
            "unpaid": 0,
            "partial": 0,
        })
        bucket["total"] += 1
        if paid >= fee:
            pay_toliq += 1
            bucket["paid"] += 1
        elif paid > 0:
            pay_qisman += 1
            bucket["partial"] += 1
        else:
            pay_tolamagan += 1
            bucket["unpaid"] += 1

    pay_category_breakdown = []
    for row in pay_category_map.values():
        total = int(row["total"] or 0)
        paid = int(row["paid"] or 0)
        unpaid = int(row["unpaid"] or 0)
        partial = int(row.get("partial") or 0)
        pay_category_breakdown.append({
            "name": row["name"],
            "total": total,
            "paid": paid,
            "unpaid": unpaid,
            "partial": partial,
            "paid_percent": round(paid / total * 100, 1) if total else 0,
            "unpaid_percent": round(unpaid / total * 100, 1) if total else 0,
            "partial_percent": round(partial / total * 100, 1) if total else 0,
        })
    pay_category_breakdown.sort(key=lambda x: (-x["total"], x["name"]))

    def _display_name(user):
        if not user:
            return "—"
        full_name = (
            getattr(user, "get_full_name", lambda: "")()
            or f"{getattr(user, 'ism', '')} {getattr(user, 'familya', '')}".strip()
            or getattr(user, "email", "")
            or getattr(user, "username", "")
        )
        return full_name.strip() or "—"

    def _group_initials(name):
        parts = [part for part in (name or "").replace("·", " ").split() if part]
        if not parts:
            return "G"
        if len(parts) == 1:
            return parts[0][:2].upper()
        return (parts[0][0] + parts[1][0]).upper()

    def _activity_dt(day, clock=None):
        if not day:
            return timezone.now()
        try:
            dt = datetime.combine(day, clock or time.min)
            if timezone.is_naive(dt):
                return timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except Exception:
            return timezone.now()

    # ── Top 5 guruh (4N+1 → 4 bulk queries) ───────────────────
    top_groups = []
    group_rows = list(groups_qs.filter(is_closed=False).select_related("oqituvchi", "category_obj"))
    _group_ids = [g.id for g in group_rows]

    _tg_enroll_map = dict(
        Enrollment.objects.filter(
            group_id__in=_group_ids,
            is_active=True,
            student__is_archived=False,
        ).values("group_id").annotate(cnt=Count("id")).values_list("group_id", "cnt")
    )
    _tg_rev_map = dict(
        pay_qs.filter(group_id__in=_group_ids)
        .values("group_id").annotate(s=Sum("summa")).values_list("group_id", "s")
    )
    _tg_att_map = {
        row["group_id"]: row
        for row in Attendance.objects.filter(
            group_id__in=_group_ids,
            date__range=(d_from, d_to),
        ).values("group_id").annotate(
            tot=Count("id"),
            pres=Count("id", filter=present_filter),
        )
    }

    for g in group_rows:
        enrolled = _tg_enroll_map.get(g.id, 0)
        raw_capacity = int(getattr(g, "max_students", 0) or 0)
        capacity = max(raw_capacity, enrolled) if (raw_capacity or enrolled) else 0
        fill_percent = round(enrolled / capacity * 100, 1) if capacity else 0
        revenue_sum = int(_tg_rev_map.get(g.id) or 0)
        _att = _tg_att_map.get(g.id, {})
        g_att_tot = _att.get("tot", 0)
        g_att_pre = _att.get("pres", 0)
        if fill_percent >= 95:
            status = "To'ldirilgan"
        elif enrolled:
            status = "Faol"
        else:
            status = "To'ldirilmoqda"
        top_groups.append({
            "id": g.id,
            "name": g.nom,
            "initials": _group_initials(g.nom),
            "teacher": _display_name(g.oqituvchi),
            "status": status,
            "enrolled": enrolled,
            "capacity": capacity,
            "fill_percent": fill_percent,
            "att_rate": round(g_att_pre / g_att_tot * 100, 1) if g_att_tot else 0,
            "revenue": revenue_sum,
        })
    top_groups.sort(key=lambda x: (-x["fill_percent"], -x["enrolled"], -x["revenue"], x["name"]))
    top_groups = top_groups[:5]

    # ── So'nggi faollik ────────────────────────────────────────
    recent_activity = []
    recent_payments = (
        _payments_for_scope(center, branch)
        .select_related("student", "group")
        .order_by("-paid_date", "-paid_time", "-id")[:8]
    )
    for payment in recent_payments:
        paid_at = _activity_dt(payment.paid_date, payment.paid_time)
        recent_activity.append({
            "type": "payment",
            "title": f"{_display_name(payment.student)} to'lov qildi",
            "subtitle": getattr(payment.group, "nom", "") or "Guruh",
            "amount": int(payment.summa or 0),
            "timestamp": paid_at.isoformat(),
            "_sort": paid_at,
        })

    absent_filter = Q(status__in=["absent_excused", "absent_unexcused"]) | Q(present=False, forced=False)
    recent_absences = (
        Attendance.objects.filter(
            group__center=center,
            date__range=(d_from, d_to),
            **({"group__branch": branch} if branch else {}),
        )
        .filter(absent_filter)
        .select_related("student", "group")
        .order_by("-date", "-id")[:8]
    )
    for attendance in recent_absences:
        happened_at = _activity_dt(attendance.date, time(23, 59))
        recent_activity.append({
            "type": "absence",
            "title": f"{_display_name(attendance.student)} darsga kelmadi",
            "subtitle": getattr(attendance.group, "nom", "") or "Guruh",
            "amount": None,
            "timestamp": happened_at.isoformat(),
            "_sort": happened_at,
        })

    recent_converted_leads = (
        converted_leads_qs
        .select_related("yonalish", "manba")
        .order_by("-converted_at", "-qoshilgan_sana", "-id")[:8]
    )
    for lead in recent_converted_leads:
        converted_at = lead.converted_at or lead.qoshilgan_sana or timezone.now()
        subtitle = (
            getattr(getattr(lead, "yonalish", None), "nom", "")
            or getattr(getattr(lead, "manba", None), "nom", "")
            or "CRM"
        )
        recent_activity.append({
            "type": "lead",
            "title": f"{lead.full_name or str(lead)} ro'yxatdan o'tdi",
            "subtitle": subtitle,
            "amount": None,
            "timestamp": converted_at.isoformat(),
            "_sort": converted_at,
        })

    recent_activity.sort(key=lambda x: x["_sort"], reverse=True)
    recent_activity = [
        {key: value for key, value in item.items() if key != "_sort"}
        for item in recent_activity[:7]
    ]

    # ── Top 5 o'quvchi ────────────────────────────────────────
    raw_att = list(
        att_qs.values("student__ism", "student__familya")
        .annotate(tot=Count("id"), pres=Count("id", filter=present_filter))
    )
    top_students = sorted(
        [
            {
                "name": f"{r['student__ism'] or ''} {r['student__familya'] or ''}".strip() or "Noma'lum",
                "rate": round(r["pres"] / r["tot"] * 100, 1) if r["tot"] else 0,
            }
            for r in raw_att
        ],
        key=lambda x: -x["rate"],
    )[:5]

    # ── Xavfli o'quvchilar (tanlangan davr ichida 3+ dars qoldirgan) ────
    xavfli_students = []
    try:
        absent_filter = Q(status__in=["absent_excused", "absent_unexcused"]) | Q(present=False, forced=False)
        xavfli_raw = list(
            Attendance.objects.filter(
                group__center=center,
                date__range=(d_from, d_to),
                **({"group__branch": branch} if branch else {}),
            )
            .filter(absent_filter)
            .values(
                "student_id",
                "student__ism",
                "student__familya",
                "student__telefon1",
                "group__nom",
            )
            .annotate(missed=Count("id"))
            .filter(missed__gte=3)
            .order_by("-missed")
        )
        # group by student — agar bir o'quvchi bir nechta guruhda bo'lsa birlashtiramiz
        _st_map = {}
        for row in xavfli_raw:
            sid = row["student_id"]
            if sid not in _st_map:
                _st_map[sid] = {
                    "name": f"{row['student__ism'] or ''} {row['student__familya'] or ''}".strip() or "Noma'lum",
                    "phone": row["student__telefon1"] or "—",
                    "groups": [],
                    "missed": 0,
                }
            _st_map[sid]["groups"].append(row["group__nom"] or "Guruh")
            _st_map[sid]["missed"] += row["missed"]

        # har bir student uchun jami davomatni hisoblash (foiz uchun)
        if _st_map:
            xavfli_att_raw = list(
                Attendance.objects.filter(
                    group__center=center,
                    date__range=(d_from, d_to),
                    student_id__in=list(_st_map.keys()),
                    **({"group__branch": branch} if branch else {}),
                )
                .values("student_id")
                .annotate(
                    tot=Count("id"),
                    pres=Count("id", filter=present_filter),
                )
            )
            _att_map = {r["student_id"]: r for r in xavfli_att_raw}
        else:
            _att_map = {}

        for sid, info in _st_map.items():
            att_info = _att_map.get(sid, {})
            tot = att_info.get("tot", 0)
            pres = att_info.get("pres", 0)
            rate = round(pres / tot * 100, 1) if tot else 0
            xavfli_students.append({
                "name": info["name"],
                "phone": info["phone"],
                "groups": ", ".join(info["groups"]),
                "missed": info["missed"],
                "rate": rate,
            })
        xavfli_students.sort(key=lambda x: (-x["missed"], x["rate"]))
        xavfli_students = xavfli_students[:20]
    except Exception:
        xavfli_students = []

    # ── Qarzdorlar — Qarzdorlar bo'limi bilan AYNAN bir xil hisob ──────────
    # YAGONA MANBA: center_month_debt_summary. education/views.py dagi
    # qarzdorlar_home ham AYNAN shu funksiyani ishlatadi => ikkala raqam
    # 100% teng bo'ladi (kamaysa kamayadi, ko'paysa ko'payadi).
    # d_to oyi uchun "hozirgi qarz" ko'rsatiladi (default = joriy oy).
    total_debt = 0
    total_debtors = 0
    try:
        from education.services.tuition import (
            center_month_debt_summary as _cmds, month_first_day as _mfd,
        )
        total_debt, total_debtors = _cmds(center, [_mfd(d_to)], branch=branch)
    except Exception:
        total_debt = 0
        total_debtors = 0

    # ── O'qituvchi maoshi (oylik chiqim qismi) ─────────────────────
    teacher_salary_total = 0
    try:
        from education.models import TeacherIncome
        ti_qs = TeacherIncome.objects.filter(
            center=center, attendance__date__range=(d_from, d_to)
        )
        if branch:
            ti_qs = ti_qs.filter(group__branch=branch)
        teacher_salary_total = int(ti_qs.aggregate(s=Sum("amount"))["s"] or 0)
    except Exception:
        teacher_salary_total = 0

    # ── Oldingi davr KPI deltalari (har bir karta uchun %) ─────────
    _pf, _pt = period["prev_from"], period["prev_to"]

    # Sof foyda — oldingi davr (teacher_salary ham keyin ayiriladi)
    try:
        prev_exp_qs = _expenses_for_center(center).filter(sana__date__range=(_pf, _pt))
        prev_expenses_val = int(prev_exp_qs.aggregate(s=Sum("summa"))["s"] or 0)
    except Exception:
        prev_expenses_val = 0
    prev_net_profit = prev_rev - prev_expenses_val

    # Aktiv o'quvchilar — oldingi davr oxiriga snapshot
    try:
        prev_active_students = students_qs.filter(date_joined__date__lte=_pt).count()
    except Exception:
        prev_active_students = 0

    # Lidlar — oldingi davr
    try:
        prev_leads = Lead.objects.filter(
            center=center,
            qoshilgan_sana__date__range=(_pf, _pt),
        ).count()
    except Exception:
        prev_leads = 0

    # Qarzdorlar (snapshot) — oldingi davr oxirida
    prev_total_debt = 0
    try:
        from education.services.tuition import calculate_enrollment_debt_snapshots, month_first_day, preload_enrollment_history_starts, preload_group_schedules
        from education.models import TuitionMonth as _TM_prev

        _prev_active = Enrollment.objects.filter(
            group__center=center,
            is_active=True,
            student__is_archived=False,
            is_deferred=False,
        )
        if branch:
            _prev_active = _prev_active.filter(group__branch=branch)

        # Faqat TuitionMonth yozuvi mavjud bo'lgan inactive enrollment'lar
        # (current period block dagi kabi — ALL inactive'ni fetch qilish N+1 manbai)
        _prev_inactive_ids = (
            _TM_prev.objects.filter(
                enrollment__group__center=center,
                enrollment__is_active=False,
                is_deleted=False,
                enrollment__student__is_archived=False,
            ).values_list("enrollment_id", flat=True).distinct()
        )
        if branch:
            _prev_inactive_ids = (
                _TM_prev.objects.filter(
                    enrollment__group__center=center,
                    enrollment__group__branch=branch,
                    enrollment__is_active=False,
                    is_deleted=False,
                    enrollment__student__is_archived=False,
                ).values_list("enrollment_id", flat=True).distinct()
            )
        _prev_inactive = Enrollment.objects.filter(id__in=_prev_inactive_ids)

        _prev_enrs = list(_prev_active) + list(_prev_inactive)

        if _prev_enrs:
             preload_enrollment_history_starts(_prev_enrs)
             preload_group_schedules({e.group_id for e in _prev_enrs if e.group_id})
             prev_snaps = calculate_enrollment_debt_snapshots(
                 _prev_enrs,
                 [month_first_day(_pt)],
                 cumulative_up_to=_pt
             )
             for snap in prev_snaps.values():
                 prev_total_debt += int(snap.get("net_cumulative_debt", 0) or 0)

    except Exception:
        prev_total_debt = 0

    # O'qituvchi maoshi — oldingi davr
    prev_teacher_salary = 0
    try:
        from education.models import TeacherIncome
        prev_ti_qs = TeacherIncome.objects.filter(
            center=center, attendance__date__range=(_pf, _pt)
        )
        if branch:
            prev_ti_qs = prev_ti_qs.filter(group__branch=branch)
        prev_teacher_salary = int(prev_ti_qs.aggregate(s=Sum("amount"))["s"] or 0)
    except Exception:
        prev_teacher_salary = 0

    # Hodimlar (oldingi davr boshigacha bo'lganlar — snapshot)
    try:
        prev_teachers_count = teachers_qs.filter(date_joined__date__lte=_pt).count()
    except Exception:
        prev_teachers_count = teachers_count
    try:
        prev_managers_count = User.objects.filter(
            center=center, role="manager", is_archived=False,
            date_joined__date__lte=_pt,
        ).count()
    except Exception:
        prev_managers_count = managers_count

    # ── To'lov turlari (PaymentMethod taqsimoti) ───────────────────
    pay_method_labels = []
    pay_method_counts = []
    pay_method_amounts = []
    try:
        method_rows = list(
            pay_qs.values("payment_type")
            .annotate(cnt=Count("id"), total=Sum("summa"))
            .order_by("-total")
        )
        _label_map = {"cash": "Naqd", "card": "Karta", "mixed": "Aralash"}
        for row in method_rows:
            code = (row.get("payment_type") or "").strip()
            label = _label_map.get(code.lower(), code.upper() or "Boshqa")
            pay_method_labels.append(label)
            pay_method_counts.append(int(row.get("cnt") or 0))
            pay_method_amounts.append(int(row.get("total") or 0))
    except Exception:
        pay_method_labels = []
        pay_method_counts = []
        pay_method_amounts = []

    # Kunlik davomat nazorati — o'qituvchi davomat qilmagan guruhlar (manager+director).
    from education.services.attendance_monitor import get_attendance_monitor
    try:
        attendance_monitor = get_attendance_monitor(center)
    except Exception:
        attendance_monitor = {"date": "", "summary": {}, "rows": []}

    return {
        "attendance_monitor": attendance_monitor,
        "kpis": {
            "total_students": total_students,
            "active_students": active_students,
            "new_this_month": new_this_month,
            "revenue": revenue,
            # Sof foyda = daromad − (manager kiritgan xarajatlar). O'qituvchi
            # maoshi endi alohida ayirilmaydi — u ham xarajat bo'lib, manager
            # xarajatlar bo'limiga o'zi yozib boradi.
            "net_profit": revenue - expenses,
            "expenses": expenses,
            "pay_count": pay_count,
            "total_groups": total_groups,
            "active_groups": active_groups,
            "avg_attendance": avg_attendance,
            "teachers_count": teachers_count,
            "managers_count": managers_count,
            "total_leads": total_leads,
            "conv_rate": conv_rate,
            "total_debt": total_debt,
            "total_debtors": total_debtors,
            "teacher_salary_total": teacher_salary_total,
            "changes": {
                "revenue": _pct_change(revenue, prev_rev),
                "expenses": _pct_change(expenses, prev_expenses_val),
                "students": _pct_change(new_this_month, prev_students),
                "net_profit": _pct_change(
                    revenue - expenses,
                    prev_rev - prev_expenses_val,
                ),
                "active_students": _pct_change(active_students, prev_active_students),
                "teachers": _pct_change(teachers_count, prev_teachers_count),
                "managers": _pct_change(managers_count, prev_managers_count),
                "leads": _pct_change(total_leads, prev_leads),
                "total_debt": _pct_change(total_debt, prev_total_debt),
                "teacher_salary": _pct_change(teacher_salary_total, prev_teacher_salary),
            },
        },
        "charts": {
            "monthly_labels": monthly_labels,
            "monthly_turnover": monthly_turnover,
            "monthly_expenses": monthly_expenses,
            "monthly_profit": monthly_profit,
            "monthly_students": monthly_students,
            "monthly_left": monthly_left,
            "group_names": [g["name"] for g in group_fill[:8]],
            "group_enrolled": [g["enrolled"] for g in group_fill[:8]],
            "grp_att_ranking": grp_att_ranking,
            "attendance_labels": ["Kelgan", "Kelmagan", "Sababli"],
            "attendance_counts": [att_present, att_absent, att_excused],
            "funnel": [
                funnel_new,
                funnel_contacted,
                funnel_trial,
                funnel_registered,
            ],
            "funnel_labels": [
                "Yangi lid",
                "Bog'langan",
                "Trial darsda",
                "Ro'yxatdan o'tdi",
            ],
            "source_labels": source_labels,
            "source_counts": source_counts,
            "cat_labels": cat_labels,
            "cat_counts": cat_counts,
            "pay_status_labels": ["To'lagan", "To'lamagan", "Qisman"],
            "pay_status_counts": [pay_toliq, pay_tolamagan, pay_qisman],
            "pay_category_breakdown": pay_category_breakdown,
            "pay_method_labels": pay_method_labels,
            "pay_method_counts": pay_method_counts,
            "pay_method_amounts": pay_method_amounts,
        },
        "group_fill_all": group_fill,
        "top_groups": top_groups,
        "recent_activity": recent_activity,
        "top_students": top_students,
        "xavfli_students": xavfli_students,
        "period": {
            "from": str(d_from),
            "to": str(d_to),
            "days": period["days"],
        },
        "generated_at": timezone.now().isoformat(),
    }


@login_required
def director_boshqaruv(request):
    center = _get_center(request)
    if not center:
        return redirect("core:home")
    d_from, d_to = _parse_dates(request)
    pending_purchase_requests = _build_pending_purchase_requests(center)
    # Active payment methods for payment modal
    active_payment_methods = []
    try:
        from store.models import PaymentMethod as _PM
        from store.views import _ensure_default_payment_methods as _seed_pm
        _seed_pm(center)
        active_payment_methods = list(
            _PM.objects.filter(center=center, is_active=True).values_list('nom', flat=True).order_by('nom')
        )
    except Exception:
        active_payment_methods = []

    return render(request, "core/dashboards/boshqaruv.html", {
        "center": center,
        "date_from": d_from,
        "date_to": d_to,
        "pending_purchase_requests": pending_purchase_requests["recent"],
        "pending_purchase_requests_count": pending_purchase_requests["count"],
        "active_payment_methods": active_payment_methods,
    })


def _build_pending_purchase_requests(center, limit: int = 6):
    """Center bo'yicha kutilayotgan xarid so'rovlari ro'yxati va umumiy soni."""
    try:
        from store.models import PurchaseRequest
    except Exception:
        return {"recent": [], "count": 0}
    qs = (
        PurchaseRequest.objects.filter(center=center, status=PurchaseRequest.PENDING)
        .select_related("student", "product")
        .order_by("-sana")
    )
    count = qs.count()
    recent = []
    for req in qs[:limit]:
        student = req.student
        product = req.product
        student_name = ""
        if student:
            student_name = (
                f"{getattr(student, 'familya', '') or ''} {getattr(student, 'ism', '') or ''}".strip()
                or student.email
            )
        try:
            price_chaqmoq = int(getattr(product, "narx_chaqmoq", 0) or 0) * int(req.qty or 1)
        except Exception:
            price_chaqmoq = 0
        recent.append({
            "id": req.id,
            "student_id": getattr(student, "id", None),
            "student_name": student_name or "—",
            "product_name": getattr(product, "nom", None) or "O‘chirilgan mahsulot",
            "qty": req.qty,
            "price_chaqmoq": price_chaqmoq,
            "sana": req.sana,
        })
    return {"recent": recent, "count": count}


@login_required
def director_boshqaruv_api(request):
    center = _get_center(request)
    if not center:
        return _403()
    d_from, d_to = _parse_dates(request)
    branch_id = request.GET.get("branch_id")
    branch = None
    if branch_id:
        try:
            branch = Branch.objects.get(pk=int(branch_id), center=center)
        except (Branch.DoesNotExist, ValueError, TypeError):
            branch = None

    from core.perf_cache import TTL_MEDIUM, perf_cache_get_or_set, versioned_cache_key
    _cache_key = versioned_cache_key(
        "boshqaruv_api",
        getattr(center, "id", None),
        d_from.isoformat() if d_from else "",
        d_to.isoformat() if d_to else "",
        branch_id or "",
    )
    data = perf_cache_get_or_set(
        _cache_key,
        lambda: _boshqaruv_payload(center, d_from, d_to, branch=branch),
        ttl=TTL_MEDIUM,
    )
    return JsonResponse(data)


# ── Faollik tarixi ───────────────────────────────────────────────
@login_required
def director_activity_history(request):
    """Boshqaruv → So'nggi faollik → Hammasi.

    Oxirgi 30 kun ichidagi to'lov, davomat (yo'qlik), lid konvertatsiya va
    guruhdan chiqish hodisalarini bir joyga to'playdi. Type filter va
    sahifalash bilan ko'rsatadi.
    """
    from django.core.paginator import Paginator
    from education.models import StudentGroupHistory as _SGH

    center = _get_center(request)
    if not center:
        return redirect("core:home")

    today_now = timezone.localdate()
    range_from = today_now - timedelta(days=30)

    type_filter = (request.GET.get("type") or "all").strip().lower()
    valid_types = {"all", "payment", "lead", "absence", "leave", "demo", "reward"}
    if type_filter not in valid_types:
        type_filter = "all"

    def _name(user):
        if not user:
            return "—"
        full = f"{getattr(user, 'ism', '') or ''} {getattr(user, 'familya', '') or ''}".strip()
        return full or getattr(user, "username", "—") or "—"

    def _dt(d, t=None):
        if not d:
            return timezone.now()
        try:
            dt = datetime.combine(d, t or time.min)
            if timezone.is_naive(dt):
                return timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except Exception:
            return timezone.now()

    events = []

    # ── Payments ─────────────────────────────────────────────────
    pay_qs = Payment.objects.filter(
        center=center,
        paid_date__gte=range_from,
        paid_date__lte=today_now,
    ).select_related("student", "group").order_by("-paid_date", "-paid_time", "-id")
    for p in pay_qs[:300]:
        amt = int(p.summa or 0)
        is_partial = False
        try:
            full_fee = int(getattr(getattr(p, "tuition_month", None), "fee_amount", 0) or 0)
            if full_fee and amt < full_fee:
                is_partial = True
        except Exception:
            pass
        title = f"{_name(p.student)} {'qisman ' if is_partial else ''}to'lov qildi"
        events.append({
            "type": "payment",
            "title": title,
            "subtitle": getattr(p.group, "nom", "") or "Guruh",
            "amount": amt,
            "amount_negative": False,
            "ts": _dt(p.paid_date, p.paid_time),
        })

    # ── Absences ─────────────────────────────────────────────────
    absent_q = Q(status__in=["absent_excused", "absent_unexcused"]) | Q(present=False, forced=False)
    abs_qs = (
        Attendance.objects.filter(
            group__center=center,
            date__gte=range_from,
            date__lte=today_now,
        )
        .filter(absent_q)
        .select_related("student", "group")
        .order_by("-date", "-id")
    )
    for a in abs_qs[:300]:
        events.append({
            "type": "absence",
            "title": f"{_name(a.student)} darsga kelmadi",
            "subtitle": getattr(a.group, "nom", "") or "Guruh",
            "amount": None,
            "amount_negative": False,
            "ts": _dt(a.date, time(23, 30)),
        })

    # ── Lead conversions ─────────────────────────────────────────
    converted_filter = Q(converted_to_student=True) | Q(converted_user__isnull=False)
    lead_qs = (
        Lead.objects.filter(center=center)
        .filter(converted_filter)
        .filter(
            Q(converted_at__date__gte=range_from)
            | Q(qoshilgan_sana__gte=range_from)
        )
        .select_related("yonalish", "manba")
        .order_by("-converted_at", "-qoshilgan_sana", "-id")
    )
    for l in lead_qs[:300]:
        when = l.converted_at or l.qoshilgan_sana or timezone.now()
        if hasattr(when, "tzinfo") and when.tzinfo is None:
            when = timezone.make_aware(when, timezone.get_current_timezone())
        if not hasattr(when, "tzinfo"):
            when = _dt(when, time(12, 0))
        sub = (
            getattr(getattr(l, "yonalish", None), "nom", "")
            or getattr(getattr(l, "manba", None), "nom", "")
            or "CRM"
        )
        events.append({
            "type": "lead",
            "title": f"{l.full_name or str(l)} ro'yxatdan o'tdi",
            "subtitle": sub,
            "amount": None,
            "amount_negative": False,
            "ts": when,
        })

    # ── Group leaves ─────────────────────────────────────────────
    leave_qs = (
        _SGH.objects.filter(
            center=center,
            end_date__isnull=False,
            end_date__gte=range_from,
            end_date__lte=today_now,
        )
        .select_related("student", "group")
        .order_by("-end_date", "-id")
    )
    for h in leave_qs[:300]:
        events.append({
            "type": "leave",
            "title": f"{_name(h.student)} guruhdan chiqdi",
            "subtitle": getattr(h.group, "nom", "") or "Guruh",
            "amount": None,
            "amount_negative": False,
            "ts": _dt(h.end_date, time(18, 0)),
        })

    # ── Hammasini sana bo'yicha tartiblash ──
    events.sort(key=lambda e: e["ts"], reverse=True)

    # Type counts (for chip badges) — based on ALL events, not filtered.
    counts = {"all": len(events), "payment": 0, "lead": 0, "absence": 0, "leave": 0, "demo": 0, "reward": 0}
    for e in events:
        if e["type"] in counts:
            counts[e["type"]] += 1

    # Apply type filter
    if type_filter != "all":
        events = [e for e in events if e["type"] == type_filter]

    # Pagination: 10 per page
    paginator = Paginator(events, 10)
    page_num = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_num)
    except Exception:
        page_obj = paginator.page(1)

    # Group current page by date label (Bugun / Kecha / "DD MMM YYYY")
    yest = today_now - timedelta(days=1)
    UZ_MONTHS = {
        1: "Yan", 2: "Fev", 3: "Mar", 4: "Apr", 5: "May", 6: "Iyn",
        7: "Iyl", 8: "Avg", 9: "Sen", 10: "Okt", 11: "Noy", 12: "Dek",
    }

    def _date_label(d):
        if d == today_now:
            return "Bugun"
        if d == yest:
            return "Kecha"
        return f"{d.day} {UZ_MONTHS.get(d.month, '')} {d.year}"

    def _rel_time(ts):
        if not ts:
            return ""
        delta = (timezone.now() - ts).total_seconds()
        if delta < 60:
            return "hozirgina"
        if delta < 3600:
            return f"{int(delta // 60)} daqiqa oldin"
        if delta < 86400:
            return f"{int(delta // 3600)} soat oldin"
        if hasattr(ts, "strftime"):
            return ts.strftime("%H:%M")
        return ""

    grouped = []
    current_label = None
    current_group = None
    for ev in page_obj.object_list:
        d_local = timezone.localtime(ev["ts"]).date() if hasattr(ev["ts"], "tzinfo") and ev["ts"].tzinfo else ev["ts"].date() if hasattr(ev["ts"], "date") else today_now
        lbl = _date_label(d_local)
        ev["rel_time"] = _rel_time(ev["ts"])
        if lbl != current_label:
            current_label = lbl
            current_group = {"label": lbl, "events": []}
            grouped.append(current_group)
        current_group["events"].append(ev)

    # Decorate each group with count
    for g in grouped:
        g["count"] = len(g["events"])

    # Preserve qs for paginator
    base_qs = request.GET.copy()
    if "page" in base_qs:
        base_qs.pop("page")
    base_qs_str = base_qs.urlencode()

    return render(request, "core/recent_activity_history.html", {
        "type_filter": type_filter,
        "counts": counts,
        "grouped": grouped,
        "page_obj": page_obj,
        "paginator": paginator,
        "base_qs": base_qs_str,
        "total_events_in_filter": paginator.count,
    })


# ── AI CHAT ──────────────────────────────────────────────────────

import json as _json

@login_required
def director_boshqaruv_export(request):
    """Direktor dashboardining KPI snapshot'ini Excel (.xlsx) sifatida yuklab olish."""
    from django.http import HttpResponse
    center = _get_center(request)
    if not center:
        return _403()
    user_role = getattr(request.user, "role", None)
    if user_role == "manager" and not request.user.is_superuser:
        return _403()

    d_from, d_to = _parse_dates(request)
    branch_id = request.GET.get("branch_id")
    branch = None
    if branch_id:
        try:
            branch = Branch.objects.get(pk=int(branch_id), center=center)
        except (Branch.DoesNotExist, ValueError, TypeError):
            branch = None

    data = _boshqaruv_payload(center, d_from, d_to, branch=branch)
    kpis = data.get("kpis", {})
    charts = data.get("charts", {})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"

    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    label_font = Font(bold=True, color="475569")

    ws["A1"] = f"Director Dashboard — {center.name}"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Davr"
    ws["B2"] = f"{d_from} → {d_to}"
    ws["A2"].font = label_font

    rows = [
        ("Daromad (so'm)", kpis.get("revenue", 0)),
        ("Sof foyda (so'm)", kpis.get("net_profit", 0)),
        ("Xarajatlar (so'm)", kpis.get("expenses", 0)),
        ("Qarzdorlik (so'm)", kpis.get("total_debt", 0)),
        ("Qarzdor o'quvchilar", kpis.get("total_debtors", 0)),
        ("O'qituvchi maoshi (so'm)", kpis.get("teacher_salary_total", 0)),
        ("To'lovlar soni", kpis.get("pay_count", 0)),
        ("Faol o'quvchilar", kpis.get("active_students", 0)),
        ("Jami o'quvchilar", kpis.get("total_students", 0)),
        ("Yangi o'quvchilar", kpis.get("new_this_month", 0)),
        ("O'qituvchilar", kpis.get("teachers_count", 0)),
        ("Managerlar", kpis.get("managers_count", 0)),
        ("Lidlar", kpis.get("total_leads", 0)),
        ("Konversiya (%)", kpis.get("conv_rate", 0)),
        ("O'rtacha davomat (%)", kpis.get("avg_attendance", 0)),
    ]
    start_row = 4
    ws[f"A{start_row}"] = "Ko'rsatkich"
    ws[f"B{start_row}"] = "Qiymat"
    ws[f"A{start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"B{start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{start_row}"].fill = header_fill
    ws[f"B{start_row}"].fill = header_fill
    for i, (label, value) in enumerate(rows, start=start_row + 1):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    # 2-sheet: 12 oylik trend
    ws2 = wb.create_sheet("Trend")
    ws2["A1"] = "Oy"
    ws2["B1"] = "Aylanma"
    ws2["C1"] = "Foyda"
    ws2["D1"] = "Xarajat"
    for col in ("A", "B", "C", "D"):
        ws2[f"{col}1"].font = Font(bold=True, color="FFFFFF")
        ws2[f"{col}1"].fill = header_fill
    labels = charts.get("monthly_labels", [])
    turn = charts.get("monthly_turnover", [])
    prof = charts.get("monthly_profit", [])
    exp = charts.get("monthly_expenses", [])
    for i in range(len(labels)):
        ws2.cell(row=i + 2, column=1, value=labels[i])
        ws2.cell(row=i + 2, column=2, value=turn[i] if i < len(turn) else 0)
        ws2.cell(row=i + 2, column=3, value=prof[i] if i < len(prof) else 0)
        ws2.cell(row=i + 2, column=4, value=exp[i] if i < len(exp) else 0)
    for col in ("A", "B", "C", "D"):
        ws2.column_dimensions[col].width = 16

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="boshqaruv_{d_from}_{d_to}.xlsx"'
    )
    wb.save(response)
    return response


_AI_CHAT_HISTORY_LIMIT = 30
_AI_CHAT_QUESTION_MAX = 2000


_AI_CHAT_ALLOWED_ROLES = {"director", "manager", "teacher", "student", "parent"}
_AI_CHAT_TITLES = {
    "director": "Direktor AI chat",
    "manager": "Manager AI chat",
    "teacher": "Ustoz AI chat",
    "student": "O'quvchi AI chat",
    "parent": "Ota-ona AI chat",
}


def _ai_auth_center(request):
    """Auth + center check shared by all AI endpoints. Returns (center, role) or (None, None)."""
    user = request.user
    if not user.is_authenticated:
        return None, None
    role = getattr(user, "role", None)
    if not (user.is_superuser or role in _AI_CHAT_ALLOWED_ROLES):
        return None, None
    center = getattr(request, "center", None) or getattr(user, "center", None)
    if not center:
        return None, None
    if not user.is_superuser and not getattr(center, "ai_enabled", False):
        return None, None
    if not user.is_superuser and role not in ("director", "manager"):
        role_flag = {
            "teacher": "ai_teacher_enabled",
            "student": "ai_student_enabled",
            "parent":  "ai_parent_enabled",
        }.get(role)
        if role_flag and not getattr(center, role_flag, False):
            return None, None
    return center, role


def _ai_chat_session(request, session_id=None):
    """Return (center, session) for any allowed role or None if unauthorized.

    session_id — ixtiyoriy. Berilsa, o'sha sessiyani qaytaradi (agar foydalanuvchiga tegishli bo'lsa).
    Berilmasa — eng oxirgi sessiya, bo'lmasa yangi yaratadi.
    """
    center, role = _ai_auth_center(request)
    if not center:
        return None, None

    from core.models import DirectorAIChatSession
    user = request.user

    if session_id:
        try:
            session = DirectorAIChatSession.objects.get(
                id=int(session_id), center=center, user=user
            )
            return center, session
        except (DirectorAIChatSession.DoesNotExist, ValueError, TypeError):
            return None, None

    # Most recent session
    session = (
        DirectorAIChatSession.objects
        .filter(center=center, user=user)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not session:
        title = _AI_CHAT_TITLES.get(role, "AI chat")
        session = DirectorAIChatSession.objects.create(center=center, user=user, title=title)
    return center, session


def _get_session_id(request):
    """Extract session_id from GET params or JSON body."""
    sid = request.GET.get("session_id") or request.GET.get("sid")
    if sid:
        return sid
    if request.content_type and "json" in request.content_type:
        try:
            import json as _j
            body = _j.loads(request.body or "{}")
            return body.get("session_id") or body.get("sid")
        except Exception:
            pass
    return None


@login_required
def director_boshqaruv_chat(request):
    """AI chat endpoint — POST {question, session_id?} → saves to DB."""
    if request.method != "POST":
        return JsonResponse({"error": "Faqat POST"}, status=405)

    import json as _json
    try:
        body = _json.loads(request.body or b"{}")
    except Exception:
        return JsonResponse({"error": "JSON xato"}, status=400)

    sid = body.get("session_id") or body.get("sid") or request.GET.get("session_id")
    center, session = _ai_chat_session(request, session_id=sid)
    if not center or not session:
        return _403()

    question = (body.get("question") or "").strip()
    if not question:
        return JsonResponse({"error": "Savol bo'sh"}, status=400)
    if len(question) > _AI_CHAT_QUESTION_MAX:
        question = question[:_AI_CHAT_QUESTION_MAX]

    from core.models import DirectorAIChatMessage

    # 1) Saqlangan tarixni DB dan olish (oxirgi N ta — modeldan oldin yozilgan)
    prev_qs = (
        DirectorAIChatMessage.objects
        .filter(session=session)
        .order_by("-created_at", "-id")[:_AI_CHAT_HISTORY_LIMIT]
    )
    history = [
        {"role": m.role, "content": m.content}
        for m in reversed(list(prev_qs))
    ]

    # 2) Foydalanuvchi savolini darhol saqlaymiz
    DirectorAIChatMessage.objects.create(
        session=session,
        role=DirectorAIChatMessage.ROLE_USER,
        content=question,
    )

    # 3) Rol bo'yicha mos AI funksiyasini chaqirish
    user_role = getattr(request.user, "role", None)
    if user_role in {"teacher", "student", "parent"}:
        try:
            from core.services.role_scoped_ai import answer_role_scoped_question
            answer, source = answer_role_scoped_question(
                viewer=request.user,
                question=question,
                history=history,
            )
        except Exception as e:
            answer = f"Hozircha AI javob bera olmayapti. Xato: {str(e)[:120]}"
            source = "error"
    else:
        # director / manager / superuser — markaz darajasidagi to'liq kontekst
        today = timezone.localdate()
        d_from = today.replace(day=1)
        stats = _boshqaruv_payload(center, d_from, today)
        # AI engine davrni bilishi uchun system qismini qo'shamiz
        stats["system"] = {
            "start_date": d_from.isoformat(),
            "end_date": today.isoformat(),
            "last_updated": timezone.now().isoformat(),
        }

        try:
            from core.services.ai_insights import answer_question_structured_bundle
            answer, source, _ = answer_question_structured_bundle(
                center=center,
                question=question,
                stats=stats,
                history=history,
                viewer=request.user,
            )
        except Exception as e:
            answer = f"Hozircha AI javob bera olmayapti. Xato: {str(e)[:120]}"
            source = "error"

    # 4) AI javobini saqlaymiz (xato bo'lsa ham — tarix uchun)
    DirectorAIChatMessage.objects.create(
        session=session,
        role=DirectorAIChatMessage.ROLE_ASSISTANT,
        content=answer,
        source=source or "",
    )

    # session.updated_at ni yangilash uchun save() (auto_now ishlashi uchun)
    session.save(update_fields=["updated_at"])

    return JsonResponse({"answer": answer, "source": source, "session_id": session.id})


@login_required
def director_boshqaruv_chat_history(request):
    """AI chat tarixi — GET ?session_id= → oxirgi xabarlar ro'yxati."""
    if request.method != "GET":
        return JsonResponse({"error": "Faqat GET"}, status=405)

    sid = request.GET.get("session_id") or request.GET.get("sid")
    center, session = _ai_chat_session(request, session_id=sid)
    if not center or not session:
        return _403()

    from core.models import DirectorAIChatMessage

    qs = (
        DirectorAIChatMessage.objects
        .filter(session=session)
        .order_by("created_at", "id")[:200]
    )
    messages = [
        {
            "role": m.role,
            "content": m.content,
            "created_at": m.created_at.isoformat(),
        }
        for m in qs
    ]
    return JsonResponse({"messages": messages, "session_id": session.id})


@login_required
def director_boshqaruv_chat_clear(request):
    """AI chat tarixini tozalash — POST {session_id?}."""
    if request.method != "POST":
        return JsonResponse({"error": "Faqat POST"}, status=405)

    import json as _json
    try:
        body = _json.loads(request.body or b"{}")
    except Exception:
        body = {}
    sid = body.get("session_id") or request.GET.get("session_id")
    center, session = _ai_chat_session(request, session_id=sid)
    if not center or not session:
        return _403()

    from core.models import DirectorAIChatMessage
    DirectorAIChatMessage.objects.filter(session=session).delete()
    return JsonResponse({"ok": True, "session_id": session.id})


@login_required
def director_ai_role_settings(request):
    """Director/Manager AI rollarini yoqish/o'chirish — POST JSON."""
    if request.method != "POST":
        return JsonResponse({"error": "Faqat POST"}, status=405)

    user = request.user
    role = getattr(user, "role", None)
    if not (user.is_superuser or role in ("director", "manager")):
        return _403()

    center = getattr(request, "center", None) or getattr(user, "center", None)
    if not center:
        return _403()

    if not user.is_superuser and not getattr(center, "ai_enabled", False):
        return JsonResponse({"error": "AI bu markaz uchun yoqilmagan"}, status=403)

    import json as _json
    try:
        payload = _json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "JSON talab qilinadi"}, status=400)

    allowed_flags = {"ai_teacher_enabled", "ai_student_enabled", "ai_parent_enabled"}
    updates = {k: bool(v) for k, v in payload.items() if k in allowed_flags}
    if not updates:
        return JsonResponse({"error": "Hech qanday o'zgarish yo'q"}, status=400)

    from accounts.models import Center as _Center
    _Center.objects.filter(pk=center.pk).update(**updates)
    # Refresh cached center object on request
    for k, v in updates.items():
        setattr(center, k, v)

    return JsonResponse({"ok": True, "updated": updates})


@login_required
def director_ai_sessions_list(request):
    """GET — foydalanuvchining barcha AI sessiyalari ro'yxati."""
    if request.method != "GET":
        return JsonResponse({"error": "Faqat GET"}, status=405)

    center, role = _ai_auth_center(request)
    if not center:
        return _403()

    from core.models import DirectorAIChatSession, DirectorAIChatMessage

    sessions_qs = (
        DirectorAIChatSession.objects
        .filter(center=center, user=request.user)
        .order_by("-updated_at", "-id")[:30]
    )

    result = []
    for s in sessions_qs:
        last_msg = (
            DirectorAIChatMessage.objects
            .filter(session=s, role="user")
            .order_by("-created_at")
            .values("content", "created_at")
            .first()
        )
        result.append({
            "id": s.id,
            "title": s.title or "AI Chat",
            "created_at": s.created_at.isoformat(),
            "updated_at": s.updated_at.isoformat(),
            "preview": (last_msg["content"][:60] + "…") if last_msg and len(last_msg["content"]) > 60 else (last_msg["content"] if last_msg else ""),
        })
    return JsonResponse({"sessions": result})


@login_required
def director_ai_session_new(request):
    """POST — yangi bo'sh AI sessiya yaratish."""
    if request.method != "POST":
        return JsonResponse({"error": "Faqat POST"}, status=405)

    center, role = _ai_auth_center(request)
    if not center:
        return _403()

    from core.models import DirectorAIChatSession
    title = _AI_CHAT_TITLES.get(role, "AI chat")
    session = DirectorAIChatSession.objects.create(
        center=center, user=request.user, title=title
    )
    return JsonResponse({"ok": True, "session_id": session.id, "title": session.title})
