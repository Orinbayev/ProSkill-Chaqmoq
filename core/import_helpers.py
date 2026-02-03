# core/views.py helpers for Excel Import
import secrets
import string
import datetime
import logging
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
from openpyxl import load_workbook
from accounts.models import User
from chaqmoq.models import Ledger

logger = logging.getLogger(__name__)

def _normalize_header(x):
    if not x: return None
    return str(x).strip().lower().replace(" ", "").replace("_", "").replace("'", "").replace("‘", "").replace("’", "")

def _pick_col(headers_map, *aliases):
    for a in aliases:
        if a in headers_map:
            return headers_map[a]
    return None

def _cell_to_str(v):
    if v is None: return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v)).strip()
    return str(v).strip()

def _normalize_phone(p):
    if not p: return ""
    digits = "".join(filter(str.isdigit, str(p)))
    if len(digits) == 9: return "998" + digits
    if len(digits) == 12 and digits.startswith("998"): return digits
    return digits

def _clean_for_login(text):
    if not text: return ""
    # simple slugify-like
    import re
    text = text.lower()
    text = re.sub(r'[^a-z0-9]', '', text)
    return text

def _normalize_gender(val):
    if not val: return None
    v = str(val).lower().strip()
    if v in ("male", "erkak", "m", "o'g'il", "o`g`il"): return "male"
    if v in ("female", "ayol", "f", "qiz"): return "female"
    return None

def _gen_default_password():
    return "chaqmoq123"

def _process_user_import(request, role="student"):
    """
    Generic User Import logic. 
    Improved with better error handling and performance.
    """
    # Normalize role
    role_map = {"students": "student", "teachers": "teacher", "managers": "manager"}
    role = role_map.get(role, role)
    
    # Determine redirect URL
    redirect_map = {
        "student": "core:stat_students",
        "teacher": "core:teacher_list",
        "manager": "core:stat_managers"
    }
    success_url = redirect_map.get(role, "core:home")

    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    if not center:
        messages.error(request, "Aktiv markaz aniqlanmadi.")
        return redirect(success_url)

    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Excel fayl tanlanmadi.")
        return redirect(success_url)

    try:
        wb = load_workbook(filename=f, data_only=True)
        ws = wb.active
        # Look for a sheet with headers if active one looks empty
        if ws.max_row < 2:
            for sn in wb.sheetnames:
                if wb[sn].max_row >= 2:
                    ws = wb[sn]
                    break
    except Exception as e:
        logger.error(f"Excel load error: {e}")
        messages.error(request, f"Excel faylni o'qishda xatolik: {e}")
        return redirect(success_url)

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        messages.error(request, "Excel faylda ma'lumot topilmadi.")
        return redirect(success_url)

    # Header detection (look in first 10 rows)
    headers_map = {}
    header_idx = -1
    for i in range(min(10, len(all_rows))):
        temp_map = {}
        for idx, val in enumerate(all_rows[i]):
            key = _normalize_header(val)
            if key: temp_map[key] = idx
        if _pick_col(temp_map, "ism", "name", "fish") is not None:
            header_idx = i
            headers_map = temp_map
            break

    if header_idx == -1:
        messages.error(request, "Excel ustunlari aniqlanmadi (Ism, Familya ustunlari bo'lishi shart).")
        return redirect(success_url)

    # Column mapping
    col_ism = _pick_col(headers_map, "ism", "firstname", "name")
    col_fam = _pick_col(headers_map, "familya", "familiya", "lastname")
    col_fish = _pick_col(headers_map, "fish", "f.i.sh", "fullname")
    col_otch = _pick_col(headers_map, "otchestvo", "middlename")
    col_tel1 = _pick_col(headers_map, "telefon", "telefon1", "phone", "tel")
    col_tel2 = _pick_col(headers_map, "telefon2", "phone2", "tel2")
    col_birth = _pick_col(headers_map, "tugilgankun", "birthdate", "birthdate", "tug'ilgansana", "sana")
    col_gender = _pick_col(headers_map, "jinsi", "gender")
    col_email = _pick_col(headers_map, "email", "login")
    col_pass = _pick_col(headers_map, "parol", "password")
    col_chaqmoq = _pick_col(headers_map, "chaqmoq", "coins", "ball")

    created, skipped, errors = 0, 0, 0
    problems = []
    
    # Pre-fetch existing emails to avoid heavy DB hits inside loop
    all_known_emails = set(User.objects.values_list("email", flat=True))

    for r_i, r in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not r or all(v is None for v in r): continue
        
        try:
            ism = _cell_to_str(r[col_ism]) if (col_ism is not None and col_ism < len(r)) else ""
            fam = _cell_to_str(r[col_fam]) if (col_fam is not None and col_fam < len(r)) else ""
            
            # Handle F.I.Sh column
            if col_fish is not None and not ism:
                fish = _cell_to_str(r[col_fish])
                if fish:
                    parts = fish.split()
                    ism = parts[0] if len(parts) > 0 else ""
                    fam = parts[1] if len(parts) > 1 else ""

            if not ism:
                skipped += 1
                continue

            # Email processing
            email_val = _cell_to_str(r[col_email]).lower() if (col_email is not None and col_email < len(r)) else ""
            
            # Ensure unique email
            if not email_val or email_val in all_known_emails:
                prefix = _clean_for_login(ism) or "user"
                suffix = secrets.randbelow(9000) + 1000
                cand = f"{prefix}{suffix}@chaqmoq.uz"
                while cand in all_known_emails:
                    suffix = secrets.randbelow(9000) + 1000
                    cand = f"{prefix}{suffix}@chaqmoq.uz"
                email_val = cand

            tel1 = _normalize_phone(_cell_to_str(r[col_tel1])) if (col_tel1 is not None and col_tel1 < len(r)) else ""
            
            with transaction.atomic():
                u = User.objects.create(
                    email=email_val,
                    role=role,
                    center=center,
                    ism=ism,
                    familya=fam,
                    first_name=ism,
                    last_name=fam,
                    telefon1=tel1
                )
                
                # Optional fields
                if col_otch is not None and col_otch < len(r):
                    u.otchestvo = _cell_to_str(r[col_otch])
                
                if col_gender is not None and col_gender < len(r):
                    u.gender = _normalize_gender(r[col_gender])
                
                if col_birth is not None and col_birth < len(r):
                    bv = r[col_birth]
                    if isinstance(bv, (datetime.date, datetime.datetime)):
                        u.birth_date = bv
                    elif isinstance(bv, str):
                        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y.%m.%d"):
                            try:
                                u.birth_date = datetime.datetime.strptime(bv.strip(), fmt).date()
                                break
                            except: continue

                # Password
                pv = _cell_to_str(r[col_pass]) if (col_pass is not None and col_pass < len(r)) else ""
                u.set_password(pv if pv and len(pv) >= 4 else _gen_default_password())
                u.save()
                
                # Chaqmoq / Coins
                if role == "student" and col_chaqmoq is not None and col_chaqmoq < len(r):
                    try:
                        ball = int(float(_cell_to_str(r[col_chaqmoq])))
                        if ball != 0:
                            Ledger.objects.create(
                                student=u,
                                ball=ball,
                                beruvchi=request.user,
                                rule_nom="Excel Import"
                            )
                    except: pass
                
                all_known_emails.add(email_val)
                created += 1

        except Exception as e:
            errors += 1
            problems.append(f"{r_i}-qator: {str(e)}")
            logger.error(f"Import row {r_i} error: {e}")

    # Final result message
    messages.success(request, f"Import tayyor: {created} ta yangi foydalanuvchi qo'shildi.")
    if skipped: messages.info(request, f"{skipped} ta qatorda ism yo'qligi uchun o'tkazib yuborildi.")
    if errors: messages.warning(request, f"{errors} ta qatorda texnik xatolik bo'ldi.")
    
    if problems:
        logger.warning(f"Import problems: {problems[:5]}")
        # messages.error(request, f"Xatolar: {problems[0]}") 

    return redirect(success_url)
