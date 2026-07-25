# core/views.py
from __future__ import annotations

import re
import secrets
import string
import datetime
import json
import logging
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, SetPasswordForm
from django.core.exceptions import FieldError, PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.utils.timezone import localdate
from django.views.decorators.http import require_GET, require_POST

from accounts.models import Branch, Center, User
from accounts.forms import TeacherForm, ParentForm
from chaqmoq.models import Ledger, Rule
from education.models import Group, Enrollment, TuitionMonth, Category
from store.models import Product, PurchaseRequest, Sale

from .forms import ProfileForm
from .services.user_import_service import (
    normalize_header as _normalize_header,
    pick_col as _pick_col,
    cell_to_str as _cell_to_str,
    clean_for_login as _clean_for_login,
    normalize_phone as _normalize_phone,
    gen_default_password as _gen_default_password,
    gen_unique_email as _gen_unique_gmail_like_email,
    normalize_gender as _normalize_gender,
)

U = get_user_model()
logger = logging.getLogger(__name__)


# =============================================================================
# TENANT HELPERS
# =============================================================================

def _get_center(request):
    """
    TenantMiddleware request.center qo'yadi.
    fallback: request.user.center
    """
    return getattr(request, "center", None) or getattr(getattr(request, "user", None), "center", None)


def _has_field(model, name: str) -> bool:
    try:
        return any(f.name == name for f in model._meta.fields)
    except Exception:
        return False


def _try_center_filter(qs, center, lookups: list[str]):
    """
    Turli modellarda center bog‘lanishi har xil bo‘lishi mumkin.
    Lookuplarni ketma-ket sinab ko‘ramiz.
    Birortasi ishlasa -> filtered qs.
    Hech biri ishlamasa -> qs.none() (leak bo‘lmasin).
    """
    if not center:
        return qs.none()
    for lk in lookups:
        try:
            filtered_qs = qs.filter(**{lk: center})
            if _has_field(qs.model, "is_deleted"):
                filtered_qs = filtered_qs.filter(is_deleted=False)
            return filtered_qs
        except FieldError:
            continue
    return qs.none()


def _filter_center(qs, center):
    """
    Har xil modellarda center bo‘lmasligi mumkin.
    - center field bo‘lsa -> center=center
    - student field bo‘lsa -> student__center=center
    - group field bo‘lsa -> group__center=center
    """
    if center is None:
        return qs.none()

    m = qs.model
    filtered_qs = qs.none()
    if _has_field(m, "center"):
        filtered_qs = qs.filter(center=center)
    elif _has_field(m, "student"):
        filtered_qs = qs.filter(student__center=center)
    elif _has_field(m, "group"):
        filtered_qs = qs.filter(group__center=center)

    if _has_field(m, "is_deleted"):
        filtered_qs = filtered_qs.filter(is_deleted=False)

    return filtered_qs


def _staff_only(request) -> bool:
    u = request.user
    return bool(u and (u.is_superuser or getattr(u, "role", None) in ("manager", "director")))


@login_required
def branch_list_api(request):
    """GET: center uchun barcha filiallar JSON."""
    center = getattr(request, "center", None) or getattr(request.user, "center", None)
    if not center:
        return JsonResponse({"ok": False, "error": "Markaz topilmadi"}, status=403)
    if not (request.user.role in ("director", "manager") or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q"}, status=403)

    branches = list(
        Branch.objects.filter(center=center)
        .values("id", "name", "address", "phone", "is_active", "order")
    )
    for branch in branches:
        branch["groups_count"] = Group.objects.filter(
            branch_id=branch["id"],
            is_archived=False,
        ).count()
        branch["students_count"] = (
            Enrollment.objects.filter(
                group__branch_id=branch["id"],
                is_active=True,
                is_deleted=False,
            )
            .values("student")
            .distinct()
            .count()
        )

    return JsonResponse({"ok": True, "branches": branches})


@login_required
def branch_create(request):
    """POST: yangi filial yaratish."""
    center = getattr(request, "center", None) or getattr(request.user, "center", None)
    if not center:
        return JsonResponse({"ok": False, "error": "Markaz topilmadi"}, status=403)
    if not (request.user.role == "director" or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "Faqat direktor"}, status=403)
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON xato"}, status=400)

    name = (data.get("name") or "").strip()
    if not name:
        return JsonResponse({"ok": False, "error": "Nom kiritilmagan"}, status=400)

    if Branch.objects.filter(center=center, name=name).exists():
        return JsonResponse({"ok": False, "error": "Bu nomda filial allaqachon mavjud"}, status=400)

    branch = Branch.objects.create(
        center=center,
        name=name,
        address=(data.get("address") or "").strip(),
        phone=(data.get("phone") or "").strip(),
        is_active=True,
    )
    return JsonResponse({
        "ok": True,
        "branch": {
            "id": branch.id,
            "name": branch.name,
            "address": branch.address,
            "phone": branch.phone,
            "is_active": branch.is_active,
            "groups_count": 0,
            "students_count": 0,
        },
    })


@login_required
def branch_update(request, pk):
    """POST: filialni tahrirlash."""
    center = getattr(request, "center", None) or getattr(request.user, "center", None)
    if not center:
        return JsonResponse({"ok": False, "error": "Markaz topilmadi"}, status=403)
    if not (request.user.role == "director" or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "Faqat direktor"}, status=403)
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)

    branch = get_object_or_404(Branch, pk=pk, center=center)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON xato"}, status=400)

    name = (data.get("name") or "").strip()
    if not name:
        return JsonResponse({"ok": False, "error": "Nom bo'sh bo'lmasin"}, status=400)

    if Branch.objects.filter(center=center, name=name).exclude(pk=pk).exists():
        return JsonResponse({"ok": False, "error": "Bu nomda filial allaqachon mavjud"}, status=400)

    is_active = data.get("is_active", True)
    if isinstance(is_active, str):
        is_active = is_active.strip().lower() in {"1", "true", "yes", "on"}
    else:
        is_active = bool(is_active)

    branch.name = name
    branch.address = (data.get("address") or "").strip()
    branch.phone = (data.get("phone") or "").strip()
    branch.is_active = is_active
    branch.save(update_fields=["name", "address", "phone", "is_active"])

    return JsonResponse({
        "ok": True,
        "branch": {
            "id": branch.id,
            "name": branch.name,
            "address": branch.address,
            "phone": branch.phone,
            "is_active": branch.is_active,
        },
    })


@login_required
def branch_delete(request, pk):
    """POST: filialni o'chirish."""
    center = getattr(request, "center", None) or getattr(request.user, "center", None)
    if not center:
        return JsonResponse({"ok": False}, status=403)
    if not (request.user.role == "director" or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "Faqat direktor"}, status=403)
    if request.method != "POST":
        return JsonResponse({"ok": False}, status=405)

    branch = get_object_or_404(Branch, pk=pk, center=center)

    Group.objects.filter(branch=branch).update(branch=None)
    branch.delete()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def toggle_manager_trash_access(request):
    if not (request.user.is_superuser or request.user.role == 'director'):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)
        raise PermissionDenied
    
    center = _get_center(request)
    if not center:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
        return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
        
    center.manager_can_access_trash = not center.manager_can_access_trash
    center.save()
    
    status_text = "yoqildi" if center.manager_can_access_trash else "o'chirildi"
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            "ok": True, 
            "status": center.manager_can_access_trash,
            "message": f"Managerlar uchun trash ruxsati {status_text} ✅"
        })
    
    messages.success(request, f"Managerlar uchun trash ruxsati {status_text} ✅")
    return redirect(request.META.get('HTTP_REFERER', 'core:home'))

@login_required
@require_POST
def toggle_manager_can_add_student(request):
    if not (request.user.is_superuser or request.user.role == 'director'):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)
        raise PermissionDenied
    center = _get_center(request)
    if not center:
        return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
    center.manager_can_add_student = not center.manager_can_add_student
    center.save()
    status_text = "yoqildi" if center.manager_can_add_student else "o'chirildi"
    return JsonResponse({"ok": True, "status": center.manager_can_add_student, "message": f"Managerlar o'quvchi qo'shishi {status_text} ✅"})

@login_required
@require_POST
def toggle_manager_can_remove_student(request):
    if not (request.user.is_superuser or request.user.role == 'director'):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)
        raise PermissionDenied
    center = _get_center(request)
    if not center:
        return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
    center.manager_can_remove_student = not center.manager_can_remove_student
    center.save()
    status_text = "yoqildi" if center.manager_can_remove_student else "o'chirildi"
    return JsonResponse({"ok": True, "status": center.manager_can_remove_student, "message": f"Managerlar o'quvchi o'chirishi {status_text} ✅"})

@login_required
@require_POST
def toggle_teacher_can_add_student(request):
    if not (request.user.is_superuser or request.user.role == 'director'):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)
        raise PermissionDenied
    center = _get_center(request)
    if not center:
        return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
    center.teacher_can_add_student = not center.teacher_can_add_student
    center.save()
    status_text = "yoqildi" if center.teacher_can_add_student else "o'chirildi"
    return JsonResponse({"ok": True, "status": center.teacher_can_add_student, "message": f"O'qituvchilar o'quvchi qo'shishi {status_text} ✅"})

@login_required
@require_POST
def toggle_teacher_can_remove_student(request):
    if not (request.user.is_superuser or request.user.role == 'director'):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)
        raise PermissionDenied
    center = _get_center(request)
    if not center:
        return JsonResponse({"ok": False, "error": "Center not found"}, status=404)
    center.teacher_can_remove_student = not center.teacher_can_remove_student
    center.save()
    status_text = "yoqildi" if center.teacher_can_remove_student else "o'chirildi"
    return JsonResponse({"ok": True, "status": center.teacher_can_remove_student, "message": f"O'qituvchilar o'quvchi o'chirishi {status_text} ✅"})


@login_required
def group_permissions_settings(request):
    """
    Direktor paneli: Manager va Teacher uchun
    O'quvchi qo'shish/o'chirish ruxsatlarini boshqarish sahifasi.
    """
    if not (request.user.is_superuser or request.user.role == 'director'):
        raise PermissionDenied
    center = _get_center(request)
    if not center:
        messages.error(request, "Center topilmadi.")
        return redirect('core:home')
    return render(request, 'core/group_permissions_settings.html', {'center': center})



def _assert_same_center(obj, center):
    """
    Obj'ekt boshqa centerga tegishli bo'lsa 404 qaytaramiz (security).
    """
    if not center:
        return
    if hasattr(obj, "center_id"):
        if obj.center_id != center.id:
            raise PermissionDenied("Boshqa center ma'lumotiga ruxsat yo‘q.")


# =============================================================================
# DASHBOARD STATS
# =============================================================================

def _build_stats(center):
    """
    Tenant scoped stats.
    center None bo’lsa -> hammasi 0.

    PERF v2: 7 alohida query → 3 query ga tushirildi:
      1. User role counts — 1 annotated query (aggregate bilan)
      2. Products + PurchaseRequests + Sales + Ledger — 4 query (unavoidable, different tables)
      Eski: 7 query, Yangi: ~4-5 query
    """
    if not center:
        return {
            "managers": 0,
            "teachers": 0,
            "students": 0,
            "products": 0,
            "pending_requests": 0,
            "total_chaqmoq": 0,
            "sales_today": 0,
        }

    # ✅ PERF: 3 alohida .count() → 1 aggregate query
    user_agg = U.objects.filter(center=center).aggregate(
        managers=Count("id", filter=Q(role="manager")),
        teachers=Count("id", filter=Q(role="teacher")),
        students=Count("id", filter=Q(role="student", is_archived=False)),
    )

    products_qs = _try_center_filter(Product.objects.all(), center, ["center"])
    pr_qs = _try_center_filter(PurchaseRequest.objects.all(), center, ["center", "student__center", "manager__center"])
    ledger_qs = _try_center_filter(Ledger.objects.all(), center, ["center", "student__center", "group__center"])
    sales_qs = _try_center_filter(Sale.objects.all(), center, ["center", "student__center", "manager__center"])

    if _has_field(Sale, "sana"):
        sales_today_qs = sales_qs.filter(sana__date=localdate())
    elif _has_field(Sale, "created_at"):
        sales_today_qs = sales_qs.filter(created_at__date=localdate())
    else:
        sales_today_qs = sales_qs.none()

    pending_status = getattr(PurchaseRequest, "PENDING", "pending")

    # ✅ PERF: products + pending_requests bitta query emas, lekin har biri tez
    # ledger aggregate va sales count — 2 qo’shimcha query
    return {
        "managers": user_agg["managers"] or 0,
        "teachers": user_agg["teachers"] or 0,
        "students": user_agg["students"] or 0,
        "products": products_qs.count(),
        "pending_requests": pr_qs.filter(status=pending_status).count(),
        "total_chaqmoq": ledger_qs.aggregate(s=Sum("ball"))["s"] or 0,
        "sales_today": sales_today_qs.count(),
    }


# =============================================================================
# HOME / DASHBOARDS
# =============================================================================

@login_required
def home(request):
    """
    Skeleton-first dashboard render.

    PERF v3:
      Eski: manager/teacher/student uchun _build_stats() + low_activity +
      balance + last_actions SYNC ishlatilardi (5–12 query, ~300–800ms).
      Yangi: home faqat template skeleton'ni render qiladi (0 DB query).
      Real ma'lumotlar browserda AJAX orqali keyin yuklanadi:
        - manager  → /<slug>/api/dashboard/quick-stats/  +  /api/dashboard/low-activity/
        - student  → /<slug>/api/dashboard/student-init/
        - teacher  → stats template'da ishlatilmaydi, skeleton kifoya.
    Redirectlar (director/parent/superuser) o'zgarmaydi.
    """
    u = request.user
    role = getattr(u, "role", None)
    if (not role) and u.is_superuser:
        role = "director"

    # ✅ SUPERADMIN: root URL '/' her doim global platform panelga yo'naltiradi.
    # Session'da active_center_id bo'lsa ham, '/' superadmin uchun tenant dashboard emas.
    if u.is_superuser:
        return redirect("platform_global:superadmin_dashboard")

    if role in ("director", "manager"):
        return redirect("core:director_boshqaruv")

    if role == "parent":
        return redirect("core:dashboard_parent")

    # Lazy center — faqat template {{ center }} ishlatgani uchun kerak.
    center = _get_center(request)

    if role == "teacher":
        return render(request, "core/dashboard_teacher.html", {"center": center})

    if role == "student":
        return render(request, "core/dashboard_student.html", {"center": center})

    return redirect("/admin/accounts/user/")

# =============================================================================
# TEACHERS (tenant scoped)
# =============================================================================

@login_required
def teacher_list(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    qs = User.objects.filter(role="teacher")
    qs = _try_center_filter(qs, center, ["center"])  # ✅ tenant

    # Search filter
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(Q(ism__icontains=q) | Q(familya__icontains=q) | Q(email__icontains=q) | Q(telefon1__icontains=q))

    # Sizda teacher->group relation nomi har xil bo'lishi mumkin.
    # Oldingi annotate(Count('group')) noto'g'ri bo'lish ehtimoli katta.
    # Shuning uchun xavfsiz yo'l: groups oqituvchi bo'yicha sanaladi.
    if _has_field(Group, "oqituvchi"):
        teachers = qs.annotate(
            group_count=Count("group", filter=Q(group__is_deleted=False, group__is_archived=False), distinct=True),
            student_count=Count(
                "group__enrollments__student",
                filter=Q(
                    group__enrollments__is_active=True,
                    group__enrollments__is_deleted=False,
                    group__enrollments__student__is_archived=False,
                    group__is_deleted=False,
                    group__is_archived=False
                ),
                distinct=True
            )
        )
    else:
        teachers = qs
    teachers_count = qs.count()

    return render(request, "core/teacher_list.html", {"teachers": teachers, "teachers_count": teachers_count})


@login_required
def teacher_detail(request, pk):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)

    teacher = get_object_or_404(User, pk=pk, role="teacher")
    _assert_same_center(teacher, center)

    # groups tenant + teacher filter
    groups_qs = Group.objects.all()
    groups_qs = _try_center_filter(groups_qs, center, ["center"])
    groups = groups_qs.filter(oqituvchi=teacher) if _has_field(Group, "oqituvchi") else groups_qs.none()

    # ✅ Handle Password Reset (Admin/Staff only)
    if request.method == "POST" and "new_password" in request.POST:
        if not _staff_only(request):
            raise PermissionDenied
        new_pass = request.POST.get("new_password")
        if new_pass:
            teacher.set_password(new_pass)
            teacher.save()
            messages.success(request, f"{teacher.get_full_name()} paroli muvaffaqiyatli o'zgartirildi.")
            return redirect("core:teacher_detail", pk=teacher.pk)

    return render(request, "core/teacher_detail.html", {"teacher": teacher, "groups": groups})


@login_required
def teacher_edit(request, pk):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)

    teacher = get_object_or_404(User, pk=pk)
    _assert_same_center(teacher, center)

    if request.method == "POST":
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            teacher = form.save(commit=False)

            teacher.otchestvo = form.cleaned_data.get("otchestvo")

            tel1 = form.cleaned_data.get("telefon1", "") or ""
            tel2 = form.cleaned_data.get("telefon2", "") or ""

            if tel1:
                teacher.telefon1 = "+998" + tel1.replace("+998", "").replace(" ", "").replace("-", "")
            if tel2:
                teacher.telefon2 = "+998" + tel2.replace("+998", "").replace(" ", "").replace("-", "")

            # ✅ Password Update (Manual)
            new_pass = request.POST.get("new_password")
            if new_pass and new_pass.strip():
                teacher.set_password(new_pass.strip())
                messages.success(request, "Parol yangilandi.")

            teacher.save()

            yangi_foiz = getattr(teacher, "oqituvchi_foizi", None)

            # ✅ tenant: faqat shu teacher ning o'z group/enrollmentlari
            if yangi_foiz is not None:
                if _has_field(Group, "oqituvchi") and _has_field(Group, "oqituvchi_foiz"):
                    Group.objects.filter(oqituvchi=teacher).update(oqituvchi_foiz=yangi_foiz)

                # Enrollment modelida field mavjud bo'lsa update
                if _has_field(Enrollment, "oqituvchi_foiz"):
                    Enrollment.objects.filter(group__oqituvchi=teacher).update(oqituvchi_foiz=yangi_foiz)

            return redirect("core:stat_teachers")
    else:
        form = TeacherForm(instance=teacher)

    return render(request, "core/teacher_edit.html", {"form": form, "teacher": teacher})


@login_required
def teacher_edit_ajax(request, pk):
    from django.utils.dateparse import parse_date as _parse_date
    from accounts.forms import PasswordUpdateForm

    if not _staff_only(request):
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q."}, status=403)

    center = _get_center(request)
    teacher = get_object_or_404(User, pk=pk, role="teacher")
    _assert_same_center(teacher, center)

    # ── GET: Return full teacher details + their groups ──
    if request.method == "GET":
        groups = Group.objects.filter(oqituvchi=teacher, is_archived=False).order_by("nom")
        group_data = [
            {
                "id": g.pk,
                "nom": g.nom,
                "kurs_narxi": g.kurs_narxi,
                "oqituvchi_foiz": g.oqituvchi_foiz if g.oqituvchi_foiz is not None else teacher.oqituvchi_foizi,
                "student_count": g.enrollments.filter(is_active=True, is_deleted=False).count(),
            }
            for g in groups
        ]
        return JsonResponse({
            "ok": True,
            "data": {
                "id": teacher.pk,
                "ism": teacher.ism,
                "familya": teacher.familya,
                "otchestvo": teacher.otchestvo or "",
                "email": teacher.email,
                "telefon1": teacher.telefon1 or "",
                "telefon2": teacher.telefon2 or "",
                "birth_date": teacher.birth_date.strftime("%Y-%m-%d") if teacher.birth_date else "",
                "gender": teacher.gender or "",
                "passport_id": teacher.passport_id or "",
                "jshr": teacher.jshr or "",
                "address": teacher.address or "",
                "telegram_username": teacher.telegram_username or "",
                "instagram_username": teacher.instagram_username or "",
                "avatar_url": teacher.avatar.url if teacher.avatar else None,
                "full_name": teacher.get_full_name(),
                "oqituvchi_foizi": teacher.oqituvchi_foizi or 0,
                "groups": group_data,
            }
        })

    action = request.POST.get("action")

    # ── POST: Update profile ──
    if action == "update_profile":
        teacher.ism = (request.POST.get("ism") or "").strip() or teacher.ism
        teacher.familya = (request.POST.get("familya") or "").strip() or teacher.familya
        teacher.otchestvo = (request.POST.get("otchestvo") or "").strip()

        tel1 = (request.POST.get("telefon1") or "").strip()
        tel2 = (request.POST.get("telefon2") or "").strip()

        if tel1:
            tel1_digits = "".join(filter(str.isdigit, tel1))
            if tel1_digits.startswith("998") and len(tel1_digits) == 12:
                teacher.telefon1 = "+" + tel1_digits
            elif len(tel1_digits) == 9:
                teacher.telefon1 = "+998" + tel1_digits
            else:
                teacher.telefon1 = tel1
        else:
            teacher.telefon1 = ""

        if tel2:
            tel2_digits = "".join(filter(str.isdigit, tel2))
            if tel2_digits.startswith("998") and len(tel2_digits) == 12:
                teacher.telefon2 = "+" + tel2_digits
            elif len(tel2_digits) == 9:
                teacher.telefon2 = "+998" + tel2_digits
            else:
                teacher.telefon2 = tel2
        else:
            teacher.telefon2 = ""

        teacher.passport_id = (request.POST.get("passport_id") or "").strip().upper()
        teacher.jshr = (request.POST.get("jshr") or "").strip()
        teacher.address = (request.POST.get("address") or "").strip()
        teacher.telegram_username = (request.POST.get("telegram_username") or "").strip()
        teacher.instagram_username = (request.POST.get("instagram_username") or "").strip()

        gender = (request.POST.get("gender") or "").strip()
        teacher.gender = gender or None

        birth_raw = (request.POST.get("birth_date") or "").strip()
        if birth_raw:
            teacher.birth_date = _parse_date(birth_raw)
        else:
            teacher.birth_date = None

        raw_foiz = (request.POST.get("oqituvchi_foizi") or "").strip()
        old_foiz = teacher.oqituvchi_foizi
        yangi_foiz = old_foiz
        if raw_foiz.isdigit():
            yangi_foiz = int(raw_foiz)
            teacher.oqituvchi_foizi = yangi_foiz

        if "avatar" in request.FILES:
            teacher.avatar = request.FILES["avatar"]

        teacher.save()

        # Update teacher percentage if changed across their groups & active enrollments
        if yangi_foiz != old_foiz:
            if _has_field(Group, "oqituvchi") and _has_field(Group, "oqituvchi_foiz"):
                Group.objects.filter(oqituvchi=teacher).update(oqituvchi_foiz=yangi_foiz)
            if _has_field(Enrollment, "oqituvchi_foiz"):
                Enrollment.objects.filter(group__oqituvchi=teacher).update(oqituvchi_foiz=yangi_foiz)

        return JsonResponse({
            "ok": True,
            "message": "Profil yangilandi ✅",
            "full_name": teacher.get_full_name(),
            "avatar_url": teacher.avatar.url if teacher.avatar else None,
            "oqituvchi_foizi": teacher.oqituvchi_foizi,
        })

    # ── POST: Update password ──
    if action == "update_password":
        form = PasswordUpdateForm(request.POST)
        if form.is_valid():
            teacher.set_password(form.cleaned_data["new_password"])
            teacher.save()
            return JsonResponse({"ok": True, "message": "Parol yangilandi 🔒"})
        return JsonResponse({"ok": False, "errors": form.errors}, status=400)

    return JsonResponse({"ok": False, "error": "Noto'g'ri so'rov."}, status=400)





@login_required
def teacher_delete(request, pk):
    if request.user.role not in ("manager", "director") and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo‘q.")
        return redirect("core:teacher_list")

    center = _get_center(request)

    teacher = get_object_or_404(User, pk=pk, role="teacher")
    _assert_same_center(teacher, center)

    if request.method == "POST":
        teacher.delete(deleted_by=request.user)
        messages.success(request, "O‘qituvchi o‘chirildi ✅")
        return redirect("core:teacher_list")

    return redirect("core:teacher_list")


# =============================================================================
# STUDENT last actions (tenant safe)
# =============================================================================

def _safe_user_label(user_obj) -> str:
    if not user_obj:
        return "Tizim"
    ism = getattr(user_obj, "ism", "") or ""
    fam = getattr(user_obj, "familya", "") or ""
    full = (ism + " " + fam).strip()
    return full or str(user_obj)


def _ledger_title(ledger_obj) -> str:
    rule = getattr(ledger_obj, "rule", None)
    if rule:
        for attr in ("nom", "name", "title"):
            v = getattr(rule, attr, None)
            if v:
                return str(v)
    return "Harakat"


def _student_last_actions(student_id: int, center=None):
    """
    Oxirgi 10 ta Ledger yozuvini dict ko‘rinishida qaytaradi.
    ✅ Tenant safe: center bo'lsa, ledger'ni center scope qilamiz.
    """
    field_names = {f.name for f in Ledger._meta.fields}

    sr = []
    for f in ("rule", "group", "beruvchi"):
        if f in field_names:
            sr.append(f)

    qs = Ledger.objects.filter(student_id=student_id)

    # tenant scope (Ledger'da center bo'lmasa student__center orqali)
    qs = _try_center_filter(qs, center, ["center", "student__center", "group__center"]) if center else qs

    # ordering field: created_at yoki sana
    if _has_field(Ledger, "created_at"):
        qs = qs.order_by("-created_at")
    elif _has_field(Ledger, "sana"):
        qs = qs.order_by("-sana")
    else:
        qs = qs.order_by("-id")

    qs = qs.select_related(*sr)[:10]

    out = []
    for x in qs:
        delta = int(getattr(x, "ball", 0) or 0)
        sign = "+" if delta >= 0 else "-"

        actor_obj = getattr(x, "beruvchi", None)

        grp = ""
        if getattr(x, "group", None):
            grp = getattr(x.group, "nom", "") or ""

        created_at = None
        if hasattr(x, "created_at") and getattr(x, "created_at"):
            created_at = x.created_at
        elif hasattr(x, "sana") and getattr(x, "sana"):
            created_at = x.sana
        else:
            created_at = timezone.now()

        out.append({
            "title": _ledger_title(x),
            "who": _safe_user_label(actor_obj),
            "created_at": created_at,
            "delta": delta,
            "sign": sign,
            "abs_delta": abs(delta),
            "group": grp,
        })
    return out


# =============================================================================
# PARENTS MANAGEMENT
# =============================================================================

@login_required
def stat_parents(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    q = request.GET.get("q", "").strip()
    center = _get_center(request)
    rows = U.objects.filter(role="parent")
    rows = _try_center_filter(rows, center, ["center"]) if center else rows.none()

    if q:
        rows = rows.filter(Q(ism__icontains=q) | Q(familya__icontains=q) | Q(email__icontains=q) | Q(telefon1__icontains=q))

    page_size = 9999
    paginator = Paginator(rows.order_by("ism", "familya", "id"), page_size)
    page_obj = paginator.get_page(1)
    start_index = page_obj.start_index() if page_obj.paginator.count else 0

    return render(request, "core/stats_users.html", {
        "title": "Ota-onalar",
        "page_obj": page_obj,
        "total_count": rows.count(),
        "user_kind": "parents",
        "no_pagination": True,
        "start_index": start_index,
        "page_size": page_size,
        "center": center,  # ✅ Add center to context
    })


@login_required
@require_POST
def update_center_donation_settings(request):
    if not _staff_only(request):
        raise PermissionDenied
    
    center = _get_center(request)
    if not center:
        messages.error(request, "Markaz topilmadi.")
        return redirect("core:stat_parents")

    center.donation_enabled = request.POST.get("donation_enabled") == "on"
    center.donation_card_number = request.POST.get("donation_card_number", "").strip()
    center.donation_card_holder = request.POST.get("donation_card_holder", "").strip()
    
    if "donation_qr_image" in request.FILES:
        center.donation_qr_image = request.FILES["donation_qr_image"]
    
    center.save()
    messages.success(request, "Donat sozlamalari yangilandi.")
    return redirect("core:stat_parents")


@login_required
def parent_add(request):
    if not _staff_only(request):
        raise PermissionDenied
    
    if request.method == "POST":
        form = ParentForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Ota-ona muvaffaqiyatli qo'shildi ✅")
            return redirect("core:stat_parents")
    else:
        form = ParentForm(request=request)
    
    return render(request, "core/parent_form.html", {"form": form, "title": "Ota-ona qo'shish"})


@login_required
def parent_edit(request, pk):
    if not _staff_only(request):
        raise PermissionDenied
    
    center = _get_center(request)
    parent = get_object_or_404(User, pk=pk, role="parent")
    _assert_same_center(parent, center)
    
    if request.method == "POST":
        form = ParentForm(request.POST, instance=parent, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Ota-ona ma'lumotlari yangilandi ✅")
            return redirect("core:stat_parents")
    else:
        form = ParentForm(instance=parent, request=request)
    
    return render(request, "core/parent_form.html", {"form": form, "title": "Ota-onani tahrirlash", "is_edit": True})


@login_required
def parent_delete(request, pk):
    if not _staff_only(request):
        raise PermissionDenied
        
    center = _get_center(request)
    parent = get_object_or_404(User, pk=pk, role="parent")
    _assert_same_center(parent, center)
    
    if request.method == "POST":
        parent.delete(deleted_by=request.user)
        messages.success(request, "Ota-ona o'chirildi ✅")
        return redirect("core:stat_parents")
    return redirect("core:stat_parents")


@login_required
def dashboard_parent(request):
    if getattr(request.user, "role", None) != "parent":
        return redirect("core:home")
    
    center = getattr(request.user, "center", None) or _get_center(request)
    children = request.user.children.all()

    # Calculate stats for each child
    for child in children:
        # Balance
        child.calculated_balance = Ledger.objects.filter(
            student=child
        ).filter(
            Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True)
        ).aggregate(Sum('ball'))['ball__sum'] or 0

        # Rank
        child.calculated_rank = Ledger.objects.filter(
            Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True),
            student__role='student'
        ).values('student').annotate(
            total_points=Sum('ball')
        ).filter(total_points__gt=child.calculated_balance).count() + 1
    
    # Parent can also search other students
    q = request.GET.get("q", "").strip()
    search_results = []
    if q:
        search_results = User.objects.filter(role="student", center=center, is_archived=False).filter(
            Q(ism__icontains=q) | Q(familya__icontains=q)
        ).distinct()[:10]

        for s in search_results:
            # Check if already child
            s.is_my_child = s in children

            # Calculate Stats (Same logic)
            s_balance = Ledger.objects.filter(
                student=s
            ).filter(
                Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True)
            ).aggregate(Sum('ball'))['ball__sum'] or 0
            
            s.calculated_balance = s_balance

            s.calculated_rank = Ledger.objects.filter(
                Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True),
                student__role='student'
            ).values('student').annotate(
                total_points=Sum('ball')
            ).filter(total_points__gt=s_balance).count() + 1
        
    return render(request, "core/dashboard_parent.html", {
        "children": children,
        "search_results": search_results,
        "center": center,
        "title": "Ota-ona paneli",
    })

@login_required
@require_POST
def toggle_child(request, student_id):
    if getattr(request.user, "role", None) != "parent":
        raise PermissionDenied

    center = _get_center(request)
    student_qs = User.objects.filter(role="student")
    if center:
        student_qs = student_qs.filter(center=center)
    student = get_object_or_404(student_qs, pk=student_id)

    parent = request.user
    if center and parent.center_id and student.center_id != parent.center_id:
        raise PermissionDenied

    if student in parent.children.all():
        parent.children.remove(student)
        messages.warning(request, f"{student.get_full_name()} farzandlaringiz safidan olib tashlandi.")
    else:
        parent.children.add(student)
        messages.success(request, f"{student.get_full_name()} farzandlaringiz safiga qo‘shildi.")
        
    return redirect(request.META.get('HTTP_REFERER', 'core:dashboard_parent'))


# =============================================================================
# STATS USERS (tenant scoped)
# =============================================================================

@login_required
def stat_managers(request):
    # ✅ Manager bu bo‘limga kirmasin (URL orqali ham)
    if getattr(request.user, "role", None) == "manager":
        return render(request, "no_permission.html", {"message": "Sizda bu bo‘limga kirish huquqi yo‘q."})

    if not (request.user.is_superuser or getattr(request.user, "role", None) in ("director", "admin")):
        return render(request, "core/dashboard_guest.html")

    q = request.GET.get("q", "").strip()

    center = _get_center(request)
    rows = U.objects.filter(role="manager")
    rows = _try_center_filter(rows, center, ["center"]) if center else rows.none()

    if q:
        rows = rows.filter(Q(ism__icontains=q) | Q(familya__icontains=q) | Q(email__icontains=q))

    page_size = 9999
    paginator = Paginator(rows.order_by("ism", "familya", "id"), page_size)
    page_obj = paginator.get_page(1)
    start_index = page_obj.start_index() if page_obj.paginator.count else 0

    return render(request, "core/stats_users.html", {
        "title": "Managerlar",
        "page_obj": page_obj,
        "total_count": rows.count(),
        "page_size": page_size,
        "start_index": start_index,
        "no_pagination": True,
        "user_kind": "managers",
        "center": center,
    })


@login_required
def stat_teachers(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    q = request.GET.get("q", "").strip()
    page_size = 9999

    center = _get_center(request)

    rows = U.objects.filter(role="teacher")
    rows = _try_center_filter(rows, center, ["center"]) if center else rows.none()

    if q:
        rows = rows.filter(Q(ism__icontains=q) | Q(familya__icontains=q) | Q(email__icontains=q))

    paginator = Paginator(rows.order_by("id"), page_size)
    page_obj = paginator.get_page(1)
    start_index = page_obj.start_index() if page_obj.paginator.count else 0

    return render(request, "core/stats_users.html", {
        "title": "O‘qituvchilar",
        "page_obj": page_obj,
        "total_count": rows.count(),
        "start_index": start_index,
        "page_size": page_size,
        "no_pagination": True,
        "user_kind": "teachers",
        "center": center,
    })


@login_required
def stat_students(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    q = request.GET.get("q", "").strip()
    gender = request.GET.get("gender", "").strip()
    section_id = request.GET.get("section", "").strip()
    multi_group = request.GET.get("multi_group", "").strip()   # NEW: "yes" = ko'p guruhda
    status = request.GET.get("status", "active").strip() # default active

    page_size = request.GET.get("size", "10")
    try:
        page_size = int(page_size)
    except Exception:
        page_size = 10

    center = _get_center(request)
    
    # 1. Categories for filter
    categories = Category.objects.all()
    if center:
        categories = categories.filter(Q(center=center) | Q(center__isnull=True))

    rows = U.objects.filter(role="student")
    rows = _try_center_filter(rows, center, ["center"]) if center else rows.none()

    # ✅ Calculate Counts (for Tabs)
    active_count = rows.filter(is_archived=False).count()
    archived_count = rows.filter(is_archived=True).count()

    # Apply Status Filter
    if status == "archived":
        rows = rows.filter(is_archived=True)
    else:
        rows = rows.filter(is_archived=False)

    from django.db.models import Prefetch
    from education.models import Enrollment

    rows = rows.prefetch_related(
        Prefetch("enrollments", queryset=Enrollment.objects.filter(group__center=center, is_active=True).select_related("group"))
    ).order_by("-id")

    if q:
        rows = rows.filter(Q(ism__icontains=q) | Q(familya__icontains=q) | Q(email__icontains=q))

    if gender:
        rows = rows.filter(gender=gender)

    if section_id:
        if section_id.isdigit():
            rows = rows.filter(enrollments__group__category_obj__id=int(section_id)).distinct()
        else:
            rows = rows.filter(enrollments__group__category_obj__name=section_id).distinct()

    if multi_group == 'yes':
        from django.db.models import Count as AnnotCount
        rows = rows.annotate(
            active_group_count=AnnotCount(
                'enrollments',
                filter=Q(enrollments__is_active=True, enrollments__group__center=center),
                distinct=True
            )
        ).filter(active_group_count__gt=1)

    # Cap page_size to avoid very large pages
    page_size = min(page_size, 100)

    paginator = Paginator(rows, page_size)
    page_obj = paginator.get_page(request.GET.get("page"))
    start_index = page_obj.start_index() if page_obj.paginator.count else 0

    # Compute jami_chaqmoq for current page students in ONE query (not per-row annotation)
    from chaqmoq.models import Ledger
    _student_ids = [u.id for u in page_obj]
    if _student_ids and center:
        _ledger_sums = (
            Ledger.objects
            .filter(student_id__in=_student_ids)
            .filter(Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True))
            .values("student_id")
            .annotate(_total=Sum("ball"))
        )
        _sum_map = {row["student_id"]: row["_total"] for row in _ledger_sums}
    else:
        _sum_map = {}
    for u in page_obj:
        u.jami_chaqmoq = _sum_map.get(u.id, 0)

    context = {
        "title": "O’quvchilar",
        "page_obj": page_obj,
        # paginator.count cache qiladi — rows.count() alohida query kerak emas
        "total_count": paginator.count,
        "start_index": start_index,
        "page_size": page_size,
        "categories": categories,
        "user_kind": "students",
        "current_status": status,
        "active_count": active_count,
        "archived_count": archived_count,
        "center": center,
        "multi_group_filter": multi_group,
    }

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "core/includes/student_list_table.html", context)

    return render(request, "core/stats_users.html", context)


@login_required
@require_POST
def archive_student(request, pk):
    """
    Students Soft Delete (Archive)
    - Remove from groups (or mark inactive)
    - Set is_archived = True
    """
    if not _staff_only(request):
        raise PermissionDenied

    center = _get_center(request)
    student = get_object_or_404(U, pk=pk, role="student")
    _assert_same_center(student, center)

    with transaction.atomic():
        # 1. Archive User
        student.is_archived = True
        student.archived_at = timezone.now()
        student.save(update_fields=["is_archived", "archived_at"])
        
        # 2. Deactivate Enrollments (Remove from active groups lists)
        # Enrollment has is_active field? Let's check model. 
        # Yes, education/models.py Enrollment has is_active field.
        Enrollment.objects.filter(student=student, is_active=True).update(is_active=False)

    messages.warning(request, f"O‘quvchi {student.get_full_name()} arxivga olindi.")
    return redirect("core:stat_students")


@login_required
@require_POST
def restore_student(request, pk):
    """
    Restore Student from Archive
    """
    if not _staff_only(request):
        raise PermissionDenied

    center = _get_center(request)
    student = get_object_or_404(U, pk=pk, role="student")
    _assert_same_center(student, center)

    with transaction.atomic():
        student.is_archived = False
        student.save(update_fields=["is_archived"])
        
        # ✅ Restore enrollments (activate them back)
        Enrollment.objects.filter(student=student, center=center).update(is_active=True)

    messages.success(request, f"O‘quvchi {student.get_full_name()} muvaffaqiyatli tiklandi.")
    return redirect("core:stat_students")


def _student_parent_link_payload(student):
    from accounts.services.parent_telegram_link import parent_link_status

    status = parent_link_status(student)
    linked_at = status.get("linked_at")
    return {
        "is_linked": status["is_linked"],
        "telegram_id": status["telegram_id"],
        "telegram_username": status["telegram_username"],
        "linked_at": linked_at.isoformat() if linked_at else "",
        "linked_at_display": timezone.localtime(linked_at).strftime("%d.%m.%Y %H:%M") if linked_at else "",
        "parent_id": status["parent_id"],
        "parent_name": status["parent_name"],
    }


def _get_student_for_parent_link(request, pk):
    if not _staff_only(request):
        raise PermissionDenied
    center = _get_center(request)
    student = get_object_or_404(U, pk=pk, role="student")
    _assert_same_center(student, center)
    return student


@login_required
@require_GET
def student_parent_link_status(request, pk):
    student = _get_student_for_parent_link(request, pk)
    return JsonResponse({"ok": True, "status": _student_parent_link_payload(student)})


@login_required
@require_POST
def student_parent_link_create(request, pk):
    student = _get_student_for_parent_link(request, pk)
    status = _student_parent_link_payload(student)
    if status["is_linked"]:
        return JsonResponse({"ok": True, "already_linked": True, "status": status})

    from accounts.services.parent_telegram_link import create_parent_telegram_invite

    invite = create_parent_telegram_invite(student=student, created_by=request.user)
    return JsonResponse(
        {
            "ok": True,
            "already_linked": False,
            "link": invite.link,
            "telegram_share_url": invite.telegram_share_url,
            "expires_at": invite.expires_at.isoformat(),
            "expires_at_display": timezone.localtime(invite.expires_at).strftime("%d.%m.%Y %H:%M"),
            "status": status,
        }
    )


@login_required
@require_POST
def student_parent_link_reminder(request, pk):
    student = _get_student_for_parent_link(request, pk)
    status = _student_parent_link_payload(student)
    if status["is_linked"]:
        return JsonResponse({"ok": True, "sent": False, "already_linked": True, "status": status})

    from accounts.services.parent_telegram_link import create_parent_telegram_invite

    invite = create_parent_telegram_invite(student=student, created_by=request.user)
    return JsonResponse(
        {
            "ok": True,
            "sent": False,
            "message": "Parent Telegram ID topilmadi. Linkni Telegram orqali yuboring.",
            "link": invite.link,
            "telegram_share_url": invite.telegram_share_url,
            "expires_at": invite.expires_at.isoformat(),
            "expires_at_display": timezone.localtime(invite.expires_at).strftime("%d.%m.%Y %H:%M"),
            "status": status,
        }
    )


@login_required
@require_POST
def hard_delete_student(request, pk):
    """
    Hard Delete Student (Permanent) - barcha bog'liq ma'lumotlarni ham o'chiradi
    """
    if not _staff_only(request):
        raise PermissionDenied

    center = _get_center(request)
    student = get_object_or_404(U, pk=pk, role="student")
    _assert_same_center(student, center)

    from education.models import TuitionMonth, PaymentAllocation, Payment, Enrollment, Attendance

    with transaction.atomic():
        # 1. Barcha enrollmentlarni topamiz
        enrollments = Enrollment.objects.filter(student=student)
        
        # 2. TuitionMonth va PaymentAllocationlarni o'chiramiz
        for enr in enrollments:
            PaymentAllocation.objects.filter(tuition_month__enrollment=enr).delete()
            TuitionMonth.objects.filter(enrollment=enr).delete()
        
        # 3. Paymentlarni o'chiramiz
        Payment.objects.filter(student=student).delete()
        
        # 4. Enrollmentlarni o'chiramiz
        enrollments.delete()
        
        # 5. Davomatlarni o'chiramiz
        Attendance.objects.filter(student=student).delete()
        
        # 6. Foydalanuvchini o'chiramiz
        student.hard_delete()

    messages.success(request, "Ma'lumotlar butunlay o'chirildi.")
    return redirect("core:stat_students")


# =============================================================================
# USER CRUD (tenant scoped)
# =============================================================================

def _first_day_of_month(d: date) -> date:
    return date(d.year, d.month, 1)


def _parse_month_yyyy_mm(s: str) -> date | None:
    """
    "2026-01" -> date(2026,1,1)
    xato bo'lsa None
    """
    try:
        s = (s or "").strip()
        if not s:
            return None
        y, m = s.split("-")
        y = int(y)
        m = int(m)
        if m < 1 or m > 12:
            return None
        return date(y, m, 1)
    except Exception:
        return None


@login_required
def user_view(request, pk):
    role = getattr(request.user, "role", None)
    if not (_staff_only(request) or role == "parent"):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    user = get_object_or_404(User, pk=pk)
    _assert_same_center(user, center)

    # ✅ Handle Password Reset (Admin/Staff only)
    if request.method == "POST" and "new_password" in request.POST:
        if not _staff_only(request):
            raise PermissionDenied
        new_pass = request.POST.get("new_password")
        if new_pass:
            user.set_password(new_pass)
            user.save()
            messages.success(request, f"{user.get_full_name()} paroli muvaffaqiyatli o'zgartirildi.")
            return redirect("core:user_view", pk=user.pk)

    # ✅ Chaqmoq balance (Tenant Scoped)
    balance_qs = Ledger.objects.filter(student=user)
    if center:
        balance_qs = balance_qs.filter(Q(group__center=center) | Q(rule__center=center) | Q(rule__center__isnull=True))
    balance = balance_qs.aggregate(Sum('ball'))['ball__sum'] or 0
    
    # ✅ Chaqmoq history (Tenant Scoped)
    all_actions = balance_qs.select_related('beruvchi', 'group').order_by("-id")
    page_number = request.GET.get('page', 1)
    paginator = Paginator(all_actions, 10)  # 10 items per page
    actions_page = paginator.get_page(page_number)

    # ✅ Attendance Stats (ONLY FOR THIS CENTER)
    from education.models import Attendance, Enrollment
    total_lessons = Attendance.objects.filter(student=user, group__center=center).count()
    present_days = Attendance.objects.filter(student=user, group__center=center, present=True).count()
    missed_days = total_lessons - present_days
    
    attendance_rate = 0
    if total_lessons > 0:
        attendance_rate = int((present_days / total_lessons) * 100)

    # ✅ Detailed Group Attendance (Calendar View)
    import calendar
    from datetime import date
    now = timezone.now()
    
    # Handle month/year filter
    try:
        selected_year = int(request.GET.get('year', now.year))
        selected_month = int(request.GET.get('month', now.month))
    except (ValueError, TypeError):
        selected_year, selected_month = now.year, now.month
    
    # Month list for filter
    available_months = [
        (1, "Yanvar"), (2, "Fevral"), (3, "Mart"), (4, "Aprel"), (5, "May"), (6, "Iyun"),
        (7, "Iyul"), (8, "Avgust"), (9, "Sentyabr"), (10, "Oktyabr"), (11, "Noyabr"), (12, "Dekabr")
    ]
    available_years = range(now.year - 1, now.year + 1)
    
    # Get the month grid
    month_calendar = calendar.monthcalendar(selected_year, selected_month) # [[0,0,1,2,3,4,5], ...]
    
    # ✅ Enrollments (ONLY FOR THIS CENTER)
    # Include both active and inactive so history is visible
    enrollments_qs = Enrollment.all_objects.filter(
        student=user, 
        group__center=center
    ).select_related('group', 'group__category_obj', 'group__oqituvchi').order_by("-is_active", "-id")
    
    enrollments = list(enrollments_qs)
    enrolled_group_ids = {e.group_id for e in enrollments}
    
    # O'quvchida davomat bor, lekin qandaydir sabab bilan Enrollment qattiq o'chirilgan (Hard Delete) bo'lsa,
    # ularning tarixini ham ko'rsatish uchun ularni ushlaymiz:
    from education.models import Group
    attendance_group_ids = Attendance.objects.filter(student=user, group__center=center).values_list('group_id', flat=True).distinct()
    missing_group_ids = set(attendance_group_ids) - enrolled_group_ids
    
    if missing_group_ids:
        missing_groups = Group.objects.filter(id__in=missing_group_ids).select_related('category_obj', 'oqituvchi')
        for g in missing_groups:
            # Soxta Enrollment yasab, tarixga qo'shamiz
            mock_enr = Enrollment(student=user, group=g, is_active=False)
            mock_enr.created_at = None
            enrollments.append(mock_enr)

    for enr in enrollments:
        # Stats (Total)
        enr.total = Attendance.objects.filter(student=user, group=enr.group).count()
        enr.present = Attendance.objects.filter(student=user, group=enr.group, present=True).count()
        enr.rate = int((enr.present / enr.total * 100)) if enr.total > 0 else 0
        
        # Monthly Stats
        enr.month_present = Attendance.objects.filter(
            student=user, group=enr.group, present=True, 
            date__year=selected_year, date__month=selected_month
        ).count()
        enr.month_absent = Attendance.objects.filter(
            student=user, group=enr.group, present=False, 
            date__year=selected_year, date__month=selected_month
        ).count()
        
        # Attendance records for this group in selected month
        month_atts = {att.date.day: att.present for att in Attendance.objects.filter(
            student=user, group=enr.group, date__year=selected_year, date__month=selected_month
        )}
        
        # Generate final grid for template
        grid = []
        for week in month_calendar:
            week_days = []
            for day in week:
                if day == 0:
                    week_days.append({"day": "", "status": "empty"})
                else:
                    status = "none" # default: no lesson
                    if day in month_atts:
                        status = "present" if month_atts[day] else "absent"
                    
                    week_days.append({
                        "day": day,
                        "status": status,
                        "is_today": (day == now.day and selected_month == now.month and selected_year == now.year)
                    })
            grid.append(week_days)
        enr.calendar_grid = grid
        enr.selected_month_name = next(m[1] for m in available_months if m[0] == selected_month)
    
    # ✅ Rank Calculation (Matches Reyting logic)
    # 1. Base query: students in this center (just like Reyting)
    rank_qs = Ledger.objects.filter(student__role='student')
    if center:
        rank_qs = rank_qs.filter(student__center=center)
        
    # 2. Annotate global balance for each student
    # 3. Count how many have more points than current user
    from django.db.models.functions import Coalesce
    rank = rank_qs.values('student').annotate(
        total_points=Coalesce(Sum('ball'), 0)
    ).filter(total_points__gt=balance).count() + 1

    parent_link_status = None
    can_manage_parent_link = bool(user.role == "student" and _staff_only(request))
    if user.role == "student":
        parent_link_status = _student_parent_link_payload(user)

    # Bir-bosishli "Kirish havolasi" — staff o'quvchi/ota-onaga parolsiz kirish havolasini beradi
    magic_login_url = ""
    if _staff_only(request) and user.role in ("student", "parent"):
        from accounts.magic_login import make_magic_login_url
        base = request.build_absolute_uri("/").rstrip("/")
        magic_login_url = make_magic_login_url(user, base_url=base)

    context = {
        "magic_login_url": magic_login_url,
        "u": user,
        "balance": balance,
        "rank": rank,
        "actions": actions_page,
        "total_lessons": total_lessons,
        "present_days": present_days,
        "missed_days": missed_days,
        "attendance_rate": attendance_rate,
        "enrollments": enrollments,
        "selected_month": selected_month,
        "selected_year": selected_year,
        "available_months": available_months,
        "available_years": available_years,
        "parent_link_status": parent_link_status,
        "can_manage_parent_link": can_manage_parent_link,
    }

    return render(request, "core/user_view.html", context)


@login_required
def user_delete(request, pk):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)

    user = get_object_or_404(U, pk=pk)
    _assert_same_center(user, center)

    if request.method == "POST":
        role = getattr(user, "role", "student")
        user.delete()
        if role == "manager":
            return redirect("core:stat_managers")
        elif role == "teacher":
            return redirect("core:stat_teachers")
        return redirect("core:stat_students")

    return render(request, "core/user_delete.html", {"user": user})


@login_required
def user_edit(request, pk):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)

    user = get_object_or_404(U, id=pk)
    _assert_same_center(user, center)

    all_groups = Group.objects.all()
    all_groups = _try_center_filter(all_groups, center, ["center"]) if center else Group.objects.none()

    # ✅ enrollments ham tenant: enrollment -> group__center
    enrollments = Enrollment.objects.filter(student=user, is_active=True).select_related("group")
    enrollments = _try_center_filter(enrollments, center, ["group__center"])

    # Determine default redirect based on role
    default_next = "/stat/students/"
    if user.role == "manager":
        default_next = "/stat/managers/"
    elif user.role == "teacher":
        default_next = "/stat/teachers/"
    elif user.role == "parent":
        default_next = "/stat/parents/"

    next_url = request.POST.get("next") or request.GET.get("next") or default_next

    month_str = request.POST.get("month") or request.GET.get("month") or ""
    selected_month = _parse_month_yyyy_mm(month_str) or _first_day_of_month(timezone.localdate())

    if request.method == "POST":
        user.ism = (request.POST.get("ism") or "").strip()
        user.familya = (request.POST.get("familya") or "").strip()
        user.otchestvo = (request.POST.get("otchestvo") or "").strip()
        user.email = (request.POST.get("email") or "").strip()
        user.telefon1 = (request.POST.get("telefon1") or "").strip()
        user.telefon2 = (request.POST.get("telefon2") or "").strip()

        # ✅ rolni o'zgartirishni xavfsiz qiling (xohlasangiz)
        new_role = (request.POST.get("role") or "").strip()
        if new_role:
            user.role = new_role

        password = request.POST.get("password")
        if password:
            user.set_password(password)

        # ✅ Save new student fields
        user.birth_date = request.POST.get("birth_date") or None
        user.gender = request.POST.get("gender") or None
        user.passport_id = (request.POST.get("passport_id") or "").strip() or None
        user.jshr = (request.POST.get("jshr") or "").strip() or None
        user.address = (request.POST.get("address") or "").strip() or None

        user.save()

        # 2) Mavjud enrollmentlar bo‘yicha narxlarni yangilash / o‘chirish
        for enroll in enrollments:
            if request.POST.get(f"delete_group_{enroll.id}") == "on":
                if hasattr(enroll, "is_active"):
                    enroll.is_active = False
                    enroll.save(update_fields=["is_active"])
                elif hasattr(enroll, "status"):
                    enroll.status = "inactive"
                    enroll.save(update_fields=["status"])
                else:
                    enroll.delete()
                continue

            field = f"kurs_narhi_{enroll.id}"
            new_price_raw = request.POST.get(field)

            if new_price_raw is not None and str(new_price_raw).strip() != "":
                try:
                    new_price = int(new_price_raw)
                    enroll.kurs_narhi = new_price
                    enroll.save(update_fields=["kurs_narhi"])

                    # ✅ [FIX] Narx o'zgarganda BARCHA eski TuitionMonth yozuvlarini o'chirib,
                    # FAQAT joriy oy uchun bitta yangi yozuv yaratamiz.
                    # Bu 2x/3x narx stacking muammosini oldini oladi!
                    from education.services.tuition import tuition_month_fee_field as _fee_field
                    _cur_month = localdate().replace(day=1)
                    TuitionMonth.objects.filter(enrollment=enroll).delete()
                    TuitionMonth.all_objects.update_or_create(
                        enrollment=enroll,
                        month=_cur_month,
                        defaults={
                            _fee_field(): new_price,
                            "is_deleted": False,
                            "deleted_at": None,
                            "deleted_by": None
                        }
                    )
                except ValueError:
                    pass

        # 3) Yangi guruhga qo‘shish (tenant check bilan!)
        yangi_group_id = request.POST.get("yangi_group_id")
        yangi_group_price = request.POST.get("yangi_group_price")

        if yangi_group_id:
            group = get_object_or_404(Group, id=yangi_group_id)
            # ✅ boshqa center guruhini qo'shib yubormasin
            if center and getattr(group, "center_id", None) != center.id:
                messages.error(request, "Bu guruh boshqa centerga tegishli.")
                return redirect(next_url)

            enrollment_price = int(getattr(group, "kurs_narxi", 0) or 0)
            if yangi_group_price:
                try:
                    enrollment_price = int(yangi_group_price)
                except ValueError:
                    pass

            enroll, _created = Enrollment.all_objects.get_or_create(
                student=user,
                group=group,
                defaults={
                    "center": group.center,
                    "kurs_narhi": enrollment_price,
                    "oqituvchi_foiz": group.oqituvchi_foiz or 40,
                    "monthly_price": enrollment_price,
                    "monthly_lessons": int(getattr(group, "oy_dars_soni", 0) or 0),
                    "joined_at": localdate(),
                },
            )
            if not _created and enroll.is_deleted:
                enroll.restore(restored_by=request.user)
            
            # Ensure it's active if we're re-adding
            enroll.is_active = True
            if yangi_group_price:
                enroll.kurs_narhi = enrollment_price
                enroll.monthly_price = enrollment_price
            elif _created and not getattr(enroll, "kurs_narhi", None):
                enroll.kurs_narhi = enrollment_price
                enroll.monthly_price = enrollment_price
            if not getattr(enroll, "monthly_lessons", None):
                enroll.monthly_lessons = int(getattr(group, "oy_dars_soni", 0) or 0)
            if not getattr(enroll, "joined_at", None):
                enroll.joined_at = localdate()
            enroll.save()
            from education.services.tuition import ensure_tuition_month
            # ✅ Yangi guruhga qo'shilganda joriy oy uchn avtomatik qarzdorlikni yaratish
            ensure_tuition_month(enroll, localdate())

        return redirect(next_url)

    return render(request, "core/user_edit.html", {
        "user_obj": user,
        "enrollments": enrollments,
        "groups": all_groups,
        "next": next_url,
        "month": month_str,
        "selected_month": selected_month,
    })


# =============================================================================
# EXCEL IMPORT/EXPORT (tenant scoped)
# =============================================================================

# _normalize_header, _pick_col, _cell_to_str, _clean_for_login,
# _normalize_phone, _gen_default_password, _gen_unique_gmail_like_email,
# _normalize_gender — funksiyalar core/services/user_import_service.py dan import qilingan (yuqorida).


@login_required
@require_POST
def _process_user_import(request, role="student"):
    """
    Generic User Import logic. 
    Overhauled for robustness and performance.
    """
    # Normalize role (handle plurals)
    role_map = {"students": "student", "teachers": "teacher", "managers": "manager"}
    normalized_role = role_map.get(role, role)
    
    # Determine redirect URL
    redirect_map = {
        "student": "core:stat_students",
        "teacher": "core:teacher_list",
        "manager": "core:stat_managers"
    }
    success_url = redirect_map.get(normalized_role, "core:home")

    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    if not center:
        messages.error(request, "Aktiv markaz tanlanmagan.")
        return redirect(success_url)

    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Excel fayl tanlanmadi.")
        return redirect(success_url)
    file_name = (getattr(f, "name", "") or "").lower()
    if not file_name.endswith(".xlsx"):
        messages.error(request, "Faqat .xlsx formatdagi faylga ruxsat beriladi.")
        return redirect(success_url)
    max_size_bytes = 10 * 1024 * 1024
    if getattr(f, "size", 0) > max_size_bytes:
        messages.error(request, "Fayl hajmi 10MB dan katta bo'lmasligi kerak.")
        return redirect(success_url)

    try:
        wb = load_workbook(filename=f, data_only=True)
        ws = wb.active
        # Look for a sheet with data if active one looks empty
        if ws.max_row < 2:
            for sn in wb.sheetnames:
                if wb[sn].max_row >= 2:
                    ws = wb[sn]
                    break
    except Exception as e:
        logger.error(f"Excel load error for center {center.id}: {e}")
        messages.error(request, "Excel faylni o'qishda xatolik yuz berdi.")
        return redirect(success_url)

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        messages.error(request, "Excel fayl bo'sh yoki ma'lumot yetarli emas.")
        return redirect(success_url)

    # Header detection (scan first 10 rows)
    headers_map = {}
    header_idx = -1
    for i in range(min(15, len(all_rows))):
        temp_map = {}
        for idx, val in enumerate(all_rows[i]):
            key = _normalize_header(val)
            if key: temp_map[key] = idx
        if _pick_col(temp_map, "ism", "name", "fish") is not None:
            header_idx = i
            headers_map = temp_map
            break

    if header_idx == -1:
        messages.error(request, "Excel ustunlari aniqlanmadi. Namuna fayldan foydalanishni tavsiya qilamiz.")
        return redirect(success_url)

    # Mapping columns
    col_map = {
        "ism": _pick_col(headers_map, "ism", "firstname", "name"),
        "fam": _pick_col(headers_map, "familya", "familiya", "lastname"),
        "fish": _pick_col(headers_map, "fish", "f.i.sh", "fullname"),
        "otch": _pick_col(headers_map, "otchestvo", "middlename"),
        "tel1": _pick_col(headers_map, "telefon", "telefon1", "phone", "tel"),
        "birth": _pick_col(headers_map, "tugilgankun", "birthdate", "birthdate", "tug'ilgansana", "sana"),
        "gender": _pick_col(headers_map, "jinsi", "gender", "jinsimalefemale", "pol", "jinsierkakayol"),
        "email": _pick_col(headers_map, "email", "login"),
        "pass": _pick_col(headers_map, "parol", "password"),
        "chaqmoq": _pick_col(headers_map, "chaqmoq", "coins", "coin", "ball"),
    }

    logger.info(f"User Import Debug: Role={normalized_role}")
    logger.info(f"Detected Headers: {list(headers_map.keys())}")
    logger.info(f"Column Map: {col_map}")

    created, skipped, errors, coins_added = 0, 0, 0, 0
    problems = []
    limit_reached_message = ""

    # ✅ Case-insensitive email check: hamma emaillarni lower() qilib olamiz
    all_known_emails = set(e.lower() for e in U.objects.values_list("email", flat=True) if e)

    import_rule = None
    if center and normalized_role == "student":
        import_rule, _ = Rule.objects.get_or_create(
            center=center,
            nom="Excel Import",
            defaults={
                "tur": Rule.PLUS,
                "min_baho": 1,
                "max_baho": 1000
            }
        )

    for r_i, r in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not r or all(v is None for v in r): continue
        
        try:
            ism = _cell_to_str(r[col_map["ism"]]) if (col_map["ism"] is not None and col_map["ism"] < len(r)) else ""
            fam = _cell_to_str(r[col_map["fam"]]) if (col_map["fam"] is not None and col_map["fam"] < len(r)) else ""
            
            if col_map["fish"] is not None and not ism:
                fish = _cell_to_str(r[col_map["fish"]])
                if fish:
                    parts = fish.split()
                    ism = parts[0] if len(parts) > 0 else "Noma'lum"
                    fam = parts[1] if len(parts) > 1 else "User"

            if not ism:
                skipped += 1
                continue

            # Identify or create email
            email_val = _cell_to_str(r[col_map["email"]]).lower() if (col_map["email"] is not None and col_map["email"] < len(r)) else ""
            
            # ✅ Qat'iy tekshiruv: email bo'sh bo'lsa yoki bazada bo'lsa - yangi generatsiya
            if not email_val or email_val in all_known_emails:
                prefix = _clean_for_login(ism) or "user"
                # Takrorlanmaslik uchun tasodifiy sonni oshiramiz
                suffix = secrets.randbelow(90000) + 10000 
                cand = f"{prefix}{suffix}@chaqmoq.uz"
                while cand in all_known_emails:
                    suffix = secrets.randbelow(90000) + 10000
                    cand = f"{prefix}{suffix}@chaqmoq.uz"
                email_val = cand

            tel1 = _normalize_phone(_cell_to_str(r[col_map["tel1"]])) if (col_map["tel1"] is not None and col_map["tel1"] < len(r)) else ""

            if normalized_role == "student":
                from accounts.student_limit import check_student_limit
                limit_state = check_student_limit(center, raise_error=False, actor=request.user)
                if limit_state["is_at_limit"]:
                    limit_reached_message = (
                        f"O'quvchi limiti to'ldi ({limit_state['current_count']}/{limit_state['limit']}, "
                        f"tarif: {limit_state['plan_name']}). Import to'xtatildi."
                    )
                    break
            
            with transaction.atomic():
                u = U.objects.create(
                    email=email_val, role=normalized_role, center=center,
                    ism=ism, familya=fam, first_name=ism, last_name=fam,
                    telefon1=tel1
                )
                
                # Secondary fields
                if col_map["otch"] is not None and col_map["otch"] < len(r):
                    u.otchestvo = _cell_to_str(r[col_map["otch"]])
                
                if col_map["gender"] is not None and col_map["gender"] < len(r):
                    g_raw = r[col_map["gender"]]
                    u.gender = _normalize_gender(g_raw)
                    # Debug log
                    if r_i <= header_idx + 5:
                       logger.info(f"Gender Debug Row {r_i}: Raw='{g_raw}' -> Norm='{u.gender}' Col={col_map['gender']}")
                
                if col_map["birth"] is not None and col_map["birth"] < len(r):
                    bv = r[col_map["birth"]]
                    if isinstance(bv, (datetime.date, datetime.datetime)):
                        u.birth_date = bv
                    elif isinstance(bv, str):
                        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y.%m.%d"):
                            try:
                                u.birth_date = datetime.datetime.strptime(bv.strip(), fmt).date()
                                break
                            except: continue

                # Password
                pv = _cell_to_str(r[col_map["pass"]]) if (col_map["pass"] is not None and col_map["pass"] < len(r)) else ""
                u.set_password(pv if len(pv) >= 4 else _gen_default_password())
                u.save()
                
                # Coins
                if normalized_role == "student" and col_map["chaqmoq"] is not None and col_map["chaqmoq"] < len(r):
                    try:
                        raw_val = _cell_to_str(r[col_map["chaqmoq"]])
                        if raw_val:
                            # Try simple conversion first
                            try:
                                ball = int(float(raw_val))
                            except ValueError:
                                # Fallback: extract digits using pre-imported re
                                digits = re.sub(r"[^\d.-]", "", raw_val)
                                ball = int(float(digits)) if digits else 0
                            
                            if ball != 0:
                                Ledger.objects.create(
                                    student=u, 
                                    ball=ball, 
                                    beruvchi=request.user, 
                                    rule=import_rule,   # ✅ LINK TO CENTER RULE
                                    rule_nom="Excel Import"
                                )
                                coins_added += 1
                    except Exception as e:
                        logger.error(f"Coin import error row {r_i}: {e}")

                all_known_emails.add(email_val)
                created += 1

        except Exception as e:
            errors += 1
            if len(problems) < 5:
                problems.append(f"{r_i}-qator: {str(e)[:50]}")
            logger.error(f"Import row {r_i} error: {e}")

    msg = f"🚀 Import yakunlandi! {created} ta yangi foydalanuvchi qo'shildi."
    if coins_added > 0:
        msg += f" {coins_added} ta o'quvchiga chaqmoq berildi."
    
    messages.success(request, msg)
    if limit_reached_message:
        messages.error(request, f"❌ {limit_reached_message}")
    if skipped: messages.info(request, f"ℹ️ {skipped} ta qator ism yo'qligi sabab tashlab o'tildi.")
    if errors: messages.warning(request, f"⚠️ {errors} ta qatorda texnik xatolik: {', '.join(problems[:3])}")
    
    return redirect(success_url)


    return redirect("core:stat_students")


@login_required
def students_download_template(request):
    """O‘quvchilarni import qilish uchun shablon Excel faylini yuklab beradi."""
    if not _staff_only(request):
        return HttpResponse("Ruxsat yo‘q", status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Template"

    # Headers
    headers = [
        "Ism", "Familya", "Otchestvo", 
        "Telefon", "Telefon2", 
        "Tegulgankun", "Jinsi", 
        "Email", "Parol", "Chaqmoq"
    ]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Sample row
    ws.append([
        "Ali", "Valiyev", "G‘anisher o‘g‘li", 
        "+998901234567", "", 
        "2010.05.15", "erkak", 
        "ali123@gmail.com", "password123", "50"
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="chaqmoq_students_template.xlsx"'
    wb.save(response)
    return response


@login_required
def stat_users_export_excel(request):
    """
    Generic export for different user roles.
    """
    if not _staff_only(request):
        return HttpResponse("Ruxsat yo‘q", status=403)

    role = request.GET.get("role", "student")
    
    # Normalize role
    role_map = {
        "students": "student",
        "teachers": "teacher",
        "managers": "manager",
        "parents": "parent"
    }
    role = role_map.get(role, role)
    
    center = _get_center(request)
    if not center:
        return HttpResponse("Active center tanlanmagan", status=400)

    status = request.GET.get("status", "active").strip()
    rows = U.objects.filter(role=role, center=center)
    
    if status == "archived":
        rows = rows.filter(is_archived=True)
    else:
        rows = rows.filter(is_archived=False)
        
    rows = rows.order_by("id")
    
    if role == "student":
        rows = rows.annotate(jami_chaqmoq=Sum("ledger__ball"))

    def _excel_safe(value):
        text = "" if value is None else str(value)
        if text.startswith(("=", "+", "-", "@")):
            return f"'{text}"
        return text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = role.capitalize() + "lar"

    # Define headers based on role
    # Define headers based on role
    if role == "student":
        headers = ["Ism", "Familya", "Otchestvo", "Telefon", "Telefon2", "Tug'ilgan sana", "Jinsi (Erkak/Ayol)", "Email", "Parol", "Chaqmoq"]
    else:
        # Teachers / Managers
        headers = ["Ism", "Familya", "Otchestvo", "Telefon", "Telefon2", "Tug'ilgan sana", "Email", "Parol"]
    
    ws.append(headers)

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Data
    gender_map = {"male": "Erkak", "female": "Ayol"}
    
    for u in rows:
        if role == "student":
            g_disp = gender_map.get(u.gender, "")
            
            ws.append([
                _excel_safe(u.ism),
                _excel_safe(u.familya),
                _excel_safe(u.otchestvo or ""),
                _excel_safe(u.telefon1 or ""),
                _excel_safe(u.telefon2 or ""),
                _excel_safe(u.birth_date.strftime("%Y-%m-%d") if u.birth_date else ""),
                _excel_safe(g_disp),
                _excel_safe(u.email),
                "***",
                u.jami_chaqmoq or 0,
            ])
        else:
            ws.append([
                _excel_safe(u.ism),
                _excel_safe(u.familya),
                _excel_safe(u.otchestvo or ""),
                _excel_safe(u.telefon1 or ""),
                _excel_safe(u.telefon2 or ""),
                _excel_safe(u.birth_date.strftime("%Y-%m-%d") if u.birth_date else ""),
                _excel_safe(u.email),
                "***",
            ])

    safe_center_name = "".join(ch for ch in (center.name or "") if ch.isalnum() or ch in ("-", "_", " "))
    safe_center_name = safe_center_name.strip() or f"center-{center.id}"
    filename = f"{role.capitalize()}s_{safe_center_name}"
    if status == "archived":
        filename += "_Arxiv"
    
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

@login_required
def users_download_template(request):
    """
    Generic template download with sample data.
    """
    role = request.GET.get("role", "student")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shablon"

    if role in ("student", "students"):
        headers = ["Ism", "Familya", "Otchestvo", "Telefon", "Telefon2", "Tug'ilgan sana", "Jinsi (Erkak/Ayol)", "Email", "Parol", "Chaqmoq"]
        sample = ["Amirxon", "O'rinbayev", "Temur o'g'li", "901234567", "", "2005-05-15", "Erkak", "amirxon@gmail.com", "12345678", 100]
    else:
        # Teachers / Managers
        headers = ["Ism", "Familya", "Otchestvo", "Telefon", "Telefon2", "Tug'ilgan sana", "Email", "Parol"]
        sample = ["Amirxon", "O'rinbayev", "Vali o'g'li", "901234567", "", "2000-01-01", "amirxon@gmail.com", "12345678"]

    ws.append(headers)
    ws.append(sample)
    
    # Simple style
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{role}_template.xlsx"'
    wb.save(response)
    return response

@login_required
@require_POST
def users_import_excel(request):
    """
    Generic import placeholder. For now, redirects to student import if role is student.
    In future, can be expanded for managers.
    """
    role = request.POST.get("role") or request.GET.get("role", "student")
    return _process_user_import(request, role=role)
    messages.warning(request, f"{role.capitalize()}lar uchun import hozircha mavjud emas.")
    return redirect(f"core:stat_{role}s" if role != "teacher" else "core:stat_teachers")


# =============================================================================
# PRODUCTS / REQUESTS / LEDGER (tenant scoped)
# =============================================================================

@login_required
def stat_products(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    q = request.GET.get("q", "").strip()

    rows = Product.objects.all()
    rows = _try_center_filter(rows, center, ["center"]) if center else rows.none()

    # field names nom/izoh har xil bo'lishi mumkin
    if q:
        f1 = "nom__icontains" if _has_field(Product, "nom") else "name__icontains"
        f2 = "izoh__icontains" if _has_field(Product, "izoh") else "description__icontains"
        try:
            rows = rows.filter(Q(**{f1: q}) | Q(**{f2: q}))
        except Exception:
            rows = rows.filter(Q(id__isnull=False))  # no-op fallback

    order_field = "-yaratilgan" if _has_field(Product, "yaratilgan") else "-id"
    rows = rows.order_by(order_field)

    return render(request, "core/stats_products.html", {"title": "Mahsulotlar", "rows": rows})


@login_required
def stat_requests(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)
    status = request.GET.get("status", "").strip()

    rows = PurchaseRequest.objects.select_related("student", "product", "manager")
    rows = _try_center_filter(rows, center, ["center", "student__center", "manager__center"]) if center else rows.none()

    if status in ("pending", "approved", "rejected"):
        rows = rows.filter(status=status)

    # ordering field sana/created_at
    if _has_field(PurchaseRequest, "sana"):
        rows = rows.order_by("-sana")
    elif _has_field(PurchaseRequest, "created_at"):
        rows = rows.order_by("-created_at")
    else:
        rows = rows.order_by("-id")

    return render(request, "core/stats_requests.html", {
        "title": "Kutilayotgan so‘rovlar",
        "rows": rows,
        "status": status
    })


@login_required
def stat_ledger(request):
    if not _staff_only(request):
        return render(request, "core/dashboard_guest.html")

    center = _get_center(request)

    ledger_qs = Ledger.objects.all()
    ledger_qs = _try_center_filter(ledger_qs, center, ["center", "student__center", "group__center"]) if center else ledger_qs.none()

    leaderboard = (ledger_qs
                   .values("student__id", "student__ism", "student__familya")
                   .annotate(jami=Sum("ball"))
                   .order_by("-jami"))

    last = ledger_qs.select_related("student", "rule", "group")

    if _has_field(Ledger, "sana"):
        last = last.order_by("-sana")[:50]
    elif _has_field(Ledger, "created_at"):
        last = last.order_by("-created_at")[:50]
    else:
        last = last.order_by("-id")[:50]

    jami = ledger_qs.aggregate(s=Sum("ball"))["s"] or 0

    return render(request, "core/stats_ledger.html", {
        "leaderboard": leaderboard,
        "last": last,
        "sum_all": jami
    })


# =============================================================================
# PROFILE
# =============================================================================

@login_required
def profile_view(request):
    user = request.user

    pform = ProfileForm(instance=user, user=user)
    pass_form = SetPasswordForm(user=user)
    
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "profile":
            pform = ProfileForm(request.POST, request.FILES, instance=user, user=user)
            if pform.is_valid():
                pform.save()
                messages.success(request, "✅ Profil yangilandi")
                return redirect("core:profile")
            else:
                messages.error(request, "❌ Profilni saqlashda xatolik bor")

        elif action == "password":
            pass_form = SetPasswordForm(user=user, data=request.POST)
            if pass_form.is_valid():
                u = pass_form.save()
                update_session_auth_hash(request, u)
                messages.success(request, "✅ Parol yangilandi")
                return redirect("core:profile")
            else:
                messages.error(request, "❌ Parolni o‘zgartirishda xatolik bor")

    return render(request, "core/profile_manager.html", {
        "pform": pform,
        "pass_form": pass_form,
    })


# =============================================================================
# NOTIFICATIONS
# =============================================================================
from core.models import Notification

@login_required
def notifications_view(request):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # 1. Barcha xabarlar (order_by created_at)
    qs = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    
    # 2. Per page
    per_page = request.GET.get('per_page', '10')
    paginator_size = 10
    
    if per_page == 'all':
        paginator_size = qs.count() if qs.count() > 0 else 10
    elif per_page in ['10', '20', '50', '100']:
        paginator_size = int(per_page)
    else:
        per_page = '10' # Default fallback for invalid input
        paginator_size = 10
    
    paginator = Paginator(qs, paginator_size)
    page = request.GET.get('page')
    
    try:
        notifications = paginator.page(page)
    except PageNotAnInteger:
        notifications = paginator.page(1)
    except EmptyPage:
        notifications = paginator.page(paginator.num_pages)
        
    return render(request, "core/notifications.html", {
        "notifications": notifications,
        "per_page": per_page
    })

@login_required
def notifications_mark_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    messages.success(request, "Barcha xabarlar o‘qildi deb belgilandi.")
    return redirect("core:notifications")


@login_required
def notification_read_one(request, pk):
    """Bitta xabarni o'qildi deb belgilaydi (xabar modalda ochilganda)."""
    if request.method == "POST":
        Notification.objects.filter(recipient=request.user, pk=pk).update(is_read=True)
    return JsonResponse({"ok": True})

@login_required
def notification_broadcast(request):
    # Only Directors/Managers
    if request.user.role not in ('director', 'manager') and not request.user.is_superuser:
        messages.error(request, "Sizga ruxsat yo‘q.")
        return redirect("core:home")

    if request.method == "POST":
        target = (request.POST.get("target") or "").strip()
        message = (request.POST.get("message") or "").strip()
        title = (request.POST.get("title") or "Muhim xabar").strip()

        allowed_targets = {"students", "students_parents", "teachers", "all"}
        if target not in allowed_targets:
            messages.error(request, "Noto'g'ri qabul qiluvchi turi.")
            return redirect("core:notifications")
        if not message:
            messages.error(request, "Xabar matni bo'sh bo'lmasligi kerak.")
            return redirect("core:notifications")
        
        center = _get_center(request)
        if not center:
            messages.error(request, "Markaz tanlanmagan.")
            return redirect("core:notifications")

        if target == "students":
            recipients = U.objects.filter(role="student", center=center, is_archived=False)
        elif target == "students_parents":
            recipients = U.objects.filter(role__in=["student", "parent"], center=center, is_archived=False)
        elif target == "teachers":
            recipients = U.objects.filter(role="teacher", center=center, is_archived=False)
        elif target == "all":
            recipients = U.objects.filter(center=center, is_archived=False)
        
        count = 0
        bulk_list = []
        for r in recipients:
            if r == request.user: continue
            bulk_list.append(Notification(
                recipient=r,
                sender=request.user,
                center=center,
                title=title,
                message=message,
                type='broadcast'
            ))
            count += 1
        
        if bulk_list:
            Notification.objects.bulk_create(bulk_list)
            
            # ✅ Senderga ham nusxa (history uchun)
            Notification.objects.create(
                recipient=request.user,
                sender=request.user,
                center=center,
                title=f"Yuborildi: {title}",
                message=f"Kimlarga: {target}\n{message}",
                type='broadcast',
                is_read=True
            )
            
        messages.success(request, f"{count} ta foydalanuvchiga xabar yuborildi.")
        return redirect("core:notifications")
    
    return redirect("core:notifications")


@login_required
def notification_preferences_view(request):
    """Foydalanuvchi bildirishnoma afzalliklarini boshqarish."""
    from .models import NotificationPreference

    pref, _ = NotificationPreference.objects.get_or_create(user=request.user)

    if request.method == "POST":
        pref.receive_coin = request.POST.get("receive_coin") == "on"
        pref.receive_broadcast = request.POST.get("receive_broadcast") == "on"
        pref.receive_purchase = request.POST.get("receive_purchase") == "on"
        pref.receive_system = request.POST.get("receive_system") == "on"
        pref.save()
        messages.success(request, "Bildirishnoma sozlamalari saqlandi.")
        return redirect("core:notification_preferences")

    return render(request, "core/notification_preferences.html", {"pref": pref})


def _get_low_activity_data(center, limit=10):
    if not center:
        return []

    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Count, Q, Prefetch
    from education.models import Enrollment
    from accounts.models import User

    today = timezone.localtime(timezone.now()).date()
    thirty_days_ago = today - timedelta(days=30)

    base_qs = User.objects.filter(center=center, role='student', is_archived=False)
    # N+1 fix: prefetch_related bilan guruh ma'lumotini oldindan yuklaymiz
    candidates = (
        base_qs
        .annotate(
            att_count=Count('attendance', filter=Q(attendance__date__gte=thirty_days_ago, attendance__present=True))
        )
        .prefetch_related(
            Prefetch(
                'enrollments',
                queryset=Enrollment.objects.filter(is_active=True).select_related('group'),
                to_attr='_active_enrollments',
            )
        )
        .filter(att_count__lt=8)
        .order_by('att_count')[:limit * 3]  # limit*3 - filter uchun zapas
    )

    low_list = []
    for s in candidates:
        active_enrs = getattr(s, '_active_enrollments', [])
        enr = active_enrs[0] if active_enrs else None
        reasons = []
        att_pct = round((s.att_count / 12) * 100) if s.att_count < 12 else 100

        if s.att_count < 5:
            reasons.append(f"Davomat past ({att_pct}%)")
        if enr and getattr(enr, 'jami_tolangan', 0) < getattr(enr, 'kurs_narhi', 0):
            reasons.append("To'lov kechikkan")
        if getattr(s, 'last_login', None) and (timezone.now() - s.last_login).days > 10:
            reasons.append("Kirmagan (10+ kun)")

        if not reasons:
            reasons.append("Kam faol")

        low_list.append({
            'student_id': s.id,
            'name': f"{s.ism} {s.familya}",
            'avatar': s.avatar.url if getattr(s, 'avatar', None) else f"https://ui-avatars.com/api/?name={s.ism}+{s.familya}&background=random",
            'course': enr.group.nom if enr else "Guruhsiz",
            'phone': s.telefon1 or "Kiritilmagan",
            'status': att_pct,
            'reasons': reasons,
        })
        if len(low_list) >= limit:
            break

    return low_list

@login_required
def low_activity_students(request):
    """Churn Prediction — O'quvchi ketish xavfi monitoringi."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) in ('director', 'manager')):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    center = getattr(request, 'center', None) or getattr(request.user, 'center', None)
    if not center:
        return render(request, "core/low_activity_students.html", {"page_obj": [], "stats": {}})

    from .models import ChurnRisk
    from .churn_service import run_churn_assessment

    # POST: qo'lda yangilash tugmasi
    if request.method == 'POST' and request.POST.get('action') == 'refresh':
        run_churn_assessment(center, notify_managers=True)
        messages.success(request, "Churn tahlili yangilandi va xavfli o'quvchilar uchun xabar yuborildi.")
        return redirect(request.path)

    # Agar hech qanday yozuv yo'q bo'lsa avtomatik hisoblash
    if not ChurnRisk.objects.filter(center=center).exists():
        run_churn_assessment(center, notify_managers=False)

    # Statistika
    all_risks   = ChurnRisk.objects.filter(center=center)
    total       = all_risks.count()
    high_count  = all_risks.filter(risk_level='high').count()
    medium_count= all_risks.filter(risk_level='medium').count()
    low_count   = all_risks.filter(risk_level='low').count()
    active_risk = high_count + medium_count

    last_assessed = all_risks.order_by('-assessed_at').values_list('assessed_at', flat=True).first()

    stats = {
        'total':  total,
        'high':   high_count,
        'medium': medium_count,
        'low':    low_count,
        'active': active_risk,
        'last_assessed': last_assessed,
    }

    # Filtrlar
    q            = request.GET.get('q', '').strip()
    filter_level = request.GET.get('level', '').strip()

    # Default — faqat haqiqiy xavfli (high + medium) ko'rinadi.
    # Past xavf (low) — ataylab tanlangandagina ko'rinadi.
    if filter_level not in ('high', 'medium', 'low', 'all'):
        filter_level = 'active'

    qs = all_risks.select_related('student', 'student__center')

    if filter_level == 'active':
        qs = qs.filter(risk_level__in=('high', 'medium'))
    elif filter_level in ('high', 'medium', 'low'):
        qs = qs.filter(risk_level=filter_level)
    # 'all' — hammasi

    if q:
        qs = qs.filter(
            Q(student__ism__icontains=q) | Q(student__familya__icontains=q)
        )

    # Sahifalar
    paginator   = Paginator(qs, 15)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    # Template uchun qulay ko'rinishga o'tkazish
    rows = []
    for risk in page_obj:
        s   = risk.student
        enrs = list(
            s.enrollments.filter(center=center, is_active=True)
            .select_related('group', 'group__oqituvchi')
        )
        primary = enrs[0] if enrs else None
        groups_label = ", ".join(e.group.nom for e in enrs[:3] if e.group) if enrs else "Guruhsiz"
        if len(enrs) > 3:
            groups_label += f" +{len(enrs) - 3}"
        teacher_label = ""
        if primary and primary.group and primary.group.oqituvchi:
            t = primary.group.oqituvchi
            teacher_label = f"{t.ism} {t.familya}".strip()
        avatar = (
            s.avatar.url
            if getattr(s, 'avatar', None) and s.avatar
            else f"https://ui-avatars.com/api/?name={s.ism}+{s.familya}&background=1e293b&color=94a3b8&size=48"
        )
        rows.append({
            'risk':         risk,
            'student_id':   s.id,
            'name':         f"{s.ism} {s.familya}",
            'avatar':       avatar,
            'phone':        getattr(s, 'telefon1', '') or '',
            'phone2':       getattr(s, 'telefon2', '') or '',
            'telegram_id':  getattr(s, 'telegram_id', '') or '',
            'course':       groups_label,
            'teacher':      teacher_label,
            'enroll_count': len(enrs),
        })

    return render(request, 'core/low_activity_students.html', {
        'rows':         rows,
        'page_obj':     page_obj,
        'stats':        stats,
        'q':            q,
        'filter_level': filter_level,
    })


@login_required
def dangerous_students(request):
    """
    Dars qoldirgan o'quvchilar (Attendance asosida) — to'liq ro'yxat.
    URL: /boshqaruv/xavflilar/?month=YYYY-MM&min=3
    Bu Churn'dan farqli — faqat tanlangan oyda 3+ darsni absent qilganlar.
    """
    if not (request.user.is_superuser or getattr(request.user, 'role', None) in ('director', 'manager')):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    center = getattr(request, 'center', None) or getattr(request.user, 'center', None)
    today = timezone.localdate()

    # Oy parametri (default: joriy oy)
    month_str = (request.GET.get('month') or '').strip()
    selected_month = None
    if month_str:
        try:
            year, mon = month_str.split('-')
            selected_month = date(int(year), int(mon), 1)
        except (ValueError, IndexError):
            selected_month = None
    if selected_month is None:
        selected_month = today.replace(day=1)

    import calendar as _cal
    last_day = _cal.monthrange(selected_month.year, selected_month.month)[1]
    d_from = selected_month
    d_to = selected_month.replace(day=last_day)
    if d_to > today:
        d_to = today

    # Min qoldirilgan darslar
    try:
        min_missed = max(1, int(request.GET.get('min') or 3))
    except (ValueError, TypeError):
        min_missed = 3

    # Qidiruv
    q = (request.GET.get('q') or '').strip()

    # Oy ro'yxati (oxirgi 12 oy)
    months = []
    cursor = today.replace(day=1)
    uz_months = ["Yanvar","Fevral","Mart","Aprel","May","Iyun","Iyul","Avgust","Sentyabr","Oktyabr","Noyabr","Dekabr"]
    for _ in range(12):
        months.append({
            "value": cursor.strftime('%Y-%m'),
            "label": f"{uz_months[cursor.month-1]} {cursor.year}",
        })
        # bir oy orqaga
        if cursor.month == 1:
            cursor = cursor.replace(year=cursor.year-1, month=12)
        else:
            cursor = cursor.replace(month=cursor.month-1)

    rows = []
    if center:
        from education.models import Attendance
        absent_filter = (
            Q(status__in=["absent_excused", "absent_unexcused"])
            | Q(present=False, forced=False)
        )
        att_qs = (
            Attendance.objects.filter(
                group__center=center,
                date__range=(d_from, d_to),
            )
            .filter(absent_filter)
            .values(
                "student_id",
                "student__ism",
                "student__familya",
                "student__telefon1",
                "student__telegram_id",
                "group__nom",
                "group__oqituvchi__ism",
                "group__oqituvchi__familya",
            )
            .annotate(missed=Count("id"))
            .order_by("-missed")
        )
        # Studentni guruhlash
        st_map = {}
        for r in att_qs:
            sid = r["student_id"]
            slot = st_map.setdefault(sid, {
                "id": sid,
                "name": f"{r['student__ism'] or ''} {r['student__familya'] or ''}".strip() or "Noma'lum",
                "phone": r["student__telefon1"] or "",
                "telegram_id": r["student__telegram_id"] or "",
                "groups": [],
                "teacher": "",
                "missed": 0,
            })
            slot["groups"].append(r["group__nom"] or "Guruh")
            slot["missed"] += r["missed"]
            if not slot["teacher"]:
                t_ism = r.get("group__oqituvchi__ism") or ""
                t_fam = r.get("group__oqituvchi__familya") or ""
                slot["teacher"] = f"{t_ism} {t_fam}".strip()

        # Davomat foizini hisoblash
        if st_map:
            att_total = list(
                Attendance.objects.filter(
                    group__center=center,
                    date__range=(d_from, d_to),
                    student_id__in=list(st_map.keys()),
                )
                .values("student_id")
                .annotate(
                    tot=Count("id"),
                    pres=Count("id", filter=Q(present=True) | Q(forced=True)),
                )
            )
            att_map = {r["student_id"]: r for r in att_total}
            for sid, info in st_map.items():
                a = att_map.get(sid, {})
                tot, pres = a.get("tot", 0), a.get("pres", 0)
                info["rate"] = round(pres / tot * 100, 1) if tot else 0
                info["groups_label"] = ", ".join(info["groups"][:3])
                if len(info["groups"]) > 3:
                    info["groups_label"] += f" +{len(info['groups']) - 3}"

        rows = sorted(
            [v for v in st_map.values() if v["missed"] >= min_missed],
            key=lambda x: (-x["missed"], x.get("rate", 0)),
        )

        # Qidiruv
        if q:
            q_low = q.lower()
            rows = [
                r for r in rows
                if q_low in r["name"].lower() or q_low in (r["phone"] or "")
            ]

    selected_label = f"{uz_months[selected_month.month-1]} {selected_month.year}"

    return render(request, "core/dangerous_students.html", {
        "rows": rows,
        "months": months,
        "selected_month": selected_month.strftime('%Y-%m'),
        "selected_label": selected_label,
        "min_missed": min_missed,
        "q": q,
        "total_count": len(rows),
        "d_from": d_from,
        "d_to": d_to,
    })


@login_required
def churn_notify_student(request, pk):
    """Alohida o'quvchi uchun menejerga xabar yuborish."""
    if request.method != 'POST':
        return redirect('core:low_activity_students')

    if not (request.user.is_superuser or getattr(request.user, 'role', None) in ('director', 'manager')):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    center = getattr(request, 'center', None) or getattr(request.user, 'center', None)
    from .models import ChurnRisk, Notification
    from accounts.models import User

    try:
        risk = ChurnRisk.objects.get(center=center, student_id=pk)
    except ChurnRisk.DoesNotExist:
        messages.error(request, "Yozuv topilmadi.")
        return redirect('core:low_activity_students')

    managers = User.objects.filter(center=center, role__in=('manager', 'director'))
    reasons_str = " | ".join(risk.reasons) if risk.reasons else "Kam faollik"
    title   = f"Eslatma: {risk.student.get_full_name()} bilan bog'laning"
    message = f"Ball: {risk.risk_score}/100 · {reasons_str}"

    for mgr in managers:
        Notification.objects.create(
            center=center, recipient=mgr,
            title=title, message=message, type='system',
        )

    ChurnRisk.objects.filter(pk=risk.pk).update(
        notified=True, notified_at=timezone.now()
    )
    messages.success(request, f"{risk.student.get_full_name()} uchun menejerga xabar yuborildi.")
    return redirect('core:low_activity_students')


# ══════════════════════════════════════════════════════════════════
#  GAME MANAGEMENT VIEWS
# ══════════════════════════════════════════════════════════════════

@login_required
def game_settings_view(request):
    """Director/Manager: O'yin sozlamalari sahifasi."""
    from core.models import GlobalGameConfig, CenterGameConfig, GameSuggestion, GameBallsConfig
    from core.api_views import _GAME_CONFIGS, _sp_center

    u = request.user
    if u.role not in ("director", "manager") and not u.is_superuser:
        from django.contrib import messages
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    center = _sp_center(request)
    global_disabled = set(
        GlobalGameConfig.objects.filter(is_enabled=False).values_list("game_slug", flat=True)
    )
    center_cfgs = {
        cc.game_slug: cc for cc in CenterGameConfig.objects.filter(center=center)
    }
    games = []
    for slug, cfg in _GAME_CONFIGS.items():
        cc = center_cfgs.get(slug)
        games.append({
            "slug": slug,
            "name": cfg["name"],
            "emoji": cfg["emoji"],
            "max_coins_default": cfg["max_coins"],
            "min_score_default": cfg["min_score"],
            "is_globally_enabled": slug not in global_disabled,
            "is_center_enabled": cc.is_enabled if cc else True,
            "center_max_coins": cc.max_coins if cc else 0,
            "center_min_score": cc.min_score if cc else 0,
        })

    suggestions = GameSuggestion.objects.filter(center=center).select_related("student")[:60]
    center_name = getattr(center, "name", "Markaz")

    balls_config, _ = GameBallsConfig.objects.get_or_create(
        center=center,
        defaults={"min_balls_to_convert": 100, "chaqmoq_per_conversion": 5},
    ) if center else (None, False)

    return render(request, "core/game_settings.html", {
        "games": games,
        "suggestions": suggestions,
        "center_name": center_name,
        "is_superuser": u.is_superuser,
        "balls_config": balls_config,
    })


@login_required
def game_super_admin_view(request):
    """Super Admin: global o'yin yoqish/o'chirish sahifasi."""
    from core.models import GlobalGameConfig
    from core.api_views import _GAME_CONFIGS

    if not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, "Faqat Super Admin uchun.")
        return redirect("core:home")

    existing = {g.game_slug: g.is_enabled for g in GlobalGameConfig.objects.all()}
    games = [
        {
            "slug": slug,
            "name": cfg["name"],
            "emoji": cfg["emoji"],
            "is_enabled": existing.get(slug, True),
        }
        for slug, cfg in _GAME_CONFIGS.items()
    ]
    return render(request, "core/game_super_admin.html", {"games": games})


@login_required
def game_hub_view(request):
    """O'yinlar markazi — barcha o'yinlar, stats, leaderboard."""
    from core.api_views import _GAME_CONFIGS, _sp_center, _get_global_game_cfg, _get_game_center_cfg, _resolve_game_cfg
    from core.models import StudentGameProgress, GameSession

    center = _sp_center(request)
    u = request.user
    role = getattr(u, "role", None)
    is_student = role == "student"

    today = localdate()

    global_disabled = _get_global_game_cfg()
    center_cfg_map = _get_game_center_cfg(center)

    # Per-game progress for this student
    progress_map = {}
    if is_student and center:
        progress_map = {
            p.game_slug: p
            for p in StudentGameProgress.objects.filter(center=center, student=u)
        }

    # Balls played today per game
    today_balls_map = {}
    if is_student and center:
        for row in (
            GameSession.objects.filter(center=center, student=u, played_at__date=today, balls_earned=1)
            .values("game_slug")
            .annotate(c=Sum("balls_earned"))
        ):
            today_balls_map[row["game_slug"]] = row["c"]

    # Build games list
    games = []
    enabled_count = 0
    for slug, cfg in _GAME_CONFIGS.items():
        eff = _resolve_game_cfg(slug, cfg, center_cfg_map, global_disabled)
        p = progress_map.get(slug)
        is_enabled = eff is not None
        if is_enabled:
            enabled_count += 1
        games.append({
            "slug": slug,
            "name": cfg["name"],
            "emoji": cfg["emoji"],
            "max_coins": eff["max_coins"] if eff else cfg["max_coins"],
            "is_enabled": is_enabled,
            "current_level": p.current_level if p else 1,
            "earned_today": today_balls_map.get(slug, 0),
            "played_today": today_balls_map.get(slug, 0) > 0,
        })

    # Overall stats
    games_today = 0
    streak = 0
    best_level = 1
    if is_student and center:
        games_today = GameSession.objects.filter(
            center=center, student=u, balls_earned=1, played_at__date=today
        ).values("game_slug").distinct().count()
        best_level = max((p.current_level for p in progress_map.values()), default=1)

    from core.models import StudentBallsWallet, GameBallsConfig
    wallet = StudentBallsWallet.objects.filter(center=center, student=u).first() if is_student and center else None
    total_balls = wallet.total_balls if wallet else 0
    lifetime_balls = wallet.lifetime_balls if wallet else 0
    daily_balls = games_today

    balls_config = GameBallsConfig.objects.filter(center=center).first() if center else None
    min_balls = balls_config.min_balls_to_convert if balls_config else 100
    chaqmoq_per_conv = balls_config.chaqmoq_per_conversion if balls_config else 5
    can_convert = total_balls >= min_balls
    wallet_ring_offset = round(251.3 * (1 - min(total_balls / max(min_balls, 1), 1)))

    # Streak (consecutive days with sessions)
    if is_student and center:
        d = today
        for _ in range(30):
            if GameSession.objects.filter(center=center, student=u, played_at__date=d).exists():
                streak += 1
                d -= timedelta(days=1)
            else:
                break

    # Center leaderboard (top 5 this week, balls-based)
    leaderboard = []
    if center:
        week_start = today - timedelta(days=today.weekday())
        lb_rows = list(
            GameSession.objects.filter(center=center, played_at__date__gte=week_start, balls_earned=1)
            .values("student_id")
            .annotate(wk=Count("id"))
            .order_by("-wk")[:5]
        )
        ids = [r["student_id"] for r in lb_rows]
        names = {u2.id: u2.get_full_name() or u2.email for u2 in User.objects.filter(id__in=ids).only("id", "first_name", "last_name", "email")}
        leaderboard = [
            {"name": names.get(r["student_id"], "—"), "coins": r["wk"], "is_me": is_student and r["student_id"] == u.id}
            for r in lb_rows
        ]

    # Global leaderboard for rating tab
    global_period = request.GET.get("rperiod", "week")
    if global_period == "week":
        gr_from = today - timedelta(days=today.weekday())
    elif global_period == "month":
        gr_from = today.replace(day=1)
    else:
        gr_from = None

    from core.models import StudentGameProgress as _SGP
    from django.db.models import Max as _Max
    gr_qs = GameSession.objects.filter(balls_earned=1)
    if gr_from:
        gr_qs = gr_qs.filter(played_at__date__gte=gr_from)

    gr_rows = list(
        gr_qs.values("student_id")
        .annotate(total=Count("id"))
        .order_by("-total")[:20]
    )
    gr_ids = [r["student_id"] for r in gr_rows]
    gr_users = {
        u2.id: u2
        for u2 in User.objects.filter(id__in=gr_ids)
        .select_related("center")
        .only("id", "first_name", "last_name", "email", "center")
    }
    gr_levels = {
        row["student_id"]: (row["ml"] or 1)
        for row in _SGP.objects.filter(student_id__in=gr_ids)
        .values("student_id")
        .annotate(ml=_Max("current_level"))
    }

    def _short(u3):
        full = u3.get_full_name().strip()
        if not full:
            return (u3.email or "O'quvchi").split("@")[0]
        parts = full.split()
        return f"{parts[0]} {parts[1][0]}." if len(parts) > 1 else parts[0]

    global_leaderboard = []
    my_global_rank = None
    for i, row in enumerate(gr_rows, 1):
        u2 = gr_users.get(row["student_id"])
        if not u2:
            continue
        is_me = is_student and row["student_id"] == u.id
        if is_me:
            my_global_rank = i
        global_leaderboard.append({
            "rank": i,
            "name": _short(u2),
            "center": u2.center.name if u2.center else "—",
            "balls": row["total"],
            "level": gr_levels.get(row["student_id"], 1),
            "is_me": is_me,
        })

    # My global position if outside top 20
    my_global_row = None
    if is_student and my_global_rank is None:
        my_balls_qs = GameSession.objects.filter(student=u, balls_earned=1)
        if gr_from:
            my_balls_qs = my_balls_qs.filter(played_at__date__gte=gr_from)
        my_balls = my_balls_qs.aggregate(t=Count("id"))["t"] or 0
        if my_balls > 0:
            better = (
                gr_qs.values("student_id")
                .annotate(t=Count("id"))
                .filter(t__gt=my_balls)
                .count()
            )
            my_global_row = {
                "rank": better + 1,
                "name": u.get_full_name() or u.email or "O'quvchi",
                "center": u.center.name if getattr(u, "center", None) else "—",
                "balls": my_balls,
                "level": gr_levels.get(u.id, 1),
                "is_me": True,
            }

    # Recent sessions (last 8)
    recent_sessions = []
    if is_student and center:
        recent_sessions = list(
            GameSession.objects.filter(center=center, student=u)
            .order_by("-played_at")[:8]
            .values("game_slug", "score", "coins_earned", "balls_earned", "played_at")
        )
        for s in recent_sessions:
            cfg = _GAME_CONFIGS.get(s["game_slug"], {})
            s["game_name"] = cfg.get("name", s["game_slug"])
            s["game_emoji"] = cfg.get("emoji", "🎮")

    # Ring: circumference of r=35 ≈ 219.9
    ratio = min(games_today / max(enabled_count, 1), 1)
    daily_dashoffset = round(219.9 * (1 - ratio))
    daily_pct_bar    = round(ratio * 100)

    # Daily challenge: pick an enabled game not yet played today, fallback to first enabled
    daily_challenge = None
    for g in games:
        if g["is_enabled"] and not g["played_today"]:
            cfg = _GAME_CONFIGS.get(g["slug"], {})
            daily_challenge = {"slug": g["slug"], "name": g["name"], "emoji": g["emoji"]}
            break
    if not daily_challenge:
        for g in games:
            if g["is_enabled"]:
                daily_challenge = {"slug": g["slug"], "name": g["name"], "emoji": g["emoji"]}
                break

    return render(request, "core/games_hub.html", {
        "games_json": json.dumps(games),
        "total_balls": total_balls,
        "lifetime_balls": lifetime_balls,
        "daily_balls": daily_balls,
        "min_balls": min_balls,
        "chaqmoq_per_conv": chaqmoq_per_conv,
        "can_convert": can_convert,
        "wallet_ring_offset": wallet_ring_offset,
        "total_games": 15,
        "games_today": games_today,
        "enabled_count": enabled_count,
        "streak": streak,
        "best_level": best_level,
        "daily_dashoffset": daily_dashoffset,
        "daily_pct_bar": daily_pct_bar,
        "recent_sessions": recent_sessions,
        "leaderboard": leaderboard,
        "global_leaderboard": global_leaderboard,
        "global_period": global_period,
        "my_global_row": my_global_row,
        "is_student": is_student,
        "student_name": u.get_full_name() or u.email or "O'quvchi",
        "daily_challenge": daily_challenge,
    })


@login_required
def game_play_view(request, game_slug):
    """O'quvchi: alohida o'yin sahifasi (har bir o'yin uchun)."""
    from core.api_views import _GAME_CONFIGS, _sp_center, _get_global_game_cfg, _get_game_center_cfg, _resolve_game_cfg
    from core.models import StudentGameProgress, GameSession

    if game_slug not in _GAME_CONFIGS:
        from django.http import Http404
        raise Http404

    center = _sp_center(request)
    u = request.user

    global_disabled = _get_global_game_cfg()
    center_cfg_map = _get_game_center_cfg(center)
    cfg = _GAME_CONFIGS[game_slug]
    eff = _resolve_game_cfg(game_slug, cfg, center_cfg_map, global_disabled)

    role = getattr(u, "role", None)
    is_student = role == "student"
    is_staff = u.is_superuser or role in ("director", "manager", "teacher")

    progress = None
    if is_student and center:
        progress, _ = StudentGameProgress.objects.get_or_create(
            center=center, student=u, game_slug=game_slug,
            defaults={"current_level": 1},
        )

    recent_sessions = []
    if is_student:
        recent_sessions = list(
            GameSession.objects.filter(center=center, student=u, game_slug=game_slug)
            .order_by("-played_at")[:10]
            .values("score", "coins_earned", "balls_earned", "played_at")
        )

    today = timezone.localdate()
    from core.models import StudentBallsWallet
    wallet = StudentBallsWallet.objects.filter(center=center, student=u).first() if is_student and center else None
    total_balls = wallet.total_balls if wallet else 0

    # Check if already earned ball today for this game
    from django.utils.timezone import localdate as _ld
    already_earned_ball = GameSession.objects.filter(
        student=u, center=center,
        game_slug=game_slug, balls_earned=1,
        played_at__date=_ld()
    ).exists() if is_student and center else False

    return render(request, "core/game_play.html", {
        "game_slug": game_slug,
        "game_cfg": cfg,
        "eff_cfg": eff,
        "progress": progress,
        "recent_sessions": recent_sessions,
        "total_balls": total_balls,
        "already_earned_ball": already_earned_ball,
        "is_enabled": eff is not None,
        "is_student": is_student,
        "is_staff": is_staff,
    })


@login_required
def game_questions_view(request):
    """Director/Manager: o'yin savollari boshqaruvi sahifasi."""
    from core.api_views import _GAME_CONFIGS, _sp_center
    from core.models import GameQuestion

    u = request.user
    if u.role not in ("director", "manager") and not u.is_superuser:
        from django.contrib import messages
        messages.error(request, "Ruxsat yo'q.")
        return redirect("core:home")

    center = _sp_center(request)
    games_info = []
    for slug, cfg in _GAME_CONFIGS.items():
        db_count = GameQuestion.objects.filter(center=center, game_slug=slug, is_active=True).count()
        ai_count = GameQuestion.objects.filter(center=center, game_slug=slug, is_active=True, is_ai_generated=True).count()
        games_info.append({
            "slug": slug,
            "name": cfg["name"],
            "emoji": cfg["emoji"],
            "db_count": db_count,
            "ai_count": ai_count,
        })

    return render(request, "core/game_questions.html", {
        "games_info": games_info,
        "game_configs_json": json.dumps({
            slug: {"name": cfg["name"], "emoji": cfg["emoji"]}
            for slug, cfg in _GAME_CONFIGS.items()
        }),
    })


# ═══════════════════════════════════════════════════════════════
# GLOBAL GAME RATING  — barcha markazlar orasidagi reyting
# ═══════════════════════════════════════════════════════════════
@login_required
def game_rating_view(request):
    """Barcha markazlar orasidagi global o'yin reytingi."""
    from core.models import GameSession, StudentGameProgress
    from django.db.models import Max

    period   = request.GET.get('period', 'week')
    today    = localdate()

    if period == 'week':
        date_from = today - timedelta(days=today.weekday())
    elif period == 'month':
        date_from = today.replace(day=1)
    else:
        date_from = None   # all-time

    # ── base queryset — use balls_earned for new system, fallback to coins_earned for old data ──
    qs = GameSession.objects.filter(balls_earned=1)
    if date_from:
        qs = qs.filter(played_at__date__gte=date_from)

    # ── top 100 globally (ranked by balls earned) ──
    top_rows = list(
        qs.values('student_id')
        .annotate(total_coins=Count('id'), games_count=Count('id'))
        .order_by('-total_coins')[:100]
    )

    student_ids = [r['student_id'] for r in top_rows]

    # Batch-fetch user info (center via FK)
    users = {
        u.id: u
        for u in User.objects.filter(id__in=student_ids)
        .select_related('center')
        .only('id', 'first_name', 'last_name', 'email', 'center')
    }

    # Batch-fetch max level per student
    levels = {
        row['student_id']: (row['max_lvl'] or 1)
        for row in StudentGameProgress.objects
        .filter(student_id__in=student_ids)
        .values('student_id')
        .annotate(max_lvl=Max('current_level'))
    }

    def _display_name(u):
        full = u.get_full_name().strip()
        if not full:
            return (u.email or "O'quvchi").split('@')[0]
        parts = full.split()
        return f"{parts[0]} {parts[1][0]}." if len(parts) > 1 else parts[0]

    leaderboard = []
    my_rank_in_top = None
    for i, row in enumerate(top_rows, 1):
        u = users.get(row['student_id'])
        if not u:
            continue
        entry = {
            'rank':   i,
            'name':   _display_name(u),
            'center': u.center.name if u.center else "—",
            'coins':  row['total_coins'],
            'games':  row['games_count'],
            'level':  levels.get(row['student_id'], 1),
            'is_me':  row['student_id'] == request.user.id,
        }
        if entry['is_me']:
            my_rank_in_top = i
        leaderboard.append(entry)

    # ── my position if outside top-100 ──
    my_row = None
    u = request.user
    is_student = getattr(u, 'role', None) == 'student'
    if is_student and my_rank_in_top is None:
        my_qs = GameSession.objects.filter(student=u, balls_earned=1)
        if date_from:
            my_qs = my_qs.filter(played_at__date__gte=date_from)
        my_agg   = my_qs.aggregate(total=Count('id'), cnt=Count('id'))
        my_total = my_agg['total'] or 0
        if my_total > 0:
            better_qs = GameSession.objects.filter(balls_earned=1)
            if date_from:
                better_qs = better_qs.filter(played_at__date__gte=date_from)
            my_rank = (
                better_qs.values('student_id')
                .annotate(t=Count('id'))
                .filter(t__gt=my_total)
                .count()
            ) + 1
            my_lvl  = StudentGameProgress.objects.filter(student=u).aggregate(ml=Max('current_level'))['ml'] or 1
            my_row  = {
                'rank':   my_rank,
                'name':   u.get_full_name() or u.email or "O'quvchi",
                'center': u.center.name if u.center else "—",
                'coins':  my_total,
                'games':  my_agg['cnt'] or 0,
                'level':  my_lvl,
                'is_me':  True,
            }

    # ── global stats ──
    stat_qs = GameSession.objects.filter(balls_earned=1)
    if date_from:
        stat_qs = stat_qs.filter(played_at__date__gte=date_from)

    total_players      = stat_qs.values('student_id').distinct().count()
    total_coins_global = stat_qs.aggregate(t=Count('id'))['t'] or 0

    return render(request, 'core/game_rating.html', {
        'leaderboard':        leaderboard,
        'period':             period,
        'my_row':             my_row,
        'is_student':         is_student,
        'total_players':      total_players,
        'total_coins_global': total_coins_global,
    })


# ─────────────────────────────────────────────
#  GURUH CHAT VIEWS
# ─────────────────────────────────────────────

@login_required
def chat_list_view(request):
    """Foydalanuvchining barcha guruh chatlari ro'yxati."""
    from education.models import Group, Enrollment
    from core.models import GroupChat, ChatMessage

    u = request.user
    center = getattr(u, 'center', None)
    if not center:
        return redirect('core:home')

    if u.role in ('director', 'manager'):
        groups = Group.objects.filter(center=center, is_archived=False).select_related('oqituvchi').order_by('nom')
    elif u.role == 'teacher':
        groups = Group.objects.filter(
            center=center, is_archived=False
        ).filter(
            Q(oqituvchi=u) | Q(support_teacher=u)
        ).select_related('oqituvchi').order_by('nom')
    elif u.role == 'student':
        enrolled_ids = Enrollment.objects.filter(
            student=u, is_deleted=False
        ).values_list('group_id', flat=True)
        groups = Group.objects.filter(
            center=center, id__in=enrolled_ids, is_archived=False
        ).select_related('oqituvchi').order_by('nom')
    else:
        groups = Group.objects.none()

    group_list = list(groups)
    chat_map = {
        gc.group_id: gc
        for gc in GroupChat.objects.filter(center=center, group__in=group_list)
    }

    groups_data = []
    for g in group_list:
        gc = chat_map.get(g.id)
        last_msg = None
        if gc:
            last_msg = (
                ChatMessage.objects.filter(chat=gc, is_deleted=False)
                .select_related('sender')
                .order_by('-created_at')
                .first()
            )
        groups_data.append({'group': g, 'last_msg': last_msg})

    return render(request, 'core/chat_list.html', {'groups_data': groups_data})


@login_required
def group_chat_view(request, group_id):
    """Guruh chat sahifasi."""
    from education.models import Group, Enrollment
    from core.models import GroupChat, ChatMessage

    u = request.user
    center = getattr(u, 'center', None)
    if not center:
        raise PermissionDenied

    group = get_object_or_404(Group, id=group_id, center=center, is_archived=False)

    if u.role in ('director', 'manager'):
        pass
    elif u.role == 'teacher':
        if group.oqituvchi_id != u.id and group.support_teacher_id != u.id:
            raise PermissionDenied
    elif u.role == 'student':
        if not Enrollment.objects.filter(group=group, student=u, is_deleted=False).exists():
            raise PermissionDenied
    else:
        raise PermissionDenied

    chat, _ = GroupChat.objects.get_or_create(center=center, group=group)

    from core.models import ChatPresence, ChatMessageRead

    # Update presence
    ChatPresence.objects.update_or_create(chat=chat, user=u, defaults={})

    # Online count
    cutoff = timezone.now() - timezone.timedelta(minutes=3)
    online_count = ChatPresence.objects.filter(chat=chat, last_seen__gte=cutoff).count()

    msgs_qs = list(reversed(list(
        ChatMessage.objects.filter(chat=chat, is_deleted=False)
        .select_related('sender', 'reply_to', 'reply_to__sender')
        .prefetch_related('attachments', 'reads')
        .order_by('-created_at')[:60]
    )))

    # Mark initial messages as read
    existing_reads = set(
        ChatMessageRead.objects.filter(message__in=msgs_qs, user=u)
        .values_list('message_id', flat=True)
    )
    new_reads = [
        ChatMessageRead(message=m, user=u)
        for m in msgs_qs
        if m.sender_id != u.id and m.id not in existing_reads
    ]
    if new_reads:
        ChatMessageRead.objects.bulk_create(new_reads, ignore_conflicts=True)

    def _msg_dict(m):
        atts = [
            {
                'type': a.att_type,
                'name': a.original_name,
                'url': a.file.url if a.file else '',
                'link': a.link_url,
                'size': a.file_size,
            }
            for a in m.attachments.all()
        ]
        reply = None
        if m.reply_to and not m.reply_to.is_deleted:
            reply = {
                'id': m.reply_to.id,
                'sender': m.reply_to.sender.get_full_name() or m.reply_to.sender.email,
                'sender_id': m.reply_to.sender_id,
                'body': m.reply_to.body[:80],
                'has_att': m.reply_to.attachments.exists(),
            }
        read_count = m.reads.exclude(user_id=u.id).count() if m.sender_id == u.id else 0
        return {
            'id': m.id,
            'sender_id': m.sender_id,
            'sender_name': m.sender.get_full_name() or m.sender.email,
            'sender_role': m.sender.role,
            'body': m.body,
            'reply': reply,
            'attachments': atts,
            'created_at': m.created_at.strftime('%H:%M'),
            'created_date': m.created_at.strftime('%Y-%m-%d'),
            'created_iso': m.created_at.isoformat(),
            'is_mine': m.sender_id == u.id,
            'read_count': read_count,
        }

    initial_msgs = [_msg_dict(m) for m in msgs_qs]
    last_id = msgs_qs[-1].id if msgs_qs else 0

    member_count = Enrollment.objects.filter(group=group, is_deleted=False).count()

    return render(request, 'core/group_chat.html', {
        'group': group,
        'chat': chat,
        'initial_msgs_json': initial_msgs,
        'last_id': last_id,
        'member_count': member_count,
        'online_count': online_count,
        'is_teacher': u.role in ('teacher', 'manager', 'director'),
        'me_id': u.id,
        'me_name': u.get_full_name() or u.email,
        'me_role': u.role,
    })
