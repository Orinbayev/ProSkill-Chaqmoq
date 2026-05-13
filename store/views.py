from .models import Product, PurchaseRequest, Comment
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages
from .models import Product, ProductImage, PurchaseRequest
from .services import approve_purchase, reject_purchase
from .forms import ProductForm
from chaqmoq.models import Ledger
from .services import convert_lead_to_student  # tepaga import qiling
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.shortcuts import get_object_or_404
from core.tenant import require_center, ensure_obj_center
from .models import Lead
from billing.decorators import require_feature
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta
from django.utils.dateparse import parse_date


def lead_detail(request, pk):
    center = require_center(request)
    lead = get_object_or_404(Lead, pk=pk)
    ensure_obj_center(request, lead.center_id)
    return render(request, "store/lead_detail.html", {"lead": lead})

# ✅ Mahsulotlar ro‘yxati
@login_required
@require_feature("store")
def products(request):
    """Mahsulotlar ro‘yxati va tanlab o‘chirish funksiyasi"""
    from core.tenant import get_request_center
    center = get_request_center(request)
    
    items = Product.objects.filter(Q(center=center) | Q(center__isnull=True)).order_by('-yaratilgan')

    # 🔹 Faqat manager, director yoki superuser qo‘sha / o‘chira oladi
    can_add = request.user.role in ('manager', 'director') or request.user.is_superuser

    # 🔹 Tanlangan mahsulotlarni o‘chirish (POST orqali)
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_products")
        if selected_ids:
            # ✅ Xavfsizlik: faqat shu center mahsulotlarini o'chirish
            Product.objects.filter(id__in=selected_ids, center=center).delete()
            messages.success(request, f"{len(selected_ids)} ta mahsulot o‘chirildi ✅")
            return redirect('store:products')
        else:
            messages.warning(request, "Hech qanday mahsulot tanlanmadi ❗")

    return render(
        request,
        'store/product_list.html',
        {
            'items': items,
            'can_add': can_add,
        },
    )



# ✅ Mahsulot tafsiloti
from django.contrib.auth import get_user_model
User = get_user_model()

@login_required
@require_feature("store")
def product_detail(request, pk):
    center = require_center(request)
    item = get_object_or_404(Product, pk=pk, center=center)
    user = request.user
    request_status = None

    if user.role == 'student':
        last_request = PurchaseRequest.objects.filter(
            student=user,
            product=item
        ).order_by('-sana').first()
        if last_request:
            request_status = last_request.status

    from chaqmoq.models import Ledger
    user_chaqmoq = Ledger.student_balansi(user.id, center=center)

    sotib_olganlar_soni = PurchaseRequest.objects.filter(
        product=item,
        status=PurchaseRequest.APPROVED
    ).count()

    comments = Comment.objects.filter(product=item, parent=None).prefetch_related('replies', 'user')
    yetarli = user_chaqmoq >= item.narx_chaqmoq

    return render(request, 'store/product_detail.html', {
        'item': item,
        'request_status': request_status,
        'comments': comments,
        'user_chaqmoq': user_chaqmoq,
        'yetarli': yetarli,
        'sotib_olganlar_soni': sotib_olganlar_soni,
    })



@login_required
@require_feature("store")
def add_comment(request, pk):
    if request.method == 'POST':
        item = get_object_or_404(Product, pk=pk)
        text = request.POST.get('text', '').strip()
        if text:
            Comment.objects.create(product=item, user=request.user, text=text)
        return redirect('store:product_detail', pk=pk)
    return redirect('store:product_detail', pk=pk)



@login_required
@require_feature("store")
def reply_comment(request, cid):
    parent = get_object_or_404(Comment, pk=cid)
    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        if text:
            Comment.objects.create(
                product=parent.product,
                user=request.user,
                text=text,
                parent=parent
            )
    return redirect('store:product_detail', pk=parent.product.id)

from django.db import transaction


# ✅ Mahsulotga so‘rov yuborish (student uchun)
# ✅ Mahsulotga so‘rov yuborish (student uchun)
@login_required
@require_feature("store")
def create_request(request, pk):
    """O‘quvchi mahsulot uchun so‘rov yuboradi"""
    product = get_object_or_404(Product, pk=pk)
    user = request.user  # Foydalanuvchini olamiz

    # ✅ 1. Faqat studentlarga ruxsat
    if user.role != 'student':
        messages.error(request, "Faqat o‘quvchilar sotib olishlari mumkin.")
        return redirect('store:product_detail', pk=pk)

    # ✅ 2. Agar so‘rov allaqachon yuborilgan bo‘lsa
    # ❌ faqat "pending" emas, balki "tasdiqlangan"ni ham qayta sotib olishga ruxsat beramiz
    if PurchaseRequest.objects.filter(
        student=user, product=product, status=PurchaseRequest.PENDING
    ).exists():
        messages.warning(request, "Siz bu mahsulot uchun so‘rov yuborgansiz. Tasdiqlanishini kuting.")
        return redirect('store:product_detail', pk=pk)


    # ✅ 3. Ledger orqali foydalanuvchining real chaqmoq balansini olish
    user_chaqmoq = Ledger.student_balansi(user.id, center=product.center)

    if user_chaqmoq < product.narx_chaqmoq:
        messages.error(
            request,
            f"Sizda yetarli chaqmoq mavjud emas. ({user_chaqmoq} / {product.narx_chaqmoq})"
        )
        return redirect('store:product_detail', pk=pk)

    # ✅ 4. So‘rovni yaratish
    with transaction.atomic():
        req_center = user.center if user.center else product.center
        
        PurchaseRequest.objects.create(
            center=req_center,
            student=user,
            product=product,
            qty=1,
            status=PurchaseRequest.PENDING
        )

    messages.success(
        request,
        f"So‘rovingiz yuborildi! Sizning joriy balansingiz: {user_chaqmoq - product.narx_chaqmoq} Chaqmoq."
    )
    return redirect('store:product_detail', pk=pk)


# ✅ Manager/Direktor uchun so‘rovlar ro‘yxati
@login_required
@require_feature("store")
def request_list(request):
    center = require_center(request)
    if request.user.role not in ('manager', 'director'):
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect('core:home')
    items = PurchaseRequest.objects.filter(center=center).order_by('-sana')
    return render(request, 'store/requests.html', {'items': items})


# ✅ So‘rovni tasdiqlash
@login_required
@require_feature("store")
def request_approve(request, pk):
    if request.user.role not in ('manager', 'director'):
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect(request.META.get('HTTP_REFERER', 'store:requests'))

    center = require_center(request)
    pr = get_object_or_404(PurchaseRequest, pk=pk, center=center)
    ok, msg = approve_purchase(pr, request.user)
    (messages.success if ok else messages.error)(request, msg)

    return redirect(request.META.get('HTTP_REFERER', 'store:requests'))


@login_required
@require_feature("store")
def request_reject(request, pk):
    if request.user.role not in ('manager', 'director'):
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect(request.META.get('HTTP_REFERER', 'store:requests'))

    pr = get_object_or_404(PurchaseRequest, pk=pk)
    ok, msg = reject_purchase(pr, request.user)
    (messages.success if ok else messages.error)(request, msg)

    return redirect(request.META.get('HTTP_REFERER', 'store:requests'))


# ✅ Mahsulot qo‘shish
@login_required
@require_feature("store")
def product_create(request):
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect('store:products')

    from .forms import ProductForm, ProductImageForm

    form = ProductForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        center = require_center(request)
        product = form.save(commit=False)
        product.center = center
        product.save()
        # Rasmlar yuklash
        files = request.FILES.getlist('rasmlar')
        for file in files:
            ProductImage.objects.create(product=product, rasm=file)
        messages.success(request, 'Mahsulot va rasmlar muvaffaqiyatli qo‘shildi!')
        return redirect('store:products')

    return render(request, 'store/product_form.html', {'form': form, 'title': "Mahsulot qo‘shish"})


@login_required
@require_feature("store")
def product_list(request):
    if request.method == "POST":
        return products(request)
    return redirect('store:products')


# ✅ Mahsulotni tahrirlash
@login_required
@require_feature("store")
def product_edit(request, pk):
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect('store:products')

    center = require_center(request)
    obj = get_object_or_404(Product, pk=pk, center=center)
    form = ProductForm(request.POST or None, request.FILES or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        product = form.save()
        
        # ✅ FIX: Rasmlarni tahrirlash paytida ham yuklash
        files = request.FILES.getlist('rasmlar')
        for file in files:
            ProductImage.objects.create(product=product, rasm=file)
            
        messages.success(request, 'Mahsulot saqlandi!')
        return redirect('store:products')

    return render(request, 'store/product_form.html', {'form': form, 'title': 'Mahsulotni tahrirlash'})


# ✅ Mahsulotni o‘chirish
@login_required
@require_feature("store")
def product_delete(request, pk):
    if request.user.role not in ('director',) and not request.user.is_superuser:
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect('store:products')

    center = require_center(request)
    obj = get_object_or_404(Product, pk=pk, center=center)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Mahsulot o‘chirildi!')
        return redirect('store:products')

    return render(request, 'accounts/logout_confirm.html', {})


@login_required
@require_feature("store")
def delete_product_image(request, pk):
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        return JsonResponse({"ok": False, "error": "Ruxsat yo'q"}, status=403)
    if request.method != 'POST':
        return JsonResponse({"ok": False, "error": "POST kerak"}, status=405)
    center = require_center(request)
    image = get_object_or_404(ProductImage, pk=pk, product__center=center)
    image.delete()
    return JsonResponse({"ok": True})


# 🔒 Talabaning balansini tekshirish
def _student_has_enough(user, price: int, center=None) -> bool:
    if getattr(user, 'role', None) != 'student':
        return True
    bal = Ledger.student_balansi(user.id, center=center)
    return bal >= price

from .models import Lead
from .forms import LeadForm
from .models import Lead, LeadStatus, Manba, Yonalish

from django.db.models import Count, Q

@login_required
@require_feature("leads")
def lead_list(request):
    center = require_center(request)

    status_id = (request.GET.get('status') or "").strip()
    q = (request.GET.get('q') or "").strip()

    # ✅ Tenant isolation: faqat shu center leadlari
    if request.user.is_superuser:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(center=center)

    if request.user.is_superuser:
         # Superuser uchun: faqat mavjud leadlarda ishlatilgan statuslarni chiqaramiz
         used_ids = leads.values_list('status', flat=True).distinct()
         statuses_qs = LeadStatus.objects.filter(id__in=used_ids)
    else:
         statuses_qs = LeadStatus.objects.filter(center=center)

    statuses = (
        statuses_qs
        .annotate(
            leads_count=Count('lead', distinct=True),
            converted_count=Count('lead', filter=Q(lead__converted_user__isnull=False), distinct=True),
        )
        .order_by('id')
    )

    if status_id:
        leads = leads.filter(status_id=status_id)

    if q:
        leads = leads.filter(
            Q(ism__icontains=q) |
            Q(familya__icontains=q) |
            Q(telefon1__icontains=q) |
            Q(telefon2__icontains=q)
        )

    leads = leads.order_by('-qoshilgan_sana')

    # Statistics context
    if request.user.is_superuser:
        base_leads_qs = Lead.objects.all()
    else:
        base_leads_qs = Lead.objects.filter(center=center)

    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    page_num = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', '10')

    if page_size == "all":
        paginator = Paginator(leads, max(1, leads.count()))
    else:
        try:
            page_size = int(page_size)
            if page_size < 1:
                page_size = 10
        except ValueError:
            page_size = 10
        paginator = Paginator(leads, page_size)

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'leads': page_obj.object_list,
        'page_obj': page_obj,
        'page_size': page_size,
        'statuses': statuses,
        'selected_status': status_id,
        'q': q,
        'total_count': base_leads_qs.count(),
        'total_converted': base_leads_qs.filter(converted_user__isnull=False).count(),
        'leads_count_filtered': leads.count(),
        'converted_count_filtered': leads.filter(converted_user__isnull=False).count(),
    }
    return render(request, 'store/lead_list.html', context)


@login_required
@require_feature("leads")
def lead_create(request):
    center = require_center(request)
    if request.method == 'POST':
        form = LeadForm(request.POST, center=center)
        if form.is_valid():
            lead = form.save(commit=False)
            lead.center = center
            
            # Auto-calculate age (safer calculation)
            if lead.birth_date:
                today = timezone.now().date()
                age = today.year - lead.birth_date.year - ((today.month, today.day) < (lead.birth_date.month, lead.birth_date.day))
                lead.yosh = max(0, age)
            else:
                lead.yosh = 0
                
            lead.save()

            # ✅ STATUS "Tasdiqlandi" bo'lsa avtomatik studentga o'tkazamiz
            if lead.status and (lead.status.nom or "").strip().lower() == "tasdiqlandi":
                user, password, created = convert_lead_to_student(lead, request.user)
                if created:
                    messages.success(
                        request,
                        f"✅ Yangi lead o‘quvchiga aylanib qo‘shildi! Login: {user.email} | Parol: {password}"
                    )
                else:
                    messages.info(
                        request,
                        f"ℹ️ Lead oldindan mavjud o‘quvchiga bog‘landi: {user.email}"
                    )
            else:
                messages.success(request, "✅ Yangi o‘quvchi (lead) qo‘shildi!")

            return redirect('store:lead_list')
    else:
        form = LeadForm(center=center)
    return render(request, 'store/lead_create.html', {'form': form})


# ✏️ Leadni tahrirlash

@login_required
@require_feature("leads")
def lead_edit(request, pk):
    center = require_center(request)
    if request.user.is_superuser:
        lead = get_object_or_404(Lead, pk=pk)
    else:
        lead = get_object_or_404(Lead, pk=pk, center=center)

    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead, center=center)
        if form.is_valid():
            lead = form.save(commit=False)
            
            # Auto-calculate age (safer calculation)
            if lead.birth_date:
                today = timezone.now().date()
                age = today.year - lead.birth_date.year - ((today.month, today.day) < (lead.birth_date.month, lead.birth_date.day))
                lead.yosh = max(0, age)
            else:
                lead.yosh = lead.yosh or 0
                
            lead.save()

            # ✅ STATUS "Tasdiqlandi" bo'lsa avtomatik studentga o'tkazamiz
            if lead.status and (lead.status.nom or "").strip().lower() == "tasdiqlandi":
                user, password, created = convert_lead_to_student(lead, request.user)

                if created:
                    messages.success(
                        request,
                        f"✅ Lead o‘quvchiga o‘tkazildi! Login: {user.email} | Parol: {password}"
                    )
                else:
                    messages.info(
                        request,
                        f"ℹ️ Lead oldindan mavjud o‘quvchiga bog‘landi: {user.email}"
                    )
            else:
                messages.success(request, "✏️ Ma’lumot tahrirlandi!")

            return redirect('store:lead_list')
    else:
        form = LeadForm(instance=lead, center=center)

    return render(request, 'store/lead_edit.html', {'form': form, 'lead': lead})


# 🗑️ Leadni o‘chirish
@login_required
@require_feature("leads")
def lead_delete(request, pk):
    center = require_center(request)
    if request.user.is_superuser:
        lead = get_object_or_404(Lead, pk=pk)
    else:
        lead = get_object_or_404(Lead, pk=pk, center=center)

    if request.method == 'POST':
        lead.delete()
        messages.warning(request, "🗑️ Lead o‘chirildi!")
        return redirect('store:lead_list')
    return render(request, 'store/lead_delete.html', {'lead': lead})




@login_required
@require_feature("leads")
def lead_detail(request, pk):
    center = require_center(request)
    if request.user.is_superuser:
        lead = get_object_or_404(Lead, pk=pk)
    else:
        lead = get_object_or_404(Lead, pk=pk, center=center)
    return render(request, "store/lead_detail.html", {"lead": lead})



@require_POST
@login_required
def lead_convert(request, pk):
    lead = get_object_or_404(Lead, pk=pk)

    # ruxsat
    if not (request.user.is_superuser or request.user.role in ('manager', 'director')):
        messages.error(request, "Ruxsat yo‘q.")
        return redirect('store:lead_list')

    # statusni "Tasdiqlandi" ga qo'yamiz
    tasdiq = LeadStatus.objects.filter(nom="Tasdiqlandi").first()
    if tasdiq:
        lead.status = tasdiq
        lead.save(update_fields=["status"])

    # ✅ Center logic: 
    # Agar superuser bo'lsa va leadning markazi bo'lmasa yoki boshqa center bo'lsa,
    # biz ayni vaqtda qaysi centerda turgan bo'lsak, o'sha centerni target qilib beramiz.
    
    from core.tenant import get_request_center
    current_center = get_request_center(request)
    
    # Agar superuser bo'lsa, current_center None bo'lishi mumkin (agar center pickerda bo'lsa).
    # Bunday holda lead.center ishlatiladi (default).
    
    user, password, created = convert_lead_to_student(
        lead, 
        converted_by=request.user, 
        target_center=current_center
    )

    if created:
        messages.success(request, f"✅ O‘tkazildi! Login: {user.email} | Parol: {password}")
    else:
        messages.info(request, f"ℹ️ Lead mavjud o‘quvchiga bog‘landi: {user.email}")

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or reverse("store:lead_list")
    return redirect(next_url)


# ⚙️ Lead Sozlamalari (Manba, Yo'nalish, Status)
@login_required
@require_feature("leads")
def lead_settings(request):
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, 'Ruxsat yo‘q.')
        return redirect('store:lead_list')

    manbalar = Manba.objects.filter(center=center)
    yonalishlar = Yonalish.objects.filter(center=center)
    statuses = LeadStatus.objects.filter(center=center)

    return render(request, 'store/lead_settings.html', {
        'manbalar': manbalar,
        'yonalishlar': yonalishlar,
        'statuses': statuses
    })


@require_POST
@login_required
def lead_config_add(request):
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        return redirect('store:lead_settings')

    conf_type = request.POST.get('type')  # 'manba' | 'yonalish' | 'status'
    nom = request.POST.get('nom', '').strip()

    if not nom:
        messages.error(request, "Nom kiritilmadi!")
        return redirect('store:lead_settings')

    try:
        if conf_type == 'manba':
            Manba.objects.create(center=center, nom=nom)
            messages.success(request, f"Manba qo'shildi: {nom}")
        elif conf_type == 'yonalish':
            Yonalish.objects.create(center=center, nom=nom)
            messages.success(request, f"Yo'nalish qo'shildi: {nom}")
        elif conf_type == 'status':
            LeadStatus.objects.create(center=center, nom=nom)
            messages.success(request, f"Status qo'shildi: {nom}")
    except Exception as e:
        messages.error(request, f"Xatolik: {str(e)}")

    return redirect('store:lead_settings')


@login_required
def lead_config_delete(request, type_code, pk):
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:lead_settings')

    model_map = {
        'manba': Manba,
        'yonalish': Yonalish,
        'status': LeadStatus
    }
    
    ModelClass = model_map.get(type_code)
    if ModelClass:
        obj = get_object_or_404(ModelClass, pk=pk, center=center)
        if request.method == 'POST':
            obj.delete()
            messages.success(request, "O'chirildi!")
    
    return redirect('store:lead_settings')


# ===================================
# 8️⃣ XARAJATLAR (EXPENSE) - TEST UCHUN
# ===================================
from .models import Expense
from django.db.models import Sum

@login_required
@require_feature("xarajatlar")
def expenses(request):
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')
    
    from .models import ExpenseCategory
    from django.db.models.functions import TruncMonth
    
    # 1. Base QuerySet
    items_qs = Expense.objects.filter(center=center).select_related('category', 'worker', 'product').order_by('-sana')

    # 2. Filters
    sana_dan_raw = (request.GET.get('sana_dan') or '').strip()
    sana_gacha_raw = (request.GET.get('sana_gacha') or '').strip()
    category_id = request.GET.get('category')
    payment_method = request.GET.get('payment_method')
    worker_id = request.GET.get('worker')
    q = request.GET.get('q')

    today = timezone.localdate()
    current_month_start = today.replace(day=1)
    sana_dan = parse_date(sana_dan_raw) if sana_dan_raw else None
    sana_gacha = parse_date(sana_gacha_raw) if sana_gacha_raw else None

    if not sana_dan and not sana_gacha:
        sana_dan = current_month_start
        sana_gacha = today
    else:
        if sana_dan and not sana_gacha:
            sana_gacha = today if sana_dan <= today else sana_dan
        elif sana_gacha and not sana_dan:
            sana_dan = sana_gacha.replace(day=1)
        if sana_dan and sana_gacha and sana_dan > sana_gacha:
            sana_gacha = sana_dan

    def _apply_shared_filters(qs):
        if category_id and category_id.isdigit():
            qs = qs.filter(category_id=category_id)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)
        if worker_id and worker_id.isdigit():
            qs = qs.filter(worker_id=worker_id)
        if q:
            qs = qs.filter(Q(izoh__icontains=q) | Q(receiver__icontains=q))
        return qs

    items_qs = _apply_shared_filters(items_qs).filter(sana__date__gte=sana_dan, sana__date__lte=sana_gacha)

    # 3. Aggregations
    total_sum = Expense.objects.filter(center=center).aggregate(Sum('summa'))['summa__sum'] or 0
    filtered_sum = items_qs.aggregate(Sum('summa'))['summa__sum'] or 0

    # Large expense lists were rendering every row at once. Keep totals/chart based
    # on the full filtered set, but render only one page to prevent request resets.
    from django.core.paginator import Paginator
    page_size_options = (10, 20, 50, 100)
    try:
        per_page = int((request.GET.get("per_page") or "50").strip())
    except (TypeError, ValueError):
        per_page = 50
    if per_page not in page_size_options:
        per_page = 50
    paginator = Paginator(items_qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))
    items = page_obj.object_list

    # 4. Chart Data (oxirgi 12 oy)
    month_names = {
        1: "Yanvar",
        2: "Fevral",
        3: "Mart",
        4: "Aprel",
        5: "May",
        6: "Iyun",
        7: "Iyul",
        8: "Avgust",
        9: "Sentyabr",
        10: "Oktyabr",
        11: "Noyabr",
        12: "Dekabr",
    }

    def _add_month(month_value, offset):
        year = month_value.year + (month_value.month - 1 + offset) // 12
        month = (month_value.month - 1 + offset) % 12 + 1
        return month_value.replace(year=year, month=month, day=1)

    chart_anchor_month = sana_gacha.replace(day=1)
    buckets = [_add_month(chart_anchor_month, -11 + index) for index in range(12)]
    chart_start = buckets[0]
    chart_end = _add_month(buckets[-1], 1) - timedelta(days=1)
    chart_items = _apply_shared_filters(
        Expense.objects.filter(center=center)
    ).filter(sana__date__gte=chart_start, sana__date__lte=chart_end)
    chart_rows = (
        chart_items.annotate(bucket=TruncMonth('sana'))
        .values('bucket')
        .annotate(total=Sum('summa'))
        .order_by('bucket')
    )
    totals_by_bucket = {}
    for entry in chart_rows:
        bucket = entry['bucket']
        if hasattr(bucket, 'date'):
            bucket = bucket.date()
        bucket = bucket.replace(day=1)
        totals_by_bucket[bucket] = int(entry['total'] or 0)
    chart_labels = [month_names.get(bucket.month, bucket.strftime('%B')) for bucket in buckets]
    chart_data = [totals_by_bucket.get(bucket, 0) for bucket in buckets]
    chart_kicker = "Oxirgi 12 oy"
    chart_period_label = f"{chart_labels[0]} dan {chart_labels[-1]} gacha"
        
    # 5. Context
    # Ensure "Do'kon" category exists so it shows in the filter
    ExpenseCategory.objects.get_or_create(nom="Do'kon", center=center)
    categories = ExpenseCategory.objects.filter(center=center)
    workers = User.objects.filter(center=center, role__in=['manager', 'director'])

    # Dynamic payment methods (PaymentMethod table) — auto-seed defaults so
    # the dropdown is never empty even on a fresh center.
    from .models import PaymentMethod
    _ensure_default_payment_methods(center)
    dynamic_payment_methods = list(
        PaymentMethod.objects.filter(center=center, is_active=True).order_by('nom')
    )
    
    context = {
        'items': items,
        'page_obj': page_obj,
        'items_count': paginator.count,
        'per_page': per_page,
        'page_size_options': page_size_options,
        'total_sum': total_sum,
        'filtered_sum': filtered_sum,
        'categories': categories,
        'workers': workers,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'chart_kicker': chart_kicker,
        'chart_period_label': chart_period_label,
        'selected_period_label': (
            f"{sana_dan.day}-{month_names.get(sana_dan.month, sana_dan.strftime('%B'))} {sana_dan.year}"
            if sana_dan == sana_gacha
            else f"{sana_dan.day}-{month_names.get(sana_dan.month, sana_dan.strftime('%B'))} {sana_dan.year} dan "
            f"{sana_gacha.day}-{month_names.get(sana_gacha.month, sana_gacha.strftime('%B'))} {sana_gacha.year} gacha"
        ),
        # filters
        'sana_dan': sana_dan.isoformat(),
        'sana_gacha': sana_gacha.isoformat(),
        'selected_cat': int(category_id) if category_id and category_id.isdigit() else None,
        'selected_method': payment_method,
        'selected_worker': int(worker_id) if worker_id and worker_id.isdigit() else None,
        'q': q,
        'dynamic_payment_methods': dynamic_payment_methods,
    }

    return render(request, 'store/expenses.html', context)


@login_required
@require_feature("xarajatlar")
def expenses_export_xlsx(request):
    """Xarajatlarni filtrlar bo'yicha Excel (.xlsx) formatda yuklab berish."""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    sana_dan_raw = (request.GET.get('sana_dan') or '').strip()
    sana_gacha_raw = (request.GET.get('sana_gacha') or '').strip()
    category_id = request.GET.get('category')
    payment_method = request.GET.get('payment_method')
    worker_id = request.GET.get('worker')
    q = request.GET.get('q')

    today = timezone.localdate()
    sana_dan = parse_date(sana_dan_raw) if sana_dan_raw else None
    sana_gacha = parse_date(sana_gacha_raw) if sana_gacha_raw else None

    if not sana_dan and not sana_gacha:
        sana_dan = today.replace(day=1)
        sana_gacha = today
    else:
        if sana_dan and not sana_gacha:
            sana_gacha = today if sana_dan <= today else sana_dan
        elif sana_gacha and not sana_dan:
            sana_dan = sana_gacha.replace(day=1)
        if sana_dan and sana_gacha and sana_dan > sana_gacha:
            sana_gacha = sana_dan

    items = (
        Expense.objects.filter(center=center)
        .select_related('category', 'worker', 'product')
        .filter(sana__date__gte=sana_dan, sana__date__lte=sana_gacha)
        .order_by('-sana')
    )
    if category_id and str(category_id).isdigit():
        items = items.filter(category_id=category_id)
    if payment_method:
        items = items.filter(payment_method=payment_method)
    if worker_id and str(worker_id).isdigit():
        items = items.filter(worker_id=worker_id)
    if q:
        items = items.filter(Q(izoh__icontains=q) | Q(receiver__icontains=q))

    payment_label = {'naqd': 'Naqd', 'plastik': 'Plastik', 'otkazma': "O'tkazma"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    row_font = Font(name="Calibri", size=11, color="1F2937")
    total_fill = PatternFill("solid", fgColor="FFF3CD")
    total_font = Font(name="Calibri", size=11, bold=True, color="92400E")
    title_font = Font(name="Calibri", size=14, bold=True, color="111827")
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    center_name = getattr(center, 'name', '') or ''
    title_cell.value = f"Xarajatlar hisoboti — {center_name}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:I2')
    period_cell = ws['A2']
    period_cell.value = (
        f"Davr: {sana_dan.strftime('%d.%m.%Y')} — {sana_gacha.strftime('%d.%m.%Y')} | "
        f"Jami yozuvlar: {items.count()} ta"
    )
    period_cell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    period_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["№", "Sana", "Vaqt", "Nomi (Izoh)", "Kategoriya",
               "Qabul qiluvchi", "To'lov usuli", "Summa (so'm)", "Xodim (kim tomonidan yozilgan)"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[4].height = 28

    total_summa = 0
    start_data_row = 5
    for idx, item in enumerate(items, start=1):
        row = start_data_row + idx - 1
        local_sana = timezone.localtime(item.sana) if item.sana else None
        worker = item.worker
        if worker:
            worker_name = (worker.get_full_name() or '').strip() or worker.username
        else:
            worker_name = "—"

        values = [
            idx,
            local_sana.strftime('%d.%m.%Y') if local_sana else "—",
            local_sana.strftime('%H:%M') if local_sana else "—",
            item.izoh or "—",
            item.category.nom if item.category else "—",
            item.receiver or "—",
            payment_label.get(item.payment_method, item.payment_method or "—"),
            int(item.summa or 0),
            worker_name,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = row_font
            cell.border = border
            if col_idx in (1, 2, 3, 7):
                cell.alignment = center_align
            elif col_idx == 8:
                cell.alignment = right_align
                cell.number_format = '#,##0'
            else:
                cell.alignment = left_align
        total_summa += int(item.summa or 0)

    total_row = start_data_row + items.count()
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    label_cell = ws.cell(row=total_row, column=1, value="JAMI")
    label_cell.font = total_font
    label_cell.fill = total_fill
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border = border
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = total_fill
    total_cell = ws.cell(row=total_row, column=8, value=total_summa)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = right_align
    total_cell.number_format = '#,##0'
    total_cell.border = border
    empty_cell = ws.cell(row=total_row, column=9, value="")
    empty_cell.fill = total_fill
    empty_cell.border = border

    widths = [6, 14, 10, 38, 20, 22, 14, 18, 28]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"xarajatlar_{sana_dan.isoformat()}_{sana_gacha.isoformat()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _user_can_manage_expenses(user) -> bool:
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if user.is_superuser:
        return True
    return getattr(user, "role", None) in ("manager", "director")


@login_required
@require_feature("xarajatlar")
def expense_create(request):
    """
    Yangi xarajat qo'shish
    """
    if not _user_can_manage_expenses(request.user):
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:expenses')
    if request.method == 'POST':
        center = require_center(request)
        from .models import Expense
        from django.utils import timezone
        from django.utils.dateparse import parse_date, parse_datetime

        summa_raw = (request.POST.get('summa') or '').strip()
        izoh = request.POST.get('izoh', '')
        category_id = request.POST.get('category')
        payment_method = request.POST.get('payment_method', 'naqd')
        sana_raw = (request.POST.get('sana') or '').strip()

        try:
            summa = int(summa_raw)
        except (TypeError, ValueError):
            messages.error(request, "Summa noto'g'ri kiritildi")
            return redirect('store:expenses')
        if summa <= 0:
            messages.error(request, "Summa noto'g'ri kiritildi")
            return redirect('store:expenses')

        receiver = request.POST.get('receiver', '')

        sana = None
        if sana_raw:
            parsed = parse_datetime(sana_raw)
            if parsed is None:
                # parse_date date qaytaradi — uni aware datetime'ga aylantirish kerak
                d = parse_date(sana_raw)
                if d is not None:
                    from datetime import datetime as _dt, time as _time
                    parsed = timezone.make_aware(_dt.combine(d, _time.min))
            if parsed is None:
                messages.error(request, "Sana formati noto'g'ri (YYYY-MM-DD kutiladi)")
                return redirect('store:expenses')
            # Naive datetime bo'lsa aware qilish
            if timezone.is_naive(parsed):
                parsed = timezone.make_aware(parsed)
            sana = parsed
        if sana is None:
            sana = timezone.now()

        Expense.objects.create(
            center=center,
            summa=summa,
            izoh=izoh,
            receiver=receiver,
            category_id=category_id if category_id else None,
            payment_method=payment_method,
            sana=sana,
            worker=request.user
        )
        messages.success(request, "Xarajat qo'shildi")

    return redirect('store:expenses')


@login_required
@require_feature("xarajatlar")
def expense_category_create(request):
    """
    Yangi xarajat toifasi (category) qo'shish
    """
    if not _user_can_manage_expenses(request.user):
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:expenses')
    if request.method == 'POST':
        center = require_center(request)
        from .models import ExpenseCategory

        nom = request.POST.get('nom', '').strip()

        if nom:
            ExpenseCategory.objects.create(center=center, nom=nom)
            messages.success(request, "Yangi toifa qo'shildi")
        else:
            messages.error(request, "Toifa nomi kiritilmadi")

    return redirect('store:expenses')

@login_required
@require_feature("xarajatlar")
def expense_edit(request, pk):
    if not _user_can_manage_expenses(request.user):
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:expenses')
    center = require_center(request)
    from .models import Expense, ExpenseCategory

    expense = get_object_or_404(Expense, pk=pk, center=center)

    if request.method == 'POST':
        from django.utils.dateparse import parse_date, parse_datetime

        summa_raw = (request.POST.get('summa') or '').strip()
        izoh = request.POST.get('izoh', '')
        sana_str = (request.POST.get('sana') or '').strip()
        category_id = request.POST.get('category_id') # ID string bo'lishi mumkin
        payment_method = request.POST.get('payment_method')
        receiver = request.POST.get('receiver', '')

        try:
            summa = int(summa_raw)
        except (TypeError, ValueError):
            messages.error(request, "Summa noto'g'ri")
            return redirect('store:expenses')
        if summa <= 0:
            messages.error(request, "Summa noto'g'ri")
            return redirect('store:expenses')

        expense.summa = summa
        expense.izoh = izoh
        expense.receiver = receiver
        expense.payment_method = payment_method

        if sana_str:
            parsed_sana = parse_datetime(sana_str)
            if parsed_sana is None:
                d = parse_date(sana_str)
                if d is not None:
                    from datetime import datetime as _dt, time as _time
                    from django.utils import timezone as _tz
                    parsed_sana = _tz.make_aware(_dt.combine(d, _time.min))
            if parsed_sana is None:
                messages.error(request, "Sana formati noto'g'ri (YYYY-MM-DD kutiladi)")
                return redirect('store:expenses')
            from django.utils import timezone as _tz
            if _tz.is_naive(parsed_sana):
                parsed_sana = _tz.make_aware(parsed_sana)
            expense.sana = parsed_sana

        if category_id:
            try:
                cat = ExpenseCategory.objects.get(pk=category_id, center=center)
                expense.category = cat
            except ExpenseCategory.DoesNotExist:
                 expense.category = None
        else:
            expense.category = None

        expense.save()
        messages.success(request, "Xarajat tahrirlandi")

    return redirect('store:expenses')

@login_required
@require_feature("xarajatlar")
def expense_delete(request, pk):
    if not _user_can_manage_expenses(request.user):
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:expenses')
    center = require_center(request)
    from .models import Expense

    if request.method == 'POST':
        expense = get_object_or_404(Expense, pk=pk, center=center)
        expense.delete()
        messages.success(request, "Xarajat o'chirildi")

    return redirect('store:expenses')


@login_required
@require_feature("xarajatlar")
def expense_comment(request, pk):
    """
    Faqat izohni o'zgartirish (Comment funksiyasi uchun)
    """
    if not _user_can_manage_expenses(request.user):
        messages.error(request, "Ruxsat yo'q")
        return redirect('store:expenses')
    center = require_center(request)
    from .models import Expense

    if request.method == 'POST':
        expense = get_object_or_404(Expense, pk=pk, center=center)
        izoh = request.POST.get('izoh', '').strip()
        if izoh:
            expense.izoh = izoh
            expense.save()
            messages.success(request, "Izoh yangilandi")

    return redirect('store:expenses')


# =============================================================
# To'lov usullari (PaymentMethod) — CRUD
# =============================================================
DEFAULT_PAYMENT_METHODS = ("NAQD", "KARTA", "ARALASH", "CLICK", "PAYME")


def _ensure_default_payment_methods(center):
    from .models import PaymentMethod
    if not center:
        return
    # Har page loadda 5 get_or_create (5q) ni cache bilan kamaytiramiz.
    # Birinchi marta yaratilganidan keyin 1 soat davomida DB query yo'q.
    from django.core.cache import cache as _cache
    _ck = f"store:pm_seeded:v1:{center.pk}"
    if _cache.get(_ck):
        return
    for nom in DEFAULT_PAYMENT_METHODS:
        PaymentMethod.objects.get_or_create(center=center, nom=nom, defaults={"is_active": True})
    _cache.set(_ck, 1, timeout=3600)


@login_required
@require_feature("finance")
def payment_methods(request):
    from .models import PaymentMethod
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    _ensure_default_payment_methods(center)
    methods = PaymentMethod.objects.filter(center=center).order_by('-is_active', 'nom')
    return render(request, 'store/payment_methods.html', {'methods': methods})


@login_required
@require_feature("finance")
def payment_method_create(request):
    from .models import PaymentMethod
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    if request.method == 'POST':
        nom = (request.POST.get('nom') or '').strip().upper()
        if not nom:
            messages.error(request, "To'lov usuli nomi bo'sh bo'lmasligi kerak")
        else:
            _, created = PaymentMethod.objects.get_or_create(
                center=center, nom=nom, defaults={'is_active': True}
            )
            if created:
                messages.success(request, f"\"{nom}\" qo'shildi")
            else:
                messages.info(request, f"\"{nom}\" allaqachon mavjud")

    return redirect('store:payment_methods')


@login_required
@require_feature("finance")
def payment_method_update(request, pk):
    from .models import PaymentMethod
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    method = get_object_or_404(PaymentMethod, pk=pk, center=center)
    if request.method == 'POST':
        nom = (request.POST.get('nom') or '').strip().upper()
        if not nom:
            messages.error(request, "Nom bo'sh bo'lmasligi kerak")
        else:
            method.nom = nom
            try:
                method.save(update_fields=['nom', 'updated_at'])
                messages.success(request, "Yangilandi")
            except Exception:
                messages.error(request, "Bunday nomli usul allaqachon mavjud")

    return redirect('store:payment_methods')


@login_required
@require_feature("finance")
def payment_method_toggle(request, pk):
    from .models import PaymentMethod
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    method = get_object_or_404(PaymentMethod, pk=pk, center=center)
    if request.method == 'POST':
        method.is_active = not method.is_active
        method.save(update_fields=['is_active', 'updated_at'])
        messages.success(request, "Holat o'zgartirildi")

    return redirect('store:payment_methods')


@login_required
@require_feature("finance")
def payment_method_delete(request, pk):
    from .models import PaymentMethod
    center = require_center(request)
    if request.user.role not in ('manager', 'director') and not request.user.is_superuser:
        messages.error(request, "Ruxsat yo'q")
        return redirect('core:home')

    method = get_object_or_404(PaymentMethod, pk=pk, center=center)
    if request.method == 'POST':
        method.delete()
        messages.success(request, "O'chirildi")

    return redirect('store:payment_methods')
